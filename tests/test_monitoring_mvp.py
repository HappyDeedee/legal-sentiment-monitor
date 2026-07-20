from __future__ import annotations

import asyncio
import json
import os
import re
import smtplib
import sqlite3
import sys
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path

import httpx
import pytest
from fastapi import HTTPException
from openpyxl import load_workbook

from api.monitoring.ai import _build_endpoint, _parse_json, _validate_ai_output, build_evaluation_payload, test_ai as run_ai_config_test
from api.monitoring.ai import DEFAULT_PROMPT
from api.monitoring.auth import SESSION_COOKIE_NAME
from api.monitoring.account_environment import resolve_account_profile_path
import api.monitoring.database as database_module
from api.monitoring.database import (
    acquire_account_lock,
    acquire_proxy_lock,
    authenticate_user,
    bootstrap_admin_from_env,
    ai_evaluation_trace_state,
    cleanup_ai_evaluation_traces,
    create_login_session,
    create_run,
    email_send_window_key,
    expire_login_sessions_for_account,
    finish_run,
    get_active_ai_key_profile,
    get_ai_config,
    get_conn,
    get_dashboard_summary,
    get_email_config,
    get_job,
    get_login_session,
    get_platform_login_config,
    get_report,
    get_run,
    get_runtime_setting_value,
    get_ai_evaluation_trace,
    get_social_account,
    get_user_by_email,
    has_running_run_for_job,
    init_db,
    list_ai_key_profiles,
    list_ai_rule_profiles,
    list_email_delivery_logs,
    list_email_templates,
    list_jobs,
    list_leads,
    list_login_sessions,
    list_platform_login_configs,
    list_proxy_profiles,
    list_reports,
    list_runtime_settings,
    list_runs,
    list_social_accounts,
    list_users,
    mark_selftest_jobs_internal,
    preview_crawl_run_job_id_backfill,
    record_email_delivery_log,
    record_skipped_run,
    recover_stale_runs_and_locks,
    release_account_lock,
    release_proxy_locks,
    release_run_resource_locks,
    render_email_template_preview,
    save_ai_config,
    save_ai_key_profile,
    save_ai_rule_profile,
    save_email_config,
    save_email_template,
    save_job,
    save_platform_login_config,
    save_proxy_profile,
    save_runtime_settings,
    save_social_account,
    save_user,
    save_ai_evaluation_trace,
    set_active_ai_key_profile,
    set_active_ai_rule_profile,
    update_social_account_check_state,
)
from api.monitoring.mediacrawler_login import get_mediacrawler_login_capability
from api.monitoring.login_browser import build_login_browser_command, open_login_browser, open_login_browser_with_command
import api.monitoring.account_check as account_check_module
import api.monitoring.login_qrcode as login_qrcode_module
import api.monitoring.mediacrawler_login as mediacrawler_login_module
from api.monitoring.login_state import login_window_status, record_login_window
from api.monitoring.mailer import REAL_EMAIL_BLOCKED_MESSAGE, build_report_email, render_report_email_preview, send_report, send_test_email
from api.monitoring.normalizer import collect_platform_outputs, in_time_window, normalize_content, parse_jsonl_file, resolve_window
from api.monitoring.platform_status import list_platform_status
from api.monitoring.preflight import build_job_preflight
from api.monitoring.readiness import get_readiness_status
from api.monitoring.reporting import create_report, resend_report_email, send_report_with_delivery_log
from api.monitoring.security import redact_sensitive
import api.monitoring.startup_launcher as api_monitoring_startup_launcher
from api.monitoring.avatar_cache import AVATAR_CACHE_DIR
from api.monitoring.selftest import create_sample_report
from api.monitoring.smoke import run_smoke_check
from api.monitoring.cli import run_due_jobs
from api.monitoring.doctor import run_doctor
from api.monitoring.startup_launcher import build_launch_plan
from cmd_arg import parse_cmd as parse_mediacrawler_cmd
from api.routers import monitor as monitor_router
import api.monitoring.cli as cli_module
import api.monitoring.ai as ai_module
import api.monitoring.readiness as readiness_module
import api.monitoring.runner as runner_module
import api.monitoring.scheduler as scheduler_module
from api.monitoring.runner import evaluate_new_contents, ingest_outputs
from api.monitoring.runner import run_job as run_monitor_job
from api.monitoring.scheduler import _is_due, next_run_at, scheduler_disabled_reason, scheduler_status
from tools.cdp_browser import resolve_cdp_user_data_dir
from scripts.pilot_gate_c_evidence import build_template, validate_evidence, write_template
from scripts.review_orphan_email_evidence import build_orphan_email_evidence_review, main as review_orphan_email_main


def _monitor_section(page: str, section_id: str) -> str:
    start = page.index(f'<section id="{section_id}"')
    end = page.find("\n      <section id=", start + 1)
    if end == -1:
        end = page.index("</main>", start)
    return page[start:end]


def _monitor_inline_styles(page: str) -> str:
    return "\n".join(
        part.split("</style>", 1)[0] for part in page.split("<style>")[1:]
    )


def _task_group_view(page: str) -> str:
    return _monitor_section(page, "runs")


@pytest.fixture(autouse=True)
def _clear_ai_skip_env(monkeypatch):
    monkeypatch.delenv("MONITOR_SKIP_AI_API", raising=False)


def test_root_entry_redirects_to_monitor_admin():
    from api import main as api_main

    response = asyncio.run(api_main.serve_frontend())

    assert response.status_code in {302, 307}
    assert response.headers["location"] == "/monitor"


def test_environment_check_returns_customer_safe_text(monkeypatch):
    from api import main as api_main

    class Result:
        returncode = 0
        stdout = "internal details"
        stderr = ""

    monkeypatch.setattr("api.main.subprocess.run", lambda *args, **kwargs: Result())

    result = asyncio.run(api_main.check_environment())
    visible = json.dumps(result, ensure_ascii=False)

    assert result["success"] is True
    assert result["message"] == "采集运行环境可用"
    assert result["output"] == "运行环境检查通过"
    for forbidden in ["MediaCrawler", "uv run", "main.py", "CLI"]:
        assert forbidden not in visible


def _complete_pilot_gate_c_evidence():
    payload = build_template()
    payload["status"] = "passed"
    payload["real_email_toggle"].update(
        {
            "operator": "pilot operator",
            "started_at": "2026-06-17T10:00:00+08:00",
            "ended_at": "2026-06-17T10:30:00+08:00",
        }
    )
    payload["server_like_environment"].update(
        {
            "environment_reference": "deployment log entry pilot-20260617",
            "service_started": True,
            "web_ui_admin_login": True,
            "server_side_browser_used": True,
            "local_chrome_not_required": True,
            "profile_root_persistent": True,
        }
    )
    payload["real_platform_workflow"].update(
        {
            "platform": "douyin",
            "account_reference": "platform account id 12, redacted",
            "web_qr_status_login_completed": True,
            "server_side_profile_persisted": True,
            "crawl_completed_with_server_profile": True,
            "run_reference": "run id 345, redacted logs archived",
            "report_reference": "report id 678, redacted artifact archived",
            "report_generated": True,
        }
    )
    payload["ai_fallback"].update(
        {
            "scenario": "AI disabled during pilot validation",
            "ai_unavailable_or_failure_exercised": True,
            "pending_review_or_manual_review_recorded": True,
            "report_generated": True,
            "evidence_reference": "run summary shows pending review counts",
        }
    )
    payload["smtp_validation"].update(
        {
            "recipient_reference": "operator-approved pilot recipient",
            "delivery_log_reference": "email delivery log id 901",
            "recipient_receipt_reference": "operator confirmed recipient inbox or spam folder receipt",
            "admin_toggle_enabled_for_validation": True,
            "real_smtp_send_succeeded": True,
            "delivery_recorded": True,
            "recipient_receipt_confirmed": True,
            "admin_toggle_disabled_after_validation": True,
            "default_paths_non_sending_confirmed": True,
        }
    )
    payload["redaction"].update(
        {
            "checked_surfaces": ["logs", "reports", "delivery_records", "ui_or_api"],
            "no_sensitive_values_found": True,
            "evidence_reference": "redaction checklist entry pilot-20260617",
            "notes": "all references are masked",
        }
    )
    payload["non_blocker_boundary"].update(
        {
            "non_blockers_confirmed": True,
            "historical_mutation_not_performed": True,
        }
    )
    return payload


def test_pilot_gate_c_evidence_template_is_incomplete_and_side_effect_free(tmp_path):
    target = tmp_path / "pilot_gate_c_evidence.json"

    write_template(target)
    payload = json.loads(target.read_text(encoding="utf-8"))
    issues = validate_evidence(payload)

    assert payload["schema_version"] == "pilot_gate_c_v3"
    assert payload["status"] == "incomplete"
    assert payload["smtp_validation"]["recipient_receipt_confirmed"] is False
    assert payload["smtp_validation"]["recipient_receipt_reference"] == ""
    assert issues
    issue_text = json.dumps([issue.as_dict() for issue in issues], ensure_ascii=False)
    assert "status" in issue_text
    assert "required_true" in issue_text


def test_pilot_gate_c_evidence_accepts_complete_redacted_operator_evidence():
    payload = _complete_pilot_gate_c_evidence()

    assert validate_evidence(payload) == []


def test_pilot_gate_c_evidence_rejects_missing_required_real_workflow_evidence():
    payload = _complete_pilot_gate_c_evidence()
    payload["real_platform_workflow"]["crawl_completed_with_server_profile"] = False
    payload["smtp_validation"]["delivery_log_reference"] = ""
    payload["smtp_validation"]["recipient_receipt_confirmed"] = False
    payload["smtp_validation"]["recipient_receipt_reference"] = ""

    issues = validate_evidence(payload)
    issue_paths = {issue.path for issue in issues}

    assert "real_platform_workflow.crawl_completed_with_server_profile" in issue_paths
    assert "smtp_validation.delivery_log_reference" in issue_paths
    assert "smtp_validation.recipient_receipt_confirmed" in issue_paths
    assert "smtp_validation.recipient_receipt_reference" in issue_paths


def test_pilot_gate_c_evidence_rejects_secret_like_values():
    payload = _complete_pilot_gate_c_evidence()
    payload["redaction"]["notes"] = "leaked key sk-proj-abcdefghijklmnopqrstuvwxyz"
    payload["real_platform_workflow"]["profile_path"] = r"E:\server\profiles\dy\acc_12"

    issues = validate_evidence(payload)
    codes = {issue.code for issue in issues}

    assert "sensitive_value" in codes
    assert "sensitive_key" in codes


def test_pilot_gate_c_evidence_rejects_wrong_schema_version():
    payload = _complete_pilot_gate_c_evidence()
    payload["schema_version"] = "pilot_gate_c_v0"

    issues = validate_evidence(payload)

    assert any(issue.code == "schema_version" and issue.path == "schema_version" for issue in issues)


@pytest.mark.parametrize("placeholder", ["todo", "TBD", "replace_me", "<run id>"])
def test_pilot_gate_c_evidence_rejects_placeholder_values(placeholder):
    payload = _complete_pilot_gate_c_evidence()
    payload["real_platform_workflow"]["run_reference"] = placeholder

    issues = validate_evidence(payload)

    assert any(issue.code == "required_text" and issue.path == "real_platform_workflow.run_reference" for issue in issues)


def test_pilot_gate_c_evidence_requires_all_redaction_surfaces():
    payload = _complete_pilot_gate_c_evidence()
    payload["redaction"]["checked_surfaces"] = ["logs", "reports"]

    issues = validate_evidence(payload)

    assert any(
        issue.code == "redaction_surfaces"
        and "delivery_records" in issue.message
        and "ui_or_api" in issue.message
        for issue in issues
    )


@pytest.mark.parametrize(
    "field_value",
    [
        "Authorization: Bearer abcdefghijklmnop",
        "Cookie: sessionid=abcdef123456",
        "smtp_password=secret123",
        "https://api.openai.com/v1/chat/completions",
        "https://user:pass@proxy.example.com:8080",
        r"C:\Users\Administrator\.env",
        "/app/data/account_profiles/1/dy/acc_12",
    ],
)
def test_pilot_gate_c_evidence_rejects_additional_secret_patterns(field_value):
    payload = _complete_pilot_gate_c_evidence()
    payload["redaction"]["notes"] = field_value

    issues = validate_evidence(payload)

    assert any(issue.code == "sensitive_value" and issue.path == "$.redaction.notes" for issue in issues)


def _phase_5_1_acceptance_evidence_module():
    script_path = Path(__file__).resolve().parents[1] / "scripts" / "phase_5_1_acceptance_evidence.py"
    assert script_path.is_file(), "Phase 5.1 acceptance evidence checker is missing"
    return __import__("scripts.phase_5_1_acceptance_evidence", fromlist=["*"])


def _server_like_validation_module():
    script_path = Path(__file__).resolve().parents[1] / "scripts" / "server_like_validation.py"
    assert script_path.is_file(), "server-like validation script is missing"
    return __import__("scripts.server_like_validation", fromlist=["*"])


def _stub_server_like_validation(module, monkeypatch, data_dir):
    monkeypatch.setattr(module.sys, "argv", ["server_like_validation.py"])
    monkeypatch.setattr(module.tempfile, "mkdtemp", lambda **_kwargs: str(data_dir))
    monkeypatch.setattr(module, "_free_port", lambda _host: 8123)
    monkeypatch.setattr(module, "_start_service", lambda *_args, **_kwargs: object())
    monkeypatch.setattr(module, "_stop_service", lambda _process: None)
    monkeypatch.setattr(module, "_wait_for_health", lambda _base_url: None)
    monkeypatch.setattr(module, "_verify_monitor_page", lambda _base_url: None)
    monkeypatch.setattr(module, "_login", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(module, "_assert_local_login_disabled", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(module, "_create_same_platform_accounts", lambda *_args, **_kwargs: [1, 2])
    monkeypatch.setattr(module, "_verify_profile_paths", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(module, "_verify_locks", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(
        module,
        "_api",
        lambda *_args, **_kwargs: {"accounts": [{"id": 1}, {"id": 2}]},
    )
    monkeypatch.setattr(module, "_verify_headless_browser", lambda _results: None)
    monkeypatch.setattr(module.time, "sleep", lambda _seconds: None)


def test_cr115_server_like_validation_retries_temporary_data_cleanup(tmp_path, monkeypatch, capsys):
    module = _server_like_validation_module()
    data_dir = tmp_path / "server-like-data"
    attempts = 0
    real_rmtree = module.shutil.rmtree

    def flaky_rmtree(path, *args, **kwargs):
        nonlocal attempts
        attempts += 1
        if attempts < 3:
            if kwargs.get("ignore_errors"):
                return None
            raise PermissionError("synthetic transient Windows lock")
        return real_rmtree(path, *args, **kwargs)

    _stub_server_like_validation(module, monkeypatch, data_dir)
    monkeypatch.setattr(module.shutil, "rmtree", flaky_rmtree)

    exit_code = module.main()
    result = json.loads(capsys.readouterr().out)

    assert exit_code == 0
    assert attempts == 3
    assert not data_dir.exists()
    assert any(
        item["name"] == "temporary_data_cleanup" and item["ok"] is True
        for item in result["checks"]
    )


def test_cr115_server_like_validation_reports_permanent_cleanup_failure(tmp_path, monkeypatch, capsys):
    module = _server_like_validation_module()
    data_dir = tmp_path / "server-like-data"
    attempts = 0

    def locked_rmtree(_path, *_args, **_kwargs):
        nonlocal attempts
        attempts += 1
        raise PermissionError("synthetic permanent Windows lock")

    _stub_server_like_validation(module, monkeypatch, data_dir)
    monkeypatch.setattr(module.shutil, "rmtree", locked_rmtree)

    exit_code = module.main()
    result = json.loads(capsys.readouterr().out)
    cleanup = next(item for item in result["checks"] if item["name"] == "temporary_data_cleanup")

    assert exit_code == 1
    assert result["ok"] is False
    assert attempts == 10
    assert cleanup["ok"] is False
    assert str(data_dir) not in cleanup["detail"]
    assert str(data_dir) not in json.dumps(result)
    assert result["data_dir"] == data_dir.name


@pytest.mark.parametrize("retention_mode", ["data_dir", "keep_data"])
def test_cr115_server_like_validation_preserves_requested_data_retention(
    tmp_path,
    monkeypatch,
    capsys,
    retention_mode,
):
    module = _server_like_validation_module()
    data_dir = tmp_path / "server-like-data"
    _stub_server_like_validation(module, monkeypatch, data_dir)
    argv = ["server_like_validation.py", "--keep-data"]
    if retention_mode == "data_dir":
        argv = ["server_like_validation.py", "--data-dir", str(data_dir)]
    monkeypatch.setattr(module.sys, "argv", argv)

    exit_code = module.main()
    result = json.loads(capsys.readouterr().out)

    assert exit_code == 0
    assert result["ok"] is True
    assert data_dir.exists()
    assert result["data_dir"] == str(data_dir.resolve())
    assert not any(item["name"] == "temporary_data_cleanup" for item in result["checks"])


def _phase_5_1_acceptance_action(
    *,
    action,
    trigger_source,
    account_reference,
    runtime_reference,
    resolution_reference,
    attempt_reference,
    environment_digest,
    observed_at,
    provider_mode,
):
    return {
        "action": action,
        "trigger_source": trigger_source,
        "account_reference": account_reference,
        "runtime_reference": runtime_reference,
        "resolution_reference": resolution_reference,
        "attempt_reference": attempt_reference,
        "environment_digest": environment_digest,
        "observed_at": observed_at,
        "provider_mode": provider_mode,
        "browser_source": "playwright_bundled",
        "proxy_effect": "passed",
        "fallback_used": False,
        "mismatch_count": 0,
    }


def _complete_phase_5_1_acceptance_evidence():
    module = _phase_5_1_acceptance_evidence_module()
    payload = module.build_template(now=datetime(2026, 7, 19, 12, 0, tzinfo=timezone.utc))
    qr_digest = "sha256:" + "a" * 64
    cookie_digest = "sha256:" + "b" * 64
    payload["status"] = "passed"
    payload["baseline"].update(
        {
            "commit": "a" * 40,
            "environment_type": "docker",
            "environment_reference": "phase51 acceptance docker run 20260719",
            "operator_reference": "acceptance operator record 20260719",
            "reviewer_reference": "independent reviewer record 20260719",
            "run_window_started_at": "2026-07-19T12:05:00+00:00",
            "run_window_ended_at": "2026-07-19T13:00:00+00:00",
            "checked_at": "2026-07-19T13:01:00+00:00",
        }
    )
    payload["environment"].update(
        {
            "service_started": True,
            "web_admin_login": True,
            "server_browser_owned": True,
            "browser_source": "playwright_bundled",
            "browser_family": "chromium",
            "browser_version": "126.0.0.0",
            "persistent_data_root": True,
            "persistent_profile_root": True,
            "local_window_disabled": True,
            "connect_existing_disabled": True,
            "restart_completed": True,
        }
    )
    payload["accounts"]["qr"].update(
        {
            "platform": "dy",
            "account_reference": "phase51 acceptance qr account 101",
            "acceptance_labeled": True,
        }
    )
    payload["accounts"]["cookie"].update(
        {
            "platform": "dy",
            "account_reference": "phase51 acceptance cookie account 102",
            "acceptance_labeled": True,
        }
    )
    payload["login_actions"] = [
        _phase_5_1_acceptance_action(
            action="qr_login",
            trigger_source="qrcode_login",
            account_reference=payload["accounts"]["qr"]["account_reference"],
            runtime_reference="runtime result qr 201",
            resolution_reference="resolution qr 201",
            attempt_reference="attempt qr 201",
            environment_digest=qr_digest,
            observed_at="2026-07-19T12:10:00+00:00",
            provider_mode="persistent_launch",
        ),
        _phase_5_1_acceptance_action(
            action="cookie_validation",
            trigger_source="cookie_validation",
            account_reference=payload["accounts"]["cookie"]["account_reference"],
            runtime_reference="runtime result cookie 202",
            resolution_reference="resolution cookie 202",
            attempt_reference="attempt cookie 202",
            environment_digest=cookie_digest,
            observed_at="2026-07-19T12:12:00+00:00",
            provider_mode="ephemeral_cookie_validation",
        ),
        _phase_5_1_acceptance_action(
            action="login_check",
            trigger_source="profile_validation",
            account_reference=payload["accounts"]["qr"]["account_reference"],
            runtime_reference="runtime result profile before 203",
            resolution_reference="resolution profile before 203",
            attempt_reference="attempt profile before 203",
            environment_digest=qr_digest,
            observed_at="2026-07-19T12:15:00+00:00",
            provider_mode="persistent_launch",
        ),
        _phase_5_1_acceptance_action(
            action="login_check",
            trigger_source="profile_validation",
            account_reference=payload["accounts"]["qr"]["account_reference"],
            runtime_reference="runtime result profile after 204",
            resolution_reference="resolution profile after 204",
            attempt_reference="attempt profile after 204",
            environment_digest=qr_digest,
            observed_at="2026-07-19T12:25:00+00:00",
            provider_mode="persistent_launch",
        ),
    ]
    payload["restart"].update(
        {
            "restarted_at": "2026-07-19T12:20:00+00:00",
            "lock_timestamp_before": "2026-07-19T12:09:00+00:00",
            "lock_timestamp_after": "2026-07-19T12:09:00+00:00",
            "environment_digest_before": qr_digest,
            "environment_digest_after": qr_digest,
            "profile_login_survived": True,
        }
    )
    payload["proxy"].update(
        {
            "account_bound": True,
            "region_reference": "CN_MAINLAND",
            "browser_region_proof_passed": True,
            "no_task_or_default_override": True,
        }
    )
    payload["crawl_actions"] = []
    for index, (trigger_source, observed_at) in enumerate(
        [
            ("manual", "2026-07-19T12:30:00+00:00"),
            ("scheduler", "2026-07-19T12:40:00+00:00"),
            ("cli_manual", "2026-07-19T12:50:00+00:00"),
        ],
        start=301,
    ):
        item = _phase_5_1_acceptance_action(
            action="crawl",
            trigger_source=trigger_source,
            account_reference=payload["accounts"]["qr"]["account_reference"],
            runtime_reference=f"runtime result crawl {index}",
            resolution_reference=f"resolution crawl {index}",
            attempt_reference=f"attempt crawl {index}",
            environment_digest=qr_digest,
            observed_at=observed_at,
            provider_mode="cdp_launch",
        )
        item.update(
            {
                "run_reference": f"crawl run {index}",
                "terminal_status": "completed",
                "pages": 1,
                "accepted_items": 1,
                "duration_seconds": 60,
            }
        )
        payload["crawl_actions"].append(item)
    payload["runtime_authority"].update(
        {
            "safe_results_collected": True,
            "browser_source_managed": True,
            "profile_reference_matched": True,
            "proxy_policy_account_bound": True,
            "provider_mode_launch_owned": True,
            "mismatch_evidence_empty": True,
            "fallback_used_false": True,
            "child_result_matched_before_ingest": True,
        }
    )
    payload["bounds"].update(
        {
            "serial_execution": True,
            "all_terminal": True,
            "max_pages": 1,
            "max_accepted_items": 10,
            "timeout_seconds": 300,
        }
    )
    payload["redaction"].update(
        {
            "checked_surfaces": [
                "safe_runtime_results",
                "account_api",
                "run_summaries",
                "logs",
                "evidence_file",
            ],
            "no_sensitive_values_found": True,
            "filled_evidence_outside_git": True,
            "evidence_reference": "redaction review phase51 20260719",
        }
    )
    payload["attestations"].update(
        {
            "operator_observed": True,
            "reviewer_cross_checked": True,
            "notes": "all references cross checked against the acceptance deployment",
        }
    )
    return payload


def test_phase_5_1_acceptance_evidence_template_is_incomplete_and_side_effect_free(tmp_path):
    module = _phase_5_1_acceptance_evidence_module()
    target = tmp_path / "phase_5_1_acceptance.json"

    module.write_template(target, now=datetime(2026, 7, 19, 12, 0, tzinfo=timezone.utc))
    payload = json.loads(target.read_text(encoding="utf-8"))
    issues = module.validate_evidence(payload, now=datetime(2026, 7, 19, 13, 5, tzinfo=timezone.utc))

    assert payload["schema_version"] == "phase_5_1_acceptance_v1"
    assert payload["status"] == "incomplete"
    assert payload["baseline"]["template_created_at"] == "2026-07-19T12:00:00+00:00"
    assert issues
    assert {issue.code for issue in issues} >= {"status", "required_true", "required_text"}


def test_phase_5_1_acceptance_evidence_example_matches_generated_template_shape():
    module = _phase_5_1_acceptance_evidence_module()
    example_path = Path(__file__).resolve().parents[1] / "docs" / "phase-5.1-server-like-acceptance.example.json"
    example = json.loads(example_path.read_text(encoding="utf-8"))
    generated = module.build_template(now=datetime(2026, 7, 19, 0, 0, tzinfo=timezone.utc))

    def shape(value):
        if isinstance(value, dict):
            return {key: shape(child) for key, child in value.items()}
        if isinstance(value, list):
            return [shape(value[0])] if value else []
        return type(value)

    assert shape(example) == shape(generated)
    assert example["status"] == "incomplete"
    assert module.validate_evidence(
        example,
        now=datetime(2026, 7, 19, 0, 5, tzinfo=timezone.utc),
    )


def test_phase_5_1_acceptance_evidence_accepts_complete_redacted_evidence():
    module = _phase_5_1_acceptance_evidence_module()
    payload = _complete_phase_5_1_acceptance_evidence()

    assert module.validate_evidence(
        payload,
        now=datetime(2026, 7, 19, 13, 5, tzinfo=timezone.utc),
    ) == []


def test_phase_5_1_acceptance_evidence_pure_helper_keeps_expected_commit_optional():
    module = _phase_5_1_acceptance_evidence_module()
    payload = _complete_phase_5_1_acceptance_evidence()

    assert module.validate_evidence(
        payload,
        expected_commit=None,
        now=datetime(2026, 7, 19, 13, 5, tzinfo=timezone.utc),
    ) == []


def test_phase_5_1_acceptance_evidence_requires_exact_trigger_sources():
    module = _phase_5_1_acceptance_evidence_module()
    payload = _complete_phase_5_1_acceptance_evidence()
    payload["crawl_actions"][2]["trigger_source"] = "cli"

    issues = module.validate_evidence(payload, now=datetime(2026, 7, 19, 13, 5, tzinfo=timezone.utc))

    assert any(issue.code == "trigger_sources" and issue.path == "crawl_actions" for issue in issues)


def test_phase_5_1_acceptance_evidence_requires_exact_action_counts():
    module = _phase_5_1_acceptance_evidence_module()
    payload = _complete_phase_5_1_acceptance_evidence()
    payload["login_actions"].pop()
    payload["crawl_actions"].pop()

    issues = module.validate_evidence(payload, now=datetime(2026, 7, 19, 13, 5, tzinfo=timezone.utc))
    issue_paths = {issue.path for issue in issues if issue.code == "action_count"}

    assert issue_paths == {"login_actions", "crawl_actions"}


def test_phase_5_1_acceptance_evidence_rejects_invalid_chronology():
    module = _phase_5_1_acceptance_evidence_module()
    payload = _complete_phase_5_1_acceptance_evidence()
    payload["restart"]["restarted_at"] = "2026-07-19T12:55:00+00:00"

    issues = module.validate_evidence(payload, now=datetime(2026, 7, 19, 13, 5, tzinfo=timezone.utc))

    assert any(issue.code == "chronology" and issue.path == "restart.restarted_at" for issue in issues)


@pytest.mark.parametrize(
    "lock_timestamp",
    [
        "2026-07-19T12:04:00+00:00",
        "2026-07-19T12:21:00+00:00",
    ],
)
def test_phase_5_1_acceptance_evidence_requires_lock_inside_window_before_restart(lock_timestamp):
    module = _phase_5_1_acceptance_evidence_module()
    payload = _complete_phase_5_1_acceptance_evidence()
    payload["restart"]["lock_timestamp_before"] = lock_timestamp
    payload["restart"]["lock_timestamp_after"] = lock_timestamp

    issues = module.validate_evidence(
        payload,
        now=datetime(2026, 7, 19, 13, 5, tzinfo=timezone.utc),
    )

    assert any(
        issue.code == "chronology" and issue.path == "restart.lock_timestamp_before"
        for issue in issues
    )


def test_phase_5_1_acceptance_evidence_requires_logins_before_restart_and_crawls_after_profile_recheck():
    module = _phase_5_1_acceptance_evidence_module()
    payload = _complete_phase_5_1_acceptance_evidence()
    payload["login_actions"][0]["observed_at"] = "2026-07-19T12:21:00+00:00"
    payload["login_actions"][1]["observed_at"] = "2026-07-19T12:22:00+00:00"
    payload["crawl_actions"][0]["observed_at"] = "2026-07-19T12:18:00+00:00"
    payload["crawl_actions"][1]["observed_at"] = "2026-07-19T12:22:00+00:00"

    issues = module.validate_evidence(
        payload,
        now=datetime(2026, 7, 19, 13, 5, tzinfo=timezone.utc),
    )
    chronology_paths = {issue.path for issue in issues if issue.code == "chronology"}

    assert "login_actions.0.observed_at" in chronology_paths
    assert "login_actions.1.observed_at" in chronology_paths
    assert "crawl_actions.0.observed_at" in chronology_paths
    assert "crawl_actions.1.observed_at" in chronology_paths


def test_phase_5_1_acceptance_evidence_rejects_future_check_timestamp():
    module = _phase_5_1_acceptance_evidence_module()
    payload = _complete_phase_5_1_acceptance_evidence()
    payload["baseline"]["checked_at"] = "2026-07-19T13:06:00+00:00"

    issues = module.validate_evidence(payload, now=datetime(2026, 7, 19, 13, 5, tzinfo=timezone.utc))

    assert any(issue.code == "chronology" and issue.path == "baseline.checked_at" for issue in issues)


def test_phase_5_1_acceptance_evidence_requires_stable_restart_and_trigger_digest():
    module = _phase_5_1_acceptance_evidence_module()
    payload = _complete_phase_5_1_acceptance_evidence()
    payload["restart"]["lock_timestamp_after"] = "2026-07-19T12:20:00+00:00"
    payload["crawl_actions"][1]["environment_digest"] = "sha256:" + "c" * 64

    issues = module.validate_evidence(payload, now=datetime(2026, 7, 19, 13, 5, tzinfo=timezone.utc))
    codes = {issue.code for issue in issues}

    assert "restart_lock" in codes
    assert "stable_digest" in codes


def test_phase_5_1_acceptance_evidence_requires_unique_action_references():
    module = _phase_5_1_acceptance_evidence_module()
    payload = _complete_phase_5_1_acceptance_evidence()
    payload["crawl_actions"][1]["resolution_reference"] = payload["crawl_actions"][0]["resolution_reference"]

    issues = module.validate_evidence(payload, now=datetime(2026, 7, 19, 13, 5, tzinfo=timezone.utc))

    assert any(issue.code == "duplicate_reference" and "resolution_reference" in issue.path for issue in issues)


def test_phase_5_1_acceptance_evidence_enforces_fixed_run_bounds():
    module = _phase_5_1_acceptance_evidence_module()
    payload = _complete_phase_5_1_acceptance_evidence()
    payload["bounds"]["serial_execution"] = False
    payload["crawl_actions"][0].update({"pages": 0, "accepted_items": 0, "duration_seconds": 0})
    payload["crawl_actions"][1].update({"pages": 2, "accepted_items": 11, "duration_seconds": 301})

    issues = module.validate_evidence(payload, now=datetime(2026, 7, 19, 13, 5, tzinfo=timezone.utc))
    issue_paths = {issue.path for issue in issues}

    assert "bounds.serial_execution" in issue_paths
    assert "crawl_actions.0.pages" in issue_paths
    assert "crawl_actions.0.accepted_items" in issue_paths
    assert "crawl_actions.0.duration_seconds" in issue_paths
    assert "crawl_actions.1.pages" in issue_paths
    assert "crawl_actions.1.accepted_items" in issue_paths
    assert "crawl_actions.1.duration_seconds" in issue_paths


def test_phase_5_1_acceptance_evidence_rejects_wrong_provider_modes_and_cookie_crawl():
    module = _phase_5_1_acceptance_evidence_module()
    payload = _complete_phase_5_1_acceptance_evidence()
    payload["login_actions"][0]["provider_mode"] = "cdp_launch"
    payload["crawl_actions"][0]["provider_mode"] = "persistent_launch"
    payload["crawl_actions"][1]["account_reference"] = payload["accounts"]["cookie"]["account_reference"]

    issues = module.validate_evidence(payload, now=datetime(2026, 7, 19, 13, 5, tzinfo=timezone.utc))
    codes = {issue.code for issue in issues}

    assert "login_actions" in codes
    assert "provider_mode" in codes
    assert any(
        issue.code == "account_reference"
        and issue.path == "crawl_actions.1.account_reference"
        and "QR/Profile" in issue.message
        for issue in issues
    )


def test_phase_5_1_acceptance_evidence_requires_all_redaction_surfaces():
    module = _phase_5_1_acceptance_evidence_module()
    payload = _complete_phase_5_1_acceptance_evidence()
    payload["redaction"]["checked_surfaces"] = ["safe_runtime_results", "account_api"]

    issues = module.validate_evidence(payload, now=datetime(2026, 7, 19, 13, 5, tzinfo=timezone.utc))

    assert any(
        issue.code == "redaction_surfaces"
        and "logs" in issue.message
        and "run_summaries" in issue.message
        and "evidence_file" in issue.message
        for issue in issues
    )


@pytest.mark.parametrize(
    "value",
    [
        "Cookie: sessionid=abcdef123456",
        "https://user:pass@proxy.example.com:8080",
        "http://proxy.example.com:8080",
        "地址http://user:pass@proxy.example.test:8080",
        "ws://127.0.0.1:9222/devtools/browser/abc",
        r"C:\server\profiles\dy\acc_12",
        "C:/server/profiles/dy/acc_12",
        r"\\server\share\profiles\dy\acc_12",
        "//server/share/profiles/dy/acc_12",
        "/app/monitor_data/account_profiles/1/dy/acc_12",
        "/srv/monitor/profiles/dy/acc_12",
        "/usr/local/share/monitor/profiles/acc_12",
        "/srv/客户/profile",
        "profile:/srv/data/profile",
        "MONITOR_BROWSER_ENVIRONMENT_PLAN={secret payload}",
        "--user-data-dir=/app/profile --proxy-server=http://proxy.invalid",
    ],
)
def test_phase_5_1_acceptance_evidence_rejects_sensitive_values(value):
    module = _phase_5_1_acceptance_evidence_module()
    payload = _complete_phase_5_1_acceptance_evidence()
    payload["attestations"]["notes"] = value

    issues = module.validate_evidence(payload, now=datetime(2026, 7, 19, 13, 5, tzinfo=timezone.utc))

    assert any(issue.code == "sensitive_value" and issue.path == "$.attestations.notes" for issue in issues)


def test_phase_5_1_acceptance_evidence_allows_normal_slash_delimited_attestation_text():
    module = _phase_5_1_acceptance_evidence_module()
    payload = _complete_phase_5_1_acceptance_evidence()
    payload["attestations"]["notes"] = "操作员/复核员已确认"

    assert module.validate_evidence(
        payload,
        now=datetime(2026, 7, 19, 13, 5, tzinfo=timezone.utc),
    ) == []


def test_phase_5_1_acceptance_evidence_rejects_sensitive_keys_and_placeholders():
    module = _phase_5_1_acceptance_evidence_module()
    payload = _complete_phase_5_1_acceptance_evidence()
    payload["runtime_authority"]["browser_process_argv"] = "redacted"
    payload["accounts"]["cookie"]["password"] = "redacted"
    payload["baseline"]["reviewer_reference"] = "TBD"

    issues = module.validate_evidence(payload, now=datetime(2026, 7, 19, 13, 5, tzinfo=timezone.utc))
    codes = {issue.code for issue in issues}

    assert "sensitive_key" in codes
    assert any(issue.code == "required_text" and issue.path == "baseline.reviewer_reference" for issue in issues)


def test_phase_5_1_acceptance_evidence_rejects_wrong_baseline_and_environment_enums():
    module = _phase_5_1_acceptance_evidence_module()
    payload = _complete_phase_5_1_acceptance_evidence()
    payload["baseline"]["commit"] = "27389a8"
    payload["baseline"]["environment_type"] = "windows_local"
    payload["environment"]["browser_source"] = "diagnostic_auto_detect"
    payload["environment"]["browser_version"] = "126"

    issues = module.validate_evidence(payload, now=datetime(2026, 7, 19, 13, 5, tzinfo=timezone.utc))
    issue_paths = {issue.path for issue in issues}

    assert "baseline.commit" in issue_paths
    assert "baseline.environment_type" in issue_paths
    assert "environment.browser_source" in issue_paths
    assert "environment.browser_version" in issue_paths


def test_phase_5_1_acceptance_evidence_rejects_system_managed_browser_sources():
    module = _phase_5_1_acceptance_evidence_module()
    payload = _complete_phase_5_1_acceptance_evidence()
    payload["environment"]["browser_source"] = "system_managed"
    payload["login_actions"][0]["browser_source"] = "system_managed"

    issues = module.validate_evidence(
        payload,
        now=datetime(2026, 7, 19, 13, 5, tzinfo=timezone.utc),
    )
    issue_paths = {issue.path for issue in issues if issue.code == "enum"}

    assert "environment.browser_source" in issue_paths
    assert "login_actions.0.browser_source" in issue_paths


def test_phase_5_1_acceptance_evidence_requires_distinct_accounts_and_one_browser_source():
    module = _phase_5_1_acceptance_evidence_module()
    payload = _complete_phase_5_1_acceptance_evidence()
    payload["accounts"]["cookie"]["account_reference"] = payload["accounts"]["qr"]["account_reference"]
    payload["login_actions"][1]["account_reference"] = payload["accounts"]["qr"]["account_reference"]
    payload["crawl_actions"][0]["browser_source"] = "explicit"

    issues = module.validate_evidence(
        payload,
        now=datetime(2026, 7, 19, 13, 5, tzinfo=timezone.utc),
    )

    assert any(
        issue.code == "account_reference" and issue.path == "accounts.cookie.account_reference"
        for issue in issues
    )
    assert any(
        issue.code == "browser_source" and issue.path == "crawl_actions.0.browser_source"
        for issue in issues
    )


def test_phase_5_1_acceptance_evidence_rejects_whitespace_only_account_distinction():
    module = _phase_5_1_acceptance_evidence_module()
    payload = _complete_phase_5_1_acceptance_evidence()
    payload["accounts"]["cookie"]["account_reference"] = (
        payload["accounts"]["qr"]["account_reference"] + " "
    )
    payload["login_actions"][1]["account_reference"] = payload["accounts"]["cookie"]["account_reference"]

    issues = module.validate_evidence(
        payload,
        now=datetime(2026, 7, 19, 13, 5, tzinfo=timezone.utc),
    )

    assert any(
        issue.code == "account_reference" and issue.path == "accounts.cookie.account_reference"
        for issue in issues
    )


def test_phase_5_1_acceptance_evidence_requires_explicit_deployed_commit_match():
    module = _phase_5_1_acceptance_evidence_module()
    payload = _complete_phase_5_1_acceptance_evidence()

    issues = module.validate_evidence(
        payload,
        expected_commit="f" * 40,
        now=datetime(2026, 7, 19, 13, 5, tzinfo=timezone.utc),
    )

    assert any(
        issue.code == "baseline"
        and issue.path == "baseline.commit"
        and "expected deployed commit" in issue.message
        for issue in issues
    )


def test_phase_5_1_acceptance_evidence_cli_requires_expected_commit(tmp_path, capsys):
    module = _phase_5_1_acceptance_evidence_module()
    evidence_path = tmp_path / "phase_5_1_acceptance.json"
    evidence_path.write_text(
        json.dumps(_complete_phase_5_1_acceptance_evidence()),
        encoding="utf-8",
    )

    with pytest.raises(SystemExit) as exc_info:
        module.main(["--check", str(evidence_path)])

    assert exc_info.value.code == 2
    assert "--expected-commit must be the exact lowercase 40-character deployed Git commit" in capsys.readouterr().err


@pytest.mark.parametrize(
    "invalid_commit",
    [
        "",
        "a" * 7,
        "A" * 40,
        "g" * 40,
    ],
)
def test_phase_5_1_acceptance_evidence_cli_rejects_invalid_expected_commit(
    tmp_path,
    capsys,
    invalid_commit,
):
    module = _phase_5_1_acceptance_evidence_module()
    evidence_path = tmp_path / "phase_5_1_acceptance.json"
    evidence_path.write_text(
        json.dumps(_complete_phase_5_1_acceptance_evidence()),
        encoding="utf-8",
    )

    with pytest.raises(SystemExit) as exc_info:
        module.main(
            [
                "--check",
                str(evidence_path),
                "--expected-commit",
                invalid_commit,
            ]
        )

    assert exc_info.value.code == 2
    assert "--expected-commit must be the exact lowercase 40-character deployed Git commit" in capsys.readouterr().err


def test_phase_5_1_acceptance_evidence_cli_rejects_deployed_commit_mismatch(tmp_path, capsys):
    module = _phase_5_1_acceptance_evidence_module()
    evidence_path = tmp_path / "phase_5_1_acceptance.json"
    evidence_path.write_text(
        json.dumps(_complete_phase_5_1_acceptance_evidence()),
        encoding="utf-8",
    )

    exit_code = module.main(
        ["--check", str(evidence_path), "--expected-commit", "f" * 40]
    )
    result = json.loads(capsys.readouterr().out)

    assert exit_code == 1
    assert result["ok"] is False
    assert any(
        issue["code"] == "baseline" and issue["path"] == "baseline.commit"
        for issue in result["issues"]
    )


def test_phase_5_1_acceptance_evidence_cli_accepts_matching_deployed_commit(tmp_path, capsys):
    module = _phase_5_1_acceptance_evidence_module()
    payload = _complete_phase_5_1_acceptance_evidence()
    evidence_path = tmp_path / "phase_5_1_acceptance.json"
    evidence_path.write_text(json.dumps(payload), encoding="utf-8")

    exit_code = module.main(
        [
            "--check",
            str(evidence_path),
            "--expected-commit",
            payload["baseline"]["commit"],
        ]
    )
    result = json.loads(capsys.readouterr().out)

    assert exit_code == 0
    assert result["ok"] is True
    assert result["issues"] == []


def test_phase_5_1_acceptance_evidence_cli_returns_structured_missing_file_error(tmp_path, capsys):
    module = _phase_5_1_acceptance_evidence_module()

    exit_code = module.main(
        [
            "--check",
            str(tmp_path / "missing.json"),
            "--expected-commit",
            "a" * 40,
        ]
    )
    result = json.loads(capsys.readouterr().out)

    assert exit_code == 1
    assert result["ok"] is False
    assert result["issues"] == [
        {
            "code": "input",
            "path": "evidence_file",
            "message": "evidence file could not be read",
        }
    ]


@pytest.mark.parametrize(
    ("content", "expected_issue"),
    [
        (
            "{invalid",
            {
                "code": "invalid_json",
                "path": "evidence_file",
                "message": "evidence file must contain valid JSON",
            },
        ),
        (
            "[]",
            {
                "code": "shape",
                "path": "evidence_file",
                "message": "evidence file must contain a JSON object",
            },
        ),
    ],
)
def test_phase_5_1_acceptance_evidence_cli_returns_structured_invalid_content_error(
    tmp_path,
    capsys,
    content,
    expected_issue,
):
    module = _phase_5_1_acceptance_evidence_module()
    evidence_path = tmp_path / "invalid.json"
    evidence_path.write_text(content, encoding="utf-8")

    exit_code = module.main(
        [
            "--check",
            str(evidence_path),
            "--expected-commit",
            "a" * 40,
        ]
    )
    result = json.loads(capsys.readouterr().out)

    assert exit_code == 1
    assert result["ok"] is False
    assert result["issues"] == [expected_issue]


def test_ai_endpoint_builder_handles_v1_and_full_paths():
    assert _build_endpoint("https://api.openai.com", "/v1/chat/completions") == "https://api.openai.com/v1/chat/completions"
    assert _build_endpoint("https://api.openai.com/v1", "/v1/chat/completions") == "https://api.openai.com/v1/chat/completions"
    assert _build_endpoint("https://api.openai.com/v1/chat/completions", "/v1/chat/completions") == "https://api.openai.com/v1/chat/completions"


def test_scheduler_cron_waits_until_today_fire_time_for_new_jobs():
    base = {"frequency": "cron", "cron_expr": "0 9 * * *", "email_time": "09:00", "last_run_at": None}
    assert _is_due(base, datetime(2026, 6, 11, 8, 0, 0)) is False
    assert _is_due(base, datetime(2026, 6, 11, 9, 0, 0)) is True


def test_scheduler_interval_uses_last_run_spacing():
    recent = (datetime.now(timezone.utc) - timedelta(hours=2)).isoformat()
    old = (datetime.now(timezone.utc) - timedelta(hours=7)).isoformat()
    assert _is_due({"frequency": "6h", "email_time": "00:00", "last_run_at": recent}, datetime.now()) is False
    assert _is_due({"frequency": "6h", "email_time": "00:00", "last_run_at": old}, datetime.now()) is True


def test_scheduler_next_run_at_is_visible_for_jobs():
    now = datetime(2026, 6, 12, 8, 0, 0)
    assert next_run_at({"enabled": False, "frequency": "daily", "email_time": "09:00"}, now) is None
    assert next_run_at({"enabled": True, "frequency": "daily", "email_time": "09:00", "last_run_at": None}, now).startswith("2026-06-12T09:00:00")
    assert next_run_at(
        {
            "enabled": True,
            "frequency": "daily",
            "email_time": "09:00",
            "last_run_at": "2026-06-12T09:01:00+00:00",
        },
        datetime(2026, 6, 12, 10, 0, 0),
    ).startswith("2026-06-13T09:00:00")
    assert next_run_at(
        {
            "enabled": True,
            "frequency": "6h",
            "email_time": "00:00",
            "last_run_at": "2026-06-12T06:00:00",
        },
        datetime(2026, 6, 12, 8, 0, 0),
    ).startswith("2026-06-12T12:00:00")


def test_phase_05_schema_foundation_tables_and_columns_exist():
    init_db()
    expected_tables = {
        "workspaces",
        "users",
        "user_sessions",
        "system_settings",
        "audit_logs",
        "resource_locks",
    }
    ownership_tables = {
        "monitor_jobs",
        "social_accounts",
        "proxy_profiles",
        "login_sessions",
        "crawl_runs",
        "raw_contents",
        "raw_comments",
        "ai_evaluations",
        "reports",
        "email_templates",
        "ai_key_profiles",
    }
    with get_conn() as conn:
        tables = {
            row["name"]
            for row in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()
        }
        assert expected_tables <= tables
        workspace = conn.execute("SELECT id, status FROM workspaces WHERE id=1").fetchone()
        assert dict(workspace) == {"id": 1, "status": "active"}
        for table in ownership_tables:
            columns = _table_columns(conn, table)
            assert {"workspace_id", "created_by", "updated_by"} <= columns
        assert {"profile_key", "locked_by_run_id", "locked_at", "lock_expires_at"} <= _table_columns(conn, "social_accounts")
        assert "profile_key" in _table_columns(conn, "login_sessions")
        assert {"timeout_seconds", "deadline_at", "timeout_reason", "account_id", "proxy_id"} <= _table_columns(conn, "crawl_runs")
        assert {"workspace_id", "resource_type", "resource_id", "run_id", "locked_at", "expires_at"} <= _table_columns(conn, "resource_locks")


def test_phase_05_schema_defaults_keep_existing_mvp_records_compatible(tmp_path):
    init_db()
    jobs_snapshot = _snapshot_monitor_jobs()
    snapshots = {
        "reports": _snapshot_table("reports"),
        "crawl_runs": _snapshot_table("crawl_runs"),
        "login_sessions": _snapshot_table("login_sessions"),
        "social_accounts": _snapshot_table("social_accounts"),
        "proxy_profiles": _snapshot_table("proxy_profiles"),
    }
    _clear_monitor_jobs()
    try:
        proxy = save_proxy_profile(
            {
                "name": "华东代理池",
                "provider": "manual",
                "proxy_url": "http://user:password@example.com:8080",
                "status": "standby",
            }
        )
        account = save_social_account(
            {
                "name": "海安律所抖音采集号",
                "platform": "dy",
                "login_type": "qrcode",
                "status": "standby",
                "profile_path": str(tmp_path / "dy_profile"),
                "proxy_id": proxy["id"],
            }
        )
        job = save_job(
            {
                "law_firm_name": "海安律所",
                "aliases": [],
                "exclude_words": [],
                "keywords": ["海安律所投诉"],
                "platforms": ["dy"],
                "recipients": [],
                "account_id": account["id"],
                "proxy_id": proxy["id"],
            }
        )
        session = create_login_session(
            {
                "platform": "dy",
                "account_id": account["id"],
                "login_url": "https://www.douyin.com/",
                "profile_path": account["profile_path"],
            }
        )
        run_id = create_run(job["id"], {"law_firm_name": job["law_firm_name"]})
        finish_run(run_id, "success", {"law_firm_name": job["law_firm_name"]})
        report = create_report(run_id, job, {"law_firm_name": job["law_firm_name"], "platforms": ["dy"]})

        assert get_job(job["id"])["workspace_id"] == 1
        stored_account = get_social_account(account["id"])
        stored_session = get_login_session(session["id"])
        assert stored_account["workspace_id"] == 1
        assert stored_account["profile_key"] == f"1/dy/acc_{account['id']}"
        assert stored_session["workspace_id"] == 1
        assert stored_session["profile_key"] == stored_account["profile_key"]
        assert get_run(run_id)["workspace_id"] == 1
        assert get_report(report["id"])["workspace_id"] == 1
        assert list_jobs()[0]["id"] == job["id"]
        assert list_social_accounts()[0]["id"] == account["id"]
        assert list_login_sessions(limit=1)[0]["id"] == session["id"]
        assert list_runs(limit=1)[0]["id"] == run_id
        assert list_reports(limit=1)[0]["id"] == report["id"]
    finally:
        _restore_table("reports", snapshots["reports"])
        _restore_table("crawl_runs", snapshots["crawl_runs"])
        _restore_table("login_sessions", snapshots["login_sessions"])
        _restore_table("social_accounts", snapshots["social_accounts"])
        _restore_table("proxy_profiles", snapshots["proxy_profiles"])
        _restore_monitor_jobs(jobs_snapshot)


def test_phase_5_1a_account_identity_schema_defaults_and_indexes():
    init_db()
    expected_columns = {
        "environment_region",
        "browser_platform",
        "identity_template",
        "fingerprint_seed",
        "user_agent",
        "timezone",
        "locale",
        "accept_language",
        "screen_width",
        "screen_height",
        "viewport_width",
        "viewport_height",
        "device_scale_factor",
        "is_mobile",
        "has_touch",
        "identity_generator_name",
        "identity_generator_version",
        "identity_environment_version",
        "proxy_region_snapshot",
        "browser_environment_locked_at",
        "browser_environment_lock_reason",
        "requires_relogin",
        "identity_state",
        "identity_runtime_snapshot_json",
    }
    with get_conn() as conn:
        assert expected_columns <= _table_columns(conn, "social_accounts")
        index_rows = {
            row["name"]: dict(row)
            for row in conn.execute("PRAGMA index_list(social_accounts)").fetchall()
        }
        expected_indexes = {
            "idx_social_accounts_identity_state": ["workspace_id", "identity_state"],
            "idx_social_accounts_requires_relogin": ["workspace_id", "requires_relogin"],
            "idx_social_accounts_identity_template": ["workspace_id", "identity_template"],
        }
        for name, expected in expected_indexes.items():
            assert index_rows[name]["unique"] == 0
            assert [
                row["name"]
                for row in conn.execute(f"PRAGMA index_info({name})").fetchall()
            ] == expected

    snapshot = _snapshot_table("social_accounts")
    try:
        now = datetime.now(timezone.utc).isoformat()
        with get_conn() as conn:
            cursor = conn.execute(
                """
                INSERT INTO social_accounts (name, platform, status, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                ("Phase 5.1A Draft Account", "dy", "active", now, now),
            )
            account_id = int(cursor.lastrowid)
        account = get_social_account(account_id)
        expected_defaults = {
            "environment_region": "",
            "browser_platform": "",
            "identity_template": "",
            "fingerprint_seed": "",
            "user_agent": "",
            "timezone": "",
            "locale": "",
            "accept_language": "",
            "screen_width": None,
            "screen_height": None,
            "viewport_width": None,
            "viewport_height": None,
            "device_scale_factor": None,
            "is_mobile": False,
            "has_touch": False,
            "identity_generator_name": "",
            "identity_generator_version": "",
            "identity_environment_version": "",
            "proxy_region_snapshot": "",
            "browser_environment_locked_at": None,
            "browser_environment_lock_reason": "",
            "requires_relogin": False,
            "identity_state": "draft",
            "identity_runtime_snapshot_json": "",
        }
        assert {key: account[key] for key in expected_defaults} == expected_defaults
        assert isinstance(account["requires_relogin"], bool)
        assert isinstance(account["is_mobile"], bool)
        assert isinstance(account["has_touch"], bool)
        assert account["profile_path"] == ""
        listed = next(
            item for item in list_social_accounts() if item["id"] == account["id"]
        )
        assert listed["profile_path"] == ""
        assert listed["requires_relogin"] is False
    finally:
        _restore_table("social_accounts", snapshot)


def test_phase_5_1a_additive_migration_preserves_legacy_account_without_backfill():
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row
    try:
        conn.executescript(
            """
            CREATE TABLE social_accounts (
                id INTEGER PRIMARY KEY,
                workspace_id INTEGER NOT NULL DEFAULT 1,
                name TEXT NOT NULL,
                platform TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'standby',
                profile_key TEXT NOT NULL DEFAULT '',
                profile_path TEXT NOT NULL DEFAULT '',
                proxy_id INTEGER,
                cookies_encrypted TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            INSERT INTO social_accounts (
                id, workspace_id, name, platform, status, profile_key,
                profile_path, proxy_id, cookies_encrypted, created_at, updated_at
            ) VALUES (
                7, 1, 'Legacy Account', 'dy', 'active', '1/dy/acc_7',
                'legacy/profile/path', 3, 'encrypted-value', 'before', 'before'
            );
            """
        )

        database_module._ensure_phase_51_account_identity_schema(conn)
        database_module._ensure_phase_51_account_identity_schema(conn)

        row = dict(conn.execute("SELECT * FROM social_accounts WHERE id=7").fetchone())
        assert row["status"] == "active"
        assert row["profile_key"] == "1/dy/acc_7"
        assert row["profile_path"] == "legacy/profile/path"
        assert row["proxy_id"] == 3
        assert row["cookies_encrypted"] == "encrypted-value"
        assert row["created_at"] == "before"
        assert row["updated_at"] == "before"
        assert row["identity_state"] == "draft"
        assert row["requires_relogin"] == 0
        assert row["environment_region"] == ""
        assert row["identity_template"] == ""
        assert row["fingerprint_seed"] == ""
        assert row["identity_runtime_snapshot_json"] == ""
        assert row["browser_environment_locked_at"] is None
    finally:
        conn.close()


def test_phase_5_1b_template_catalog_expands_all_documented_rows_exactly():
    from api.monitoring.account_identity import (
        IDENTITY_ENVIRONMENT_VERSION,
        IDENTITY_GENERATOR_VERSION,
        IDENTITY_TEMPLATE_CATALOG,
    )

    windows_ua = (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/127.0.6533.17 Safari/537.36"
    )
    expected = {
        "CN_WIN_CHROME_1920": {
            "browser_platform": "windows", "user_agent": windows_ua,
            "screen_width": 1920, "screen_height": 1080,
            "viewport_width": 1920, "viewport_height": 963,
            "device_scale_factor": 1, "is_mobile": False, "has_touch": False,
            "timezone": "Asia/Shanghai", "locale": "zh-CN",
            "accept_language": "zh-CN",
        },
        "CN_WIN_CHROME_1536": {
            "browser_platform": "windows", "user_agent": windows_ua,
            "screen_width": 1536, "screen_height": 864,
            "viewport_width": 1536, "viewport_height": 768,
            "device_scale_factor": 1, "is_mobile": False, "has_touch": False,
            "timezone": "Asia/Shanghai", "locale": "zh-CN",
            "accept_language": "zh-CN",
        },
        "CN_MAC_CHROME_1440": {
            "browser_platform": "macos",
            "user_agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_5) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/127.0.6533.17 Safari/537.36"
            ),
            "screen_width": 1440, "screen_height": 900,
            "viewport_width": 1440, "viewport_height": 789,
            "device_scale_factor": 2, "is_mobile": False, "has_touch": False,
            "timezone": "Asia/Shanghai", "locale": "zh-CN",
            "accept_language": "zh-CN",
        },
        "CN_ANDROID_CHROME": {
            "browser_platform": "android",
            "user_agent": (
                "Mozilla/5.0 (Linux; Android 14; Pixel 7) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/127.0.6533.17 Mobile Safari/537.36"
            ),
            "screen_width": 1080, "screen_height": 2400,
            "viewport_width": 412, "viewport_height": 915,
            "device_scale_factor": 2.625, "is_mobile": True, "has_touch": True,
            "timezone": "Asia/Shanghai", "locale": "zh-CN",
            "accept_language": "zh-CN",
        },
        "HK_DESKTOP_CHROME": {
            "browser_platform": "windows", "user_agent": windows_ua,
            "screen_width": 1920, "screen_height": 1080,
            "viewport_width": 1920, "viewport_height": 963,
            "device_scale_factor": 1, "is_mobile": False, "has_touch": False,
            "timezone": "Asia/Hong_Kong", "locale": "zh-HK",
            "accept_language": "zh-HK",
        },
        "SG_DESKTOP_CHROME": {
            "browser_platform": "windows", "user_agent": windows_ua,
            "screen_width": 1440, "screen_height": 900,
            "viewport_width": 1440, "viewport_height": 789,
            "device_scale_factor": 1, "is_mobile": False, "has_touch": False,
            "timezone": "Asia/Singapore", "locale": "en-SG",
            "accept_language": "en-SG",
        },
    }
    assert tuple(item["identity_template"] for item in IDENTITY_TEMPLATE_CATALOG) == tuple(expected)
    assert IDENTITY_GENERATOR_VERSION == "1.1"
    assert IDENTITY_ENVIRONMENT_VERSION == "v2"
    for item in IDENTITY_TEMPLATE_CATALOG:
        assert {key: item[key] for key in expected[item["identity_template"]]} == expected[item["identity_template"]]


def test_cr116_catalog_matches_pinned_playwright_chromium_metadata():
    import playwright

    from api.monitoring.account_identity import IDENTITY_TEMPLATE_CATALOG

    metadata_path = Path(playwright.__file__).resolve().parent / "driver" / "package" / "browsers.json"
    metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    chromium = next(
        item
        for item in metadata["browsers"]
        if item.get("name") == "chromium" and item.get("installByDefault") is True
    )
    expected_version = str(chromium["browserVersion"])
    catalog_versions = {
        item["user_agent"].split("Chrome/", 1)[1].split(" ", 1)[0]
        for item in IDENTITY_TEMPLATE_CATALOG
    }

    assert catalog_versions == {expected_version}
    assert all(item["accept_language"] == item["locale"] for item in IDENTITY_TEMPLATE_CATALOG)


def test_phase_5_1b_generator_uses_documented_hmac_and_family_filters():
    import hashlib
    import hmac

    from api.monitoring.account_identity import (
        AccountIdentityError,
        IDENTITY_TEMPLATE_CATALOG,
        generate_account_identity,
    )

    salt = b"phase-5.1b-fixed-test-salt"
    generated = generate_account_identity(
        workspace_id=1,
        platform="dy",
        account_id=41,
        proxy_region_snapshot="CN_MAINLAND",
        template_family="auto",
        seed_salt=salt,
    )
    selection_message = b"1|dy|41|CN_MAINLAND|auto"
    selection_seed = hmac.new(salt, selection_message, hashlib.sha256).hexdigest()[:32]
    candidates = [item for item in IDENTITY_TEMPLATE_CATALOG if item["region"] == "CN_MAINLAND"]
    expected_template = candidates[int(selection_seed[:8], 16) % len(candidates)]["identity_template"]
    expected_fingerprint = hmac.new(
        salt,
        f"1|dy|41|CN_MAINLAND|{expected_template}".encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()[:32]
    assert generated["identity_template"] == expected_template
    assert generated["fingerprint_seed"] == expected_fingerprint
    assert generate_account_identity(
        workspace_id=1, platform="dy", account_id=41,
        proxy_region_snapshot="CN_MAINLAND", template_family="auto", seed_salt=salt,
    ) == generated
    assert generate_account_identity(
        workspace_id=1, platform="dy", account_id=42,
        proxy_region_snapshot="CN_MAINLAND", template_family="auto", seed_salt=salt,
    )["fingerprint_seed"] != generated["fingerprint_seed"]

    assert generate_account_identity(
        workspace_id=1, platform="dy", account_id=41,
        proxy_region_snapshot="CN_MAINLAND", template_family="windows_chrome_desktop", seed_salt=salt,
    )["identity_template"] in {"CN_WIN_CHROME_1920", "CN_WIN_CHROME_1536"}
    assert generate_account_identity(
        workspace_id=1, platform="dy", account_id=41,
        proxy_region_snapshot="CN_MAINLAND", template_family="mac_chrome_desktop", seed_salt=salt,
    )["identity_template"] == "CN_MAC_CHROME_1440"
    assert generate_account_identity(
        workspace_id=1, platform="dy", account_id=41,
        proxy_region_snapshot="CN_MAINLAND", template_family="android_chrome", seed_salt=salt,
    )["identity_template"] == "CN_ANDROID_CHROME"
    assert generate_account_identity(
        workspace_id=1, platform="dy", account_id=41,
        proxy_region_snapshot="HK", template_family="auto", seed_salt=salt,
    )["identity_template"] == "HK_DESKTOP_CHROME"
    with pytest.raises(AccountIdentityError) as exc_info:
        generate_account_identity(
            workspace_id=1, platform="dy", account_id=41,
            proxy_region_snapshot="HK", template_family="mac_chrome_desktop", seed_salt=salt,
        )
    assert exc_info.value.reason == "account_identity_contradiction"


def test_phase_5_1b_deployment_seed_derivation_is_domain_separated(monkeypatch):
    import base64
    import hashlib
    import hmac

    from api.monitoring import account_identity

    deployment_key = bytes(range(32))
    encoded_key = base64.urlsafe_b64encode(deployment_key)
    monkeypatch.delenv("MONITOR_ACCOUNT_IDENTITY_SEED_SALT", raising=False)
    monkeypatch.setattr(account_identity, "load_or_create_secret_key", lambda: encoded_key)

    assert account_identity.resolve_account_identity_seed_salt() == hmac.new(
        deployment_key,
        b"MediaCrawler/account-identity/seed/v1",
        hashlib.sha256,
    ).digest()

    monkeypatch.setattr(account_identity, "load_or_create_secret_key", lambda: base64.urlsafe_b64encode(b"short"))
    with pytest.raises(account_identity.AccountIdentityError) as exc_info:
        account_identity.resolve_account_identity_seed_salt()
    assert exc_info.value.reason == "account_identity_missing"


def test_phase_5_1b_validator_fails_closed_for_missing_and_contradictory_fields():
    from api.monitoring.account_identity import (
        AccountIdentityError,
        generate_account_identity,
        validate_account_identity,
    )

    generated = generate_account_identity(
        workspace_id=1,
        platform="dy",
        account_id=51,
        proxy_region_snapshot="CN_MAINLAND",
        template_family="windows_chrome_desktop",
        seed_salt=b"phase-5.1b-validator-salt",
    )
    account = {
        **generated,
        "id": 51,
        "workspace_id": 1,
        "platform": "dy",
        "proxy_id": None,
        "identity_state": "generated",
        "requires_relogin": False,
    }
    validator_salt = b"phase-5.1b-validator-salt"
    assert validate_account_identity(account, seed_salt=validator_salt) == account

    for field, value, reason in (
        ("timezone", "", "account_identity_missing"),
        ("device_scale_factor", None, "account_identity_missing"),
        ("device_scale_factor", 0, "account_identity_contradiction"),
        ("viewport_width", account["screen_width"] + 1, "account_identity_contradiction"),
        ("is_mobile", True, "account_identity_contradiction"),
        ("has_touch", True, "account_identity_contradiction"),
        ("proxy_region_snapshot", "HK", "account_identity_contradiction"),
        ("fingerprint_seed", "not-a-seed", "account_identity_contradiction"),
    ):
        with pytest.raises(AccountIdentityError) as exc_info:
            validate_account_identity({**account, field: value}, seed_salt=validator_salt)
        assert exc_info.value.reason == reason

    with pytest.raises(AccountIdentityError) as exc_info:
        validate_account_identity({**account, "fingerprint_seed": "0" * 32}, seed_salt=validator_salt)
    assert exc_info.value.reason == "account_identity_contradiction"

    with pytest.raises(AccountIdentityError) as exc_info:
        validate_account_identity(
            {**account, "proxy_id": 99},
            bound_proxy_exists=False,
            seed_salt=validator_salt,
        )
    assert exc_info.value.reason == "account_identity_missing"
    with pytest.raises(AccountIdentityError) as exc_info:
        validate_account_identity(
            {**account, "identity_state": "locked"},
            task_proxy_id=7,
            seed_salt=validator_salt,
        )
    assert exc_info.value.reason == "account_identity_locked_proxy_override"
    with pytest.raises(AccountIdentityError) as exc_info:
        validate_account_identity(
            {**account, "identity_state": "active", "requires_relogin": True},
            seed_salt=validator_salt,
        )
    assert exc_info.value.reason == "account_identity_requires_relogin"


def test_cr116_old_identity_catalog_version_requires_explicit_relogin():
    from api.monitoring.account_identity import (
        AccountIdentityError,
        generate_account_identity,
        validate_account_identity,
    )

    generated = generate_account_identity(
        workspace_id=1,
        platform="dy",
        account_id=52,
        proxy_region_snapshot="CN_MAINLAND",
        template_family="windows_chrome_desktop",
        seed_salt=b"cr116-old-catalog-salt",
    )
    account = {
        **generated,
        "id": 52,
        "workspace_id": 1,
        "platform": "dy",
        "proxy_id": None,
        "identity_state": "validated",
        "requires_relogin": False,
        "identity_generator_version": "1.0",
        "identity_environment_version": "v1",
    }

    with pytest.raises(AccountIdentityError) as exc_info:
        validate_account_identity(account, seed_salt=b"cr116-old-catalog-salt")

    assert exc_info.value.reason == "account_identity_requires_relogin"
    assert exc_info.value.fields == (
        "identity_generator_version",
        "identity_environment_version",
    )


def test_phase_5_1b_new_account_generation_is_transactional_and_updates_do_not_backfill(monkeypatch):
    monkeypatch.setenv("MONITOR_ACCOUNT_IDENTITY_SEED_SALT", "phase-5.1b-database-salt")
    init_db()
    snapshot = _snapshot_table("social_accounts")
    try:
        account = save_social_account(
            {
                "name": "Phase 5.1B Generated",
                "platform": "dy",
                "login_type": "qrcode",
                "status": "standby",
                "proxy_region_snapshot": "CN_MAINLAND",
                "identity_template_family": "windows_chrome_desktop",
                "user_agent": "request-controlled-user-agent",
                "fingerprint_seed": "request-controlled-seed",
            }
        )
        stored = get_social_account(account["id"], masked=False)
        assert stored["identity_state"] == "generated"
        assert stored["environment_region"] == "CN_MAINLAND"
        assert stored["proxy_region_snapshot"] == "CN_MAINLAND"
        assert stored["identity_template"] in {"CN_WIN_CHROME_1920", "CN_WIN_CHROME_1536"}
        assert stored["user_agent"] != "request-controlled-user-agent"
        assert stored["fingerprint_seed"] != "request-controlled-seed"
        assert stored["identity_generator_name"] == "mediacrawler_account_identity"
        assert stored["identity_generator_version"] == "1.1"
        assert stored["identity_environment_version"] == "v2"

        with get_conn() as conn:
            conn.execute(
                """
                UPDATE social_accounts SET environment_region='', browser_platform='', identity_template='',
                    fingerprint_seed='', user_agent='', timezone='', locale='', accept_language='',
                    screen_width=NULL, screen_height=NULL, viewport_width=NULL, viewport_height=NULL,
                    device_scale_factor=NULL, is_mobile=0, has_touch=0, identity_generator_name='',
                    identity_generator_version='', identity_environment_version='', proxy_region_snapshot='',
                    identity_state='draft' WHERE id=?
                """,
                (account["id"],),
            )
        save_social_account(
            {
                **get_social_account(account["id"], masked=False),
                "name": "Phase 5.1B Legacy Draft Update",
                "proxy_region_snapshot": "HK",
                "identity_template_family": "auto",
            },
            account["id"],
        )
        legacy = get_social_account(account["id"], masked=False)
        assert legacy["identity_state"] == "draft"
        assert legacy["identity_template"] == ""
        assert legacy["proxy_region_snapshot"] == ""

        with get_conn() as conn:
            before_count = conn.execute("SELECT COUNT(*) FROM social_accounts").fetchone()[0]
        with pytest.raises(ValueError, match="account_identity_contradiction"):
            save_social_account(
                {
                    "name": "Unsupported Region",
                    "platform": "dy",
                    "login_type": "qrcode",
                    "proxy_region_snapshot": "JP",
                }
            )
        with get_conn() as conn:
            after_count = conn.execute("SELECT COUNT(*) FROM social_accounts").fetchone()[0]
        assert after_count == before_count
    finally:
        _restore_table("social_accounts", snapshot)


def test_phase_5_1b_customer_account_view_redacts_identity_internals():
    view = monitor_router._customer_view_social_account(
        {
            "id": 71,
            "identity_state": "generated",
            "identity_template": "CN_WIN_CHROME_1920",
            "environment_region": "CN_MAINLAND",
            "browser_platform": "windows",
            "timezone": "Asia/Shanghai",
            "locale": "zh-CN",
            "identity_generator_name": "mediacrawler_account_identity",
            "identity_generator_version": "1.0",
            "fingerprint_seed": "0123456789abcdef0123456789abcdef",
            "identity_runtime_snapshot_json": '{"profile_path":"C:/secret","cookies":"secret"}',
            "cookies_encrypted": "secret",
            "profile_path": "C:/secret-profile",
            "profile_runtime_path": "C:/secret-runtime",
            "proxy_url": "http://user:pass@example.com:8080",
        }
    )
    for forbidden in (
        "fingerprint_seed",
        "identity_runtime_snapshot_json",
        "cookies_encrypted",
        "profile_path",
        "profile_runtime_path",
        "proxy_url",
    ):
        assert forbidden not in view
    assert view["identity_state"] == "generated"
    assert view["identity_template"] == "CN_WIN_CHROME_1920"
    assert view["identity_generator_version"] == "1.0"


def test_phase_5_1b_account_ui_only_submits_safe_identity_choices_for_new_accounts():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    account_dialog = page[page.index('id="account_dialog"') : page.index('<section id="proxies">')]
    assert 'id="social_account_proxy_region"' in account_dialog
    assert 'id="social_account_identity_template_family"' in account_dialog
    assert '<option value="CN_MAINLAND"' in account_dialog
    assert '<option value="auto"' in account_dialog
    for forbidden_id in (
        "social_account_user_agent",
        "social_account_timezone",
        "social_account_locale",
        "social_account_viewport_width",
        "social_account_screen_width",
        "social_account_device_scale_factor",
        "social_account_is_mobile",
        "social_account_has_touch",
        "social_account_fingerprint_seed",
    ):
        assert f'id="{forbidden_id}"' not in account_dialog
    assert "if(!id || accountIdentityPreloginState(existing)){" in page
    assert "payload.proxy_region_snapshot=val('social_account_proxy_region')" in page
    assert "payload.identity_template_family=val('social_account_identity_template_family')" in page
    draft_login = page[page.index("async function startLoginSessionForDraft()") : page.index("async function startLoginSessionForAccount(")]
    assert "proxy_region_snapshot:val('social_account_proxy_region')" in draft_login
    assert "identity_template_family:val('social_account_identity_template_family')" in draft_login
    assert "renderAccountIdentityStatus(a)" in page
    assert "setAccountIdentityControlsLocked(!(prelogin ||" in page
    assert "setAccountIdentityControlsLocked(false)" in page


def test_phase_5_1b_test_tripwire_blocks_real_account_playwright_by_default(account_identity_tripwire):
    assert account_identity_tripwire["blocked"] is True
    assert account_identity_tripwire["required_flags"] == (
        "TEST_ALLOW_REAL_ACCOUNT_IDENTITY",
        "TEST_ALLOW_REAL_PLATFORM_LOGIN",
    )
    with pytest.raises(AssertionError, match="TEST_ALLOW_REAL_ACCOUNT_IDENTITY"):
        asyncio.run(account_check_module.async_playwright().start())
    with pytest.raises(AssertionError, match="TEST_ALLOW_REAL_PLATFORM_LOGIN"):
        asyncio.run(login_qrcode_module.async_playwright().start())


def test_phase_5_1b_test_tripwire_policy_requires_explicit_opt_ins(account_identity_tripwire):
    policy = account_identity_tripwire["policy"]
    assert policy({}) == {
        "blocked": True,
        "proxy_allowed": False,
    }
    assert policy(
        {
            "TEST_ALLOW_REAL_ACCOUNT_IDENTITY": "true",
            "TEST_ALLOW_REAL_PLATFORM_LOGIN": "true",
        }
    ) == {"blocked": False, "proxy_allowed": False}
    assert policy(
        {
            "TEST_ALLOW_REAL_ACCOUNT_IDENTITY": "true",
            "TEST_ALLOW_REAL_PLATFORM_LOGIN": "true",
            "TEST_ALLOW_REAL_PROXY": "true",
        }
    ) == {"blocked": False, "proxy_allowed": True}


def test_phase_5_1c_identity_lifecycle_locks_only_after_verified_success(monkeypatch):
    from api.monitoring.database import (
        complete_social_account_identity_login,
        list_audit_logs,
        prepare_social_account_identity_login,
    )

    monkeypatch.setenv("MONITOR_ACCOUNT_IDENTITY_SEED_SALT", "phase-5.1c-lifecycle-salt")
    init_db()
    snapshots = {
        "audit_logs": _snapshot_table("audit_logs"),
        "social_accounts": _snapshot_table("social_accounts"),
    }
    try:
        account = save_social_account(
            {"name": "Phase 5.1C QR", "platform": "dy", "login_type": "qrcode"}
        )
        assert account["identity_state"] == "generated"

        prepared = prepare_social_account_identity_login(
            account["id"], trigger_source="qrcode_login", user_id=7
        )
        assert prepared["identity_state"] == "login_in_progress"
        assert prepared["browser_environment_locked_at"] is None

        continued = prepare_social_account_identity_login(
            account["id"],
            trigger_source="qrcode_verification",
            user_id=7,
            allow_prepared_validation=True,
        )
        assert continued["identity_state"] == "login_in_progress"

        active = complete_social_account_identity_login(
            account["id"],
            ok=True,
            trigger_source="qrcode_login",
            lock_reason="qrcode_login_success",
            user_id=7,
        )
        assert active["identity_state"] == "active"
        assert active["requires_relogin"] is False
        assert active["browser_environment_locked_at"]
        assert active["browser_environment_lock_reason"] == "qrcode_login_success"

        audits = [
            item for item in list_audit_logs(limit=30)
            if item["resource_type"] == "social_account" and int(item["resource_id"]) == account["id"]
        ]
        actions = [item["action_type"] for item in reversed(audits)]
        assert "identity_generated" in actions
        assert "identity_validated" in actions
        assert "identity_locked" in actions
        assert "identity_activated" in actions
        allowed_detail_keys = {
            "trigger_source",
            "account_id",
            "workspace_id",
            "platform",
            "identity_template",
            "old_identity_state",
            "new_identity_state",
            "old_proxy_region_snapshot",
            "new_proxy_region_snapshot",
            "old_proxy_id",
            "new_proxy_id",
            "reason",
        }
        for item in audits:
            details = json.loads(item["details_json"])
            assert set(details) == allowed_detail_keys
            visible = json.dumps(details, ensure_ascii=False).lower()
            for forbidden in (
                "fingerprint_seed",
                "cookie",
                "proxy_url",
                "profile_path",
                "identity_runtime_snapshot_json",
                "cdp",
                "novnc",
            ):
                assert forbidden not in visible
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_phase_5_1c_failure_recovery_distinguishes_unlocked_and_maintenance(monkeypatch):
    from api.monitoring.account_identity import AccountIdentityError
    from api.monitoring.database import (
        complete_social_account_identity_login,
        prepare_social_account_identity_login,
    )

    monkeypatch.setenv("MONITOR_ACCOUNT_IDENTITY_SEED_SALT", "phase-5.1c-failure-salt")
    init_db()
    snapshot = _snapshot_table("social_accounts")
    try:
        account = save_social_account(
            {"name": "Phase 5.1C Failure", "platform": "dy", "login_type": "qrcode"}
        )
        prepare_social_account_identity_login(account["id"], trigger_source="qrcode_login")
        failed = complete_social_account_identity_login(
            account["id"],
            ok=False,
            trigger_source="qrcode_login",
            failure_reason="qrcode_failed",
        )
        assert failed["identity_state"] == "validated"
        assert failed["browser_environment_locked_at"] is None

        prepare_social_account_identity_login(account["id"], trigger_source="profile_validation")
        active = complete_social_account_identity_login(
            account["id"],
            ok=True,
            trigger_source="profile_validation",
            lock_reason="profile_validation_success",
        )
        original_lock = active["browser_environment_locked_at"]
        prepare_social_account_identity_login(account["id"], trigger_source="qrcode_maintenance")
        restored = complete_social_account_identity_login(
            account["id"],
            ok=False,
            trigger_source="qrcode_maintenance",
            failure_reason="timeout",
        )
        assert restored["identity_state"] == "active"
        assert restored["browser_environment_locked_at"] == original_lock

        prepare_social_account_identity_login(account["id"], trigger_source="qrcode_maintenance")
        with pytest.raises(AccountIdentityError, match="account_identity_login_conflict"):
            prepare_social_account_identity_login(account["id"], trigger_source="qrcode_duplicate")
        complete_social_account_identity_login(
            account["id"], ok=False, trigger_source="qrcode_maintenance", failure_reason="cancelled"
        )

        with get_conn() as conn:
            conn.execute(
                "UPDATE social_accounts SET identity_state='requires_relogin', requires_relogin=1 WHERE id=?",
                (account["id"],),
            )
        with pytest.raises(AccountIdentityError, match="account_identity_requires_relogin"):
            prepare_social_account_identity_login(account["id"], trigger_source="qrcode_login")
    finally:
        _restore_table("social_accounts", snapshot)


def test_phase_5_1c_configuration_change_and_reset_preserve_login_material(monkeypatch):
    from api.monitoring.account_identity import AccountIdentityError
    from api.monitoring.database import (
        apply_social_account_identity_configuration,
        complete_social_account_identity_login,
        prepare_social_account_identity_login,
        reset_social_account_identity,
    )

    monkeypatch.setenv("MONITOR_ACCOUNT_IDENTITY_SEED_SALT", "phase-5.1c-reset-salt")
    init_db()
    snapshots = {
        "audit_logs": _snapshot_table("audit_logs"),
        "proxy_profiles": _snapshot_table("proxy_profiles"),
        "social_accounts": _snapshot_table("social_accounts"),
    }
    try:
        proxy = save_proxy_profile(
            {"name": "Phase 5.1C SG", "proxy_url": "http://fixture.invalid:8080", "status": "active"}
        )
        account = save_social_account(
            {
                "name": "Phase 5.1C Cookie",
                "platform": "dy",
                "login_type": "cookie",
                "cookies": "sessionid=phase51c_fixture",
            }
        )
        with get_conn() as conn:
            conn.execute(
                """
                UPDATE social_accounts SET platform_account_id='fixture_uid',
                    platform_account_name='Fixture Account' WHERE id=?
                """,
                (account["id"],),
            )
        changed = apply_social_account_identity_configuration(
            account["id"],
            proxy_id=None,
            proxy_region_snapshot="HK",
            template_family="auto",
            trigger_source="admin_prelogin_change",
            user_id=8,
        )
        assert changed["identity_state"] == "validated"
        assert changed["proxy_region_snapshot"] == "HK"
        assert changed["identity_template"] == "HK_DESKTOP_CHROME"

        prepare_social_account_identity_login(account["id"], trigger_source="cookie_validation")
        with pytest.raises(AccountIdentityError, match="account_identity_login_conflict"):
            apply_social_account_identity_configuration(
                account["id"],
                proxy_id=proxy["id"],
                proxy_region_snapshot="SG",
                template_family="auto",
                trigger_source="admin_change_during_login",
            )
        active = complete_social_account_identity_login(
            account["id"],
            ok=True,
            trigger_source="cookie_validation",
            lock_reason="cookie_validation_success",
        )
        locked_template = active["identity_template"]
        locked_proxy_id = active["proxy_id"]
        marked = apply_social_account_identity_configuration(
            account["id"],
            proxy_id=proxy["id"],
            proxy_region_snapshot="SG",
            template_family="auto",
            trigger_source="admin_locked_change",
            user_id=8,
        )
        assert marked["identity_state"] == "requires_relogin"
        assert marked["requires_relogin"] is True
        assert marked["identity_template"] == locked_template
        assert marked["proxy_id"] == locked_proxy_id
        assert marked["proxy_region_snapshot"] == "HK"

        with get_conn() as conn:
            conn.execute("UPDATE social_accounts SET locked_by_run_id=999 WHERE id=?", (account["id"],))
        with pytest.raises(AccountIdentityError, match="account_identity_reset_blocked"):
            reset_social_account_identity(
                account["id"],
                proxy_id=proxy["id"],
                proxy_region_snapshot="SG",
                template_family="auto",
                user_id=8,
            )
        with get_conn() as conn:
            conn.execute("UPDATE social_accounts SET locked_by_run_id=NULL WHERE id=?", (account["id"],))

        reset = reset_social_account_identity(
            account["id"],
            proxy_id=proxy["id"],
            proxy_region_snapshot="SG",
            template_family="auto",
            user_id=8,
        )
        stored = get_social_account(account["id"], masked=False)
        assert reset["identity_state"] == "validated"
        assert reset["status"] == "standby"
        assert reset["proxy_id"] == proxy["id"]
        assert reset["proxy_region_snapshot"] == "SG"
        assert reset["identity_template"] == "SG_DESKTOP_CHROME"
        assert reset["browser_environment_locked_at"] is None
        assert reset["browser_environment_lock_reason"] == ""
        assert reset["requires_relogin"] is False
        assert stored["cookies"] == "sessionid=phase51c_fixture"
        assert stored["profile_key"] == account["profile_key"]
        assert stored["platform_account_id"] == "fixture_uid"
        assert stored["platform_account_name"] == "Fixture Account"
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_phase_5_1c_cookie_check_owns_prepare_and_completion(monkeypatch):
    monkeypatch.setenv("MONITOR_ACCOUNT_IDENTITY_SEED_SALT", "phase-5.1c-check-salt")
    init_db()
    snapshots = {
        "audit_logs": _snapshot_table("audit_logs"),
        "social_accounts": _snapshot_table("social_accounts"),
    }
    try:
        account = save_social_account(
            {
                "name": "Phase 5.1C Checked Cookie",
                "platform": "dy",
                "login_type": "cookie",
                "cookies": "sessionid=fake_cookie",
            }
        )

        async def fake_cookie_check(_account, _timeout_ms):
            assert get_social_account(account["id"])["identity_state"] == "login_in_progress"
            return {"ok": True, "status": "valid", "message": "ok", "identity": {}}

        monkeypatch.setattr(account_check_module, "_check_cookie_account", fake_cookie_check)
        result = asyncio.run(account_check_module.check_social_account_login(account["id"], actor_id=9))
        assert result["ok"] is True
        assert result["account"]["identity_state"] == "active"
        assert result["account"]["browser_environment_lock_reason"] == "cookie_validation_success"
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_phase_5_1c_qr_verification_reuses_prepared_identity(monkeypatch):
    from api.monitoring.database import (
        complete_social_account_identity_login,
        prepare_social_account_identity_login,
    )

    monkeypatch.setenv("MONITOR_ACCOUNT_IDENTITY_SEED_SALT", "phase-5.1c-qr-salt")
    init_db()
    snapshots = {
        "login_sessions": _snapshot_table("login_sessions"),
        "social_accounts": _snapshot_table("social_accounts"),
    }
    try:
        account = save_social_account(
            {"name": "Phase 5.1C QR Verify", "platform": "dy", "login_type": "qrcode"}
        )
        prepare_social_account_identity_login(account["id"], trigger_source="qrcode_login")
        session = create_login_session(
            {"platform": "dy", "account_id": account["id"], "status": "success"}
        )
        captured = {}

        async def fake_check(account_id, timeout_ms=15000, allow_draft=False, identity_prepared=False, actor_id=None):
            captured.update(
                account_id=account_id,
                allow_draft=allow_draft,
                identity_prepared=identity_prepared,
            )
            completed = complete_social_account_identity_login(
                account_id,
                ok=True,
                trigger_source="qrcode_login",
                lock_reason="qrcode_login_success",
            )
            return {"ok": True, "account": completed}

        monkeypatch.setattr(monitor_router, "check_social_account_login", fake_check)
        verified_session, account_status = asyncio.run(monitor_router._verify_successful_login_session(session))
        assert verified_session["status"] == "success"
        assert account_status["identity_state"] == "active"
        assert captured == {
            "account_id": account["id"],
            "allow_draft": True,
            "identity_prepared": True,
        }
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_phase_5_1c_qr_start_prepares_and_delete_recovers_identity(monkeypatch):
    monkeypatch.setenv("MONITOR_ACCOUNT_IDENTITY_SEED_SALT", "phase-5.1c-qr-route-salt")
    init_db()
    snapshots = {
        "login_sessions": _snapshot_table("login_sessions"),
        "social_accounts": _snapshot_table("social_accounts"),
    }
    seen = {}
    try:
        account = save_social_account(
            {"name": "Phase 5.1C QR Route", "platform": "dy", "login_type": "qrcode"}
        )

        monkeypatch.setattr(monitor_router, "_open_login_window_for_command", lambda *args: {})

        async def fake_start(session_id, platform, command):
            seen["start_state"] = get_social_account(account["id"])["identity_state"]
            return {
                "ok": True,
                "status": "waiting_qrcode",
                "qr_image": "data:image/png;base64,phase51c",
                "message": "请扫码登录",
            }

        async def fake_close(session_id):
            seen["closed_session_id"] = int(session_id)

        monkeypatch.setattr(monitor_router, "start_qrcode_login_session_with_profile", fake_start)
        monkeypatch.setattr(monitor_router, "close_qrcode_login_session", fake_close)

        created = asyncio.run(
            monitor_router.create_platform_login_session(
                {"platform": "dy", "account_id": account["id"]},
                {"id": 13, "workspace_id": 1},
            )
        )

        assert seen["start_state"] == "login_in_progress"
        assert created["account_status"]["identity_state"] == "login_in_progress"
        assert get_social_account(account["id"])["identity_state"] == "login_in_progress"

        session_id = int(created["session"]["id"])
        assert asyncio.run(
            monitor_router.remove_login_session(
                session_id,
                {"id": 13, "workspace_id": 1},
            )
        ) == {"ok": True}
        assert seen["closed_session_id"] == session_id
        assert get_login_session(session_id) is None
        assert get_social_account(account["id"])["identity_state"] == "validated"
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_phase_5_1c_visible_browser_prepares_then_recovers_sync_failure(monkeypatch):
    monkeypatch.setenv("MONITOR_ACCOUNT_IDENTITY_SEED_SALT", "phase-5.1c-visible-route-salt")
    init_db()
    snapshot = _snapshot_table("social_accounts")
    seen = {}
    try:
        account = save_social_account(
            {"name": "Phase 5.1C Visible", "platform": "dy", "login_type": "qrcode"}
        )

        async def fake_command(platform, payload, **kwargs):
            seen["command_state"] = get_social_account(account["id"])["identity_state"]
            return {
                "platform": platform,
                "account_id": account["id"],
                "profile_key": account["profile_key"],
                "profile_path": account["profile_path"],
            }

        def fake_open(command):
            seen["open_state"] = get_social_account(account["id"])["identity_state"]
            raise RuntimeError("fixture browser start failed")

        monkeypatch.setattr(monitor_router, "_login_browser_command_for_payload", fake_command)
        monkeypatch.setattr(monitor_router, "open_login_browser_with_command", fake_open)

        with pytest.raises(HTTPException) as exc_info:
            asyncio.run(
                monitor_router.platform_login_browser(
                    "dy",
                    {"account_id": account["id"]},
                    {"id": 14, "workspace_id": 1},
                )
            )

        assert exc_info.value.status_code == 500
        assert seen == {
            "command_state": "login_in_progress",
            "open_state": "login_in_progress",
        }
        assert get_social_account(account["id"])["identity_state"] == "validated"
    finally:
        _restore_table("social_accounts", snapshot)


def test_phase_5_1c_admin_check_continues_persisted_login_owner(monkeypatch):
    from api.monitoring.database import prepare_social_account_identity_login

    monkeypatch.setenv("MONITOR_ACCOUNT_IDENTITY_SEED_SALT", "phase-5.1c-admin-check-salt")
    init_db()
    snapshot = _snapshot_table("social_accounts")
    captured = {}
    try:
        account = save_social_account(
            {"name": "Phase 5.1C Admin Check", "platform": "dy", "login_type": "qrcode"}
        )
        prepare_social_account_identity_login(account["id"], trigger_source="visible_browser_login")

        async def fake_check(
            account_id,
            timeout_ms=15000,
            allow_draft=False,
            identity_prepared=False,
            actor_id=None,
        ):
            captured.update(
                account_id=account_id,
                allow_draft=allow_draft,
                identity_prepared=identity_prepared,
                actor_id=actor_id,
            )
            return {
                "ok": False,
                "status": "invalid",
                "message": "fixture",
                "account": get_social_account(account_id),
            }

        monkeypatch.setattr(monitor_router, "check_social_account_login", fake_check)

        asyncio.run(
            monitor_router.check_social_account(
                account["id"],
                {"id": 15, "workspace_id": 1},
            )
        )

        assert captured == {
            "account_id": account["id"],
            "allow_draft": False,
            "identity_prepared": True,
            "actor_id": 15,
        }
    finally:
        _restore_table("social_accounts", snapshot)


def test_phase_5_1c_update_and_reset_routes_use_lifecycle_authority(monkeypatch):
    from api.monitoring.database import (
        complete_social_account_identity_login,
        prepare_social_account_identity_login,
    )

    monkeypatch.setenv("MONITOR_ACCOUNT_IDENTITY_SEED_SALT", "phase-5.1c-route-reset-salt")
    init_db()
    snapshots = {
        "audit_logs": _snapshot_table("audit_logs"),
        "social_accounts": _snapshot_table("social_accounts"),
    }
    try:
        account = save_social_account(
            {"name": "Phase 5.1C Route Reset", "platform": "dy", "login_type": "qrcode"}
        )
        prepare_social_account_identity_login(account["id"], trigger_source="profile_validation")
        active = complete_social_account_identity_login(
            account["id"],
            ok=True,
            trigger_source="profile_validation",
            lock_reason="profile_validation_success",
        )
        original_template = active["identity_template"]

        updated = asyncio.run(
            monitor_router.update_social_account(
                account["id"],
                {
                    "name": account["name"],
                    "platform": "dy",
                    "login_type": "qrcode",
                    "status": "active",
                    "proxy_id": None,
                    "proxy_region_snapshot": "HK",
                    "identity_template_family": "auto",
                    "notes": "",
                    "last_error": "",
                },
                {"id": 16, "workspace_id": 1},
            )
        )["account"]

        assert updated["identity_state"] == "requires_relogin"
        assert updated["requires_relogin"] is True
        assert updated["identity_template"] == original_template
        assert updated["proxy_region_snapshot"] == "CN_MAINLAND"

        with get_conn() as conn:
            conn.execute(
                "UPDATE social_accounts SET locked_by_run_id=999 WHERE id=?",
                (account["id"],),
            )
        with pytest.raises(HTTPException) as exc_info:
            asyncio.run(
                monitor_router.reset_social_account_identity_route(
                    account["id"],
                    {
                        "proxy_id": None,
                        "proxy_region_snapshot": "HK",
                        "identity_template_family": "auto",
                    },
                    {"id": 16, "workspace_id": 1},
                )
            )
        assert exc_info.value.status_code == 409
        assert get_social_account(account["id"])["identity_state"] == "requires_relogin"

        with get_conn() as conn:
            conn.execute(
                "UPDATE social_accounts SET locked_by_run_id=NULL WHERE id=?",
                (account["id"],),
            )
        reset = asyncio.run(
            monitor_router.reset_social_account_identity_route(
                account["id"],
                {
                    "proxy_id": None,
                    "proxy_region_snapshot": "HK",
                    "identity_template_family": "auto",
                },
                {"id": 16, "workspace_id": 1},
            )
        )["account"]

        assert reset["identity_state"] == "validated"
        assert reset["status"] == "standby"
        assert reset["proxy_region_snapshot"] == "HK"
        assert reset["identity_template"] == "HK_DESKTOP_CHROME"
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_phase_5_1c_qr_and_profile_checks_use_distinct_lock_reasons(monkeypatch):
    from api.monitoring.database import prepare_social_account_identity_login

    monkeypatch.setenv("MONITOR_ACCOUNT_IDENTITY_SEED_SALT", "phase-5.1c-lock-reason-salt")
    init_db()
    snapshot = _snapshot_table("social_accounts")
    try:
        async def fake_profile_check(account, timeout_ms):
            return {"ok": True, "status": "valid", "message": "ok", "identity": {}}

        monkeypatch.setattr(account_check_module, "_check_profile_account", fake_profile_check)

        qr_account = save_social_account(
            {"name": "Phase 5.1C QR Reason", "platform": "dy", "login_type": "qrcode"}
        )
        prepare_social_account_identity_login(qr_account["id"], trigger_source="qrcode_login")
        qr_result = asyncio.run(
            account_check_module.check_social_account_login(
                qr_account["id"],
                allow_draft=True,
                identity_prepared=True,
            )
        )
        assert qr_result["account"]["browser_environment_lock_reason"] == "qrcode_login_success"

        profile_account = save_social_account(
            {"name": "Phase 5.1C Profile Reason", "platform": "dy", "login_type": "qrcode"}
        )
        profile_result = asyncio.run(
            account_check_module.check_social_account_login(profile_account["id"])
        )
        assert profile_result["account"]["browser_environment_lock_reason"] == "profile_validation_success"
    finally:
        _restore_table("social_accounts", snapshot)


@pytest.mark.parametrize("route_name", ["submit", "request"])
def test_phase_5_1c_verification_code_terminal_failure_recovers_identity(route_name, monkeypatch):
    from api.monitoring.database import (
        complete_social_account_identity_login,
        prepare_social_account_identity_login,
    )

    monkeypatch.setenv("MONITOR_ACCOUNT_IDENTITY_SEED_SALT", "phase-5.1c-code-route-salt")
    init_db()
    snapshots = {
        "login_sessions": _snapshot_table("login_sessions"),
        "social_accounts": _snapshot_table("social_accounts"),
    }
    try:
        account = save_social_account(
            {"name": f"Phase 5.1C Code {route_name}", "platform": "dy", "login_type": "qrcode"}
        )
        prepare_social_account_identity_login(account["id"], trigger_source="qrcode_login")
        session = create_login_session(
            {
                "platform": "dy",
                "account_id": account["id"],
                "status": "needs_verification",
            }
        )

        async def fake_code_action(*args, **kwargs):
            return {
                "active": False,
                "success": False,
                "status": "platform_error",
                "message": "fixture verification failed",
            }

        async def fake_check(
            account_id,
            timeout_ms=15000,
            allow_draft=False,
            identity_prepared=False,
            actor_id=None,
        ):
            assert identity_prepared is True
            completed = complete_social_account_identity_login(
                account_id,
                ok=False,
                trigger_source="qrcode_login",
                failure_reason="platform_error",
            )
            return {
                "ok": False,
                "status": "platform_error",
                "message": "fixture verification failed",
                "account": completed,
            }

        monkeypatch.setattr(monitor_router, "check_social_account_login", fake_check)
        monkeypatch.setattr(monitor_router, "list_platform_status", lambda: [])
        if route_name == "submit":
            monkeypatch.setattr(monitor_router, "submit_qrcode_login_verification_code", fake_code_action)
            result = asyncio.run(
                monitor_router.submit_login_session_verification_code(
                    session["id"],
                    {"code": "123456"},
                    {"id": 17, "workspace_id": 1},
                )
            )
        else:
            monkeypatch.setattr(monitor_router, "request_qrcode_login_verification_code", fake_code_action)
            result = asyncio.run(
                monitor_router.request_login_session_verification_code(
                    session["id"],
                    {"id": 17, "workspace_id": 1},
                )
            )

        assert result["session"]["status"] == "platform_error"
        assert get_social_account(account["id"])["identity_state"] == "validated"
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_phase_5_1c_confirm_draft_cannot_replace_locked_proxy(monkeypatch):
    from api.monitoring.database import (
        complete_social_account_identity_login,
        prepare_social_account_identity_login,
    )

    monkeypatch.setenv("MONITOR_ACCOUNT_IDENTITY_SEED_SALT", "phase-5.1c-confirm-salt")
    init_db()
    snapshots = {
        "proxy_profiles": _snapshot_table("proxy_profiles"),
        "social_accounts": _snapshot_table("social_accounts"),
    }
    try:
        proxy = save_proxy_profile(
            {"name": "Phase 5.1C Confirm Proxy", "proxy_url": "http://fixture.invalid:8080"}
        )
        draft = database_module.create_draft_social_account(
            {"name": "Phase 5.1C Confirm", "platform": "dy"}
        )
        prepare_social_account_identity_login(draft["id"], trigger_source="qrcode_login")
        complete_social_account_identity_login(
            draft["id"],
            ok=True,
            trigger_source="qrcode_login",
            lock_reason="qrcode_login_success",
        )

        confirmed = asyncio.run(
            monitor_router.confirm_account(
                draft["id"],
                {
                    "name": draft["name"],
                    "login_type": "qrcode",
                    "status": "active",
                    "proxy_id": proxy["id"],
                },
                {"id": 18, "workspace_id": 1},
            )
        )["account"]

        assert confirmed["is_draft"] is False
        assert confirmed["identity_state"] == "requires_relogin"
        assert confirmed["requires_relogin"] is True
        assert confirmed["proxy_id"] is None
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_phase_5_1c_reset_rejects_persisted_resetting_state(monkeypatch):
    from api.monitoring.account_identity import AccountIdentityError
    from api.monitoring.database import reset_social_account_identity

    monkeypatch.setenv("MONITOR_ACCOUNT_IDENTITY_SEED_SALT", "phase-5.1c-resetting-salt")
    init_db()
    snapshot = _snapshot_table("social_accounts")
    try:
        account = save_social_account(
            {"name": "Phase 5.1C Resetting", "platform": "dy", "login_type": "qrcode"}
        )
        with get_conn() as conn:
            conn.execute(
                "UPDATE social_accounts SET identity_state='resetting' WHERE id=?",
                (account["id"],),
            )
        with pytest.raises(AccountIdentityError, match="account_identity_reset_blocked"):
            reset_social_account_identity(
                account["id"],
                proxy_id=None,
                proxy_region_snapshot="CN_MAINLAND",
                template_family="auto",
            )
        assert get_social_account(account["id"])["identity_state"] == "resetting"
    finally:
        _restore_table("social_accounts", snapshot)


def test_phase_5_1c_locked_proxy_clear_audit_records_explicit_null(monkeypatch):
    from api.monitoring.account_identity import identity_template_family
    from api.monitoring.database import (
        apply_social_account_identity_configuration,
        complete_social_account_identity_login,
        list_audit_logs,
        prepare_social_account_identity_login,
    )

    monkeypatch.setenv("MONITOR_ACCOUNT_IDENTITY_SEED_SALT", "phase-5.1c-audit-null-salt")
    init_db()
    snapshots = {
        "audit_logs": _snapshot_table("audit_logs"),
        "proxy_profiles": _snapshot_table("proxy_profiles"),
        "social_accounts": _snapshot_table("social_accounts"),
    }
    try:
        proxy = save_proxy_profile(
            {"name": "Phase 5.1C Audit Proxy", "proxy_url": "http://fixture.invalid:8080"}
        )
        account = save_social_account(
            {
                "name": "Phase 5.1C Audit Null",
                "platform": "dy",
                "login_type": "qrcode",
                "proxy_id": proxy["id"],
            }
        )
        prepare_social_account_identity_login(account["id"], trigger_source="profile_validation")
        complete_social_account_identity_login(
            account["id"],
            ok=True,
            trigger_source="profile_validation",
            lock_reason="profile_validation_success",
        )
        apply_social_account_identity_configuration(
            account["id"],
            proxy_id=None,
            proxy_region_snapshot=account["proxy_region_snapshot"],
            template_family=identity_template_family(account["identity_template"]),
            trigger_source="admin_locked_proxy_clear",
        )

        audit = next(
            item
            for item in list_audit_logs(limit=30)
            if item["action_type"] == "identity_requires_relogin"
            and int(item["resource_id"]) == account["id"]
        )
        details = json.loads(audit["details_json"])
        assert details["old_proxy_id"] == proxy["id"]
        assert details["new_proxy_id"] is None
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_cr113_qr_draft_route_forwards_safe_identity_choices(monkeypatch):
    captured = {}

    def fake_create_draft(payload):
        captured.update(payload)
        raise ValueError("stop after capture")

    monkeypatch.setattr(monitor_router, "create_draft_social_account", fake_create_draft)
    with pytest.raises(HTTPException) as exc_info:
        asyncio.run(
            monitor_router.create_platform_login_session(
                {
                    "name": "CR-113 Draft",
                    "platform": "dy",
                    "proxy_id": "",
                    "proxy_region_snapshot": "HK",
                    "identity_template_family": "windows_chrome_desktop",
                    "user_agent": "must-not-forward",
                },
                {"id": 10, "workspace_id": 1},
            )
        )
    assert exc_info.value.status_code == 400
    assert captured["proxy_region_snapshot"] == "HK"
    assert captured["identity_template_family"] == "windows_chrome_desktop"
    assert "user_agent" not in captured


def test_phase_5_1c_account_ui_exposes_safe_reset_relogin_flow_only():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    account_dialog = page[page.index('id="account_dialog"') : page.index('<section id="proxies">')]
    for marker in (
        'id="social_account_identity_status"',
        'id="account_identity_change_button"',
        'id="account_identity_reset_button"',
        "function accountIdentityStateLabel(",
        "function beginAccountIdentityChange(",
        "async function resetCurrentAccountIdentity(",
        "'/social-accounts/'+accountId+'/identity/reset'",
    ):
        assert marker in account_dialog or marker in page
    assert "account_identity_requires_relogin" in page
    assert "requires_relogin" in page
    bulk_toolbar = page[
        page.index("function updateAccountBulkToolbar(") :
        page.index("function accountLedgerTable(")
    ]
    identity_controls = page[
        page.index("function setAccountIdentityControlsLocked(") :
        page.index("function renderAccountIdentityStatus(")
    ]
    assert "const hasBlocked=" in bulk_toolbar
    assert "if(checkBtn && hasBlocked)" in bulk_toolbar
    assert "hasBlocked" not in identity_controls
    for forbidden in ("fingerprint_seed", "identity_runtime_snapshot_json", "profile_runtime_path"):
        assert forbidden not in account_dialog


def _phase_5_1d_persisted_account(*, state="generated", proxy_id=41, requires_relogin=False):
    from api.monitoring.account_identity import generate_account_identity

    account_id = 5101
    return {
        **generate_account_identity(
            workspace_id=1,
            platform="dy",
            account_id=account_id,
            proxy_region_snapshot="CN_MAINLAND",
            template_family="windows_chrome_desktop",
        ),
        "id": account_id,
        "workspace_id": 1,
        "platform": "dy",
        "profile_key": f"1/dy/acc_{account_id}",
        "profile_path": r"C:\legacy\shared-dy-profile",
        "proxy_id": proxy_id,
        "identity_state": state,
        "requires_relogin": requires_relogin,
        "is_draft": False,
    }


def _phase_5_1d_active_proxy(proxy_id=41, *, status="active"):
    return {
        "id": proxy_id,
        "workspace_id": 1,
        "status": status,
        "proxy_url": "http://phase51d-user:phase51d-pass@proxy.invalid:8080",
    }


def _phase_5_1d_safe_snapshot():
    return {
        "contract_version": 1,
        "resolution_id": "resolution-5101",
        "attempt_id": "attempt-5101",
        "action": "login_check",
        "trigger_source": "profile_validation",
        "account": {
            "workspace_id": 1,
            "account_id": 5101,
            "platform": "dy",
            "identity_state": "validated",
        },
        "browser": {
            "family": "chromium",
            "version": "126.0.6478.183",
            "source": "playwright_bundled",
        },
        "profile": {"profile_key": "1/dy/acc_5101", "mode": "persistent"},
        "proxy": {
            "policy": "account_bound",
            "proxy_id": 41,
            "region": "CN_MAINLAND",
            "effect_proof": "passed",
        },
        "requested": {
            "identity_template": "CN_WIN_CHROME_1920",
            "browser_platform": "windows",
            "user_agent": "Mozilla/5.0 Chrome/126.0.6478.183",
            "timezone": "Asia/Shanghai",
            "locale": "zh-CN",
            "accept_language": "zh-CN,zh;q=0.9",
            "screen_width": 1920,
            "screen_height": 1080,
            "viewport_width": 1920,
            "viewport_height": 963,
            "device_scale_factor": 1.0,
            "is_mobile": False,
            "has_touch": False,
            "proxy_region_snapshot": "CN_MAINLAND",
        },
        "effective": {
            "user_agent": "Mozilla/5.0 Chrome/126.0.6478.183",
            "timezone": "Asia/Shanghai",
            "locale": "zh-CN",
            "accept_language": "zh-CN,zh;q=0.9",
            "screen_width": 1920,
            "screen_height": 1080,
            "viewport_width": 1920,
            "viewport_height": 963,
            "device_scale_factor": 1.0,
            "is_mobile": False,
            "has_touch": False,
            "proxy_region_snapshot": "CN_MAINLAND",
        },
        "provider": {
            "name": "playwright",
            "mode": "persistent_launch",
            "version": "1.45.0",
        },
        "probes": {
            "navigator_user_agent": "Mozilla/5.0 Chrome/126.0.6478.183",
            "navigator_language": "zh-CN",
            "navigator_languages": ["zh-CN", "zh"],
            "timezone": "Asia/Shanghai",
            "screen_width": 1920,
            "screen_height": 1080,
            "viewport_width": 1920,
            "viewport_height": 963,
            "device_scale_factor": 1.0,
            "max_touch_points": 0,
            "is_mobile": False,
            "webdriver": True,
        },
        "unsupported_fields": ["canvas", "webgl", "fonts", "plugins"],
        "mismatch_evidence": [],
        "fallback_used": False,
        "ok": True,
        "reason": "",
        "validated_at": "2026-07-19T04:00:00+00:00",
    }


def test_phase_5_1d_provider_resolves_exact_account_environment(tmp_path, monkeypatch):
    from dataclasses import FrozenInstanceError

    from api.monitoring import account_environment
    from api.monitoring.browser_environment_provider import resolve_account_browser_environment
    from tools.browser_environment import (
        browser_context_options,
        browser_environment_plan_from_json,
        browser_environment_plan_to_json,
    )

    profile_root = (tmp_path / "account-profiles").resolve()
    explicit_browser = tmp_path / "managed-chromium.exe"
    bundled_browser = tmp_path / "playwright-chromium.exe"
    explicit_browser.write_bytes(b"synthetic browser fixture")
    bundled_browser.write_bytes(b"synthetic Playwright fixture")
    monkeypatch.setattr(account_environment, "ACCOUNT_PROFILE_ROOT", profile_root)
    monkeypatch.setenv("MONITOR_BROWSER_EXECUTABLE", str(explicit_browser))
    monkeypatch.setenv("MONITOR_BROWSER_PROXY_PROBE_URL", "https://probe.invalid/region")

    account = _phase_5_1d_persisted_account()
    plan = resolve_account_browser_environment(
        account,
        action="login_check",
        trigger_source="profile_validation",
        headless=True,
        launch_mode="persistent_launch",
        proxy=_phase_5_1d_active_proxy(),
        playwright_executable_path=str(bundled_browser),
    )

    assert plan.workspace_id == account["workspace_id"]
    assert plan.account_id == account["id"]
    assert plan.platform == account["platform"]
    assert plan.identity_state == account["identity_state"]
    assert plan.identity_template == account["identity_template"]
    assert plan.profile_key == account["profile_key"]
    assert Path(plan.profile_path) == profile_root / "1" / "dy" / "acc_5101"
    assert plan.profile_path != account["profile_path"]
    assert plan.profile_mode == "persistent"
    assert plan.browser_executable_path == str(explicit_browser.resolve())
    assert plan.browser_source == "explicit"
    assert plan.browser_family == "chromium"
    assert plan.proxy_policy == "account_bound"
    assert plan.proxy_id == 41
    assert plan.proxy_region == "CN_MAINLAND"
    assert plan.proxy_url == _phase_5_1d_active_proxy()["proxy_url"]
    assert plan.provider_name == "playwright"
    assert plan.launch_mode == "persistent_launch"
    assert plan.headless is True

    with pytest.raises(FrozenInstanceError):
        plan.action = "crawl"
    serialized_plan = browser_environment_plan_to_json(plan)
    assert browser_environment_plan_from_json(serialized_plan) == plan
    unsafe_plan = json.loads(serialized_plan)
    unsafe_plan["unexpected"] = "process default"
    with pytest.raises(ValueError) as exc_info:
        browser_environment_plan_from_json(json.dumps(unsafe_plan))
    assert getattr(exc_info.value, "reason", "") == "account_identity_provider_unsupported"

    context_options = browser_context_options(plan)
    assert context_options == {
        "user_data_dir": plan.profile_path,
        "executable_path": str(explicit_browser.resolve()),
        "headless": True,
        "proxy": {
            "server": "http://proxy.invalid:8080",
            "username": "phase51d-user",
            "password": "phase51d-pass",
        },
        "user_agent": account["user_agent"],
        "timezone_id": account["timezone"],
        "locale": account["locale"],
        "extra_http_headers": {"Accept-Language": account["accept_language"]},
        "viewport": {"width": account["viewport_width"], "height": account["viewport_height"]},
        "screen": {"width": account["screen_width"], "height": account["screen_height"]},
        "device_scale_factor": account["device_scale_factor"],
        "is_mobile": account["is_mobile"],
        "has_touch": account["has_touch"],
    }


def test_phase_5_1d_restart_resolution_is_fresh_without_mutating_account_identity(tmp_path, monkeypatch):
    from dataclasses import asdict

    from api.monitoring import account_environment
    from api.monitoring.browser_environment_provider import resolve_account_browser_environment
    from tools.browser_environment import reset_browser_environment_cache_for_tests

    profile_root = (tmp_path / "restart-profiles").resolve()
    browser_path = tmp_path / "playwright-chromium.exe"
    browser_path.write_bytes(b"synthetic Playwright fixture")
    monkeypatch.setattr(account_environment, "ACCOUNT_PROFILE_ROOT", profile_root)
    monkeypatch.delenv("MONITOR_BROWSER_EXECUTABLE", raising=False)
    account = _phase_5_1d_persisted_account(state="active", proxy_id=None)
    original_account = json.loads(json.dumps(account))

    def resolve():
        return resolve_account_browser_environment(
            account,
            action="crawl",
            trigger_source="scheduler",
            headless=True,
            launch_mode="cdp_launch",
            playwright_executable_path=str(browser_path),
        )

    before_restart = resolve()
    reset_browser_environment_cache_for_tests()
    after_restart = resolve()

    assert before_restart.resolution_id != after_restart.resolution_id
    assert before_restart.attempt_id != after_restart.attempt_id
    comparable_before = {
        key: value for key, value in asdict(before_restart).items() if key not in {"resolution_id", "attempt_id"}
    }
    comparable_after = {
        key: value for key, value in asdict(after_restart).items() if key not in {"resolution_id", "attempt_id"}
    }
    assert comparable_before == comparable_after
    assert account == original_account


@pytest.mark.parametrize(
    ("case", "expected_reason"),
    [
        ("invalid_explicit_executable", "account_identity_provider_unsupported"),
        ("missing_identity_field", "account_identity_missing"),
        ("requires_relogin", "account_identity_requires_relogin"),
        ("locked_task_proxy_override", "account_identity_locked_proxy_override"),
        ("inactive_bound_proxy", "account_identity_missing"),
        ("missing_proxy_probe", "account_identity_provider_unsupported"),
        ("diagnostic_auto_detect_for_locked_identity", "account_identity_provider_unsupported"),
    ],
)
def test_phase_5_1d_provider_fails_closed_for_invalid_coverage(case, expected_reason, tmp_path, monkeypatch):
    from api.monitoring.browser_environment_provider import resolve_account_browser_environment

    bundled_browser = tmp_path / "playwright-chromium.exe"
    bundled_browser.write_bytes(b"synthetic Playwright fixture")
    monkeypatch.delenv("MONITOR_BROWSER_EXECUTABLE", raising=False)
    monkeypatch.setenv("MONITOR_BROWSER_PROXY_PROBE_URL", "https://probe.invalid/region")
    account = _phase_5_1d_persisted_account()
    proxy = _phase_5_1d_active_proxy()
    task_proxy_id = None
    diagnostic = False

    if case == "invalid_explicit_executable":
        monkeypatch.setenv("MONITOR_BROWSER_EXECUTABLE", str(tmp_path / "missing-browser.exe"))
    elif case == "missing_identity_field":
        account["timezone"] = ""
    elif case == "requires_relogin":
        account["requires_relogin"] = True
        account["identity_state"] = "requires_relogin"
    elif case == "locked_task_proxy_override":
        account["identity_state"] = "active"
        task_proxy_id = 99
    elif case == "inactive_bound_proxy":
        proxy["status"] = "disabled"
    elif case == "missing_proxy_probe":
        monkeypatch.delenv("MONITOR_BROWSER_PROXY_PROBE_URL", raising=False)
    elif case == "diagnostic_auto_detect_for_locked_identity":
        account["identity_state"] = "locked"
        diagnostic = True

    with pytest.raises(ValueError) as exc_info:
        resolve_account_browser_environment(
            account,
            action="login_check",
            trigger_source="profile_validation",
            headless=True,
            launch_mode="persistent_launch",
            proxy=proxy,
            task_proxy_id=task_proxy_id,
            playwright_executable_path=str(bundled_browser),
            diagnostic=diagnostic,
        )
    assert getattr(exc_info.value, "reason", "") == expected_reason


@pytest.mark.parametrize(
    "forbidden_key",
    [
        "cookie",
        "cookies",
        "cookie_value",
        "proxy_url",
        "proxy_credentials",
        "profile_path",
        "browser_executable_path",
        "executable_path",
        "cdp_url",
        "websocket_url",
        "debug_port",
        "novnc_token",
        "command",
        "environment",
        "fingerprint_seed",
        "probe_url",
        "external_ip",
        "exception",
    ],
)
def test_phase_5_1d_snapshot_recursively_rejects_forbidden_keys(forbidden_key):
    from tools.browser_environment import BrowserEnvironmentError, validate_safe_runtime_snapshot

    snapshot = _phase_5_1d_safe_snapshot()
    snapshot["probes"]["nested"] = {"evidence": [{forbidden_key: "secret-or-path"}]}

    with pytest.raises(BrowserEnvironmentError) as exc_info:
        validate_safe_runtime_snapshot(snapshot)
    assert exc_info.value.reason == "account_identity_snapshot_unsafe"


def test_phase_5_1d_snapshot_rejects_more_than_64_kib_before_persistence():
    from tools.browser_environment import BrowserEnvironmentError, validate_safe_runtime_snapshot

    snapshot = _phase_5_1d_safe_snapshot()
    snapshot["requested"]["user_agent"] = "X" * 70000

    with pytest.raises(BrowserEnvironmentError) as exc_info:
        validate_safe_runtime_snapshot(snapshot)

    assert exc_info.value.reason == "account_identity_snapshot_unsafe"
    assert exc_info.value.fields == ("snapshot_size",)


@pytest.mark.parametrize(
    "mismatch",
    [
        {"field": "undocumented_field", "requested": "a", "effective": "b"},
        {"field": "timezone", "requested": "Asia/Shanghai", "effective": "UTC", "detail": "raw"},
        {"field": "timezone", "requested": {"proxy_url": "http://secret"}, "effective": "UTC"},
        {"field": "timezone", "requested": "https://probe.invalid", "effective": "UTC"},
        {"field": "timezone", "requested": r"C:\secret\profile", "effective": "UTC"},
        {"field": "timezone", "requested": r"\\server\share\profile", "effective": "UTC"},
        {"field": "timezone", "requested": "/srv/account/profile", "effective": "UTC"},
        {"field": "timezone", "requested": ["Asia/Shanghai"], "effective": "UTC"},
        {"field": "screen_width", "requested": "1920", "effective": 1920},
        {"field": "is_mobile", "requested": 0, "effective": False},
    ],
)
def test_phase_5_1d_snapshot_rejects_unsafe_or_untyped_mismatch_evidence(mismatch):
    from tools.browser_environment import BrowserEnvironmentError, validate_safe_runtime_snapshot

    snapshot = _phase_5_1d_safe_snapshot()
    snapshot["ok"] = False
    snapshot["reason"] = "account_identity_snapshot_mismatch"
    snapshot["mismatch_evidence"] = [mismatch]

    with pytest.raises(BrowserEnvironmentError) as exc_info:
        validate_safe_runtime_snapshot(snapshot)
    assert exc_info.value.reason == "account_identity_snapshot_unsafe"


def test_phase_5_1d_snapshot_allows_field_scoped_safe_mismatch_and_bounded_summary():
    from api.monitoring.browser_environment_provider import safe_browser_environment_summary
    from tools.browser_environment import validate_safe_runtime_snapshot

    snapshot = _phase_5_1d_safe_snapshot()
    snapshot["ok"] = False
    snapshot["reason"] = "account_identity_snapshot_mismatch"
    snapshot["mismatch_evidence"] = [
        {"field": "timezone", "requested": "Asia/Shanghai", "effective": "UTC"}
    ]
    assert validate_safe_runtime_snapshot(snapshot) == snapshot

    summary = safe_browser_environment_summary(json.dumps(snapshot))
    assert summary == {
        "provider": {"name": "playwright", "mode": "persistent_launch"},
        "browser": {
            "family": "chromium",
            "version": "126.0.6478.183",
            "source": "playwright_bundled",
        },
        "profile": {"profile_key": "1/dy/acc_5101", "mode": "persistent"},
        "proxy": {"policy": "account_bound", "effect_proof": "passed"},
        "status": {
            "ok": False,
            "reason": "account_identity_snapshot_mismatch",
            "validated_at": "2026-07-19T04:00:00+00:00",
            "fallback_used": False,
            "unsupported_field_count": 4,
            "mismatch_fields": ["timezone"],
        },
    }
    serialized = json.dumps(summary, ensure_ascii=False)
    for forbidden in (
        "user_agent",
        "requested",
        "effective",
        "probes",
        "proxy_id",
        "CN_MAINLAND",
        '"account_id"',
        '"workspace_id"',
    ):
        assert forbidden not in serialized


def test_phase_5_1d_admin_account_view_exposes_only_safe_runtime_summary():
    snapshot = _phase_5_1d_safe_snapshot()
    snapshot["ok"] = False
    snapshot["reason"] = "account_identity_snapshot_mismatch"
    snapshot["mismatch_evidence"] = [
        {"field": "timezone", "requested": "Asia/Shanghai", "effective": "UTC"}
    ]
    raw = {
        "id": 5101,
        "name": "受管抖音账号",
        "platform": "dy",
        "identity_runtime_snapshot_json": json.dumps(snapshot),
        "cookies": "sessionid=secret-cookie",
        "fingerprint_seed": "secret-seed",
        "profile_path": r"E:\server\profiles\dy\acc_5101",
        "profile_runtime_path": r"E:\server\profiles\dy\acc_5101",
        "proxy_url": "http://user:password@proxy.invalid:8080",
        "browser_executable_path": r"C:\browser\chrome.exe",
        "cdp_url": "http://127.0.0.1:9222",
        "debug_port": 9222,
        "command": "chrome.exe --remote-debugging-port=9222",
        "requested_user_agent": "secret requested UA",
        "runtime_probes": {"user_agent": "secret effective UA"},
    }

    view = monitor_router._customer_view_social_account(raw)
    summary = view["identity_runtime_summary"]
    visible = json.dumps(view, ensure_ascii=False)

    assert summary["provider"] == {"name": "playwright", "mode": "persistent_launch"}
    assert summary["proxy"] == {"policy": "account_bound", "effect_proof": "passed"}
    assert summary["status"] == {
        "ok": False,
        "reason": "account_identity_snapshot_mismatch",
        "validated_at": "2026-07-19T04:00:00+00:00",
        "fallback_used": False,
        "unsupported_field_count": 4,
        "mismatch_fields": ["timezone"],
    }
    assert "identity_runtime_snapshot_json" not in view
    for forbidden in (
        "secret-cookie",
        "secret-seed",
        "E:\\server",
        "user:password",
        "chrome.exe",
        "127.0.0.1",
        "9222",
        "secret requested UA",
        "secret effective UA",
        '"requested"',
        '"effective"',
        '"probes"',
    ):
        assert forbidden not in visible


def test_phase_5_1d_admin_account_drawer_renders_compact_runtime_summary():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    start = page.index("function accountRuntimeEnvironmentSummary(account)")
    runtime_summary_function = page[start : page.index("function accountIdentityPreloginState", start)]

    for marker in (
        "identity_runtime_summary",
        "summary.provider",
        "summary.proxy",
        "summary.status",
        "validated_at",
        "fallback_used",
        "unsupported_field_count",
        "mismatch_fields",
    ):
        assert marker in runtime_summary_function
    for marker in (
        "playwright:'Playwright'",
        "persistent_launch:'持久化启动'",
        "ephemeral_cookie_validation:'Cookie 临时验活'",
        "cdp_launch:'CDP 启动'",
        "timezone:'时区'",
        "proxy_effect:'代理证明'",
    ):
        assert marker in runtime_summary_function
    assert "accountSummaryItem('运行实况', accountRuntimeEnvironmentSummary(account))" in page
    for forbidden in (
        "identity_runtime_snapshot_json",
        "requested_user_agent",
        "runtime_probes",
        "browser_executable_path",
        "cdp_url",
        "debug_port",
        "command_line",
    ):
        assert forbidden not in runtime_summary_function


def test_phase_5_1d_deployment_examples_expose_only_operator_provider_settings():
    examples = [
        Path(".env.example"),
        Path("deploy/docker/monitor.env.example"),
        Path("deploy/systemd/legal-sentiment-monitor.env.example"),
    ]
    for path in examples:
        content = path.read_text(encoding="utf-8")
        assert "MONITOR_BROWSER_EXECUTABLE=" in content, path
        assert "MONITOR_BROWSER_PROXY_PROBE_URL=" in content, path
        assert "MONITOR_BROWSER_PROXY_PROBE_TIMEOUT_MS=30000" in content, path
        assert "MONITOR_BROWSER_ENVIRONMENT_PLAN" not in content, path
        assert "MONITOR_BROWSER_ENVIRONMENT_RESULT_PATH" not in content, path


def test_phase_5_1d_snapshot_persistence_is_safe_and_account_scoped(monkeypatch):
    from api.monitoring.database import update_social_account_identity_runtime_snapshot
    from tools.browser_environment import BrowserEnvironmentError

    monkeypatch.setenv("MONITOR_ACCOUNT_IDENTITY_SEED_SALT", "phase-5.1d-persistence-salt")
    init_db()
    snapshot_rows = _snapshot_table("social_accounts")
    try:
        account = save_social_account(
            {
                "name": "Phase 5.1D Snapshot Account",
                "platform": "dy",
                "login_type": "qrcode",
                "status": "standby",
                "proxy_region_snapshot": "CN_MAINLAND",
                "identity_template_family": "windows_chrome_desktop",
            }
        )
        safe_snapshot = _phase_5_1d_safe_snapshot()
        safe_snapshot["account"] = {
            "workspace_id": account["workspace_id"],
            "account_id": account["id"],
            "platform": account["platform"],
            "identity_state": account["identity_state"],
        }
        safe_snapshot["profile"]["profile_key"] = account["profile_key"]
        with get_conn() as conn:
            before = dict(conn.execute("SELECT * FROM social_accounts WHERE id=?", (account["id"],)).fetchone())

        stored = update_social_account_identity_runtime_snapshot(account["id"], safe_snapshot)
        with get_conn() as conn:
            after = dict(conn.execute("SELECT * FROM social_accounts WHERE id=?", (account["id"],)).fetchone())

        assert json.loads(stored["identity_runtime_snapshot_json"]) == safe_snapshot
        for field in before:
            if field not in {"identity_runtime_snapshot_json", "updated_at"}:
                assert after[field] == before[field]

        mismatched = json.loads(json.dumps(safe_snapshot))
        mismatched["account"]["account_id"] = account["id"] + 1
        with pytest.raises(BrowserEnvironmentError) as exc_info:
            update_social_account_identity_runtime_snapshot(account["id"], mismatched)
        assert exc_info.value.reason == "account_identity_snapshot_mismatch"

        unsafe = json.loads(json.dumps(safe_snapshot))
        unsafe["probes"]["profile_path"] = r"C:\secret\profile"
        with pytest.raises(BrowserEnvironmentError) as exc_info:
            update_social_account_identity_runtime_snapshot(account["id"], unsafe)
        assert exc_info.value.reason == "account_identity_snapshot_unsafe"
    finally:
        _restore_table("social_accounts", snapshot_rows)


def _phase_5_1d_plan(tmp_path, monkeypatch, *, proxy_bound=False, action="login_check", launch_mode="persistent_launch"):
    from api.monitoring import account_environment
    from api.monitoring.browser_environment_provider import resolve_account_browser_environment

    profile_root = (tmp_path / "managed-profiles").resolve()
    browser_path = tmp_path / "playwright-chromium.exe"
    browser_path.write_bytes(b"synthetic Playwright fixture")
    monkeypatch.setattr(account_environment, "ACCOUNT_PROFILE_ROOT", profile_root)
    monkeypatch.delenv("MONITOR_BROWSER_EXECUTABLE", raising=False)
    account = _phase_5_1d_persisted_account(proxy_id=41 if proxy_bound else None)
    proxy = _phase_5_1d_active_proxy() if proxy_bound else None
    if proxy_bound:
        monkeypatch.setenv("MONITOR_BROWSER_PROXY_PROBE_URL", "https://probe.invalid/region")
    else:
        monkeypatch.delenv("MONITOR_BROWSER_PROXY_PROBE_URL", raising=False)
    return resolve_account_browser_environment(
        account,
        action=action,
        trigger_source="cookie_validation" if action == "cookie_validation" else "profile_validation",
        headless=True,
        launch_mode=launch_mode,
        proxy=proxy,
        playwright_executable_path=str(browser_path),
    )


def _phase_5_1d_effective_probe(plan):
    return {
        "user_agent": plan.user_agent,
        "timezone": plan.timezone,
        "language": plan.locale,
        "languages": [part.split(";", 1)[0].strip() for part in plan.accept_language.split(",")],
        "screen_width": plan.screen_width,
        "screen_height": plan.screen_height,
        "viewport_width": plan.viewport_width,
        "viewport_height": plan.viewport_height,
        "device_scale_factor": plan.device_scale_factor,
        "max_touch_points": 1 if plan.has_touch else 0,
        "is_mobile": plan.is_mobile,
        "webdriver": True,
    }


def test_phase_5_1d_proxy_proof_finishes_before_context_is_returned(tmp_path, monkeypatch):
    from tools.browser_environment import launch_managed_browser_context

    plan = _phase_5_1d_plan(tmp_path, monkeypatch, proxy_bound=True)
    events = []
    captured = {}

    class FakeResponse:
        ok = True
        status = 200

        async def json(self):
            return {"region": "CN_MAINLAND"}

    class FakeProofPage:
        async def goto(self, url, **kwargs):
            events.append(("probe", url, kwargs["timeout"]))
            return FakeResponse()

        async def close(self):
            events.append(("probe_closed",))

    class FakeContext:
        pages = []
        browser = None

        def on(self, event, callback):
            captured["request_handler"] = callback

        async def new_page(self):
            return FakeProofPage()

    context = FakeContext()

    class FakeChromium:
        async def launch_persistent_context(self, **kwargs):
            captured["launch"] = kwargs
            events.append(("launch",))
            return context

    class FakePlaywright:
        chromium = FakeChromium()

    session = asyncio.run(launch_managed_browser_context(FakePlaywright(), plan))

    assert session.context is context
    assert session.browser is None
    assert [item[0] for item in events] == ["launch", "probe", "probe_closed"]
    assert captured["launch"]["user_data_dir"] == plan.profile_path
    assert captured["launch"]["executable_path"] == plan.browser_executable_path
    assert captured["launch"]["proxy"] == {
        "server": "http://proxy.invalid:8080",
        "username": "phase51d-user",
        "password": "phase51d-pass",
    }
    assert getattr(context, "_monitor_proxy_effect_proof") == "passed"
    assert getattr(context, "_monitor_browser_environment_plan") is plan


@pytest.mark.parametrize(
    ("case", "expected_reason"),
    [
        ("absent_url", "account_identity_provider_unsupported"),
        ("http_failure", "account_identity_proxy_proof_failed"),
        ("malformed_body", "account_identity_proxy_proof_failed"),
        ("timeout", "account_identity_proxy_proof_failed"),
        ("wrong_region", "account_identity_snapshot_mismatch"),
    ],
)
def test_phase_5_1d_proxy_proof_fails_before_platform_navigation(case, expected_reason, tmp_path, monkeypatch):
    from tools.browser_environment import BrowserEnvironmentError, launch_managed_browser_context

    plan = _phase_5_1d_plan(tmp_path, monkeypatch, proxy_bound=True)
    if case == "absent_url":
        monkeypatch.delenv("MONITOR_BROWSER_PROXY_PROBE_URL", raising=False)
    platform_navigation = []

    class FakeResponse:
        status = 503 if case == "http_failure" else 200
        ok = case != "http_failure"

        async def json(self):
            if case == "malformed_body":
                return ["CN_MAINLAND"]
            return {"region": "HK" if case == "wrong_region" else "CN_MAINLAND"}

    class FakeProofPage:
        async def goto(self, *args, **kwargs):
            if case == "timeout":
                raise TimeoutError("synthetic timeout")
            return FakeResponse()

        async def close(self):
            pass

    class FakeContext:
        pages = []
        browser = None

        def on(self, *args):
            pass

        async def new_page(self):
            return FakeProofPage()

    class FakeChromium:
        async def launch_persistent_context(self, **kwargs):
            return FakeContext()

    class FakePlaywright:
        chromium = FakeChromium()

    with pytest.raises(BrowserEnvironmentError) as exc_info:
        asyncio.run(launch_managed_browser_context(FakePlaywright(), plan))
    assert exc_info.value.reason == expected_reason
    failure_result = getattr(exc_info.value, "browser_environment_result", None)
    assert failure_result is not None
    assert failure_result.ok is False
    assert failure_result.snapshot["proxy"]["effect_proof"] == "failed"
    assert failure_result.snapshot["effective"] == {}
    assert failure_result.snapshot["probes"] == {}
    assert platform_navigation == []


@pytest.mark.parametrize(
    ("field", "mutated_value"),
    [
        ("user_agent", "Mozilla/5.0 Chrome/125.0.0.0"),
        ("timezone", "UTC"),
        ("locale", "en-US"),
        ("accept_language", "en-US,en;q=0.9"),
        ("screen_width", 1280),
        ("viewport_height", 720),
        ("device_scale_factor", 2.0),
        ("has_touch", True),
        ("is_mobile", True),
        ("browser_version", "125.0.0.0"),
    ],
)
def test_phase_5_1d_page_probe_records_field_scoped_mismatch(field, mutated_value, tmp_path, monkeypatch):
    from tools.browser_environment import launch_managed_browser_context, verify_managed_page

    plan = _phase_5_1d_plan(tmp_path, monkeypatch)
    probe = _phase_5_1d_effective_probe(plan)
    browser_version = plan.browser_version
    request_headers = {"accept-language": plan.accept_language}
    if field == "locale":
        probe["language"] = mutated_value
    elif field == "accept_language":
        request_headers["accept-language"] = mutated_value
    elif field == "has_touch":
        probe["max_touch_points"] = 1
    elif field == "browser_version":
        browser_version = mutated_value
    else:
        probe[field] = mutated_value

    class FakeRequest:
        headers = request_headers

    class FakePage:
        async def goto(self, *args, **kwargs):
            context.request_handler(FakeRequest())

        async def evaluate(self, script):
            return probe

    class FakeBrowser:
        version = browser_version

    class FakeContext:
        pages = []
        browser = FakeBrowser()

        def on(self, event, callback):
            self.request_handler = callback

    context = FakeContext()

    class FakeChromium:
        async def launch_persistent_context(self, **kwargs):
            return context

    class FakePlaywright:
        chromium = FakeChromium()

    session = asyncio.run(launch_managed_browser_context(FakePlaywright(), plan))
    page = FakePage()
    asyncio.run(page.goto("https://platform.invalid"))
    result = asyncio.run(verify_managed_page(session.context, page))

    assert result is not None
    assert result.ok is False
    assert result.reason == "account_identity_snapshot_mismatch"
    assert result.snapshot["mismatch_evidence"] == [
        {
            "field": field,
            "requested": getattr(plan, field if field != "browser_version" else "browser_version"),
            "effective": mutated_value,
        }
    ]
    assert set(result.snapshot["mismatch_evidence"][0]) == {"field", "requested", "effective"}


def test_phase_5_1d_profile_cookie_and_visible_login_share_exact_plan(tmp_path, monkeypatch):
    from api.monitoring import account_environment

    profile_root = (tmp_path / "shared-profiles").resolve()
    browser_path = tmp_path / "playwright-chromium.exe"
    browser_path.write_bytes(b"synthetic Playwright fixture")
    monkeypatch.setattr(account_environment, "ACCOUNT_PROFILE_ROOT", profile_root)
    monkeypatch.delenv("MONITOR_BROWSER_EXECUTABLE", raising=False)
    account = _phase_5_1d_persisted_account(proxy_id=None)
    account["cookies"] = "sessionid=synthetic"
    account["login_type"] = "cookie"
    profile_path = profile_root / "1" / "dy" / "acc_5101"
    profile_path.mkdir(parents=True)
    launches = []

    class FakeRequest:
        headers = {"accept-language": account["accept_language"]}

    class FakePage:
        def set_default_timeout(self, timeout):
            pass

        async def goto(self, *args, **kwargs):
            self.context.request_handler(FakeRequest())

        async def wait_for_timeout(self, timeout):
            pass

        async def evaluate(self, script):
            return _phase_5_1d_effective_probe(self.context._monitor_browser_environment_plan)

    class FakeContext:
        def __init__(self, browser):
            self.browser = browser
            self.pages = [FakePage()]
            self.pages[0].context = self
            self.closed = False
            self.cookies_added = []

        def on(self, event, callback):
            self.request_handler = callback

        async def new_page(self):
            page = FakePage()
            page.context = self
            self.pages.append(page)
            return page

        async def add_cookies(self, cookies):
            self.cookies_added.extend(cookies)

        async def cookies(self):
            return []

        async def close(self):
            self.closed = True

    class FakeBrowser:
        version = account["user_agent"].split("Chrome/", 1)[1].split(" ", 1)[0]

        def __init__(self):
            self.closed = False
            self.context = None

        async def new_context(self, **kwargs):
            launches.append(("new_context", kwargs))
            self.context = FakeContext(self)
            return self.context

        async def close(self):
            self.closed = True

    class FakeChromium:
        executable_path = str(browser_path)

        async def launch_persistent_context(self, **kwargs):
            launches.append(("persistent", kwargs))
            return FakeContext(FakeBrowser())

        async def launch(self, **kwargs):
            launches.append(("ephemeral", kwargs))
            return FakeBrowser()

    class FakePlaywright:
        chromium = FakeChromium()
        stopped = False

        async def stop(self):
            self.stopped = True

    class FakeFactory:
        async def start(self):
            return FakePlaywright()

    async def fake_verify(*args, **kwargs):
        return {"ok": True, "status": "valid", "message": "ok"}

    async def fake_identity(*args, **kwargs):
        return {}

    monkeypatch.setattr(account_check_module, "async_playwright", lambda: FakeFactory())
    monkeypatch.setattr(account_check_module, "_verify_collectable_login", fake_verify)
    monkeypatch.setattr(account_check_module, "_extract_platform_identity", fake_identity)
    monkeypatch.setattr(account_check_module, "get_proxy_profile", lambda *args, **kwargs: None, raising=False)
    monkeypatch.setattr(monitor_router, "get_social_account", lambda *args, **kwargs: account)
    monkeypatch.setattr(monitor_router, "get_proxy_profile", lambda *args, **kwargs: None)
    monkeypatch.setattr(
        monitor_router,
        "_playwright_chromium_executable_path",
        lambda: str(browser_path),
        raising=False,
    )

    profile_account = {**account, "login_type": "qrcode"}
    profile_result = asyncio.run(account_check_module._check_profile_account(profile_account, 1000))
    cookie_result = asyncio.run(account_check_module._check_cookie_account(account, 1000))
    visible_command = asyncio.run(
        monitor_router._login_browser_command_for_payload(
            "dy",
            {"account_id": account["id"]},
            action="login_check",
            trigger_source="visible_browser_login",
            headless=False,
        )
    )

    profile_plan = profile_result["_browser_environment_plan"]
    cookie_plan = cookie_result["_browser_environment_plan"]
    visible_plan = visible_command["_browser_environment_plan"]
    shared_fields = (
        "workspace_id",
        "account_id",
        "platform",
        "identity_template",
        "browser_executable_path",
        "browser_family",
        "browser_source",
        "profile_key",
        "proxy_policy",
        "proxy_id",
        "proxy_region",
        "user_agent",
        "timezone",
        "locale",
        "accept_language",
        "screen_width",
        "screen_height",
        "viewport_width",
        "viewport_height",
        "device_scale_factor",
        "is_mobile",
        "has_touch",
    )
    for field in shared_fields:
        assert getattr(profile_plan, field) == getattr(cookie_plan, field) == getattr(visible_plan, field)
    assert profile_plan.profile_mode == "persistent"
    assert cookie_plan.profile_mode == "ephemeral_cookie_validation"
    assert cookie_plan.launch_mode == "ephemeral_cookie_validation"
    assert visible_plan.headless is False
    assert launches[0][0] == "persistent"
    assert launches[1][0] == "ephemeral"
    assert all(str(profile_path) not in json.dumps(options) for kind, options in launches if kind != "persistent")
    assert cookie_result["_browser_session_closed"] is True


def test_phase_5_1d_qr_provider_mismatch_stops_before_qrcode_prepare(tmp_path, monkeypatch):
    from api.monitoring.login_browser import build_managed_login_browser_command

    plan = _phase_5_1d_plan(tmp_path, monkeypatch, action="qr_login")
    events = []
    probe = _phase_5_1d_effective_probe(plan)
    probe["timezone"] = "UTC"

    class FakeRequest:
        headers = {"accept-language": plan.accept_language}

    class FakePage:
        def set_default_timeout(self, timeout):
            pass

        async def goto(self, *args, **kwargs):
            events.append("platform_goto")
            context.request_handler(FakeRequest())

        async def evaluate(self, script):
            return probe

    class FakeBrowser:
        version = plan.browser_version

    class FakeContext:
        browser = FakeBrowser()
        pages = [FakePage()]
        closed = False

        def on(self, event, callback):
            self.request_handler = callback

        async def close(self):
            self.closed = True

    context = FakeContext()

    class FakeChromium:
        async def launch_persistent_context(self, **kwargs):
            events.append("launch")
            return context

    class FakePlaywright:
        chromium = FakeChromium()

        async def stop(self):
            events.append("stop")

    class FakeFactory:
        async def start(self):
            return FakePlaywright()

    async def forbidden_prepare(*args, **kwargs):
        events.append("qrcode_prepare")

    monkeypatch.setattr(login_qrcode_module, "async_playwright", lambda: FakeFactory())
    monkeypatch.setattr(login_qrcode_module, "_prepare_login_page", forbidden_prepare)
    result = asyncio.run(
        login_qrcode_module._start_qrcode_login_session_with_profile_once(
            951001,
            "dy",
            build_managed_login_browser_command(plan),
            1000,
        )
    )

    assert result["ok"] is False
    assert result["_browser_environment_result"].reason == "account_identity_snapshot_mismatch"
    assert events[:2] == ["launch", "platform_goto"]
    assert "qrcode_prepare" not in events
    assert context.closed is True


def test_phase_5_1d_browser_disconnect_after_proof_is_typed_and_has_no_fallback(tmp_path, monkeypatch):
    from tools.browser_environment import (
        BrowserEnvironmentError,
        launch_managed_browser_context,
        verify_managed_page,
    )

    plan = _phase_5_1d_plan(tmp_path, monkeypatch)

    class FakeBrowser:
        version = plan.browser_version

    class FakeContext:
        browser = FakeBrowser()
        pages = []

        def on(self, event, callback):
            self.request_handler = callback

    context = FakeContext()

    class FakeChromium:
        launches = 0

        async def launch_persistent_context(self, **kwargs):
            self.launches += 1
            return context

    chromium = FakeChromium()

    class FakePlaywright:
        pass

    playwright = FakePlaywright()
    playwright.chromium = chromium

    class ClosedPage:
        async def evaluate(self, script):
            raise RuntimeError("Target page, context or browser has been closed")

    session = asyncio.run(launch_managed_browser_context(playwright, plan))
    with pytest.raises(BrowserEnvironmentError) as exc_info:
        asyncio.run(verify_managed_page(session.context, ClosedPage()))

    assert exc_info.value.reason == "account_identity_provider_browser_crashed"
    assert exc_info.value.browser_environment_result.ok is False
    assert exc_info.value.browser_environment_result.snapshot["fallback_used"] is False
    assert chromium.launches == 1


def test_phase_5_1d_failed_direct_proof_persists_before_lifecycle_recovery(tmp_path, monkeypatch):
    from api.monitoring.browser_environment_provider import (
        persist_account_browser_environment_result as real_persist,
        resolve_account_browser_environment,
    )
    from tools.browser_environment import browser_environment_failure_result

    monkeypatch.setenv("MONITOR_ACCOUNT_IDENTITY_SEED_SALT", "phase-5.1d-direct-order-salt")
    browser_path = tmp_path / "playwright-chromium.exe"
    browser_path.write_bytes(b"synthetic Playwright fixture")
    init_db()
    rows = _snapshot_table("social_accounts")
    observed_states = []
    try:
        account = save_social_account(
            {"name": "Phase 5.1D Direct Failure", "platform": "dy", "login_type": "qrcode"}
        )

        async def fake_profile_check(prepared_account, timeout_ms):
            plan = resolve_account_browser_environment(
                prepared_account,
                action="login_check",
                trigger_source="profile_validation",
                headless=True,
                launch_mode="persistent_launch",
                proxy=None,
                playwright_executable_path=str(browser_path),
            )
            provider_result = browser_environment_failure_result(
                plan,
                "account_identity_snapshot_mismatch",
                proxy_effect="not_applicable",
            )
            return {
                "ok": True,
                "status": "valid",
                "message": "business check must not activate",
                "identity": {},
                "_browser_environment_plan": plan,
                "_browser_environment_result": provider_result,
                "_browser_session_closed": True,
            }

        def recording_persist(account_id, plan, result):
            observed_states.append(get_social_account(account_id)["identity_state"])
            return real_persist(account_id, plan, result)

        monkeypatch.setattr(account_check_module, "_check_profile_account", fake_profile_check)
        monkeypatch.setattr(account_check_module, "persist_account_browser_environment_result", recording_persist)

        checked = asyncio.run(account_check_module.check_social_account_login(account["id"]))
        stored = get_social_account(account["id"], masked=False)

        assert observed_states == ["login_in_progress"]
        assert checked["ok"] is False
        assert stored["identity_state"] == "validated"
        assert stored["browser_environment_locked_at"] is None
        assert json.loads(stored["identity_runtime_snapshot_json"])["ok"] is False
    finally:
        _restore_table("social_accounts", rows)


def _phase_5_1d_result_for_plan(plan, *, ok=True, reason=""):
    from tools.browser_environment import BrowserEnvironmentResult

    effective = {
        "user_agent": plan.user_agent,
        "timezone": plan.timezone,
        "locale": plan.locale,
        "accept_language": plan.accept_language,
        "screen_width": plan.screen_width,
        "screen_height": plan.screen_height,
        "viewport_width": plan.viewport_width,
        "viewport_height": plan.viewport_height,
        "device_scale_factor": float(plan.device_scale_factor),
        "is_mobile": plan.is_mobile,
        "has_touch": plan.has_touch,
        "proxy_region_snapshot": plan.proxy_region,
    }
    probes = {
        "navigator_user_agent": plan.user_agent,
        "navigator_language": plan.locale,
        "navigator_languages": [part.split(";", 1)[0].strip() for part in plan.accept_language.split(",")],
        "timezone": plan.timezone,
        "screen_width": plan.screen_width,
        "screen_height": plan.screen_height,
        "viewport_width": plan.viewport_width,
        "viewport_height": plan.viewport_height,
        "device_scale_factor": float(plan.device_scale_factor),
        "max_touch_points": 1 if plan.has_touch else 0,
        "is_mobile": plan.is_mobile,
        "webdriver": True,
    }
    snapshot = {
        "contract_version": 1,
        "resolution_id": plan.resolution_id,
        "attempt_id": plan.attempt_id,
        "action": plan.action,
        "trigger_source": plan.trigger_source,
        "account": {
            "workspace_id": plan.workspace_id,
            "account_id": plan.account_id,
            "platform": plan.platform,
            "identity_state": plan.identity_state,
        },
        "browser": {
            "family": plan.browser_family,
            "version": plan.browser_version,
            "source": plan.browser_source,
        },
        "profile": {"profile_key": plan.profile_key, "mode": plan.profile_mode},
        "proxy": {
            "policy": plan.proxy_policy,
            "proxy_id": plan.proxy_id,
            "region": plan.proxy_region,
            "effect_proof": "passed" if plan.proxy_policy == "account_bound" else "not_applicable",
        },
        "requested": {
            "identity_template": plan.identity_template,
            "browser_platform": plan.browser_platform,
            "user_agent": plan.user_agent,
            "timezone": plan.timezone,
            "locale": plan.locale,
            "accept_language": plan.accept_language,
            "screen_width": plan.screen_width,
            "screen_height": plan.screen_height,
            "viewport_width": plan.viewport_width,
            "viewport_height": plan.viewport_height,
            "device_scale_factor": float(plan.device_scale_factor),
            "is_mobile": plan.is_mobile,
            "has_touch": plan.has_touch,
            "proxy_region_snapshot": plan.proxy_region,
        },
        "effective": effective if ok else {},
        "provider": {"name": "playwright", "mode": plan.launch_mode, "version": "1.45.0"},
        "probes": probes if ok else {},
        "unsupported_fields": ["canvas", "webgl", "fonts", "plugins"],
        "mismatch_evidence": [],
        "fallback_used": False,
        "ok": ok,
        "reason": reason,
        "validated_at": datetime.now(timezone.utc).isoformat(),
    }
    return BrowserEnvironmentResult(ok=ok, reason=reason, snapshot=snapshot)


def test_phase_5_1d_runner_handoff_is_bounded_and_not_in_command_or_summary(tmp_path, monkeypatch):
    from dataclasses import replace

    from tools.browser_environment import (
        PLAN_ENV_NAME,
        RESULT_PATH_ENV_NAME,
        browser_environment_plan_from_json,
        plan_from_environment,
        reset_browser_environment_cache_for_tests,
    )

    plan = _phase_5_1d_plan(tmp_path, monkeypatch, proxy_bound=True, action="crawl", launch_mode="cdp_launch")
    result_path = tmp_path / "attempt" / "browser-environment-result.json"
    account_binding = {
        "account_id": plan.account_id,
        "account_name": "synthetic account",
        "platform": plan.platform,
        "login_type": "qrcode",
        "profile_key": plan.profile_key,
        "profile_path": plan.profile_path,
        "proxy_id": plan.proxy_id,
        "proxy_name": "synthetic proxy",
        "proxy_url": plan.proxy_url,
    }
    monkeypatch.setenv("HTTP_PROXY", "http://process-default.invalid:8888")
    monkeypatch.setenv("MONITOR_CDP_USER_DATA_DIR", r"C:\generic\profile")
    monkeypatch.setenv("MONITOR_CDP_CONNECT_EXISTING", "false")

    env = runner_module._build_crawler_env(account_binding, plan, result_path)
    cmd = runner_module._build_crawler_cmd(
        {"keywords": ["Phase 5.1D"], "target_type": "search"},
        "dy",
        tmp_path,
        account_binding,
        plan,
    )

    assert len(env[PLAN_ENV_NAME].encode("utf-8")) <= 8192
    assert browser_environment_plan_from_json(env[PLAN_ENV_NAME]) == plan
    assert env[RESULT_PATH_ENV_NAME] == str(result_path.resolve())
    for forbidden_env in (
        "HTTP_PROXY",
        "HTTPS_PROXY",
        "ALL_PROXY",
        "MONITOR_CDP_USER_DATA_DIR",
        "MONITOR_CDP_USER_DATA_DIR_DY",
        "MONITOR_ACTIVE_ACCOUNT_ID",
        "MONITOR_ACTIVE_PROXY_ID",
    ):
        assert forbidden_env not in env
    serialized_cmd = " ".join(cmd)
    for forbidden in (PLAN_ENV_NAME, plan.profile_path, plan.proxy_url, plan.resolution_id, plan.attempt_id):
        assert forbidden not in serialized_cmd
    assert _cmd_value(cmd, "--cdp_connect_existing") == "false"
    assert _cmd_value(cmd, "--headless") == "true"

    long_plan = replace(
        plan,
        user_agent="Mozilla/5.0 " + ("X" * 900),
        proxy_url="http://" + ("u" * 700) + ":" + ("p" * 700) + "@proxy.invalid:8080",
    )
    long_payload = runner_module._build_crawler_env(account_binding, long_plan, result_path)[PLAN_ENV_NAME]
    parsed_long = browser_environment_plan_from_json(long_payload)
    assert parsed_long.user_agent == long_plan.user_agent
    assert parsed_long.proxy_url == long_plan.proxy_url
    assert len(long_payload.encode("utf-8")) <= 8192

    reset_browser_environment_cache_for_tests()
    monkeypatch.setenv(PLAN_ENV_NAME, env[PLAN_ENV_NAME])
    monkeypatch.setenv("HTTP_PROXY", plan.proxy_url)
    monkeypatch.setenv("HTTPS_PROXY", plan.proxy_url)
    parsed = plan_from_environment(required=True)
    assert parsed == plan
    assert PLAN_ENV_NAME not in os.environ
    assert "HTTP_PROXY" not in os.environ
    assert "HTTPS_PROXY" not in os.environ
    reset_browser_environment_cache_for_tests()


@pytest.mark.parametrize(
    ("case", "expected_reason"),
    [
        ("missing", "account_identity_child_result_missing"),
        ("malformed", "account_identity_snapshot_unsafe"),
        ("temp_only", "account_identity_child_result_missing"),
        ("wrong_attempt", "account_identity_snapshot_mismatch"),
        ("stale", "account_identity_snapshot_mismatch"),
    ],
)
def test_phase_5_1d_runner_rejects_missing_unsafe_or_stale_child_result(case, expected_reason, tmp_path, monkeypatch):
    from dataclasses import replace

    from tools.browser_environment import BrowserEnvironmentError

    plan = _phase_5_1d_plan(tmp_path, monkeypatch, action="crawl", launch_mode="cdp_launch")
    result_path = tmp_path / "result.json"
    started_at = datetime.now(timezone.utc)
    if case == "malformed":
        result_path.write_text("{not-json", encoding="utf-8")
    elif case == "temp_only":
        result_path.with_name(f"{result_path.name}.tmp.123").write_text(
            json.dumps(_phase_5_1d_result_for_plan(plan).snapshot),
            encoding="utf-8",
        )
    elif case == "wrong_attempt":
        wrong = _phase_5_1d_result_for_plan(replace(plan, attempt_id="attempt-wrong"))
        result_path.write_text(
            json.dumps({"ok": wrong.ok, "reason": wrong.reason, "snapshot": wrong.snapshot}),
            encoding="utf-8",
        )
    elif case == "stale":
        stale = _phase_5_1d_result_for_plan(plan)
        stale.snapshot["validated_at"] = (started_at - timedelta(minutes=5)).isoformat()
        result_path.write_text(
            json.dumps({"ok": stale.ok, "reason": stale.reason, "snapshot": stale.snapshot}),
            encoding="utf-8",
        )

    with pytest.raises(BrowserEnvironmentError) as exc_info:
        runner_module._load_managed_child_result(result_path, plan, started_at)
    assert exc_info.value.reason == expected_reason


def test_phase_5_1d_runner_accepts_only_atomic_fresh_bound_child_result(tmp_path, monkeypatch):
    from tools.browser_environment import RESULT_PATH_ENV_NAME, write_browser_environment_result

    plan = _phase_5_1d_plan(tmp_path, monkeypatch, action="crawl", launch_mode="cdp_launch")
    result = _phase_5_1d_result_for_plan(plan)
    destination = tmp_path / "atomic-result.json"
    monkeypatch.setenv(RESULT_PATH_ENV_NAME, str(destination))
    started_at = datetime.now(timezone.utc) - timedelta(seconds=1)

    write_browser_environment_result(result)
    loaded = runner_module._load_managed_child_result(destination, plan, started_at)

    assert loaded == result
    assert destination.is_file()
    assert not list(tmp_path.glob("atomic-result.json.tmp.*"))


def test_phase_5_1d_runner_resolves_after_locks_and_persists_before_ingest(tmp_path, monkeypatch):
    from dataclasses import replace

    base_plan = _phase_5_1d_plan(tmp_path, monkeypatch, action="crawl", launch_mode="cdp_launch")
    Path(base_plan.profile_path).mkdir(parents=True, exist_ok=True)
    account = _phase_5_1d_persisted_account(state="active", proxy_id=None)
    binding = {
        "account_id": account["id"],
        "account_name": "Phase 5.1D Runner",
        "platform": "dy",
        "login_type": "qrcode",
        "profile_key": account["profile_key"],
        "profile_path": base_plan.profile_path,
        "proxy_id": None,
        "_account": account,
        "_proxy": None,
        "task_proxy_id": None,
    }
    events = []
    lock_state = {"held": False}
    plans = []

    monkeypatch.setattr(runner_module, "_resolve_platform_account_binding", lambda *args: binding)
    monkeypatch.setattr(runner_module, "_raise_if_stop_requested", lambda *args: None)
    monkeypatch.setattr(runner_module, "_raise_if_deadline_passed", lambda *args: None)
    monkeypatch.setattr(runner_module, "_remaining_run_seconds", lambda *args: 30)
    monkeypatch.setattr(runner_module, "_run_timeout_seconds", lambda *args: 30)
    monkeypatch.setattr(runner_module, "_run_deadline_at", lambda *args: "")
    monkeypatch.setattr(runner_module, "_lock_expires_at", lambda *args: "future")
    monkeypatch.setattr(runner_module, "_ensure_login_window_closed", lambda *args: None)
    monkeypatch.setattr(runner_module, "set_run_resource_bindings", lambda *args: events.append("bind"))

    def acquire(*args):
        lock_state["held"] = True
        events.append("lock")
        return True

    def release(*args):
        events.append("unlock")
        lock_state["held"] = False

    monkeypatch.setattr(runner_module, "acquire_account_lock", acquire)
    monkeypatch.setattr(runner_module, "release_account_lock", release)

    def resolve(binding_arg, trigger_source):
        assert lock_state["held"] is True
        events.append(f"resolve:{trigger_source}")
        plan = replace(
            base_plan,
            resolution_id=f"resolution-{trigger_source}",
            attempt_id=f"attempt-base-{trigger_source}",
            trigger_source=trigger_source,
        )
        plans.append(plan)
        return plan

    monkeypatch.setattr(runner_module, "_resolve_runner_browser_plan", resolve, raising=False)

    def fake_attempt(job_arg, platform, out_dir, account_binding):
        assert lock_state["held"] is True
        events.append("attempt")
        plan = job_arg["_browser_environment_plan"]
        return _phase_5_1d_result_for_plan(plan)

    monkeypatch.setattr(runner_module, "_run_crawler_attempt", fake_attempt)

    def persist(account_id, plan, result):
        assert lock_state["held"] is True
        events.append("persist")
        return result.snapshot

    monkeypatch.setattr(runner_module, "persist_account_browser_environment_result", persist, raising=False)
    monkeypatch.setattr(runner_module, "collect_platform_outputs", lambda *args: (events.append("collect") or ([], [])))
    monkeypatch.setattr(
        runner_module,
        "ingest_outputs",
        lambda *args: {
            "raw_contents": 0,
            "filtered_contents": 0,
            "excluded_contents": 0,
            "new_contents": 0,
            "content_db_ids": [],
        },
    )
    monkeypatch.setattr(runner_module, "_update_collection_progress", lambda *args, **kwargs: None)
    monkeypatch.setattr(runner_module, "_finalize_collection_progress", lambda *args, **kwargs: None)
    monkeypatch.setenv("MONITOR_CRAWLER_MAX_RETRIES", "0")

    manual = asyncio.run(
        runner_module.run_platform({"id": 9511, "keywords": [], "_trigger_source": "manual"}, 9511, "dy", tmp_path / "manual")
    )
    scheduler = asyncio.run(
        runner_module.run_platform({"id": 9512, "keywords": [], "_trigger_source": "scheduler"}, 9512, "dy", tmp_path / "scheduler")
    )

    assert manual["new_contents"] == scheduler["new_contents"] == 0
    assert [plan.trigger_source for plan in plans] == ["manual", "scheduler"]
    comparable = [
        {key: value for key, value in plan.__dict__.items() if key not in {"resolution_id", "attempt_id", "trigger_source"}}
        for plan in plans
    ]
    assert comparable[0] == comparable[1]
    assert events.index("lock") < events.index("resolve:manual") < events.index("attempt") < events.index("persist") < events.index("collect") < events.index("unlock")


def test_phase_5_1d_runner_missing_child_result_blocks_ingest(tmp_path, monkeypatch):
    base_plan = _phase_5_1d_plan(tmp_path, monkeypatch, action="crawl", launch_mode="cdp_launch")
    Path(base_plan.profile_path).mkdir(parents=True, exist_ok=True)
    binding = {
        "account_id": base_plan.account_id,
        "account_name": "Phase 5.1D Missing Result",
        "platform": "dy",
        "login_type": "qrcode",
        "profile_key": base_plan.profile_key,
        "profile_path": base_plan.profile_path,
        "proxy_id": None,
        "_account": _phase_5_1d_persisted_account(state="active", proxy_id=None),
        "_proxy": None,
        "task_proxy_id": None,
    }
    monkeypatch.setattr(runner_module, "_resolve_platform_account_binding", lambda *args: binding)
    monkeypatch.setattr(runner_module, "_resolve_runner_browser_plan", lambda *args: base_plan, raising=False)
    monkeypatch.setattr(runner_module, "_ensure_login_window_closed", lambda *args: None)
    monkeypatch.setattr(runner_module, "_raise_if_stop_requested", lambda *args: None)
    monkeypatch.setattr(runner_module, "_raise_if_deadline_passed", lambda *args: None)
    monkeypatch.setattr(runner_module, "_remaining_run_seconds", lambda *args: 30)
    monkeypatch.setattr(runner_module, "_lock_expires_at", lambda *args: "future")
    monkeypatch.setattr(runner_module, "acquire_account_lock", lambda *args: True)
    monkeypatch.setattr(runner_module, "release_account_lock", lambda *args: None)
    monkeypatch.setattr(runner_module, "set_run_resource_bindings", lambda *args: None)
    monkeypatch.setattr(runner_module, "_run_crawler_attempt", lambda *args: None)
    monkeypatch.setattr(
        runner_module,
        "collect_platform_outputs",
        lambda *args: (_ for _ in ()).throw(AssertionError("ingest must not run without a child result")),
    )
    monkeypatch.setattr(runner_module, "_update_collection_progress", lambda *args, **kwargs: None)
    monkeypatch.setenv("MONITOR_CRAWLER_MAX_RETRIES", "0")

    with pytest.raises(RuntimeError, match="child result"):
        asyncio.run(
            runner_module.run_platform(
                {"id": 9521, "keywords": [], "_trigger_source": "manual"},
                9521,
                "dy",
                tmp_path,
            )
        )


def test_phase_5_1d_runner_provider_failure_stops_retry_chain_after_persistence(tmp_path, monkeypatch):
    base_plan = _phase_5_1d_plan(tmp_path, monkeypatch, action="crawl", launch_mode="cdp_launch")
    Path(base_plan.profile_path).mkdir(parents=True, exist_ok=True)
    binding = {
        "account_id": base_plan.account_id,
        "account_name": "Phase 5.1D Provider Failure",
        "platform": "dy",
        "login_type": "qrcode",
        "profile_key": base_plan.profile_key,
        "profile_path": base_plan.profile_path,
        "proxy_id": None,
        "_account": _phase_5_1d_persisted_account(state="active", proxy_id=None),
        "_proxy": None,
        "task_proxy_id": None,
    }
    calls = {"attempt": 0, "persist": 0}

    monkeypatch.setattr(runner_module, "_resolve_platform_account_binding", lambda *args: binding)
    monkeypatch.setattr(runner_module, "_resolve_runner_browser_plan", lambda *args: base_plan, raising=False)
    monkeypatch.setattr(runner_module, "_ensure_login_window_closed", lambda *args: None)
    monkeypatch.setattr(runner_module, "_raise_if_stop_requested", lambda *args: None)
    monkeypatch.setattr(runner_module, "_raise_if_deadline_passed", lambda *args: None)
    monkeypatch.setattr(runner_module, "_remaining_run_seconds", lambda *args: 30)
    monkeypatch.setattr(runner_module, "_run_timeout_seconds", lambda *args: 30)
    monkeypatch.setattr(runner_module, "_run_deadline_at", lambda *args: "")
    monkeypatch.setattr(runner_module, "_lock_expires_at", lambda *args: "future")
    monkeypatch.setattr(runner_module, "acquire_account_lock", lambda *args: True)
    monkeypatch.setattr(runner_module, "release_account_lock", lambda *args: None)
    monkeypatch.setattr(runner_module, "set_run_resource_bindings", lambda *args: None)
    monkeypatch.setattr(runner_module, "_update_collection_progress", lambda *args, **kwargs: None)
    monkeypatch.setattr(runner_module, "_finalize_collection_progress", lambda *args, **kwargs: None)
    monkeypatch.setenv("MONITOR_CRAWLER_MAX_RETRIES", "2")

    def fail_provider(job_arg, *args):
        calls["attempt"] += 1
        attempt_plan = job_arg["_browser_environment_plan"]
        result = _phase_5_1d_result_for_plan(
            attempt_plan,
            ok=False,
            reason="account_identity_snapshot_mismatch",
        )
        return runner_module.ManagedCrawlerOutcome(result, 0, False)

    def persist(*args):
        calls["persist"] += 1
        return args[2].snapshot

    monkeypatch.setattr(runner_module, "_run_crawler_attempt", fail_provider)
    monkeypatch.setattr(runner_module, "persist_account_browser_environment_result", persist, raising=False)
    monkeypatch.setattr(
        runner_module,
        "collect_platform_outputs",
        lambda *args: (_ for _ in ()).throw(AssertionError("provider failure must block ingest")),
    )

    with pytest.raises(RuntimeError, match="failed after 1 attempt.*account_identity_snapshot_mismatch"):
        asyncio.run(
            runner_module.run_platform(
                {"id": 9522, "keywords": [], "_trigger_source": "manual"},
                9522,
                "dy",
                tmp_path,
            )
        )

    assert calls == {"attempt": 1, "persist": 1}


@pytest.mark.parametrize(
    ("case", "expected_reason"),
    [
        ("locked_task_proxy_override", "account_identity_locked_proxy_override"),
        ("missing_profile", "account_identity_requires_relogin"),
        ("connect_existing", "account_identity_provider_unsupported"),
    ],
)
def test_phase_5_1d_runner_resolver_fails_before_spawn(case, expected_reason, tmp_path, monkeypatch):
    from api.monitoring import account_environment

    profile_root = (tmp_path / "runner-profiles").resolve()
    monkeypatch.setattr(account_environment, "ACCOUNT_PROFILE_ROOT", profile_root)
    monkeypatch.delenv("MONITOR_CDP_CONNECT_EXISTING", raising=False)
    account = _phase_5_1d_persisted_account(state="active", proxy_id=None)
    binding = {
        "account_id": account["id"],
        "platform": account["platform"],
        "profile_key": account["profile_key"],
        "profile_path": account["profile_path"],
        "proxy_id": None,
        "task_proxy_id": None,
        "_account": account,
        "_proxy": None,
    }
    derived = profile_root / "1" / "dy" / "acc_5101"
    if case != "missing_profile":
        derived.mkdir(parents=True)
    if case == "locked_task_proxy_override":
        binding["task_proxy_id"] = 99
    elif case == "connect_existing":
        monkeypatch.setenv("MONITOR_CDP_CONNECT_EXISTING", "true")

    with pytest.raises(ValueError) as exc_info:
        runner_module._resolve_runner_browser_plan(binding, "manual")
    assert getattr(exc_info.value, "reason", "") == expected_reason


def test_phase_5_1d_runner_ignores_stored_generic_profile_path(tmp_path, monkeypatch):
    from api.monitoring import account_environment

    profile_root = (tmp_path / "runner-profiles").resolve()
    monkeypatch.setattr(account_environment, "ACCOUNT_PROFILE_ROOT", profile_root)
    monkeypatch.delenv("MONITOR_CDP_CONNECT_EXISTING", raising=False)
    account = _phase_5_1d_persisted_account(state="active", proxy_id=None)
    account["profile_path"] = r"C:\generic\shared-profile"
    derived = profile_root / "1" / "dy" / "acc_5101"
    derived.mkdir(parents=True)
    binding = {
        "account_id": account["id"],
        "platform": account["platform"],
        "profile_key": account["profile_key"],
        "profile_path": account["profile_path"],
        "proxy_id": None,
        "task_proxy_id": None,
        "_account": account,
        "_proxy": None,
    }

    plan = runner_module._resolve_runner_browser_plan(binding, "manual")

    assert plan is not None
    assert Path(plan.profile_path) == derived
    assert plan.profile_path != account["profile_path"]


def test_phase_5_1d_cdp_prepares_exact_commands_before_navigation_and_writes_result(tmp_path, monkeypatch):
    from tools.browser_environment import (
        RESULT_PATH_ENV_NAME,
        bind_managed_context,
        prepare_managed_page,
        verify_managed_page,
    )

    plan = _phase_5_1d_plan(tmp_path, monkeypatch, action="crawl", launch_mode="cdp_launch")
    result_path = tmp_path / "cdp-result.json"
    monkeypatch.setenv(RESULT_PATH_ENV_NAME, str(result_path))
    events = []

    class FakeRequest:
        headers = {"accept-language": plan.accept_language}

    class FakeCDPSession:
        async def send(self, method, params):
            events.append((method, params))

    class FakeBrowser:
        version = plan.browser_version

    class FakeContext:
        browser = FakeBrowser()

        def on(self, event, callback):
            self.request_handler = callback

        async def new_cdp_session(self, page):
            return FakeCDPSession()

    context = FakeContext()

    class FakePage:
        async def goto(self, url):
            assert [event[0] for event in events] == [
                "Emulation.setUserAgentOverride",
                "Emulation.setTimezoneOverride",
                "Emulation.setLocaleOverride",
                "Emulation.setDeviceMetricsOverride",
                "Emulation.setTouchEmulationEnabled",
            ]
            events.append(("goto", url))
            context.request_handler(FakeRequest())

        async def evaluate(self, script):
            return _phase_5_1d_effective_probe(plan)

    page = FakePage()
    bind_managed_context(context, plan)
    asyncio.run(prepare_managed_page(context, page))
    asyncio.run(page.goto("https://platform.invalid"))
    result = asyncio.run(verify_managed_page(context, page))

    assert result is not None and result.ok is True
    assert result_path.is_file()
    stored = json.loads(result_path.read_text(encoding="utf-8"))
    assert stored["snapshot"]["resolution_id"] == plan.resolution_id
    assert events[0][1] == {
        "userAgent": plan.user_agent,
        "acceptLanguage": plan.accept_language,
        "platform": plan.browser_platform,
    }
    assert events[3][1] == {
        "width": plan.viewport_width,
        "height": plan.viewport_height,
        "deviceScaleFactor": plan.device_scale_factor,
        "mobile": plan.is_mobile,
        "screenWidth": plan.screen_width,
        "screenHeight": plan.screen_height,
    }


def test_cr116_persistent_context_uses_cdp_browser_version_proof(tmp_path, monkeypatch):
    from tools.browser_environment import launch_managed_browser_context, verify_managed_page

    plan = _phase_5_1d_plan(tmp_path, monkeypatch)
    events = []

    class FakeRequest:
        headers = {"accept-language": plan.accept_language}

    class FakeCDPSession:
        async def send(self, method):
            events.append(("send", method))
            return {"product": f"HeadlessChrome/{plan.browser_version}"}

        async def detach(self):
            events.append(("detach",))

    class FakePage:
        async def goto(self, *args, **kwargs):
            context.request_handler(FakeRequest())

        async def evaluate(self, script):
            return _phase_5_1d_effective_probe(plan)

    page = FakePage()

    class FakeContext:
        pages = []
        browser = None

        def on(self, event, callback):
            self.request_handler = callback

        async def new_cdp_session(self, used_page):
            assert used_page is page
            events.append(("attach",))
            return FakeCDPSession()

    context = FakeContext()

    class FakeChromium:
        async def launch_persistent_context(self, **kwargs):
            return context

    class FakePlaywright:
        chromium = FakeChromium()

    session = asyncio.run(launch_managed_browser_context(FakePlaywright(), plan))
    asyncio.run(page.goto("https://platform.invalid"))
    result = asyncio.run(verify_managed_page(session.context, page))

    assert result is not None and result.ok is True
    assert result.snapshot["browser"]["version"] == plan.browser_version
    assert events == [
        ("attach",),
        ("send", "Browser.getVersion"),
        ("detach",),
    ]


def test_cr116_persistent_context_rejects_malformed_cdp_browser_version():
    from tools.browser_environment import _effective_browser_version

    events = []

    class FakeCDPSession:
        async def send(self, method):
            events.append(("send", method))
            return {"product": "HeadlessChrome/not-a-version"}

        async def detach(self):
            events.append(("detach",))

    class FakeContext:
        browser = None

        async def new_cdp_session(self, page):
            events.append(("attach",))
            return FakeCDPSession()

    with pytest.raises(TypeError, match="missing browser version"):
        asyncio.run(_effective_browser_version(FakeContext(), object()))

    assert events == [
        ("attach",),
        ("send", "Browser.getVersion"),
        ("detach",),
    ]


def test_cr116_persistent_context_detach_failure_is_fail_closed():
    from tools.browser_environment import _effective_browser_version

    class FakeCDPSession:
        async def send(self, method):
            return {"product": "HeadlessChrome/127.0.6533.17"}

        async def detach(self):
            raise RuntimeError("synthetic detach failure")

    class FakeContext:
        browser = None

        async def new_cdp_session(self, page):
            return FakeCDPSession()

    with pytest.raises(RuntimeError, match="synthetic detach failure"):
        asyncio.run(_effective_browser_version(FakeContext(), object()))


@pytest.mark.parametrize(
    "failing_method",
    [
        "Emulation.setUserAgentOverride",
        "Emulation.setTimezoneOverride",
        "Emulation.setLocaleOverride",
        "Emulation.setDeviceMetricsOverride",
        "Emulation.setTouchEmulationEnabled",
    ],
)
def test_phase_5_1d_cdp_command_failure_blocks_navigation(failing_method, tmp_path, monkeypatch):
    from tools.browser_environment import (
        BrowserEnvironmentError,
        bind_managed_context,
        prepare_managed_page,
    )

    plan = _phase_5_1d_plan(tmp_path, monkeypatch, action="crawl", launch_mode="cdp_launch")
    navigations = []

    class FakeCDPSession:
        async def send(self, method, params):
            if method == failing_method:
                raise RuntimeError("synthetic CDP failure")

    class FakeContext:
        def on(self, event, callback):
            pass

        async def new_cdp_session(self, page):
            return FakeCDPSession()

    class FakePage:
        async def goto(self, url):
            navigations.append(url)

    context = FakeContext()
    page = FakePage()
    bind_managed_context(context, plan)

    with pytest.raises(BrowserEnvironmentError) as exc_info:
        asyncio.run(prepare_managed_page(context, page))
    assert exc_info.value.reason == "account_identity_provider_unsupported"
    assert exc_info.value.browser_environment_result.ok is False
    assert navigations == []


def test_phase_5_1d_context_and_page_bindings_survive_reused_object_ids(tmp_path, monkeypatch):
    from dataclasses import replace

    import tools.browser_environment as browser_environment

    first_plan = _phase_5_1d_plan(
        tmp_path,
        monkeypatch,
        action="crawl",
        launch_mode="cdp_launch",
    )
    second_plan = replace(
        first_plan,
        resolution_id="resolution-second",
        attempt_id="attempt-second",
        user_agent=f"{first_plan.user_agent} second",
    )
    sent_commands = {"first": [], "second": []}

    class FakeCDPSession:
        def __init__(self, label):
            self.label = label

        async def send(self, method, params):
            sent_commands[self.label].append((method, params))

    class FakeContext:
        def __init__(self, label):
            self.label = label

        def on(self, event, callback):
            pass

        async def new_cdp_session(self, page):
            return FakeCDPSession(self.label)

    class FakePage:
        pass

    first_context = FakeContext("first")
    second_context = FakeContext("second")
    first_page = FakePage()
    second_page = FakePage()

    monkeypatch.setattr(browser_environment, "id", lambda value: 1, raising=False)
    browser_environment.bind_managed_context(first_context, first_plan)
    browser_environment.bind_managed_context(second_context, second_plan)

    asyncio.run(browser_environment.prepare_managed_page(first_context, first_page))
    asyncio.run(browser_environment.prepare_managed_page(second_context, second_page))

    assert sent_commands["first"][0] == (
        "Emulation.setUserAgentOverride",
        {
            "userAgent": first_plan.user_agent,
            "acceptLanguage": first_plan.accept_language,
            "platform": first_plan.browser_platform,
        },
    )
    assert sent_commands["second"][0] == (
        "Emulation.setUserAgentOverride",
        {
            "userAgent": second_plan.user_agent,
            "acceptLanguage": second_plan.accept_language,
            "platform": second_plan.browser_platform,
        },
    )


def test_phase_5_1d_cdp_manager_launches_only_exact_managed_environment(tmp_path, monkeypatch):
    from tools.browser_environment import managed_proxy_formats, reset_browser_environment_cache_for_tests
    from tools.cdp_browser import CDPBrowserManager
    import tools.cdp_browser as cdp_browser_module

    plan = _phase_5_1d_plan(tmp_path, monkeypatch, proxy_bound=True, action="crawl", launch_mode="cdp_launch")
    Path(plan.profile_path).mkdir(parents=True, exist_ok=True)
    reset_browser_environment_cache_for_tests()
    monkeypatch.setenv(
        "MONITOR_BROWSER_ENVIRONMENT_PLAN",
        __import__("tools.browser_environment", fromlist=["browser_environment_plan_to_json"])
        .browser_environment_plan_to_json(plan),
    )
    parsed_plan = cdp_browser_module.plan_from_environment(required=True)
    assert parsed_plan == plan
    proxy, _ = managed_proxy_formats()
    captured = {}

    class FakeResponse:
        ok = True

        async def json(self):
            return {"region": plan.proxy_region}

    class FakeProofPage:
        async def goto(self, *args, **kwargs):
            captured["probe_before_return"] = True
            return FakeResponse()

        async def close(self):
            pass

    class FakeContext:
        pages = []

        def on(self, event, callback):
            self.request_handler = callback

        async def new_page(self):
            return FakeProofPage()

    context = FakeContext()

    class FakeBrowser:
        contexts = [context]
        version = plan.browser_version

        def is_connected(self):
            return True

    context.browser = FakeBrowser()

    class FakeLauncher:
        browser_process = None

        def detect_browser_paths(self):
            raise AssertionError("managed CDP must not auto-detect a browser")

        def find_available_port(self, start_port):
            return 9333

        def launch_browser(self, **kwargs):
            captured["launch"] = kwargs
            self.browser_process = object()
            return self.browser_process

        def wait_for_browser_ready(self, *args, **kwargs):
            return True

        def cleanup(self):
            pass

    manager = CDPBrowserManager(plan=plan)
    manager.launcher = FakeLauncher()
    original_sleep = asyncio.sleep
    monkeypatch.setattr(manager, "_register_cleanup_handlers", lambda: None)
    monkeypatch.setattr(manager, "_test_cdp_connection", lambda *args: original_sleep(0, result=True))

    async def fake_connect(playwright):
        manager.browser = context.browser

    monkeypatch.setattr(manager, "_connect_via_cdp", fake_connect)
    monkeypatch.setattr(cdp_browser_module.asyncio, "sleep", lambda *args: original_sleep(0))
    monkeypatch.setattr(cdp_browser_module.config, "CDP_CONNECT_EXISTING", False)

    returned = asyncio.run(
        manager.launch_and_connect(
            playwright=object(),
            playwright_proxy=proxy,
            user_agent="caller-must-not-win",
            headless=False,
        )
    )

    assert returned is context
    assert captured["probe_before_return"] is True
    assert captured["launch"] == {
        "browser_path": plan.browser_executable_path,
        "debug_port": 9333,
        "headless": plan.headless,
        "user_data_dir": plan.profile_path,
        "proxy_server": plan.proxy_url,
        "user_agent": plan.user_agent,
        "language": plan.locale,
        "window_size": (plan.viewport_width, plan.viewport_height),
        "device_scale_factor": plan.device_scale_factor,
        "managed": True,
    }
    reset_browser_environment_cache_for_tests()


@pytest.mark.parametrize(
    ("module_name", "class_name"),
    [
        ("media_platform.douyin.core", "DouYinCrawler"),
        ("media_platform.kuaishou.core", "KuaishouCrawler"),
        ("media_platform.xhs.core", "XiaoHongShuCrawler"),
        ("media_platform.bilibili.core", "BilibiliCrawler"),
        ("media_platform.weibo.core", "WeiboCrawler"),
        ("media_platform.tieba.core", "TieBaCrawler"),
        ("media_platform.zhihu.core", "ZhihuCrawler"),
    ],
)
def test_phase_5_1d_all_platforms_rethrow_managed_cdp_failure_without_standard_fallback(
    module_name,
    class_name,
    tmp_path,
    monkeypatch,
):
    import importlib

    from tools.browser_environment import BrowserEnvironmentError

    plan = _phase_5_1d_plan(tmp_path, monkeypatch, action="crawl", launch_mode="cdp_launch")
    module = importlib.import_module(module_name)
    monkeypatch.setattr(module, "plan_from_environment", lambda required=False: plan, raising=False)
    crawler = getattr(module, class_name)()
    crawler.browser_environment_plan = plan
    fallback_called = []

    class FailingManager:
        def __init__(self, *args, **kwargs):
            pass

        async def launch_and_connect(self, **kwargs):
            raise BrowserEnvironmentError("account_identity_provider_browser_crashed", "cdp")

    async def forbidden_standard(*args, **kwargs):
        fallback_called.append(True)
        return object()

    monkeypatch.setattr(module, "CDPBrowserManager", FailingManager)
    monkeypatch.setattr(crawler, "launch_browser", forbidden_standard)

    with pytest.raises(BrowserEnvironmentError) as exc_info:
        asyncio.run(crawler.launch_browser_with_cdp(object(), None, plan.user_agent, headless=True))
    assert exc_info.value.reason == "account_identity_provider_browser_crashed"
    assert fallback_called == []


@pytest.mark.parametrize(
    ("module_name", "class_name", "client_factory", "is_tieba"),
    [
        ("media_platform.douyin.core", "DouYinCrawler", "create_douyin_client", False),
        ("media_platform.kuaishou.core", "KuaishouCrawler", "create_ks_client", False),
        ("media_platform.xhs.core", "XiaoHongShuCrawler", "create_xhs_client", False),
        ("media_platform.bilibili.core", "BilibiliCrawler", "create_bilibili_client", False),
        ("media_platform.weibo.core", "WeiboCrawler", "create_weibo_client", False),
        ("media_platform.tieba.core", "TieBaCrawler", "create_tieba_client", True),
        ("media_platform.zhihu.core", "ZhihuCrawler", "create_zhihu_client", False),
    ],
)
@pytest.mark.parametrize("launch_mode", ["persistent_launch", "cdp_launch"])
def test_phase_5_1d_all_platforms_use_one_managed_plan_before_first_navigation(
    module_name,
    class_name,
    client_factory,
    is_tieba,
    launch_mode,
    tmp_path,
    monkeypatch,
):
    import importlib

    plan = _phase_5_1d_plan(
        tmp_path,
        monkeypatch,
        proxy_bound=True,
        action="crawl",
        launch_mode=launch_mode,
    )
    module = importlib.import_module(module_name)
    events = []
    captured = {"proxy_calls": 0, "pool_calls": 0}
    browser_proxy = {"server": "http://managed-proxy.invalid:8080"}
    http_proxy = "http://managed-proxy.invalid:8080"

    monkeypatch.setattr(module, "plan_from_environment", lambda required=False: plan, raising=False)
    monkeypatch.setattr(module.config, "ENABLE_IP_PROXY", True)
    monkeypatch.setattr(module.config, "ENABLE_CDP_MODE", launch_mode == "persistent_launch")
    monkeypatch.setattr(module.config, "CRAWLER_TYPE", "phase_5_1d_noop")
    monkeypatch.setattr(module.config, "HEADLESS", False)
    monkeypatch.setattr(module.config, "CDP_HEADLESS", False)

    def fake_managed_proxy_formats():
        captured["proxy_calls"] += 1
        return browser_proxy, http_proxy

    async def forbidden_create_ip_pool(*args, **kwargs):
        captured["pool_calls"] += 1
        raise AssertionError("managed crawl must not create a dynamic proxy pool")

    monkeypatch.setattr(module, "managed_proxy_formats", fake_managed_proxy_formats, raising=False)
    monkeypatch.setattr(module, "create_ip_pool", forbidden_create_ip_pool)

    class FakePage:
        async def goto(self, url, **kwargs):
            events.append(("navigate", url))

    page = FakePage()

    class FakeContext:
        async def new_page(self):
            return page

    context = FakeContext()

    class FakeManagedSession:
        browser = object()

        def __init__(self):
            self.context = context

    async def fake_launch_managed(playwright, used_plan):
        events.append(("managed_launch", used_plan))
        return FakeManagedSession()

    async def fake_prepare(used_context, used_page):
        assert used_context is context
        assert used_page is page
        events.append(("prepare",))

    async def fake_verify(used_context, used_page):
        assert used_context is context
        assert used_page is page
        events.append(("verify",))
        return None

    monkeypatch.setattr(module, "launch_managed_browser_context", fake_launch_managed, raising=False)
    monkeypatch.setattr(module, "prepare_managed_page", fake_prepare, raising=False)
    monkeypatch.setattr(module, "verify_managed_page", fake_verify, raising=False)

    class FakePlaywrightManager:
        async def __aenter__(self):
            return object()

        async def __aexit__(self, exc_type, exc, tb):
            return False

    monkeypatch.setattr(module, "async_playwright", lambda: FakePlaywrightManager())

    async def no_sleep(*args, **kwargs):
        return None

    monkeypatch.setattr(module.asyncio, "sleep", no_sleep)

    crawler = getattr(module, class_name)()
    captured["crawler"] = crawler

    async def fake_launch_cdp(playwright, used_proxy, user_agent, headless=True):
        captured["cdp_launch"] = {
            "proxy": used_proxy,
            "user_agent": user_agent,
            "headless": headless,
        }
        events.append(("cdp_launch",))
        return context

    monkeypatch.setattr(crawler, "launch_browser_with_cdp", fake_launch_cdp)

    class FakeClient:
        async def pong(self, *args, **kwargs):
            return True

        async def update_cookies(self, *args, **kwargs):
            return None

    async def fake_create_client(*args, **kwargs):
        captured["http_proxy"] = args[0]
        return FakeClient()

    monkeypatch.setattr(crawler, client_factory, fake_create_client)

    if is_tieba:
        async def fake_inject():
            return None

        async def fake_navigate():
            events.append(("navigate", crawler.index_url))

        monkeypatch.setattr(crawler, "_inject_anti_detection_scripts", fake_inject)
        monkeypatch.setattr(crawler, "_navigate_to_tieba_via_baidu", fake_navigate)

    asyncio.run(crawler.start())

    assert captured["proxy_calls"] == 1
    assert captured["pool_calls"] == 0
    assert captured["http_proxy"] == http_proxy
    assert crawler.browser_environment_plan == plan
    assert crawler.user_agent == plan.user_agent
    if class_name == "WeiboCrawler":
        assert crawler.mobile_user_agent == plan.user_agent

    event_names = [event[0] for event in events]
    assert event_names.index("prepare") < event_names.index("navigate") < event_names.index("verify")
    if launch_mode == "cdp_launch":
        assert "managed_launch" not in event_names
        assert captured["cdp_launch"] == {
            "proxy": browser_proxy,
            "user_agent": plan.user_agent,
            "headless": plan.headless,
        }
    else:
        assert "cdp_launch" not in event_names
        managed_event = next(event for event in events if event[0] == "managed_launch")
        assert managed_event[1] == plan


def test_phase_5_1d_all_platform_cores_use_shared_managed_adapters():
    core_paths = [
        Path("media_platform/douyin/core.py"),
        Path("media_platform/kuaishou/core.py"),
        Path("media_platform/xhs/core.py"),
        Path("media_platform/bilibili/core.py"),
        Path("media_platform/weibo/core.py"),
        Path("media_platform/tieba/core.py"),
        Path("media_platform/zhihu/core.py"),
    ]
    for path in core_paths:
        source = path.read_text(encoding="utf-8")
        assert "plan_from_environment" in source, path
        assert "managed_proxy_formats" in source, path
        assert "launch_managed_browser_context" in source, path
        assert "prepare_managed_page" in source, path
        assert "verify_managed_page" in source, path
        assert "if self.browser_environment_plan is None and config.ENABLE_IP_PROXY" in source, path


def test_phase_1_bootstrap_admin_login_session_and_user_management():
    from api.routers import auth as auth_router

    init_db()
    snapshots = {
        "audit_logs": _snapshot_table("audit_logs"),
        "user_sessions": _snapshot_table("user_sessions"),
        "users": _snapshot_table("users"),
    }
    try:
        with get_conn() as conn:
            conn.execute("DELETE FROM user_sessions")
            conn.execute("DELETE FROM audit_logs")
            conn.execute("DELETE FROM users")

        admin = bootstrap_admin_from_env("admin@example.com", "StrongPass123!", "Admin")
        assert admin
        assert admin["email"] == "admin@example.com"
        assert admin["role"] == "administrator"
        assert "password_hash" not in admin
        assert authenticate_user("admin@example.com", "StrongPass123!")["id"] == admin["id"]
        assert authenticate_user("admin@example.com", "wrong-password") is None

        login_response = _FakeResponse()
        login_result = asyncio.run(
            auth_router.login(
                {"email": "admin@example.com", "password": "StrongPass123!"},
                _FakeRequest(),
                login_response,
            )
        )
        session_cookie = login_response.cookies[SESSION_COOKIE_NAME]["value"]
        assert login_result["user"]["role"] == "administrator"
        assert login_result["user"]["menu_permissions"]["users_permissions"] is True
        assert session_cookie
        session_user = asyncio.run(auth_router.session(get_user_for_test_session(session_cookie)))["user"]
        assert session_user["email"] == "admin@example.com"

        normal = asyncio.run(
            auth_router.create_user(
                {
                    "email": "user1@example.com",
                    "display_name": "User One",
                    "password": "UserPass123!",
                    "role": "normal",
                },
                admin,
            )
        )["user"]
        assert normal["role"] == "normal"
        assert normal["status"] == "active"
        assert authenticate_user("user1@example.com", "UserPass123!")["role"] == "normal"

        disabled = asyncio.run(auth_router.update_user(int(normal["id"]), {"status": "disabled"}, admin))["user"]
        assert disabled["status"] == "disabled"
        assert authenticate_user("user1@example.com", "UserPass123!") is None
        assert any(user["email"] == "user1@example.com" for user in list_users())
        assert get_user_by_email("user1@example.com")["status"] == "disabled"

        logout_response = _FakeResponse()
        logout_result = asyncio.run(auth_router.logout(logout_response, admin, session_cookie))
        assert logout_result["ok"] is True
        assert SESSION_COOKIE_NAME in logout_response.deleted_cookies
    finally:
        _restore_table("user_sessions", snapshots["user_sessions"])
        _restore_table("audit_logs", snapshots["audit_logs"])
        _restore_table("users", snapshots["users"])


def test_phase_1_http_routes_enforce_sessions_roles_and_owner_scope():
    from api import main as api_main

    init_db()
    snapshots = {
        "reports": _snapshot_table("reports"),
        "crawl_runs": _snapshot_table("crawl_runs"),
        "raw_contents": _snapshot_table("raw_contents"),
        "ai_evaluations": _snapshot_table("ai_evaluations"),
        "audit_logs": _snapshot_table("audit_logs"),
        "user_sessions": _snapshot_table("user_sessions"),
        "users": _snapshot_table("users"),
    }
    jobs_snapshot = _snapshot_monitor_jobs()
    try:
        with get_conn() as conn:
            for table in ["ai_evaluations", "raw_contents", "reports", "crawl_runs", "user_sessions", "audit_logs", "users"]:
                conn.execute(f"DELETE FROM {table}")
        _clear_monitor_jobs()

        admin = bootstrap_admin_from_env("admin@example.com", "AdminPass123!", "Admin")
        user1 = save_user(
            {
                "email": "user1@example.com",
                "display_name": "User One",
                "password": "UserPass123!",
                "role": "normal",
            },
            actor_id=int(admin["id"]),
        )
        user2 = save_user(
            {
                "email": "user2@example.com",
                "display_name": "User Two",
                "password": "UserPass456!",
                "role": "normal",
            },
            actor_id=int(admin["id"]),
        )
        job1 = save_job(
            {
                "law_firm_name": "海安律所",
                "keywords": ["海安律所避雷"],
                "platforms": ["dy"],
                "recipients": [],
            },
            actor=user1,
        )
        job2 = save_job(
            {
                "law_firm_name": "恒泰律所",
                "keywords": ["恒泰律所投诉"],
                "platforms": ["dy"],
                "recipients": [],
            },
            actor=user2,
        )
        run1 = create_run(job1["id"], {"job_id": job1["id"], "law_firm_name": job1["law_firm_name"], "platforms": ["dy"]})
        run2 = create_run(job2["id"], {"job_id": job2["id"], "law_firm_name": job2["law_firm_name"], "platforms": ["dy"]})
        finish_run(run1, "success", {"job_id": job1["id"], "law_firm_name": job1["law_firm_name"], "platforms": ["dy"]})
        finish_run(run2, "success", {"job_id": job2["id"], "law_firm_name": job2["law_firm_name"], "platforms": ["dy"]})
        report1 = create_report(run1, job1, {"job_id": job1["id"], "law_firm_name": job1["law_firm_name"], "platforms": ["dy"]})
        report2 = create_report(run2, job2, {"job_id": job2["id"], "law_firm_name": job2["law_firm_name"], "platforms": ["dy"]})

        transport = httpx.ASGITransport(app=api_main.app)

        async def exercise() -> None:
            async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as anonymous:
                assert (await anonymous.get("/api/monitor/jobs")).status_code == 401

            async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as normal_client:
                login = await normal_client.post(
                    "/api/auth/login",
                    json={"email": "user1@example.com", "password": "UserPass123!"},
                )
                assert login.status_code == 200
                session = await normal_client.get("/api/auth/session")
                assert session.status_code == 200
                permissions = session.json()["user"]["menu_permissions"]
                assert permissions["overview"] is True
                assert permissions["platform_accounts"] is False
                assert (await normal_client.get("/api/monitor/social-accounts")).status_code == 403

                jobs_response = await normal_client.get("/api/monitor/jobs")
                assert jobs_response.status_code == 200
                job_ids = {item["id"] for item in jobs_response.json()["jobs"]}
                assert job_ids == {job1["id"]}
                assert (await normal_client.get(f"/api/monitor/jobs/{job2['id']}/preflight")).status_code == 404

                runs_response = await normal_client.get("/api/monitor/runs")
                assert {item["id"] for item in runs_response.json()["runs"]} == {run1}
                reports_response = await normal_client.get("/api/monitor/reports")
                assert {item["id"] for item in reports_response.json()["reports"]} == {report1["id"]}
                assert (await normal_client.get(f"/api/monitor/reports/{report2['id']}")).status_code == 404

            async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as admin_client:
                login = await admin_client.post(
                    "/api/auth/login",
                    json={"email": "admin@example.com", "password": "AdminPass123!"},
                )
                assert login.status_code == 200
                assert (await admin_client.get("/api/monitor/social-accounts")).status_code == 200
                jobs_response = await admin_client.get("/api/monitor/jobs")
                assert {item["id"] for item in jobs_response.json()["jobs"]} == {job1["id"], job2["id"]}
                reports_response = await admin_client.get("/api/monitor/reports")
                assert {item["id"] for item in reports_response.json()["reports"]} == {report1["id"], report2["id"]}

        asyncio.run(exercise())
    finally:
        with get_conn() as conn:
            user_ids = [
                row["id"]
                for row in conn.execute(
                    "SELECT id FROM users WHERE email IN ('admin@example.com','user1@example.com','user2@example.com')"
                ).fetchall()
            ]
            if user_ids:
                placeholders = ",".join("?" for _ in user_ids)
                run_ids = [
                    row["id"]
                    for row in conn.execute(
                        f"SELECT id FROM crawl_runs WHERE created_by IN ({placeholders})",
                        user_ids,
                    ).fetchall()
                ]
                if run_ids:
                    run_placeholders = ",".join("?" for _ in run_ids)
                    conn.execute(f"DELETE FROM ai_evaluations WHERE run_id IN ({run_placeholders})", run_ids)
                    conn.execute(f"DELETE FROM raw_contents WHERE run_id IN ({run_placeholders})", run_ids)
                    conn.execute(f"DELETE FROM reports WHERE run_id IN ({run_placeholders})", run_ids)
                    conn.execute(f"DELETE FROM crawl_runs WHERE id IN ({run_placeholders})", run_ids)
                conn.execute(f"DELETE FROM user_sessions WHERE user_id IN ({placeholders})", user_ids)
                conn.execute(f"DELETE FROM audit_logs WHERE user_id IN ({placeholders})", user_ids)
                conn.execute(f"DELETE FROM users WHERE id IN ({placeholders})", user_ids)
        _restore_table("users", snapshots["users"])
        _restore_monitor_jobs(jobs_snapshot)
        _restore_table("crawl_runs", snapshots["crawl_runs"])
        _restore_table("reports", snapshots["reports"])
        _restore_table("raw_contents", snapshots["raw_contents"])
        _restore_table("ai_evaluations", snapshots["ai_evaluations"])
        _restore_table("user_sessions", snapshots["user_sessions"])
        _restore_table("audit_logs", snapshots["audit_logs"])


def test_phase_13a_operations_home_data_layer_scopes_real_aggregates():
    from api import main as api_main

    init_db()
    snapshots = {
        "reports": _snapshot_table("reports"),
        "crawl_runs": _snapshot_table("crawl_runs"),
        "raw_contents": _snapshot_table("raw_contents"),
        "ai_evaluations": _snapshot_table("ai_evaluations"),
        "audit_logs": _snapshot_table("audit_logs"),
        "user_sessions": _snapshot_table("user_sessions"),
        "users": _snapshot_table("users"),
        "social_accounts": _snapshot_table("social_accounts"),
        "proxy_profiles": _snapshot_table("proxy_profiles"),
        "ai_key_profiles": _snapshot_table("ai_key_profiles"),
        "login_sessions": _snapshot_table("login_sessions"),
    }
    jobs_snapshot = _snapshot_monitor_jobs()
    try:
        with get_conn() as conn:
            for table in [
                "ai_evaluations",
                "raw_contents",
                "reports",
                "crawl_runs",
                "user_sessions",
                "audit_logs",
                "users",
                "social_accounts",
                "proxy_profiles",
                "ai_key_profiles",
                "login_sessions",
            ]:
                conn.execute(f"DELETE FROM {table}")
        _clear_monitor_jobs()

        admin = bootstrap_admin_from_env("phase13a-admin@example.com", "AdminPass123!", "Phase 13A Admin")
        user1 = save_user(
            {
                "email": "phase13a-user1@example.com",
                "display_name": "Phase 13A User One",
                "password": "UserPass123!",
                "role": "normal",
            },
            actor_id=int(admin["id"]),
        )
        user2 = save_user(
            {
                "email": "phase13a-user2@example.com",
                "display_name": "Phase 13A User Two",
                "password": "UserPass456!",
                "role": "normal",
            },
            actor_id=int(admin["id"]),
        )
        job1 = save_job(
            {
                "law_firm_name": "海安律所",
                "keywords": ["海安律所避雷"],
                "platforms": ["dy"],
                "recipients": ["ops@example.com"],
                "enabled": True,
            },
            actor=user1,
        )
        job2 = save_job(
            {
                "law_firm_name": "恒泰律所",
                "keywords": ["恒泰律所投诉"],
                "platforms": ["ks"],
                "recipients": ["ops@example.com"],
                "enabled": False,
            },
            actor=user2,
        )
        run1 = create_run(job1["id"], {"job_id": job1["id"], "law_firm_name": job1["law_firm_name"], "platforms": ["dy"]})
        run2 = create_run(job2["id"], {"job_id": job2["id"], "law_firm_name": job2["law_firm_name"], "platforms": ["ks"]})
        finish_run(run1, "success", {"job_id": job1["id"], "law_firm_name": job1["law_firm_name"], "platforms": ["dy"]})
        finish_run(run2, "failed", {"job_id": job2["id"], "law_firm_name": job2["law_firm_name"], "platforms": ["ks"]}, "平台登录态失效")
        report1 = create_report(
            run1,
            job1,
            {
                "job_id": job1["id"],
                "law_firm_name": job1["law_firm_name"],
                "platforms": ["dy"],
                "negative_count": 1,
                "high_count": 1,
                "pending_review_count": 0,
            },
        )
        report2 = create_report(
            run2,
            job2,
            {
                "job_id": job2["id"],
                "law_firm_name": job2["law_firm_name"],
                "platforms": ["ks"],
                "negative_count": 0,
                "high_count": 0,
                "pending_review_count": 1,
            },
        )
        with get_conn() as conn:
            now = datetime.now(timezone.utc).isoformat()
            rows = [
                (job1, run1, user1, "dy", "phase13a-user1-negative", "海安律所避雷", "海安律所退费争议"),
                (job2, run2, user2, "ks", "phase13a-user2-pending", "恒泰律所投诉", "恒泰律所沟通争议"),
            ]
            raw_ids = []
            for job, run_id, owner, platform, content_id, keyword, title in rows:
                conn.execute(
                    """
                    INSERT INTO raw_contents (
                        workspace_id, platform, content_id, job_id, run_id,
                        law_firm_name, source_keyword, title, description,
                        author_name, content_url, cover_url, publish_time,
                        comment_count, raw_json, first_seen_at, last_seen_at,
                        created_by, updated_by
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        1,
                        platform,
                        content_id,
                        job["id"],
                        run_id,
                        job["law_firm_name"],
                        keyword,
                        title,
                        "服务争议",
                        "用户",
                        "https://example.com/" + content_id,
                        "",
                        int(datetime.now(timezone.utc).timestamp()),
                        0,
                        "{}",
                        now,
                        now,
                        owner["id"],
                        owner["id"],
                    ),
                )
                raw_ids.append(int(conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]))
            conn.execute(
                """
                INSERT INTO ai_evaluations (
                    workspace_id, raw_content_id, run_id, status, is_related,
                    is_negative, risk_level, reason, evidence_quotes,
                    recommended_action, raw_response, created_at, created_by,
                    updated_by
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (1, raw_ids[0], run1, "ok", 1, 1, "high", "疑似负面", '["退费争议"]', "人工复核", "{}", now, user1["id"], user1["id"]),
            )
            conn.execute(
                """
                INSERT INTO ai_evaluations (
                    workspace_id, raw_content_id, run_id, status, is_related,
                    is_negative, risk_level, reason, evidence_quotes,
                    recommended_action, raw_response, created_at, created_by,
                    updated_by
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (1, raw_ids[1], run2, "pending_review", 0, 0, "low", "待人工复核", "[]", "人工复核", "{}", now, user2["id"], user2["id"]),
            )
            conn.execute("UPDATE reports SET email_status='sent', email_error='' WHERE id=?", (report1["id"],))
            conn.execute("UPDATE reports SET email_status='failed', email_error='SMTP 配置未完成' WHERE id=?", (report2["id"],))

        save_proxy_profile({"name": "Phase 13A Proxy", "provider": "manual", "proxy_url": "http://user:pass@example.com:8080", "status": "active"})
        save_social_account({"name": "Phase 13A Account", "platform": "dy", "login_type": "qrcode", "status": "active"})
        save_ai_key_profile({"name": "Phase 13A AI", "provider": "openai", "base_url": "https://api.example.com", "api_key": "sk-test", "model": "test", "is_active": True})
        create_login_session({"platform": "dy", "account_id": None, "login_url": "https://www.douyin.com/"})

        admin_summary = get_dashboard_summary(actor=admin)
        user1_summary = get_dashboard_summary(actor=user1)
        user2_summary = get_dashboard_summary(actor=user2)

        assert admin_summary["jobs_total"] == 2
        assert admin_summary["operations_home"]["task_health"]["total"] == 2
        assert admin_summary["operations_home"]["run_activity"]["failed_recent"] == 1
        assert admin_summary["operations_home"]["report_activity"]["total"] == 2
        assert admin_summary["operations_home"]["email_delivery"]["failed"] == 1
        assert admin_summary["operations_home"]["lead_metrics"]["suspected_negative"] == 1
        assert admin_summary["operations_home"]["lead_metrics"]["pending_review"] == 1
        assert admin_summary["operations_home"]["resource_health"]["scope"] == "workspace"
        assert admin_summary["operations_home"]["resource_health"]["social_accounts_total"] == 1

        assert user1_summary["operations_home"]["task_health"]["total"] == 1
        assert user1_summary["operations_home"]["run_activity"]["failed_recent"] == 0
        assert user1_summary["operations_home"]["report_activity"]["total"] == 1
        assert user1_summary["operations_home"]["email_delivery"]["sent"] == 1
        assert user1_summary["operations_home"]["email_delivery"]["failed"] == 0
        assert user1_summary["operations_home"]["lead_metrics"]["suspected_negative"] == 1
        assert user1_summary["operations_home"]["lead_metrics"]["pending_review"] == 0
        assert user1_summary["operations_home"]["resource_health"]["scope"] == "business_safe"
        assert "social_accounts_total" not in user1_summary["operations_home"]["resource_health"]

        assert user2_summary["operations_home"]["task_health"]["active"] == 0
        assert user2_summary["operations_home"]["task_health"]["paused"] == 1
        assert user2_summary["operations_home"]["run_activity"]["failed_recent"] == 1
        assert user2_summary["operations_home"]["email_delivery"]["failed"] == 1
        assert user2_summary["operations_home"]["lead_metrics"]["pending_review"] == 1

        transport = httpx.ASGITransport(app=api_main.app)

        async def exercise_dashboard_contract() -> None:
            async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as normal_client:
                login = await normal_client.post(
                    "/api/auth/login",
                    json={"email": "phase13a-user1@example.com", "password": "UserPass123!"},
                )
                assert login.status_code == 200
                response = await normal_client.get("/api/monitor/dashboard")
                assert response.status_code == 200
                payload = response.json()
                assert payload["summary"]["jobs_total"] == 1
                assert payload["operations_home"] == payload["summary"]["operations_home"]
                assert payload["operations_home"]["task_health"]["total"] == 1
                assert payload["operations_home"]["resource_health"]["scope"] == "business_safe"
                assert "social_accounts_total" not in payload["operations_home"]["resource_health"]

        asyncio.run(exercise_dashboard_contract())
    finally:
        _restore_monitor_jobs(jobs_snapshot)
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_phase_4_normal_user_task_api_ignores_advanced_resource_fields():
    from api import main as api_main

    init_db()
    snapshots = {
        "users": _snapshot_table("users"),
        "user_sessions": _snapshot_table("user_sessions"),
        "monitor_jobs": _snapshot_table("monitor_jobs"),
        "job_keywords": _snapshot_table("job_keywords"),
        "job_platforms": _snapshot_table("job_platforms"),
        "job_recipients": _snapshot_table("job_recipients"),
        "ai_key_profiles": _snapshot_table("ai_key_profiles"),
        "email_templates": _snapshot_table("email_templates"),
        "proxy_profiles": _snapshot_table("proxy_profiles"),
        "social_accounts": _snapshot_table("social_accounts"),
    }
    try:
        with get_conn() as conn:
            for table in ["job_recipients", "job_platforms", "job_keywords", "monitor_jobs", "user_sessions", "users"]:
                conn.execute(f"DELETE FROM {table}")

        admin = save_user(
            {
                "email": "phase4-admin@example.com",
                "display_name": "Phase4 Admin",
                "password": "AdminPass123!",
                "role": "administrator",
            }
        )
        save_user(
            {
                "email": "phase4-user@example.com",
                "display_name": "Phase4 User",
                "password": "UserPass123!",
                "role": "normal",
            },
            actor_id=int(admin["id"]),
        )
        profile = save_ai_key_profile(
            {
                "name": "Phase4 AI",
                "provider": "openai",
                "base_url": "https://example.com",
                "api_key": "sk-phase4",
                "model": "phase4-model",
            }
        )
        template = save_email_template({"name": "Phase4 Template", "subject_template": "日报 {law_firm_name}", "html_template": "{report_body}"})
        proxy = save_proxy_profile({"name": "Phase4 Proxy", "provider": "manual", "proxy_url": "http://user:pass@127.0.0.1:8081"})
        account = save_social_account({"name": "Phase4 Account", "platform": "dy", "status": "active", "proxy_id": proxy["id"]})
        payload = {
            "law_firm_name": "海安律所",
            "aliases": ["海安律师事务所"],
            "keywords": ["海安律所避雷", "海安律所退费"],
            "platforms": ["dy"],
            "recipients": ["target@example.com"],
            "enable_comments": True,
            "enable_sub_comments": True,
            "time_window_type": "recent_7d",
            "frequency": "daily",
            "email_time": "09:00",
            "max_items": 80,
            "start_page": 2,
            "max_pages": 3,
            "target_type": "detail",
            "output_mode": "excel",
            "browser_mode": "local_window",
            "ai_profile_id": profile["id"],
            "email_template_id": template["id"],
            "account_id": account["id"],
            "proxy_id": proxy["id"],
        }
        transport = httpx.ASGITransport(app=api_main.app)

        async def exercise() -> None:
            async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as normal_client:
                login = await normal_client.post(
                    "/api/auth/login",
                    json={"email": "phase4-user@example.com", "password": "UserPass123!"},
                )
                assert login.status_code == 200
                created = await normal_client.post("/api/monitor/jobs", json=payload)
                assert created.status_code == 200
                normal_job = created.json()["job"]
                assert normal_job["target_type"] == "search"
                assert normal_job["output_mode"] == "internal"
                assert normal_job["browser_mode"] == "server_qrcode"
                assert normal_job["ai_profile_id"] is None
                assert normal_job["email_template_id"] is None
                assert normal_job["account_id"] is None
                assert normal_job["proxy_id"] is None
                assert normal_job["max_items"] == 80
                assert normal_job["start_page"] == 2
                assert normal_job["max_pages"] == 3

            async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as admin_client:
                login = await admin_client.post(
                    "/api/auth/login",
                    json={"email": "phase4-admin@example.com", "password": "AdminPass123!"},
                )
                assert login.status_code == 200
                created = await admin_client.post("/api/monitor/jobs", json=payload)
                assert created.status_code == 200
                admin_job = created.json()["job"]
                assert admin_job["target_type"] == "detail"
                assert admin_job["output_mode"] == "excel"
                assert admin_job["browser_mode"] == "local_window"
                assert admin_job["ai_profile_id"] == profile["id"]
                assert admin_job["email_template_id"] == template["id"]
                assert admin_job["account_id"] == account["id"]
                assert admin_job["proxy_id"] == proxy["id"]

        asyncio.run(exercise())
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_phase_2_runtime_settings_storage_validation_and_environment_locks(monkeypatch):
    init_db()
    snapshot = _snapshot_table("system_settings")
    audit_snapshot = _snapshot_table("audit_logs")
    try:
        with get_conn() as conn:
            conn.execute("DELETE FROM system_settings")
            conn.execute("DELETE FROM audit_logs")

        settings = list_runtime_settings()
        assert settings["global_crawl_concurrency"]["value"] == 2
        assert settings["crawler_timeout_seconds"]["value"] == 900
        assert settings["scheduler_disabled"]["value"] is False

        updated = save_runtime_settings(
            {
                "global_crawl_concurrency": 3,
                "crawler_timeout_seconds": 1200,
                "scheduler_disabled": True,
            },
            actor_id=123,
        )
        assert updated["global_crawl_concurrency"]["value"] == 3
        assert get_runtime_setting_value("crawler_timeout_seconds") == 1200
        assert updated["scheduler_disabled"]["source"] == "database"
        with get_conn() as conn:
            audit = conn.execute("SELECT action_type, resource_type FROM audit_logs ORDER BY id DESC LIMIT 1").fetchone()
        assert dict(audit) == {"action_type": "update_runtime_settings", "resource_type": "system_settings"}

        with pytest.raises(ValueError, match="at most"):
            save_runtime_settings({"global_crawl_concurrency": 99})

        monkeypatch.setenv("MONITOR_CRAWLER_TIMEOUT_SECONDS", "1800")
        locked = list_runtime_settings()["crawler_timeout_seconds"]
        assert locked["value"] == 1800
        assert locked["is_locked"] is True
        assert locked["source"] == "environment"
        with pytest.raises(ValueError, match="locked"):
            save_runtime_settings({"crawler_timeout_seconds": 900})
    finally:
        _restore_table("system_settings", snapshot)
        _restore_table("audit_logs", audit_snapshot)


def test_phase_2_runtime_settings_api_is_admin_only():
    from api import main as api_main

    init_db()
    snapshots = {
        "audit_logs": _snapshot_table("audit_logs"),
        "system_settings": _snapshot_table("system_settings"),
        "user_sessions": _snapshot_table("user_sessions"),
        "users": _snapshot_table("users"),
    }
    try:
        with get_conn() as conn:
            for table in ["audit_logs", "system_settings", "user_sessions", "users"]:
                conn.execute(f"DELETE FROM {table}")
        admin = bootstrap_admin_from_env("admin@example.com", "AdminPass123!", "Admin")
        save_user(
            {
                "email": "user1@example.com",
                "display_name": "User One",
                "password": "UserPass123!",
                "role": "normal",
            },
            actor_id=int(admin["id"]),
        )
        transport = httpx.ASGITransport(app=api_main.app)

        async def exercise() -> None:
            async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as normal_client:
                login = await normal_client.post("/api/auth/login", json={"email": "user1@example.com", "password": "UserPass123!"})
                assert login.status_code == 200
                assert (await normal_client.get("/api/monitor/runtime-settings")).status_code == 403

            async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as admin_client:
                login = await admin_client.post("/api/auth/login", json={"email": "admin@example.com", "password": "AdminPass123!"})
                assert login.status_code == 200
                get_response = await admin_client.get("/api/monitor/runtime-settings")
                assert get_response.status_code == 200
                assert get_response.json()["settings"]["scheduler_tick_seconds"]["value"] == 60
                put_response = await admin_client.put(
                    "/api/monitor/runtime-settings",
                    json={"scheduler_tick_seconds": 30, "crawler_retry_count": 2},
                )
                assert put_response.status_code == 200
                data = put_response.json()["settings"]
                assert data["scheduler_tick_seconds"]["value"] == 30
                assert data["crawler_retry_count"]["value"] == 2

        asyncio.run(exercise())
    finally:
        _restore_table("users", snapshots["users"])
        _restore_table("user_sessions", snapshots["user_sessions"])
        _restore_table("system_settings", snapshots["system_settings"])
        _restore_table("audit_logs", snapshots["audit_logs"])


def test_phase_2_scheduler_uses_runtime_settings(monkeypatch):
    init_db()
    snapshot = _snapshot_table("system_settings")
    try:
        with get_conn() as conn:
            conn.execute("DELETE FROM system_settings")
        save_runtime_settings({"scheduler_tick_seconds": 30})
        monkeypatch.delenv("MONITOR_DISABLE_SCHEDULER", raising=False)
        monkeypatch.delenv("WEB_CONCURRENCY", raising=False)
        monkeypatch.delenv("UVICORN_WORKERS", raising=False)

        status = scheduler_status()

        assert status["enabled"] is True
        assert status["tick_seconds"] == 30
        assert "30 秒" in status["message"]

        save_runtime_settings({"scheduler_disabled": True})

        assert "暂停自动调度" in scheduler_disabled_reason()
        assert scheduler_status()["enabled"] is False
    finally:
        _restore_table("system_settings", snapshot)


def test_phase_2_qrcode_timeout_and_ttl_use_runtime_settings():
    init_db()
    snapshot = _snapshot_table("system_settings")
    try:
        with get_conn() as conn:
            conn.execute("DELETE FROM system_settings")
        save_runtime_settings({"login_qr_timeout_seconds": 12, "login_session_ttl_seconds": 60})
        handle = login_qrcode_module.LoginSessionHandle(
            platform="dy",
            playwright=object(),
            context=object(),
            page=object(),
            profile_path="browser_data/test",
            created_at=datetime.now(timezone.utc) - timedelta(seconds=61),
        )

        assert login_qrcode_module._login_qr_timeout_ms() == 12000
        assert login_qrcode_module._session_expired(handle) is True
    finally:
        _restore_table("system_settings", snapshot)


def test_phase_2_run_platform_uses_settings_for_retry_and_deadline(tmp_path, monkeypatch):
    init_db()
    snapshot = _snapshot_table("system_settings")
    job = save_job(
        {
            "law_firm_name": "运行策略测试律所",
            "keywords": ["运行策略测试律所避雷"],
            "platforms": ["dy"],
            "recipients": [],
            "enable_comments": False,
            "time_window_type": "recent_1d",
        }
    )
    run_id = create_run(job["id"], {"job_id": job["id"]}, timeout_seconds=120)
    calls: list[dict[str, Any]] = []

    def fake_run_attempt(job_arg, platform_arg, out_dir, proxy_binding=None):
        calls.append({"timeout": job_arg.get("_crawler_timeout_seconds"), "out_dir": out_dir})
        if len(calls) == 1:
            raise RuntimeError("temporary network error")
        json_dir = out_dir / "douyin" / "json"
        json_dir.mkdir(parents=True)
        (json_dir / "search_contents_runtime_settings.json").write_text(
            json.dumps(
                [
                    {
                        "aweme_id": "pytest_runtime_settings_retry_001",
                        "title": "运行策略测试律所避雷",
                        "desc": "第二次成功",
                        "create_time": int(datetime.now(timezone.utc).timestamp()),
                    }
                ],
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )

    try:
        with get_conn() as conn:
            conn.execute("DELETE FROM system_settings")
        save_runtime_settings({"crawler_retry_count": 1, "crawler_retry_delay_seconds": 0})
        monkeypatch.setattr(runner_module, "list_platform_status", lambda: [{"platform": "dy", "login_window_open": False}])
        monkeypatch.setattr(runner_module, "_run_crawler_attempt", fake_run_attempt)

        result = asyncio.run(runner_module.run_platform(job, run_id, "dy", tmp_path))

        assert len(calls) == 2
        assert all(1 <= int(call["timeout"]) <= 120 for call in calls)
        assert result["attempts"] == 2
        assert result["max_retries"] == 1
        assert result["timeout_seconds"] == 120
        assert result["new_contents"] == 1
    finally:
        _cleanup_test_records(job["id"], "pytest_runtime_settings_retry_001")
        _restore_table("system_settings", snapshot)


def test_phase_2_run_job_marks_run_timeout_with_partial_results(monkeypatch):
    init_db()
    snapshots = {
        "reports": _snapshot_table("reports"),
        "crawl_runs": _snapshot_table("crawl_runs"),
        "raw_contents": _snapshot_table("raw_contents"),
        "raw_comments": _snapshot_table("raw_comments"),
        "ai_evaluations": _snapshot_table("ai_evaluations"),
        "system_settings": _snapshot_table("system_settings"),
    }
    jobs_snapshot = _snapshot_monitor_jobs()
    try:
        with get_conn() as conn:
            for table in ["reports", "crawl_runs", "raw_contents", "raw_comments", "ai_evaluations", "system_settings"]:
                conn.execute(f"DELETE FROM {table}")
        save_runtime_settings({"crawler_timeout_seconds": 60})
        job = save_job(
            {
                "law_firm_name": "超时测试律所",
                "keywords": ["超时测试律所避雷"],
                "platforms": ["dy", "ks"],
                "recipients": [],
                "enable_comments": False,
                "time_window_type": "recent_1d",
            }
        )

        async def fake_run_platform(job_arg, run_id, platform, run_dir):
            if platform == "dy":
                return {
                    "status": "success",
                    "raw_contents": 1,
                    "filtered_contents": 1,
                    "excluded_contents": 0,
                    "new_contents": 1,
                    "content_db_ids": [],
                }
            raise runner_module.CrawlerTimedOut("任务达到系统运行时间上限（60 秒），已停止未完成的采集进程")

        monkeypatch.setattr(runner_module, "run_platform", fake_run_platform)
        monkeypatch.setattr(runner_module, "create_report", lambda run_id, job, summary: {"id": 9876, "run_id": run_id, "job_id": job["id"], "summary": summary})
        monkeypatch.setattr(
            runner_module,
            "send_report_with_delivery_log",
            lambda job, report, send_type="auto": (False, "未配置收件人", report, None),
        )

        result = asyncio.run(runner_module.run_job(job["id"]))
        run = get_run(result["run_id"])

        assert result["status"] == "timeout"
        assert run["status"] == "timeout"
        assert run["timeout_seconds"] == 60
        assert run["deadline_at"]
        assert "系统运行时间上限" in (run["timeout_reason"] or "")
        assert result["summary"]["new_contents"] == 1
        assert result["summary"]["platform_results"]["dy"]["status"] == "success"
        assert result["summary"]["platform_results"]["ks"]["status"] == "timeout"
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)
        _restore_monitor_jobs(jobs_snapshot)


def test_custom_window_and_millisecond_timestamps():
    start, end = resolve_window({"time_window_type": "custom", "custom_start": "2026-06-10", "custom_end": "2026-06-11"})
    assert start.isoformat().startswith("2026-06-10T00:00:00")
    assert end.isoformat().startswith("2026-06-11T23:59:59")
    publish_ms = int(datetime(2026, 6, 11, 12, tzinfo=timezone.utc).timestamp() * 1000)
    assert in_time_window({"publish_time": publish_ms}, {"time_window_type": "custom", "custom_start": "2026-06-11", "custom_end": "2026-06-11"})


def test_platform_normalization_keeps_cover_and_keyword():
    job = {"law_firm_name": "测试律所", "keywords": ["测试律所避雷"]}
    xhs = normalize_content(
        "xhs",
        {
            "note_id": "x1",
            "title": "测试律所避雷",
            "desc": "退费争议",
            "note_url": "https://example.com/xhs",
            "image_list": '[{"url":"https://example.com/cover.jpg"}]',
            "time": 1781180000,
        },
        job,
    )
    assert xhs
    assert xhs["cover_url"] == "https://example.com/cover.jpg"
    assert xhs["source_keyword"] == "测试律所避雷"


def test_collect_platform_outputs_supports_json_and_jsonl(tmp_path):
    json_dir = tmp_path / "douyin" / "json"
    jsonl_dir = tmp_path / "douyin" / "jsonl"
    json_dir.mkdir(parents=True)
    jsonl_dir.mkdir(parents=True)
    (json_dir / "search_contents_2026-06-12.json").write_text('[{"aweme_id":"json_1"}]', encoding="utf-8")
    (jsonl_dir / "search_contents_2026-06-12.jsonl").write_text('{"aweme_id":"jsonl_1"}\nnot-json\n{"aweme_id":"jsonl_2"}\n', encoding="utf-8")
    (jsonl_dir / "search_comments_2026-06-12.jsonl").write_text('{"comment_id":"c1","aweme_id":"jsonl_1"}\n', encoding="utf-8")

    contents, comments = collect_platform_outputs(tmp_path, "dy")

    assert [item["aweme_id"] for item in contents] == ["json_1", "jsonl_1", "jsonl_2"]
    assert parse_jsonl_file(jsonl_dir / "search_contents_2026-06-12.jsonl")[0]["aweme_id"] == "jsonl_1"
    assert comments[0]["comment_id"] == "c1"


def test_platform_status_reports_profile_and_login_error(tmp_path):
    profile = tmp_path / "browser_data" / "cdp_dy_user_data_dir"
    profile.mkdir(parents=True)
    state = profile / "state"
    state.write_text("ok", encoding="utf-8")
    profile_time = datetime.now(timezone.utc) - timedelta(minutes=10)
    run_time = datetime.now(timezone.utc)
    os.utime(state, (profile_time.timestamp(), profile_time.timestamp()))
    statuses = list_platform_status(
        tmp_path,
        [
            {
                "finished_at": run_time.isoformat(),
                "summary": {
                    "platform_results": {
                        "dy": {"error": "MediaCrawler exited with 1；检测到登录态失效，请先重新登录该平台账号"}
                    }
                }
            }
        ],
    )
    dy = next(item for item in statuses if item["platform"] == "dy")
    ks = next(item for item in statuses if item["platform"] == "ks")
    assert dy["profile_exists"] is True
    assert dy["needs_login"] is True
    assert ks["profile_exists"] is False


def test_platform_status_ignores_login_error_older_than_successful_login_session(tmp_path):
    init_db()
    snapshots = {
        "login_sessions": _snapshot_table("login_sessions"),
        "platform_login_configs": _snapshot_table("platform_login_configs"),
    }
    profile = tmp_path / "browser_data" / "cdp_dy_user_data_dir"
    profile.mkdir(parents=True)
    state = profile / "state"
    state.write_text("ok", encoding="utf-8")
    error_time = datetime(2026, 6, 12, 9, 0, tzinfo=timezone.utc)
    try:
        session = create_login_session(
            {
                "platform": "dy",
                "login_url": "https://www.douyin.com/",
                "profile_path": str(profile),
            }
        )
        monitor_router.update_login_session_status(
            int(session["id"]),
            "success",
            "登录成功，Profile 已保存。",
        )
        with get_conn() as conn:
            conn.execute(
                "UPDATE login_sessions SET updated_at=? WHERE id=?",
                ((error_time + timedelta(minutes=10)).isoformat(), session["id"]),
            )

        statuses = list_platform_status(
            tmp_path,
            [
                {
                    "finished_at": error_time.isoformat(),
                    "summary": {
                        "platform_results": {
                            "dy": {"error": "MediaCrawler exited with 1；检测到登录态失效，请先重新登录该平台账号"}
                        }
                    },
                }
            ],
        )
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)
    dy = next(item for item in statuses if item["platform"] == "dy")

    assert dy["profile_exists"] is True
    assert dy["needs_login"] is False
    assert dy["last_error"] == ""


def test_platform_status_keeps_fresh_login_error_when_browser_profile_was_touched(tmp_path):
    init_db()
    snapshot = _snapshot_table("platform_login_configs")
    profile = tmp_path / "browser_data" / "cdp_ks_user_data_dir"
    profile.mkdir(parents=True)
    state = profile / "state"
    state.write_text("browser touched during failed login", encoding="utf-8")
    error_time = datetime.now(timezone.utc) - timedelta(minutes=5)
    touched_time = error_time + timedelta(seconds=5)
    os.utime(state, (touched_time.timestamp(), touched_time.timestamp()))
    try:
        with get_conn() as conn:
            conn.execute(
                "UPDATE platform_login_configs SET updated_at=? WHERE platform='ks'",
                ((error_time - timedelta(minutes=1)).isoformat(),),
            )
        statuses = list_platform_status(
            tmp_path,
            [
                {
                    "finished_at": error_time.isoformat(),
                    "summary": {
                        "platform_results": {
                            "ks": {
                                "error": "MediaCrawler exited with 1；检测到登录态失效，请先重新登录该平台账号"
                            }
                        }
                    },
                }
            ],
        )
    finally:
        _restore_table("platform_login_configs", snapshot)
    ks = next(item for item in statuses if item["platform"] == "ks")

    assert ks["profile_exists"] is True
    assert ks["needs_login"] is True
    assert "登录态失效" in ks["last_error"]


def test_platform_status_ignores_login_error_older_than_cookie_config(tmp_path, monkeypatch):
    init_db()
    snapshot = _snapshot_table("platform_login_configs")
    try:
        save_platform_login_config("dy", {"login_type": "cookie", "cookies": "sessionid=secret-cookie"})
        with get_conn() as conn:
            conn.execute(
                "UPDATE platform_login_configs SET updated_at=? WHERE platform='dy'",
                ("2026-06-12T09:10:00+00:00",),
            )
        statuses = list_platform_status(
            tmp_path,
            [
                {
                    "finished_at": "2026-06-12T09:00:00+00:00",
                    "summary": {
                        "platform_results": {
                            "dy": {"error": "MediaCrawler exited with 1；检测到登录态失效，请先重新登录该平台账号"}
                        }
                    },
                }
            ],
        )
    finally:
        _restore_table("platform_login_configs", snapshot)

    dy = next(item for item in statuses if item["platform"] == "dy")

    assert dy["login_type"] == "cookie"
    assert dy["has_cookies"] is True
    assert dy["profile_exists"] is False
    assert dy["needs_login"] is False
    assert dy["last_error"] == ""


def test_platform_status_ignores_legacy_phone_login_config(tmp_path):
    init_db()
    snapshot = _snapshot_table("platform_login_configs")
    try:
        with get_conn() as conn:
            conn.execute(
                """
                INSERT INTO platform_login_configs (platform, login_type, cookies_encrypted, login_phone_encrypted, updated_at)
                VALUES ('xhs', 'phone', '', '', ?)
                ON CONFLICT(platform) DO UPDATE SET login_type='phone', login_phone_encrypted='', updated_at=excluded.updated_at
                """,
                ("2026-06-13T08:00:00+00:00",),
            )
        statuses = list_platform_status(tmp_path, [])
    finally:
        _restore_table("platform_login_configs", snapshot)

    xhs = next(item for item in statuses if item["platform"] == "xhs")

    assert xhs["login_type"] == "qrcode"
    assert "has_login_phone" not in xhs
    assert "login_phone" not in xhs
    assert xhs["login_material_ready"] is False
    assert "网页登录态" in xhs["login_material_error"]
    assert xhs["needs_login"] is True
    assert xhs["login_ready"] is False


def test_platform_status_clears_closed_login_window_error(tmp_path, monkeypatch):
    profile = tmp_path / "browser_data" / "cdp_dy_user_data_dir"
    profile.mkdir(parents=True)
    (profile / "state").write_text("ok", encoding="utf-8")
    monkeypatch.setattr("api.monitoring.platform_status.login_window_status", lambda platform: {"is_open": False})
    statuses = list_platform_status(
        tmp_path,
        [
            {
                "finished_at": datetime.now(timezone.utc).isoformat(),
                "summary": {
                    "platform_results": {
                        "dy": {"error": "抖音登录窗口未关闭，请关闭窗口后再运行采集"}
                    }
                },
            }
        ],
    )
    dy = next(item for item in statuses if item["platform"] == "dy")

    assert dy["last_error"] == ""
    assert dy["needs_login"] is False


def test_platform_status_reports_open_login_window(tmp_path, monkeypatch):
    browser_data = tmp_path / "profiles"
    (browser_data / "cdp_dy_user_data_dir").mkdir(parents=True)
    monkeypatch.setenv("MONITOR_BROWSER_DATA_DIR", str(browser_data))
    monkeypatch.setattr("api.monitoring.login_state.LOGIN_STATE_DIR", tmp_path / "login_windows")
    monkeypatch.setattr("api.monitoring.login_state._pid_exists", lambda pid: pid == 12345)
    record_login_window("dy", 12345, 9323, str(browser_data / "cdp_dy_user_data_dir"))

    dy = next(item for item in list_platform_status(tmp_path, []) if item["platform"] == "dy")

    assert dy["login_window_open"] is True
    assert dy["login_window_pid"] == 12345


def test_login_window_status_removes_stale_pid_record(tmp_path, monkeypatch):
    monkeypatch.setattr("api.monitoring.login_state.LOGIN_STATE_DIR", tmp_path / "login_windows")
    monkeypatch.setattr("api.monitoring.login_state._pid_exists", lambda pid: False)
    record_login_window("dy", 12345, 9323, str(tmp_path / "profile"))

    status = login_window_status("dy")

    assert status["is_open"] is False
    assert status["pid"] is None
    assert status["opened_at"]
    assert status["closed_at"]
    assert (tmp_path / "login_windows" / "dy.json").exists()


def test_platform_status_clears_login_error_after_closed_login_window_profile_update(tmp_path, monkeypatch):
    login_state_dir = tmp_path / "login_windows"
    monkeypatch.setattr("api.monitoring.login_state.LOGIN_STATE_DIR", login_state_dir)
    monkeypatch.setattr("api.monitoring.platform_status.LOGIN_STATE_DIR", login_state_dir, raising=False)
    monkeypatch.setattr("api.monitoring.login_state._pid_exists", lambda pid: False)
    browser_data = tmp_path / "browser_data"
    profile = browser_data / "cdp_ks_user_data_dir"
    profile.mkdir(parents=True)
    state = profile / "state"
    state.write_text("manual login refreshed profile", encoding="utf-8")
    monkeypatch.setenv("MONITOR_BROWSER_DATA_DIR", str(browser_data))
    error_time = datetime.now(timezone.utc) - timedelta(minutes=5)
    opened_at = error_time + timedelta(minutes=1)
    refreshed_at = opened_at + timedelta(minutes=1)
    closed_at = refreshed_at + timedelta(minutes=1)
    os.utime(state, (refreshed_at.timestamp(), refreshed_at.timestamp()))
    login_state_dir.mkdir(parents=True)
    (login_state_dir / "ks.json").write_text(
        json.dumps(
            {
                "platform": "ks",
                "pid": 12345,
                "debug_port": 9324,
                "profile_path": str(profile),
                "opened_at": opened_at.isoformat(),
                "closed_at": closed_at.isoformat(),
            }
        ),
        encoding="utf-8",
    )

    statuses = list_platform_status(
        tmp_path,
        [
            {
                "finished_at": error_time.isoformat(),
                "summary": {
                    "platform_results": {
                        "ks": {"error": "MediaCrawler exited with 1；检测到登录态失效，请先重新登录该平台账号"}
                    }
                },
            }
        ],
    )
    ks = next(item for item in statuses if item["platform"] == "ks")

    assert ks["profile_exists"] is True
    assert ks["needs_login"] is False
    assert ks["last_error"] == ""


def test_platform_status_supports_custom_browser_data_dir(tmp_path, monkeypatch):
    browser_data = tmp_path / "profiles"
    (browser_data / "cdp_dy_user_data_dir").mkdir(parents=True)
    monkeypatch.setenv("MONITOR_BROWSER_DATA_DIR", str(browser_data))

    statuses = list_platform_status(tmp_path, [])
    dy = next(item for item in statuses if item["platform"] == "dy")

    assert dy["profile_path"] == str((browser_data / "cdp_dy_user_data_dir").resolve())
    assert dy["profile_exists"] is True


def test_platform_status_uses_active_account_profile_when_present(tmp_path):
    init_db()
    snapshot = _snapshot_table("social_accounts")
    try:
        account = save_social_account(
            {
                "name": "海安律所抖音采集号",
                "platform": "dy",
                "login_type": "qrcode",
                "status": "active",
            }
        )
        expected_path = resolve_account_profile_path(f"1/dy/acc_{account['id']}")
        expected_path.mkdir(parents=True, exist_ok=True)
        (expected_path / "state").write_text("ok", encoding="utf-8")

        statuses = list_platform_status(recent_runs=[])
    finally:
        _restore_table("social_accounts", snapshot)
    dy = next(item for item in statuses if item["platform"] == "dy")

    assert dy["profile_key"] == f"1/dy/acc_{account['id']}"
    assert dy["profile_path"] == str(expected_path)
    assert dy["profile_exists"] is True
    assert dy["using_account_profile"] is True
    assert dy["active_account_id"] == account["id"]
    assert dy["active_account_name"] == "海安律所抖音采集号"


def test_cdp_browser_uses_same_custom_profile_root_as_status(tmp_path, monkeypatch):
    browser_data = tmp_path / "profiles"
    monkeypatch.setenv("MONITOR_BROWSER_DATA_DIR", str(browser_data))

    expected = browser_data / "cdp_dy_user_data_dir"

    assert Path(resolve_cdp_user_data_dir("dy")) == expected
    dy_status = next(item for item in list_platform_status(tmp_path, []) if item["platform"] == "dy")
    assert dy_status["profile_path"] == str(expected.resolve())


def test_cdp_browser_can_use_explicit_account_profile(tmp_path, monkeypatch):
    account_profile = tmp_path / "account_profiles" / "dy_1"
    monkeypatch.setenv("MONITOR_BROWSER_DATA_DIR", str(tmp_path / "platform_profiles"))
    monkeypatch.setenv("MONITOR_CDP_USER_DATA_DIR_DY", str(account_profile))

    assert Path(resolve_cdp_user_data_dir("dy")) == account_profile


def test_login_browser_command_uses_monitor_profile_root(tmp_path, monkeypatch):
    browser_data = tmp_path / "profiles"
    fake_browser = tmp_path / "chrome.exe"
    fake_browser.write_text("", encoding="utf-8")
    monkeypatch.setenv("MONITOR_BROWSER_DATA_DIR", str(browser_data))
    monkeypatch.setattr("api.monitoring.login_browser.BrowserLauncher.detect_browser_paths", lambda self: [str(fake_browser)])

    command = build_login_browser_command("xhs")

    assert command["profile_path"] == str((browser_data / "cdp_xhs_user_data_dir").resolve())
    assert command["debug_port"] == 9325
    assert command["login_url"].startswith("https://www.xiaohongshu.com")


def test_login_browser_message_reminds_to_close_window(tmp_path, monkeypatch):
    fake_browser = tmp_path / "chrome.exe"
    fake_browser.write_text("", encoding="utf-8")
    monkeypatch.setenv("MONITOR_BROWSER_DATA_DIR", str(tmp_path / "profiles"))
    monkeypatch.setattr("api.monitoring.login_state.LOGIN_STATE_DIR", tmp_path / "login_windows")
    monkeypatch.setattr("api.monitoring.login_state._pid_exists", lambda pid: pid == 12345)
    monkeypatch.setattr("api.monitoring.login_browser.BrowserLauncher.detect_browser_paths", lambda self: [str(fake_browser)])

    class FakeProcess:
        pid = 12345

    monkeypatch.setattr("api.monitoring.login_browser.subprocess.Popen", lambda *args, **kwargs: FakeProcess())

    result = open_login_browser("dy")

    assert result["pid"] == 12345
    assert login_window_status("dy")["pid"] == 12345
    assert "关闭该窗口" in result["message"]
    assert "运行采集" in result["message"]


def test_login_browser_route_can_use_social_account_profile(tmp_path, monkeypatch):
    init_db()
    snapshots = {
        "social_accounts": _snapshot_table("social_accounts"),
    }
    fake_browser = tmp_path / "chrome.exe"
    fake_browser.write_text("", encoding="utf-8")
    seen: dict[str, Any] = {}
    try:
        account = save_social_account(
            {
                "name": "海安律所抖音采集号",
                "platform": "dy",
                "login_type": "qrcode",
                "status": "standby",
                "profile_path": str(tmp_path / "account_profile"),
            }
        )
        monkeypatch.setattr("api.monitoring.login_browser.BrowserLauncher.detect_browser_paths", lambda self: [str(fake_browser)])

        def fake_open_login_browser_with_command(command):
            seen["profile_path"] = command["profile_path"]
            return {**command, "pid": 12345, "message": "ok"}

        monkeypatch.setattr(monitor_router, "open_login_browser_with_command", fake_open_login_browser_with_command)

        result = asyncio.run(monitor_router.platform_login_browser("dy", {"account_id": account["id"]}))

        assert result["pid"] == 12345
        assert seen["profile_path"] == str(resolve_account_profile_path(f"1/dy/acc_{account['id']}"))
        assert result["profile_path"] == "网页登录态已配置"
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_login_browser_command_supports_per_platform_port_env(tmp_path, monkeypatch):
    browser_data = tmp_path / "profiles"
    fake_browser = tmp_path / "chrome.exe"
    fake_browser.write_text("", encoding="utf-8")
    monkeypatch.setenv("MONITOR_BROWSER_DATA_DIR", str(browser_data))
    monkeypatch.setenv("MONITOR_LOGIN_DEBUG_PORT_DY", "19323")
    monkeypatch.setattr("api.monitoring.login_browser.BrowserLauncher.detect_browser_paths", lambda self: [str(fake_browser)])

    command = build_login_browser_command("dy")

    assert command["debug_port"] == 19323
    assert command["profile_path"] == str((browser_data / "cdp_dy_user_data_dir").resolve())


def test_windows_oneclick_launcher_separates_bind_host_and_browser_url():
    default_plan = build_launch_plan("0.0.0.0", 8080)
    explicit_plan = build_launch_plan("0.0.0.0", 8080, "http://10.0.0.12:8080/monitor")

    assert default_plan.bind_host == "0.0.0.0"
    assert default_plan.probe_url == "http://127.0.0.1:8080/api/health"
    assert default_plan.browser_url == "http://127.0.0.1:8080/monitor"
    assert explicit_plan.browser_url == "http://10.0.0.12:8080/monitor"
    assert default_plan.command[:4] == [sys.executable, "-m", "uvicorn", "api.main:app"]


def test_windows_oneclick_launcher_opens_browser_after_health(monkeypatch):
    seen = {}

    class FakeProcess:
        def __init__(self):
            self.poll_calls = 0
            self.returncode = None

        def poll(self):
            self.poll_calls += 1
            if self.poll_calls < 2:
                return None
            return 0

        def terminate(self):
            seen["terminated"] = True

        def wait(self, timeout):
            return 0

        def kill(self):
            seen["killed"] = True

    class FakeResponse:
        def read(self):
            return b'{"status":"ok"}'

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

    def fake_popen(*args, **kwargs):
        seen["command"] = args[0]
        seen["env_host"] = kwargs["env"]["MONITOR_HOST"]
        seen["env_port"] = kwargs["env"]["MONITOR_PORT"]
        seen["env_browser_url"] = kwargs["env"].get("MONITOR_BROWSER_URL")
        return FakeProcess()

    monkeypatch.setattr("api.monitoring.startup_launcher.subprocess.Popen", fake_popen)
    monkeypatch.setattr("api.monitoring.startup_launcher.urlopen", lambda url, timeout=2: FakeResponse())
    monkeypatch.setattr("api.monitoring.startup_launcher.webbrowser.open", lambda url: seen.setdefault("opened", url))

    plan = api_monitoring_startup_launcher.start_oneclick("0.0.0.0", 8080, "http://10.0.0.12:8080/monitor", 1)

    assert plan.browser_url == "http://10.0.0.12:8080/monitor"
    assert seen["env_host"] == "0.0.0.0"
    assert seen["env_port"] == "8080"
    assert seen["env_browser_url"] == "http://10.0.0.12:8080/monitor"
    assert seen["opened"] == "http://10.0.0.12:8080/monitor"


def test_job_validation_rejects_operator_input_errors():
    base = {
        "law_firm_name": "校验测试律所",
        "keywords": ["校验测试律所避雷"],
        "platforms": ["dy"],
        "enable_comments": False,
        "time_window_type": "recent_1d",
        "frequency": "daily",
        "email_time": "09:00",
        "enabled": True,
    }
    with pytest.raises(ValueError, match="invalid recipient email"):
        save_job({**base, "recipients": ["bad-email"]})
    with pytest.raises(ValueError, match="cron_expr is required"):
        save_job({**base, "recipients": [], "frequency": "cron", "cron_expr": ""})
    with pytest.raises(ValueError, match="custom_start must be before custom_end"):
        save_job(
            {
                **base,
                "recipients": [],
                "time_window_type": "custom",
                "custom_start": "2026-06-12",
                "custom_end": "2026-06-11",
            }
        )
    with pytest.raises(ValueError, match="email_time must be HH:MM"):
        save_job({**base, "recipients": [], "email_time": "25:00"})
    with pytest.raises(ValueError, match="测试数据模板"):
        save_job({**base, "law_firm_name": "请改成目标律所名称", "recipients": []})
    with pytest.raises(ValueError, match="测试数据模板"):
        save_job({**base, "keywords": ["目标律所避雷"], "recipients": []})


def test_job_advanced_collect_config_persists_and_validates(tmp_path):
    init_db()
    snapshots = {
        "monitor_jobs": _snapshot_table("monitor_jobs"),
        "job_keywords": _snapshot_table("job_keywords"),
        "job_platforms": _snapshot_table("job_platforms"),
        "job_recipients": _snapshot_table("job_recipients"),
        "ai_key_profiles": _snapshot_table("ai_key_profiles"),
        "email_templates": _snapshot_table("email_templates"),
        "proxy_profiles": _snapshot_table("proxy_profiles"),
        "social_accounts": _snapshot_table("social_accounts"),
    }
    try:
        profile = save_ai_key_profile(
            {
                "name": "海安 AI 接入",
                "provider": "openai",
                "base_url": "https://example.com",
                "api_key": "sk-test-advanced",
                "model": "test-model",
            }
        )
        template = save_email_template({"name": "海安日报模板", "subject_template": "日报 {law_firm_name}", "html_template": "{report_body}"})
        proxy = save_proxy_profile({"name": "华东代理", "provider": "manual", "proxy_url": "http://user:pass@127.0.0.1:8081"})
        account = save_social_account({"name": "抖音采集号", "platform": "dy", "status": "active", "proxy_id": proxy["id"]})
        job = save_job(
            {
                "law_firm_name": "海安律所",
                "aliases": ["海安律师事务所"],
                "keywords": ["海安律所避雷"],
                "platforms": ["dy"],
                "recipients": ["target@example.com"],
                "enable_comments": True,
                "enable_sub_comments": True,
                "time_window_type": "recent_1d",
                "frequency": "daily",
                "email_time": "09:00",
                "target_type": "detail",
                "max_pages": 3,
                "max_items": 12,
                "start_page": 2,
                "output_mode": "excel",
                "browser_mode": "profile",
                "ai_profile_id": profile["id"],
                "email_template_id": template["id"],
                "account_id": account["id"],
                "proxy_id": proxy["id"],
            }
        )
        stored = get_job(job["id"])
        cmd = runner_module._build_crawler_cmd(stored, "dy", tmp_path)

        assert stored["enable_sub_comments"] is True
        assert stored["target_type"] == "detail"
        assert stored["max_pages"] == 3
        assert stored["max_items"] == 12
        assert stored["start_page"] == 2
        assert stored["output_mode"] == "excel"
        assert stored["browser_mode"] == "profile"
        assert stored["ai_profile_id"] == profile["id"]
        assert stored["email_template_id"] == template["id"]
        assert stored["account_id"] == account["id"]
        assert stored["proxy_id"] == proxy["id"]
        assert _cmd_value(cmd, "--type") == "detail"
        assert _cmd_value(cmd, "--save_data_option") == "excel"
        assert _cmd_value(cmd, "--start") == "2"
        assert _cmd_value(cmd, "--get_sub_comment") == "true"
        assert _cmd_value(cmd, "--crawler_max_notes_count") == "30"
        assert _cmd_value(cmd, "--specified_id") == "海安律所避雷"

        for patch, message in [
            ({"target_type": "bad"}, "target_type must be one of"),
            ({"output_mode": "bad"}, "output_mode must be one of"),
            ({"browser_mode": "bad"}, "browser_mode must be one of"),
            ({"max_pages": 0}, "max_pages must be between"),
            ({"account_id": 99999999}, "social account not found"),
        ]:
            with pytest.raises(ValueError, match=message):
                save_job({**stored, **patch, "recipients": ["target@example.com"]}, stored["id"])
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_runner_command_maps_creator_mode_to_platform_user_collection(tmp_path):
    init_db()
    snapshot = _snapshot_table("platform_login_configs")
    try:
        cmd = runner_module._build_crawler_cmd(
            {
                "keywords": ["MS4wLjABAAAAhaian"],
                "enable_comments": False,
                "enable_sub_comments": False,
                "time_window_type": "recent_1d",
                "target_type": "creator",
                "max_pages": 1,
                "max_items": 8,
                "start_page": 1,
                "output_mode": "internal",
            },
            "dy",
            tmp_path,
        )
    finally:
        _restore_table("platform_login_configs", snapshot)

    assert _cmd_value(cmd, "--type") == "creator"
    assert _cmd_value(cmd, "--save_data_option") == "json"
    assert _cmd_value(cmd, "--crawler_max_notes_count") == "10"
    assert _cmd_value(cmd, "--creator_id") == "MS4wLjABAAAAhaian"
    assert "--specified_id" not in cmd


def test_ai_and_email_config_validation_rejects_bad_inputs():
    init_db()
    with pytest.raises(ValueError, match="invalid AI provider"):
        save_ai_config({"provider": "bad", "temperature": 0})
    with pytest.raises(ValueError, match="temperature must be between 0 and 2"):
        save_ai_config({"provider": "openai", "temperature": 9})
    with pytest.raises(ValueError, match="smtp_port must be between 1 and 65535"):
        save_email_config({"smtp_port": 70000})
    with pytest.raises(ValueError, match="invalid email encryption"):
        save_email_config({"encryption": "tls"})
    with pytest.raises(ValueError, match="invalid recipient email"):
        save_email_config({"default_recipients": ["bad-email"]})


def test_platform_login_config_defaults_masking_and_validation():
    init_db()
    snapshot = _snapshot_table("platform_login_configs")
    try:
        configs = list_platform_login_configs()
        dy = next(item for item in configs if item["platform"] == "dy")
        ks = next(item for item in configs if item["platform"] == "ks")

        assert dy["login_type"] == "qrcode"
        assert dy["supported_login_types"] == ["qrcode", "cookie"]
        assert "phone" not in ks["supported_login_types"]
        assert dy["login_capability_source"] == "平台采集服务"
        assert "暂未开放手机号登录" in ks["unsupported_reason"]

        saved = save_platform_login_config("dy", {"login_type": "cookie", "cookies": "sessionid=secret-cookie"})
        raw = get_platform_login_config("dy", masked=False)

        assert saved["login_type"] == "cookie"
        assert saved["has_cookies"] is True
        assert "secret-cookie" not in saved["cookies"]
        assert raw["cookies"] == "sessionid=secret-cookie"

        with pytest.raises(ValueError, match="暂未开放手机号登录"):
            save_platform_login_config("ks", {"login_type": "phone"})
        with pytest.raises(ValueError, match="Cookie 登录需要先填写 Cookie"):
            save_platform_login_config("xhs", {"login_type": "cookie"})
        with pytest.raises(ValueError, match="暂未开放手机号登录"):
            save_platform_login_config("dy", {"login_type": "phone", "clear_login_phone": True})
        with pytest.raises(ValueError, match="unsupported platform"):
            save_platform_login_config("wb", {"login_type": "qrcode"})
    finally:
        _restore_table("platform_login_configs", snapshot)


def test_runner_command_uses_platform_login_config_for_cookie_mode(tmp_path):
    init_db()
    snapshot = _snapshot_table("platform_login_configs")
    try:
        save_platform_login_config("dy", {"login_type": "cookie", "cookies": "sessionid=secret-cookie"})
        cmd = runner_module._build_crawler_cmd(
            {"keywords": ["海安律所避雷"], "enable_comments": False, "time_window_type": "recent_1d"},
            "dy",
            tmp_path,
        )
    finally:
        _restore_table("platform_login_configs", snapshot)

    assert _cmd_value(cmd, "--lt") == "cookie"
    assert _cmd_value(cmd, "--cookies") == "sessionid=secret-cookie"


def test_runner_command_uses_bound_account_cookie_login_parameter(tmp_path):
    init_db()
    snapshot = _snapshot_table("social_accounts")
    try:
        account = save_social_account(
            {
                "name": "海安律所小红书采集号",
                "platform": "xhs",
                "login_type": "cookie",
                "status": "active",
                "cookies": "web_session=account-cookie",
            }
        )
        cmd = runner_module._build_crawler_cmd(
            {"keywords": ["海安律所退费"], "enable_comments": False, "time_window_type": "recent_1d"},
            "xhs",
            tmp_path,
            {"login_type": account["login_type"], "cookies": "web_session=account-cookie"},
        )
    finally:
        _restore_table("social_accounts", snapshot)

    assert _cmd_value(cmd, "--lt") == "cookie"
    assert _cmd_value(cmd, "--cookies") == "web_session=account-cookie"


def test_runner_command_defaults_to_qrcode_login(tmp_path):
    init_db()
    snapshot = _snapshot_table("platform_login_configs")
    try:
        cmd = runner_module._build_crawler_cmd(
            {"keywords": ["海安律所避雷"], "enable_comments": False, "time_window_type": "recent_1d"},
            "xhs",
            tmp_path,
        )
    finally:
        _restore_table("platform_login_configs", snapshot)

    assert _cmd_value(cmd, "--lt") == "qrcode"
    assert "--cookies" not in cmd


def test_mediacrawler_cli_accepts_login_phone(monkeypatch):
    import config

    original = getattr(config, "LOGIN_PHONE", "")
    try:
        result = asyncio.run(
            parse_mediacrawler_cmd(
                [
                    "--platform",
                    "xhs",
                    "--lt",
                    "phone",
                    "--type",
                    "search",
                    "--keywords",
                    "海安律所投诉",
                    "--login_phone",
                    "13800138000",
                ]
            )
        )

        assert result.login_phone == "13800138000"
        assert config.LOGIN_PHONE == "13800138000"
    finally:
        config.LOGIN_PHONE = original


def test_runner_injects_bound_active_proxy_without_leaking_secret(tmp_path):
    init_db()
    snapshots = {
        "proxy_profiles": _snapshot_table("proxy_profiles"),
        "social_accounts": _snapshot_table("social_accounts"),
    }
    try:
        proxy = save_proxy_profile(
            {
                "name": "华东采集代理",
                "provider": "manual",
                "proxy_url": "http://user:pass@127.0.0.1:8081",
                "status": "active",
                "max_concurrency": 1,
            }
        )
        account = save_social_account(
            {
                "name": "抖音采集号",
                "platform": "dy",
                "login_type": "qrcode",
                "status": "active",
                "proxy_id": proxy["id"],
            }
        )
        binding = runner_module._resolve_platform_proxy_binding("dy")
        env = runner_module._build_crawler_env(binding)
        summary = runner_module._proxy_summary(binding)
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)

    assert binding["account_id"] == account["id"]
    assert env["HTTP_PROXY"] == "http://user:pass@127.0.0.1:8081"
    assert env["HTTPS_PROXY"] == "http://user:pass@127.0.0.1:8081"
    assert env["MONITOR_ACTIVE_PROXY_ID"] == str(proxy["id"])
    assert summary["proxy_id"] == proxy["id"]
    assert "user:pass" not in summary["proxy_url"]
    assert "[REDACTED]" in summary["proxy_url"]


def test_runner_injects_active_account_profile_for_cdp(tmp_path):
    init_db()
    snapshot = _snapshot_table("social_accounts")
    try:
        account = save_social_account(
            {
                "name": "抖音采集号",
                "platform": "dy",
                "login_type": "qrcode",
                "status": "active",
                "profile_path": str(tmp_path / "dy_account_profile"),
            }
        )
        binding = runner_module._resolve_platform_account_binding("dy")
        env = runner_module._build_crawler_env(binding)
        summary = runner_module._account_summary(binding)
    finally:
        _restore_table("social_accounts", snapshot)

    assert binding["account_id"] == account["id"]
    expected_profile_path = str(resolve_account_profile_path(f"1/dy/acc_{account['id']}"))
    assert binding["profile_key"] == f"1/dy/acc_{account['id']}"
    assert binding["profile_path"] == expected_profile_path
    assert env["MONITOR_CDP_USER_DATA_DIR"] == expected_profile_path
    assert env["MONITOR_CDP_USER_DATA_DIR_DY"] == expected_profile_path
    assert env["MONITOR_ACTIVE_ACCOUNT_ID"] == str(account["id"])
    assert summary["account_name"] == "抖音采集号"
    assert summary["profile_key"] == f"1/dy/acc_{account['id']}"
    assert "profile_path" not in summary


def test_crawler_command_uses_platform_search_terms_only(tmp_path):
    init_db()
    snapshot = _snapshot_table("platform_login_configs")
    try:
        job = {
            "law_firm_name": "海安律所",
            "aliases": ["海安律师事务所", "海安律师"],
            "exclude_words": ["招聘", "广告合作"],
            "keywords": ["海安律所避雷", "海安律所退费", "海安律所投诉"],
            "enable_comments": False,
            "enable_sub_comments": False,
            "time_window_type": "recent_1d",
            "target_type": "search",
            "max_pages": 1,
            "max_items": 20,
            "start_page": 1,
            "output_mode": "internal",
        }
        cmd = runner_module._build_crawler_cmd(job, "dy", tmp_path)
    finally:
        _restore_table("platform_login_configs", snapshot)

    assert _cmd_value(cmd, "--keywords") == "海安律所避雷,海安律所退费,海安律所投诉"
    command_text = " ".join(cmd)
    assert "海安律师事务所" not in command_text
    assert "海安律师" not in command_text
    assert "招聘" not in command_text
    assert "广告合作" not in command_text


def test_runner_prefers_job_bound_account_and_proxy(tmp_path):
    init_db()
    snapshots = {
        "proxy_profiles": _snapshot_table("proxy_profiles"),
        "social_accounts": _snapshot_table("social_accounts"),
    }
    try:
        account_proxy = save_proxy_profile(
            {
                "name": "账号默认代理",
                "provider": "manual",
                "proxy_url": "http://account:pass@127.0.0.1:8081",
                "status": "active",
                "max_concurrency": 1,
            }
        )
        job_proxy = save_proxy_profile(
            {
                "name": "任务指定代理",
                "provider": "manual",
                "proxy_url": "http://job:pass@127.0.0.1:8082",
                "status": "active",
                "max_concurrency": 1,
            }
        )
        fallback_account = save_social_account(
            {
                "name": "抖音备用号",
                "platform": "dy",
                "login_type": "qrcode",
                "status": "active",
                "profile_path": str(tmp_path / "fallback_profile"),
            }
        )
        bound_account = save_social_account(
            {
                "name": "海安律所抖音采集号",
                "platform": "dy",
                "login_type": "qrcode",
                "status": "active",
                "profile_path": str(tmp_path / "bound_profile"),
                "proxy_id": account_proxy["id"],
            }
        )
        binding = runner_module._resolve_platform_account_binding(
            "dy",
            {"account_id": bound_account["id"], "proxy_id": job_proxy["id"]},
        )
        env = runner_module._build_crawler_env(binding)
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)

    assert fallback_account["id"] != bound_account["id"]
    assert binding["account_id"] == bound_account["id"]
    assert binding["profile_key"] == f"1/dy/acc_{bound_account['id']}"
    assert binding["profile_path"] == str(resolve_account_profile_path(binding["profile_key"]))
    assert binding["proxy_id"] == account_proxy["id"]
    assert binding["task_proxy_id"] == job_proxy["id"]
    assert env["HTTP_PROXY"] == "http://account:pass@127.0.0.1:8081"


def test_phase_5_account_lock_blocks_concurrent_profile_use():
    init_db()
    snapshots = {
        "social_accounts": _snapshot_table("social_accounts"),
        "crawl_runs": _snapshot_table("crawl_runs"),
    }
    jobs_snapshot = _snapshot_monitor_jobs()
    run1 = None
    run2 = None
    try:
        _clear_monitor_jobs()
        account = save_social_account({"name": "抖音锁定号", "platform": "dy", "login_type": "qrcode", "status": "active"})
        job = save_job(
            {
                "law_firm_name": "锁定测试律所",
                "keywords": ["锁定测试律所避雷"],
                "platforms": ["dy"],
                "recipients": [],
            }
        )
        run1 = create_run(job["id"], timeout_seconds=120)
        run2 = create_run(job["id"], timeout_seconds=120)
        deadline = get_run(run1)["deadline_at"]

        assert acquire_account_lock(account["id"], run1, deadline) is True
        assert acquire_account_lock(account["id"], run2, deadline) is False

        release_account_lock(account["id"], run1)

        assert acquire_account_lock(account["id"], run2, deadline) is True
    finally:
        for run_id in [run1, run2]:
            if run_id:
                release_run_resource_locks(run_id)
        _restore_monitor_jobs(jobs_snapshot)
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_phase_5_proxy_resource_locks_respect_max_concurrency():
    init_db()
    snapshots = {
        "proxy_profiles": _snapshot_table("proxy_profiles"),
        "crawl_runs": _snapshot_table("crawl_runs"),
        "resource_locks": _snapshot_table("resource_locks"),
    }
    jobs_snapshot = _snapshot_monitor_jobs()
    run1 = None
    run2 = None
    try:
        _clear_monitor_jobs()
        proxy = save_proxy_profile(
            {
                "name": "单并发代理",
                "provider": "manual",
                "proxy_url": "http://user:pass@127.0.0.1:8081",
                "status": "active",
                "max_concurrency": 1,
            }
        )
        job = save_job(
            {
                "law_firm_name": "代理并发测试律所",
                "keywords": ["代理并发测试律所避雷"],
                "platforms": ["dy"],
                "recipients": [],
            }
        )
        run1 = create_run(job["id"], timeout_seconds=120)
        run2 = create_run(job["id"], timeout_seconds=120)
        deadline = get_run(run1)["deadline_at"]

        assert acquire_proxy_lock(proxy["id"], run1, deadline) is True
        assert acquire_proxy_lock(proxy["id"], run2, deadline) is False

        release_proxy_locks(run1)

        assert acquire_proxy_lock(proxy["id"], run2, deadline) is True
    finally:
        for run_id in [run1, run2]:
            if run_id:
                release_run_resource_locks(run_id)
        _restore_monitor_jobs(jobs_snapshot)
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_phase_5_recovery_marks_expired_running_run_before_releasing_locks():
    init_db()
    snapshots = {
        "proxy_profiles": _snapshot_table("proxy_profiles"),
        "social_accounts": _snapshot_table("social_accounts"),
        "crawl_runs": _snapshot_table("crawl_runs"),
        "resource_locks": _snapshot_table("resource_locks"),
    }
    jobs_snapshot = _snapshot_monitor_jobs()
    try:
        _clear_monitor_jobs()
        proxy = save_proxy_profile(
            {
                "name": "恢复代理",
                "provider": "manual",
                "proxy_url": "http://user:pass@127.0.0.1:8081",
                "status": "active",
                "max_concurrency": 1,
            }
        )
        account = save_social_account(
            {
                "name": "恢复账号",
                "platform": "dy",
                "login_type": "qrcode",
                "status": "active",
                "proxy_id": proxy["id"],
            }
        )
        job = save_job(
            {
                "law_firm_name": "恢复测试律所",
                "keywords": ["恢复测试律所避雷"],
                "platforms": ["dy"],
                "recipients": [],
                "account_id": account["id"],
                "proxy_id": proxy["id"],
            }
        )
        run_id = create_run(job["id"], timeout_seconds=120)
        expired_at = (datetime.now(timezone.utc) - timedelta(seconds=30)).isoformat()
        with get_conn() as conn:
            conn.execute("UPDATE crawl_runs SET deadline_at=? WHERE id=?", (expired_at, run_id))
        assert acquire_account_lock(account["id"], run_id, expired_at) is True
        assert acquire_proxy_lock(proxy["id"], run_id, expired_at) is True

        recovered = recover_stale_runs_and_locks("pytest_recovery")
        run = get_run(run_id)
        account_after = get_social_account(account["id"], masked=False)
        with get_conn() as conn:
            proxy_locks = conn.execute("SELECT COUNT(*) AS n FROM resource_locks WHERE run_id=?", (run_id,)).fetchone()["n"]

        assert recovered["recovered_runs"] == 1
        assert run["status"] == "timeout"
        assert account_after["locked_by_run_id"] is None
        assert proxy_locks == 0
    finally:
        _restore_monitor_jobs(jobs_snapshot)
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_phase_7_1_run_identity_legacy_lookup_and_dry_run_backfill():
    init_db()
    snapshots = {"crawl_runs": _snapshot_table("crawl_runs")}
    jobs_snapshot = _snapshot_monitor_jobs()
    try:
        _clear_monitor_jobs()
        job = save_job(
            {
                "law_firm_name": "Phase71身份律所",
                "keywords": ["Phase71身份律所投诉"],
                "platforms": ["dy"],
                "recipients": [],
            }
        )
        run_id = create_run(job["id"], {"job_id": job["id"], "law_firm_name": job["law_firm_name"]})
        with get_conn() as conn:
            persisted = conn.execute("SELECT job_id FROM crawl_runs WHERE id=?", (run_id,)).fetchone()["job_id"]
            conn.execute("UPDATE crawl_runs SET job_id=NULL WHERE id=?", (run_id,))
            conn.execute(
                "INSERT INTO crawl_runs (workspace_id, job_id, status, started_at, summary) VALUES (1, NULL, 'running', ?, ?)",
                (
                    datetime.now(timezone.utc).isoformat(),
                    json.dumps({"job_id": 99999999, "law_firm_name": "未知律所"}, ensure_ascii=False),
                ),
            )

        preview = preview_crawl_run_job_id_backfill()
        applied = preview_crawl_run_job_id_backfill(apply=True)

        assert persisted == job["id"]
        assert any(item == {"run_id": run_id, "job_id": job["id"]} for item in preview["resolvable"])
        assert any(item["summary_job_id"] == 99999999 for item in preview["unresolved"])
        assert applied["applied"] == 1
        assert get_run(run_id)["job_id"] == job["id"]
        assert has_running_run_for_job(job["id"]) is True
        assert preview_crawl_run_job_id_backfill()["unresolved"]
    finally:
        _restore_monitor_jobs(jobs_snapshot)
        _restore_table("crawl_runs", snapshots["crawl_runs"])


def test_phase_7_1_finish_run_is_idempotent_and_releases_locks():
    init_db()
    snapshots = {
        "crawl_runs": _snapshot_table("crawl_runs"),
        "social_accounts": _snapshot_table("social_accounts"),
        "resource_locks": _snapshot_table("resource_locks"),
        "proxy_profiles": _snapshot_table("proxy_profiles"),
    }
    jobs_snapshot = _snapshot_monitor_jobs()
    try:
        _clear_monitor_jobs()
        account = save_social_account({"name": "Phase71终态账号", "platform": "dy", "login_type": "qrcode", "status": "active"})
        proxy = save_proxy_profile({"name": "Phase71终态代理", "provider": "manual", "proxy_url": "http://u:p@127.0.0.1:8000", "status": "active", "max_concurrency": 1})
        job = save_job({"law_firm_name": "Phase71终态律所", "keywords": ["Phase71终态律所"], "platforms": ["dy"], "recipients": []})
        run_id = create_run(job["id"], {"job_id": job["id"], "phase_7_1_lifecycle": True}, timeout_seconds=120)
        deadline = get_run(run_id)["deadline_at"]
        assert acquire_account_lock(account["id"], run_id, deadline) is True
        assert acquire_proxy_lock(proxy["id"], run_id, deadline) is True

        finish_run(run_id, "success", {"job_id": job["id"], "phase": "terminal:success"})
        finish_run(run_id, "failed", {"job_id": job["id"], "phase": "terminal:failed"}, "late writer should not reopen")
        release_run_resource_locks(run_id)
        release_run_resource_locks(run_id)

        run = get_run(run_id)
        account_after = get_social_account(account["id"], masked=False)
        with get_conn() as conn:
            proxy_locks = conn.execute("SELECT COUNT(*) AS n FROM resource_locks WHERE run_id=?", (run_id,)).fetchone()["n"]

        assert run["status"] == "success"
        assert run["summary"]["terminal_status"] == "success"
        assert account_after["locked_by_run_id"] is None
        assert proxy_locks == 0
    finally:
        _restore_monitor_jobs(jobs_snapshot)
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_phase_7_1_stale_recovery_marks_new_lifecycle_rows_interrupted_not_legacy():
    init_db()
    snapshots = {"crawl_runs": _snapshot_table("crawl_runs"), "system_settings": _snapshot_table("system_settings")}
    jobs_snapshot = _snapshot_monitor_jobs()
    try:
        _clear_monitor_jobs()
        with get_conn() as conn:
            conn.execute("DELETE FROM system_settings")
        save_runtime_settings({"stale_run_heartbeat_grace_seconds": 60})
        job = save_job({"law_firm_name": "Phase71中断律所", "keywords": ["Phase71中断律所"], "platforms": ["dy"], "recipients": []})
        old_heartbeat = (datetime.now(timezone.utc) - timedelta(minutes=20)).isoformat()
        future_deadline = (datetime.now(timezone.utc) + timedelta(hours=1)).isoformat()
        phase_run = create_run(
            job["id"],
            {
                "job_id": job["id"],
                "phase_7_1_lifecycle": True,
                "phase": "ai_evaluating",
                "progress_updated_at": old_heartbeat,
                "last_safe_result": {"evaluated": 250},
            },
            timeout_seconds=3600,
        )
        legacy_run = create_run(job["id"], {"job_id": job["id"], "progress_updated_at": old_heartbeat}, timeout_seconds=3600)
        with get_conn() as conn:
            conn.execute("UPDATE crawl_runs SET deadline_at=? WHERE id IN (?, ?)", (future_deadline, phase_run, legacy_run))

        recovered = recover_stale_runs_and_locks("pytest_phase_7_1")

        assert recovered["interrupted_runs"] == 1
        assert get_run(phase_run)["status"] == "interrupted"
        assert get_run(phase_run)["display_status"] == "执行中断"
        assert get_run(legacy_run)["status"] == "running"
    finally:
        _restore_monitor_jobs(jobs_snapshot)
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_ai_and_email_test_paths_reuse_config_validation():
    init_db()
    with pytest.raises(ValueError, match="invalid AI provider"):
        asyncio.run(run_ai_config_test({"provider": "bad", "base_url": "https://example.com", "api_key": "sk-test", "model": "m"}))
    with pytest.raises(ValueError, match="temperature must be between 0 and 2"):
        asyncio.run(run_ai_config_test({"provider": "openai", "base_url": "https://example.com", "api_key": "sk-test", "model": "m", "temperature": 9}))
    with pytest.raises(ValueError, match="smtp_port must be between 1 and 65535"):
        send_test_email({"smtp_port": 70000, "smtp_host": "smtp.example.com", "sender": "a@example.com", "default_recipients": ["b@example.com"]})
    with pytest.raises(ValueError, match="invalid email encryption"):
        send_test_email({"encryption": "tls", "smtp_host": "smtp.example.com", "sender": "a@example.com", "default_recipients": ["b@example.com"]})
    with pytest.raises(ValueError, match="invalid recipient email"):
        send_test_email({"smtp_host": "smtp.example.com", "sender": "a@example.com", "target": "bad-email"})


def test_report_email_uses_specific_attachment_mime_types(tmp_path):
    html_path = tmp_path / "report.html"
    xlsx_path = tmp_path / "report.xlsx"
    md_path = tmp_path / "report.md"
    html_path.write_text("<h1>日报</h1>", encoding="utf-8")
    xlsx_path.write_bytes(b"fake-xlsx")
    md_path.write_text("# 日报", encoding="utf-8")

    msg = build_report_email(
        {"sender": "sender@example.com"},
        ["target@example.com"],
        "测试日报",
        {"html_path": str(html_path), "excel_path": str(xlsx_path), "markdown_path": str(md_path)},
    )
    attachment_types = [part.get_content_type() for part in msg.iter_attachments()]

    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in attachment_types
    assert "text/markdown" in attachment_types
    assert msg["To"] == "target@example.com"


def test_active_email_template_uses_report_summary_values(tmp_path):
    init_db()
    snapshot = _snapshot_table("email_templates")
    html_path = tmp_path / "report.html"
    xlsx_path = tmp_path / "report.xlsx"
    md_path = tmp_path / "report.md"
    html_path.write_text("<article>真实报告正文</article>", encoding="utf-8")
    xlsx_path.write_bytes(b"fake-xlsx")
    md_path.write_text("# 日报", encoding="utf-8")
    try:
        save_email_template(
            {
                "name": "企业日报模板",
                "subject_template": "日报 {law_firm_name} {negative_count}",
                "html_template": "<main>{law_firm_name}|{new_contents}|{negative_count}|{high_count}|{pending_review_count}|{platforms}|{report_html}</main>",
                "is_active": True,
            }
        )
        msg = build_report_email(
            {"sender": "sender@example.com"},
            ["target@example.com"],
            "测试日报",
            {
                "law_firm_name": "海安律所",
                "summary": {
                    "law_firm_name": "海安律所",
                    "new_contents": 9,
                    "negative_count": 3,
                    "high_count": 1,
                    "pending_review_count": 2,
                    "platforms": ["dy", "ks", "xhs"],
                },
                "html_path": str(html_path),
                "excel_path": str(xlsx_path),
                "markdown_path": str(md_path),
            },
        )
        html_body = _email_html_body(msg)
    finally:
        _restore_table("email_templates", snapshot)

    assert "海安律所|9|3|1|2|抖音 / 快手 / 小红书" in html_body
    assert "真实报告正文" in html_body


def test_active_email_template_uses_job_values_for_html_body(tmp_path):
    init_db()
    snapshot = _snapshot_table("email_templates")
    html_path = tmp_path / "report.html"
    xlsx_path = tmp_path / "report.xlsx"
    md_path = tmp_path / "report.md"
    html_path.write_text("<article>正文</article>", encoding="utf-8")
    xlsx_path.write_bytes(b"fake-xlsx")
    md_path.write_text("# 日报", encoding="utf-8")
    try:
        save_email_template(
            {
                "name": "企业日报模板",
                "subject_template": "日报 {law_firm_name}",
                "html_template": "<main>{law_firm_name}|{platforms}|{report_html}</main>",
                "is_active": True,
            }
        )
        msg = build_report_email(
            {"sender": "sender@example.com"},
            ["target@example.com"],
            "测试日报",
            {
                "summary": {"platforms": ["dy", "xhs"]},
                "html_path": str(html_path),
                "excel_path": str(xlsx_path),
                "markdown_path": str(md_path),
            },
            {"law_firm_name": "海安律所"},
        )
        html_body = _email_html_body(msg)
    finally:
        _restore_table("email_templates", snapshot)

    assert "海安律所|抖音 / 小红书" in html_body
    assert "<article>正文</article>" in html_body


def test_active_email_template_supports_report_body_alias(tmp_path):
    init_db()
    snapshot = _snapshot_table("email_templates")
    html_path = tmp_path / "report.html"
    xlsx_path = tmp_path / "report.xlsx"
    md_path = tmp_path / "report.md"
    html_path.write_text("<article>报告正文别名</article>", encoding="utf-8")
    xlsx_path.write_bytes(b"fake-xlsx")
    md_path.write_text("# 日报", encoding="utf-8")
    try:
        save_email_template(
            {
                "name": "企业日报模板",
                "subject_template": "日报 {law_firm_name}",
                "html_template": "<main>{law_firm_name}|{report_body}</main>",
                "is_active": True,
            }
        )
        msg = build_report_email(
            {"sender": "sender@example.com"},
            ["target@example.com"],
            "测试日报",
            {
                "summary": {"platforms": ["dy"]},
                "html_path": str(html_path),
                "excel_path": str(xlsx_path),
                "markdown_path": str(md_path),
            },
            {"law_firm_name": "海安律所"},
        )
        html_body = _email_html_body(msg)
    finally:
        _restore_table("email_templates", snapshot)

    assert "海安律所|<article>报告正文别名</article>" in html_body


def test_phase_17_2_email_template_body_guardrails_block_missing_report_body(tmp_path):
    init_db()
    snapshot = _snapshot_table("email_templates")
    html_path = tmp_path / "report.html"
    xlsx_path = tmp_path / "report.xlsx"
    md_path = tmp_path / "report.md"
    html_path.write_text("<article>真实报告正文</article>", encoding="utf-8")
    xlsx_path.write_bytes(b"fake-xlsx")
    md_path.write_text("# 日报", encoding="utf-8")
    try:
        with pytest.raises(ValueError, match=r"\{report_html\}"):
            save_email_template(
                {
                    "name": "缺正文模板",
                    "subject_template": "日报 {law_firm_name}",
                    "html_template": "<main>只有外壳</main>",
                    "is_active": True,
                }
            )

        preview = render_email_template_preview(
            {
                "subject_template": "日报 {law_firm_name}",
                "html_template": "<main>只有外壳</main>",
                "law_firm_name": "海安律所",
            }
        )
        assert preview["has_report_body_placeholder"] is False
        assert "保存会被阻止" in preview["body_guardrail"]
        assert "预览使用样例数据" in preview["sample_data_note"]
        assert "高风险线索" in preview["html"]

        with get_conn() as conn:
            conn.execute("UPDATE email_templates SET is_active=0")
            conn.execute(
                """
                INSERT INTO email_templates (name, subject_template, html_template, is_active, created_at, updated_at)
                VALUES ('历史缺正文模板', '日报 {law_firm_name}', '<main>历史模板外壳</main>', 1, ?, ?)
                """,
                (datetime.now(timezone.utc).isoformat(), datetime.now(timezone.utc).isoformat()),
            )
        msg = build_report_email(
            {"sender": "sender@example.com"},
            ["target@example.com"],
            "测试日报",
            {
                "summary": {"platforms": ["dy"]},
                "html_path": str(html_path),
                "excel_path": str(xlsx_path),
                "markdown_path": str(md_path),
            },
            {"law_firm_name": "海安律所"},
        )
        html_body = _email_html_body(msg)
    finally:
        _restore_table("email_templates", snapshot)

    assert "历史模板外壳" in html_body
    assert "真实报告正文" in html_body


def test_report_email_preview_reuses_active_email_template(tmp_path):
    init_db()
    snapshot = _snapshot_table("email_templates")
    html_path = tmp_path / "report.html"
    xlsx_path = tmp_path / "report.xlsx"
    md_path = tmp_path / "report.md"
    html_path.write_text("<article>真实报告正文</article>", encoding="utf-8")
    xlsx_path.write_bytes(b"fake-xlsx")
    md_path.write_text("# 日报", encoding="utf-8")
    try:
        save_email_template(
            {
                "name": "企业日报模板",
                "subject_template": "日报 {law_firm_name} {negative_count}",
                "html_template": "<main>{law_firm_name}|{negative_count}|{report_html}</main>",
                "is_active": True,
            }
        )
        preview = render_report_email_preview(
            {"law_firm_name": "海安律所"},
            {
                "summary": {"negative_count": 3, "platforms": ["dy"]},
                "html_path": str(html_path),
                "excel_path": str(xlsx_path),
                "markdown_path": str(md_path),
            },
            {"sender": "sender@example.com"},
        )
    finally:
        _restore_table("email_templates", snapshot)

    assert preview["subject"] == "日报 海安律所 3"
    assert "海安律所|3|<article>真实报告正文</article>" in preview["html"]


def test_job_bound_email_template_takes_precedence_for_email_and_preview(tmp_path):
    init_db()
    snapshot = _snapshot_table("email_templates")
    html_path = tmp_path / "report.html"
    xlsx_path = tmp_path / "report.xlsx"
    md_path = tmp_path / "report.md"
    html_path.write_text("<article>报告正文</article>", encoding="utf-8")
    xlsx_path.write_bytes(b"fake-xlsx")
    md_path.write_text("# 日报", encoding="utf-8")
    try:
        active = save_email_template(
            {
                "name": "默认邮件模板",
                "subject_template": "默认 {law_firm_name}",
                "html_template": "<main>默认模板|{law_firm_name}|{report_html}</main>",
                "is_active": True,
            }
        )
        bound = save_email_template(
            {
                "name": "海安任务模板",
                "subject_template": "绑定 {law_firm_name} {new_contents}",
                "html_template": "<main>绑定模板|{law_firm_name}|{new_contents}|{report_body}</main>",
                "is_active": False,
            }
        )
        job = {"law_firm_name": "海安律所", "email_template_id": bound["id"]}
        report = {
            "summary": {"new_contents": 5, "platforms": ["dy"]},
            "html_path": str(html_path),
            "excel_path": str(xlsx_path),
            "markdown_path": str(md_path),
        }

        preview = render_report_email_preview(job, report, {"sender": "sender@example.com"})
        msg = build_report_email(
            {"sender": "sender@example.com"},
            ["target@example.com"],
            preview["subject"],
            report,
            job,
        )
        html_body = _email_html_body(msg)

        assert active["is_active"] is True
        assert preview["subject"] == "绑定 海安律所 5"
        assert "绑定模板|海安律所|5|<article>报告正文</article>" in preview["html"]
        assert "绑定模板|海安律所|5|<article>报告正文</article>" in html_body
        assert "默认模板" not in html_body
    finally:
        _restore_table("email_templates", snapshot)


def test_email_template_preview_supports_report_body_alias():
    preview = asyncio.run(
        monitor_router.email_template_preview(
            {
                "subject_template": "日报 {law_firm_name}",
                "html_template": "<main>{law_firm_name}|{report_body}</main>",
                "law_firm_name": "海安律所",
            }
        )
    )["preview"]

    assert preview["subject"] == "日报 海安律所"
    assert "海安律所|" in preview["html"]
    assert "高风险线索" in preview["html"]


def test_report_email_preview_api_returns_actual_email_body():
    init_db()
    snapshots = {
        "email_templates": _snapshot_table("email_templates"),
        "monitor_jobs": _snapshot_monitor_jobs(),
    }
    _clear_monitor_jobs()
    job = save_job(
        {
            "law_firm_name": "海安律所",
            "aliases": [],
            "keywords": ["海安律所避雷"],
            "exclude_words": [],
            "platforms": ["dy"],
            "recipients": ["target@example.com"],
            "enabled": False,
        }
    )
    run_id = create_run(job["id"])
    try:
        save_email_template(
            {
                "name": "企业日报模板",
                "subject_template": "日报 {law_firm_name} {new_contents}",
                "html_template": "<main>{law_firm_name}|{new_contents}|{report_html}</main>",
                "is_active": True,
            }
        )
        report = create_report(
            run_id,
            job,
            {"platforms": ["dy"], "failed_platforms": [], "new_contents": 2, "negative_count": 0, "high_count": 0},
        )

        result = asyncio.run(monitor_router.report_email_preview(int(report["id"])))["preview"]

        assert result["subject"] == "日报 海安律所 2"
        assert "海安律所|2|" in result["html"]
        assert "AI 结果仅用于舆情线索筛查" in result["html"]
    finally:
        _cleanup_test_records(job["id"], "")
        _restore_table("email_templates", snapshots["email_templates"])
        _restore_monitor_jobs(snapshots["monitor_jobs"])


def test_report_download_media_types_are_specific(tmp_path):
    assert (
        monitor_router._report_download_media_type("excel", tmp_path / "report.xlsx")
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert monitor_router._report_download_media_type("markdown", tmp_path / "report.md") == "text/markdown"
    assert monitor_router._report_download_media_type("html", tmp_path / "report.html") == "text/html"


def test_report_path_guard_rejects_files_outside_report_dir(tmp_path, monkeypatch):
    reports_dir = tmp_path / "reports"
    reports_dir.mkdir()
    safe_path = reports_dir / "report.html"
    unsafe_path = tmp_path / "secret.txt"
    safe_path.write_text("ok", encoding="utf-8")
    unsafe_path.write_text("secret", encoding="utf-8")
    monkeypatch.setattr(monitor_router, "MONITOR_DATA_DIR", tmp_path)

    assert monitor_router._safe_report_path(str(safe_path)) == safe_path
    with pytest.raises(HTTPException) as exc:
        monitor_router._safe_report_path(str(unsafe_path))

    assert exc.value.status_code == 403


def test_sensitive_text_is_redacted():
    text = "Authorization: Bearer sk-secret123456789 api_key=abc123 password=hunter2 cookie=session=abc token=mytoken 密码：明文密码 代理地址=http://user:pass@example.com"
    proxy_text = "proxy=http://user:pass@127.0.0.1:8081"
    path_text = r"Browser path: C:\Program Files\Google\Chrome\Application\chrome.exe and E:\myproject\MediaCrawler\monitor_data\runs\job_1\run_2\crawler.log"
    redacted = redact_sensitive(text)
    redacted_proxy = redact_sensitive(proxy_text)
    redacted_path = redact_sensitive(path_text)
    customer_safe = monitor_router.customer_safe_text(path_text)

    assert "sk-secret123456789" not in redacted
    assert "abc123" not in redacted
    assert "hunter2" not in redacted
    assert "session=abc" not in redacted
    assert "mytoken" not in redacted
    assert "明文密码" not in redacted
    assert "user:pass@example.com" not in redacted
    assert "user:pass" not in redacted_proxy
    assert "http://[REDACTED]@127.0.0.1:8081" in redacted_proxy
    assert "[REDACTED]" in redacted
    assert "Program Files" not in redacted_path
    assert r"Chrome\Application" not in redacted_path
    assert r"E:\myproject" not in redacted_path
    assert "运行日志" in customer_safe
    assert "Program Files" not in customer_safe
    assert r"Chrome\Application" not in customer_safe
    assert r"E:\myproject" not in customer_safe


def test_phase_9_admin_resource_operations_are_audited_without_secrets():
    init_db()
    snapshots = {
        "audit_logs": _snapshot_table("audit_logs"),
        "social_accounts": _snapshot_table("social_accounts"),
        "proxy_profiles": _snapshot_table("proxy_profiles"),
        "ai_key_profiles": _snapshot_table("ai_key_profiles"),
        "email_configs": _snapshot_singleton_table("email_configs"),
        "email_templates": _snapshot_table("email_templates"),
    }
    admin = {"id": 901, "workspace_id": 1, "role": "administrator"}
    try:
        with get_conn() as conn:
            conn.execute("DELETE FROM audit_logs")

        proxy = asyncio.run(
            monitor_router.create_proxy(
                {
                    "name": "Phase9 Proxy",
                    "provider": "manual",
                    "proxy_url": "http://secret-user:secret-pass@127.0.0.1:8081",
                    "status": "active",
                    "max_concurrency": 1,
                    "last_error": "代理地址=http://secret-user:secret-pass@127.0.0.1:8081",
                },
                admin=admin,
            )
        )["proxy"]
        account = asyncio.run(
            monitor_router.create_social_account(
                {
                    "name": "Phase9 Account",
                    "platform": "dy",
                    "login_type": "cookie",
                    "cookies": "sessionid=secret-cookie",
                    "status": "active",
                    "proxy_id": proxy["id"],
                },
                admin=admin,
            )
        )["account"]
        profile = asyncio.run(
            monitor_router.create_ai_profile(
                {
                    "name": "Phase9 AI",
                    "provider": "openai",
                    "base_url": "https://example.com",
                    "api_key": "sk-phase9-secret-key",
                    "model": "phase9-model",
                },
                admin=admin,
            )
        )["profile"]
        asyncio.run(
            monitor_router.update_email_config(
                {
                    "smtp_host": "smtp.example.com",
                    "smtp_port": 465,
                    "encryption": "ssl",
                    "sender": "sender@example.com",
                    "username": "sender@example.com",
                    "password": "smtp-secret",
                    "default_recipients": ["target@example.com"],
                },
                admin=admin,
            )
        )
        template = asyncio.run(
            monitor_router.create_email_template(
                {"name": "Phase9 Template", "subject_template": "日报 {law_firm_name}", "html_template": "{report_body}"},
                admin=admin,
            )
        )["template"]

        with get_conn() as conn:
            rows = [dict(row) for row in conn.execute("SELECT * FROM audit_logs ORDER BY id").fetchall()]
        actions = {row["action_type"] for row in rows}
        assert {
            "create_proxy",
            "create_social_account",
            "create_ai_profile",
            "update_email_config",
            "create_email_template",
        } <= actions
        serialized = json.dumps(rows, ensure_ascii=False)
        for forbidden in ["secret-pass", "secret-cookie", "sk-phase9-secret-key", "smtp-secret"]:
            assert forbidden not in serialized
        assert all(row["user_id"] == admin["id"] for row in rows)
        assert str(proxy["id"]) in {row["resource_id"] for row in rows}
        assert str(account["id"]) in {row["resource_id"] for row in rows}
        assert str(profile["id"]) in {row["resource_id"] for row in rows}
        assert str(template["id"]) in {row["resource_id"] for row in rows}
    finally:
        _restore_table("audit_logs", snapshots["audit_logs"])
        _restore_table("social_accounts", snapshots["social_accounts"])
        _restore_table("proxy_profiles", snapshots["proxy_profiles"])
        _restore_table("ai_key_profiles", snapshots["ai_key_profiles"])
        _restore_singleton_table("email_configs", snapshots["email_configs"])
        _restore_table("email_templates", snapshots["email_templates"])


def test_run_summary_and_log_api_redact_sensitive_values(tmp_path, monkeypatch):
    init_db()
    job = save_job(
        {
            "law_firm_name": "脱敏测试律所",
            "aliases": [],
            "exclude_words": [],
            "keywords": ["脱敏测试律所避雷"],
            "platforms": ["dy"],
            "recipients": [],
            "enable_comments": False,
            "time_window_type": "recent_1d",
            "frequency": "daily",
            "email_time": "09:00",
            "enabled": False,
        }
    )
    run_id = create_run(job["id"])
    secret_text = "api_key=abc123 password=hunter2 cookie=session=abc"
    finish_run(
        run_id,
        "failed",
        {"platform_results": {"dy": {"status": "failed", "error": secret_text}}},
        secret_text,
    )
    run_dir = tmp_path / "runs" / "job_1" / f"run_{run_id}_pytest" / "dy"
    run_dir.mkdir(parents=True)
    path_text = r"Browser path: C:\Program Files\Google\Chrome\Application\chrome.exe profile_path=E:\myproject\MediaCrawler\monitor_data\profiles\dy\acc_1"
    (run_dir / "crawler.log").write_text(f"{secret_text}\n{path_text}", encoding="utf-8")
    monkeypatch.setattr(monitor_router, "MONITOR_DATA_DIR", tmp_path)
    monkeypatch.setattr(database_module, "MONITOR_DATA_DIR", tmp_path)

    with get_conn() as conn:
        row = conn.execute("SELECT summary, error_message FROM crawl_runs WHERE id=?", (run_id,)).fetchone()
    logs = asyncio.run(monitor_router.run_logs(run_id))["logs"]
    _cleanup_test_records(job["id"], "")

    assert "abc123" not in row["summary"]
    assert "hunter2" not in row["error_message"]
    assert logs
    assert "session=abc" not in logs[0]["content"]
    assert "Program Files" not in logs[0]["content"]
    assert r"Chrome\Application" not in logs[0]["content"]
    assert r"E:\myproject" not in logs[0]["content"]
    assert "profile_path" not in logs[0]["content"]
    assert "[REDACTED]" in logs[0]["content"]


def test_phase_7_1_ai_invalid_json_exception_and_timeout_fallback_to_pending_review(monkeypatch):
    async def run_case(case_name: str, fake_eval):
        init_db()
        snapshots = {
            "crawl_runs": _snapshot_table("crawl_runs"),
            "raw_contents": _snapshot_table("raw_contents"),
            "raw_comments": _snapshot_table("raw_comments"),
            "ai_evaluations": _snapshot_table("ai_evaluations"),
            "system_settings": _snapshot_table("system_settings"),
        }
        jobs_snapshot = _snapshot_monitor_jobs()
        content_id = f"pytest_phase_7_1_ai_{case_name}"
        try:
            _clear_monitor_jobs()
            with get_conn() as conn:
                for table in ["crawl_runs", "raw_contents", "raw_comments", "ai_evaluations", "system_settings"]:
                    conn.execute(f"DELETE FROM {table}")
            save_runtime_settings({"ai_item_timeout_seconds": 5, "ai_item_retry_count": 0})
            job = save_job(
                {
                    "law_firm_name": f"Phase71AI{case_name}律所",
                    "keywords": [f"Phase71AI{case_name}律所投诉"],
                    "platforms": ["dy"],
                    "recipients": [],
                    "enable_comments": False,
                    "time_window_type": "recent_1d",
                }
            )
            run_id = create_run(job["id"], {"job_id": job["id"], "phase_7_1_lifecycle": True}, timeout_seconds=120)
            item = {
                "aweme_id": content_id,
                "title": f"Phase71AI{case_name}律所投诉",
                "desc": "需要人工复核",
                "create_time": int(datetime.now(timezone.utc).timestamp()),
            }
            ingested = ingest_outputs(job, run_id, "dy", [item], [])
            monkeypatch.setattr(runner_module, "evaluate_content", fake_eval)

            summary = await evaluate_new_contents(job, run_id, ingested["content_db_ids"])

            with get_conn() as conn:
                row = conn.execute(
                    "SELECT status, reason FROM ai_evaluations WHERE run_id=?",
                    (run_id,),
                ).fetchone()
            assert summary["pending_review_count"] == 1
            assert summary["ai_failed_fallback_evaluations"] == 1
            assert summary["ai_unresolved_items"] == 0
            assert row["status"] == "pending_review"
            assert "复核" in row["reason"]
        finally:
            _cleanup_test_records(job["id"], content_id)
            _restore_monitor_jobs(jobs_snapshot)
            for table, snapshot in snapshots.items():
                _restore_table(table, snapshot)

    async def invalid_json(job, content, comments):
        return {"not": "contract"}

    async def raises(job, content, comments):
        raise RuntimeError("provider crashed api_key=secret")

    async def slow(job, content, comments):
        await asyncio.sleep(0.02)
        return {"status": "ok"}

    async def tiny_timeout(run_id):
        return 0.001

    original_timeout = runner_module._ai_item_timeout_seconds
    try:
        asyncio.run(run_case("invalid", invalid_json))
        asyncio.run(run_case("exception", raises))
        monkeypatch.setattr(runner_module, "_ai_item_timeout_seconds", lambda run_id: 0.001)
        asyncio.run(run_case("timeout", slow))
    finally:
        monkeypatch.setattr(runner_module, "_ai_item_timeout_seconds", original_timeout)


def test_phase_7_1_ai_interruption_generates_partial_report_and_terminal_state(monkeypatch):
    init_db()
    snapshots = {
        "reports": _snapshot_table("reports"),
        "crawl_runs": _snapshot_table("crawl_runs"),
        "raw_contents": _snapshot_table("raw_contents"),
        "raw_comments": _snapshot_table("raw_comments"),
        "ai_evaluations": _snapshot_table("ai_evaluations"),
        "email_delivery_logs": _snapshot_table("email_delivery_logs"),
    }
    jobs_snapshot = _snapshot_monitor_jobs()
    content_ids = []
    try:
        _clear_monitor_jobs()
        with get_conn() as conn:
            for table in ["reports", "crawl_runs", "raw_contents", "raw_comments", "ai_evaluations", "email_delivery_logs"]:
                conn.execute(f"DELETE FROM {table}")
        job = save_job(
            {
                "law_firm_name": "Phase71部分报告律所",
                "keywords": ["Phase71部分报告律所投诉"],
                "platforms": ["dy"],
                "recipients": [],
                "enable_comments": False,
                "time_window_type": "recent_1d",
            }
        )

        async def fake_run_platform(job_arg, run_id, platform, run_dir):
            items = []
            for idx in range(1, 6):
                cid = f"pytest_phase_7_1_partial_{idx}"
                content_ids.append(cid)
                items.append(
                    {
                        "aweme_id": cid,
                        "title": f"Phase71部分报告律所投诉 {idx}",
                        "desc": "收费争议",
                        "create_time": int(datetime.now(timezone.utc).timestamp()),
                    }
                )
            return ingest_outputs(job_arg, run_id, platform, items, [])

        eval_calls = 0

        async def fake_evaluate(job_arg, content, comments):
            nonlocal eval_calls
            eval_calls += 1
            if str(content.get("content_id") or "").endswith(("_3", "_4", "_5")):
                raise asyncio.CancelledError()
            return {
                "status": "ok",
                "is_related": True,
                "is_negative": True,
                "risk_level": "high",
                "reason": "命中投诉",
                "evidence_quotes": [content.get("title")],
                "recommended_action": "人工复核",
                "raw_response": "{}",
            }

        monkeypatch.setattr(runner_module, "run_platform", fake_run_platform)
        monkeypatch.setattr(runner_module, "evaluate_content", fake_evaluate)
        monkeypatch.setattr(
            runner_module,
            "send_report_with_delivery_log",
            lambda job, report, send_type="auto": (False, "未配置收件人", report, None),
        )

        result = asyncio.run(run_monitor_job(job["id"]))
        run = get_run(result["run_id"])
        reports = list_reports(10)
        report = next(item for item in reports if item["run_id"] == result["run_id"])

        assert result["status"] == "success"
        assert run["status"] == "success"
        assert result["summary"]["pending_review_count"] >= 3
        assert result["summary"]["ai_unresolved_items"] == 0
        assert report["summary"]["pending_review_count"] >= 3
        assert Path(report["html_path"]).exists()
    finally:
        for cid in content_ids:
            _cleanup_test_records(job["id"], cid)
        _restore_monitor_jobs(jobs_snapshot)
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_ai_test_requires_contract_shaped_output():
    valid = _validate_ai_output(
        {
            "is_related": True,
            "is_negative": True,
            "risk_level": "high",
            "reason": "命中投诉",
            "evidence_quotes": ["退费争议"],
            "recommended_action": "人工复核",
        }
    )
    assert valid["risk_level"] == "high"
    tolerant = _validate_ai_output(
        {
            "result": {
                "is_related": "true",
                "is_negative": "false",
                "risk_level": "高风险",
                "reason": "命中投诉",
                "evidence_quotes": "退费争议",
                "recommended_action": "人工复核",
            }
        }
    )
    assert tolerant["is_related"] is True
    assert tolerant["is_negative"] is False
    assert tolerant["risk_level"] == "high"
    assert tolerant["evidence_quotes"] == ["退费争议"]
    with pytest.raises(ValueError, match="AI 输出缺少字段"):
        _validate_ai_output({"is_related": True})
    with pytest.raises(ValueError, match="risk_level"):
        _validate_ai_output(
            {
                "is_related": True,
                "is_negative": True,
                "risk_level": "urgent",
                "reason": "命中投诉",
                "evidence_quotes": ["退费争议"],
                "recommended_action": "人工复核",
            }
        )


def test_ai_json_parser_accepts_fenced_json_with_prefix_text():
    parsed = _parse_json(
        """
        下面是判断结果：
        ```json
        {"is_related": true, "is_negative": false, "risk_level": "low", "reason": "普通内容", "evidence_quotes": [], "recommended_action": "无需处理"}
        ```
        """
    )

    assert parsed["is_related"] is True
    assert parsed["risk_level"] == "low"


def test_ai_evaluation_failure_redacts_provider_endpoint(monkeypatch):
    monkeypatch.delenv("MONITOR_SKIP_AI_API", raising=False)
    monkeypatch.setattr(ai_module, "get_active_ai_key_profile", lambda masked=False: None)
    monkeypatch.setattr(
        ai_module,
        "get_ai_config",
        lambda masked=False: {
            "provider": "openai",
            "base_url": "https://deedee.tech",
            "api_key": "sk-test",
            "model": "test-model",
            "temperature": 0,
            "prompt": DEFAULT_PROMPT,
        },
    )

    async def fake_call_openai(cfg, prompt, payload):
        request = httpx.Request("POST", "https://deedee.tech/v1/chat/completions")
        response = httpx.Response(502, request=request)
        raise httpx.HTTPStatusError(
            "Server error '502 Bad Gateway' for url 'https://deedee.tech/v1/chat/completions'",
            request=request,
            response=response,
        )

    monkeypatch.setattr(ai_module, "_call_openai", fake_call_openai)

    result = asyncio.run(
        ai_module.evaluate_content(
            {"law_firm_name": "海安律所", "aliases": [], "exclude_words": []},
            {
                "platform": "dy",
                "source_keyword": "海安律所避雷",
                "title": "海安律所退费投诉",
                "description": "退费迟迟没有处理",
            },
            [],
        )
    )

    assert result["status"] == "pending_review"
    assert "deedee.tech" not in result["reason"]
    assert "v1/chat/completions" not in result["reason"]
    assert "[AI_ENDPOINT_REDACTED]" in result["reason"]


def test_ai_test_uses_haian_sample_payload(monkeypatch):
    init_db()
    seen: dict[str, Any] = {}

    async def fake_call_openai(cfg, prompt, payload):
        seen.update(payload)
        return json.dumps(
            {
                "is_related": True,
                "is_negative": True,
                "risk_level": "medium",
                "reason": "命中退费投诉",
                "evidence_quotes": ["退费拖了很久"],
                "recommended_action": "人工复核",
            },
            ensure_ascii=False,
        )

    monkeypatch.setattr("api.monitoring.ai._call_openai", fake_call_openai)

    result = asyncio.run(
        run_ai_config_test(
            {
                "provider": "openai",
                "base_url": "https://example.com",
                "api_key": "sk-test",
                "model": "test-model",
                "temperature": 0,
            }
        )
    )

    assert result["risk_level"] == "medium"
    assert seen["law_firm_name"] == "海安律所"
    assert seen["source_keyword"] == "海安律所避雷"
    assert "海安律所" in seen["title"]
    assert seen["content_url"].startswith("https://www.douyin.com/video/")
    assert seen["cover_url"]
    assert seen["comment_summary"]["sample_count"] == 2


def test_ai_test_accepts_editable_sample_context(monkeypatch):
    init_db()
    seen: dict[str, Any] = {}

    async def fake_call_openai(cfg, prompt, payload):
        seen.update(payload)
        return json.dumps(
            {
                "is_related": True,
                "is_negative": True,
                "risk_level": "medium",
                "reason": "样例命中",
                "evidence_quotes": [payload["source_keyword"]],
                "recommended_action": "人工复核",
            },
            ensure_ascii=False,
        )

    monkeypatch.setattr("api.monitoring.ai._call_openai", fake_call_openai)

    result = asyncio.run(
        run_ai_config_test(
            {
                "provider": "openai",
                "base_url": "https://example.com",
                "api_key": "sk-test",
                "model": "test-model",
                "temperature": 0,
                "prompt": DEFAULT_PROMPT,
                "sample_law_firm_name": "测试律所",
                "sample_platform": "xhs",
                "sample_source_keyword": "测试律所投诉",
                "sample_title": "测试律所投诉样例",
                "sample_text": "收费争议需要复核",
                "sample_comments": "自定义评论一\n自定义评论二",
            }
        )
    )

    assert result["risk_level"] == "medium"
    assert seen["law_firm_name"] == "测试律所"
    assert seen["platform"] == "小红书"
    assert seen["platform_code"] == "xhs"
    assert seen["source_keyword"] == "测试律所投诉"
    assert seen["title"] == "测试律所投诉样例"
    assert seen["description"] == "收费争议需要复核"
    assert seen["comments"] == ["自定义评论一", "自定义评论二"]


def test_ai_evaluation_payload_includes_content_and_comment_context():
    payload = build_evaluation_payload(
        {
            "law_firm_name": "海安律所",
            "aliases": ["海安律师事务所"],
            "exclude_words": ["招聘"],
        },
        {
            "platform": "xhs",
            "platform_label": "小红书",
            "source_keyword": "海安律所退费",
            "title": "海安律所退费沟通记录",
            "description": "咨询退费一直没有明确回复。",
            "author_name": "海安用户",
            "content_url": "https://www.xiaohongshu.com/explore/haian-note",
            "cover_url": "https://example.com/cover.jpg",
            "publish_time": 1781280000,
            "comment_count": 12,
        },
        [
            {"content": "我也想知道退费怎么处理", "author_name": "评论用户A", "create_time": 1781280100},
            {"content": "先保留合同和聊天记录", "author_name": "评论用户B", "create_time": 1781280200},
        ],
    )

    assert payload["law_firm_name"] == "海安律所"
    assert payload["platform"] == "小红书"
    assert payload["platform_code"] == "xhs"
    assert payload["content_url"].endswith("haian-note")
    assert payload["cover_url"].endswith("cover.jpg")
    assert payload["author_name"] == "海安用户"
    assert payload["comment_count"] == 12
    assert payload["comments"] == ["我也想知道退费怎么处理", "先保留合同和聊天记录"]
    assert payload["comment_samples"][0]["author_name"] == "评论用户A"
    assert payload["comment_summary"]["declared_count"] == 12
    assert payload["comment_summary"]["observed_count"] == 2
    assert "退费" in payload["comment_summary"]["sample_text"]


def test_phase_7_2_cr096_valid_ai_output_is_not_forced_unrelated_by_source_keyword_gate(monkeypatch):
    init_db()
    profile_snapshot = _snapshot_table("ai_key_profiles")
    raw_response = {
        "is_related": True,
        "is_negative": True,
        "risk_level": "high",
        "reason": "模型误把搜索词当作目标证据",
        "evidence_quotes": ["北京海安律所退费"],
        "recommended_action": "人工复核",
    }

    async def fake_call_openai(cfg, prompt, payload):
        assert payload["source_keyword"] == "北京海安律所退费"
        assert "海安律所" not in payload["title"]
        assert "海安律所" not in payload["description"]
        assert payload["comments"] == []
        return json.dumps(raw_response, ensure_ascii=False)

    try:
        save_ai_key_profile(
            {
                "name": "CR045 校准 AI",
                "provider": "openai",
                "base_url": "https://ai.example.com",
                "api_key": "sk-profile",
                "model": "profile-model",
                "temperature": 0,
                "prompt": DEFAULT_PROMPT,
                "is_active": True,
            }
        )
        monkeypatch.setattr("api.monitoring.ai._call_openai", fake_call_openai)

        result = asyncio.run(
            ai_module.evaluate_content(
                {"law_firm_name": "海安律所", "aliases": ["海安律师事务所"], "exclude_words": []},
                {
                    "platform": "dy",
                    "platform_label": "抖音",
                    "source_keyword": "北京海安律所退费",
                    "title": "考研课程退费维权记录",
                    "description": "报名课程后想退费，沟通很久没有解决。",
                    "author_name": "教育消费者",
                    "comment_count": 0,
                },
                [],
            )
        )
    finally:
        _restore_table("ai_key_profiles", profile_snapshot)

    assert result["status"] == "ok"
    assert result["is_related"] is True
    assert result["is_negative"] is True
    assert result["risk_level"] == "high"
    assert result["reason"] == raw_response["reason"]
    assert result["evidence_quotes"] == raw_response["evidence_quotes"]
    assert result["recommended_action"] == raw_response["recommended_action"]


def test_phase_7_2_cr096_semantic_law_firm_reference_survives_format_only_postprocessing(monkeypatch):
    init_db()
    profile_snapshot = _snapshot_table("ai_key_profiles")
    raw_response = {
        "is_related": True,
        "is_negative": True,
        "risk_level": "high",
        "reason": "标题明确指向北京海安律师事务所并包含被骗陈述",
        "evidence_quotes": ["我被北京海安律师事务所骗了"],
        "recommended_action": "人工复核",
    }

    async def fake_call_openai(cfg, prompt, payload):
        assert payload["law_firm_name"] == "北京海安律所"
        assert payload.get("aliases") == []
        assert "北京海安律师事务所骗了" in payload["title"]
        assert "source_keyword" in prompt
        return json.dumps(raw_response, ensure_ascii=False)

    try:
        save_ai_key_profile(
            {
                "name": "CR096 北京海安 AI",
                "provider": "openai",
                "base_url": "https://ai.example.com",
                "api_key": "sk-profile",
                "model": "profile-model",
                "temperature": 0,
                "prompt": DEFAULT_PROMPT,
                "is_active": True,
            }
        )
        monkeypatch.setattr("api.monitoring.ai._call_openai", fake_call_openai)

        result = asyncio.run(
            ai_module.evaluate_content(
                {"law_firm_name": "北京海安律所", "aliases": [], "exclude_words": []},
                {
                    "platform": "dy",
                    "platform_label": "抖音",
                    "source_keyword": "北京海安律所",
                    "title": "我被北京海安律师事务所骗了",
                    "description": "我也被这家律所骗了2980元。",
                    "author_name": "当事人",
                    "comment_count": 0,
                },
                [],
            )
        )
    finally:
        _restore_table("ai_key_profiles", profile_snapshot)

    assert result["status"] == "ok"
    assert result["is_related"] is True
    assert result["is_negative"] is True
    assert result["risk_level"] == "high"
    assert result["reason"] == raw_response["reason"]
    assert result["evidence_quotes"] == raw_response["evidence_quotes"]


def test_phase_7_2_target_evidence_in_title_or_comments_remains_negative(monkeypatch):
    init_db()
    profile_snapshot = _snapshot_table("ai_key_profiles")

    async def fake_call_openai(cfg, prompt, payload):
        return json.dumps(
            {
                "is_related": True,
                "is_negative": True,
                "risk_level": "medium",
                "reason": "标题或评论明确提到目标律所并有退费投诉",
                "evidence_quotes": [payload["title"]] if "海安律所" in payload["title"] else [payload["comments"][0]],
                "recommended_action": "人工复核",
            },
            ensure_ascii=False,
        )

    try:
        save_ai_key_profile(
            {
                "name": "CR045 真阳性 AI",
                "provider": "openai",
                "base_url": "https://ai.example.com",
                "api_key": "sk-profile",
                "model": "profile-model",
                "temperature": 0,
                "prompt": DEFAULT_PROMPT,
                "is_active": True,
            }
        )
        monkeypatch.setattr("api.monitoring.ai._call_openai", fake_call_openai)

        title_result = asyncio.run(
            ai_module.evaluate_content(
                {"law_firm_name": "海安律所", "aliases": ["海安律师事务所"], "exclude_words": []},
                {
                    "platform": "dy",
                    "platform_label": "抖音",
                    "source_keyword": "海安律所退费",
                    "title": "海安律所退费投诉记录",
                    "description": "沟通迟迟没有明确答复。",
                    "author_name": "当事人",
                    "comment_count": 0,
                },
                [],
            )
        )
        comment_result = asyncio.run(
            ai_module.evaluate_content(
                {"law_firm_name": "海安律所", "aliases": ["海安律师事务所"], "exclude_words": []},
                {
                    "platform": "dy",
                    "platform_label": "抖音",
                    "source_keyword": "海安律所退费",
                    "title": "退费投诉记录",
                    "description": "正文没有点名目标。",
                    "author_name": "当事人",
                    "comment_count": 1,
                },
                [{"content": "评论补充：海安律所一直没处理退费。", "author_name": "评论用户"}],
            )
        )
    finally:
        _restore_table("ai_key_profiles", profile_snapshot)

    assert title_result["is_related"] is True
    assert title_result["is_negative"] is True
    assert title_result["risk_level"] == "medium"
    assert comment_result["is_related"] is True
    assert comment_result["is_negative"] is True
    assert comment_result["evidence_quotes"] == ["评论补充：海安律所一直没处理退费。"]


def test_phase_7_2_cr096_homonym_geography_valid_ai_output_is_preserved(monkeypatch):
    init_db()
    profile_snapshot = _snapshot_table("ai_key_profiles")

    async def fake_call_openai(cfg, prompt, payload):
        return json.dumps(
            {
                "is_related": True,
                "is_negative": True,
                "risk_level": "high",
                "reason": "模型误把地名海安当作目标律所",
                "evidence_quotes": ["海安本地退费纠纷"],
                "recommended_action": "人工复核",
            },
            ensure_ascii=False,
        )

    try:
        save_ai_key_profile(
            {
                "name": "CR045 同名地名 AI",
                "provider": "openai",
                "base_url": "https://ai.example.com",
                "api_key": "sk-profile",
                "model": "profile-model",
                "temperature": 0,
                "prompt": DEFAULT_PROMPT,
                "is_active": True,
            }
        )
        monkeypatch.setattr("api.monitoring.ai._call_openai", fake_call_openai)

        result = asyncio.run(
            ai_module.evaluate_content(
                {"law_firm_name": "海安律所", "aliases": ["海安律师事务所"], "exclude_words": []},
                {
                    "platform": "xhs",
                    "platform_label": "小红书",
                    "source_keyword": "海安律所退费",
                    "title": "海安本地培训机构退费经历",
                    "description": "江苏海安一家培训机构退费慢，准备咨询律师。",
                    "author_name": "海安生活",
                    "comment_count": 0,
                },
                [],
            )
        )
    finally:
        _restore_table("ai_key_profiles", profile_snapshot)

    assert result["is_related"] is True
    assert result["is_negative"] is True
    assert result["risk_level"] == "high"
    assert result["reason"] == "模型误把地名海安当作目标律所"
    assert result["evidence_quotes"] == ["海安本地退费纠纷"]


def test_ai_offline_check_does_not_call_provider_or_update_test_status(monkeypatch):
    init_db()
    ai_snapshot = _snapshot_singleton_table("ai_configs")
    called = False

    async def fake_call_openai(cfg, prompt, payload):
        nonlocal called
        called = True
        raise RuntimeError("offline check must not call provider")

    try:
        save_ai_config(
            {
                "provider": "openai",
                "base_url": "https://saved.example.com",
                "api_key": "sk-saved",
                "model": "saved-model",
                "temperature": 0,
                "prompt": DEFAULT_PROMPT,
            }
        )
        before = get_ai_config()
        monkeypatch.setattr("api.monitoring.ai._call_openai", fake_call_openai)

        result = asyncio.run(
            monitor_router.ai_config_offline_check(
                {
                    "provider": "openai",
                    "base_url": "https://example.com",
                    "api_key": "sk-test",
                    "model": "test-model",
                    "temperature": 0,
                    "prompt": DEFAULT_PROMPT,
                }
            )
        )["result"]
        after = get_ai_config()

        assert called is False
        assert result["mode"] == "offline"
        assert result["endpoint"] == "https://example.com/v1/chat/completions"
        assert result["api_key_present"] is True
        assert result["sample_payload"]["law_firm_name"] == "海安律所"
        assert result["sample_payload"]["source_keyword"] == "海安律所避雷"
        assert after["base_url"] == before["base_url"]
        assert after["last_test_status"] == before["last_test_status"]
    finally:
        _restore_singleton_table("ai_configs", ai_snapshot)


def test_ai_profiles_can_be_selected_and_used_for_evaluation(monkeypatch):
    init_db()
    profile_snapshot = _snapshot_table("ai_key_profiles")
    seen: dict[str, Any] = {}

    async def fake_call_openai(cfg, prompt, payload):
        seen.update(cfg)
        return json.dumps(
            {
                "is_related": True,
                "is_negative": True,
                "risk_level": "high",
                "reason": "命中投诉",
                "evidence_quotes": ["投诉"],
                "recommended_action": "人工复核",
            },
            ensure_ascii=False,
        )

    try:
        profile = save_ai_key_profile(
            {
                "name": "主力 OpenAI",
                "provider": "openai",
                "base_url": "https://ai.example.com",
                "api_key": "sk-profile",
                "model": "profile-model",
                "temperature": 0,
                "prompt": DEFAULT_PROMPT,
                "is_active": True,
            }
        )
        monkeypatch.setattr("api.monitoring.ai._call_openai", fake_call_openai)

        result = asyncio.run(
            ai_module.evaluate_content(
                {"law_firm_name": "海安律所"},
                {"platform": "dy", "title": "海安律所投诉", "description": "退费迟迟没有处理"},
                [],
            )
        )

        assert profile["is_active"] is True
        assert get_active_ai_key_profile()["id"] == profile["id"]
        assert seen["model"] == "profile-model"
        assert result["risk_level"] == "high"
    finally:
        _restore_table("ai_key_profiles", profile_snapshot)


def test_ai_rule_test_uses_global_evaluation_prompt_with_active_profile(monkeypatch):
    init_db()
    ai_snapshot = _snapshot_singleton_table("ai_configs")
    profile_snapshot = _snapshot_table("ai_key_profiles")
    email_snapshot = _snapshot_singleton_table("email_configs")
    settings_snapshot = _snapshot_table("system_settings")
    settings_snapshot = _snapshot_table("system_settings")
    settings_snapshot = _snapshot_table("system_settings")
    seen: dict[str, Any] = {}

    async def fake_call_openai(cfg, prompt, payload):
        seen["prompt"] = prompt
        seen["model"] = cfg.get("model")
        return json.dumps(
            {
                "is_related": True,
                "is_negative": False,
                "risk_level": "low",
                "reason": "按当前规则判断",
                "evidence_quotes": [],
                "recommended_action": "人工复核",
            },
            ensure_ascii=False,
        )

    try:
        save_ai_config(
            {
                "provider": "openai",
                "base_url": "",
                "api_key": "",
                "model": "",
                "temperature": 0,
                "prompt": "全局评估规则 Prompt",
            }
        )
        save_ai_key_profile(
            {
                "name": "默认 AI 接入",
                "provider": "openai",
                "base_url": "https://profile.example.com",
                "api_key": "sk-profile",
                "model": "profile-model",
                "temperature": 0,
                "prompt": "不应使用的接入 Prompt",
                "is_active": True,
            }
        )
        monkeypatch.setattr("api.monitoring.ai._call_openai", fake_call_openai)

        result = asyncio.run(run_ai_config_test({}))

        assert result["risk_level"] == "low"
        assert seen["model"] == "profile-model"
        assert seen["prompt"] == "全局评估规则 Prompt"
    finally:
        _restore_singleton_table("ai_configs", ai_snapshot)
        _restore_table("ai_key_profiles", profile_snapshot)


def test_job_bound_ai_profile_takes_precedence_over_active_profile(monkeypatch):
    init_db()
    profile_snapshot = _snapshot_table("ai_key_profiles")
    seen: dict[str, Any] = {}

    async def fake_call_openai(cfg, prompt, payload):
        seen.update(cfg)
        return json.dumps(
            {
                "is_related": True,
                "is_negative": True,
                "risk_level": "medium",
                "reason": "命中退费投诉",
                "evidence_quotes": ["退费投诉"],
                "recommended_action": "人工复核",
            },
            ensure_ascii=False,
        )

    try:
        active = save_ai_key_profile(
            {
                "name": "默认 AI 接入",
                "provider": "openai",
                "base_url": "https://active.example.com",
                "api_key": "sk-active",
                "model": "active-model",
                "temperature": 0,
                "prompt": DEFAULT_PROMPT,
                "is_active": True,
            }
        )
        bound = save_ai_key_profile(
            {
                "name": "海安任务 AI 接入",
                "provider": "openai",
                "base_url": "https://bound.example.com",
                "api_key": "sk-bound",
                "model": "bound-model",
                "temperature": 0,
                "prompt": DEFAULT_PROMPT,
            }
        )
        monkeypatch.setattr("api.monitoring.ai._call_openai", fake_call_openai)

        result = asyncio.run(
            ai_module.evaluate_content(
                {"law_firm_name": "海安律所", "ai_profile_id": bound["id"]},
                {"platform": "dy", "title": "海安律所退费", "description": "投诉退费迟迟没有处理"},
                [],
            )
        )

        assert active["is_active"] is True
        assert seen["model"] == "bound-model"
        assert seen["base_url"] == "https://bound.example.com"
        assert result["risk_level"] == "medium"
    finally:
        _restore_table("ai_key_profiles", profile_snapshot)


def test_evaluate_content_sends_enriched_payload_to_provider(monkeypatch):
    init_db()
    profile_snapshot = _snapshot_table("ai_key_profiles")
    seen_payload: dict[str, Any] = {}

    async def fake_call_openai(cfg, prompt, payload):
        seen_payload.update(payload)
        return json.dumps(
            {
                "is_related": True,
                "is_negative": True,
                "risk_level": "medium",
                "reason": "评论和正文均提到退费投诉",
                "evidence_quotes": ["退费一直没有回复"],
                "recommended_action": "人工复核",
            },
            ensure_ascii=False,
        )

    try:
        save_ai_key_profile(
            {
                "name": "海安律所 OpenAI 接入",
                "provider": "openai",
                "base_url": "https://ai.example.com",
                "api_key": "sk-profile",
                "model": "profile-model",
                "temperature": 0,
                "prompt": DEFAULT_PROMPT,
                "is_active": True,
            }
        )
        monkeypatch.setattr("api.monitoring.ai._call_openai", fake_call_openai)

        result = asyncio.run(
            ai_module.evaluate_content(
                {"law_firm_name": "海安律所", "aliases": ["海安律师"], "exclude_words": []},
                {
                    "platform": "dy",
                    "platform_label": "抖音",
                    "source_keyword": "海安律所投诉",
                    "title": "海安律所投诉记录",
                    "description": "退费一直没有回复。",
                    "author_name": "海安用户",
                    "content_url": "https://www.douyin.com/video/haian-complaint",
                    "cover_url": "https://example.com/haian.jpg",
                    "publish_time": 1781280000,
                    "comment_count": 6,
                },
                [
                    {"content": "我也在等退费", "author_name": "评论用户A", "create_time": 1781280100},
                    {"content": "投诉后有人处理吗", "author_name": "评论用户B", "create_time": 1781280200},
                ],
            )
        )

        assert result["risk_level"] == "medium"
        assert seen_payload["law_firm_name"] == "海安律所"
        assert seen_payload["content_url"].endswith("haian-complaint")
        assert seen_payload["cover_url"].endswith("haian.jpg")
        assert seen_payload["author_name"] == "海安用户"
        assert seen_payload["publish_time"] == 1781280000
        assert seen_payload["comment_summary"]["declared_count"] == 6
        assert seen_payload["comment_summary"]["observed_count"] == 2
        assert seen_payload["comment_samples"][1]["author_name"] == "评论用户B"
        assert "投诉" in seen_payload["comments"][1]
    finally:
        _restore_table("ai_key_profiles", profile_snapshot)


def test_ai_profile_offline_check_uses_profile_without_calling_provider(monkeypatch):
    init_db()
    profile_snapshot = _snapshot_table("ai_key_profiles")
    called = False

    async def fake_call_openai(cfg, prompt, payload):
        nonlocal called
        called = True
        raise RuntimeError("offline profile check must not call provider")

    try:
        profile = save_ai_key_profile(
            {
                "name": "海安律所 OpenAI 接入",
                "provider": "openai",
                "base_url": "https://profile.example.com",
                "api_key": "sk-profile",
                "model": "profile-model",
                "temperature": 0,
                "prompt": DEFAULT_PROMPT,
                "is_active": False,
            }
        )
        monkeypatch.setattr("api.monitoring.ai._call_openai", fake_call_openai)

        result = asyncio.run(monitor_router.ai_profile_offline_check(int(profile["id"])))["result"]

        assert called is False
        assert result["mode"] == "offline"
        assert result["endpoint"] == "https://profile.example.com/v1/chat/completions"
        assert result["model"] == "profile-model"
        assert result["api_key_present"] is True
    finally:
        _restore_table("ai_key_profiles", profile_snapshot)


def test_ai_connection_test_returns_request_and_model_text(monkeypatch):
    monkeypatch.delenv("MONITOR_SKIP_AI_API", raising=False)

    async def fake_ping_openai(cfg):
        assert cfg["model"] == "deepseek-test"
        return {"choices": [{"message": {"content": "pong from model"}}]}

    monkeypatch.setattr(ai_module, "_ping_openai", fake_ping_openai)

    result = asyncio.run(
        ai_module.test_ai_connection(
            {
                "provider": "openai",
                "base_url": "https://api.example.com",
                "api_key": "sk-test",
                "model": "deepseek-test",
                "temperature": 0,
            }
        )
    )

    assert result["ok"] is True
    assert result["protocol"] == "openai"
    assert result["request_message"] == "hi"
    assert result["response_text"] == "pong from model"
    assert result["response_preview"] == "pong from model"
    assert "返回文本" in result["message"]


def test_ai_connection_test_reports_empty_response_shape(monkeypatch):
    monkeypatch.delenv("MONITOR_SKIP_AI_API", raising=False)

    async def fake_ping_openai(cfg):
        return {"id": "chatcmpl-empty", "choices": [{"message": {"content": ""}}]}

    monkeypatch.setattr(ai_module, "_ping_openai", fake_ping_openai)

    with pytest.raises(ValueError) as exc:
        asyncio.run(
            ai_module.test_ai_connection(
                {
                    "provider": "openai",
                    "base_url": "https://api.example.com",
                    "api_key": "sk-test",
                    "model": "deepseek-test",
                    "temperature": 0,
                }
            )
        )

    assert "没有返回文本" in str(exc.value)
    assert "chatcmpl-empty" in str(exc.value)


def test_anthropic_connection_test_uses_content_blocks_and_larger_token_budget(monkeypatch):
    captured = {}

    class FakeResponse:
        def raise_for_status(self):
            return None

        def json(self):
            return {"content": [{"type": "text", "text": "Hi"}]}

    class FakeClient:
        def __init__(self, *args, **kwargs):
            captured["client_kwargs"] = kwargs

        async def __aenter__(self):
            return self

        async def __aexit__(self, exc_type, exc, tb):
            return None

        async def post(self, url, headers, json):
            captured["url"] = url
            captured["headers"] = headers
            captured["json"] = json
            return FakeResponse()

    monkeypatch.setattr(ai_module.httpx, "AsyncClient", FakeClient)

    result = asyncio.run(
        ai_module.test_ai_connection(
            {
                "provider": "anthropic",
                "base_url": "https://api.example.com/anthropic",
                "api_key": "sk-test",
                "model": "compatible-model",
                "temperature": 0,
            }
        )
    )

    assert result["response_text"] == "Hi"
    assert result["protocol"] == "anthropic"
    assert captured["url"] == "https://api.example.com/anthropic/v1/messages"
    assert captured["json"]["max_tokens"] == ai_module.AI_CONNECTION_TEST_MAX_TOKENS
    assert captured["json"]["messages"] == [
        {"role": "user", "content": [{"type": "text", "text": "hi"}]}
    ]


def test_ai_model_text_extractor_supports_anthropic_and_compatible_shapes():
    assert ai_module._extract_model_text({"content": [{"text": "Hi there"}]}) == "Hi there"
    assert ai_module._extract_model_text({"content": "Hi from string"}) == "Hi from string"
    assert ai_module._extract_model_text({"choices": [{"message": {"content": [{"type": "text", "text": "Hi from array"}]}}]}) == "Hi from array"
    assert ai_module._extract_model_text({"choices": [{"text": "Hi from choice"}]}) == "Hi from choice"


def test_ai_model_list_fetches_openai_compatible_models(monkeypatch):
    captured = {}

    class FakeResponse:
        def raise_for_status(self):
            return None

        def json(self):
            return {"data": [{"id": "deepseek-v4-flash"}, {"id": "deepseek-reasoner"}]}

    class FakeClient:
        def __init__(self, *args, **kwargs):
            captured["client_kwargs"] = kwargs

        async def __aenter__(self):
            return self

        async def __aexit__(self, exc_type, exc, tb):
            return None

        async def get(self, url, headers):
            captured["url"] = url
            captured["headers"] = headers
            return FakeResponse()

    monkeypatch.setattr(ai_module.httpx, "AsyncClient", FakeClient)

    result = asyncio.run(
        ai_module.list_ai_models(
            {
                "provider": "openai",
                "base_url": "https://api.example.com",
                "api_key": "sk-test",
            }
        )
    )

    assert result["models"] == ["deepseek-v4-flash", "deepseek-reasoner"]
    assert result["endpoint"] == "https://api.example.com/v1/models"
    assert captured["headers"]["Authorization"] == "Bearer sk-test"


def test_ai_model_list_fetches_anthropic_compatible_models(monkeypatch):
    captured = {}

    class FakeResponse:
        def raise_for_status(self):
            return None

        def json(self):
            return {"data": [{"id": "claude-compatible"}, {"name": "custom-model"}]}

    class FakeClient:
        def __init__(self, *args, **kwargs):
            captured["client_kwargs"] = kwargs

        async def __aenter__(self):
            return self

        async def __aexit__(self, exc_type, exc, tb):
            return None

        async def get(self, url, headers):
            captured["url"] = url
            captured["headers"] = headers
            return FakeResponse()

    monkeypatch.setattr(ai_module.httpx, "AsyncClient", FakeClient)

    result = asyncio.run(
        ai_module.list_ai_models(
            {
                "provider": "anthropic",
                "base_url": "https://api.example.com/anthropic",
                "api_key": "sk-test",
            }
        )
    )

    assert result["models"] == ["claude-compatible", "custom-model"]
    assert result["endpoint"] == "https://api.example.com/anthropic/v1/models"
    assert captured["headers"]["x-api-key"] == "sk-test"
    assert captured["headers"]["Authorization"] == "Bearer sk-test"
    assert captured["headers"]["anthropic-version"] == "2023-06-01"


def test_ai_model_list_falls_back_from_adapter_path_to_parent_models(monkeypatch):
    attempts = []

    class FakeResponse:
        def __init__(self, url: str, ok: bool):
            self.url = url
            self.ok = ok

        def raise_for_status(self):
            if self.ok:
                return None
            request = httpx.Request("GET", self.url)
            response = httpx.Response(404, request=request)
            raise httpx.HTTPStatusError("not found", request=request, response=response)

        def json(self):
            return {"data": [{"id": "deepseek-v4-flash"}]}

    class FakeClient:
        def __init__(self, *args, **kwargs):
            pass

        async def __aenter__(self):
            return self

        async def __aexit__(self, exc_type, exc, tb):
            return None

        async def get(self, url, headers):
            attempts.append(url)
            return FakeResponse(url, ok=url == "https://api.example.com/models")

    monkeypatch.setattr(ai_module.httpx, "AsyncClient", FakeClient)

    result = asyncio.run(
        ai_module.list_ai_models(
            {
                "provider": "anthropic",
                "base_url": "https://api.example.com/anthropic",
                "api_key": "sk-test",
            }
        )
    )

    assert result["models"] == ["deepseek-v4-flash"]
    assert result["endpoint"] == "https://api.example.com/models"
    assert attempts[:3] == [
        "https://api.example.com/anthropic/v1/models",
        "https://api.example.com/anthropic/models",
        "https://api.example.com/models",
    ]


def test_ai_model_list_requires_connection_fields(monkeypatch):
    with pytest.raises(ValueError) as exc:
        asyncio.run(ai_module.list_ai_models({"provider": "openai", "base_url": "", "api_key": ""}))

    assert "AI 接入未配置完整" in str(exc.value)


def test_ai_profile_real_test_respects_skip_env_and_records_result(monkeypatch):
    init_db()
    profile_snapshot = _snapshot_table("ai_key_profiles")
    try:
        profile = save_ai_key_profile(
            {
                "name": "海安律所 Anthropic Profile",
                "provider": "anthropic",
                "base_url": "https://anthropic.example.com",
                "api_key": "sk-ant",
                "model": "claude-test",
                "temperature": 0,
                "prompt": DEFAULT_PROMPT,
                "is_active": False,
            }
        )
        monkeypatch.setenv("MONITOR_SKIP_AI_API", "true")

        with pytest.raises(HTTPException) as exc:
            asyncio.run(monitor_router.test_ai_profile(int(profile["id"])))

        assert exc.value.status_code == 400
        assert "未启用" in exc.value.detail
        tested = next(item for item in list_ai_key_profiles() if item["id"] == profile["id"])
        assert tested["last_test_status"] == "failed"
        assert "未启用" in tested["last_test_error"]
    finally:
        _restore_table("ai_key_profiles", profile_snapshot)


def test_readiness_and_doctor_prefer_active_ai_profile(monkeypatch):
    init_db()
    ai_snapshot = _snapshot_singleton_table("ai_configs")
    profile_snapshot = _snapshot_table("ai_key_profiles")
    try:
        save_ai_config(
            {
                "provider": "openai",
                "base_url": "",
                "api_key": "",
                "model": "",
                "temperature": 0,
                "prompt": DEFAULT_PROMPT,
            }
        )
        profile = save_ai_key_profile(
            {
                "name": "海安律所当前 AI 接入",
                "provider": "openai",
                "base_url": "https://profile.example.com",
                "api_key": "sk-profile",
                "model": "profile-model",
                "temperature": 0,
                "prompt": DEFAULT_PROMPT,
                "is_active": True,
            }
        )
        profile = monitor_router.mark_ai_key_profile_test_result(int(profile["id"]), True)
        assert profile["last_test_status"] == "success"

        readiness_check = next(check for check in get_readiness_status()["checks"] if check["key"] == "ai_config")
        doctor_check = next(check for check in run_doctor()["checks"] if check["key"] == "ai_config")

        assert readiness_check["ok"] is True
        assert "profile-model" in readiness_check["message"]
        assert doctor_check["ok"] is True
        assert "profile-model" in doctor_check["message"]
    finally:
        _restore_singleton_table("ai_configs", ai_snapshot)
        _restore_table("ai_key_profiles", profile_snapshot)


def test_email_templates_preview_and_pool_configs_are_persisted():
    init_db()
    snapshots = {
        "email_templates": _snapshot_table("email_templates"),
        "social_accounts": _snapshot_table("social_accounts"),
        "proxy_profiles": _snapshot_table("proxy_profiles"),
    }
    try:
        preview = render_email_template_preview(
            {
                "subject_template": "日报 {law_firm_name} {date}",
                "html_template": "<h1>{law_firm_name}</h1><section>{report_html}</section>",
            }
        )
        template = save_email_template(
            {
                "name": "企业日报模板",
                "subject_template": "日报 {law_firm_name}",
                "html_template": "<main>{report_html}</main>",
                "is_active": True,
            }
        )
        proxy = save_proxy_profile(
            {
                "name": "华东代理池",
                "provider": "manual",
                "proxy_url": "http://user:pass@127.0.0.1:8081",
                "status": "active",
                "max_concurrency": 2,
            }
        )
        account = save_social_account(
            {
                "name": "抖音一号",
                "platform": "dy",
                "login_type": "qrcode",
                "status": "active",
                "proxy_id": proxy["id"],
            }
        )
        summary = get_dashboard_summary()

        assert preview["subject"].startswith("日报 海安律所")
        assert "{report_html}" not in preview["html"]
        assert "海安律所退费投诉" in preview["html"]
        assert template["is_active"] is True
        assert list_email_templates()[0]["id"] == template["id"]
        assert list_proxy_profiles()[0]["proxy_url"].startswith("htt")
        assert account["platform"] == "dy"
        assert account["proxy_name"] == "华东代理池"
        assert account["proxy_status"] == "active"
        assert list_social_accounts()[0]["name"] == "抖音一号"
        assert list_social_accounts()[0]["proxy_name"] == "华东代理池"
        assert summary["social_accounts_total"] >= 1
        assert summary["proxy_profiles_total"] >= 1
        with pytest.raises(ValueError, match="proxy not found"):
            save_social_account(
                {
                    "name": "海安律所异常代理账号",
                    "platform": "dy",
                    "login_type": "qrcode",
                    "status": "standby",
                    "proxy_id": 99999999,
                }
            )
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_social_account_pool_redacts_signed_avatar_urls():
    init_db()
    snapshot = _snapshot_table("social_accounts")
    try:
        account = save_social_account({"name": "签名头像账号", "platform": "dy", "login_type": "qrcode", "status": "active"})
        update_social_account_check_state(
            int(account["id"]),
            True,
            "登录态有效",
            identity={
                "platform_account_name": "签名头像账号",
                "platform_avatar_url": "https://p3-pc-sign.douyinpic.com/avatar.jpeg?x-expires=1799999999&x-signature=secretSig&lk3s=abc",
                "platform_home_url": "https://www.douyin.com/user/self?from_nav=1",
            },
        )

        listed = list_social_accounts(masked=True)
        visible = json.dumps(listed, ensure_ascii=False)
        target = next(item for item in listed if item["id"] == account["id"])
    finally:
        _restore_table("social_accounts", snapshot)

    assert target["platform_avatar_url"] == "https://p3-pc-sign.douyinpic.com/avatar.jpeg"
    assert target["platform_home_url"] == "https://www.douyin.com/user/self"
    for forbidden in ["x-signature", "x-expires", "lk3s", "secretSig", "from_nav"]:
        assert forbidden not in visible


def test_social_account_api_uses_local_avatar_endpoint_without_exposing_source(monkeypatch, tmp_path):
    import api.routers.monitor as monitor_router

    init_db()
    snapshot = _snapshot_table("social_accounts")
    avatar_root = tmp_path / "avatars"
    monkeypatch.setattr(monitor_router, "AVATAR_CACHE_DIR", avatar_root)
    try:
        account = save_social_account({"name": "签名头像账号", "platform": "dy", "login_type": "qrcode", "status": "active"})
        update_social_account_check_state(
            int(account["id"]),
            True,
            "登录态有效",
            identity={
                "platform_account_name": "签名头像账号",
                "platform_avatar_url": "https://p3-pc-sign.douyinpic.com/avatar.jpeg?x-expires=1799999999&x-signature=secretSig&lk3s=abc",
                "platform_home_url": "https://www.douyin.com/user/self?from_nav=1",
            },
        )

        account_view = monitor_router._customer_view_social_account(list_social_accounts(masked=False)[0])
        visible = json.dumps(account_view, ensure_ascii=False)
        avatar_url = account_view["platform_avatar_url"]
    finally:
        _restore_table("social_accounts", snapshot)

    assert avatar_url == f"/api/monitor/social-accounts/{account['id']}/avatar"
    assert "douyinpic.com" not in avatar_url
    assert "x-signature" not in visible
    assert "secretSig" not in visible
    assert "from_nav" not in visible
    assert not list(avatar_root.glob("account_*"))


def test_social_account_avatar_endpoint_is_admin_only_and_caches_signed_source(monkeypatch, tmp_path):
    import api.main as api_main
    import api.monitoring.avatar_cache as avatar_cache_module
    import api.routers.monitor as monitor_router

    init_db()
    snapshots = {
        "users": _snapshot_table("users"),
        "user_sessions": _snapshot_table("user_sessions"),
        "social_accounts": _snapshot_table("social_accounts"),
    }
    avatar_root = tmp_path / "avatars"
    monkeypatch.setattr(avatar_cache_module, "AVATAR_CACHE_DIR", avatar_root)
    monkeypatch.setattr(monitor_router, "AVATAR_CACHE_DIR", avatar_root)

    class FakeResponse:
        headers = {"Content-Type": "image/jpeg"}

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

        def read(self, _size):
            return b"\xff\xd8\xffcached-avatar"

    def fake_urlopen(request, timeout=10.0):
        assert "x-signature=secretSig" in request.full_url
        return FakeResponse()

    monkeypatch.setattr(avatar_cache_module.urllib.request, "urlopen", fake_urlopen)
    try:
        account = save_social_account({"name": "签名头像账号", "platform": "dy", "login_type": "qrcode", "status": "active"})
        update_social_account_check_state(
            int(account["id"]),
            True,
            "登录态有效",
            identity={
                "platform_account_name": "签名头像账号",
                "platform_avatar_url": "https://p3-pc-sign.douyinpic.com/avatar.jpeg?x-expires=1799999999&x-signature=secretSig&lk3s=abc",
            },
        )
        save_user(
            {
                "email": "avatar-admin@example.com",
                "display_name": "Avatar Admin",
                "password": "AdminPass123!",
                "role": "administrator",
                "status": "active",
            },
            actor_id=1,
        )
        save_user(
            {
                "email": "avatar-user@example.com",
                "display_name": "Avatar User",
                "password": "UserPass123!",
                "role": "normal",
                "status": "active",
            },
            actor_id=1,
        )
        transport = httpx.ASGITransport(app=api_main.app)

        async def exercise() -> None:
            async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as client:
                normal_login = await client.post("/api/auth/login", json={"email": "avatar-user@example.com", "password": "UserPass123!"})
                assert normal_login.status_code == 200
                assert (await client.get(f"/api/monitor/social-accounts/{account['id']}/avatar")).status_code == 403

            async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as client:
                admin_login = await client.post("/api/auth/login", json={"email": "avatar-admin@example.com", "password": "AdminPass123!"})
                assert admin_login.status_code == 200
                accounts_response = await client.get("/api/monitor/social-accounts")
                assert accounts_response.status_code == 200
                accounts_visible = json.dumps(accounts_response.json(), ensure_ascii=False)
                assert f"/api/monitor/social-accounts/{account['id']}/avatar" in accounts_visible
                assert "douyinpic.com" not in accounts_visible
                assert "secretSig" not in accounts_visible

                response = await client.get(f"/api/monitor/social-accounts/{account['id']}/avatar")
                assert response.status_code == 200
                assert response.content == b"\xff\xd8\xffcached-avatar"
                traversal = await client.get("/api/monitor/social-accounts/avatar/..%2Fsecret.key")
                assert traversal.status_code == 404

        asyncio.run(exercise())
        cached_files = list(avatar_root.glob("account_*"))
        assert len(cached_files) == 1
        assert cached_files[0].read_bytes() == b"\xff\xd8\xffcached-avatar"
    finally:
        for table, snapshot_rows in snapshots.items():
            _restore_table(table, snapshot_rows)


def test_social_account_api_view_hides_profile_paths_and_cookie_values():
    raw = {
        "id": 1,
        "name": "安全账号",
        "platform": "dy",
        "login_type": "qrcode",
        "status": "active",
        "profile_key": "1/dy/acc_1",
        "profile_path": r"E:\server\profiles\dy\acc_1",
        "profile_runtime_path": r"E:\server\profiles\dy\acc_1",
        "profile_configured": True,
        "cookies": "sessionid=secret",
        "has_cookies": True,
        "platform_avatar_url": "https://p3.example.com/avatar.jpeg?x-signature=secret&x-expires=1799999999",
        "platform_home_url": "https://www.douyin.com/user/self?from_nav=1",
    }

    view = monitor_router._customer_view_social_account(raw)
    visible = json.dumps(view, ensure_ascii=False)

    assert view["profile_configured"] is True
    assert view["has_cookies"] is True
    assert "profile_path" not in view
    assert "profile_runtime_path" not in view
    assert "cookies" not in view
    assert "E:\\" not in visible
    assert "sessionid" not in visible
    assert "x-signature" not in visible
    assert "x-expires" not in visible
    assert "from_nav" not in visible


def test_init_db_creates_default_email_template_when_empty():
    init_db()
    snapshot = _snapshot_table("email_templates")
    try:
        with get_conn() as conn:
            conn.execute("DELETE FROM email_templates")
        init_db()
        templates = list_email_templates()

        assert len(templates) == 1
        assert templates[0]["name"] == "标准舆情日报模板"
        assert templates[0]["is_active"] is True
        assert "{report_html}" in templates[0]["html_template"]
    finally:
        _restore_table("email_templates", snapshot)


def test_login_sessions_are_persisted_for_server_side_login_flow():
    init_db()
    snapshot = _snapshot_table("login_sessions")
    try:
        session = create_login_session(
            {
                "platform": "dy",
                "account_id": None,
                "login_url": "https://www.douyin.com/",
                "profile_path": "browser_data/cdp_dy_user_data_dir",
            }
        )
        listed = list_login_sessions()
        summary = get_dashboard_summary()

        assert session["status"] == "preparing"
        assert get_login_session(session["id"])["platform"] == "dy"
        assert listed[0]["id"] == session["id"]
        assert summary["login_sessions_total"] >= 1
        with pytest.raises(ValueError, match="unsupported platform"):
            create_login_session({"platform": "wb"})
    finally:
        _restore_table("login_sessions", snapshot)


def test_login_sessions_can_be_expired_for_same_account():
    init_db()
    snapshot = _snapshot_table("login_sessions")
    try:
        first = create_login_session(
            {
                "platform": "dy",
                "account_id": 10001,
                "login_url": "https://www.douyin.com/",
                "profile_path": "browser_data/account_10001",
            }
        )
        other = create_login_session(
            {
                "platform": "dy",
                "account_id": 10002,
                "login_url": "https://www.douyin.com/",
                "profile_path": "browser_data/account_10002",
            }
        )

        expired = expire_login_sessions_for_account(10001, "dy", "browser_data/account_10001")

        assert expired == [first["id"]]
        assert get_login_session(first["id"])["status"] == "timeout"
        assert get_login_session(other["id"])["status"] == "preparing"
    finally:
        _restore_table("login_sessions", snapshot)


def test_login_session_routes_create_pollable_session(monkeypatch):
    init_db()
    snapshots = {"login_sessions": _snapshot_table("login_sessions"), "social_accounts": _snapshot_table("social_accounts")}
    try:
        account = _login_test_account("dy")
        monkeypatch.setattr(
            monitor_router,
            "build_login_browser_command",
            lambda platform: {
                "platform": platform,
                "platform_label": "抖音",
                "login_url": "https://www.douyin.com/",
                "profile_path": "browser_data/cdp_dy_user_data_dir",
                "debug_port": 9323,
                "browser_path": "chrome",
            },
        )
        async def fake_start_qrcode_login_session_with_profile(session_id, platform, command):
            return {
                "ok": True,
                "qr_image": "data:image/png;base64,abc",
                "message": "请扫码登录",
                "profile_path": "browser_data/cdp_dy_user_data_dir",
            }

        async def fake_poll_qrcode_login_session(session_id):
            return {"active": True, "success": False, "message": "等待扫码确认。"}

        monkeypatch.setattr(monitor_router, "start_qrcode_login_session_with_profile", fake_start_qrcode_login_session_with_profile)
        monkeypatch.setattr(monitor_router, "poll_qrcode_login_session", fake_poll_qrcode_login_session)
        monkeypatch.setattr(
            monitor_router,
            "list_platform_status",
            lambda: [
                {
                    "platform": "dy",
                    "platform_label": "抖音",
                    "profile_path": "browser_data/cdp_dy_user_data_dir",
                    "login_ready": False,
                    "login_window_open": False,
                }
            ],
        )

        created = asyncio.run(monitor_router.create_platform_login_session({"platform": "dy", "account_id": account["id"]}))
        session_id = created["session"]["id"]
        polled = asyncio.run(monitor_router.login_session(session_id))

        assert created["capabilities"]["manual_browser_fallback"] is True
        assert created["capabilities"]["login_capability_source"] == "平台采集服务"
        assert created["capabilities"]["login_boundary"] == "复用平台采集服务登录能力"
        assert "验证码" in created["capabilities"]["captcha_policy"]
        assert created["capabilities"]["qr_image_supported"] is True
        assert created["session"]["status"] == "waiting_qrcode"
        assert created["session"]["qr_image"].startswith("data:image")
        assert polled["session"]["status"] == "waiting_scan"
        assert polled["platform_status"]["platform"] == "dy"
        assert polled["capabilities"]["login_capability_source"] == "平台采集服务"
        assert polled["capabilities"]["login_boundary"] == "复用平台采集服务登录能力"
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_login_session_route_falls_back_when_qrcode_unavailable(monkeypatch):
    init_db()
    snapshots = {"login_sessions": _snapshot_table("login_sessions"), "social_accounts": _snapshot_table("social_accounts")}
    try:
        account = _login_test_account("dy")
        monkeypatch.setattr(
            monitor_router,
            "build_login_browser_command",
            lambda platform: {
                "platform": platform,
                "platform_label": "抖音",
                "login_url": "https://www.douyin.com/",
                "profile_path": "browser_data/cdp_dy_user_data_dir",
                "debug_port": 9323,
                "browser_path": "chrome",
            },
        )

        async def fake_start_qrcode_login_session_with_profile(session_id, platform, command):
            return {
                "ok": False,
                "message": "没有在页面中找到登录二维码，请使用网页登录窗口处理。当前页面：登录页",
                "diagnostic_image": "data:image/png;base64,diagnostic",
                "profile_path": "browser_data/cdp_dy_user_data_dir",
            }

        monkeypatch.setattr(monitor_router, "start_qrcode_login_session_with_profile", fake_start_qrcode_login_session_with_profile)

        created = asyncio.run(monitor_router.create_platform_login_session({"platform": "dy", "account_id": account["id"]}))

        assert created["capabilities"]["manual_browser_fallback"] is True
        assert created["capabilities"]["qr_image_supported"] is False
        assert created["capabilities"]["diagnostic_image_supported"] is False
        assert created["capabilities"]["diagnostic_image"] == ""
        assert created["session"]["status"] == "qrcode_failed"
        assert "网页登录窗口处理" in created["session"]["message"]
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_login_session_route_keeps_preparing_session_pending_before_qr_handle_exists(monkeypatch):
    init_db()
    snapshot = _snapshot_table("login_sessions")
    try:
        session = create_login_session(
            {
                "platform": "dy",
                "account_id": 10003,
                "login_url": "https://www.douyin.com/",
                "profile_key": "1/dy/acc_10003",
                "profile_path": "browser_data/account_10003",
                "message": "正在生成登录二维码。",
            }
        )

        async def fake_poll_qrcode_login_session(session_id):
            return {
                "active": False,
                "success": False,
                "status": "qrcode_failed",
                "message": "二维码浏览器会话不在运行，请重新生成二维码或打开登录窗口。",
            }

        monkeypatch.setattr(monitor_router, "poll_qrcode_login_session", fake_poll_qrcode_login_session)
        monkeypatch.setattr(monitor_router, "list_platform_status", lambda: [])

        polled = asyncio.run(monitor_router.login_session(int(session["id"])))
        refreshed = get_login_session(int(session["id"]))

        assert polled["session"]["status"] == "preparing"
        assert polled["session"]["message"] == "正在生成登录二维码。"
        assert refreshed["status"] == "preparing"
    finally:
        _restore_table("login_sessions", snapshot)


def test_login_session_route_expires_stale_preparing_session_without_qr_handle(monkeypatch):
    init_db()
    snapshot = _snapshot_table("login_sessions")
    try:
        session = create_login_session(
            {
                "platform": "dy",
                "account_id": 10004,
                "login_url": "https://www.douyin.com/",
                "profile_key": "1/dy/acc_10004",
                "profile_path": "browser_data/account_10004",
                "message": "正在生成登录二维码。",
            }
        )
        old = (datetime.now(timezone.utc) - timedelta(seconds=120)).isoformat()
        with get_conn() as conn:
            conn.execute("UPDATE login_sessions SET created_at=?, updated_at=? WHERE id=?", (old, old, session["id"]))

        async def fake_poll_qrcode_login_session(session_id):
            return {
                "active": False,
                "success": False,
                "status": "qrcode_failed",
                "message": "二维码浏览器会话不在运行，请重新生成二维码或打开登录窗口。",
            }

        monkeypatch.setattr(monitor_router, "poll_qrcode_login_session", fake_poll_qrcode_login_session)
        monkeypatch.setattr(monitor_router, "list_platform_status", lambda: [])

        polled = asyncio.run(monitor_router.login_session(int(session["id"])))

        assert polled["session"]["status"] == "qrcode_failed"
        assert "二维码浏览器会话不在运行" in polled["session"]["message"]
    finally:
        _restore_table("login_sessions", snapshot)


@pytest.mark.parametrize("platform", ["dy", "xhs", "ks"])
def test_login_session_failure_reconciles_successful_same_account_check_for_supported_platforms(platform, monkeypatch):
    init_db()
    snapshots = {"login_sessions": _snapshot_table("login_sessions"), "social_accounts": _snapshot_table("social_accounts")}
    try:
        account = _login_test_account(platform)
        session = create_login_session(
            {
                "platform": platform,
                "account_id": account["id"],
                "login_url": get_mediacrawler_login_capability(platform)["login_url"],
                "profile_key": account["profile_key"],
                "profile_path": account["profile_path"],
                "qr_image": "data:image/png;base64,existing",
            }
        )
        monitor_router.update_login_session_status(
            int(session["id"]),
            "waiting_scan",
            "二维码已生成，请扫码登录。",
            "data:image/png;base64,existing",
        )

        async def fake_poll_qrcode_login_session(session_id):
            return {
                "active": False,
                "success": False,
                "status": "qrcode_failed",
                "message": "二维码浏览器会话不在运行，请重新生成二维码或打开登录窗口。",
            }

        async def fake_check_social_account_login(
            account_id,
            timeout_ms=15000,
            allow_draft=False,
            identity_prepared=False,
            actor_id=None,
        ):
            return {
                "ok": True,
                "account": update_social_account_check_state(
                    account_id,
                    True,
                    "登录态有效，可供采集任务使用。",
                    identity={"nickname": "已验活账号"},
                ),
            }

        monkeypatch.setattr(monitor_router, "poll_qrcode_login_session", fake_poll_qrcode_login_session)
        monkeypatch.setattr(monitor_router, "check_social_account_login", fake_check_social_account_login)
        monkeypatch.setattr(monitor_router, "list_platform_status", lambda: [])

        polled = asyncio.run(monitor_router.login_session(int(session["id"])))
        refreshed = get_social_account(int(account["id"]))
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)

    assert polled["session"]["status"] == "success"
    assert "通过验活" in polled["session"]["message"]
    assert polled["account_status"]["status"] == "active"
    assert refreshed["status"] == "active"
    assert refreshed["last_error"] == ""


def test_login_session_failure_reconciliation_keeps_failure_when_account_check_fails(monkeypatch):
    init_db()
    snapshots = {"login_sessions": _snapshot_table("login_sessions"), "social_accounts": _snapshot_table("social_accounts")}
    try:
        account = _login_test_account("xhs")
        session = create_login_session(
            {
                "platform": "xhs",
                "account_id": account["id"],
                "login_url": "https://www.xiaohongshu.com",
                "profile_key": account["profile_key"],
                "profile_path": account["profile_path"],
            }
        )
        monitor_router.update_login_session_status(int(session["id"]), "waiting_scan", "二维码已生成，请扫码登录。")

        async def fake_poll_qrcode_login_session(session_id):
            return {
                "active": False,
                "success": False,
                "status": "qrcode_failed",
                "message": "二维码浏览器会话不在运行，请重新生成二维码或打开登录窗口。",
            }

        async def fake_check_social_account_login(
            account_id,
            timeout_ms=15000,
            allow_draft=False,
            identity_prepared=False,
            actor_id=None,
        ):
            return {
                "ok": False,
                "message": "登录态无效或已失效，请重新扫码登录。",
                "account": update_social_account_check_state(account_id, False, "登录态无效或已失效，请重新扫码登录。"),
            }

        monkeypatch.setattr(monitor_router, "poll_qrcode_login_session", fake_poll_qrcode_login_session)
        monkeypatch.setattr(monitor_router, "check_social_account_login", fake_check_social_account_login)
        monkeypatch.setattr(monitor_router, "list_platform_status", lambda: [])

        polled = asyncio.run(monitor_router.login_session(int(session["id"])))
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)

    assert polled["session"]["status"] == "qrcode_failed"
    assert "二维码浏览器会话不在运行" in polled["session"]["message"]
    assert polled["account_status"]["status"] == "limited"


def test_xhs_account_check_reports_legacy_profile_path_hint(tmp_path, monkeypatch):
    init_db()
    snapshot = _snapshot_table("social_accounts")
    try:
        legacy_path = tmp_path / "xhs_legacy_profile"
        legacy_path.mkdir()
        account = save_social_account({"name": "小红书采集号", "platform": "xhs", "login_type": "qrcode", "status": "standby"})
        with get_conn() as conn:
            conn.execute("UPDATE social_accounts SET profile_path=? WHERE id=?", (str(legacy_path), account["id"]))
        checked = asyncio.run(account_check_module.check_social_account_login(int(account["id"])))
    finally:
        _restore_table("social_accounts", snapshot)

    assert checked["ok"] is False
    assert checked["status"] == "missing_profile"
    assert "旧版网页登录态目录" in checked["message"]
    assert "重新扫码登录" in checked["message"]


def test_phase_6_production_mode_hides_local_login_window(monkeypatch):
    monkeypatch.setenv("MONITOR_ALLOW_LOCAL_LOGIN_WINDOW", "false")

    caps = monitor_router._login_capability_response("dy")

    assert caps["primary_login_flow"] == "server_qrcode"
    assert caps["manual_browser_fallback"] is False
    assert caps["local_login_window_allowed"] is False
    with pytest.raises(HTTPException) as exc:
        asyncio.run(monitor_router.platform_login_browser("dy", {}))
    assert exc.value.status_code == 403
    assert "网页登录二维码" in exc.value.detail


def test_cr108_qr_login_is_blocked_when_same_profile_login_window_is_open(monkeypatch, tmp_path):
    init_db()
    snapshots = {
        "login_sessions": _snapshot_table("login_sessions"),
        "social_accounts": _snapshot_table("social_accounts"),
    }
    try:
        account = _login_test_account("dy", tmp_path)
        profile_path = str(resolve_account_profile_path(f"1/dy/acc_{account['id']}"))
        started = {"called": False}

        monkeypatch.setattr(
            monitor_router,
            "build_login_browser_command",
            lambda platform: {
                "platform": platform,
                "platform_label": "抖音",
                "login_url": "https://www.douyin.com/",
                "profile_path": str(tmp_path / "default_profile"),
                "debug_port": 9323,
                "browser_path": "chrome",
            },
        )
        login_state_dir = tmp_path / "login_windows"
        monkeypatch.setattr("api.monitoring.login_state.LOGIN_STATE_DIR", login_state_dir)
        monkeypatch.setattr("api.monitoring.login_state._pid_exists", lambda pid: True)
        record_login_window("dy", 12345, 9323, profile_path)

        async def fake_start_qrcode_login_session_with_profile(session_id, platform, command):
            started["called"] = True
            return {"ok": True, "qr_image": "data:image/png;base64,abc", "message": "请扫码登录"}

        monkeypatch.setattr(
            monitor_router,
            "start_qrcode_login_session_with_profile",
            fake_start_qrcode_login_session_with_profile,
        )

        created = asyncio.run(monitor_router.create_platform_login_session({"platform": "dy", "account_id": account["id"]}))

        assert started["called"] is False
        assert created["session"]["status"] == "needs_verification"
        assert "登录窗口正在使用该账号" in created["session"]["message"]
        assert "TargetClosedError" not in created["session"]["message"]
        assert profile_path not in created["session"]["message"]
        assert created["capabilities"]["local_login_window_allowed"] is True
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_cr108_qr_login_is_blocked_when_same_profile_key_login_window_is_open(monkeypatch, tmp_path):
    init_db()
    snapshots = {
        "login_sessions": _snapshot_table("login_sessions"),
        "social_accounts": _snapshot_table("social_accounts"),
    }
    try:
        account = _login_test_account("dy", tmp_path)
        started = {"called": False}

        monkeypatch.setattr(
            monitor_router,
            "build_login_browser_command",
            lambda platform: {
                "platform": platform,
                "platform_label": "抖音",
                "login_url": "https://www.douyin.com/",
                "profile_path": str(tmp_path / "default_profile"),
                "debug_port": 9323,
                "browser_path": "chrome",
            },
        )
        monkeypatch.setattr(
            monitor_router,
            "login_window_status",
            lambda platform: {
                "is_open": True,
                "pid": 12345,
                "debug_port": 9323,
                "profile_key": account["profile_key"],
                "profile_path": str(tmp_path / "different_runtime_path"),
            },
        )

        async def fake_start_qrcode_login_session_with_profile(session_id, platform, command):
            started["called"] = True
            return {"ok": True, "qr_image": "data:image/png;base64,abc", "message": "请扫码登录"}

        monkeypatch.setattr(
            monitor_router,
            "start_qrcode_login_session_with_profile",
            fake_start_qrcode_login_session_with_profile,
        )

        created = asyncio.run(monitor_router.create_platform_login_session({"platform": "dy", "account_id": account["id"]}))

        assert started["called"] is False
        assert created["session"]["status"] == "needs_verification"
        assert "登录窗口正在使用该账号" in created["session"]["message"]
        assert account["profile_key"] not in created["session"]["message"]
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_cr108_opening_login_window_supersedes_active_qr_session(monkeypatch, tmp_path):
    init_db()
    snapshots = {
        "login_sessions": _snapshot_table("login_sessions"),
        "social_accounts": _snapshot_table("social_accounts"),
    }
    closed: list[int] = []
    seen: dict[str, Any] = {}
    try:
        account = _login_test_account("dy", tmp_path)
        profile_path = str(resolve_account_profile_path(f"1/dy/acc_{account['id']}"))
        session = create_login_session(
            {
                "platform": "dy",
                "account_id": account["id"],
                "login_url": "https://www.douyin.com/",
                "profile_key": account["profile_key"],
                "profile_path": profile_path,
                "message": "二维码已生成，请扫码登录。",
            }
        )
        monitor_router.update_login_session_status(int(session["id"]), "waiting_qrcode", "二维码已生成，请扫码登录。")

        monkeypatch.setattr(
            monitor_router,
            "build_login_browser_command",
            lambda platform: {
                "platform": platform,
                "platform_label": "抖音",
                "login_url": "https://www.douyin.com/",
                "profile_path": str(tmp_path / "default_profile"),
                "debug_port": 9323,
                "browser_path": "chrome",
            },
        )

        async def fake_close_qrcode_login_session(session_id):
            closed.append(int(session_id))

        def fake_open_login_browser_with_command(command):
            seen["profile_path"] = command["profile_path"]
            return {**command, "pid": 45678, "message": "ok"}

        monkeypatch.setattr(monitor_router, "close_qrcode_login_session", fake_close_qrcode_login_session)
        monkeypatch.setattr(monitor_router, "open_login_browser_with_command", fake_open_login_browser_with_command)

        result = asyncio.run(monitor_router.platform_login_browser("dy", {"account_id": account["id"]}))
        refreshed = get_login_session(int(session["id"]))

        assert closed == [int(session["id"])]
        assert refreshed["status"] == "timeout"
        assert "登录窗口" in refreshed["message"]
        assert result["pid"] == 45678
        assert seen["profile_path"] == profile_path
        assert result["message"].startswith("已切换到登录窗口")
        assert result["profile_path"] == "网页登录态已配置"
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_account_login_session_does_not_inherit_default_platform_success(monkeypatch, tmp_path):
    init_db()
    snapshots = {
        "login_sessions": _snapshot_table("login_sessions"),
        "social_accounts": _snapshot_table("social_accounts"),
    }
    try:
        account = save_social_account(
            {
                "name": "海安律所抖音采集号",
                "platform": "dy",
                "login_type": "qrcode",
                "status": "standby",
                "profile_path": str(tmp_path / "account_profile"),
            }
        )
        session = create_login_session(
            {
                "platform": "dy",
                "account_id": account["id"],
                "login_url": "https://www.douyin.com",
                "profile_path": account["profile_path"],
                "message": "TargetClosedError: 浏览器会话被关闭或 Profile 正被占用，请稍后重试。",
            }
        )
        session = monitor_router.update_login_session_status(
            int(session["id"]),
            "waiting_manual_browser",
            "TargetClosedError: 浏览器会话被关闭或 Profile 正被占用，请稍后重试。",
        )

        async def fake_poll_qrcode_login_session(session_id):
            return {"active": False, "success": False, "message": "二维码浏览器会话不在运行，请重新生成二维码或打开登录窗口。"}

        monkeypatch.setattr(monitor_router, "poll_qrcode_login_session", fake_poll_qrcode_login_session)
        monkeypatch.setattr(
            monitor_router,
            "list_platform_status",
            lambda: [
                {
                    "platform": "dy",
                    "platform_label": "抖音",
                    "profile_path": str(tmp_path / "default_profile"),
                    "active_account_id": None,
                    "login_ready": True,
                    "login_window_open": False,
                }
            ],
        )

        polled = asyncio.run(monitor_router.login_session(int(session["id"])))
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)

    assert polled["session"]["status"] == "qrcode_failed"
    assert "TargetClosedError" in polled["session"]["message"]
    assert polled["platform_status"]["login_ready"] is True


def test_terminal_login_session_lookup_does_not_downgrade_checked_account(monkeypatch, tmp_path):
    init_db()
    snapshots = {
        "login_sessions": _snapshot_table("login_sessions"),
        "social_accounts": _snapshot_table("social_accounts"),
    }
    try:
        account = save_social_account(
            {
                "name": "海安律所小红书采集号",
                "platform": "xhs",
                "login_type": "qrcode",
                "status": "active",
                "profile_path": str(tmp_path / "xhs_profile"),
            }
        )
        update_social_account_check_state(int(account["id"]), True, "登录态有效")
        session = create_login_session(
            {
                "platform": "xhs",
                "account_id": account["id"],
                "login_url": "https://www.xiaohongshu.com",
                "profile_path": account["profile_path"],
                "message": "二维码已过期，请重新生成。",
            }
        )
        monitor_router.update_login_session_status(int(session["id"]), "expired", "二维码已过期，请重新生成。")

        async def fake_poll_qrcode_login_session(session_id):
            return {"active": False, "success": False, "expired": True, "message": "二维码已过期，请重新生成。"}

        monkeypatch.setattr(monitor_router, "poll_qrcode_login_session", fake_poll_qrcode_login_session)
        monkeypatch.setattr(monitor_router, "list_platform_status", lambda: [])

        polled = asyncio.run(monitor_router.login_session(int(session["id"])))
        refreshed = get_social_account(int(account["id"]))
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)

    assert polled["session"]["status"] == "timeout"
    assert refreshed["status"] == "active"
    assert refreshed["last_error"] == ""
    assert refreshed["last_checked_at"]


def test_default_login_session_does_not_turn_manual_failure_into_success(monkeypatch, tmp_path):
    init_db()
    snapshot = _snapshot_table("login_sessions")
    try:
        profile_path = str(tmp_path / "default_profile")
        session = create_login_session(
            {
                "platform": "dy",
                "login_url": "https://www.douyin.com",
                "profile_path": profile_path,
                "message": "TargetClosedError: 浏览器会话被关闭或 Profile 正被占用，请稍后重试。",
            }
        )
        monitor_router.update_login_session_status(
            int(session["id"]),
            "waiting_manual_browser",
            "TargetClosedError: 浏览器会话被关闭或 Profile 正被占用，请稍后重试。",
        )

        async def fake_poll_qrcode_login_session(session_id):
            return {"active": False, "success": False, "message": "二维码浏览器会话不在运行，请重新生成二维码或打开登录窗口。"}

        monkeypatch.setattr(monitor_router, "poll_qrcode_login_session", fake_poll_qrcode_login_session)
        monkeypatch.setattr(
            monitor_router,
            "list_platform_status",
            lambda: [
                {
                    "platform": "dy",
                    "platform_label": "抖音",
                    "profile_path": profile_path,
                    "active_account_id": None,
                    "login_ready": True,
                    "login_window_open": False,
                }
            ],
        )

        polled = asyncio.run(monitor_router.login_session(int(session["id"])))
    finally:
        _restore_table("login_sessions", snapshot)

    assert polled["session"]["status"] == "qrcode_failed"
    assert "TargetClosedError" in polled["session"]["message"]
    assert polled["platform_status"]["login_ready"] is True


def test_waiting_qrcode_session_does_not_inherit_platform_success(monkeypatch, tmp_path):
    init_db()
    snapshot = _snapshot_table("login_sessions")
    try:
        profile_path = str(tmp_path / "default_profile")
        session = create_login_session(
            {
                "platform": "xhs",
                "login_url": "https://www.xiaohongshu.com",
                "profile_path": profile_path,
                "qr_image": "data:image/png;base64,qr",
                "message": "请扫码登录",
            }
        )
        monitor_router.update_login_session_status(int(session["id"]), "waiting_qrcode", "二维码已生成，请扫码登录。", "data:image/png;base64,qr")

        async def fake_poll_qrcode_login_session(session_id):
            return {"active": True, "success": False, "qr_image": "data:image/png;base64,qr", "message": "二维码已生成，请扫码登录。"}

        monkeypatch.setattr(monitor_router, "poll_qrcode_login_session", fake_poll_qrcode_login_session)
        monkeypatch.setattr(
            monitor_router,
            "list_platform_status",
            lambda: [
                {
                    "platform": "xhs",
                    "platform_label": "小红书",
                    "profile_path": profile_path,
                    "active_account_id": None,
                    "login_ready": True,
                    "login_window_open": False,
                }
            ],
        )

        polled = asyncio.run(monitor_router.login_session(int(session["id"])))
    finally:
        _restore_table("login_sessions", snapshot)

    assert polled["session"]["status"] == "waiting_scan"
    assert polled["session"]["qr_image"].startswith("data:image")
    assert polled["platform_status"]["login_ready"] is True


def test_qrcode_lookup_falls_back_to_visible_page_candidate(monkeypatch):
    async def no_adapter_qrcode(login_adapter):
        return ""

    async def no_selector_qrcode(page, selector):
        return ""

    async def candidate_qrcode(page, platform):
        return "ZmFrZS1xcmNvZGU="

    monkeypatch.setattr(login_qrcode_module, "_find_qrcode_with_mediacrawler_adapter", no_adapter_qrcode)
    monkeypatch.setattr(login_qrcode_module, "_find_qrcode_with_mediacrawler_util", no_selector_qrcode)
    monkeypatch.setattr(login_qrcode_module, "_find_visible_qrcode_candidate_screenshot", candidate_qrcode)

    image = asyncio.run(login_qrcode_module._find_login_qrcode(object(), "xhs", 1000, object()))

    assert image == "ZmFrZS1xcmNvZGU="


def test_login_session_route_maps_manual_verification_then_qrcode(monkeypatch):
    init_db()
    snapshots = {"login_sessions": _snapshot_table("login_sessions"), "social_accounts": _snapshot_table("social_accounts")}
    try:
        account = _login_test_account("ks")
        monkeypatch.setattr(
            monitor_router,
            "build_login_browser_command",
            lambda platform: {
                "platform": platform,
                "platform_label": "快手",
                "login_url": "https://www.kuaishou.com/?isHome=1",
                "profile_path": "browser_data/cdp_ks_user_data_dir",
                "debug_port": 9324,
                "browser_path": "chrome",
            },
        )

        async def fake_start_qrcode_login_session_with_profile(session_id, platform, command):
            return {
                "ok": True,
                "needs_verification": True,
                "verification_type": "slider",
                "verification_label": "滑块验证",
                "verification_detail": "请拖动滑块完成拼图",
                "qr_image": "",
                "verification_image": "data:image/png;base64,ks-verification",
                "message": "平台要求先完成滑块验证，当前不会自动处理验证码。",
                "profile_path": command["profile_path"],
            }

        async def fake_poll_qrcode_login_session(session_id):
            return {
                "active": True,
                "success": False,
                "qr_image": "data:image/png;base64,ks-qr",
                "message": "二维码已生成，请扫码登录。",
            }

        monkeypatch.setattr(monitor_router, "start_qrcode_login_session_with_profile", fake_start_qrcode_login_session_with_profile)
        monkeypatch.setattr(monitor_router, "poll_qrcode_login_session", fake_poll_qrcode_login_session)
        monkeypatch.setattr(
            monitor_router,
            "list_platform_status",
            lambda: [
                {
                    "platform": "ks",
                    "platform_label": "快手",
                    "profile_path": "browser_data/cdp_ks_user_data_dir",
                    "login_ready": False,
                    "login_window_open": False,
                }
            ],
        )

        created = asyncio.run(monitor_router.create_platform_login_session({"platform": "ks", "account_id": account["id"]}))
        polled = asyncio.run(monitor_router.login_session(int(created["session"]["id"])))

        assert created["session"]["status"] == "needs_verification"
        assert created["capabilities"]["qr_image_supported"] is False
        assert created["capabilities"]["verification_image_supported"] is False
        assert created["capabilities"]["verification_image"] == ""
        assert created["capabilities"]["verification_type"] == "slider"
        assert created["capabilities"]["verification_label"] == "滑块验证"
        assert "拖动滑块" in created["capabilities"]["verification_detail"]
        assert "滑块" in created["session"]["message"]
        assert polled["session"]["status"] == "waiting_scan"
        assert polled["session"]["qr_image"].startswith("data:image")
        assert polled["capabilities"]["qr_image_supported"] is True
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_mediacrawler_login_capability_contract_is_explicit():
    expected_classes = {
        "dy": "media_platform.douyin.login.DouYinLogin",
        "ks": "media_platform.kuaishou.login.KuaishouLogin",
        "xhs": "media_platform.xhs.login.XiaoHongShuLogin",
    }
    for platform, class_path in expected_classes.items():
        capability = get_mediacrawler_login_capability(platform)

        assert capability["source"] == "MediaCrawler"
        assert capability["boundary"] == "media_crawler_only"
        assert capability["captcha_policy"] == "report_only"
        assert capability["login_engine"] == "MediaCrawler platform login class"
        assert capability["login_class"] == class_path
        assert capability["bridge_role"] == "capture_qrcode_and_forward_status_only"
        assert capability["qrcode_capture_method"] == "tools.utils.find_login_qrcode"
        assert capability["qrcode_prepare_method"].endswith(".prepare_qrcode_login")
        assert capability["qrcode_flow_steps"]
        assert "不实现独立平台登录爬虫" in capability["unsupported_behaviors"]


def test_login_session_response_exposes_mediacrawler_contract(monkeypatch):
    init_db()
    snapshots = {"login_sessions": _snapshot_table("login_sessions"), "social_accounts": _snapshot_table("social_accounts")}
    try:
        account = _login_test_account("xhs")
        monkeypatch.setattr(
            monitor_router,
            "build_login_browser_command",
            lambda platform: {
                "platform": platform,
                "platform_label": "小红书",
                "login_url": "https://www.xiaohongshu.com",
                "profile_path": "browser_data/cdp_xhs_user_data_dir",
                "debug_port": 9325,
                "browser_path": "chrome",
            },
        )

        async def fake_start_qrcode_login_session_with_profile(session_id, platform, command):
            return {
                "ok": True,
                "qr_image": "data:image/png;base64,xhs",
                "message": "请扫码登录",
                "profile_path": command["profile_path"],
            }

        monkeypatch.setattr(monitor_router, "start_qrcode_login_session_with_profile", fake_start_qrcode_login_session_with_profile)

        created = asyncio.run(monitor_router.create_platform_login_session({"platform": "xhs", "account_id": account["id"]}))
        caps = created["capabilities"]

        assert caps["login_capability_source"] == "平台采集服务"
        assert caps["login_boundary"] == "复用平台采集服务登录能力"
        assert "验证码" in caps["captcha_policy"]
        assert caps["login_class"] == ""
        assert caps["qrcode_capture_method"] == "页面二维码回传"
        assert caps["qrcode_prepare_method"] == "平台登录会话"
        assert caps["qrcode_flow_steps"]
        assert "不自动处理滑块、图形验证码或短信验证码" in caps["unsupported_behaviors"]
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_login_qrcode_bridge_uses_mediacrawler_login_adapter(monkeypatch):
    calls: list[str] = []

    class FakeLoginAdapter:
        def __init__(self, login_type, browser_context, context_page):
            calls.append(f"init:{login_type}")

        async def prepare_qrcode_login(self, timeout_ms=10000):
            calls.append(f"prepare:{timeout_ms}")

        async def capture_qrcode(self):
            calls.append("capture")
            return "data:image/png;base64,abc"

    class FakePage:
        async def wait_for_selector(self, *args, **kwargs):
            raise RuntimeError("selector unavailable")

    monkeypatch.setitem(login_qrcode_module.MEDIACRAWLER_LOGIN_CLASSES, "dy", FakeLoginAdapter)
    adapter = login_qrcode_module._build_mediacrawler_login_adapter("dy", object(), FakePage())

    asyncio.run(login_qrcode_module._prepare_login_page("dy", FakePage(), 1234, adapter))
    image = asyncio.run(login_qrcode_module._find_login_qrcode(FakePage(), "dy", 1234, adapter))

    assert image == "data:image/png;base64,abc"
    assert calls == ["init:qrcode", "prepare:1234", "capture"]


def test_login_session_route_keeps_manual_verification_status_when_window_is_open(monkeypatch):
    init_db()
    snapshot = _snapshot_table("login_sessions")
    try:
        session = create_login_session(
            {
                "platform": "ks",
                "login_url": "https://www.kuaishou.com/?isHome=1",
                "profile_path": "browser_data/cdp_ks_user_data_dir",
                "message": "等待完成人工验证",
            }
        )
        session = monitor_router.update_login_session_status(
            int(session["id"]),
            "waiting_verification",
            "等待完成人工验证",
        )

        async def fake_poll_qrcode_login_session(session_id):
            return {
                "active": True,
                "success": False,
                "needs_verification": True,
                "verification_type": "sms",
                "verification_label": "短信验证码",
                "verification_detail": "请输入验证码",
                "message": "平台要求先完成短信验证码，当前不会自动处理验证码。",
            }

        monkeypatch.setattr(monitor_router, "poll_qrcode_login_session", fake_poll_qrcode_login_session)
        monkeypatch.setattr(
            monitor_router,
            "list_platform_status",
            lambda: [
                {
                    "platform": "ks",
                    "platform_label": "快手",
                    "profile_path": session["profile_path"],
                    "login_ready": True,
                    "login_window_open": True,
                }
            ],
        )

        polled = asyncio.run(monitor_router.login_session(int(session["id"])))

        assert polled["session"]["status"] == "needs_verification"
        assert "短信验证码" in polled["session"]["message"]
        assert polled["capabilities"]["verification_type"] == "sms"
        assert polled["capabilities"]["verification_label"] == "短信验证码"
    finally:
        _restore_table("login_sessions", snapshot)


def test_qrcode_data_url_is_preserved():
    raw = "data:image/png;base64,abc123"

    assert login_qrcode_module._as_data_url(raw) == raw


def test_mediacrawler_qrcode_util_fetches_remote_image_with_browser_request():
    class FakeResponse:
        ok = True
        headers = {"content-type": "image/jpeg"}

        async def body(self):
            return b"fake-image"

    class FakeRequest:
        async def get(self, url, headers=None):
            assert url == "https://example.com/qrcode.jpg"
            return FakeResponse()

    class FakeContext:
        request = FakeRequest()

    class FakeElement:
        async def get_attribute(self, attr):
            if attr == "src":
                return "https://example.com/qrcode.jpg"
            return ""

        async def evaluate(self, script):
            return "https://example.com/qrcode.jpg"

    class FakePage:
        context = FakeContext()

        async def wait_for_selector(self, selector):
            assert selector == "img.qrcode"
            return FakeElement()

    result = asyncio.run(login_qrcode_module.utils.find_login_qrcode(FakePage(), "img.qrcode"))

    assert result == "ZmFrZS1pbWFnZQ=="


def test_mediacrawler_qrcode_util_uses_element_screenshot_fallback():
    class FakeElement:
        async def get_attribute(self, attr):
            return ""

        async def evaluate(self, script):
            return ""

        async def screenshot(self):
            return b"png-bytes"

    class FakePage:
        async def wait_for_selector(self, selector):
            return FakeElement()

    result = asyncio.run(login_qrcode_module.utils.find_login_qrcode(FakePage(), "img.qrcode"))

    assert result == "cG5nLWJ5dGVz"


def test_qrcode_bridge_uses_mediacrawler_selectors():
    assert login_qrcode_module.MEDIACRAWLER_LOGIN_FLOWS["ks"]["login_button_selector"] == "xpath=//p[text()='登录']"
    assert login_qrcode_module.MEDIACRAWLER_LOGIN_FLOWS["ks"]["qrcode_selector"] == "xpath=//div[@class='qrcode-img']//img"
    assert login_qrcode_module.MEDIACRAWLER_LOGIN_FLOWS["xhs"]["login_button_selector"] == "xpath=//*[@id='app']/div[1]/div[2]/div[1]/ul/div[1]/button"
    assert "qrcode" in login_qrcode_module.MEDIACRAWLER_LOGIN_FLOWS["xhs"]["qrcode_selector"]
    assert login_qrcode_module.MEDIACRAWLER_LOGIN_FLOWS["ks"]["login_state"]["cookie_rules"] == {"passToken": None}
    assert login_qrcode_module.MEDIACRAWLER_LOGIN_FLOWS["xhs"]["login_state"]["session_cookie"] == "web_session"
    source = Path(login_qrcode_module.__file__).read_text(encoding="utf-8")
    assert "button:has-text('登录')" not in source
    assert "xpath=//li[contains(@class,'user-info-item')]" not in source
    assert "xpath=//div[contains(@class,'user')]" not in source
    assert "kwai-captcha" not in source
    assert "geetest" not in source
    assert "请按住滑块" not in source
    assert "验证码已发送" not in source
    assert "_image_from_selector" not in source
    assert "_fetch_image_data_url" not in source
    assert "AutomationControlled" not in source
    assert "navigator, 'webdriver'" not in source


def test_login_capabilities_are_sourced_from_mediacrawler():
    dy = get_mediacrawler_login_capability("dy")
    ks = get_mediacrawler_login_capability("ks")
    xhs = get_mediacrawler_login_capability("xhs")

    assert dy["source"] == "MediaCrawler"
    assert dy["boundary"] == "media_crawler_only"
    assert dy["captcha_policy"] == "report_only"
    assert dy["qrcode_selector"] == "xpath=//div[@id='animate_qrcode_container']//img"
    assert dy["login_state"]["cookie_rules"] == {"LOGIN_STATUS": "1"}
    assert dy["login_state"]["local_storage_rules"] == {"HasUserLogin": "1"}
    assert "phone" in dy["mediacrawler_supported_login_types"]
    assert "phone" not in dy["supported_login_types"]
    assert ks["login_button_selector"] == "xpath=//p[text()='登录']"
    assert ks["login_state"]["cookie_rules"] == {"passToken": None}
    assert "phone" not in ks["supported_login_types"]
    assert ks["manual_verification"]["labels"]["slider"] == "滑块验证"
    assert "请拖动滑块完成拼图" in ks["manual_verification"]["text_markers"]["slider"]
    assert "[class*='kwai-captcha']" in ks["manual_verification"]["selectors"]["slider"]
    assert xhs["login_button_selector"] == "xpath=//*[@id='app']/div[1]/div[2]/div[1]/ul/div[1]/button"
    assert xhs["login_state"]["session_cookie"] == "web_session"
    assert ".geetest_panel" in xhs["manual_verification"]["selectors"]["slider"]


def test_qrcode_finder_prefers_mediacrawler_util(monkeypatch):
    seen: dict[str, str] = {}

    async def fake_find_login_qrcode(page, selector):
        seen["selector"] = selector
        return "data:image/png;base64,abc"

    monkeypatch.setattr(login_qrcode_module.utils, "find_login_qrcode", fake_find_login_qrcode)

    result = asyncio.run(login_qrcode_module._find_login_qrcode(object(), "xhs", 3000))

    assert "qrcode" in seen["selector"]
    assert result == "data:image/png;base64,abc"


def test_qrcode_start_prefers_qrcode_before_manual_verification(monkeypatch, tmp_path):
    events: list[str] = []

    class FakePage:
        def __init__(self):
            self.context = None

        def set_default_timeout(self, timeout):
            events.append("timeout")

        async def goto(self, *args, **kwargs):
            events.append("goto")

    class FakeContext:
        def __init__(self):
            self.pages = [FakePage()]
            self.pages[0].context = self

        async def new_page(self):
            page = FakePage()
            page.context = self
            self.pages.append(page)
            return page

        async def close(self):
            events.append("close")

        async def cookies(self):
            return []

    class FakeChromium:
        async def launch_persistent_context(self, **kwargs):
            return FakeContext()

    class FakePlaywright:
        chromium = FakeChromium()

        async def stop(self):
            events.append("stop")

    class FakePlaywrightFactory:
        async def start(self):
            return FakePlaywright()

    async def fake_prepare_login_page(platform, page, timeout, login_adapter=None):
        events.append("prepare")

    async def fake_detect_manual_verification(platform, page):
        events.append("verify")
        return {"needs_verification": True, "verification_type": "slider", "verification_label": "滑块验证", "verification_detail": "请拖动滑块"}

    async def fake_find_login_qrcode(page, platform, timeout, login_adapter=None):
        events.append("find_qr")
        return "data:image/png;base64,should-not-happen"

    monkeypatch.setattr(login_qrcode_module, "async_playwright", lambda: FakePlaywrightFactory())
    monkeypatch.setattr(login_qrcode_module, "_build_mediacrawler_login_adapter", lambda platform, context, page: object())
    monkeypatch.setattr(login_qrcode_module, "_prepare_login_page", fake_prepare_login_page)
    monkeypatch.setattr(login_qrcode_module, "_detect_manual_verification", fake_detect_manual_verification)
    monkeypatch.setattr(login_qrcode_module, "_find_login_qrcode", fake_find_login_qrcode)

    result = asyncio.run(
        login_qrcode_module.start_qrcode_login_session_with_profile(
            888001,
            "ks",
            {
                "profile_path": str(tmp_path / "ks_profile"),
                "browser_path": "chrome",
            },
        )
    )

    try:
        assert result["ok"] is True
        assert result["qr_image"] == "data:image/png;base64,should-not-happen"
        assert "needs_verification" not in result
        assert "verify" not in events
        assert "find_qr" in events
    finally:
        asyncio.run(login_qrcode_module.close_qrcode_login_session(888001))


def test_qrcode_start_has_outer_timeout_and_closes_half_initialized_browser(monkeypatch, tmp_path):
    events: list[str] = []

    class FakePage:
        def __init__(self):
            self.context = None

        def set_default_timeout(self, timeout):
            events.append(f"timeout:{timeout}")

        async def goto(self, *args, **kwargs):
            events.append("goto")

    class FakeContext:
        def __init__(self):
            self.pages = [FakePage()]
            self.pages[0].context = self

        async def new_page(self):
            page = FakePage()
            page.context = self
            self.pages.append(page)
            return page

        async def close(self):
            events.append("close")

        async def cookies(self):
            return []

    class FakeChromium:
        async def launch_persistent_context(self, **kwargs):
            events.append("launch")
            return FakeContext()

    class FakePlaywright:
        chromium = FakeChromium()

        async def stop(self):
            events.append("stop")

    class FakePlaywrightFactory:
        async def start(self):
            events.append("start")
            return FakePlaywright()

    async def hanging_prepare_login_page(platform, page, timeout, login_adapter=None):
        events.append("prepare")
        await asyncio.Event().wait()

    monkeypatch.setattr(login_qrcode_module, "async_playwright", lambda: FakePlaywrightFactory())
    monkeypatch.setattr(login_qrcode_module, "_build_mediacrawler_login_adapter", lambda platform, context, page: object())
    monkeypatch.setattr(login_qrcode_module, "_prepare_login_page", hanging_prepare_login_page)

    result = asyncio.run(
        asyncio.wait_for(
            login_qrcode_module.start_qrcode_login_session_with_profile(
                888002,
                "dy",
                {
                    "profile_path": str(tmp_path / "dy_profile"),
                    "browser_path": "chrome",
                },
                timeout_ms=5,
            ),
            timeout=0.2,
        )
    )

    assert result["ok"] is False
    assert result["status"] == "qrcode_failed"
    assert "超时" in result["message"]
    assert "close" in events
    assert "stop" in events
    assert 888002 not in login_qrcode_module.ACTIVE_LOGIN_SESSIONS


def test_qrcode_login_defaults_to_server_headless_browser(monkeypatch):
    monkeypatch.delenv("MONITOR_LOGIN_QR_HEADLESS", raising=False)

    assert login_qrcode_module._login_qr_headless() is True


def test_qrcode_manual_verification_detects_slider_text_and_selector():
    class FakeLocator:
        def __init__(self, text: str = "", visible: bool = False):
            self.text = text
            self.visible = visible

        @property
        def first(self):
            return self

        async def inner_text(self, timeout=0):
            return self.text

        async def count(self):
            return 1 if self.visible else 0

        async def is_visible(self, timeout=0):
            return self.visible

    class TextPage:
        def locator(self, selector):
            return FakeLocator("请通过验证，向右拖动滑块", False)

    class SelectorPage:
        def locator(self, selector):
            if selector == "body":
                return FakeLocator("登录", False)
            return FakeLocator("", "captcha" in selector)

    assert asyncio.run(login_qrcode_module._needs_manual_verification("ks", TextPage())) is True
    assert asyncio.run(login_qrcode_module._needs_manual_verification("ks", SelectorPage())) is True


def test_qrcode_manual_verification_detects_challenge_url():
    class FakePage:
        url = "https://www.kuaishou.com/captcha/challenge"

    assert asyncio.run(login_qrcode_module._needs_manual_verification("ks", FakePage())) is True


def test_qrcode_manual_verification_detects_kuaishou_slider_copy():
    class FakeLocator:
        def __init__(self, text: str = ""):
            self.text = text

        @property
        def first(self):
            return self

        async def inner_text(self, timeout=0):
            return self.text

        async def count(self):
            return 0

        async def is_visible(self, timeout=0):
            return False

    class FakePage:
        url = "https://www.kuaishou.com/?isHome=1"

        def locator(self, selector):
            if selector == "body":
                return FakeLocator("请拖动滑块完成拼图")
            return FakeLocator("")

    assert asyncio.run(login_qrcode_module._needs_manual_verification("ks", FakePage())) is True


def test_qrcode_manual_verification_classifies_sms_code():
    class FakeLocator:
        def __init__(self, text: str = ""):
            self.text = text

        @property
        def first(self):
            return self

        async def inner_text(self, timeout=0):
            return self.text

        async def count(self):
            return 0

        async def is_visible(self, timeout=0):
            return False

    class FakePage:
        url = "https://www.douyin.com/"

        def locator(self, selector):
            if selector == "body":
                return FakeLocator("请输入验证码，验证码已发送")
            return FakeLocator("")

    result = asyncio.run(login_qrcode_module._detect_manual_verification("dy", FakePage()))

    assert result["needs_verification"] is True
    assert result["verification_type"] == "sms"
    assert result["verification_label"] == "短信验证码"


def test_login_session_verification_code_route_submits_sms_code(monkeypatch):
    init_db()
    snapshot = _snapshot_table("login_sessions")
    seen: dict[str, Any] = {}
    try:
        session = create_login_session(
            {
                "platform": "dy",
                "login_url": "https://www.douyin.com",
                "status": "needs_verification",
                "message": "平台要求先完成短信验证码。",
            }
        )

        async def fake_submit_qrcode_login_verification_code(session_id, code):
            seen["session_id"] = session_id
            seen["code"] = code
            return {
                "active": True,
                "success": False,
                "status": "waiting_confirm",
                "verification_type": "sms",
                "verification_label": "短信验证码",
                "message": "短信验证码已提交，请等待平台确认登录结果。",
            }

        monkeypatch.setattr(monitor_router, "submit_qrcode_login_verification_code", fake_submit_qrcode_login_verification_code)
        monkeypatch.setattr(monitor_router, "list_platform_status", lambda: [])

        result = asyncio.run(
            monitor_router.submit_login_session_verification_code(int(session["id"]), {"code": "123456"})
        )
    finally:
        _restore_table("login_sessions", snapshot)

    assert seen == {"session_id": int(session["id"]), "code": "123456"}
    assert result["session"]["status"] == "waiting_confirm"
    assert "短信验证码已提交" in result["session"]["message"]
    assert result["capabilities"]["verification_type"] == "sms"


def test_login_session_verification_code_request_route_sends_sms_code(monkeypatch):
    init_db()
    snapshot = _snapshot_table("login_sessions")
    seen: dict[str, Any] = {}
    try:
        session = create_login_session(
            {
                "platform": "dy",
                "login_url": "https://www.douyin.com",
                "status": "needs_verification",
                "message": "平台要求先完成短信验证码。",
            }
        )

        async def fake_request_qrcode_login_verification_code(session_id):
            seen["session_id"] = session_id
            return {
                "active": True,
                "success": False,
                "status": "needs_verification",
                "needs_verification": True,
                "verification_type": "sms",
                "verification_label": "短信验证码",
                "verification_detail": "短信验证码发送请求已提交",
                "message": "短信验证码发送请求已提交，请查收后输入验证码。",
            }

        monkeypatch.setattr(
            monitor_router,
            "request_qrcode_login_verification_code",
            fake_request_qrcode_login_verification_code,
        )
        monkeypatch.setattr(monitor_router, "list_platform_status", lambda: [])

        result = asyncio.run(monitor_router.request_login_session_verification_code(int(session["id"])))
    finally:
        _restore_table("login_sessions", snapshot)

    assert seen == {"session_id": int(session["id"])}
    assert result["session"]["status"] == "needs_verification"
    assert "短信验证码发送请求已提交" in result["session"]["message"]
    assert result["capabilities"]["verification_type"] == "sms"


def test_qrcode_manual_sms_verification_submission_fills_server_page(monkeypatch):
    class FakeInput:
        def __init__(self, page):
            self.page = page

        async def is_visible(self, timeout=0):
            return True

        async def fill(self, value):
            self.page.filled = value

    class FakeButton:
        def __init__(self, page):
            self.page = page

        async def is_visible(self, timeout=0):
            return True

        async def click(self, timeout=0):
            self.page.clicked = True

    class FakeLocator:
        def __init__(self, page, kind: str):
            self.page = page
            self.kind = kind

        async def count(self):
            return 1 if self.kind in {"input", "button"} else 0

        def nth(self, index):
            return FakeInput(self.page) if self.kind == "input" else FakeButton(self.page)

    class FakePage:
        filled = ""
        clicked = False

        def locator(self, selector):
            if selector == "input[placeholder*='验证码']":
                return FakeLocator(self, "input")
            if selector.startswith("xpath=//button"):
                return FakeLocator(self, "button")
            return FakeLocator(self, "")

        async def wait_for_timeout(self, timeout):
            pass

    async def fake_is_logged_in(platform, context, page, baseline):
        return False

    page = FakePage()
    session_id = 901234
    login_qrcode_module.ACTIVE_LOGIN_SESSIONS[session_id] = login_qrcode_module.LoginSessionHandle(
        platform="dy",
        playwright=object(),
        context=object(),
        page=page,
        profile_path="browser_data/test",
        created_at=datetime.now(timezone.utc),
    )
    monkeypatch.setattr(login_qrcode_module, "_is_logged_in", fake_is_logged_in)
    try:
        result = asyncio.run(login_qrcode_module.submit_qrcode_login_verification_code(session_id, "123456"))
    finally:
        login_qrcode_module.ACTIVE_LOGIN_SESSIONS.pop(session_id, None)

    assert page.filled == "123456"
    assert page.clicked is True
    assert result["status"] == "waiting_confirm"
    assert "短信验证码已提交" in result["message"]


def test_qrcode_manual_sms_verification_submission_prefers_second_verify_overlay():
    class FakeInput:
        def __init__(self, page, source: str):
            self.page = page
            self.source = source

        async def is_visible(self, timeout=0):
            return True

        async def fill(self, value):
            self.page.filled.append((self.source, value))

    class FakeButton:
        def __init__(self, page, source: str):
            self.page = page
            self.source = source

        async def is_visible(self, timeout=0):
            return True

        async def click(self, timeout=0):
            if self.source == "page":
                self.page.page_button_clicked = True
                raise RuntimeError("second verify overlay intercepts pointer events")
            self.page.overlay_button_clicked = True

    class FakeLocator:
        def __init__(self, page, kind: str):
            self.page = page
            self.kind = kind

        async def count(self):
            return 1 if self.kind else 0

        def nth(self, index):
            if self.kind == "overlay_input":
                return FakeInput(self.page, "overlay")
            if self.kind == "page_input":
                return FakeInput(self.page, "page")
            if self.kind == "overlay_button":
                return FakeButton(self.page, "overlay")
            return FakeButton(self.page, "page")

    class FakePage:
        def __init__(self):
            self.filled: list[tuple[str, str]] = []
            self.page_button_clicked = False
            self.overlay_button_clicked = False

        def locator(self, selector):
            if selector.startswith("#uc-second-verify") and "input" in selector:
                return FakeLocator(self, "overlay_input")
            if "uc-second-verify" in selector and "button" in selector:
                return FakeLocator(self, "overlay_button")
            if selector == "input[placeholder*='验证码']":
                return FakeLocator(self, "page_input")
            if selector.startswith("xpath=//button"):
                return FakeLocator(self, "page_button")
            return FakeLocator(self, "")

    page = FakePage()
    result = asyncio.run(login_qrcode_module._submit_manual_verification_code(page, "112233"))

    assert result is True
    assert page.filled == [("overlay", "112233")]
    assert page.overlay_button_clicked is True
    assert page.page_button_clicked is False


def test_qrcode_manual_sms_verification_submission_prefers_exact_overlay_verify_text():
    class FakeInput:
        def __init__(self, page):
            self.page = page

        async def is_visible(self, timeout=0):
            return True

        async def fill(self, value):
            self.page.filled = value

    class FakeButton:
        def __init__(self, page, label: str):
            self.page = page
            self.label = label

        async def is_visible(self, timeout=0):
            return True

        async def click(self, timeout=0):
            self.page.clicked_labels.append(self.label)
            if self.label != "验证":
                raise RuntimeError(f"wrong button clicked: {self.label}")

    class FakeLocator:
        def __init__(self, page, kind: str):
            self.page = page
            self.kind = kind

        async def count(self):
            if self.kind == "overlay_input":
                return 1
            if self.kind == "overlay_exact_verify":
                return 1
            if self.kind == "overlay_wide_buttons":
                return 3
            return 0

        def nth(self, index):
            if self.kind == "overlay_input":
                return FakeInput(self.page)
            if self.kind == "overlay_exact_verify":
                return FakeButton(self.page, "验证")
            return FakeButton(self.page, ["获取验证码", "重新获取验证码", "验证"][index])

    class FakePage:
        def __init__(self):
            self.filled = ""
            self.clicked_labels: list[str] = []

        def locator(self, selector):
            if selector.startswith("#uc-second-verify") and "input" in selector:
                return FakeLocator(self, "overlay_input")
            if "@id='uc-second-verify'" in selector and "normalize-space(.)='验证'" in selector and " or contains" not in selector:
                return FakeLocator(self, "overlay_exact_verify")
            if "uc-second-verify" in selector and "contains(., '确认')" in selector:
                return FakeLocator(self, "overlay_wide_buttons")
            return FakeLocator(self, "")

    page = FakePage()
    result = asyncio.run(login_qrcode_module._submit_manual_verification_code(page, "445566"))

    assert result is True
    assert page.filled == "445566"
    assert page.clicked_labels == ["验证"]


def test_qrcode_manual_sms_verification_request_clicks_send_code():
    class FakeButton:
        def __init__(self, page):
            self.page = page

        async def is_visible(self, timeout=0):
            return True

        async def click(self, timeout=0):
            self.page.clicked = True

    class FakeLocator:
        def __init__(self, page, kind: str):
            self.page = page
            self.kind = kind

        async def count(self):
            return 1 if self.kind == "button" else 0

        def nth(self, index):
            return FakeButton(self.page)

    class FakePage:
        def __init__(self):
            self.clicked = False

        def locator(self, selector):
            if "接收短信验证码" in selector or "发送短信验证" in selector:
                return FakeLocator(self, "button")
            return FakeLocator(self, "")

        async def wait_for_timeout(self, timeout):
            pass

    page = FakePage()
    session_id = 901235
    login_qrcode_module.ACTIVE_LOGIN_SESSIONS[session_id] = login_qrcode_module.LoginSessionHandle(
        platform="dy",
        playwright=object(),
        context=object(),
        page=page,
        profile_path="browser_data/test",
        created_at=datetime.now(timezone.utc),
    )
    try:
        result = asyncio.run(login_qrcode_module.request_qrcode_login_verification_code(session_id))
    finally:
        login_qrcode_module.ACTIVE_LOGIN_SESSIONS.pop(session_id, None)

    assert page.clicked is True
    assert result["status"] == "needs_verification"
    assert result["verification_type"] == "sms"
    assert "短信验证码发送请求已提交" in result["message"]


def test_xhs_login_state_requires_session_change_when_login_modal_visible():
    class FakeContext:
        async def cookies(self):
            return [{"name": "web_session", "value": "new-session"}]

    class FakePage:
        async def is_visible(self, selector, timeout=0):
            if selector == "div.login-container, .login-modal, img.qrcode-img":
                return True
            return False

    result = asyncio.run(login_qrcode_module._is_logged_in("xhs", FakeContext(), FakePage(), "old-session"))

    assert result is False


def test_xhs_login_state_succeeds_after_session_change():
    class FakeContext:
        async def cookies(self):
            return [{"name": "web_session", "value": "logged-session"}]

    result = asyncio.run(login_qrcode_module._is_logged_in("xhs", FakeContext(), object(), "guest-session"))

    assert result is True


def test_qrcode_login_state_timeout_falls_back_to_cookie_rules(monkeypatch):
    class FakeContext:
        pages = []

        async def cookies(self):
            return [{"name": "LOGIN_STATUS", "value": "1"}]

    class FakePage:
        async def is_visible(self, selector, timeout=0):
            return False

    async def hanging_login_state(platform, context, page, login_baseline=""):
        await asyncio.Event().wait()

    monkeypatch.setenv("MONITOR_LOGIN_QR_POLL_TIMEOUT_MS", "5")
    monkeypatch.setattr(login_qrcode_module, "call_mediacrawler_check_login_state", hanging_login_state)

    result = asyncio.run(
        asyncio.wait_for(
            login_qrcode_module._is_logged_in("dy", FakeContext(), FakePage(), ""),
            timeout=0.2,
        )
    )

    assert result is True


def test_mediacrawler_login_state_check_uses_platform_method(monkeypatch):
    calls: list[str] = []

    class FakeLogin:
        def __init__(self, login_type, browser_context, context_page):
            calls.append(login_type)

        async def check_login_state(self):
            calls.append("check")
            return True

    monkeypatch.setitem(mediacrawler_login_module.MEDIACRAWLER_LOGIN_CLASSES, "dy", FakeLogin)

    result = asyncio.run(mediacrawler_login_module.call_mediacrawler_check_login_state("dy", object(), object()))

    assert result is True
    assert calls == ["qrcode", "check"]


def test_account_collectable_login_requires_mediacrawler_pong(monkeypatch):
    async def fake_login_state(platform, context, page, login_baseline=""):
        return True

    async def fake_pong(platform, context, page, timeout_ms):
        return {"ok": False, "message": "采集前验活未通过。"}

    monkeypatch.setattr(account_check_module, "call_mediacrawler_check_login_state", fake_login_state)
    monkeypatch.setattr(account_check_module, "_check_mediacrawler_client_pong", fake_pong)

    result = asyncio.run(account_check_module._verify_collectable_login("xhs", object(), object(), 1000, "guest-session"))

    assert result["ok"] is False
    assert result["status"] == "client_check_failed"
    assert "采集前验活" in result["message"]


def test_login_session_route_marks_existing_profile_success(monkeypatch):
    init_db()
    snapshots = {"login_sessions": _snapshot_table("login_sessions"), "social_accounts": _snapshot_table("social_accounts")}
    try:
        account = _login_test_account("dy")
        monkeypatch.setattr(
            monitor_router,
            "build_login_browser_command",
            lambda platform: {
                "platform": platform,
                "platform_label": "抖音",
                "login_url": "https://www.douyin.com/",
                "profile_path": "browser_data/cdp_dy_user_data_dir",
                "debug_port": 9323,
                "browser_path": "chrome",
            },
        )

        async def fake_start_qrcode_login_session_with_profile(session_id, platform, command):
            return {
                "ok": True,
                "already_logged_in": True,
                "qr_image": "",
                "message": "当前 Profile 已经登录，不需要重新扫码。",
                "profile_path": "browser_data/cdp_dy_user_data_dir",
            }

        monkeypatch.setattr(monitor_router, "start_qrcode_login_session_with_profile", fake_start_qrcode_login_session_with_profile)
        async def fake_check_social_account_login(
            account_id,
            timeout_ms=15000,
            allow_draft=False,
            identity_prepared=False,
            actor_id=None,
        ):
            return {
                "ok": True,
                "account": get_social_account(account_id),
            }

        monkeypatch.setattr(monitor_router, "check_social_account_login", fake_check_social_account_login)

        created = asyncio.run(monitor_router.create_platform_login_session({"platform": "dy", "account_id": account["id"]}))

        assert created["session"]["status"] == "success"
        assert created["capabilities"]["manual_browser_fallback"] is True
        assert created["capabilities"]["qr_image_supported"] is False
        assert "通过验活" in created["session"]["message"]
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_login_session_success_requires_account_check(monkeypatch):
    init_db()
    snapshots = {"login_sessions": _snapshot_table("login_sessions"), "social_accounts": _snapshot_table("social_accounts")}
    try:
        account = _login_test_account("xhs")

        async def fake_start_qrcode_login_session_with_profile(session_id, platform, command):
            return {
                "ok": True,
                "already_logged_in": True,
                "qr_image": "",
                "message": "当前 Profile 已经登录，不需要重新扫码。",
                "profile_path": command["profile_path"],
            }

        async def fake_check_social_account_login(
            account_id,
            timeout_ms=15000,
            allow_draft=False,
            identity_prepared=False,
            actor_id=None,
        ):
            return {
                "ok": False,
                "message": "登录态无效或已失效，请重新扫码登录。",
                "account": update_social_account_check_state(account_id, False, "登录态无效或已失效，请重新扫码登录。"),
            }

        monkeypatch.setattr(monitor_router, "start_qrcode_login_session_with_profile", fake_start_qrcode_login_session_with_profile)
        monkeypatch.setattr(monitor_router, "check_social_account_login", fake_check_social_account_login)

        created = asyncio.run(monitor_router.create_platform_login_session({"platform": "xhs", "account_id": account["id"]}))

        assert created["session"]["status"] == "platform_error"
        assert "重新扫码登录" in created["session"]["message"]
        assert created["account_status"]["status"] == "limited"
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_login_session_uses_social_account_profile(monkeypatch, tmp_path):
    init_db()
    snapshots = {
        "login_sessions": _snapshot_table("login_sessions"),
        "social_accounts": _snapshot_table("social_accounts"),
    }
    seen: dict[str, Any] = {}
    try:
        account_profile = tmp_path / "account_profile"
        account = save_social_account(
            {
                "name": "海安律所抖音采集号",
                "platform": "dy",
                "login_type": "qrcode",
                "status": "standby",
                "profile_path": str(account_profile),
            }
        )
        monkeypatch.setattr(
            monitor_router,
            "build_login_browser_command",
            lambda platform: {
                "platform": platform,
                "platform_label": "抖音",
                "login_url": "https://www.douyin.com/",
                "profile_path": str(tmp_path / "global_profile"),
                "debug_port": 9323,
                "browser_path": "chrome",
            },
        )

        async def fake_start_qrcode_login_session_with_profile(session_id, platform, command):
            seen["profile_path"] = command["profile_path"]
            return {
                "ok": True,
                "qr_image": "data:image/png;base64,abc",
                "message": "请扫码登录",
                "profile_path": command["profile_path"],
            }

        monkeypatch.setattr(monitor_router, "start_qrcode_login_session_with_profile", fake_start_qrcode_login_session_with_profile)

        created = asyncio.run(monitor_router.create_platform_login_session({"platform": "dy", "account_id": account["id"]}))

        assert seen["profile_path"] == str(resolve_account_profile_path(f"1/dy/acc_{account['id']}"))
        assert created["session"]["account_id"] == account["id"]
        assert created["session"]["profile_key"] == f"1/dy/acc_{account['id']}"
        assert created["session"]["profile_path"] == "网页登录态已配置"
        assert created["session"]["qr_image"].startswith("data:image")
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_login_session_list_account_id_can_reopen_account_profile(monkeypatch, tmp_path):
    init_db()
    snapshots = {
        "login_sessions": _snapshot_table("login_sessions"),
        "social_accounts": _snapshot_table("social_accounts"),
    }
    seen: dict[str, Any] = {}
    try:
        account_profile = tmp_path / "haian_dy_account_profile"
        account = save_social_account(
            {
                "name": "海安律所抖音采集号",
                "platform": "dy",
                "login_type": "qrcode",
                "status": "standby",
                "profile_path": str(account_profile),
            }
        )
        session = create_login_session(
            {
                "platform": "dy",
                "account_id": account["id"],
                "login_url": "https://www.douyin.com/",
                "profile_path": str(account_profile),
                "message": "二维码生成失败，请使用登录窗口兜底",
            }
        )

        monkeypatch.setattr(
            monitor_router,
            "build_login_browser_command",
            lambda platform: {
                "platform": platform,
                "platform_label": "抖音",
                "login_url": "https://www.douyin.com/",
                "profile_path": str(tmp_path / "default_profile"),
                "debug_port": 9323,
                "browser_path": "chrome",
            },
        )

        def fake_open_login_browser_with_command(command):
            seen["profile_path"] = command["profile_path"]
            return {**command, "pid": 23456, "message": "ok"}

        monkeypatch.setattr(monitor_router, "open_login_browser_with_command", fake_open_login_browser_with_command)

        listed = list_login_sessions(limit=1)[0]
        result = asyncio.run(monitor_router.platform_login_browser(listed["platform"], {"account_id": listed["account_id"]}))

        assert listed["id"] == session["id"]
        assert listed["account_id"] == account["id"]
        assert result["pid"] == 23456
        assert seen["profile_path"] == str(resolve_account_profile_path(f"1/dy/acc_{account['id']}"))
        assert result["profile_path"] == "网页登录态已配置"
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_phase_5_social_account_profile_key_drives_runtime_path():
    init_db()
    snapshot = _snapshot_table("social_accounts")
    try:
        account = save_social_account(
            {
                "name": "海安律所小红书采集号",
                "platform": "xhs",
                "login_type": "qrcode",
                "status": "standby",
            }
        )
        expected_key = f"1/xhs/acc_{account['id']}"
        expected_path = str(resolve_account_profile_path(expected_key))
        assert account["profile_key"] == expected_key
        assert account["profile_path"] == ""
        assert account["profile_configured"] is True
        assert get_social_account(account["id"], masked=False)["profile_path"] == expected_path
        assert account["login_capability_source"] == "平台采集服务"
        assert account["login_boundary"] == "media_crawler_only"
        assert account["captcha_policy"] == "report_only"
        assert "qrcode" in account["supported_login_types"]

        updated = save_social_account({**account, "name": "新展示名", "profile_path": str(Path("ignored"))}, int(account["id"]))

        assert updated["profile_key"] == expected_key
        assert get_social_account(account["id"], masked=False)["profile_path"] == expected_path
    finally:
        _restore_table("social_accounts", snapshot)


def test_phase_5_profile_key_resolver_rejects_path_traversal(tmp_path):
    assert resolve_account_profile_path("1/dy/acc_123", root=tmp_path) == tmp_path / "1" / "dy" / "acc_123"
    for bad_key in ["../dy/acc_1", "1/dy/../../x", "1/dy/acc_bad", "1/DY/acc_1"]:
        with pytest.raises(ValueError, match="invalid account profile key"):
            resolve_account_profile_path(bad_key, root=tmp_path)


def test_social_account_login_type_must_follow_mediacrawler_capability():
    init_db()
    snapshot = _snapshot_table("social_accounts")
    try:
        with pytest.raises(ValueError, match="暂未开放手机号登录"):
            save_social_account(
                {
                    "name": "海安律所快手采集号",
                    "platform": "ks",
                    "login_type": "phone",
                    "status": "standby",
                }
            )

        with pytest.raises(ValueError, match="暂未开放手机号登录"):
            save_social_account(
                {
                    "name": "海安律所抖音采集号",
                    "platform": "dy",
                    "login_type": "phone",
                    "status": "standby",
                }
            )

        account = save_social_account({"name": "海安律所抖音采集号", "platform": "dy", "login_type": "qrcode", "status": "standby"})
        assert account["login_capability_source"] == "平台采集服务"
        assert account["login_boundary"] == "media_crawler_only"
        assert account["supported_login_types"] == ["qrcode", "cookie"]
    finally:
        _restore_table("social_accounts", snapshot)


def test_qrcode_poll_success_closes_browser_session(monkeypatch):
    class DummyContext:
        closed = False

        async def close(self):
            self.closed = True

    class DummyPlaywright:
        stopped = False

        async def stop(self):
            self.stopped = True

    context = DummyContext()
    playwright = DummyPlaywright()
    handle = login_qrcode_module.LoginSessionHandle(
        platform="dy",
        playwright=playwright,
        context=context,
        page=object(),
        profile_path="browser_data/cdp_dy_user_data_dir",
        created_at=datetime.now(timezone.utc),
    )
    login_qrcode_module.ACTIVE_LOGIN_SESSIONS[99999] = handle

    async def fake_is_logged_in(platform, context, page, login_baseline=""):
        return True

    monkeypatch.setattr(login_qrcode_module, "_is_logged_in", fake_is_logged_in)

    result = asyncio.run(login_qrcode_module.poll_qrcode_login_session(99999))

    assert result["success"] is True
    assert 99999 not in login_qrcode_module.ACTIVE_LOGIN_SESSIONS
    assert context.closed is True
    assert playwright.stopped is True


def test_qrcode_poll_timeout_returns_pending_state(monkeypatch):
    class DummyContext:
        async def close(self):
            pass

    class DummyPlaywright:
        async def stop(self):
            pass

    handle = login_qrcode_module.LoginSessionHandle(
        platform="dy",
        playwright=DummyPlaywright(),
        context=DummyContext(),
        page=object(),
        profile_path="browser_data/cdp_dy_user_data_dir",
        created_at=datetime.now(timezone.utc),
    )
    login_qrcode_module.ACTIVE_LOGIN_SESSIONS[99998] = handle

    async def fake_is_logged_in(platform, context, page, login_baseline=""):
        return False

    async def hanging_find_login_qrcode(page, platform, timeout, login_adapter=None):
        await asyncio.Event().wait()

    monkeypatch.setenv("MONITOR_LOGIN_QR_POLL_TIMEOUT_MS", "5")
    monkeypatch.setattr(login_qrcode_module, "_is_logged_in", fake_is_logged_in)
    monkeypatch.setattr(login_qrcode_module, "_find_login_qrcode", hanging_find_login_qrcode)

    try:
        result = asyncio.run(
            asyncio.wait_for(
                login_qrcode_module.poll_qrcode_login_session(99998),
                timeout=0.2,
            )
        )
    finally:
        asyncio.run(login_qrcode_module.close_qrcode_login_session(99998))

    assert result["active"] is True
    assert result["success"] is False
    assert result["status"] == "waiting_confirm"
    assert "正在等待平台确认" in result["message"]


def test_qrcode_poll_login_state_timeout_still_checks_cookie_success(monkeypatch):
    class DummyContext:
        pages = []

        async def close(self):
            pass

        async def cookies(self):
            return [{"name": "LOGIN_STATUS", "value": "1"}]

    class DummyPage:
        async def is_visible(self, selector, timeout=0):
            return False

    class DummyPlaywright:
        async def stop(self):
            pass

    handle = login_qrcode_module.LoginSessionHandle(
        platform="dy",
        playwright=DummyPlaywright(),
        context=DummyContext(),
        page=DummyPage(),
        profile_path="browser_data/cdp_dy_user_data_dir",
        created_at=datetime.now(timezone.utc),
    )
    login_qrcode_module.ACTIVE_LOGIN_SESSIONS[99997] = handle

    async def hanging_login_state(platform, context, page, login_baseline=""):
        await asyncio.Event().wait()

    monkeypatch.setenv("MONITOR_LOGIN_QR_POLL_TIMEOUT_MS", "5")
    monkeypatch.setattr(login_qrcode_module, "call_mediacrawler_check_login_state", hanging_login_state)

    try:
        result = asyncio.run(
            asyncio.wait_for(
                login_qrcode_module.poll_qrcode_login_session(99997),
                timeout=0.2,
            )
        )
    finally:
        login_qrcode_module.ACTIVE_LOGIN_SESSIONS.pop(99997, None)

    assert result["success"] is True
    assert result["status"] == "success"
    assert 99997 not in login_qrcode_module.ACTIVE_LOGIN_SESSIONS


def test_ai_skip_env_prevents_external_ai_calls(monkeypatch):
    init_db()
    called = False

    async def fake_call_openai(cfg, prompt, payload):
        nonlocal called
        called = True
        raise RuntimeError("AI provider should not be called")

    monkeypatch.setenv("MONITOR_SKIP_AI_API", "true")
    monkeypatch.setattr("api.monitoring.ai._call_openai", fake_call_openai)

    result = asyncio.run(
        ai_module.evaluate_content(
            {"law_firm_name": "海安律所"},
            {"platform": "dy", "title": "海安律所投诉", "description": "退费迟迟没有处理"},
            [],
        )
    )

    assert called is False
    assert result["status"] == "pending_review"
    assert "未启用" in result["reason"]
    with pytest.raises(ValueError, match="未启用"):
        asyncio.run(
            run_ai_config_test(
                {
                    "provider": "openai",
                    "base_url": "https://example.com",
                    "api_key": "sk-test",
                    "model": "test-model",
                    "temperature": 0,
                }
            )
        )


def test_ai_test_route_skip_env_does_not_save_payload(monkeypatch):
    init_db()
    ai_snapshot = _snapshot_singleton_table("ai_configs")
    try:
        save_ai_config(
            {
                "provider": "openai",
                "base_url": "https://saved.example.com",
                "api_key": "sk-saved",
                "model": "saved-model",
                "temperature": 0,
                "prompt": DEFAULT_PROMPT,
            }
        )
        before = get_ai_config()
        monkeypatch.setenv("MONITOR_SKIP_AI_API", "true")

        with pytest.raises(HTTPException) as exc:
            asyncio.run(
                monitor_router.test_ai_config(
                    {
                        "provider": "openai",
                        "base_url": "https://changed.example.com",
                        "api_key": "sk-changed",
                        "model": "changed-model",
                        "temperature": 0,
                        "prompt": "changed",
                    }
                )
            )
        after = get_ai_config()

        assert exc.value.status_code == 400
        assert "未启用" in str(exc.value.detail)
        assert after["base_url"] == before["base_url"]
        assert after["model"] == before["model"]
        assert after["last_test_status"] == before["last_test_status"]
    finally:
        _restore_singleton_table("ai_configs", ai_snapshot)


def test_ai_skip_env_warns_without_blocking_preflight(monkeypatch):
    monkeypatch.setenv("MONITOR_SKIP_AI_API", "true")
    cfg = {
        "provider": "openai",
        "base_url": "https://example.com",
        "api_key": "sk-********test",
        "model": "test-model",
        "last_test_status": "success",
        "last_test_at": "2026-06-12T00:00:00+00:00",
    }
    job = {
        "id": 1,
        "enabled": True,
        "law_firm_name": "海安律所",
        "keywords": ["海安律所避雷"],
        "platforms": ["dy"],
        "recipients": ["target@example.com"],
    }
    monkeypatch.setattr(
        "api.monitoring.preflight.list_platform_status",
        lambda: [
            {
                "platform": "dy",
                "platform_label": "抖音",
                "login_type": "qrcode",
                "profile_exists": True,
                "has_cookies": False,
                "needs_login": False,
                "login_ready": True,
                "login_window_open": False,
            }
        ],
    )
    monkeypatch.setattr("api.monitoring.preflight.get_ai_config", lambda masked=True: cfg)
    monkeypatch.setattr(
        "api.monitoring.preflight.get_email_config",
        lambda masked=True: {"smtp_host": "smtp.example.com", "sender": "sender@example.com", "last_test_status": "success"},
    )

    preflight = build_job_preflight(job, [])

    assert readiness_module._ai_ready(cfg) is False
    assert "未启用" in readiness_module._ai_message(cfg)
    actions = readiness_module._next_actions([{"key": "ai_config", "ok": False}], [], set(), set())
    assert any("未启用" in action for action in actions)
    assert preflight["can_run"] is True
    assert any("未启用" in item for item in preflight["warnings"])


def test_job_preflight_uses_active_ai_profile_before_legacy_config(monkeypatch):
    init_db()
    profile_snapshot = _snapshot_table("ai_key_profiles")
    ai_snapshot = _snapshot_singleton_table("ai_configs")
    try:
        save_ai_config({"provider": "openai", "base_url": "", "api_key": "", "model": ""})
        profile = save_ai_key_profile(
            {
                "name": "海安律所当前 AI 接入",
                "provider": "openai",
                "base_url": "https://ai.example.com",
                "api_key": "sk-profile",
                "model": "profile-model",
                "temperature": 0,
                "prompt": DEFAULT_PROMPT,
                "is_active": True,
            }
        )
        with get_conn() as conn:
            conn.execute(
                "UPDATE ai_key_profiles SET last_test_status='success', last_test_at=?, last_test_error='' WHERE id=?",
                ("2026-06-12T00:00:00+00:00", profile["id"]),
            )
        monkeypatch.setattr(
            "api.monitoring.preflight.list_platform_status",
            lambda: [
                {
                    "platform": "dy",
                    "platform_label": "抖音",
                    "login_type": "qrcode",
                    "profile_exists": True,
                    "has_cookies": False,
                    "needs_login": False,
                    "login_ready": True,
                    "login_window_open": False,
                }
            ],
        )
        monkeypatch.setattr(
            "api.monitoring.preflight.get_email_config",
            lambda masked=True: {"smtp_host": "smtp.example.com", "sender": "sender@example.com", "last_test_status": "success"},
        )
        job = {
            "id": 1,
            "enabled": True,
            "law_firm_name": "海安律所",
            "keywords": ["海安律所避雷"],
            "platforms": ["dy"],
            "recipients": ["target@example.com"],
        }

        preflight = build_job_preflight(job, [])
        ai_check = next(item for item in preflight["checks"] if item["key"] == "ai_config")

        assert ai_check["ok"] is True
        assert "默认 AI 接入最近测试通过" in ai_check["message"]
        assert not any("AI" in warning for warning in preflight["warnings"])
    finally:
        _restore_table("ai_key_profiles", profile_snapshot)
        _restore_singleton_table("ai_configs", ai_snapshot)


def test_ai_email_test_results_are_persisted_for_readiness(monkeypatch):
    init_db()
    ai_snapshot = _snapshot_singleton_table("ai_configs")
    profile_snapshot = _snapshot_table("ai_key_profiles")
    email_snapshot = _snapshot_singleton_table("email_configs")
    settings_snapshot = _snapshot_table("system_settings")

    async def fake_ai_test(payload):
        return {
            "is_related": True,
            "is_negative": True,
            "risk_level": "medium",
            "reason": "测试通过",
            "evidence_quotes": ["测试"],
            "recommended_action": "继续",
        }

    def fake_send_test_email(payload, allow_real_send=None):
        return None

    try:
        _restore_table("ai_key_profiles", [])
        monkeypatch.setattr(monitor_router.ai, "test_ai", fake_ai_test)
        result = asyncio.run(
            monitor_router.test_ai_config(
                {
                    "provider": "openai",
                    "base_url": "https://example.com",
                    "api_key": "sk-test",
                    "model": "test-model",
                    "temperature": 0,
                    "prompt": DEFAULT_PROMPT,
                }
            )
        )
        assert result["config"]["last_test_status"] == "success"
        ai_check = next(check for check in get_readiness_status()["checks"] if check["key"] == "ai_config")
        assert ai_check["ok"] is True

        save_ai_config(
            {
                "provider": "openai",
                "base_url": "https://example.com",
                "api_key": "",
                "model": "test-model",
                "temperature": 0,
                "prompt": DEFAULT_PROMPT,
            }
        )
        assert get_ai_config()["last_test_status"] == "success"
        ai_check = next(check for check in get_readiness_status()["checks"] if check["key"] == "ai_config")
        assert ai_check["ok"] is True

        save_ai_config({"provider": "openai", "base_url": "https://example.com", "api_key": "sk-test", "model": "changed"})
        assert get_ai_config()["last_test_status"] == "untested"
        ai_check = next(check for check in get_readiness_status()["checks"] if check["key"] == "ai_config")
        assert ai_check["ok"] is False

        monkeypatch.setattr(monitor_router, "send_test_email", fake_send_test_email)
        save_runtime_settings({"real_email_delivery": True}, actor_id=1)
        result = asyncio.run(
            monitor_router.test_email(
                {
                    "smtp_host": "smtp.example.com",
                    "smtp_port": 465,
                    "encryption": "ssl",
                    "sender": "sender@example.com",
                    "username": "sender@example.com",
                    "password": "smtp-password",
                    "default_recipients": ["target@example.com"],
                },
                admin={"id": 1, "workspace_id": 1, "role": "administrator"},
            )
        )
        assert result["config"]["last_test_status"] == "success"
        email_check = next(check for check in get_readiness_status()["checks"] if check["key"] == "email_config")
        assert email_check["ok"] is True

        save_email_config(
            {
                "smtp_host": "smtp.example.com",
                "smtp_port": 465,
                "encryption": "ssl",
                "sender": "sender@example.com",
                "username": "sender@example.com",
                "password": "",
                "default_recipients": ["target@example.com"],
            }
        )
        assert get_email_config()["last_test_status"] == "success"
        email_check = next(check for check in get_readiness_status()["checks"] if check["key"] == "email_config")
        assert email_check["ok"] is True

        save_email_config({"smtp_host": "smtp.example.com", "sender": "changed@example.com", "default_recipients": ["target@example.com"]})
        assert get_email_config()["last_test_status"] == "untested"
        email_check = next(check for check in get_readiness_status()["checks"] if check["key"] == "email_config")
        assert email_check["ok"] is False
    finally:
        _restore_table("system_settings", settings_snapshot)
        _restore_table("ai_key_profiles", profile_snapshot)
        _restore_singleton_table("ai_configs", ai_snapshot)
        _restore_singleton_table("email_configs", email_snapshot)


def test_ai_test_can_reuse_saved_config_when_payload_is_empty(monkeypatch):
    init_db()
    ai_snapshot = _snapshot_singleton_table("ai_configs")

    async def fake_ai_test(payload):
        assert payload == {}
        return {
            "is_related": True,
            "is_negative": True,
            "risk_level": "medium",
            "reason": "测试通过",
            "evidence_quotes": ["测试"],
            "recommended_action": "继续",
        }

    try:
        save_ai_config(
            {
                "provider": "openai",
                "base_url": "https://example.com",
                "api_key": "sk-test",
                "model": "test-model",
                "temperature": 0,
                "prompt": DEFAULT_PROMPT,
            }
        )
        monkeypatch.setattr(monitor_router.ai, "test_ai", fake_ai_test)

        result = asyncio.run(monitor_router.test_ai_config({}))

        assert result["config"]["last_test_status"] == "success"
        assert result["config"]["base_url"] == "https://example.com"
        assert result["config"]["model"] == "test-model"
    finally:
        _restore_singleton_table("ai_configs", ai_snapshot)


def test_ai_config_api_exposes_default_prompt():
    init_db()
    result = asyncio.run(monitor_router.ai_config())

    assert result["default_prompt"] == DEFAULT_PROMPT
    assert "负面" in result["default_prompt"]
    assert result["prompt_sections"]["role"]
    assert any(item["field"] == "risk_level" for item in result["output_schema"])
    assert any(item["field"] == "recommended_action" for item in result["output_schema"])


def test_ai_evaluation_config_test_preserves_editable_sample_context(monkeypatch):
    seen: dict[str, Any] = {}

    async def fake_ai_test(payload):
        seen.update(payload)
        return {
            "is_related": True,
            "is_negative": True,
            "risk_level": "high",
            "reason": payload["sample_law_firm_name"],
            "evidence_quotes": [payload["sample_title"]],
            "recommended_action": "人工复核",
        }

    monkeypatch.setattr(monitor_router.ai, "ai_api_disabled", lambda: False)
    monkeypatch.setattr(monitor_router.ai, "test_ai", fake_ai_test)

    result = asyncio.run(
        monitor_router.test_ai_evaluation_config(
            {
                "prompt": "只按当前样例判断",
                "sample_law_firm_name": "平安",
                "sample_platform": "dy",
                "sample_source_keyword": "平安律师避雷",
                "sample_title": "平安律师避雷：退费拖了很久",
                "sample_text": "我想曝光一下。",
                "sample_comments": "扫码后仍然没有确认\n评论区有人补充投诉",
            }
        )
    )

    assert result["result"]["reason"] == "平安"
    assert seen["sample_law_firm_name"] == "平安"
    assert seen["sample_title"] == "平安律师避雷：退费拖了很久"
    assert seen["sample_comments"] == "扫码后仍然没有确认\n评论区有人补充投诉"


def test_ai_rule_profiles_can_be_managed_and_selected():
    init_db()
    ai_snapshot = _snapshot_singleton_table("ai_configs")
    rule_snapshot = _snapshot_table("ai_rule_profiles")

    try:
        rule_a = save_ai_rule_profile({"name": "海安默认规则", "prompt": "规则 A", "is_active": True})
        rule_b = save_ai_rule_profile({"name": "投诉高敏规则", "prompt": "规则 B"})

        listed = list_ai_rule_profiles()
        assert any(item["id"] == rule_a["id"] for item in listed)
        assert any(item["id"] == rule_b["id"] for item in listed)
        assert get_ai_config()["prompt"] == "规则 A"

        active = set_active_ai_rule_profile(rule_b["id"])

        assert active["is_active"] is True
        assert get_ai_config()["prompt"] == "规则 B"
        assert next(item for item in list_ai_rule_profiles() if item["id"] == rule_a["id"])["is_active"] is False
    finally:
        _restore_table("ai_rule_profiles", rule_snapshot)
        _restore_singleton_table("ai_configs", ai_snapshot)


def test_ai_rule_profile_routes_expose_profiles_and_test_status(monkeypatch):
    init_db()
    ai_snapshot = _snapshot_singleton_table("ai_configs")
    rule_snapshot = _snapshot_table("ai_rule_profiles")

    async def fake_ai_test(payload):
        assert payload["prompt"] == "规则测试 Prompt"
        return {
            "is_related": True,
            "is_negative": True,
            "risk_level": "medium",
            "reason": "测试通过",
            "evidence_quotes": ["测试"],
            "recommended_action": "继续复核",
        }

    try:
        monkeypatch.setattr(monitor_router.ai, "test_ai", fake_ai_test)
        created = asyncio.run(
            monitor_router.create_ai_rule_profile(
                {"name": "规则测试", "prompt": "规则测试 Prompt", "is_active": True}
            )
        )["profile"]

        listed = asyncio.run(monitor_router.ai_rule_profiles())
        assert any(item["id"] == created["id"] for item in listed["profiles"])
        assert listed["output_schema"]

        tested = asyncio.run(monitor_router.test_ai_rule_profile(created["id"], {}))

        assert tested["result"]["risk_level"] == "medium"
        refreshed = next(item for item in list_ai_rule_profiles() if item["id"] == created["id"])
        assert refreshed["last_test_status"] == "success"
    finally:
        _restore_table("ai_rule_profiles", rule_snapshot)
        _restore_singleton_table("ai_configs", ai_snapshot)


def test_failed_ai_test_is_recorded_after_saving_valid_config(monkeypatch):
    init_db()
    ai_snapshot = _snapshot_singleton_table("ai_configs")

    async def failing_ai_test(payload):
        raise RuntimeError("provider rejected request")

    try:
        monkeypatch.setattr(monitor_router.ai, "test_ai", failing_ai_test)
        with pytest.raises(HTTPException):
            asyncio.run(
                monitor_router.test_ai_config(
                    {
                        "provider": "openai",
                        "base_url": "https://example.com",
                        "api_key": "sk-test",
                        "model": "test-model",
                        "temperature": 0,
                    }
                )
            )
        cfg = get_ai_config()
        assert cfg["last_test_status"] == "failed"
        assert "RuntimeError" in cfg["last_test_error"]
    finally:
        _restore_singleton_table("ai_configs", ai_snapshot)


def test_ingest_dedupes_and_report_keeps_pending_review(monkeypatch):
    asyncio.run(_dedupe_and_report_check(monkeypatch))


def test_dedupe_is_isolated_per_monitor_job():
    init_db()
    base = {
        "aliases": [],
        "exclude_words": [],
        "keywords": ["同ID测试律所避雷"],
        "platforms": ["dy"],
        "recipients": [],
        "enable_comments": False,
        "time_window_type": "recent_1d",
        "frequency": "daily",
        "email_time": "09:00",
        "enabled": True,
    }
    job_a = save_job({**base, "law_firm_name": "同ID测试律所A"})
    job_b = save_job({**base, "law_firm_name": "同ID测试律所B"})
    now_ts = int(datetime.now(timezone.utc).timestamp())
    item = {
        "aweme_id": "pytest_shared_content_001",
        "title": "同ID测试律所避雷",
        "desc": "服务争议",
        "create_time": now_ts,
    }

    run_a1 = create_run(job_a["id"])
    first_a = ingest_outputs(job_a, run_a1, "dy", [item], [])
    run_a2 = create_run(job_a["id"])
    second_a = ingest_outputs(job_a, run_a2, "dy", [item], [])
    run_b1 = create_run(job_b["id"])
    first_b = ingest_outputs(job_b, run_b1, "dy", [item], [])

    _cleanup_test_records(job_a["id"], "pytest_shared_content_001")
    _cleanup_test_records(job_b["id"], "pytest_shared_content_001")

    assert first_a["new_contents"] == 1
    assert second_a["new_contents"] == 0
    assert first_b["new_contents"] == 1


def test_exclude_words_filter_before_insert():
    init_db()
    job = save_job(
        {
            "law_firm_name": "排除测试律所",
            "aliases": [],
            "exclude_words": ["招聘"],
            "keywords": ["排除测试律所避雷"],
            "platforms": ["dy"],
            "recipients": [],
            "enable_comments": False,
            "time_window_type": "recent_1d",
            "frequency": "daily",
            "email_time": "09:00",
            "enabled": False,
        }
    )
    now_ts = int(datetime.now(timezone.utc).timestamp())
    run_id = create_run(job["id"])
    result = ingest_outputs(
        job,
        run_id,
        "dy",
        [
            {
                "aweme_id": "pytest_exclude_keep_001",
                "title": "排除测试律所避雷",
                "desc": "服务争议",
                "create_time": now_ts,
            },
            {
                "aweme_id": "pytest_exclude_drop_001",
                "title": "排除测试律所招聘",
                "desc": "招聘信息",
                "create_time": now_ts,
            },
        ],
        [],
    )

    _cleanup_test_records(job["id"], "pytest_exclude_keep_001")
    _cleanup_test_records(job["id"], "pytest_exclude_drop_001")

    assert result["raw_contents"] == 2
    assert result["filtered_contents"] == 1
    assert result["excluded_contents"] == 1
    assert result["new_contents"] == 1


def test_unrelated_negative_is_not_reported_as_risk(monkeypatch):
    asyncio.run(_unrelated_negative_check(monkeypatch))


def test_report_includes_platform_status_and_failure_reason():
    init_db()
    job = save_job(
        {
            "law_firm_name": "报告失败测试律所",
            "aliases": [],
            "exclude_words": [],
            "keywords": ["报告失败测试律所避雷"],
            "platforms": ["dy", "ks"],
            "recipients": [],
            "enable_comments": False,
            "time_window_type": "recent_1d",
            "frequency": "daily",
            "email_time": "09:00",
            "enabled": False,
        }
    )
    run_id = create_run(job["id"])
    report = create_report(
        run_id,
        job,
        {
            "platforms": ["dy", "ks"],
            "failed_platforms": ["ks"],
            "platform_results": {
                "dy": {
                    "status": "success",
                    "raw_contents": 2,
                    "new_contents": 1,
                    "proxy": {"proxy_id": 8, "proxy_name": "华东采集代理", "provider": "manual"},
                },
                "ks": {"status": "failed", "error": "检测到登录态失效"},
            },
            "new_contents": 1,
            "negative_count": 0,
            "high_count": 0,
        },
    )
    html = Path(report["html_path"]).read_text(encoding="utf-8")
    markdown = Path(report["markdown_path"]).read_text(encoding="utf-8")
    _cleanup_test_records(job["id"], "")

    assert "平台采集状态" in html
    assert "华东采集代理 / manual #8" in html
    assert "快手" in html
    assert "检测到登录态失效" in html
    assert "平台采集状态" in markdown
    assert "代理：华东采集代理 / manual #8" in markdown
    assert "快手：失败" in markdown


def test_cr041_report_outputs_redact_signed_media_urls_and_run_view_paths():
    init_db()
    snapshots = {
        "reports": _snapshot_table("reports"),
        "crawl_runs": _snapshot_table("crawl_runs"),
        "raw_contents": _snapshot_table("raw_contents"),
        "raw_comments": _snapshot_table("raw_comments"),
        "ai_evaluations": _snapshot_table("ai_evaluations"),
    }
    jobs_snapshot = _snapshot_monitor_jobs()
    try:
        _clear_monitor_jobs()
        with get_conn() as conn:
            for table in ["reports", "crawl_runs", "raw_contents", "raw_comments", "ai_evaluations"]:
                conn.execute(f"DELETE FROM {table}")
        job = save_job(
            {
                "law_firm_name": "CR041脱敏律所",
                "aliases": [],
                "exclude_words": [],
                "keywords": ["CR041脱敏律所投诉"],
                "platforms": ["dy"],
                "recipients": [],
                "enable_comments": False,
                "time_window_type": "recent_1d",
                "frequency": "daily",
                "email_time": "09:00",
                "enabled": False,
            }
        )
        run_id = create_run(
            job["id"],
            {
                "job_id": job["id"],
                "run_dir": r"E:\myproject\MediaCrawler-worktrees\cr041-pilot-evidence\data_server_like\runs\job_1\run_1_real",
                "platform_results": {
                    "dy": {
                        "status": "success",
                        "raw_contents": 1,
                        "new_contents": 1,
                        "debug_path": r"C:\Users\Administrator\AppData\Local\profile",
                    }
                },
            },
        )
        ingested = ingest_outputs(
            job,
            run_id,
            "dy",
            [
                {
                    "aweme_id": "pytest_cr041_signed_media",
                    "title": "CR041脱敏律所投诉",
                    "desc": "收费争议需要复核",
                    "aweme_url": "https://www.douyin.com/video/7356421?modal_id=7356421&previous_page=app_code_link",
                    "cover_url": "https://p3-sign.douyinpic.com/tos-cn-i-0813/cover.jpeg?x-expires=1799999999&x-signature=secretSig&lk3s=abc",
                    "create_time": int(datetime.now(timezone.utc).timestamp()),
                }
            ],
            [],
        )
        with get_conn() as conn:
            conn.execute(
                """
                INSERT INTO ai_evaluations (
                    workspace_id, raw_content_id, run_id, status,
                    is_related, is_negative, risk_level, reason,
                    evidence_quotes, recommended_action, raw_response, created_at
                )
                VALUES (?, ?, ?, 'pending_review', 1, 1, 'medium', ?, ?, ?, '', ?)
                """,
                (
                    job.get("workspace_id") or 1,
                    ingested["content_db_ids"][0],
                    run_id,
                    "AI 未配置，进入人工复核",
                    json.dumps(["收费争议需要复核"], ensure_ascii=False),
                    "人工复核",
                    datetime.now(timezone.utc).isoformat(),
                ),
            )
        report = create_report(
            run_id,
            job,
            {
                "platforms": ["dy"],
                "failed_platforms": [],
                "new_contents": ingested["new_contents"],
                "negative_count": 1,
                "high_count": 0,
            },
        )
        html = Path(report["html_path"]).read_text(encoding="utf-8")
        markdown = Path(report["markdown_path"]).read_text(encoding="utf-8")
        workbook = load_workbook(report["excel_path"])
        excel_text = "\n".join(
            str(cell.value or "")
            for row in workbook.active.iter_rows()
            for cell in row
        )
        run_view = monitor_router._customer_view_run(get_run(run_id) or {})
        visible = json.dumps(run_view, ensure_ascii=False)
    finally:
        _restore_monitor_jobs(jobs_snapshot)
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)

    combined_report = "\n".join([html, markdown, excel_text])
    for forbidden in ["x-signature", "x-expires", "lk3s", "secretSig", "modal_id", "previous_page"]:
        assert forbidden not in combined_report
    assert "https://www.douyin.com/video/7356421" in combined_report
    assert "媒体链接已脱敏" in combined_report
    assert "E:\\" not in visible
    assert "C:\\" not in visible
    assert "run_dir" not in visible


def test_leads_api_lists_pending_review_items():
    result = asyncio.run(create_sample_report())
    try:
        leads = list_leads(50)
        api_result = asyncio.run(monitor_router.leads(risk="pending"))["leads"]
        report_result = asyncio.run(monitor_router.reports(risk="pending"))["reports"]
        no_risk_reports = asyncio.run(monitor_router.reports(risk="none"))["reports"]
    finally:
        _cleanup_test_records(result["job"]["id"], f"selftest_negative_{result['run_id']}")
        _cleanup_test_records(result["job"]["id"], f"selftest_excluded_{result['run_id']}")

    assert any(item["content_id"] == f"selftest_negative_{result['run_id']}" for item in leads)
    assert any(item["content_id"] == f"system-check-{result['run_id']}" for item in api_result)
    assert all("selftest" not in item["content_id"] for item in api_result)
    assert all(item["eval_status"] == "pending_review" for item in api_result)
    assert any(item["id"] == result["report"]["id"] for item in report_result)
    assert next(item for item in report_result if item["id"] == result["report"]["id"])["summary"]["pending_review_count"] == 1
    assert all(item["id"] != result["report"]["id"] for item in no_risk_reports)


def test_phase_7_2_missing_ai_evaluation_is_limited_context_not_no_risk():
    init_db()
    snapshots = {
        "reports": _snapshot_table("reports"),
        "crawl_runs": _snapshot_table("crawl_runs"),
        "raw_contents": _snapshot_table("raw_contents"),
        "raw_comments": _snapshot_table("raw_comments"),
        "ai_evaluations": _snapshot_table("ai_evaluations"),
    }
    jobs_snapshot = _snapshot_monitor_jobs()
    content_id = "pytest_phase_7_2_missing_eval"
    try:
        _clear_monitor_jobs()
        with get_conn() as conn:
            for table in ["reports", "crawl_runs", "raw_contents", "raw_comments", "ai_evaluations"]:
                conn.execute(f"DELETE FROM {table}")
        job = save_job(
            {
                "law_firm_name": "Phase72缺评律所",
                "keywords": ["Phase72缺评律所退费"],
                "platforms": ["dy"],
                "recipients": [],
                "enable_comments": False,
                "time_window_type": "recent_1d",
                "enabled": True,
            }
        )
        run_id = create_run(job["id"], {"job_id": job["id"]})
        ingested = ingest_outputs(
            job,
            run_id,
            "dy",
            [{"aweme_id": content_id, "title": "Phase72缺评律所退费", "desc": "AI 未完成", "create_time": int(datetime.now(timezone.utc).timestamp())}],
            [],
        )
        finish_run(run_id, "timeout", {"job_id": job["id"], **ingested}, "任务达到系统运行时间上限")
        report = create_report(run_id, job, {"job_id": job["id"], "law_firm_name": job["law_firm_name"], "platforms": ["dy"], **ingested})

        leads_all = asyncio.run(monitor_router.leads(report_id=report["id"], limit=0))["leads"]
        leads_none = asyncio.run(monitor_router.leads(report_id=report["id"], risk="none", limit=0))["leads"]
        leads_unevaluated = asyncio.run(monitor_router.leads(report_id=report["id"], risk="unevaluated", limit=0))["leads"]
        no_risk_reports = asyncio.run(monitor_router.reports(risk="none", limit=0))["reports"]
        unevaluated_reports = asyncio.run(monitor_router.reports(risk="unevaluated", limit=0))["reports"]
        hydrated_report = get_report(report["id"])
        hydrated_run = get_run(run_id)
    finally:
        _restore_monitor_jobs(jobs_snapshot)
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)

    assert leads_all[0]["content_id"] == content_id
    assert leads_all[0]["evaluation_missing"] is True
    assert leads_all[0]["lead_status"] == "limited_context"
    assert leads_all[0]["eval_status"] == "limited_context"
    assert not leads_none
    assert [item["content_id"] for item in leads_unevaluated] == [content_id]
    assert all(item["id"] != report["id"] for item in no_risk_reports)
    assert any(item["id"] == report["id"] for item in unevaluated_reports)
    assert hydrated_report["summary"]["unevaluated_count"] == 1
    assert hydrated_report["summary"]["no_risk_count"] == 0
    assert hydrated_run["summary"]["unevaluated_count"] == 1


def test_phase_7_2_lead_filters_split_unrelated_no_risk_pending_and_unevaluated():
    init_db()
    snapshots = {
        "reports": _snapshot_table("reports"),
        "crawl_runs": _snapshot_table("crawl_runs"),
        "raw_contents": _snapshot_table("raw_contents"),
        "raw_comments": _snapshot_table("raw_comments"),
        "ai_evaluations": _snapshot_table("ai_evaluations"),
    }
    jobs_snapshot = _snapshot_monitor_jobs()
    content_ids = [
        "pytest_phase_7_2_unrelated",
        "pytest_phase_7_2_no_risk",
        "pytest_phase_7_2_pending",
        "pytest_phase_7_2_unevaluated",
    ]
    try:
        _clear_monitor_jobs()
        with get_conn() as conn:
            for table in ["reports", "crawl_runs", "raw_contents", "raw_comments", "ai_evaluations"]:
                conn.execute(f"DELETE FROM {table}")
        job = save_job(
            {
                "law_firm_name": "Phase72分桶律所",
                "keywords": ["Phase72分桶律所"],
                "platforms": ["dy"],
                "recipients": [],
                "enable_comments": False,
                "time_window_type": "recent_1d",
                "enabled": True,
            }
        )
        run_id = create_run(job["id"], {"job_id": job["id"]})
        now_ts = int(datetime.now(timezone.utc).timestamp())
        ingested = ingest_outputs(
            job,
            run_id,
            "dy",
            [{"aweme_id": content_id, "title": content_id, "create_time": now_ts} for content_id in content_ids],
            [],
        )
        with get_conn() as conn:
            rows = conn.execute(
                "SELECT id, content_id FROM raw_contents WHERE run_id=? ORDER BY id",
                (run_id,),
            ).fetchall()
            now = datetime.now(timezone.utc).isoformat()
            payloads = {
                "pytest_phase_7_2_unrelated": ("ok", 0, 1, "high", "其他机构负面", "不处理"),
                "pytest_phase_7_2_no_risk": ("ok", 1, 0, "low", "相关但未见风险", "无需处理"),
                "pytest_phase_7_2_pending": ("pending_review", 1, 1, "high", "AI 未完成", "人工复核"),
            }
            for row in rows:
                if row["content_id"] not in payloads:
                    continue
                status, is_related, is_negative, risk_level, reason, action = payloads[row["content_id"]]
                conn.execute(
                    """
                    INSERT INTO ai_evaluations (
                        workspace_id, raw_content_id, run_id, status, is_related, is_negative,
                        risk_level, reason, evidence_quotes, recommended_action, raw_response, created_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '', ?)
                    """,
                    (job.get("workspace_id") or 1, row["id"], run_id, status, is_related, is_negative, risk_level, reason, "[]", action, now),
                )
        finish_run(run_id, "success", {"job_id": job["id"], **ingested})
        report = create_report(run_id, job, {"job_id": job["id"], "law_firm_name": job["law_firm_name"], "platforms": ["dy"], **ingested})

        unrelated = asyncio.run(monitor_router.leads(report_id=report["id"], risk="unrelated", limit=0))["leads"]
        no_risk = asyncio.run(monitor_router.leads(report_id=report["id"], risk="none", limit=0))["leads"]
        pending = asyncio.run(monitor_router.leads(report_id=report["id"], risk="pending", limit=0))["leads"]
        unevaluated = asyncio.run(monitor_router.leads(report_id=report["id"], risk="unevaluated", limit=0))["leads"]
        negative = asyncio.run(monitor_router.leads(report_id=report["id"], risk="negative", limit=0))["leads"]
        high = asyncio.run(monitor_router.leads(report_id=report["id"], risk="high", limit=0))["leads"]
        report_view = get_report(report["id"])
    finally:
        _restore_monitor_jobs(jobs_snapshot)
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)

    assert [item["content_id"] for item in unrelated] == ["pytest_phase_7_2_unrelated"]
    assert unrelated[0]["lead_status"] == "unrelated"
    assert [item["content_id"] for item in no_risk] == ["pytest_phase_7_2_no_risk"]
    assert no_risk[0]["lead_status"] == "no_risk"
    assert [item["content_id"] for item in pending] == ["pytest_phase_7_2_pending"]
    assert pending[0]["lead_status"] == "pending_review"
    assert [item["content_id"] for item in unevaluated] == ["pytest_phase_7_2_unevaluated"]
    assert unevaluated[0]["lead_status"] == "limited_context"
    assert negative == []
    assert high == []
    assert report_view["summary"]["negative_count"] == 0
    assert report_view["summary"]["high_count"] == 0
    assert report_view["summary"]["unrelated_count"] == 1
    assert report_view["summary"]["no_risk_count"] == 1
    assert report_view["summary"]["pending_review_count"] == 1
    assert report_view["summary"]["unevaluated_count"] == 1


def test_phase_7_2_cr096_calibration_fixtures_preserve_valid_ai_output(monkeypatch):
    init_db()
    snapshots = {
        "reports": _snapshot_table("reports"),
        "crawl_runs": _snapshot_table("crawl_runs"),
        "raw_contents": _snapshot_table("raw_contents"),
        "raw_comments": _snapshot_table("raw_comments"),
        "ai_evaluations": _snapshot_table("ai_evaluations"),
        "ai_key_profiles": _snapshot_table("ai_key_profiles"),
    }
    jobs_snapshot = _snapshot_monitor_jobs()

    async def noisy_positive_model(cfg, prompt, payload):
        if payload["content_url"].endswith("comment-only"):
            quote = payload["comments"][0]
            risk = "high"
        elif "海安律所" in payload["title"]:
            quote = payload["title"]
            risk = "medium"
        else:
            quote = payload["source_keyword"]
            risk = "high"
        return json.dumps(
            {
                "is_related": True,
                "is_negative": True,
                "risk_level": risk,
                "reason": "模拟模型正向输出",
                "evidence_quotes": [quote],
                "recommended_action": "人工复核",
            },
            ensure_ascii=False,
        )

    try:
        _clear_monitor_jobs()
        with get_conn() as conn:
            for table in ["reports", "crawl_runs", "raw_contents", "raw_comments", "ai_evaluations"]:
                conn.execute(f"DELETE FROM {table}")
        save_ai_key_profile(
            {
                "name": "CR045 校准模型",
                "provider": "openai",
                "base_url": "https://ai.example.com",
                "api_key": "sk-profile",
                "model": "profile-model",
                "temperature": 0,
                "prompt": DEFAULT_PROMPT,
                "is_active": True,
            }
        )
        monkeypatch.setattr("api.monitoring.ai._call_openai", noisy_positive_model)
        job = save_job(
            {
                "law_firm_name": "海安律所",
                "aliases": ["海安律师事务所"],
                "keywords": ["北京海安律所退费"],
                "platforms": ["dy"],
                "recipients": [],
                "enable_comments": True,
                "time_window_type": "recent_1d",
                "enabled": True,
            }
        )
        run_id = create_run(job["id"], {"job_id": job["id"]})
        now_ts = int(datetime.now(timezone.utc).timestamp())
        ingested = ingest_outputs(
            job,
            run_id,
            "dy",
            [
                {
                    "aweme_id": "pytest_cr045_keyword_only",
                    "source_keyword": "北京海安律所退费",
                    "title": "教育课程退款避坑记录",
                    "desc": "报名后沟通退费，没有提到任何目标律所。",
                    "aweme_url": "https://example.com/keyword-only",
                    "create_time": now_ts,
                },
                {
                    "aweme_id": "pytest_cr045_title_target",
                    "source_keyword": "北京海安律所退费",
                    "title": "海安律所退费沟通记录",
                    "desc": "收费争议需要复核。",
                    "aweme_url": "https://example.com/title-target",
                    "create_time": now_ts,
                },
                {
                    "aweme_id": "pytest_cr045_comment_only",
                    "source_keyword": "北京海安律所退费",
                    "title": "律师服务退费讨论",
                    "desc": "正文没有点名目标。",
                    "aweme_url": "https://example.com/comment-only",
                    "comment_count": 1,
                    "create_time": now_ts,
                },
                {
                    "aweme_id": "pytest_cr045_geography_only",
                    "source_keyword": "北京海安律所退费",
                    "title": "海安本地培训退费投诉",
                    "desc": "江苏海安一家培训机构退款慢。",
                    "aweme_url": "https://example.com/geography-only",
                    "create_time": now_ts,
                },
            ],
            [
                {
                    "comment_id": "pytest_cr045_comment_only_c1",
                    "aweme_id": "pytest_cr045_comment_only",
                    "content": "补充一下：海安律所一直没处理退款。",
                    "nickname": "评论用户",
                    "create_time": now_ts,
                }
            ],
        )
        eval_summary = asyncio.run(evaluate_new_contents(job, run_id, ingested["content_db_ids"]))
        finish_run(run_id, "success", {"job_id": job["id"], **ingested, **eval_summary})
        report = create_report(run_id, job, {"job_id": job["id"], "platforms": ["dy"], **ingested, **eval_summary})
        high = asyncio.run(monitor_router.leads(report_id=report["id"], risk="high", limit=0))["leads"]
        suspected = asyncio.run(monitor_router.leads(report_id=report["id"], risk="negative", limit=0))["leads"]
        unrelated = asyncio.run(monitor_router.leads(report_id=report["id"], risk="unrelated", limit=0))["leads"]
        report_view = get_report(report["id"])
        with get_conn() as conn:
            rows = {
                row["content_id"]: row
                for row in conn.execute(
                    """
                    SELECT c.content_id, e.is_related, e.is_negative, e.risk_level, e.reason, e.evidence_quotes
                    FROM raw_contents c
                    JOIN ai_evaluations e ON e.raw_content_id=c.id
                    WHERE c.run_id=?
                    """,
                    (run_id,),
                ).fetchall()
            }
    finally:
        _restore_monitor_jobs(jobs_snapshot)
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)

    assert eval_summary["negative_count"] == 4
    assert eval_summary["high_count"] == 3
    assert report_view["summary"]["negative_count"] == 4
    assert report_view["summary"]["suspected_negative_count"] == 1
    assert report_view["summary"]["high_count"] == 3
    assert {item["content_id"] for item in high} == {
        "pytest_cr045_keyword_only",
        "pytest_cr045_comment_only",
        "pytest_cr045_geography_only",
    }
    assert {item["content_id"] for item in suspected} == {"pytest_cr045_title_target"}
    assert unrelated == []
    assert rows["pytest_cr045_keyword_only"]["is_related"] == 1
    assert rows["pytest_cr045_keyword_only"]["is_negative"] == 1
    assert rows["pytest_cr045_keyword_only"]["risk_level"] == "high"
    assert rows["pytest_cr045_geography_only"]["is_related"] == 1
    assert rows["pytest_cr045_geography_only"]["is_negative"] == 1
    assert rows["pytest_cr045_geography_only"]["risk_level"] == "high"
    assert rows["pytest_cr045_title_target"]["risk_level"] == "medium"
    assert "source_keyword" in DEFAULT_PROMPT
    assert "不能单独证明" in DEFAULT_PROMPT
    assert json.loads(rows["pytest_cr045_keyword_only"]["evidence_quotes"]) == ["北京海安律所退费"]


def test_cr050_report_center_risk_filters_do_not_mix_high_and_suspected_negative():
    init_db()
    snapshots = {
        "reports": _snapshot_table("reports"),
        "crawl_runs": _snapshot_table("crawl_runs"),
        "raw_contents": _snapshot_table("raw_contents"),
        "raw_comments": _snapshot_table("raw_comments"),
        "ai_evaluations": _snapshot_table("ai_evaluations"),
    }
    jobs_snapshot = _snapshot_monitor_jobs()

    def create_risk_report(job: dict[str, Any], content_id: str, risk_level: str) -> dict[str, Any]:
        run_id = create_run(job["id"], {"job_id": job["id"]})
        ingested = ingest_outputs(
            job,
            run_id,
            "dy",
            [
                {
                    "aweme_id": content_id,
                    "title": f"{job['law_firm_name']} {content_id}",
                    "desc": "服务争议",
                    "create_time": int(datetime.now(timezone.utc).timestamp()),
                }
            ],
            [],
        )
        with get_conn() as conn:
            conn.execute(
                """
                INSERT INTO ai_evaluations (
                    workspace_id, raw_content_id, run_id, status, is_related, is_negative,
                    risk_level, reason, evidence_quotes, recommended_action, raw_response, created_at
                )
                VALUES (?, ?, ?, 'ok', 1, 1, ?, ?, ?, '人工复核', '{}', ?)
                """,
                (
                    job.get("workspace_id") or 1,
                    ingested["content_db_ids"][0],
                    run_id,
                    risk_level,
                    f"{risk_level} 风险线索",
                    json.dumps(["服务争议"], ensure_ascii=False),
                    datetime.now(timezone.utc).isoformat(),
                ),
            )
        finish_run(run_id, "success", {"job_id": job["id"], **ingested})
        return create_report(
            run_id,
            job,
            {
                "job_id": job["id"],
                "law_firm_name": job["law_firm_name"],
                "platforms": ["dy"],
                "failed_platforms": [],
                **ingested,
            },
        )

    try:
        _clear_monitor_jobs()
        with get_conn() as conn:
            for table in ["reports", "crawl_runs", "raw_contents", "raw_comments", "ai_evaluations"]:
                conn.execute(f"DELETE FROM {table}")
        job = save_job(
            {
                "law_firm_name": "CR050筛选律所",
                "keywords": ["CR050筛选律所投诉"],
                "platforms": ["dy"],
                "recipients": [],
                "enable_comments": False,
                "time_window_type": "recent_1d",
                "enabled": True,
            }
        )
        high_report = create_risk_report(job, "pytest_cr050_high", "high")
        suspected_report = create_risk_report(job, "pytest_cr050_suspected", "medium")

        high_leads = asyncio.run(monitor_router.leads(report_id=high_report["id"], risk="high", limit=0))["leads"]
        high_as_negative = asyncio.run(monitor_router.leads(report_id=high_report["id"], risk="negative", limit=0))["leads"]
        suspected_leads = asyncio.run(monitor_router.leads(report_id=suspected_report["id"], risk="negative", limit=0))["leads"]
        suspected_as_high = asyncio.run(monitor_router.leads(report_id=suspected_report["id"], risk="high", limit=0))["leads"]
        high_reports = asyncio.run(monitor_router.reports(risk="high", limit=0))["reports"]
        negative_reports = asyncio.run(monitor_router.reports(risk="negative", limit=0))["reports"]
        hydrated_high = get_report(high_report["id"])
        hydrated_suspected = get_report(suspected_report["id"])
    finally:
        _restore_monitor_jobs(jobs_snapshot)
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)

    assert [item["content_id"] for item in high_leads] == ["pytest_cr050_high"]
    assert high_leads[0]["lead_status"] == "high_risk"
    assert high_as_negative == []
    assert [item["content_id"] for item in suspected_leads] == ["pytest_cr050_suspected"]
    assert suspected_leads[0]["lead_status"] == "suspected_negative"
    assert suspected_as_high == []
    assert high_report["id"] in {item["id"] for item in high_reports}
    assert suspected_report["id"] not in {item["id"] for item in high_reports}
    assert suspected_report["id"] in {item["id"] for item in negative_reports}
    assert high_report["id"] not in {item["id"] for item in negative_reports}
    assert hydrated_high["summary"]["negative_count"] == 1
    assert hydrated_high["summary"]["suspected_negative_count"] == 0
    assert hydrated_high["summary"]["high_count"] == 1
    assert hydrated_suspected["summary"]["negative_count"] == 1
    assert hydrated_suspected["summary"]["suspected_negative_count"] == 1
    assert hydrated_suspected["summary"]["high_count"] == 0


def test_phase_7_2_timeout_finalization_creates_pending_review_fallback_rows(monkeypatch):
    init_db()
    snapshots = {
        "reports": _snapshot_table("reports"),
        "crawl_runs": _snapshot_table("crawl_runs"),
        "raw_contents": _snapshot_table("raw_contents"),
        "raw_comments": _snapshot_table("raw_comments"),
        "ai_evaluations": _snapshot_table("ai_evaluations"),
        "email_delivery_logs": _snapshot_table("email_delivery_logs"),
    }
    jobs_snapshot = _snapshot_monitor_jobs()
    content_ids = ["pytest_phase_7_2_timeout_a", "pytest_phase_7_2_timeout_b"]
    try:
        _clear_monitor_jobs()
        with get_conn() as conn:
            for table in ["reports", "crawl_runs", "raw_contents", "raw_comments", "ai_evaluations", "email_delivery_logs"]:
                conn.execute(f"DELETE FROM {table}")
        job = save_job(
            {
                "law_firm_name": "Phase72超时律所",
                "keywords": ["Phase72超时律所投诉"],
                "platforms": ["dy"],
                "recipients": [],
                "enable_comments": False,
                "time_window_type": "recent_1d",
                "enabled": True,
            }
        )

        async def fake_run_platform(job_arg, run_id, platform, run_dir):
            ingested = ingest_outputs(
                job_arg,
                run_id,
                platform,
                [
                    {"aweme_id": content_ids[0], "title": "Phase72超时律所投诉 A", "create_time": int(datetime.now(timezone.utc).timestamp())},
                    {"aweme_id": content_ids[1], "title": "Phase72超时律所投诉 B", "create_time": int(datetime.now(timezone.utc).timestamp())},
                ],
                [],
            )
            return {**ingested, "status": "success"}

        async def fake_evaluate_new_contents(job_arg, run_id, candidate_ids):
            raise runner_module.CrawlerTimedOut("AI 评估达到系统运行时间上限")

        monkeypatch.setattr(runner_module, "run_platform", fake_run_platform)
        monkeypatch.setattr(runner_module, "evaluate_new_contents", fake_evaluate_new_contents)
        monkeypatch.setattr(
            runner_module,
            "send_report_with_delivery_log",
            lambda job, report, send_type="auto": (False, "未配置收件人", report, None),
        )

        first = asyncio.run(run_monitor_job(job["id"]))
        summary = first["summary"]
        with get_conn() as conn:
            rows = conn.execute(
                "SELECT raw_content_id, status, reason FROM ai_evaluations WHERE run_id=? ORDER BY raw_content_id",
                (first["run_id"],),
            ).fetchall()
        second_created = runner_module._mark_unresolved_candidates_pending_review(
            first["run_id"],
            [int(row["raw_content_id"]) for row in rows],
        )
    finally:
        _restore_monitor_jobs(jobs_snapshot)
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)

    assert first["status"] == "timeout"
    assert len(rows) == 2
    assert {row["status"] for row in rows} == {"pending_review"}
    assert summary["pending_review_count"] == 2
    assert summary["ai_unresolved_items"] == 0
    assert summary["ai_finalization_fallback"]["known_unresolved_candidate_ids"] == 2
    assert summary["ai_finalization_fallback"]["pending_review_rows_created"] == 2
    assert summary["ai_finalization_fallback"]["limited_context_rows_left_unchanged"] == 0
    assert first["report"]["summary"]["pending_review_count"] == 2
    assert second_created == 0


def test_phase_7_run_job_generates_report_without_ai_or_email(monkeypatch):
    init_db()
    monkeypatch.setenv("MONITOR_SKIP_AI_API", "true")
    job = save_job(
        {
            "law_firm_name": "海安律所",
            "aliases": ["海安律师事务所"],
            "exclude_words": [],
            "keywords": ["海安律所避雷"],
            "platforms": ["dy"],
            "recipients": [],
            "enable_comments": False,
            "time_window_type": "recent_1d",
            "frequency": "daily",
            "email_time": "09:00",
            "enabled": True,
        }
    )
    content_id = "pytest_phase7_no_ai_email_001"
    now_ts = int(datetime.now(timezone.utc).timestamp())

    async def fake_run_platform(job_arg, run_id, platform, run_dir):
        return ingest_outputs(
            job_arg,
            run_id,
            platform,
            [
                {
                    "aweme_id": content_id,
                    "title": "海安律所退费投诉",
                    "desc": "收费争议需要人工复核",
                    "create_time": now_ts,
                    "share_url": "https://www.douyin.com/video/phase7",
                }
            ],
            [],
        )

    def fake_send_report(job_arg, report):
        return False, "SMTP 配置未完成"

    try:
        monkeypatch.setattr(runner_module, "run_platform", fake_run_platform)
        monkeypatch.setattr("api.monitoring.reporting.send_report", fake_send_report)

        result = asyncio.run(run_monitor_job(job["id"]))
        run = get_run(int(result["run_id"]))
        report = get_report(int(result["report"]["id"]))
        html = Path(result["report"]["html_path"]).read_text(encoding="utf-8")
    finally:
        _cleanup_test_records(job["id"], content_id)

    assert result["status"] == "success"
    assert run and run["status"] == "success"
    assert result["report"]
    assert report and report["email_status"] == "failed"
    assert result["summary"]["new_contents"] == 1
    assert result["summary"]["pending_review_count"] == 1
    assert result["summary"]["email_status"] == "failed"
    assert "SMTP 配置未完成" in result["summary"]["email_error"]
    assert "待人工复核" in html
    assert "AI 结果仅用于舆情线索筛查" in html


def test_selftest_report_generates_downloadable_artifacts():
    asyncio.run(_selftest_report_check())


def test_internal_selftest_jobs_are_hidden_from_operator_job_list():
    init_db()
    job = save_job(
        {
            "law_firm_name": "海安律所",
            "aliases": [],
            "exclude_words": [],
            "keywords": ["海安律所避雷"],
            "platforms": ["dy"],
            "recipients": [],
            "enable_comments": False,
            "time_window_type": "recent_1d",
            "frequency": "daily",
            "email_time": "09:00",
            "enabled": False,
            "is_internal": True,
        }
    )
    visible_jobs = list_jobs()
    all_jobs = list_jobs(include_internal=True)
    _cleanup_test_records(job["id"], "")

    assert all(j["id"] != job["id"] for j in visible_jobs)
    assert any(j["id"] == job["id"] and j["is_internal"] for j in all_jobs)


def test_selftest_jobs_are_hidden_by_run_summary_marker():
    init_db()
    job = save_job(
        {
            "law_firm_name": "海安律所",
            "aliases": [],
            "exclude_words": [],
            "keywords": ["海安律所避雷"],
            "platforms": ["dy"],
            "recipients": [],
            "enable_comments": False,
            "time_window_type": "recent_1d",
            "frequency": "daily",
            "email_time": "09:00",
            "enabled": False,
        }
    )
    run_id = create_run(job["id"])
    try:
        finish_run(run_id, "selftest", {"selftest": True, "law_firm_name": "海安律所"})
        mark_selftest_jobs_internal()
        visible_jobs = list_jobs()
        all_jobs = list_jobs(include_internal=True)
    finally:
        _cleanup_test_records(job["id"], "")

    assert all(j["id"] != job["id"] for j in visible_jobs)
    assert any(j["id"] == job["id"] and j["is_internal"] for j in all_jobs)


def test_readiness_status_reports_checks():
    init_db()
    status = get_readiness_status()
    keys = {check["key"] for check in status["checks"]}

    assert {"platform_profiles", "account_alerts", "proxy_alerts", "ai_config", "email_config", "selftest_report", "real_report"} <= keys
    assert isinstance(status["ready"], bool)
    assert isinstance(status["next_actions"], list)
    assert len(status["platforms"]) == 3
    assert all("label" in check and "ok" in check and "message" in check for check in status["checks"])


def test_readiness_platform_profiles_only_require_douyin_but_preflight_checks_selected_platform(monkeypatch):
    init_db()
    job = {
        "id": 123,
        "enabled": True,
        "keywords": ["海安律所避雷"],
        "platforms": ["ks"],
        "recipients": ["target@example.com"],
    }
    monkeypatch.setattr(
        readiness_module,
        "list_platform_status",
        lambda: [
            {"platform": "dy", "platform_label": "抖音", "profile_exists": True, "needs_login": False},
            {"platform": "ks", "platform_label": "快手", "profile_exists": True, "needs_login": True},
            {"platform": "xhs", "platform_label": "小红书", "profile_exists": True, "needs_login": False},
        ],
    )
    monkeypatch.setattr(
        "api.monitoring.preflight.list_platform_status",
        lambda: [
            {"platform": "dy", "platform_label": "抖音", "profile_exists": True, "needs_login": False},
            {"platform": "ks", "platform_label": "快手", "profile_exists": True, "needs_login": True},
            {"platform": "xhs", "platform_label": "小红书", "profile_exists": True, "needs_login": False},
        ],
    )

    status = get_readiness_status()
    platform_check = next(check for check in status["checks"] if check["key"] == "platform_profiles")
    preflight = build_job_preflight(job, [])

    assert platform_check["ok"] is True
    assert "抖音登录配置可用" in platform_check["message"]
    assert "快手" in platform_check["message"]
    assert any("扩展平台资源" in action and "快手" in action for action in status["next_actions"])
    assert preflight["can_run"] is False
    assert any("重新登录" in blocker and "快手" in blocker for blocker in preflight["blockers"])


def test_readiness_and_preflight_warn_when_login_window_is_still_open(monkeypatch):
    init_db()
    job = {
        "id": 123,
        "enabled": True,
        "keywords": ["测试律所避雷"],
        "platforms": ["dy"],
        "recipients": ["target@example.com"],
    }
    statuses = [
        {"platform": "dy", "platform_label": "抖音", "profile_exists": True, "needs_login": False, "login_window_open": True},
        {"platform": "ks", "platform_label": "快手", "profile_exists": True, "needs_login": False, "login_window_open": False},
        {"platform": "xhs", "platform_label": "小红书", "profile_exists": True, "needs_login": False, "login_window_open": False},
    ]
    monkeypatch.setattr(readiness_module, "list_platform_status", lambda: statuses)
    monkeypatch.setattr("api.monitoring.preflight.list_platform_status", lambda: statuses)

    readiness = get_readiness_status()
    preflight = build_job_preflight(job, [])

    platform_check = next(check for check in readiness["checks"] if check["key"] == "platform_profiles")
    assert platform_check["ok"] is False
    assert "登录窗口未关闭" in platform_check["message"]
    assert any("关闭" in action and "抖音" in action for action in readiness["next_actions"])
    assert preflight["can_run"] is False
    assert any("关闭登录窗口" in blocker for blocker in preflight["blockers"])


def test_readiness_and_preflight_block_missing_web_profile(monkeypatch):
    statuses = [
        {
            "platform": "dy",
            "platform_label": "抖音",
            "login_type": "qrcode",
            "profile_exists": True,
            "needs_login": False,
            "login_ready": True,
            "login_window_open": False,
        },
        {
            "platform": "ks",
            "platform_label": "快手",
            "login_type": "qrcode",
            "profile_exists": True,
            "needs_login": False,
            "login_ready": True,
            "login_window_open": False,
        },
        {
            "platform": "xhs",
            "platform_label": "小红书",
            "login_type": "qrcode",
            "profile_exists": False,
            "login_material_ready": False,
            "needs_login": True,
            "login_ready": False,
            "login_window_open": False,
        },
    ]
    monkeypatch.setattr(readiness_module, "list_platform_status", lambda: statuses)
    monkeypatch.setattr("api.monitoring.preflight.list_platform_status", lambda: statuses)

    readiness = get_readiness_status()
    preflight = build_job_preflight(
        {"id": 1009, "enabled": True, "keywords": ["海安律所投诉"], "platforms": ["xhs"], "recipients": ["target@example.com"]},
        [],
    )
    platform_check = next(check for check in readiness["checks"] if check["key"] == "platform_profiles")

    assert platform_check["ok"] is True
    assert "抖音登录配置可用" in platform_check["message"]
    assert "小红书网页登录态待准备" in platform_check["message"]
    assert any("扩展平台资源" in action and "小红书网页登录态待准备" in action for action in readiness["next_actions"])
    assert preflight["can_run"] is False
    assert any("请先重新登录小红书账号" in blocker for blocker in preflight["blockers"])


def test_job_preflight_blocks_active_account_with_disabled_proxy(monkeypatch):
    init_db()
    snapshots = {
        "proxy_profiles": _snapshot_table("proxy_profiles"),
        "social_accounts": _snapshot_table("social_accounts"),
    }
    statuses = [
        {"platform": "dy", "platform_label": "抖音", "login_type": "qrcode", "profile_exists": True, "needs_login": False, "login_window_open": False},
        {"platform": "ks", "platform_label": "快手", "login_type": "qrcode", "profile_exists": True, "needs_login": False, "login_window_open": False},
        {"platform": "xhs", "platform_label": "小红书", "login_type": "qrcode", "profile_exists": True, "needs_login": False, "login_window_open": False},
    ]
    monkeypatch.setattr("api.monitoring.preflight.list_platform_status", lambda: statuses)
    try:
        proxy = save_proxy_profile(
            {
                "name": "海安律所停用代理",
                "provider": "manual",
                "proxy_url": "http://user:pass@127.0.0.1:8081",
                "status": "disabled",
                "max_concurrency": 1,
            }
        )
        save_social_account(
            {
                "name": "海安律所抖音采集号",
                "platform": "dy",
                "login_type": "qrcode",
                "status": "active",
                "proxy_id": proxy["id"],
            }
        )

        preflight = build_job_preflight(
            {"id": 1010, "enabled": True, "keywords": ["海安律所避雷"], "platforms": ["dy"], "recipients": ["target@example.com"]},
            [],
        )

        assert preflight["can_run"] is False
        assert any("绑定代理已停用" in blocker and "抖音" in blocker for blocker in preflight["blockers"])
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_phase_9_readiness_reports_account_and_proxy_alerts(monkeypatch):
    statuses = [
        {
            "platform": "dy",
            "platform_label": "抖音",
            "login_type": "qrcode",
            "profile_exists": True,
            "needs_login": True,
            "login_window_open": False,
            "last_error": "登录态失效 password=hunter2",
            "active_proxy_error": "代理超时 密码：secret",
        }
    ]
    monkeypatch.setattr(readiness_module, "list_platform_status", lambda: statuses)

    readiness = get_readiness_status()
    account_check = next(check for check in readiness["checks"] if check["key"] == "account_alerts")
    proxy_check = next(check for check in readiness["checks"] if check["key"] == "proxy_alerts")

    assert account_check["ok"] is False
    assert proxy_check["ok"] is False
    visible = json.dumps({"checks": readiness["checks"], "actions": readiness["next_actions"]}, ensure_ascii=False)
    assert "登录态失效" in visible
    assert "hunter2" not in visible
    assert "secret" not in visible
    assert any("平台账号页" in action for action in readiness["next_actions"])
    assert any("代理资源页" in action for action in readiness["next_actions"])


def test_job_preflight_warns_active_account_with_limited_proxy_error(monkeypatch):
    init_db()
    snapshots = {
        "proxy_profiles": _snapshot_table("proxy_profiles"),
        "social_accounts": _snapshot_table("social_accounts"),
    }
    statuses = [
        {"platform": "xhs", "platform_label": "小红书", "login_type": "qrcode", "profile_exists": True, "needs_login": False, "login_window_open": False},
    ]
    monkeypatch.setattr("api.monitoring.preflight.list_platform_status", lambda: statuses)
    try:
        proxy = save_proxy_profile(
            {
                "name": "海安律所受限代理",
                "provider": "manual",
                "proxy_url": "http://user:pass@127.0.0.1:8081",
                "status": "limited",
                "max_concurrency": 1,
                "last_error": "timeout with password=hunter2",
            }
        )
        save_social_account(
            {
                "name": "海安律所小红书采集号",
                "platform": "xhs",
                "login_type": "qrcode",
                "status": "active",
                "proxy_id": proxy["id"],
            }
        )

        preflight = build_job_preflight(
            {"id": 1011, "enabled": True, "keywords": ["海安律所投诉"], "platforms": ["xhs"], "recipients": ["target@example.com"]},
            [],
        )

        assert preflight["can_run"] is True
        assert any("绑定代理状态为受限" in warning for warning in preflight["warnings"])
        assert any("最近有错误" in warning and "hunter2" not in warning for warning in preflight["warnings"])
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_doctor_reports_deployment_diagnostics():
    init_db()
    status = run_doctor()
    keys = {check["key"] for check in status["checks"]}

    assert {
        "project_files",
        "uv",
        "data_dir",
        "database",
        "disk_space",
        "retention_settings",
        "backup_set",
        "gitignore_runtime_data",
        "platform_login",
        "browser_profiles",
        "resource_alerts",
        "ai_config",
        "email_config",
        "reports",
    } <= keys
    assert "readiness" in status
    assert "paths" in status
    assert status["paths"]["monitor_data_dir"]
    assert isinstance(status["recommendations"], list)
    login_check = next(check for check in status["checks"] if check["key"] == "platform_login")
    assert login_check["ok"] is True
    assert "平台采集服务" in login_check["message"]
    capabilities = login_check["capabilities"]
    assert all(item["bridge_role"] == "capture_qrcode_and_forward_status_only" for item in capabilities)
    assert all(str(item["login_class"]).startswith("media_platform.") for item in capabilities)
    assert all(str(item["qrcode_prepare_method"]).endswith(".prepare_qrcode_login") for item in capabilities)
    assert all(item["qrcode_capture_method"] == "tools.utils.find_login_qrcode" for item in capabilities)
    retention_check = next(check for check in status["checks"] if check["key"] == "retention_settings")
    assert "运行日志保留" in retention_check["message"]
    disk_check = next(check for check in status["checks"] if check["key"] == "disk_space")
    assert "GB" in disk_check["message"]
    backup_check = next(check for check in status["checks"] if check["key"] == "backup_set")
    assert "数据库" in backup_check["message"]


def test_phase_9_doctor_resource_alerts_are_customer_safe(monkeypatch):
    statuses = [
        {
            "platform": "dy",
            "platform_label": "抖音",
            "login_type": "qrcode",
            "profile_exists": True,
            "needs_login": True,
            "login_window_open": False,
            "last_error": "cookie=session-secret",
            "active_proxy_error": "proxy=http://user:pass@127.0.0.1:8081",
        }
    ]
    monkeypatch.setattr("api.monitoring.doctor.list_platform_status", lambda: statuses)

    status = run_doctor()
    check = next(item for item in status["checks"] if item["key"] == "resource_alerts")

    assert check["ok"] is False
    visible = json.dumps({"check": check, "tips": status["recommendations"]}, ensure_ascii=False)
    assert "session-secret" not in visible
    assert "user:pass" not in visible
    assert "平台账号和代理资源页" in visible


def test_doctor_checks_gitignore_runtime_data(monkeypatch, tmp_path):
    (tmp_path / ".gitignore").write_text("/browser_data/\n.env\n", encoding="utf-8")
    monkeypatch.setattr("api.monitoring.doctor.PROJECT_ROOT", tmp_path)

    status = run_doctor()
    check = next(item for item in status["checks"] if item["key"] == "gitignore_runtime_data")

    assert check["ok"] is False
    assert "/monitor_data/" in check["message"]
    assert "*.log" in check["message"]


def test_env_examples_cover_monitor_runtime_knobs():
    required = [
        "MONITOR_DATA_DIR",
        "MONITOR_BROWSER_DATA_DIR",
        "MONITOR_CRAWLER_HEADLESS",
        "MONITOR_CDP_CONNECT_EXISTING",
        "MONITOR_LOGIN_QR_HEADLESS",
        "MONITOR_LOGIN_QR_TIMEOUT_MS",
        "MONITOR_LOGIN_QR_TTL_SECONDS",
        "MONITOR_CDP_DEBUG_PORT_DY",
        "MONITOR_CDP_DEBUG_PORT_KS",
        "MONITOR_CDP_DEBUG_PORT_XHS",
        "MONITOR_CRAWLER_TIMEOUT_SECONDS",
        "MONITOR_CRAWLER_MAX_RETRIES",
        "MONITOR_CRAWLER_RETRY_DELAY_SECONDS",
        "MONITOR_JOB_LOCK_TTL_SECONDS",
        "MONITOR_SKIP_AI_API",
        "MONITOR_DISABLE_SCHEDULER",
    ]
    paths = [
        Path(".env.example"),
        Path("deploy/systemd/legal-sentiment-monitor.env.example"),
    ]
    for path in paths:
        text = path.read_text(encoding="utf-8")
        missing = [name for name in required if name not in text]
        assert not missing, f"{path} missing {missing}"


def test_doctor_flags_non_mediacrawler_login_boundary(monkeypatch):
    def fake_capabilities():
        return [
            {
                "platform": "dy",
                "source": "Custom",
                "boundary": "custom",
                "bridge_role": "custom_login",
                "login_class": "custom.DouyinLogin",
                "qrcode_prepare_method": "custom.prepare",
                "qrcode_capture_method": "custom.capture",
                "qrcode_supported": True,
                "login_state": {"cookie_rules": {"LOGIN_STATUS": "1"}},
                "manual_verification": {"text_markers": {"captcha": ["验证"]}},
            },
            {
                "platform": "ks",
                "source": "MediaCrawler",
                "boundary": "media_crawler_only",
                "bridge_role": "capture_qrcode_and_forward_status_only",
                "login_class": "media_platform.kuaishou.login.KuaishouLogin",
                "qrcode_prepare_method": "KuaishouLogin.prepare_qrcode_login",
                "qrcode_capture_method": "tools.utils.find_login_qrcode",
                "qrcode_supported": True,
                "login_state": {"cookie_rules": {"passToken": None}},
                "manual_verification": {"text_markers": {"captcha": ["验证"]}},
            },
            {
                "platform": "xhs",
                "source": "MediaCrawler",
                "boundary": "media_crawler_only",
                "bridge_role": "capture_qrcode_and_forward_status_only",
                "login_class": "media_platform.xhs.login.XiaoHongShuLogin",
                "qrcode_prepare_method": "XiaoHongShuLogin.prepare_qrcode_login",
                "qrcode_capture_method": "tools.utils.find_login_qrcode",
                "qrcode_supported": True,
                "login_state": {"session_cookie": "web_session"},
                "manual_verification": {"text_markers": {"captcha": ["验证"]}},
            },
        ]

    monkeypatch.setattr("api.monitoring.doctor.list_mediacrawler_login_capabilities", fake_capabilities)

    status = run_doctor()
    login_check = next(check for check in status["checks"] if check["key"] == "platform_login")

    assert login_check["ok"] is False
    assert "抖音登录能力来源异常" in login_check["message"]
    assert "抖音边界不是 media_crawler_only" in login_check["message"]
    assert "抖音登录桥接角色不是只回传二维码和状态" in login_check["message"]
    assert "抖音缺少平台登录适配" in login_check["message"]
    assert "抖音缺少二维码准备能力" in login_check["message"]
    assert "抖音二维码获取能力异常" in login_check["message"]


def test_doctor_report_check_uses_full_report_history(monkeypatch):
    reports = [
        {
            "id": 30,
            "summary": {
                "platform_results": {
                    "dy": {"status": "success", "raw_contents": 1},
                    "ks": {"status": "success", "raw_contents": 1},
                    "xhs": {"status": "success", "raw_contents": 1},
                }
            },
        },
        {"id": 1, "summary": {"selftest": True}},
    ]
    monkeypatch.setattr("api.monitoring.doctor.list_reports", lambda limit=100: reports)

    status = run_doctor()
    report_check = next(check for check in status["checks"] if check["key"] == "reports")

    assert report_check["ok"] is True
    assert "系统自检报告和抖音采集报告" in report_check["message"]


def test_doctor_report_check_does_not_accept_partial_real_report(monkeypatch):
    reports = [
        {"id": 30, "summary": {"platform_results": {"dy": {"status": "success", "raw_contents": 1}}}},
        {"id": 1, "summary": {"selftest": True}},
    ]
    monkeypatch.setattr("api.monitoring.doctor.list_reports", lambda limit=100: reports)

    status = run_doctor()
    report_check = next(check for check in status["checks"] if check["key"] == "reports")

    assert report_check["ok"] is True
    assert "抖音采集报告" in report_check["message"]
    assert not any("selftest-report" in tip for tip in status["recommendations"])


def test_doctor_does_not_defer_job_recommendation_for_optional_platform_login(monkeypatch):
    monkeypatch.setattr(
        "api.monitoring.doctor.list_platform_status",
        lambda: [
            {"platform": "dy", "platform_label": "抖音", "profile_exists": True, "needs_login": False, "login_ready": True},
            {"platform": "ks", "platform_label": "快手", "profile_exists": True, "needs_login": True, "login_ready": False},
            {"platform": "xhs", "platform_label": "小红书", "profile_exists": True, "needs_login": False, "login_ready": True},
        ],
    )
    monkeypatch.setattr(
        readiness_module,
        "list_platform_status",
        lambda: [
            {"platform": "dy", "platform_label": "抖音", "profile_exists": True, "needs_login": False, "login_ready": True},
            {"platform": "ks", "platform_label": "快手", "profile_exists": True, "needs_login": True, "login_ready": False},
            {"platform": "xhs", "platform_label": "小红书", "profile_exists": True, "needs_login": False, "login_ready": True},
        ],
    )
    monkeypatch.setattr("api.monitoring.doctor.list_jobs", lambda: [{"id": 1007, "enabled": False}])

    status = run_doctor()

    browser_check = next(check for check in status["checks"] if check["key"] == "browser_profiles")
    assert browser_check["ok"] is True
    assert "扩展平台待维护" in browser_check["message"]
    assert any(tip == "在任务管理页创建并启用至少一个监控任务。" for tip in status["recommendations"])
    assert not any("登录态恢复后" in tip for tip in status["recommendations"])


def test_doctor_reports_ai_skip_mode(monkeypatch):
    init_db()
    monkeypatch.setenv("MONITOR_SKIP_AI_API", "true")

    status = run_doctor()
    ai_check = next(check for check in status["checks"] if check["key"] == "ai_config")

    assert ai_check["ok"] is False
    assert "未启用" in ai_check["message"]
    assert any("未启用" in tip for tip in status["recommendations"])


def test_doctor_lists_optional_login_window_as_maintenance(monkeypatch):
    init_db()
    statuses = [
        {"platform": "dy", "platform_label": "抖音", "profile_exists": True, "needs_login": False, "login_ready": True, "login_window_open": False},
        {"platform": "ks", "platform_label": "快手", "profile_exists": True, "needs_login": False, "login_ready": False, "login_window_open": True},
        {"platform": "xhs", "platform_label": "小红书", "profile_exists": True, "needs_login": False, "login_ready": True, "login_window_open": False},
    ]
    monkeypatch.setattr("api.monitoring.doctor.list_platform_status", lambda: statuses)
    monkeypatch.setattr(readiness_module, "list_platform_status", lambda: statuses)

    status = run_doctor()
    browser_check = next(check for check in status["checks"] if check["key"] == "browser_profiles")

    assert browser_check["ok"] is True
    assert "扩展平台待维护" in browser_check["message"]
    assert "快手登录窗口待关闭" in browser_check["message"]
    assert not any("关闭" in tip and "快手" in tip for tip in status["recommendations"] if not tip.startswith("扩展平台资源"))


def test_doctor_blocks_when_required_login_window_is_still_open(monkeypatch):
    init_db()
    statuses = [
        {"platform": "dy", "platform_label": "抖音", "profile_exists": True, "needs_login": False, "login_ready": False, "login_window_open": True},
        {"platform": "ks", "platform_label": "快手", "profile_exists": True, "needs_login": False, "login_ready": True, "login_window_open": False},
        {"platform": "xhs", "platform_label": "小红书", "profile_exists": True, "needs_login": False, "login_ready": True, "login_window_open": False},
    ]
    monkeypatch.setattr("api.monitoring.doctor.list_platform_status", lambda: statuses)
    monkeypatch.setattr(readiness_module, "list_platform_status", lambda: statuses)

    status = run_doctor()
    browser_check = next(check for check in status["checks"] if check["key"] == "browser_profiles")

    assert browser_check["ok"] is False
    assert "登录窗口未关闭" in browser_check["message"]
    assert "抖音" in browser_check["message"]
    assert any("关闭" in tip and "抖音" in tip for tip in status["recommendations"])


def test_doctor_api_exposes_deployment_diagnostics():
    init_db()
    status = asyncio.run(monitor_router.doctor())

    assert "checks" in status
    assert "readiness" in status
    assert "recommendations" in status
    assert "paths" in status
    visible = json.dumps(status, ensure_ascii=False)
    for forbidden in [
        "MediaCrawler",
        "media_platform.",
        "media_crawler_only",
        "tools.utils",
        "prepare_qrcode_login",
        "MONITOR_SKIP_AI_API",
        "selftest",
        "Profile",
        "离线模式",
        "离线自检",
        "uv 命令",
        "uv.EXE",
        "main.py",
        "docs/deployment_runbook.md",
    ]:
        assert forbidden not in visible


def test_readiness_dashboard_and_checklist_are_customer_safe():
    init_db()
    readiness = asyncio.run(monitor_router.readiness())
    dashboard = asyncio.run(monitor_router.dashboard())
    checklist = asyncio.run(monitor_router.system_checklist())

    assert "latest_system_check_report_id" in readiness
    assert "latest_selftest_report_id" not in readiness
    assert "latest_system_check_report_id" in dashboard["readiness"]
    assert "latest_system_check_report_id" in checklist
    visible = json.dumps({"readiness": readiness, "dashboard": dashboard, "checklist": checklist}, ensure_ascii=False)
    for forbidden in [
        "MediaCrawler",
        "MONITOR_SKIP_AI_API",
        "selftest",
        "Profile",
        "离线模式",
        "离线自检",
        "html_path",
        "markdown_path",
        "excel_path",
        "E:\\",
        "main.py",
        "debug_port",
    ]:
        assert forbidden not in visible


def test_smoke_check_generates_selftest_artifacts_and_summaries():
    result = asyncio.run(run_smoke_check())
    selftest = result["selftest"]
    artifacts = selftest["artifacts"]
    try:
        report = get_report(selftest["report_id"])
    finally:
        _cleanup_test_records(selftest["job_id"], f"selftest_negative_{selftest['run_id']}")
        _cleanup_test_records(selftest["job_id"], f"selftest_excluded_{selftest['run_id']}")

    assert result["ok"] is True
    assert report is not None
    assert artifacts["html"]["exists"] is True
    assert artifacts["excel"]["exists"] is True
    assert artifacts["markdown"]["exists"] is True
    assert artifacts["html"]["download_url"].endswith(f"/download?type=html")
    assert "failed_checks" in result["doctor"]
    assert "next_actions" in result["readiness"]
    assert "不调用真实平台" in result["note"]


def test_smoke_api_returns_local_smoke_result():
    result = asyncio.run(monitor_router.smoke())["result"]
    system_check = result["system_check"]
    try:
        report = get_report(system_check["report_id"])
    finally:
        _cleanup_test_records(report["job_id"], f"selftest_negative_{system_check['run_id']}")
        _cleanup_test_records(report["job_id"], f"selftest_excluded_{system_check['run_id']}")

    assert result["ok"] is True
    assert report is not None
    assert result["system_check"]["artifacts"]["markdown"]["download_url"].endswith("type=markdown")
    visible = json.dumps(result, ensure_ascii=False)
    for forbidden in ["MediaCrawler", "selftest", "本地自测", "smoke", "MONITOR_SKIP_AI_API"]:
        assert forbidden not in visible


def test_system_check_report_api_returns_operator_summary():
    result = asyncio.run(monitor_router.report_system_check())["result"]
    report = get_report(result["report_id"])
    try:
        assert result["ok"] is True
        assert result["artifacts"]["html"]["download_url"].endswith("type=html")
        assert result["artifacts"]["excel"]["exists"] is True
        assert result["artifacts"]["markdown"]["exists"] is True
        visible = json.dumps(result, ensure_ascii=False)
        for forbidden in ["MediaCrawler", "selftest", "本地自测", "html_path", "markdown_path", "excel_path"]:
            assert forbidden not in visible
    finally:
        _cleanup_test_records(report["job_id"], f"selftest_negative_{result['run_id']}")
        _cleanup_test_records(report["job_id"], f"selftest_excluded_{result['run_id']}")


def test_legacy_selftest_report_route_is_customer_safe():
    result = asyncio.run(monitor_router.report_selftest())["result"]
    report = get_report(result["report_id"])
    try:
        assert result["ok"] is True
        assert result["artifacts"]["html"]["download_url"].endswith("type=html")
        visible = json.dumps(result, ensure_ascii=False)
        for forbidden in ["MediaCrawler", "selftest", "本地自测", "html_path", "markdown_path", "excel_path", "E:\\"]:
            assert forbidden not in visible
    finally:
        _cleanup_test_records(report["job_id"], f"selftest_negative_{result['run_id']}")
        _cleanup_test_records(report["job_id"], f"selftest_excluded_{result['run_id']}")


def test_cli_smoke_command_runs_local_smoke(monkeypatch):
    result = asyncio.run(cli_module._run_command(cli_module.build_parser().parse_args(["smoke"])))
    selftest = result["selftest"]
    try:
        report = get_report(selftest["report_id"])
    finally:
        _cleanup_test_records(selftest["job_id"], f"selftest_negative_{selftest['run_id']}")
        _cleanup_test_records(selftest["job_id"], f"selftest_excluded_{selftest['run_id']}")

    assert result["ok"] is True
    assert report is not None
    assert result["selftest"]["artifacts"]["excel"]["exists"] is True


def test_scheduler_is_disabled_for_multi_worker_env(monkeypatch):
    monkeypatch.setenv("WEB_CONCURRENCY", "2")
    monkeypatch.setattr(scheduler_module, "_apscheduler", None)
    monkeypatch.setattr(scheduler_module, "_scheduler_task", None)

    assert "多 worker" in scheduler_disabled_reason()
    asyncio.run(scheduler_module.start_scheduler())
    status = run_doctor()
    scheduler_check = next(check for check in status["checks"] if check["key"] == "scheduler_mode")

    assert scheduler_module._apscheduler is None
    assert scheduler_module._scheduler_task is None
    assert scheduler_check["ok"] is False
    assert "多 worker" in scheduler_check["message"]


def test_scheduler_status_api_exposes_internal_mode(monkeypatch):
    monkeypatch.delenv("MONITOR_DISABLE_SCHEDULER", raising=False)
    monkeypatch.delenv("WEB_CONCURRENCY", raising=False)
    monkeypatch.delenv("UVICORN_WORKERS", raising=False)

    status = scheduler_status()
    api_status = asyncio.run(monitor_router.monitor_scheduler_status())

    assert status["enabled"] is True
    assert status["mode"] == "internal"
    assert "60 秒" in status["message"]
    assert api_status["enabled"] is True


def test_scheduler_tick_skips_template_jobs_and_continues(monkeypatch):
    calls: list[int] = []
    skipped: list[tuple[int, str]] = []
    schedule_updates: list[tuple[int, str | None]] = []
    jobs = [
        {
            "id": 1,
            "enabled": True,
            "law_firm_name": "请改成目标律所名称",
            "keywords": ["目标律所避雷"],
            "platforms": ["dy"],
            "frequency": "daily",
            "email_time": "00:00",
            "last_run_at": None,
        },
        {
            "id": 2,
            "enabled": True,
            "law_firm_name": "海安律所",
            "keywords": ["海安律所避雷"],
            "platforms": ["dy"],
            "frequency": "daily",
            "email_time": "00:00",
            "last_run_at": None,
        },
    ]

    monkeypatch.setattr(scheduler_module, "list_jobs", lambda: jobs)
    monkeypatch.setattr(
        "api.monitoring.preflight.list_platform_status",
        lambda: [
            {
                "platform": "dy",
                "platform_label": "抖音",
                "login_type": "qrcode",
                "profile_exists": True,
                "needs_login": False,
                "login_ready": True,
                "login_window_open": False,
            }
        ],
    )
    monkeypatch.setattr(scheduler_module, "set_job_schedule_state", lambda job_id, value: schedule_updates.append((job_id, value)))
    monkeypatch.setattr(scheduler_module, "launch_job", lambda job_id, source="scheduler": calls.append(job_id))
    monkeypatch.setattr(scheduler_module, "record_skipped_run", lambda job_id, reason, summary=None: skipped.append((job_id, reason)))

    asyncio.run(scheduler_module.tick())

    assert calls == [2]
    assert skipped and skipped[0][0] == 1
    assert [job_id for job_id, _ in schedule_updates] == [1, 2]


def test_scheduler_tick_blocks_preflight_and_records_skipped_run(monkeypatch):
    init_db()
    jobs_snapshot = _snapshot_monitor_jobs()
    runs_snapshot = _snapshot_table("crawl_runs")
    try:
        _clear_monitor_jobs()
        job = save_job(
            {
                "law_firm_name": "海安律所",
                "aliases": [],
                "exclude_words": [],
                "keywords": ["海安律所避雷"],
                "platforms": ["dy"],
                "recipients": ["target@example.com"],
                "enable_comments": False,
                "time_window_type": "recent_1d",
                "frequency": "daily",
                "email_time": "00:00",
                "enabled": True,
            }
        )
        calls: list[int] = []
        monkeypatch.setattr(scheduler_module, "launch_job", lambda job_id, source="scheduler": calls.append(job_id))
        monkeypatch.setattr(
            "api.monitoring.preflight.list_platform_status",
            lambda: [
                {"platform": "dy", "platform_label": "抖音", "profile_exists": True, "needs_login": False, "login_window_open": True},
                {"platform": "ks", "platform_label": "快手", "profile_exists": True, "needs_login": False, "login_window_open": False},
                {"platform": "xhs", "platform_label": "小红书", "profile_exists": True, "needs_login": False, "login_window_open": False},
            ],
        )

        asyncio.run(scheduler_module.tick())
        runs = [run for run in list_runs(0) if run["job_id"] == job["id"]]

        assert calls == []
        assert runs
        assert runs[0]["status"] == "skipped"
        assert "运行前检查未通过" in (runs[0]["error_message"] or "")
        assert "关闭登录窗口" in (runs[0]["error_message"] or "")
        assert runs[0]["summary"]["skip_type"] == "preflight_blocked"
    finally:
        _restore_monitor_jobs(jobs_snapshot)
        _restore_table("crawl_runs", runs_snapshot)


def test_record_skipped_run_deduplicates_recent_same_reason():
    init_db()
    jobs_snapshot = _snapshot_monitor_jobs()
    runs_snapshot = _snapshot_table("crawl_runs")
    try:
        _clear_monitor_jobs()
        job = save_job(
            {
                "law_firm_name": "海安律所",
                "aliases": [],
                "exclude_words": [],
                "keywords": ["海安律所退费"],
                "platforms": ["dy"],
                "recipients": [],
                "enable_comments": False,
                "time_window_type": "recent_1d",
                "frequency": "daily",
                "email_time": "09:00",
                "enabled": True,
            }
        )
        first = record_skipped_run(job["id"], "同一原因", {"law_firm_name": "海安律所"})
        second = record_skipped_run(job["id"], "同一原因", {"law_firm_name": "海安律所"})
        third = record_skipped_run(job["id"], "另一个原因", {"law_firm_name": "海安律所"})

        assert second == first
        assert third != first
    finally:
        _restore_monitor_jobs(jobs_snapshot)
        _restore_table("crawl_runs", runs_snapshot)


def test_skipped_run_has_operator_display_fields():
    init_db()
    jobs_snapshot = _snapshot_monitor_jobs()
    runs_snapshot = _snapshot_table("crawl_runs")
    try:
        _clear_monitor_jobs()
        job = save_job(
            {
                "law_firm_name": "海安律所",
                "aliases": [],
                "exclude_words": [],
                "keywords": ["海安律所投诉"],
                "platforms": ["dy"],
                "recipients": ["target@example.com"],
                "enable_comments": False,
                "time_window_type": "recent_1d",
                "frequency": "daily",
                "email_time": "09:00",
                "enabled": True,
            }
        )
        reason = "运行前检查未通过：请先重新登录再运行采集：抖音"
        run_id = record_skipped_run(
            job["id"],
            reason,
            {
                "law_firm_name": "海安律所",
                "platforms": ["dy"],
                "keywords": ["海安律所投诉"],
                "skip_type": "preflight_blocked",
            },
        )
        run = get_run(run_id)
        dashboard = get_dashboard_summary()

        assert run["status"] == "skipped"
        assert run["display_status"] == "预检拦截"
        assert run["display_error"] == reason
        assert run["display_law_firm_name"] == "海安律所"
        assert dashboard["skipped_runs_recent"] >= 1
    finally:
        _restore_monitor_jobs(jobs_snapshot)
        _restore_table("crawl_runs", runs_snapshot)


def test_job_preflight_warns_but_allows_missing_ai_email(monkeypatch):
    init_db()
    job = save_job(
        {
            "law_firm_name": "预检测试律所",
            "aliases": [],
            "exclude_words": [],
            "keywords": ["预检测试律所避雷"],
            "platforms": ["dy"],
            "recipients": [],
            "enable_comments": False,
            "time_window_type": "recent_1d",
            "frequency": "daily",
            "email_time": "09:00",
            "enabled": False,
        }
    )
    ai_snapshot = _snapshot_singleton_table("ai_configs")
    profile_snapshot = _snapshot_table("ai_key_profiles")
    email_snapshot = _snapshot_singleton_table("email_configs")

    try:
        _restore_table("ai_key_profiles", [])
        save_ai_config({"provider": "openai", "base_url": "", "api_key": "", "model": ""})
        save_email_config({"smtp_host": "", "sender": "", "default_recipients": []})
        monkeypatch.setattr(
            "api.monitoring.preflight.list_platform_status",
            lambda: [
                {"platform": "dy", "platform_label": "抖音", "profile_exists": True, "needs_login": False},
                {"platform": "ks", "platform_label": "快手", "profile_exists": True, "needs_login": False},
                {"platform": "xhs", "platform_label": "小红书", "profile_exists": True, "needs_login": False},
            ],
        )

        preflight = build_job_preflight(job, [])
        api_result = asyncio.run(monitor_router.job_preflight(job["id"]))["preflight"]
    finally:
        _restore_table("ai_key_profiles", profile_snapshot)
        _restore_singleton_table("ai_configs", ai_snapshot)
        _restore_singleton_table("email_configs", email_snapshot)
        _cleanup_test_records(job["id"], "")

    assert preflight["can_run"] is True
    assert preflight["ready"] is False
    assert any("AI" in item for item in preflight["warnings"])
    assert any("收件人" in item for item in preflight["warnings"])
    assert api_result["can_run"] is True


def test_job_preflight_blocks_missing_platform_search_terms_not_missing_ai_email(monkeypatch):
    init_db()
    ai_snapshot = _snapshot_singleton_table("ai_configs")
    profile_snapshot = _snapshot_table("ai_key_profiles")
    email_snapshot = _snapshot_singleton_table("email_configs")
    job = {
        "id": 1008,
        "law_firm_name": "海安律所",
        "aliases": ["海安律师事务所"],
        "exclude_words": ["招聘"],
        "keywords": [],
        "platforms": ["dy"],
        "recipients": [],
        "enabled": True,
        "target_type": "search",
        "output_mode": "internal",
    }
    monkeypatch.setattr(
        "api.monitoring.preflight.list_platform_status",
        lambda: [{"platform": "dy", "platform_label": "抖音", "profile_exists": True, "needs_login": False, "login_window_open": False}],
    )

    try:
        _restore_table("ai_key_profiles", [])
        save_ai_config({"provider": "openai", "base_url": "", "api_key": "", "model": ""})
        save_email_config({"smtp_host": "", "sender": "", "default_recipients": []})
        preflight = build_job_preflight(job, [])
    finally:
        _restore_table("ai_key_profiles", profile_snapshot)
        _restore_singleton_table("ai_configs", ai_snapshot)
        _restore_singleton_table("email_configs", email_snapshot)

    assert preflight["can_run"] is False
    assert any("未配置平台搜索词" in blocker for blocker in preflight["blockers"])
    assert any("AI" in warning for warning in preflight["warnings"])
    assert any("收件人" in warning for warning in preflight["warnings"])


def test_job_preflight_uses_bound_ai_profile_and_email_template(monkeypatch):
    init_db()
    snapshots = {
        "monitor_jobs": _snapshot_monitor_jobs(),
        "ai_key_profiles": _snapshot_table("ai_key_profiles"),
        "email_templates": _snapshot_table("email_templates"),
        "email_configs": _snapshot_singleton_table("email_configs"),
    }
    try:
        _clear_monitor_jobs()
        save_email_config(
            {
                "smtp_host": "smtp.example.com",
                "smtp_port": 465,
                "sender": "sender@example.com",
                "default_recipients": [],
            }
        )
        profile = save_ai_key_profile(
            {
                "name": "海安任务 AI 接入",
                "provider": "openai",
                "base_url": "https://ai.example.com",
                "api_key": "sk-profile",
                "model": "profile-model",
                "temperature": 0,
                "prompt": DEFAULT_PROMPT,
                "is_active": False,
            }
        )
        template = save_email_template(
            {
                "name": "海安任务模板",
                "subject_template": "日报 {law_firm_name}",
                "html_template": "<main>{report_body}</main>",
                "is_active": False,
            }
        )
        job = save_job(
            {
                "law_firm_name": "海安律所",
                "aliases": [],
                "exclude_words": [],
                "keywords": ["海安律所避雷"],
                "platforms": ["dy"],
                "recipients": ["target@example.com"],
                "enable_comments": False,
                "time_window_type": "recent_1d",
                "frequency": "daily",
                "email_time": "09:00",
                "enabled": False,
                "ai_profile_id": profile["id"],
                "email_template_id": template["id"],
            }
        )
        monkeypatch.setattr(
            "api.monitoring.preflight.list_platform_status",
            lambda: [
                {"platform": "dy", "platform_label": "抖音", "profile_exists": True, "needs_login": False, "login_window_open": False},
                {"platform": "ks", "platform_label": "快手", "profile_exists": True, "needs_login": False, "login_window_open": False},
                {"platform": "xhs", "platform_label": "小红书", "profile_exists": True, "needs_login": False, "login_window_open": False},
            ],
        )

        preflight = build_job_preflight(job, [])
        ai_check = next(item for item in preflight["checks"] if item["key"] == "ai_config")
        email_check = next(item for item in preflight["checks"] if item["key"] == "email_config")
        template_check = next(item for item in preflight["checks"] if item["key"] == "email_template")

        assert ai_check["severity"] == "warning"
        assert "任务绑定 AI 接入" in ai_check["message"]
        assert "未测试通过" in ai_check["message"]
        assert "邮件配置未测试通过" in email_check["message"]
        assert "任务收件人优先" in email_check["message"]
        assert "发件人不会自动成为收件人" in email_check["message"]
        assert template_check["severity"] == "ok"
        assert "任务绑定邮件模板可用" in template_check["message"]
        assert "系统会插入本次运行生成的报告正文" in template_check["message"]
    finally:
        _restore_monitor_jobs(snapshots["monitor_jobs"])
        _restore_table("ai_key_profiles", snapshots["ai_key_profiles"])
        _restore_table("email_templates", snapshots["email_templates"])
        _restore_singleton_table("email_configs", snapshots["email_configs"])


def test_job_preflight_warns_when_bound_profiles_are_missing(monkeypatch):
    init_db()
    snapshots = {
        "monitor_jobs": _snapshot_monitor_jobs(),
        "email_configs": _snapshot_singleton_table("email_configs"),
    }
    try:
        _clear_monitor_jobs()
        save_email_config(
            {
                "smtp_host": "smtp.example.com",
                "smtp_port": 465,
                "sender": "sender@example.com",
                "default_recipients": [],
                "last_test_status": "success",
            }
        )
        job = save_job(
            {
                "law_firm_name": "海安律所",
                "aliases": [],
                "exclude_words": [],
                "keywords": ["海安律所退费"],
                "platforms": ["dy"],
                "recipients": ["target@example.com"],
                "enable_comments": False,
                "time_window_type": "recent_1d",
                "frequency": "daily",
                "email_time": "09:00",
                "enabled": False,
            }
        )
        job = {**job, "ai_profile_id": 99999901, "email_template_id": 99999902}
        monkeypatch.setattr(
            "api.monitoring.preflight.list_platform_status",
            lambda: [
                {"platform": "dy", "platform_label": "抖音", "profile_exists": True, "needs_login": False, "login_window_open": False},
                {"platform": "ks", "platform_label": "快手", "profile_exists": True, "needs_login": False, "login_window_open": False},
                {"platform": "xhs", "platform_label": "小红书", "profile_exists": True, "needs_login": False, "login_window_open": False},
            ],
        )

        preflight = build_job_preflight(job, [])

        assert preflight["can_run"] is True
        assert any("任务绑定的 AI 接入已不存在" in item for item in preflight["warnings"])
        assert any("任务绑定的邮件模板已不存在" in item for item in preflight["warnings"])
        assert any(item["key"] == "email_template" and item["severity"] == "warning" for item in preflight["checks"])
    finally:
        _restore_monitor_jobs(snapshots["monitor_jobs"])
        _restore_singleton_table("email_configs", snapshots["email_configs"])


def test_job_preflight_blocks_already_running_job():
    job = {"id": 123, "enabled": True, "keywords": ["测试"], "platforms": ["dy"], "recipients": ["a@example.com"]}
    preflight = build_job_preflight(job, [123])

    assert preflight["can_run"] is False
    assert any("正在运行" in item for item in preflight["blockers"])


def test_manual_run_blocks_when_preflight_has_blockers(monkeypatch):
    init_db()
    jobs_snapshot = _snapshot_monitor_jobs()
    try:
        _clear_monitor_jobs()
        job = save_job(
            {
                "law_firm_name": "海安律所",
                "aliases": [],
                "exclude_words": [],
                "keywords": ["海安律所避雷"],
                "platforms": ["dy"],
                "recipients": ["target@example.com"],
                "enable_comments": False,
                "time_window_type": "recent_1d",
                "frequency": "daily",
                "email_time": "09:00",
                "enabled": False,
            }
        )
        monkeypatch.setattr(
            "api.monitoring.preflight.list_platform_status",
            lambda: [
                {"platform": "dy", "platform_label": "抖音", "profile_exists": True, "needs_login": False, "login_window_open": True},
                {"platform": "ks", "platform_label": "快手", "profile_exists": True, "needs_login": False, "login_window_open": False},
                {"platform": "xhs", "platform_label": "小红书", "profile_exists": True, "needs_login": False, "login_window_open": False},
            ],
        )
        called = False

        def fake_launch_job(job_id, source="manual"):
            nonlocal called
            called = True
            return {"started": True}

        monkeypatch.setattr(monitor_router, "launch_job", fake_launch_job)

        with pytest.raises(HTTPException) as exc:
            asyncio.run(monitor_router.run_job_now(job["id"]))

        assert exc.value.status_code == 400
        assert "运行前检查未通过" in str(exc.value.detail)
        assert called is False
    finally:
        _restore_monitor_jobs(jobs_snapshot)


def test_manual_run_allows_preflight_warnings_and_returns_preflight(monkeypatch):
    init_db()
    jobs_snapshot = _snapshot_monitor_jobs()
    ai_snapshot = _snapshot_singleton_table("ai_configs")
    email_snapshot = _snapshot_singleton_table("email_configs")
    try:
        _clear_monitor_jobs()
        save_ai_config({"provider": "openai", "base_url": "", "api_key": "", "model": ""})
        save_email_config({"smtp_host": "", "sender": "", "default_recipients": []})
        job = save_job(
            {
                "law_firm_name": "海安律所",
                "aliases": [],
                "exclude_words": [],
                "keywords": ["海安律所退费"],
                "platforms": ["dy"],
                "recipients": [],
                "enable_comments": False,
                "time_window_type": "recent_1d",
                "frequency": "daily",
                "email_time": "09:00",
                "enabled": False,
            }
        )
        monkeypatch.setattr(
            "api.monitoring.preflight.list_platform_status",
            lambda: [
                {"platform": "dy", "platform_label": "抖音", "profile_exists": True, "needs_login": False, "login_window_open": False},
                {"platform": "ks", "platform_label": "快手", "profile_exists": True, "needs_login": False, "login_window_open": False},
                {"platform": "xhs", "platform_label": "小红书", "profile_exists": True, "needs_login": False, "login_window_open": False},
            ],
        )

        monkeypatch.setattr(
            monitor_router,
            "launch_job",
            lambda job_id, source="manual": {"started": True, "status": "queued", "job_id": job_id, "source": source},
        )

        result = asyncio.run(monitor_router.run_job_now(job["id"]))

        assert result["started"] is True
        assert result["source"] == "manual"
        assert result["preflight"]["can_run"] is True
        assert result["preflight"]["warnings"]
    finally:
        _restore_monitor_jobs(jobs_snapshot)
        _restore_singleton_table("ai_configs", ai_snapshot)
        _restore_singleton_table("email_configs", email_snapshot)


def test_resume_job_blocks_when_preflight_has_blockers(monkeypatch):
    init_db()
    jobs_snapshot = _snapshot_monitor_jobs()
    try:
        _clear_monitor_jobs()
        job = save_job(
            {
                "law_firm_name": "海安律所",
                "aliases": [],
                "exclude_words": [],
                "keywords": ["海安律所避雷"],
                "platforms": ["dy"],
                "recipients": ["target@example.com"],
                "enable_comments": False,
                "time_window_type": "recent_1d",
                "frequency": "daily",
                "email_time": "09:00",
                "enabled": False,
            }
        )
        monkeypatch.setattr(
            "api.monitoring.preflight.list_platform_status",
            lambda: [
                {"platform": "dy", "platform_label": "抖音", "profile_exists": True, "needs_login": False, "login_window_open": True},
                {"platform": "ks", "platform_label": "快手", "profile_exists": True, "needs_login": False, "login_window_open": False},
                {"platform": "xhs", "platform_label": "小红书", "profile_exists": True, "needs_login": False, "login_window_open": False},
            ],
        )

        with pytest.raises(HTTPException) as exc:
            asyncio.run(monitor_router.resume_job(job["id"]))

        assert exc.value.status_code == 400
        assert "启用前检查未通过" in str(exc.value.detail)
        assert get_job(job["id"])["enabled"] is False
    finally:
        _restore_monitor_jobs(jobs_snapshot)


def test_resume_job_allows_warnings_and_returns_preflight(monkeypatch):
    init_db()
    jobs_snapshot = _snapshot_monitor_jobs()
    ai_snapshot = _snapshot_singleton_table("ai_configs")
    email_snapshot = _snapshot_singleton_table("email_configs")
    try:
        _clear_monitor_jobs()
        save_ai_config({"provider": "openai", "base_url": "", "api_key": "", "model": ""})
        save_email_config({"smtp_host": "", "sender": "", "default_recipients": []})
        job = save_job(
            {
                "law_firm_name": "海安律所",
                "aliases": [],
                "exclude_words": [],
                "keywords": ["海安律所避雷"],
                "platforms": ["dy"],
                "recipients": [],
                "enable_comments": False,
                "time_window_type": "recent_1d",
                "frequency": "daily",
                "email_time": "09:00",
                "enabled": False,
            }
        )
        monkeypatch.setattr(
            "api.monitoring.preflight.list_platform_status",
            lambda: [
                {"platform": "dy", "platform_label": "抖音", "profile_exists": True, "needs_login": False, "login_window_open": False},
                {"platform": "ks", "platform_label": "快手", "profile_exists": True, "needs_login": False, "login_window_open": False},
                {"platform": "xhs", "platform_label": "小红书", "profile_exists": True, "needs_login": False, "login_window_open": False},
            ],
        )

        result = asyncio.run(monitor_router.resume_job(job["id"]))

        assert result["ok"] is True
        assert result["job"]["enabled"] is True
        assert result["preflight"]["can_run"] is True
        assert result["preflight"]["warnings"]
    finally:
        _restore_monitor_jobs(jobs_snapshot)
        _restore_singleton_table("ai_configs", ai_snapshot)
        _restore_singleton_table("email_configs", email_snapshot)


def test_job_preflight_and_launcher_block_template_placeholders(monkeypatch):
    job = {
        "id": 123,
        "enabled": True,
        "law_firm_name": "请改成目标律所名称",
        "keywords": ["目标律所避雷"],
        "platforms": ["dy"],
        "recipients": ["a@example.com"],
    }

    preflight = build_job_preflight(job, [])

    assert preflight["can_run"] is False
    assert any("测试数据模板" in item for item in preflight["blockers"])
    monkeypatch.setattr(scheduler_module, "get_job", lambda job_id: job)
    with pytest.raises(ValueError, match="测试数据模板"):
        scheduler_module.launch_job(123)


def test_scheduler_stop_job_requests_runner_stop(monkeypatch):
    scheduler_module._running_jobs.add(24680)
    scheduler_module._job_tasks.pop(24680, None)
    calls: list[int] = []

    def fake_request_stop_job(job_id):
        calls.append(job_id)
        return 2

    try:
        monkeypatch.setattr(scheduler_module, "request_stop_job", fake_request_stop_job)
        result = scheduler_module.stop_job(24680)
    finally:
        scheduler_module._running_jobs.discard(24680)
        scheduler_module._job_tasks.pop(24680, None)

    assert result["stopped"] is True
    assert result["terminated_processes"] == 2
    assert calls == [24680]


def test_refresh_jobs_schedule_api_recomputes_next_run_at():
    init_db()
    job = save_job(
        {
            "law_firm_name": "调度刷新测试律所",
            "aliases": [],
            "exclude_words": [],
            "keywords": ["调度刷新测试律所避雷"],
            "platforms": ["dy"],
            "recipients": [],
            "enable_comments": False,
            "time_window_type": "recent_1d",
            "frequency": "daily",
            "email_time": "23:59",
            "enabled": True,
        }
    )
    try:
        result = asyncio.run(monitor_router.refresh_jobs_schedule())
        refreshed = next(item for item in result["jobs"] if item["id"] == job["id"])
    finally:
        _cleanup_test_records(job["id"], "")

    assert refreshed["next_run_at"]
    assert refreshed["next_run_at"].endswith("23:59:00")


def test_monitor_page_uses_tob_information_architecture_without_customer_facing_engine_traces():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")

    assert "总览" in page
    assert "舆情监控" in page
    assert "任务中心" in page
    assert 'data-tab="reports"' not in page
    assert 'data-shortcut-tab="reports"' not in page
    assert '<section id="reports"' not in page
    assert "资源管理" in page
    assert "系统配置" in page
    assert "dashboard_metrics" in page
    assert "/dashboard" in page
    assert "企业级律所舆情监控" in page
    assert "系统运行状态" in page
    assert "调度器状态" in page
    assert "loadSchedulerStatus" in page
    assert "scheduler-status" in page
    assert "平台账号" in page
    assert "账号资源" in page
    assert "账号详情" in page
    assert "账号列表" in page
    assert "查看账号可用性、登录方式和最近异常" in page
    assert "平台账号概览" not in page
    assert "账号资源台账" not in page
    assert "account_platform_overview" not in page
    assert ".account-console { display:grid; grid-template-columns:minmax(0,1fr);" in page
    assert ".account-list-panel" in page
    assert "account_modal" in page
    assert "drawer-backdrop" in page
    assert "openNewSocialAccountModal" in page
    assert "openSocialAccountModal" in page
    assert "closeSocialAccountModal" in page
    assert 'onclick="openNewSocialAccountModal()">新增账号' in page
    assert 'onclick="openNewSocialAccountModal()">添加账号' not in page
    assert "保存账号" in page
    assert "account_modal_actions" in page
    assert "account_save_button" in page
    assert "account_delete_button" in page
    assert "deleteCurrentSocialAccount" in page
    assert "updateAccountLoginPrerequisites" in page
    assert "qrButton.disabled=!nameReady" in page
    assert "checkSocialAccountLogin" in page
    assert "检测登录态" in page
    assert "checkCurrentSocialAccountLogin" not in page
    assert "account_check_result" not in page
    assert "checkSelectedSocialAccountLogins" in page
    assert "/check-login" in page
    assert "updateAccountModalActions" in page
    assert "这是一个新账号。请先填写基础资料并保存账号，再按需完成扫码或 Cookie 登录。" in page
    assert "account_login_type_filter" in page
    assert "selectedSocialAccountIds" in page
    assert "account_bulk_bar" in page
    assert "account_bulk_toolbar" in page
    assert "account_bulk_count" in page
    assert "account_bulk_check_btn" in page
    assert "未选择账号" in page
    assert "请先勾选账号" in page
    assert "updateAccountBulkToolbar" in page
    assert "toggleAllFilteredAccounts" in page
    assert "toggleAccountSelection" in page
    assert "批量操作只用于可用性维护" in page
    assert "批量停用" in page
    assert "批量启用" in page
    assert "toggleAccountActionMenu" in page
    assert "reloginSocialAccount" in page
    assert "setAccountPlatformLocked" in page
    assert "deleteSelectedSocialAccounts" in page
    assert "startLoginSessionFromSelected" not in page
    assert "openSelectedAccountLoginBrowser" not in page
    assert "accountLedgerTable" in page
    assert 'return `<div class="table-wrap"><table class="account-table">' in page
    assert '<th class="col-actions">操作</th>' in page
    assert "查看账号可用性、登录方式和最近异常" in page
    assert "startLoginSessionForAccount" in page
    assert "openCurrentAccountLoginBrowser" in page
    assert "openLoginSessionBrowser" in page
    assert "localLoginWindowAllowed" in page
    assert "account_local_login_button" in page
    assert "打开登录窗口" in page
    assert "打开登录窗口兜底" not in page
    assert "先按平台和状态定位账号资源" not in page
    assert "1. 基础资料" in page
    assert "2. 登录维护" in page
    assert "4. 完成账号设置" in page
    assert "高级设置" in page
    assert "登录态来源" in page
    assert "social_account_login_source" in page
    assert "social_account_error_summary" in page
    assert "updateAccountDerivedFields" in page
    assert "绑定代理" in page
    assert "不绑定代理" in page
    assert "renderProxySelectOptions" in page
    assert "accountProxyLabel" in page
    assert "plainAccountProxyLabel" in page
    assert "扫码登录" in page
    assert "生成登录二维码" in page
    assert "打开登录窗口" in page
    assert "Cookie 登录" in page
    assert "social_account_cookie_input" in page
    assert "saveCurrentPlatformCookieLogin" in page
    assert "手机号登录" not in page
    assert "social_account_phone" not in page
    assert "saveCurrentPlatformPhoneLogin" not in page
    assert "renderLoginModePanel" in page
    assert "handleSocialLoginTypeChange" in page
    assert "selectSocialLoginType" in page
    assert "supportedSocialLoginTypes" in page
    assert "social_login_method_options" in page
    assert "login-method-option" in page
    assert 'id="social_account_login_type" onchange="handleSocialLoginTypeChange()" style="display:none"' in page
    assert "platform-login-panel" in page
    assert "panel.style.display = active ? 'block' : 'none'" in page
    assert "login-card-grid" in page
    assert "登录状态与平台登录设置" not in page
    assert "运行线索" not in page
    assert "3. 登录记录" in page
    assert "这里只展示最近登录结果，登录和检测操作请在上方维护区或账号列表中完成" in page
    assert "查看状态" not in page
    assert "刷新记录" not in page
    assert "刷新当前账号" not in page
    assert "已生成登录态" in page
    assert "account_metrics" in page
    assert "renderAccountList" in page
    assert "social-accounts" in page
    assert "loadAccountsPool" in page
    assert "二维码已生成，系统正在自动确认登录结果" in page
    assert "点击“生成二维码并登录”后，系统会自动确认扫码和登录结果。" in page
    assert "系统正在自动确认登录结果，每 3 秒刷新一次。" in page
    assert "登录成功，账号已保存" in page
    assert "生成二维码" in page
    assert "手机扫码确认" in page
    assert "保存登录态" in page
    assert "needs_verification" in page
    assert "waiting_scan" in page
    assert "waiting_confirm" in page
    assert "等待验证" in page
    assert "平台要求先完成验证，请按文字提示处理" in page
    assert "平台验证页面截图" not in page
    assert "verification_image" not in page
    assert "verification_label" in page
    assert "verification_detail" in page
    assert "diagnostic_image" not in page
    assert "登录页面诊断截图" not in page
    assert "我已处理，继续确认" in page
    assert "login-sessions" in page
    assert "pollLoginSession" in page
    assert "代理资源" in page
    assert "proxy_resource_summary" in page
    assert "proxy_resource_count" in page
    assert "proxy_search" in page
    assert "proxy_status_filter" in page
    assert "clearProxyFilters" in page
    assert "renderProxyResourceSummary" in page
    assert "renderProxyProfilesTable" in page
    assert "proxies" in page
    assert "loadProxyPool" in page
    assert "platformStatusTable" in page
    assert "platform_login_config_table" not in page
    assert "loadPlatformLoginConfigs" in page
    assert "platform-login-configs" in page
    assert "savePlatformLoginConfig" not in page
    assert "下一步处理" in page
    assert "尚未完成平台采集" in page
    assert "readiness_actions" in page
    assert "renderReadinessActions" in page
    assert "platformLoginActionNote" in page
    assert "action-card" in page
    assert "去账号池处理登录" in page
    assert "检查 AI 接入" in page
    assert "配置测试邮件" in page
    assert "运行抖音采集" in page
    assert "查看任务中心" in page
    assert "switchTab" in page
    assert "已运行但未采到内容" in page
    assert "采集无结果" in page
    assert "系统诊断" in page
    assert "Refresh schedule times" in page
    assert "refreshJobSchedule" in page
    assert "jobs/refresh-schedule" in page
    assert "toast('任务已保存'); resetJobForm(); closeJobDrawer(); await Promise.all([loadJobs(), loadDashboard(), loadDoctor()]);" in page
    assert "toast('调度时间已刷新');" in page
    assert "await Promise.all([loadJobs(), loadSchedulerStatus(), loadDashboard()]);" in page
    assert "打开登录窗口" in page
    assert "用于默认登录态维护；账号资源请在账号详情里发起登录。" not in page
    assert "平台默认登录态" not in page
    assert "如平台需要额外确认，系统会提示下一步操作。" in page
    assert "login-browser" in page
    assert "openPlatformLoginBrowser" in page
    assert "正在运行的任务 ID" in page
    assert "运行 ID" in page
    assert "任务 ID" in page
    assert "run_log_drawer" in page
    assert "copyCurrentRunLogs" in page
    assert "downloadCurrentRunLogs" in page
    assert "全部运行记录" in page
    assert "skipped" in page
    assert "runStatusBadge" in page
    assert "runDisplayError" in page
    assert "jobActions" in page
    assert "toggleJobActionMenu" in page
    assert "/jobs/'+id+'/stop" in page
    assert "/runs/'+id+'/stop" in page
    assert "await Promise.all([loadRuns(), loadSchedulerStatus(), loadDashboard()]);" in page
    assert "任务正在运行，请先停止后再删除" in page
    assert "startRunPolling" in page
    assert "api('/doctor')" in page
    assert "运行系统诊断" in page
    assert "runSmokeCheck" in page
    assert "smoke_result" in page
    assert "api('/smoke'" in page
    assert "renderSmokeResult" in page
    assert "正在运行系统诊断，请稍候。" in page
    assert "formatBytes" in page
    assert "preflight" in page
    assert "运行前提示" in page
    assert "填入海安律所样例" in page
    assert "1. 目标" in page
    assert "2. 采集内容" in page
    assert "3. 调度" in page
    assert "4. 报告" in page
    assert "normal_task_wizard_steps" in page
    assert "normal_task_wizard_hint" in page
    assert "wizard-step" in page
    assert "wizard-section" in page
    assert "normal-only" in page
    assert "普通用户只需填写目标律所、平台搜索词、采集范围、调度和收件邮箱。" in page
    assert "账号、代理、AI 接入、邮件模板和浏览器方式由管理员维护。" in page
    assert "管理员高级采集设置" in page
    assert "admin-only-job-field" in page
    assert "applyJobFormRoleMode" in page
    assert "isAdminUser" in page
    assert "document.querySelectorAll('.admin-only-job-field')" in page
    assert "document.querySelectorAll('.normal-only')" in page
    assert "采集条数是内容数量上限" in page
    assert "部分平台可能返回少于所选范围的结果" in page
    assert "任务运行超时由管理员在运行策略中统一控制" in page
    assert "过滤与去重" in page
    assert "平台搜索词（多行）" in page
    assert "监控对象" not in page
    assert "这里决定系统采什么内容" not in page
    assert "关键词栏" not in page
    assert "每 6 小时，起点" in page
    assert "每 12 小时，起点" in page
    assert "fill_sample_job_btn" in page
    assert "addEventListener('click', fillSampleJobTemplate)" in page
    assert "fillSampleJobTemplate" in page
    assert "hasJobTemplatePlaceholders" in page
    assert "请先将律所名称和平台搜索词改成真实内容" in page
    assert "恢复默认规则" in page
    assert "default_prompt" in page
    assert "resetAIPrompt" in page
    assert "基础信息" in page
    assert "评估规则列表" in page
    assert "openNewAIRuleModal" in page
    assert "loadAIRuleProfiles" in page
    assert "ai-rule-profiles" in page
    assert "ai_rule_modal" in page
    assert "rule-modal-flow" in page
    assert "rule-modal-layout" not in page
    assert "rule-editor-stack" in page
    assert "rule-test-stack" in page
    assert "rule-side-stack" not in page
    assert "ai_rule_active_label" not in page
    assert "rule-accordion" in page
    assert "testAIRuleFromModal" in page
    assert "saveAIRuleFromModal" in page
    assert "activateAIRuleProfile" in page
    assert "deleteAIRuleProfile" in page
    assert "aiRuleActions" in page
    assert "toggleAIRuleActionMenu" in page
    assert "测试规则" in page
    assert "规则配置" in page
    assert "角色定位" in page
    assert "相关性判断" in page
    assert "疑似负面判断" in page
    assert "风险等级规则" in page
    assert "证据摘录规则" in page
    assert "处理建议规则" in page
    assert "生成后的 Prompt 预览" in page
    assert "composeAIPromptFromRules" in page
    assert "parsePromptSections" in page
    assert "applyPromptToRuleFields" in page
    assert "renderAIOutputSchema" in page
    assert "prompt_sections" in page
    assert "output_schema" in page
    ai_rule_modal = page[page.index('id="ai_rule_modal"') : page.index('id="email"')]
    assert ai_rule_modal.index("基础信息") < ai_rule_modal.index("规则配置") < ai_rule_modal.index("固定输出字段") < ai_rule_modal.index("测试样例") < ai_rule_modal.index("测试结果")
    assert "测试样例" in page
    assert "ai_sample_law_firm_name" in page
    assert "ai_sample_platform" in page
    assert "ai_sample_source_keyword" in page
    assert "ai_sample_title" in page
    assert "ai_sample_text" in page
    assert "ai_sample_comments" in page
    assert "海安律所避雷：退费拖了很久" in page
    assert "sample_law_firm_name:val('ai_sample_law_firm_name')" in page
    assert "sample_platform:val('ai_sample_platform')" in page
    assert "sample_source_keyword:val('ai_sample_source_keyword')" in page
    assert "sample_title:val('ai_sample_title')" in page
    assert "sample_text:val('ai_sample_text')" in page
    assert "sample_comments:val('ai_sample_comments')" in page
    assert "is_related(boolean), is_negative(boolean), risk_level(high|medium|low|irrelevant)" in page
    assert "AI 接入资源" in page
    assert "接口协议" in page
    assert "Provider" not in page
    assert "ai_resource_summary" in page
    assert "ai_resource_count" in page
    assert "ai_profile_search" in page
    assert "ai_profile_provider_filter" in page
    assert "ai_profile_test_filter" in page
    assert "clearAIProfileFilters" in page
    assert "renderAIResourceSummary" in page
    assert "renderAIProfilesTable" in page
    assert "获取模型列表" in page
    assert "可手动输入，或获取列表后选择" in page
    assert "ai_profile_model_options" in page
    assert "toggleAIProfileModelOptions" in page
    assert "selectAIProfileModel" in page
    assert "loadAIProfileModels" in page
    assert "ai-profiles/models" in page
    assert "models" in page
    assert "ai-profiles" in page
    assert "loadAIProfiles" in page
    assert "activateAIProfile" in page
    assert "testAIProfile" in page
    assert "testAIProfile" in page
    assert "ai-profiles/'+id+'/connection-test" in page
    assert "ai-profiles/'+id+'/connection-test" in page
    assert "ai_connection_test_modal" in page
    assert "ai_result" not in page
    assert "openAIConnectionTestModal" in page
    assert "runAIConnectionTest" in page
    assert "closeAIConnectionTestModal" in page
    assert "测试 AI 接入" in page
    assert "开始测试" in page
    assert "测试消息" in page
    assert "模型返回" in page
    assert "模型已返回文本" in page
    assert "连接测试" in page
    assert "连接测试" in page
    assert "ai-evaluation-config/test" in page
    assert "testAI" in page
    assert "ai_profile_model_status" not in page
    assert "model-combobox-status" not in page
    assert "部分服务不提供模型列表；获取失败时仍可手动填写模型名称。" not in page
    assert "可获取当前接入下的模型列表；如果服务不支持列表接口，可保持手动填写。" not in page
    assert "HTML 邮件模板" in page
    assert "email_template_summary" in page
    assert "email_template_resource_summary" in page
    assert "email_template_search" in page
    assert "email_template_status_filter" in page
    assert "clearEmailTemplateFilters" in page
    assert "renderEmailTemplateResourceSummary" in page
    assert "renderEmailTemplatesTable" in page
    assert "mail_config_modal" in page
    assert "openMailConfigModal" in page
    assert "closeMailConfigModal" in page
    assert "mail_test_modal" in page
    assert "openMailTestModal" in page
    assert "closeMailTestModal" in page
    assert "mail_test_console" in page
    assert "mail_test_start_btn" in page
    assert "email_subject_summary" in page
    assert "真实邮件：已关闭" in page
    assert "开启后测试邮件、手动重发和自动交付会真实提交 SMTP" in page
    assert "email_validation_status" not in page
    assert "email-validation-compact" not in page
    assert "email-toolbar-switch" in page
    assert "email_validation_summary" not in page
    assert "real_email_delivery_toggle" in page
    assert "real_email_delivery_toggle_label" in page
    assert "setRealEmailDelivery(this.checked)" in page
    assert "email_validation_open_btn" not in page
    assert "email_validation_close_btn" not in page
    assert "SMTP 已接受不代表收件箱已收到" in page
    assert "renderEmailValidationWindow" in page
    assert "emailRecipientSourceLabel" in page
    email_section = page[page.index('<section id="email"') : page.index('<div id="mail_config_backdrop"')]
    assert email_section.count("openMailConfigModal()") == 1
    assert email_section.count("openMailTestModal()") == 1
    assert "SMTP 与发送默认值" not in email_section
    assert "真实邮件发送状态尚未读取" not in email_section
    assert "smtp_password_status" in page
    assert "已保存密码" in page
    assert "如需更换请重新输入" in page
    assert "email-templates/preview" in page
    assert "email_template_preview" in page
    assert "scheduleEmailPreview" in page
    assert "线索明细" not in page
    task_center_section = _monitor_section(page, "runs")
    delivery_drawer = page[page.index('id="email_delivery_history_drawer"') : page.index('id="email_template_drawer_backdrop"')]
    assert "<label>线索状态</label>" not in task_center_section
    assert 'id="report_risk"' not in task_center_section
    assert '<label>线索状态</label>' not in page
    assert 'id="lead_status_filter"' not in page
    assert "筛选线索" not in page
    assert "<label>风险</label>" not in task_center_section
    assert "<label>报告范围</label>" in page
    assert "全部报告和线索" not in task_center_section
    assert "线索状态：" not in page
    assert "data-report-lead-panel" not in task_center_section
    assert "leads_table" not in task_center_section
    assert "report-hint" not in task_center_section
    assert "选择报告后查看正文" not in task_center_section
    assert "点击报告列表中的“预览”" not in task_center_section
    assert "查看交付历史" not in task_center_section
    assert "点击报告行的邮件状态或“更多 > 查看交付历史”会打开悬浮窗" not in task_center_section
    assert 'id="email_delivery_history"' not in task_center_section
    assert "data-report-delivery-panel" not in task_center_section
    assert "data-report-delivery-panel" in delivery_drawer
    assert 'id="email_delivery_history"' in delivery_drawer
    assert 'id="email_delivery_history_scope"' in delivery_drawer
    assert 'id="email_delivery_history_count"' in delivery_drawer
    assert "report_leads_drawer" not in page
    assert "report_leads_backdrop" not in page
    assert "closeReportLeadsDrawer" not in page
    assert "leads_scope_hint" not in page
    assert "线索限定在当前报告，不作为全局线索工作台。" not in page
    assert "线索限定在当前运行，不作为全局线索工作台。" not in page
    assert "viewRunLeads" not in page
    assert "leadScopeForRun" not in page
    assert "reloadCurrentLeadDrawer" not in page
    assert "risk:val('lead_status_filter')" not in page
    assert "risk:val('report_risk')" not in page
    assert "report-workspace" not in page
    assert "action-menu-host" in page
    assert "report-action-menu" not in page
    assert "data-report-menu-button" not in page
    assert "positionReportActionMenu" not in page
    assert "renderReportsTable" not in page
    assert "currentReports" not in page
    assert "reportActions" not in page
    assert 'title="交付历史可从更多菜单或运行详情查看"' not in page
    assert '请从运行详情的邮件交付区域打开' in page
    assert "下载 Markdown" in page
    assert "api('/leads?" not in page
    assert "jumpToReportAiEvaluations" in page
    assert "run_detail_report_filter" in page
    assert "setRunDetailAiFilters" in page
    assert "report_id: reportId" in page
    assert "当前筛选条件下的线索" not in page
    assert "待人工复核" in page
    assert "待复核" in page
    assert "运行系统诊断" in page
    assert "reports/system-check" in page
    assert "loadRuns(), loadReadiness(), loadDoctor(), loadDashboard()" in page
    assert "重发邮件" in page
    assert "系统自检报告已生成" not in page


def test_frontend_sms_verification_has_manual_code_submission():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")

    assert "function submitLoginVerificationCode" in page
    assert "/verification-code" in page
    assert "autocomplete=\"one-time-code\"" in page
    assert "提交短信验证码" in page
    assert "capabilities.verification_type === 'sms'" in page


def test_frontend_sms_verification_has_send_request():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")

    assert "function requestLoginVerificationCode" in page
    assert "/verification-code/request" in page
    assert "发送短信验证码" in page
    assert "submitLoginVerificationCode" in page
    assert "capabilities.verification_type === 'sms'" in page


def test_frontend_sms_verification_preserves_input_during_polling():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")

    assert "function currentLoginVerificationCode(sessionId)" in page
    assert "const existingVerificationCode=currentLoginVerificationCode(session.id)" in page
    assert "value=\"${esc(existingVerificationCode)}\"" in page
    assert "if(normalizeLoginSessionStatus(data.session.status)==='needs_verification') return true;" in page
    assert "toast(cleanCustomerText(data.session.message||'短信验证码已提交，请等待平台确认登录结果。'))" in page


def test_frontend_sms_verification_panel_is_structured_and_inline_validated():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")

    assert ".login-verification-panel" in page
    assert ".login-verification-actions" in page
    assert ".login-verification-code-row" in page
    assert ".login-verification-error" in page
    assert "function setLoginVerificationError" in page
    assert "id=\"login_verification_error_${Number(session.id)}\"" in page
    assert "aria-describedby=\"login_verification_error_${Number(session.id)}\"" in page
    assert "setLoginVerificationError(sessionId, '请输入 4-8 位短信验证码')" in page
    assert "toast('请输入 4-8 位短信验证码')" not in page
    assert "const displayTitle=canSubmitLoginVerificationCode && status==='needs_verification' ? '平台要求完成短信验证码'" in page
    assert "平台要求完成短信验证码" in page
    assert "请先发送验证码，收到短信后输入并提交。" in page
    assert "1. 发送短信验证码" in page
    assert "2. 提交短信验证码" in page
    assert "发送中..." in page
    assert "capabilities.verification_type === 'sms'" in page


def test_monitor_page_uses_consistent_buttons_tables_and_modal_actions():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    monitor_css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    frontend_source = page + "\n" + monitor_css

    assert "white-space: nowrap" in frontend_source or "white-space:nowrap" in frontend_source
    assert "word-break: keep-all" in frontend_source or "word-break:keep-all" in frontend_source
    assert "min-height: 36px" in frontend_source or "min-height:36px" in frontend_source
    assert ".row > * { flex:0 1 auto; }" in page
    assert ".row > button, .row > a { flex:0 0 auto; }" in page
    assert ".wide-actions { display:inline-flex; gap:6px; align-items:center; flex-wrap:nowrap;" in page
    assert "td.col-actions, th.col-actions" in page
    assert "right: 0;" in monitor_css
    assert "th.col-actions" in monitor_css
    assert "function tableColumnClass(header)" in page
    assert "if(['操作','详情'].includes(header)) classes.push('col-actions');" in page
    assert "if(header==='状态') classes.push('col-status');" in page
    assert "Math.max(Number(options.minWidth||0) || 920, (headers||[]).length * 112)" in page
    assert "class=\"form-actions\"" in page
    assert ".form-actions { position:relative;" in page
    assert ".account-flow-actions { position:relative;" in page
    assert ".ai-test-actions { position:relative;" in page
    assert ".rule-modal-actions { position:relative;" in page
    assert ".resource-modal-actions { position:relative;" in page
    assert "drawer-fixed-footer" in frontend_source
    assert ".resource-summary-grid { display:grid;" in page
    assert ".resource-toolbar { display:flex;" in page
    assert ".mail-test-modal" in page
    assert ".test-console" in page
    assert ".config-section .section-note, .config-section .field-hint { display:none; }" in page
    assert ".ui-icon svg" in page
    assert '<symbol id="icon-dashboard"' in page
    assert '<use href="#icon-monitor">' in page
    assert ".page-toolbar {" in frontend_source
    assert "report-workspace" not in page
    assert ".schema-item { grid-template-columns:1fr; }" in page or ".schema-item {" in monitor_css
    assert ".action-menu-host" in frontend_source
    assert ".action-menu.active" in frontend_source
    assert ".report-action-menu {" not in frontend_source
    assert "function enhanceFilterSelects(root=document)" in page
    assert "root.querySelectorAll('.page-filter-region select:not([data-filter-select-enhanced])')" in page
    assert "className='filter-select-menu'" in page
    assert "document.body.appendChild(menu)" in page
    assert "activeFilterSelect.dispatchEvent(new Event('change', {bubbles:true}))" in page
    assert ".filter-select-menu {\n  position: fixed;" in monitor_css
    assert "z-index: calc(var(--z-modal) + 12);" in monitor_css
    assert ".filter-select-native" in monitor_css
    assert ".filter-select-button" in monitor_css
    assert "el.matches('.page-filter-region select[data-filter-select-enhanced=\"1\"]')" in page
    assert "function enhanceFilterDateInputs(root=document)" in page
    assert "root.querySelectorAll('.page-filter-region input[type=\"date\"]:not([data-filter-date-enhanced])')" in page
    assert "className='filter-date-menu'" in page
    assert "data-filter-date-value" in page
    assert "const input=activeFilterDateInput;" in page
    assert "input.dispatchEvent(new Event('change', {bubbles:true}))" in page
    assert "activeFilterDateInput.dispatchEvent(new Event('change', {bubbles:true}))" not in page
    assert ".filter-date-menu {\n  position: absolute;" in monitor_css
    assert ".filter-date-native" in monitor_css
    assert ".filter-date-button" in monitor_css
    assert "el.matches('.page-filter-region input[type=\"date\"][data-filter-date-enhanced=\"1\"]')" in page
    assert "openReportMenuId" not in page
    assert "resourceStat(label, value)" in page
    assert "renderProxyProfilesTable" in page
    assert "renderAIProfilesTable" in page
    assert "renderEmailTemplatesTable" in page
    assert "addEventListener('input', renderProxyProfilesTable)" in page
    assert "addEventListener('input', renderAIProfilesTable)" in page
    assert "addEventListener('input', renderEmailTemplatesTable)" in page
    assert "oldHtml = btn ? btn.innerHTML : ''" in page
    assert "btn.innerHTML = oldHtml" in page
    assert "<div class=\"wide-actions\"><button class=\"secondary\" onclick=\"switchTab('accounts')\">管理账号</button></div>" in page
    assert "jobActions(j, running)" in page
    assert "leadLinks(item)" not in page
    assert "resendReportEmail" in page
    assert "resend-email" in page
    assert "邮件预览" in page
    assert "report_email_subject" in page
    assert "email-preview" in page
    assert "邮件标题：" in page
    assert "download?type=html" in page


def test_cr071_drawer_modal_selects_reuse_filter_dropdown_mechanism():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")

    assert "function refreshEnhancedFilterSelects(root=document)" in page
    assert "refreshEnhancedFilterSelects(dialog)" in page
    for opener_name in [
        "openJobDrawer",
        "openProxyDrawer",
        "openAIProfileDrawer",
        "openAIRuleModal",
        "openMailConfigModal",
        "openEmailTemplateDrawer",
    ]:
        start = page.index(f"function {opener_name}")
        end = page.find("\n    function ", start + 1)
        snippet = page[start:end]
        assert "openDrawerChrome(" in snippet
        assert "refreshEnhancedFilterSelects(drawer)" in snippet
    assert "function refreshEnhancedFilterDateInputs(root=document)" in page
    assert "refreshEnhancedFilterDateInputs(document.getElementById('job_drawer'))" in page
    assert ".modal-filter-region.page-filter-region" in page
    assert ".modal-filter-region.page-filter-region" in css

    enhanced_select_ids = [
        "job_account_id",
        "job_proxy_id",
        "job_target_type",
        "job_output_mode",
        "job_browser_mode",
        "time_window_type",
        "frequency",
        "job_ai_profile_id",
        "job_email_template_id",
        "social_account_platform",
        "social_account_status",
        "social_account_proxy_id",
        "proxy_status",
        "ai_profile_provider",
        "ai_sample_platform",
        "encryption",
        "email_template_preset",
    ]
    for select_id in enhanced_select_ids:
        marker = f'<select id="{select_id}"'
        select_index = page.index(marker)
        region_index = page.rfind("page-filter-region modal-filter-region", 0, select_index)
        assert region_index != -1, select_id
        assert select_index - region_index < 1400, select_id

    for function_name in [
        "renderJobAccountOptions",
        "renderProxySelectOptions",
        "renderJobAIProfileOptions",
        "renderJobEmailTemplateOptions",
    ]:
        start = page.index(f"function {function_name}")
        end = page.find("\n    function ", start + 1)
        snippet = page[start:end]
        assert "syncFilterSelectButton(select);" in snippet

    ai_profile_drawer = page[
        page.index('id="ai_profile_drawer"') : page.index('id="ai_rule_modal_backdrop"')
    ]
    assert 'id="ai_profile_model"' in ai_profile_drawer
    assert '<select id="ai_profile_model"' not in ai_profile_drawer
    assert 'id="ai_profile_model_options" class="model-options"' in ai_profile_drawer

    for date_id in ["custom_start", "custom_end"]:
        marker = f'<input id="{date_id}" type="date">'
        date_index = page.index(marker)
        region_index = page.rfind("page-filter-region modal-filter-region", 0, date_index)
        assert region_index != -1, date_id
        assert date_index - region_index < 500, date_id

    assert "filter-date-enhanced" in css
    assert "filter-select-button filter-date-button" in page
    assert ".filter-date-menu {\n  position: absolute;" in css
    assert "top='calc(100% + 4px)'" in page
    assert "download?type=excel" in page
    assert "download?type=markdown" in page
    assert "下一步处理" in page
    assert "next_actions" in page
    forbidden = [
        "MediaCrawler",
        "MONITOR_SKIP_AI_API",
        "selftest",
        "上线验收",
        "待验收",
        "项目进展",
        "MVP自测",
        "登录窗口测试律所",
        "first commit",
        "生成自测报告",
        "本地冒烟自检",
        "部署诊断",
        "运行三平台采集",
    ]
    assert not [word for word in forbidden if word in page]


def test_phase_11a_monitor_static_boundary_and_tokens_are_available():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    js = Path("api/webui/monitor/monitor.js").read_text(encoding="utf-8")

    stylesheet = '<link rel="stylesheet" href="/static/monitor/monitor.css">'
    module_script = '<script type="module" src="/static/monitor/monitor.js"></script>'

    assert stylesheet in page
    assert module_script in page
    assert page.index(stylesheet) < page.index("<style>")
    assert page.index("</script>") < page.index(module_script)

    for token in [
        "--color-neutral-0",
        "--color-primary-600",
        "--color-status-success-text",
        "--color-status-warning-text",
        "--color-status-danger-text",
        "--color-status-info-text",
        "--color-navigation-bg",
        "--font-family-sans",
        "--font-size-md",
        "--font-weight-semibold",
        "--line-height-base",
        "--space-4",
        "--space-page-x",
        "--radius-control-medium",
        "--shadow-elevation-1",
        "--z-floating-menu",
        "--transition-duration-base",
        "--breakpoint-mobile-max",
        "--breakpoint-tablet-min",
        "--breakpoint-desktop-min",
    ]:
        assert token in css

    for legacy_alias in [
        "--bg:",
        "--surface:",
        "--line:",
        "--text:",
        "--muted:",
        "--primary:",
        "--radius:",
    ]:
        assert legacy_alias not in css

    for forbidden_js in ["console.", "globalThis"]:
        assert forbidden_js not in js


def test_phase_11b_base_layout_styles_live_in_monitor_css():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")

    for selector in [
        ".shell {",
        ".shell > aside",
        ".brand {",
        "nav {",
        ".nav-group {",
        "nav button {",
        "header {",
        ".header-title strong",
        ".metric-card {",
        "button.primary",
        "button.secondary",
        "button.danger",
        ".toolbar {",
        ".toolbar-actions {",
    ]:
        assert selector in css

    inline_style = _monitor_inline_styles(page)
    inline_base_style = inline_style.split("@media", 1)[0]
    for migrated_selector in [
        ".shell {",
        ".shell > aside",
        ".brand {",
        "nav { display:grid",
        "nav button {",
        "header {",
        "button.primary, button.secondary, button.danger",
        ".toolbar {",
        "\n    .toolbar-actions {",
    ]:
        assert migrated_selector not in inline_base_style


def test_phase_11c_interaction_helpers_and_floating_menus():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    js = Path("api/webui/monitor/monitor.js").read_text(encoding="utf-8")
    inline_style = _monitor_inline_styles(page)

    for helper in [
        "root.MonitorUI = Object.freeze",
        "showToast",
        "showLoading",
        "renderEmptyState",
        "closeFloatingMenus",
        "positionFloatingMenu",
        "ensurePortalRoot",
    ]:
        assert helper in js

    assert "console." not in js
    assert "import " not in js

    for selector in [
        ".toast {",
        ".monitor-loading",
        ".empty-state",
        ".drawer-backdrop {",
        ".drawer {",
        ".modal-close",
        ".action-menu,",
        ".account-action-menu {",
        ".monitor-portal-root",
    ]:
        assert selector in css

    assert "position: fixed;" in css
    assert ".account-action-menu { position:absolute" not in inline_style
    assert ".action-menu { position:absolute" not in inline_style
    assert ".report-action-menu {" not in inline_style
    assert ".report-action-menu" not in css

    for marker in [
        "function positionFloatingMenu(triggerEl, menuEl",
        "window.MonitorUI.positionFloatingMenu",
        "function positionActiveFloatingMenus()",
        "document.addEventListener('keydown'",
        "event.key === 'Escape'",
        "document.addEventListener('monitor:close-floating-menus', closeFloatingMenus)",
        "window.addEventListener('resize', positionActiveFloatingMenus)",
        "window.addEventListener('scroll', positionActiveFloatingMenus, true)",
        "closeFloatingMenus();",
        "data-account-menu-button",
        "data-job-menu-button",
        "data-ai-rule-menu-button",
        'id="account_action_menu"',
        'id="job_action_menu"',
        'id="ai_rule_action_menu"',
        "function renderAccountActionMenu()",
        "function renderJobActionMenu()",
        "function renderAIRuleActionMenu()",
        "positionFloatingMenu(anchor, menu",
    ]:
        assert marker in page

    assert 'id="account_menu_${id}"' not in page
    assert 'id="job_menu_${id}"' not in page
    assert 'id="ai_rule_menu_${id}"' not in page
    assert "data-report-menu-button" not in page
    assert 'id="report_action_menu"' not in page
    assert "data-proxy-menu-button" not in page
    assert "data-ai-profile-menu-button" not in page
    assert "data-email-template-menu-button" not in page


def test_phase_11d_responsive_foundation_and_mobile_navigation():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    inline_style = _monitor_inline_styles(page)

    for marker in [
        'id="mobile_nav_toggle"',
        'class="mobile-nav-toggle"',
        'aria-controls="primary_navigation"',
        'aria-expanded="false"',
        'id="mobile_nav_backdrop"',
        'class="mobile-nav-backdrop"',
        'id="primary_sidebar"',
        'id="primary_navigation"',
        'class="row header-actions"',
    ]:
        assert marker in page

    for marker in [
        "function setMobileNavOpen(open)",
        "function closeMobileNav()",
        "function toggleMobileNav()",
        "document.body.classList.toggle('mobile-nav-open'",
        "const sidebar=document.getElementById('primary_sidebar')",
        "sidebar.style.setProperty('left', open ? '0px' : '-340px', 'important')",
        "document.getElementById('mobile_nav_toggle')?.addEventListener('click'",
        "document.getElementById('mobile_nav_backdrop')?.addEventListener('click', closeMobileNav)",
        "if(open && window.innerWidth >= 768) open=false",
        "if(window.innerWidth >= 768) closeMobileNav()",
        "closeMobileNav();",
        "if(!canMenu('mail_templates')) return;",
    ]:
        assert marker in page

    assert "@media (max-width: 1279px)" in css
    assert "@media (min-width: 768px) and (max-width: 1279px)" in css
    assert "@media (max-width: 767px)" in css
    assert "@media (max-width: 1100px)" not in inline_style
    assert "@media (max-width: 720px)" not in inline_style
    inline_header_actions_block = inline_style.split("body header .header-actions {", 1)[1].split("}", 1)[0]
    assert "display:contents !important;" in inline_header_actions_block
    assert "width:auto !important;" in inline_header_actions_block
    assert "grid-template-areas:\n          \"nav title refresh account\"\n          \"status status status status\";" in inline_style
    inline_title_strong_block = inline_style.split("body header .header-title strong {", 1)[1].split("}", 1)[0]
    assert "display:block;" in inline_title_strong_block
    assert "white-space:nowrap;" in inline_title_strong_block
    assert "word-break:keep-all;" in inline_title_strong_block
    inline_top_status_block = inline_style.split("body header #top_status {", 1)[1].split("}", 1)[0]
    assert "grid-area:status;" in inline_top_status_block
    assert "display:flex !important;" in inline_top_status_block
    inline_account_area_block = inline_style.split("body header .account-area {", 1)[1].split("}", 1)[0]
    assert "grid-area:account;" in inline_account_area_block
    assert "width:auto;" in inline_account_area_block

    for selector in [
        ".mobile-nav-toggle",
        ".mobile-nav-backdrop",
        "body.mobile-nav-open",
        "body.mobile-nav-open .shell > aside",
        ".nav-sublist",
        ".table-wrap",
        ".page-toolbar .toolbar-actions",
        ".drawer {",
        ".template-modal,",
        ".form-drawer,",
    ]:
        assert selector in css

    assert "grid-template-columns: minmax(0, 1fr);" in css
    assert "height: 100dvh;" in css
    tablet_nav_block = css.split("@media (min-width: 768px) and (max-width: 1279px)", 1)[1].split(
        "@media (max-width: 767px)", 1
    )[0]
    mobile_nav_block = css.split("@media (max-width: 767px)", 1)[1]
    assert "body.sidebar-collapsed .shell {\n    grid-template-columns: 68px minmax(0, 1fr);" in tablet_nav_block
    assert (
        ".mobile-nav-toggle,\n"
        "  .mobile-nav-backdrop,\n"
        "  body.mobile-nav-open .mobile-nav-backdrop {\n"
        "    display: none !important;"
    ) in tablet_nav_block
    assert ".sidebar-collapse-button {\n    display: flex !important;" in tablet_nav_block
    assert "body.mobile-nav-open {" in mobile_nav_block
    assert "body.mobile-nav-open #primary_sidebar {\n    left: 0 !important;\n    transform: none !important;" in mobile_nav_block
    assert ".mobile-nav-toggle {\n    display: inline-flex;" in mobile_nav_block
    assert "grid-template-columns: 40px minmax(0, 1fr) 36px 44px;" in mobile_nav_block
    assert (
        'grid-template-areas:\n'
        '      "nav title refresh account"\n'
        '      "status status status status";'
    ) in mobile_nav_block
    assert "body header .mobile-nav-toggle > span:not(.mobile-nav-icon) {\n    display: none;" in mobile_nav_block
    assert "body header .header-actions {\n    display: contents !important;" in mobile_nav_block
    assert "body header .header-title {\n    grid-area: title;" in mobile_nav_block
    assert "display: block;" in mobile_nav_block
    assert "writing-mode: horizontal-tb;" in mobile_nav_block
    top_status_block = mobile_nav_block.split("body header #top_status {", 1)[1].split("}", 1)[0]
    assert "display: flex !important;" in top_status_block
    assert "grid-area: status;" in top_status_block
    assert "overflow: visible;" in top_status_block
    assert "flex-wrap: wrap;" in top_status_block
    assert "body header #global_refresh_button {\n    grid-area: refresh;" in mobile_nav_block
    assert ".account-area {\n    display: inline-flex;" in mobile_nav_block
    assert "grid-area: account;" in mobile_nav_block
    assert "body header .header-actions > .account-area {\n    min-width: 0;\n    width: auto;" in mobile_nav_block
    assert "body:not(.mobile-nav-open) #primary_sidebar {\n    left: -340px !important;\n    transform: none !important;" in mobile_nav_block
    assert "max-height: calc(100dvh - 18px);" in css
    assert ".table-wrap table" in css
    assert "min-width: 760px !important;" in css


def test_phase_21a_global_shell_and_design_tokens_refinement():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    js = Path("api/webui/monitor/monitor.js").read_text(encoding="utf-8")

    assert "/* Phase 21 formal console refinement layer: visual-only, no workflow changes. */" in css
    for token in [
        "--phase21-bg: #f3f6f8;",
        "--phase21-surface: #ffffff;",
        "--phase21-line: #d6e0e8;",
        "--phase21-text: #17212f;",
        "--phase21-muted: #586a7d;",
        "--phase21-accent: #0f766e;",
        "--phase21-nav-hover: #edf1f5;",
        "--phase21-nav-active: #e8fbf6;",
        "--phase21-success-bg: #edf8f3;",
        "--phase21-warning-bg: #fff7e8;",
        "--phase21-danger-bg: #fff0f2;",
        "--phase21-info-bg: #edf5fb;",
        "--phase21-shadow-overlay:",
        "--phase21-focus:",
    ]:
        assert token in css

    for selector in [
        "html body {",
        "body .shell {",
        "body .content {",
        "body .shell > aside {",
        "body header {",
        "body input:focus-visible,",
        "body button.primary {",
        "body button.secondary,",
        "body button.danger {",
        "body button:disabled {",
        "body .status.ok,",
        "body .status.warn,",
        "body .status.bad,",
        "body .empty,\nbody .empty-state,",
        "body .skeleton-block,",
        "body .drawer {",
        "body .modal-close:hover {",
        "body .refresh-icon-button {",
    ]:
        assert selector in css

    assert "body .shell > aside {\n  position: fixed;" in css
    assert "overflow-x: hidden;" in css
    assert "body * {\n  min-width: 0;\n}" in css
    assert "body .wide-actions,\nbody .toolbar-actions," in css
    assert "background: linear-gradient(180deg, #15847b 0%, var(--phase21-accent-strong) 100%);" in css
    assert "box-shadow: var(--phase21-focus);" in css
    assert "cursor: progress;" in css
    assert "border-top-color: var(--phase21-accent);" in css

    for forbidden in [
        "tailwind",
        "alpine",
        "petite-vue",
        "react.production",
        "vue.global",
    ]:
        assert forbidden not in page.lower()
        assert forbidden not in js.lower()


def test_cr038_sticky_drawer_close_controls_are_shared_and_preserved():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    frontend_source = page + "\n" + css

    for drawer_id in [
        "job_drawer",
        "account_dialog",
        "proxy_drawer",
        "ai_profile_drawer",
        "mail_config_modal",
        "email_template_drawer",
        "run_log_drawer",
        "report_preview_drawer",
    ]:
        assert f'id="{drawer_id}"' in page

    for close_handler in [
        "closeJobDrawer()",
        "closeSocialAccountModal()",
        "closeProxyDrawer()",
        "closeAIProfileDrawer()",
        "closeMailConfigModal()",
        "closeEmailTemplateDrawer()",
        "closeRunLogDrawer()",
        "closeReportPreviewDrawer()",
    ]:
        assert f'onclick="{close_handler}" title="关闭"' in page

    for backdrop_id, close_handler in [
        ("job_drawer_backdrop", "closeJobDrawer()"),
        ("account_modal", "closeSocialAccountModal(event)"),
        ("proxy_drawer_backdrop", "closeProxyDrawer()"),
        ("ai_profile_drawer_backdrop", "closeAIProfileDrawer()"),
        ("mail_config_backdrop", "closeMailConfigModal()"),
        ("email_template_drawer_backdrop", "closeEmailTemplateDrawer()"),
        ("run_log_backdrop", "closeRunLogDrawer()"),
        ("report_preview_backdrop", "closeReportPreviewDrawer()"),
    ]:
        assert f'id="{backdrop_id}" class="drawer-backdrop" onclick="{close_handler}"' in page

    assert ".modal-head,\n.drawer-head {" in css
    sticky_header_block = css[css.index(".modal-head,\n.drawer-head {") : css.index(".modal-head h3,")]
    for marker in [
        "position: sticky;",
        "top: 0;",
        "z-index: 30;",
        "background: var(--color-neutral-0);",
        "border-bottom: 1px solid var(--color-neutral-200);",
        "box-shadow: 0 8px 18px rgba(15, 23, 42, 0.07);",
        "flex: 0 0 auto;",
        "margin: 0;",
        "padding: var(--drawer-padding-y) var(--drawer-padding-x) var(--space-3);",
        "border-radius: calc(var(--radius-modal-medium) - 1px) calc(var(--radius-modal-medium) - 1px) 0 0;",
    ]:
        assert marker in sticky_header_block

    assert "--drawer-padding-x: 18px;" in css
    assert "--drawer-padding-y: 18px;" in css
    assert "--drawer-padding-x: 14px;" in css
    assert "--drawer-padding-y: 14px;" in css
    assert "padding: var(--space-4) var(--drawer-padding-x) var(--drawer-padding-y);" in css
    assert "function normalizeDrawerScrollBodies(root=document)" in page
    assert "body.className='drawer-scroll-body';" in page
    assert ".modal-head > div,\n.drawer-head > div {\n  min-width: 0;" in css
    assert "flex: 0 0 34px;" in css
    assert "--z-floating-menu: 55;" in css
    assert "z-index: var(--z-floating-menu);" in css
    assert "z-index:45;" in page

    for fixed_footer in [
        ".form-actions { position:relative;",
        ".account-flow-actions { position:relative;",
        ".ai-test-actions { position:relative;",
        ".rule-modal-actions { position:relative;",
        ".resource-modal-actions { position:relative;",
    ]:
        assert fixed_footer in frontend_source

    for marker in [
        "const footerSelector='.form-actions, .resource-modal-actions, .account-flow-actions, .ai-test-actions, .rule-modal-actions';",
        "body.querySelectorAll(footerSelector).forEach(action=>{",
        "action.classList.add('drawer-fixed-footer');",
        "drawer.appendChild(action);",
        ".drawer > .drawer-fixed-footer {",
        "body .drawer > .drawer-fixed-footer {",
    ]:
        assert marker in frontend_source

    footer_block = css[css.index(".drawer > .drawer-fixed-footer {") : css.index(".action-menu-host,")]
    for marker in [
        "position: relative;",
        "flex: 0 0 auto;",
        "bottom: auto;",
        "top: auto;",
        "width: 100%;",
        "margin: 0;",
        "padding: 14px var(--drawer-padding-x);",
        "border-top: 1px solid var(--color-neutral-200);",
        "border-radius: 0 0 calc(var(--radius-modal-medium) - 1px) calc(var(--radius-modal-medium) - 1px);",
    ]:
        assert marker in footer_block
    assert "top: calc(var(--drawer-padding-y) + 80px);" not in css

    normalized_drawer_ids = [
        "job_drawer",
        "account_dialog",
        "proxy_drawer",
        "ai_connection_test_modal",
        "ai_profile_drawer",
        "ai_rule_modal",
        "mail_config_modal",
        "mail_test_modal",
        "email_template_drawer",
    ]
    footer_action_classes = [
        "form-actions",
        "resource-modal-actions",
        "account-flow-actions",
        "ai-test-actions",
        "rule-modal-actions",
    ]
    for drawer_id in normalized_drawer_ids:
        drawer_start = page.index(f'id="{drawer_id}"')
        next_drawer = page.find('class="drawer', drawer_start + 1)
        drawer_html = page[drawer_start : next_drawer if next_drawer != -1 else len(page)]
        footer_positions = [
            drawer_html.find(f'class="{action_class}')
            for action_class in footer_action_classes
            if f'class="{action_class}' in drawer_html
        ]
        if footer_positions:
            first_footer_position = min(footer_positions)
            assert 'class="drawer-scroll-body"' not in drawer_html[:first_footer_position]

    assert "document.addEventListener('keydown', event => {" in page
    assert "if(event.key === 'Escape'){" in page
    assert "closeAllOverlays();" in page


def test_cr073_scrollable_drawer_scrollbars_preserve_corner_radius():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")

    for drawer_id in [
        "job_drawer",
        "account_dialog",
        "proxy_drawer",
        "ai_profile_drawer",
        "ai_rule_modal",
        "mail_config_modal",
        "email_template_drawer",
        "run_detail_drawer",
        "run_log_drawer",
        "report_preview_drawer",
        "email_delivery_history_drawer",
    ]:
        assert f'id="{drawer_id}"' in page

    drawer_block = css[css.index(".drawer {") : css.index(".drawer-scroll-body {")]
    for marker in [
        "display: flex;",
        "flex-direction: column;",
        "padding: 0;",
        "overflow: hidden;",
        "border-radius: var(--radius-modal-medium);",
    ]:
        assert marker in drawer_block

    drawer_body_block = css[
        css.index(".drawer-scroll-body {") : css.index(".drawer-scroll-body::-webkit-scrollbar {")
    ]
    for marker in [
        "flex: 1 1 auto;",
        "min-height: 0;",
        "padding: var(--space-4) var(--drawer-padding-x) var(--drawer-padding-y);",
        "overflow: auto;",
        "scrollbar-color: rgba(100, 116, 139, 0.36) transparent;",
        "scrollbar-width: thin;",
    ]:
        assert marker in drawer_body_block

    scrollbar_block = css[
        css.index(".drawer-scroll-body::-webkit-scrollbar {") : css.index(".drawer.active {")
    ]
    for marker in [
        ".drawer-scroll-body::-webkit-scrollbar-track {",
        "margin-bottom: calc(var(--radius-modal-medium) + 2px);",
        "background: transparent;",
        ".drawer-scroll-body::-webkit-scrollbar-thumb {",
        "border: 3px solid transparent;",
        "border-radius: 999px;",
        "background-clip: content-box;",
        ".drawer-scroll-body::-webkit-scrollbar-corner {",
    ]:
        assert marker in scrollbar_block
    assert ".drawer::before {" not in css
    assert ".drawer::-webkit-scrollbar" not in css

    sticky_header_block = css[css.index(".modal-head,\n.drawer-head {") : css.index(".modal-head h3,")]
    assert "position: sticky;" in sticky_header_block
    assert "border-radius: calc(var(--radius-modal-medium) - 1px) calc(var(--radius-modal-medium) - 1px) 0 0;" in sticky_header_block

    for marker in [
        "function normalizeDrawerScrollBodies(root=document)",
        "root.querySelectorAll('.drawer').forEach(drawer=>{",
        "drawer.querySelector(':scope > .drawer-scroll-body')",
        "drawer.querySelector(':scope > .drawer-head, :scope > .modal-head')",
        "body.className='drawer-scroll-body';",
        "const topChromeSelector='.run-detail-toolbar, .detail-tabs.run-detail-tabs, .report-leads-toolbar, .drawer-actions';",
        "node.matches(topChromeSelector)",
        "node.classList.add('drawer-fixed-toolbar');",
        "body.querySelectorAll(`:scope > ${topChromeSelector.split(',').join(', :scope > ')}`).forEach(chrome=>{",
        "drawer.insertBefore(chrome, body);",
        "drawer.insertBefore(body, insertAfter.nextSibling);",
        "const footerSelector='.form-actions, .resource-modal-actions, .account-flow-actions, .ai-test-actions, .rule-modal-actions';",
        "body.querySelectorAll(footerSelector).forEach(action=>{",
        "action.classList.add('drawer-fixed-footer');",
        "normalizeDrawerScrollBodies();",
    ]:
        assert marker in page

    footer_block = css[css.index(".drawer > .drawer-fixed-footer {") : css.index(".action-menu-host,")]
    for marker in [
        "flex: 0 0 auto;",
        "bottom: auto;",
        "top: auto;",
        "margin: 0;",
        "z-index: 24;",
    ]:
        assert marker in footer_block

    modal_close_block = css[css.index(".modal-close {") : css.index(".modal-close:hover {")]
    assert "flex: 0 0 34px;" in modal_close_block
    assert "position: relative;" in modal_close_block
    assert "margin-left" not in modal_close_block
    assert "right:" not in modal_close_block


def test_cr082_drawer_scrollbar_boundaries_are_rechecked_on_open():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    frontend_source = page + "\n" + css
    inline_style = _monitor_inline_styles(page)

    for marker in [
        "function openDrawerChrome(backdropId, drawerId, options={})",
        "normalizeDrawerScrollBodies();",
        "drawer.classList.toggle('has-fixed-footer', !!drawer.querySelector(':scope > .drawer-fixed-footer'));",
        "drawer.classList.toggle('has-fixed-toolbar', !!drawer.querySelector(':scope > .drawer-fixed-toolbar'));",
        "body.setAttribute('data-scroll-owner', 'drawer-content');",
        "if(options.lockBody !== false) document.body.style.overflow='hidden';",
    ]:
        assert marker in page

    for opener in [
        "openDrawerChrome('account_modal', 'account_dialog')",
        "openDrawerChrome('proxy_drawer_backdrop', 'proxy_drawer')",
        "openDrawerChrome('job_drawer_backdrop', 'job_drawer')",
        "openDrawerChrome('ai_rule_modal_backdrop', 'ai_rule_modal')",
        "openDrawerChrome('ai_connection_test_backdrop', 'ai_connection_test_modal', {lockBody:false})",
        "openDrawerChrome('ai_profile_drawer_backdrop', 'ai_profile_drawer')",
        "openDrawerChrome('mail_config_backdrop', 'mail_config_modal')",
        "openDrawerChrome('mail_test_backdrop', 'mail_test_modal')",
        "openDrawerChrome('email_template_drawer_backdrop', 'email_template_drawer')",
        "openDrawerChrome('run_log_backdrop', 'run_log_drawer')",
        "openDrawerChrome('run_detail_backdrop', 'run_detail_drawer')",
        "openDrawerChrome('report_preview_backdrop', 'report_preview_drawer')",
        "openDrawerChrome('email_delivery_history_backdrop', 'email_delivery_history_drawer')",
    ]:
        assert opener in page

    for forbidden in [
        "document.getElementById('mail_config_backdrop').classList.add('active')",
        "document.getElementById('mail_config_modal').classList.add('active')",
        "document.getElementById('job_drawer_backdrop').classList.add('active')",
        "document.getElementById('run_detail_drawer').classList.add('active')",
    ]:
        assert forbidden not in page

    for marker in [
        "min-height: 0;",
        "overflow: hidden;",
        "isolation: isolate;",
        "position: relative;",
        "overflow-x: hidden;",
        "overflow-y: auto;",
        "overscroll-behavior: contain;",
        "scrollbar-gutter: stable;",
        ".drawer > .drawer-fixed-toolbar {",
        ".drawer.has-fixed-footer > .drawer-scroll-body {",
    ]:
        assert marker in frontend_source
        assert marker.replace(": ", ":").replace("; ", ";") in inline_style.replace(": ", ":").replace("; ", ";")

    for marker in [
        "flex: 0 0 auto;",
        "z-index: 25;",
        ".drawer > .drawer-fixed-toolbar.drawer-actions",
        ".drawer > .drawer-fixed-toolbar.run-detail-toolbar,",
        ".drawer > .drawer-fixed-toolbar.detail-tabs",
    ]:
        assert marker in frontend_source

    for chrome_class in [
        "run-detail-toolbar",
        "detail-tabs run-detail-tabs",
        "report-leads-toolbar",
        "drawer-actions",
    ]:
        assert chrome_class in page

    assert ".drawer::-webkit-scrollbar" not in css
    assert "top: calc(var(--drawer-padding-y) + 80px);" not in frontend_source


def test_phase_12a_navigation_groups_and_login_landing():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    inline_style = _monitor_inline_styles(page)

    for forbidden in [
        "nav-popover",
        "data-menu=",
        "toggleNavMenu",
        "closeNavMenus",
    ]:
        assert forbidden not in page
        assert forbidden not in css

    for marker in [
        'class="nav-group nav-group-collapsible"',
        'data-nav-group="resources"',
        'data-nav-group-toggle="resources"',
        'id="nav_group_resources"',
        'data-nav-group="settings"',
        'data-nav-group-toggle="settings"',
        'id="nav_group_settings"',
        'aria-expanded="true"',
        'class="nav-sublist"',
        'class="nav-caret"',
        'class="account-area"',
        'id="sidebar_collapse_button"',
        'class="sidebar-collapse-button"',
        'aria-label="收起侧边导航"',
        "function routeToOperationsHome()",
        "routeToOperationsHome();",
        "const NAV_GROUPS = {",
        "function navButtonLabel(button)",
        "function syncNavigationTooltips()",
        "function setSidebarCollapsed(collapsed)",
        "function syncResponsiveNavigationState()",
        "document.getElementById('sidebar_collapse_button')?.addEventListener('click'",
        "document.body.classList.toggle('sidebar-collapsed', !!collapsed)",
        "if(btn.dataset.navGroupToggle && document.body.classList.contains('sidebar-collapsed'))",
        "if(window.innerWidth < 768) collapsed=false",
        "if(window.innerWidth >= 768 && window.innerWidth < 1280)",
        "syncResponsiveNavigationState();",
        "function toggleNavGroup(group)",
        "function setNavGroupExpanded(group, expanded)",
        "function expandNavGroupForTab(tab)",
        "activateNavTab(dashboardButton, {skipLoad:true})",
        "activateNavTab(btn, options={})",
        "if(!options.skipLoad) loadSectionData(btn.dataset.tab)",
    ]:
        assert marker in page

    for selector in [
        ".nav-group-collapsible",
        ".nav-group-toggle",
        ".nav-caret",
        ".nav-sublist",
        ".nav-group.is-collapsed .nav-sublist",
        ".account-area",
        ".sidebar-collapse-button",
        "body.sidebar-collapsed .shell",
        "body.sidebar-collapsed nav button",
        "body.sidebar-collapsed nav button::after",
    ]:
        assert selector in css

    assert "--color-navigation-hover-bg: #edf1f5;" in css
    assert "--phase21-nav-hover: #edf1f5;" in css
    assert "body nav button:hover:not(.active) {\n  background: var(--phase21-nav-hover);" in css
    assert "body nav button.active {\n  background: var(--phase21-nav-active);" in css
    assert "color: var(--phase21-accent);" in css
    assert "--phase21-accent: #0f766e;" in css
    assert "--phase21-nav-active: #e8fbf6;" in css
    assert "nav button.active,\nnav button:hover" not in css
    assert ".nav-sublist button.active,\n.nav-sublist button:hover" not in css
    assert "nav button.active {\n  background: var(--color-navigation-active-bg);" in css
    assert "body .nav-sublist button.sub:hover:not(.active) {\n  background: var(--phase21-nav-hover);" in css
    assert "body .sidebar-collapse-button:hover {\n  color: var(--phase21-text);\n  background: var(--phase21-nav-hover);" in css
    collapsed_sidebar_hover_block = css[
        css.index("body.sidebar-collapsed .sidebar-collapse-button:hover {") : css.index(
            "body .mobile-nav-toggle:hover", css.index("body.sidebar-collapsed .sidebar-collapse-button:hover {")
        )
    ]
    assert "background: var(--phase21-nav-hover);" in collapsed_sidebar_hover_block
    assert "color: var(--phase21-text);" in collapsed_sidebar_hover_block
    assert "var(--phase21-accent-soft)" not in collapsed_sidebar_hover_block
    assert "var(--phase21-accent-strong)" not in collapsed_sidebar_hover_block
    assert ".nav-sublist button.active {\n  background: var(--color-navigation-active-bg);" in css
    nav_sublist_hover_block = page[
        page.index(".nav-sublist button:hover:not(.active) {") : page.index(
            ".nav-sublist button.active {"
        )
    ]
    nav_sublist_active_block = page[
        page.index(".nav-sublist button.active {") : page.index(
            "header {", page.index(".nav-sublist button.active {")
        )
    ]
    assert "background:#edf1f5;" in nav_sublist_hover_block
    assert "color:var(--text);" in nav_sublist_hover_block
    assert "background:#e8fbf6;" not in nav_sublist_hover_block
    assert "color:#0f766e;" not in nav_sublist_hover_block
    assert "background:#e8fbf6;" in nav_sublist_active_block
    assert "color:#0f766e;" in nav_sublist_active_block
    assert ".nav-popover" not in inline_style
    assert 'id="mobile_account_area"' not in page
    assert 'id="mobile_current_user_badge"' not in page
    assert '<button class="secondary" onclick="logout()">退出</button>' not in page


def test_phase_21_header_account_menu_matches_compact_user_control():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    header = page[page.index("<header>") : page.index("</header>", page.index("<header>"))]

    assert 'id="account_menu_button"' in header
    assert 'class="account-menu-button"' in header
    assert 'aria-haspopup="menu"' in header
    assert 'aria-controls="account_menu"' in header
    assert 'id="account_menu"' in header
    assert "刷新当前页" in header
    assert "返回运营首页" in header
    assert "退出登录" in header
    assert 'onclick="logoutFromAccountMenu()"' in header
    assert '<button class="secondary" onclick="logout()">退出</button>' not in header

    for snippet in [
        "function toggleSessionMenu(event)",
        "function closeSessionMenu()",
        "function refreshFromAccountMenu()",
        "function returnToOperationsHomeFromAccountMenu()",
        "function logoutFromAccountMenu()",
        "document.getElementById('account_menu_button')?.addEventListener('click', toggleSessionMenu)",
        "if(!event.target.closest('.account-area'))",
    ]:
        assert snippet in page

    for selector in [
        ".account-menu-button",
        ".account-avatar",
        ".account-menu-copy",
        ".account-menu",
        ".account-menu.active",
        ".account-menu-summary",
        ".account-menu .account-menu-danger",
    ]:
        assert selector in css


def test_phase_21b_navigation_hierarchy_keeps_task_loop_and_subpages_distinct():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")

    nav = page[page.index('<nav id="primary_navigation">') : page.index("</nav>", page.index('<nav id="primary_navigation">'))]

    for first_level in [
        'data-tab="dashboard" data-menu-key="overview"',
        'data-tab="jobs" data-menu-key="monitoring"',
        'data-tab="runs" data-menu-key="run_center"',
    ]:
        assert first_level in nav

    for group_marker in [
        'data-nav-group="resources"',
        'data-nav-group-toggle="resources"',
        'id="nav_group_resources"',
        'data-nav-group="settings"',
        'data-nav-group-toggle="settings"',
        'id="nav_group_settings"',
    ]:
        assert group_marker in nav

    for sub_marker in [
        '<button class="sub" data-tab="accounts"',
        '<button class="sub" data-tab="proxies"',
        '<button class="sub" data-tab="ai"',
        '<button class="sub" data-tab="ai_rules"',
        '<button class="sub" data-tab="email"',
        '<button class="sub" data-tab="email_templates"',
        '<button class="sub" data-tab="runtime"',
        '<button class="sub" data-tab="doctor"',
    ]:
        assert sub_marker in nav

    for icon_marker in [
        'href="#icon-dashboard"',
        'href="#icon-monitor"',
        'href="#icon-run"',
        'href="#icon-resource"',
        'href="#icon-account"',
        'href="#icon-proxy"',
        'href="#icon-ai"',
        'href="#icon-settings"',
        'href="#icon-rules"',
        'href="#icon-mail"',
        'href="#icon-template"',
        'href="#icon-runtime"',
        'href="#icon-doctor"',
    ]:
        assert icon_marker in nav

    for css_marker in [
        "body #primary_navigation > .nav-group:not(.nav-group-collapsible) > button {\n  min-height: 44px;",
        "body .nav-group-toggle {\n  min-height: 40px;",
        "body .nav-sublist {\n  position: relative;",
        "body .nav-sublist::before {",
        "background: rgba(148, 163, 184, 0.28);",
        "body .nav-sublist button.sub {\n  position: relative;",
        "font-weight: 560;",
        "body .nav-group-toggle.active::before {\n  display: none;",
        "body.sidebar-collapsed .nav-sublist {\n  display: grid !important;",
    ]:
        assert css_marker in css

    assert "body nav button:hover:not(.active)" in css
    assert "body .nav-sublist button.sub:hover:not(.active)" in css
    assert "var(--phase21-nav-hover)" in css
    assert "var(--phase21-nav-active)" in css
    assert "body.sidebar-collapsed nav button::after" in css

    tablet_block = css.split("@media (min-width: 768px) and (max-width: 1279px)", 1)[1].split(
        "@media (max-width: 767px)", 1
    )[0]
    tablet_nav_block = tablet_block
    for tablet_marker in [
        "body.sidebar-collapsed nav {\n    gap: 4px;",
        "min-height: 0;",
        "max-height: 100%;",
        "overflow-y: auto;",
        "overscroll-behavior: contain;",
        "body.sidebar-collapsed nav button,\n  body.sidebar-collapsed nav button.sub,\n  body.sidebar-collapsed .nav-group-toggle {\n    width: 42px;\n    height: 36px;",
        "overflow-x: hidden;",
        "body.sidebar-collapsed nav button::after,\n  body.sidebar-collapsed .sidebar-collapse-button::after {\n    content: none;",
        "body.sidebar-collapsed .sidebar-collapse-button {\n    width: 42px;\n    height: 38px;",
    ]:
        assert tablet_marker in tablet_block

    assert "body.sidebar-collapsed .shell > aside {" in tablet_block
    assert "grid-template-rows: auto minmax(0, 1fr) auto;" in tablet_block
    assert "body .shell > aside {\n    position: fixed;" in tablet_nav_block
    assert "top: 0;" in tablet_nav_block
    assert "inset: 0 auto 0 0;" in tablet_nav_block
    assert "width: 272px;" in tablet_nav_block
    assert "body.sidebar-collapsed .shell > aside {\n    width: 68px;" in tablet_nav_block


def test_phase_21b_mobile_header_and_tablet_side_rail_resilience():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    inline_style = _monitor_inline_styles(page)

    mobile_nav_block = css.split("@media (max-width: 767px)", 1)[1]
    inline_mobile_block = inline_style.split("@media (max-width:767px)", 1)[1]
    inline_tablet_nav_block = inline_style.split("@media (min-width:768px) and (max-width:1279px)", 1)[1].split(
        "@media (max-width:767px)", 1
    )[0]
    tablet_nav_block = css.split("@media (min-width: 768px) and (max-width: 1279px)", 1)[1].split(
        "@media (max-width: 767px)", 1
    )[0]

    for source in [mobile_nav_block, inline_mobile_block]:
        compact_source = source.replace(" ", "")
        assert "grid-template-columns:40pxminmax(0,1fr)36px44px;" in compact_source
        assert '"nav title refresh account"' in source
        assert '"status status status status"' in source
        assert "bodyheader.mobile-nav-toggle>span:not(.mobile-nav-icon)" in compact_source
        assert "display:none;" in compact_source
        assert "display:block;" in compact_source
        assert "writing-mode:horizontal-tb;" in compact_source
        assert "white-space:nowrap;" in compact_source
        assert "word-break:keep-all;" in compact_source
        assert "display:contents!important;" in compact_source
        assert "width:auto!important;" in compact_source
        assert "min-width:0!important;" in compact_source
        top_status_compact = compact_source.split("bodyheader#top_status{", 1)[1].split("}", 1)[0]
        assert "overflow:visible;" in top_status_compact
        assert "flex-wrap:wrap;" in top_status_compact

    for source in [mobile_nav_block]:
        assert "body:not(.mobile-nav-open) #primary_sidebar" in source
        assert "left: -340px !important;" in source
        assert "body .table-wrap" in source
        assert "overflow-x: auto;" in source
        assert "max-width: 100%;" in source

    assert "grid-template-rows: auto minmax(0, 1fr) auto;" in tablet_nav_block
    assert "min-height: 0;" in tablet_nav_block
    assert "overflow-y: auto;" in tablet_nav_block
    assert "overflow-x: hidden;" in tablet_nav_block
    assert "overscroll-behavior: contain;" in tablet_nav_block
    assert "body.sidebar-collapsed .shell > aside {\n    width: 68px;" in tablet_nav_block
    assert ".sidebar-collapse-button {\n    display: flex !important;" in tablet_nav_block
    assert "body.sidebar-collapsed nav button::after,\n  body.sidebar-collapsed .sidebar-collapse-button::after {\n    content: none;" in tablet_nav_block

    inline_tablet_compact = inline_tablet_nav_block.replace(" ", "")
    assert "body.sidebar-collapsed.shell{grid-template-columns:68pxminmax(0,1fr)!important;" in inline_tablet_compact
    assert "body.sidebar-collapsed#primary_sidebar" in inline_tablet_compact
    assert "width:68px!important;" in inline_tablet_compact
    assert "max-width:68px!important;" in inline_tablet_compact
    assert "bodyheadermobile-nav-toggle" not in inline_tablet_compact


def test_phase_21d_monitoring_tasks_and_task_drawer_visual_pass_preserves_workflow():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    jobs_section = _monitor_section(page, "jobs")
    job_drawer = page[page.index('id="job_drawer"') : page.index('<section id="accounts">')]

    for marker in [
        "openNewJobDrawer()",
        "refreshJobSchedule(this)",
        'data-shortcut-tab="runs" data-shortcut-grouped="1"',
        'id="job_search"',
        'id="job_platform_filter"',
        'id="job_status_filter"',
        "clearJobFilters()",
        'id="jobs_table"',
        "renderJobsTable()",
        "jobActions",
        "toggleJobActionMenu",
        "/jobs/'+id+'/run",
        "/jobs/'+id+'/stop",
        "toggleJob",
        "/jobs/'+id+(enabled?'/pause':'/resume')",
        "deleteJob",
        'id="job_drawer_backdrop" class="drawer-backdrop" onclick="closeJobDrawer()"',
    ]:
        assert marker in page

    for marker in [
        'id="job_drawer" class="drawer form-drawer"',
        'id="job_form" class="panel"',
        'id="normal_task_wizard_hint"',
        'id="normal_task_wizard_steps"',
        'class="wizard-step"',
        'class="config-section wizard-section" data-wizard-step="target"',
        'id="law_firm_name"',
        'id="aliases"',
        'data-wizard-step="content"',
        'id="keywords"',
        'name="platform" value="dy"',
        'name="platform" value="ks"',
        'name="platform" value="xhs"',
        'id="job_max_items"',
        'id="job_start_page"',
        'id="job_max_pages"',
        'id="enable_comments"',
        'id="enable_sub_comments"',
        'id="exclude_words"',
        'class="config-section admin-advanced-section admin-only-job-field"',
        'id="job_account_id"',
        'id="job_proxy_id"',
        'id="job_target_type"',
        'id="job_output_mode"',
        'id="job_browser_mode"',
        'data-wizard-step="schedule"',
        'id="time_window_type"',
        'id="frequency"',
        'id="email_time"',
        'id="custom_start" type="date"',
        'id="custom_end" type="date"',
        'id="cron_expr"',
        'id="enabled"',
        'id="job_ai_profile_id"',
        '使用系统配置中的 AI 评估规则',
        'data-wizard-step="report"',
        'id="recipients"',
        'id="job_email_template_id"',
        'id="save_job_btn"',
        'id="fill_sample_job_btn"',
        'id="reset_job_btn"',
        'onclick="closeJobDrawer()">关闭</button>',
    ]:
        assert marker in job_drawer

    for marker in [
        "applyJobFormRoleMode()",
        "document.querySelectorAll('.admin-only-job-field')",
        "document.querySelectorAll('.normal-only')",
        "refreshEnhancedFilterSelects(document.getElementById('job_drawer'))",
        "refreshEnhancedFilterDateInputs(document.getElementById('job_drawer'))",
        "toast('任务已保存'); resetJobForm(); closeJobDrawer(); await Promise.all([loadJobs(), loadDashboard(), loadDoctor()]);",
        "fillSampleJobTemplate",
        "clearJobFilters",
        "resetJobForm",
        "closeJobDrawer()",
    ]:
        assert marker in page

    for selector in [
        "body #jobs > .panel {",
        "body #jobs .toolbar {",
        "body #job_drawer .panel {",
        "body #job_drawer textarea {",
        "body .wizard-steps {",
        "body .wizard-step {",
        "body .config-section h3::before {",
        "body .admin-advanced-section {",
        "body .admin-advanced-section h3::before {",
        "body .form-actions,",
        "body #jobs_table .table-wrap table {",
    ]:
        assert selector in css

    assert "min-width: 980px;" in css
    assert "采集规则说明" not in jobs_section
def test_phase_21e_platform_accounts_visual_pass_preserves_account_workflow():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    accounts_section = _monitor_section(page, "accounts")
    account_dialog = page[page.index('id="account_dialog"') : page.index('<section id="proxies">')]

    for marker in [
        'data-page-entry="accounts"',
        'onclick="openNewSocialAccountModal()">新增账号',
        'data-shortcut-tab="jobs" data-menu-key="monitoring"',
        'id="account_metrics"',
        'id="account_bulk_toolbar"',
        'id="account_bulk_count"',
        'id="account_bulk_check_btn"',
        'id="account_bulk_disable_btn"',
        'id="account_bulk_enable_btn"',
        'id="account_bulk_delete_btn"',
        'id="account_search"',
        'id="account_platform_filter"',
        'id="account_status_filter"',
        'id="account_login_type_filter"',
        'id="account_attention_filter"',
        'id="account_bulk_bar"',
        'id="social_accounts_table"',
        "toggleAllFilteredAccounts",
        "toggleAccountSelection",
        "toggleAccountActionMenu",
        "checkSelectedSocialAccountLogins",
        "setSelectedSocialAccountStatus",
        "deleteSelectedSocialAccounts",
        "accountLedgerTable",
        "startLoginSessionForAccount",
        "reloginSocialAccount",
        "deleteSocialAccount",
    ]:
        assert marker in page

    for marker in [
        'id="account_modal" class="drawer-backdrop" onclick="closeSocialAccountModal(event)"',
        'id="account_dialog" class="drawer form-drawer account-detail-card"',
        'id="account_detail_title"',
        'onclick="closeSocialAccountModal()"',
        'id="social_account_id"',
        'id="account_detail_empty"',
        'id="account_detail_summary"',
        'class="account-modal-grid"',
        'class="account-section"',
        "1. 基础资料",
        'id="social_account_name"',
        'id="social_account_platform"',
        'id="social_account_status"',
        'id="social_account_proxy_id"',
        'id="social_account_login_source"',
        'id="social_account_notes"',
        'id="social_account_error_summary"',
        'id="social_account_last_error"',
        "2. 登录维护",
        'id="login_status_badge"',
        'id="social_account_login_type"',
        'id="social_login_method_options"',
        'id="login_panel_qrcode"',
        'id="account_qrcode_button"',
        'id="account_local_login_button"',
        'id="login_session_result"',
        'id="login_panel_cookie"',
        'id="social_account_cookie_input"',
        'id="social_account_clear_cookie"',
        'id="account_cookie_save_button"',
        'id="cookie_login_result"',
        "3. 登录记录",
        'id="current_account_login_sessions"',
        'id="account_modal_actions"',
        "4. 完成账号设置",
        'id="account_save_button"',
        'id="account_delete_button"',
    ]:
        assert marker in account_dialog or marker in page

    for marker in [
        "handleSocialLoginTypeChange",
        "selectSocialLoginType",
        "supportedSocialLoginTypes",
        "renderLoginModePanel",
        "startLoginSessionFromForm",
        "openCurrentAccountLoginBrowser",
        "saveCurrentPlatformCookieLogin",
        "loadLoginSessions",
        "renderLoginSessionResult",
        "updateAccountLoginPrerequisites",
        "updateAccountModalActions",
        "saveSocialAccount",
        "deleteCurrentSocialAccount",
        "renderProxySelectOptions",
    ]:
        assert marker in page

    for selector in [
        "body #accounts .account-list-panel {",
        "body #accounts .account-ledger-head {",
        "body #accounts .account-bulk-toolbar {",
        "body #accounts .account-bulk-count {",
        "body #accounts .account-filter {",
        "body #accounts .account-quick-filter {",
        "body #account_dialog.account-detail-card {",
        "body #account_dialog .drawer-scroll-body {",
        "body .account-detail-empty {",
        "body .account-detail-summary div {",
        "body .account-modal-grid {",
        "body .account-section {",
        "body .login-card {",
        "body .login-card h3 {",
        "body .login-actions {",
        "body #account_modal_actions.account-flow-actions {",
    ]:
        assert selector in css

    account_scroll_block = css[
        css.index("body #account_dialog .drawer-scroll-body {") : css.index(
            "body .account-detail-empty {", css.index("body #account_dialog .drawer-scroll-body {")
        )
    ]
    assert "overflow-x: hidden;" in account_scroll_block
    for marker in [
        "body #account_dialog .account-modal-grid,",
        "body #account_dialog .account-section,",
        "body #account_dialog .login-card-grid,",
        "body #account_dialog .login-actions,",
        "max-width: 100%;",
        "min-width: 0;",
    ]:
        assert marker in account_scroll_block
    assert "body #account_dialog > .drawer-fixed-footer," in css
    assert "body #account_dialog .resource-modal-actions {" not in account_scroll_block

    for forbidden in [
        "手机号登录",
        "social_account_phone",
        "saveCurrentPlatformPhoneLogin",
        "startLoginSessionFromSelected",
        "openSelectedAccountLoginBrowser",
    ]:
        assert forbidden not in accounts_section


def test_phase_21f_proxy_resources_visual_pass_preserves_proxy_workflow():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    proxies_section = _monitor_section(page, "proxies")
    proxy_drawer = page[page.index('id="proxy_drawer"') : page.index('<section id="ai">')]
    header = page[page.index("<header>") : page.index("</header>", page.index("<header>"))]

    for marker in [
        'data-page-entry="proxies"',
        'onclick="openProxyDrawer()">新增代理',
        'data-shortcut-tab="accounts" data-menu-key="platform_accounts"',
        'id="proxy_resource_summary"',
        'id="proxy_resource_count"',
        'id="proxy_search"',
        'id="proxy_status_filter"',
        "clearProxyFilters()",
        'id="proxy_profiles_table"',
        "loadProxyPool",
        "renderProxyResourceSummary",
        "filteredProxyProfiles",
        "renderProxyProfilesTable",
        "proxyIssueCell",
        "editProxyProfile",
        "deleteProxyProfile",
    ]:
        assert marker in page

    for marker in [
        'id="proxy_drawer_backdrop" class="drawer-backdrop" onclick="closeProxyDrawer()"',
        'id="proxy_drawer" class="drawer form-drawer"',
        'id="proxy_drawer_title"',
        'onclick="closeProxyDrawer()"',
        'id="proxy_id"',
        'id="proxy_name"',
        'id="proxy_provider"',
        'id="proxy_status"',
        'id="proxy_url" type="password"',
        'id="proxy_max_concurrency"',
        'id="proxy_notes"',
        'id="proxy_last_error"',
        "resetProxyForm()",
        'id="proxy_save_button"',
        "saveProxyProfile()",
        "refreshEnhancedFilterSelects(document.getElementById('proxy_drawer'))",
    ]:
        assert marker in proxy_drawer or marker in page

    for selector in [
        "body #proxies > .panel {",
        "body #proxies .account-ledger-head {",
        "body #proxies .resource-toolbar {",
        "body #proxy_profiles_table .proxy-resource-table table {",
        "body .proxy-mask-code {",
        "body .proxy-issue-cell {",
        "body .proxy-error-line {",
        "body #proxy_drawer {",
        "body #proxy_drawer .drawer-scroll-body {",
        "body #proxy_drawer .resource-modal-actions {",
        "body #proxy_drawer > .drawer-fixed-footer,",
    ]:
        assert selector in css

    assert "proxy-resource-table" in page
    assert "proxy-mask-code code" in page
    assert "min-width: 980px;" in css
    assert "overflow-x: hidden;" in css[
        css.index("body #proxy_drawer .drawer-scroll-body {") : css.index(
            "body #proxy_drawer .panel {", css.index("body #proxy_drawer .drawer-scroll-body {")
        )
    ]

    assert 'id="global_refresh_button"' in header
    assert "refreshActiveSection(this)" in header
    for duplicate_refresh in [
        ">刷新代理</button>",
        "refreshProxy",
        "onclick=\"loadProxyPool()\"",
    ]:
        assert duplicate_refresh not in proxies_section


def test_phase_21g_ai_access_visual_pass_preserves_ai_workflow():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    ai_section = _monitor_section(page, "ai")
    ai_drawer = page[page.index('id="ai_profile_drawer"') : page.index('id="ai_rule_modal_backdrop"')]
    ai_test_modal = page[page.index('id="ai_connection_test_modal"') : page.index('id="ai_profile_drawer_backdrop"')]
    header = page[page.index("<header>") : page.index("</header>", page.index("<header>"))]

    for marker in [
        'data-page-entry="ai"',
        'onclick="openAIProfileDrawer()">新增 AI 接入',
        'data-shortcut-tab="ai_rules" data-menu-key="ai_rules"',
        'id="ai_resource_summary"',
        'id="ai_resource_count"',
        'id="ai_profile_search"',
        'id="ai_profile_provider_filter"',
        'id="ai_profile_test_filter"',
        "clearAIProfileFilters()",
        'id="ai_profiles_table"',
        "loadAIProfiles",
        "renderAIResourceSummary",
        "filteredAIProfiles",
        "renderAIProfilesTable",
        "editAIProfile",
        "testAIProfile",
        "activateAIProfile",
        "deleteAIProfile",
    ]:
        assert marker in page

    for marker in [
        'id="ai_profile_drawer_backdrop" class="drawer-backdrop" onclick="closeAIProfileDrawer()"',
        'id="ai_profile_drawer" class="drawer form-drawer"',
        'id="ai_profile_drawer_title"',
        'onclick="closeAIProfileDrawer()"',
        'id="ai_profile_id"',
        'id="ai_profile_name"',
        'id="ai_profile_provider"',
        'id="ai_profile_model"',
        'class="secondary combo-trigger model-load-trigger"',
        'id="ai_profile_model_options" class="model-options"',
        'id="ai_profile_base_url"',
        'id="ai_profile_api_key" type="password"',
        'id="ai_profile_model_load_btn"',
        'id="ai_profile_temperature"',
        'id="ai_profile_active"',
        "resetAIProfileForm()",
        'id="ai_profile_save_button"',
        "saveAIProfile()",
        "loadAIProfileModels(this)",
        "toggleAIProfileModelOptions",
        "selectAIProfileModel",
            "const drawer=openDrawerChrome('ai_profile_drawer_backdrop', 'ai_profile_drawer');",
            "refreshEnhancedFilterSelects(drawer)",
    ]:
        assert marker in ai_drawer or marker in page

    for marker in [
        'id="ai_connection_test_modal" class="drawer ai-test-modal"',
        'onclick="closeAIConnectionTestModal()"',
        'id="ai_test_profile_name"',
        'id="ai_test_profile_meta"',
        'id="ai_test_status_badge"',
        'id="ai_test_model"',
        'id="ai_test_provider"',
        'id="ai_test_message"',
        'id="ai_test_console"',
        'id="ai_test_start_btn"',
        "runAIConnectionTest()",
        "ai-profiles/'+id+'/connection-test",
    ]:
        assert marker in ai_test_modal or marker in page

    for selector in [
        "body #ai > .panel {",
        "body #ai .account-ledger-head {",
        "body #ai .resource-toolbar {",
        "body #ai_profiles_table .ai-resource-table table {",
        "body #ai_profiles_table .ai-resource-table th:nth-child(4),",
        "body #ai_profiles_table .ai-resource-table th:nth-child(5),",
        "body .ai-model-chip,",
        "body .ai-base-url {",
        "body .ai-test-state {",
        "body #ai_profile_drawer,",
        "body #ai_profile_drawer .drawer-scroll-body,",
        "body #ai_profile_drawer .panel {",
        "body #ai_connection_test_modal .ai-test-meta {",
        "body #ai_profile_drawer > .drawer-fixed-footer,",
    ]:
        assert selector in css

    assert "ai-resource-table" in page
    assert "ai-model-chip" in page
    assert "ai-base-url code" in page
    assert "ai-key-mask" in page
    assert "ai-test-state" in page
    assert "min-width: 1560px;" in css
    assert "table-layout: fixed;" in css
    assert "text-overflow: ellipsis;" in css
    ai_chip_block = css[
        css.index("body .ai-model-chip,") : css.index(
            "body .ai-key-mask {", css.index("body .ai-key-mask {") + 1
        )
    ]
    assert "white-space: nowrap;" in ai_chip_block
    assert "overflow: hidden;" in ai_chip_block
    assert "body #ai_profile_drawer .model-combobox {" in css[
        css.index("@media (max-width: 767px)") :
    ]
    assert "overflow-x: hidden;" in css[
        css.index("body #ai_profile_drawer .drawer-scroll-body,") : css.index(
            "body #ai_profile_drawer .panel {", css.index("body #ai_profile_drawer .drawer-scroll-body,")
        )
    ]

    assert 'id="global_refresh_button"' in header
    assert "refreshActiveSection(this)" in header
    for duplicate_refresh in [
        ">刷新 AI 接入</button>",
        "refreshAIProfile",
        "onclick=\"loadAIProfiles()\"",
    ]:
        assert duplicate_refresh not in ai_section

    assert '<select id="ai_profile_model"' not in ai_drawer
    assert 'id="ai_profile_model_toggle"' not in ai_drawer


def test_phase_21h_ai_evaluation_rules_visual_pass_preserves_rule_workflow():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    ai_rules_section = _monitor_section(page, "ai_rules")
    ai_rule_modal = page[page.index('id="ai_rule_modal"') : page.index('id="email"')]
    header = page[page.index("<header>") : page.index("</header>", page.index("<header>"))]

    for marker in [
        'data-page-entry="ai_rules"',
        'onclick="openNewAIRuleModal()">新增评估规则',
        'data-shortcut-tab="ai" data-menu-key="ai_access"',
        'class="panel ai-rule-panel"',
        'class="account-ledger-head ai-rule-ledger-head"',
        'id="ai_rules_summary" class="account-bulk-count ai-rule-summary"',
        'id="ai_rules_table"',
        "loadAIRuleProfiles",
        "renderAIRuleProfiles",
        "aiRuleActions",
        "toggleAIRuleActionMenu",
        "testAIRuleProfile",
        "activateAIRuleProfile",
        "deleteAIRuleProfile",
        "aiRuleTestStateCell",
    ]:
        assert marker in page

    for marker in [
        'id="ai_rule_modal_backdrop" class="drawer-backdrop" onclick="closeAIRuleModal()"',
        'id="ai_rule_modal" class="drawer template-modal"',
        'id="ai_rule_modal_title"',
        'onclick="closeAIRuleModal()"',
        'class="rule-section-card rule-section-card-basic"',
        'class="rule-section-card rule-section-card-config"',
        'class="rule-section-card rule-section-card-schema"',
        'class="rule-section-card rule-section-card-sample"',
        'class="rule-section-card rule-section-card-result"',
        'id="ai_rule_name"',
        'id="ai_rule_role"',
        'id="ai_rule_relevance"',
        'id="ai_rule_negative"',
        'id="ai_rule_risk"',
        'id="ai_rule_evidence"',
        'id="ai_rule_action"',
        'id="ai_prompt_preview"',
        'id="ai_output_schema_table"',
        'id="ai_sample_law_firm_name"',
        'id="ai_sample_platform"',
        'id="ai_sample_source_keyword"',
        'id="ai_sample_title"',
        'id="ai_sample_text"',
        'id="ai_sample_comments"',
        'id="ai_rules_result" class="resultbox ai-rule-result is-idle"',
        'id="ai_rule_test_button"',
        'onclick="testAIRuleFromModal()"',
        "resetAIPrompt()",
        'id="ai_rule_save_button"',
        "saveAIRuleFromModal()",
    ]:
        assert marker in ai_rule_modal or marker in page

    for marker in [
        "测试规则",
        "设为默认",
        "删除规则",
        "基础信息",
        "规则配置",
        "固定输出字段",
        "测试样例",
        "测试结果",
        "角色定位",
        "相关性判断",
        "疑似负面判断",
        "风险等级规则",
        "证据摘录规则",
        "处理建议规则",
        "生成后的 Prompt 预览",
        "恢复默认规则",
        "保存规则",
        "关闭",
    ]:
        assert marker in page

    for selector in [
        "body #ai_rules .ai-rule-panel {",
        "body #ai_rules .ai-rule-summary {",
        "body #ai_rules_table .ai-rule-table table {",
        "body #ai_rules_table .ai-rule-table th:nth-child(6),",
        "body #ai_rules_table .ai-rule-table td.col-actions .wide-actions {",
        "body .ai-rule-name-cell,",
        "body .ai-rule-test-cell {",
        "body .ai-rule-test-cell.is-ok {",
        "body .ai-rule-test-cell.is-bad {",
        "body .ai-rule-status-chip {",
        "body .rule-modal-flow {",
        "body .rule-editor-stack,",
        "body .rule-test-stack {",
        "body .rule-section-card {",
        "body .rule-accordion {",
        "body .rule-accordion summary:focus-visible {",
        "body .rule-section-card-config .advanced {",
        "body .prompt-preview {",
        "body .ai-rule-result {",
        "body .ai-rule-result.is-loading {",
        "body .ai-rule-result.is-ok {",
        "body .ai-rule-result.is-bad {",
    ]:
        assert selector in css

    assert "ai-rule-table" in page
    assert "table-layout: fixed;" in css
    assert "min-width: 920px !important;" in css
    assert "width: 120px;" in css
    assert "min-width: 120px;" in css
    assert "min-width: 112px;" in css
    assert "grid-template-columns: minmax(0, 1.16fr) minmax(360px, 0.84fr);" in css
    assert "body .rule-modal-flow {\n    grid-template-columns: minmax(0, 1fr);" in css[
        css.index("@media (max-width: 1279px)") :
    ]
    assert "body .rule-test-stack {\n    grid-template-columns: minmax(0, 1fr);" in css[
        css.index("@media (max-width: 767px)") :
    ]
    assert "body .rule-accordion textarea {\n  margin: 0 10px 9px;" in css
    assert "body .ai-rule-result {\n  min-height: 204px;" in css
    assert "body .prompt-preview {\n  max-height: 260px;" in css
    assert "body .rule-section-card-sample .grid {\n  grid-template-columns: minmax(0, 1fr) 180px;" in css
    assert "body .modal-select-compact {\n  min-width: 0;" in css

    assert 'id="global_refresh_button"' in header
    assert "refreshActiveSection(this)" in header
    for duplicate_refresh in [
        ">刷新规则</button>",
        "refreshAIRules",
        "onclick=\"loadAIRuleProfiles()\"",
    ]:
        assert duplicate_refresh not in ai_rules_section

    assert "rule-modal-layout" not in page
    assert "rule-side-stack" not in page


def test_phase_21i_mail_configuration_visual_pass_preserves_delivery_workflow():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    email_section = page[page.index('<section id="email"') : page.index('<div id="mail_config_backdrop"')]
    mail_config_modal = page[page.index('<div id="mail_config_backdrop"') : page.index('<div id="mail_test_backdrop"')]
    mail_test_modal = page[page.index('<div id="mail_test_backdrop"') : page.index('<section id="email_templates">')]
    header = page[page.index("<header>") : page.index("</header>", page.index("<header>"))]

    for marker in [
        'data-page-entry="email"',
        "邮件配置",
        "onclick=\"openMailConfigModal()\"",
        "onclick=\"openMailTestModal()\"",
        'data-shortcut-tab="runs" data-shortcut-grouped="1"',
        'id="real_email_delivery_toggle"',
        "setRealEmailDelivery(this.checked)",
        "真实邮件：已关闭",
        "email-toolbar-switch",
        "mail-config-summary-grid",
        'id="email_test_status_card"',
        "mail-config-stat",
        'id="email_status_label"',
        'id="email_sender_summary"',
        'id="email_recipient_summary"',
        "0 个兜底收件人",
        'id="email_subject_summary"',
    ]:
        assert marker in email_section

    assert email_section.count("openMailConfigModal()") == 1
    assert email_section.count("openMailTestModal()") == 1
    assert "SMTP 与发送默认值" not in email_section
    assert "email_validation_status" not in email_section

    for marker in [
        'id="mail_config_backdrop" class="drawer-backdrop" onclick="closeMailConfigModal()"',
        'id="mail_config_modal" class="drawer form-drawer"',
        'aria-label="邮件配置"',
        "编辑邮件配置",
        "mail-config-panel",
        "mail-config-card",
        "SMTP 连接",
        'id="smtp_host"',
        'id="smtp_port"',
        'id="encryption"',
        'id="sender"',
        'id="smtp_username"',
        'id="smtp_password"',
        'id="smtp_password_status"',
        "发送默认值",
        'id="default_recipients"',
        'id="subject_template"',
        "取消",
        'id="mail_config_save_button"',
        "保存配置",
        "closeMailConfigModal()",
        "saveEmail()",
    ]:
        assert marker in mail_config_modal

    for marker in [
        'id="mail_test_backdrop" class="drawer-backdrop" onclick="closeMailTestModal()"',
        'id="mail_test_modal" class="drawer mail-test-modal"',
        'aria-label="邮件测试"',
        "发送测试邮件",
        'id="mail_test_console"',
        'id="mail_test_start_btn"',
        "开始测试",
        "关闭",
        "testEmail()",
    ]:
        assert marker in mail_test_modal

    for marker in [
        "const statusCard=document.getElementById('email_test_status_card');",
        "statusCard.classList.remove('is-ok','is-bad','is-idle');",
        "个兜底收件人",
        "renderEmailValidationWindow(emailValidationWindowState)",
        "refreshEnhancedFilterSelects(drawer)",
        "openDrawerChrome('mail_config_backdrop', 'mail_config_modal')",
        "openDrawerChrome('mail_test_backdrop', 'mail_test_modal')",
        "api('/email-config'",
        "api('/email-config/test'",
        "请先开启真实邮件发送",
        "SMTP 已接受不代表收件箱已收到，仍需人工确认收件箱或垃圾箱。",
        "clearActionButtonLoading(btn, '重新测试')",
    ]:
        assert marker in page

    for selector in [
        "body #email .page-head {",
        "body #email .email-toolbar-switch {",
        "body #email .email-toolbar-switch:has(input:checked) {",
        "body .mail-config-summary-grid {",
        "body .mail-config-stat {",
        "body .mail-config-stat::after {",
        "body .mail-config-stat.is-ok::after {",
        "body .mail-config-stat.is-bad::after {",
        "body .mail-config-stat.is-idle::after {",
        "body #email_subject_summary {",
        "body .mail-config-panel {",
        "body .mail-config-card {",
        "body .mail-default-card textarea {",
        "body #smtp_password_status {",
        "body #mail_test_modal .drawer-scroll-body {",
        "body #mail_test_modal .test-console {",
    ]:
        assert selector in css

    assert "font-family: var(--font-family-sans);" in css
    assert "overflow-wrap: anywhere;" in css
    assert 'id="global_refresh_button"' in header
    assert "refreshActiveSection(this)" in header
    for duplicate_refresh in [
        ">刷新邮件配置</button>",
        "refreshMailConfig",
        "onclick=\"loadEmail()\"",
    ]:
        assert duplicate_refresh not in email_section

    assert ".drawer.has-fixed-footer > .drawer-scroll-body" in css
    assert "action.classList.add('drawer-fixed-footer');" in page
    assert "drawer.appendChild(action);" in page


def test_phase_21j_mail_templates_visual_pass_preserves_template_workflow():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    email_templates_section = _monitor_section(page, "email_templates")
    template_drawer = page[page.index('id="email_template_drawer_backdrop"') : page.index('<div id="run_log_backdrop"')]
    header = page[page.index("<header>") : page.index("</header>", page.index("<header>"))]

    for marker in [
        'data-page-entry="email_templates"',
        "邮件模板",
        'onclick="openEmailTemplateDrawer()"',
        'data-shortcut-tab="email" data-menu-key="mail_config"',
        'id="email_template_resource_summary"',
        'class="panel email-template-panel"',
        'id="email_template_search"',
        'id="email_template_status_filter"',
        "全部模板",
        "当前使用",
        "备用",
        "clearEmailTemplateFilters()",
        'id="email_templates_table"',
        "renderEmailTemplateResourceSummary",
        "renderEmailTemplatesTable",
        "email-template-table",
        "email-template-name-cell",
        "email-template-subject-cell",
        "email-template-updated-cell",
        "email-template-state-chip",
        "editEmailTemplate",
        "deleteEmailTemplate",
    ]:
        assert marker in email_templates_section or marker in page

    for marker in [
        'id="email_template_drawer_backdrop" class="drawer-backdrop" onclick="closeEmailTemplateDrawer()"',
        'id="email_template_drawer" class="drawer template-modal email-template-modal"',
        'aria-label="邮件模板"',
        'id="email_template_drawer_title"',
        'class="template-editor-stack"',
        'class="template-section-card template-section-card-basic"',
        'class="template-section-card template-section-card-html"',
        'class="template-section-card template-section-card-variables"',
        'class="template-active-toggle"',
        'id="email_template_active"',
        'id="email_template_name"',
        'id="email_template_subject"',
        'id="email_template_preset"',
        'id="email_html_template"',
        'id="email_template_guardrail"',
        "HTML 模板必须包含 {report_html} 或 {report_body}，否则无法保存。",
        'onclick="copyText(\'{report_body}\')"',
        'onclick="copyText(\'{report_html}\')"',
        'data-feedback-bound="1" onclick="copyText',
        'id="email_template_save_button"',
        "saveEmailTemplate()",
        'id="email_template_preview_button"',
        "previewEmailTemplate(true, this)",
        "resetEmailTemplateForm()",
        'class="template-preview-panel"',
        'class="template-preview-head"',
        'id="email_preview_subject"',
        'id="email_template_preview" class="preview-frame"',
        "closeEmailTemplateDrawer()",
    ]:
        assert marker in template_drawer or marker in page

    for marker in [
        "function copyText(text)",
        "navigator.clipboard.writeText(value)",
        "toast('变量已复制')",
        "emailTemplateHasReportBodyPlaceholder",
        "updateEmailTemplateGuardrail()",
        "applyEmailTemplatePreset",
        "email-templates/preview",
        "templateDrawerActive",
        "refreshEnhancedFilterSelects(drawer)",
        "openDrawerChrome('email_template_drawer_backdrop', 'email_template_drawer')",
    ]:
        assert marker in page

    for selector in [
        "body #email_templates .email-template-panel {",
        "body #email_templates_table .email-template-table table {",
        "body .email-template-name-cell,",
        "body .email-template-state-chip {",
        "body #email_template_drawer .template-editor-layout {",
        "body .template-editor-stack {",
        "body .template-section-card {",
        "body .template-section-head {",
        "body .template-active-toggle {",
        "body .template-section-card-html textarea {",
        "body .template-variable-row {",
        "body .template-variable-row button {",
        "body .template-variable-row button:focus-visible {",
        "body .template-preview-head {",
        "body .template-preview-subject {",
        "body .template-preview-panel iframe {",
    ]:
        assert selector in css

    assert "grid-template-columns: minmax(0, 1.08fr) minmax(360px, .92fr);" in css
    assert "grid-template-columns: minmax(0, 1fr);" in css[css.index("@media (max-width: 1279px)") :]
    assert 'id="global_refresh_button"' in header
    assert "refreshActiveSection(this)" in header
    for duplicate_refresh in [
        ">刷新邮件模板</button>",
        "refreshEmailTemplates",
        "onclick=\"loadEmailTemplates()\"",
    ]:
        assert duplicate_refresh not in email_templates_section

    assert "点击复制到编辑器；`{report_body}` 和 `{report_html}` 至少保留一个。" not in page
    assert "正文占位符已保留" not in email_templates_section
    assert "email-template-updated-cell" in page
    assert "formatCompactTime(t.updated_at)" in page
    assert "title=\"${esc(formatTime(t.updated_at))}\"" in page
    assert "esc(formatCompactTime(t.updated_at))" in page


def test_phase_21k_runtime_strategy_visual_pass_preserves_runtime_workflow():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    runtime_section = _monitor_section(page, "runtime")
    header = page[page.index("<header>") : page.index("</header>", page.index("<header>"))]

    for marker in [
        'data-page-entry="runtime"',
        "运行策略",
        "runtime-save-button",
        "saveRuntimeSettings()",
        'data-shortcut-tab="doctor" data-menu-key="system_diagnostics"',
        'class="runtime-guardrail"',
        "保存后会影响后续采集、登录会话、调度检查或清理任务。",
        'id="runtime_settings_table"',
    ]:
        assert marker in runtime_section

    for marker in [
        "function renderRuntimeSettings(settings)",
        "const groups=['Crawling','Login','Scheduler','Retention'];",
        "const groupNotes={",
        "runtime-settings-card",
        "runtime-settings-card-head",
        "runtime-settings-table",
        "table(['设置','当前值','调整','范围','生效','锁定'], rows, {className:'runtime-settings-table', minWidth:1040})",
        "function runtimeSettingRow(item)",
        "data-runtime-key",
        "runtime-setting-name",
        "runtime-current-value",
        "runtime-setting-control",
        "runtime-range",
        "runtime-scope-chip",
        "runtime-lock-state",
        "runtime-lock-pill",
        "function runtimeSettingGroupHint(group)",
        "function runtimeSettingSourceLabel(source)",
        "function runtimeLockReasonLabel(reason)",
        "由部署配置控制",
        "'scheduler recovery':'调度恢复检查'",
        "document.querySelectorAll('[data-runtime-key]').forEach",
        "api('/runtime-settings',{method:'PUT'",
        "toast('运行策略已保存')",
    ]:
        assert marker in page

    for selector in [
        "body #runtime .panel {",
        "body #runtime .runtime-guardrail {",
        "body #runtime_settings_table {",
        "body .runtime-settings-card {",
        "body .runtime-settings-card-head {",
        "body #runtime_settings_table .runtime-settings-table {",
        "body #runtime_settings_table .runtime-settings-table table {",
        "body .runtime-setting-name,",
        "body .runtime-current-value,",
        "body .runtime-setting-control select,",
        "body .runtime-range {",
        "body .runtime-scope-chip {",
        "body .runtime-save-button {",
    ]:
        assert selector in css

    assert 'id="global_refresh_button"' in header
    assert "refreshActiveSection(this)" in header
    for duplicate_refresh in [
        ">刷新运行策略</button>",
        "refreshRuntimeSettings",
        'onclick="loadRuntimeSettings()"',
    ]:
        assert duplicate_refresh not in runtime_section

    for forbidden_customer_copy in [
        "MONITOR_GLOBAL_CRAWL_CONCURRENCY",
        "MONITOR_SCHEDULER_TICK_SECONDS",
        "deployment configuration",
        "<div class=\"small\">${esc(item.key)}</div>",
        "<div class=\"small\">${esc(item.source||'default')}</div>",
    ]:
        assert forbidden_customer_copy not in runtime_section


def test_phase_12b_page_entry_and_role_flow_shortcuts():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    dashboard_section = _monitor_section(page, "dashboard")

    for entry in [
        'data-page-entry="dashboard"',
        'data-page-entry="jobs"',
        'data-page-entry="runs"',
        'data-page-entry="accounts"',
        'data-page-entry="proxies"',
        'data-page-entry="ai"',
        'data-page-entry="ai_rules"',
        'data-page-entry="email"',
        'data-page-entry="email_templates"',
        'data-page-entry="runtime"',
        'data-page-entry="doctor"',
    ]:
        assert entry in page

    for marker in [
        "运营首页",
        "任务监控",
        "任务中心",
        "运行记录",
        "资源支撑",
        "系统配置",
        'id="task_loop_shortcuts"',
        'class="task-loop-shortcuts"',
        'data-shortcut-action="new-job"',
        'data-shortcut-target="task_center_panel"',
        'data-shortcut-grouped="1"',
        'data-shortcut-grouped="0"',
        'id="task_center_panel"',
        'id="task_group_toggle"',
        "async function refreshActiveSection(button=null)",
        "function navigateShortcut(tab, options={})",
        "function normalizeConsoleTab(tab)",
        "function setTaskCenterGrouped(checked, options={})",
        "function loadTaskCenter()",
        "function clearRunFilters()",
        "function refreshTaskCenter()",
        "function bindShortcutButtons(root=document)",
        "root.querySelectorAll('[data-shortcut-tab]').forEach",
        "document.querySelectorAll('[data-shortcut-tab][data-menu-key]')",
        "btn.classList.toggle('is-hidden', !canMenu(btn.dataset.menuKey))",
        "document.querySelectorAll('.admin-entry[data-menu-key]')",
        "if(options.action==='new-job')",
        "openNewJobDrawer()",
        "loadReadiness();",
        "loadSchedulerStatus();",
        "loadPlatformStatus();",
        "addPermittedLoad(loads, '系统状态', 'system_diagnostics', loadSystemChecklist);",
    ]:
        assert marker in page

    for selector in [
        ".page-entry",
        ".page-title-block",
        ".page-kicker",
        ".page-filter-region",
        ".task-loop-shortcuts",
        ".shortcut-card",
        ".shortcut-primary",
        ".report-task-group",
    ]:
        assert selector in css

    assert "刷新全局状态" not in page
    assert "刷新当前页面" not in page
    assert 'id="global_refresh_button"' in page
    assert 'aria-label="Refresh current page"' in page
    assert ".refresh-icon-button" in css
    assert ".refresh-icon-button.is-loading .refresh-icon" in css

    # Normal-user shortcuts and page actions must not expose administrator
    # resource entries when menu permissions hide them.
    assert 'class="shortcut-card admin-entry" type="button" data-shortcut-tab="accounts" data-menu-key="platform_accounts"' in page
    assert 'data-shortcut-tab="accounts" data-menu-key="platform_accounts"><span>04</span><strong>资源处理</strong>' in page
    assert 'data-shortcut-tab="doctor" data-menu-key="system_diagnostics"' in page
    assert 'data-shortcut-tab="email" data-menu-key="mail_config"' in page
    assert 'data-shortcut-tab="ai_rules" data-menu-key="ai_rules"' in page

    assert "/api/monitor/dashboard" not in page
    assert "email_delivery_logs" not in dashboard_section
    assert "job_snapshot_json" not in dashboard_section
    assert "crawl_runs.visibility" not in dashboard_section


def test_phase_13b_operations_home_desktop_visual_metrics():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    js = Path("api/webui/monitor/monitor.js").read_text(encoding="utf-8")
    echarts_vendor = Path("api/webui/monitor/vendor/echarts.min.js")
    dashboard_section = _monitor_section(page, "dashboard")

    for marker in [
        'class="operations-home"',
        'id="operations_home_meta"',
        'id="dashboard_metrics" class="operations-metric-grid"',
        'id="operations_home_breakdowns" class="operations-breakdown-grid"',
        'id="operations_home_attention" class="operations-chart-card operations-issues-chart"',
        'id="operations_home_resource"',
        'id="operations_home_admin_health" class="operations-admin-health admin-entry" data-menu-key="system_diagnostics"',
        'class="operations-context-bar"',
        'class="operations-chart-card operations-main-trend operations-cockpit-trend"',
        'class="operations-trend-head"',
        'class="operations-trend-controls"',
        'class="operations-window-toggle" role="tablist" aria-label="时间窗口"',
        "renderOperationsChartSurface('trend', 'trend', '监控走势')",
        "renderOperationsChartSurface('issues', 'issues', '问题分布')",
        "renderOperationsChartSurface('platforms', 'platforms', '平台分布')",
        "renderOperationsChartSurface('delivery', 'delivery', '交付 / 复核')",
        "renderOperationsChartSurface('resource', 'resource', '资源健康')",
        'class="operations-breakdown-card operations-platform-chart"',
        'class="operations-breakdown-card operations-delivery-chart"',
        'class="operations-chart-fallback"',
        'class="operations-visual-track"',
        'class="task-loop-shortcuts operations-quick-dock"',
        "operations-visual-segment is-",
        "const operationsOverviewState = {",
        "charts: new Map()",
        "async function loadOperationsTrendBuckets(days=7){",
        "async function setOperationsTrendWindow(days=7){",
        "function operationsOverviewViewModel(home, trendPayload=null){",
        "function renderOperationsTrendChart(model){",
        "function renderOperationsBreakdowns(model){",
        "function renderOperationsIssueChart(model){",
        "function renderOperationsChartSurface",
        "function scheduleOperationsChartRender(model)",
        "function renderOperationsECharts(model)",
        "function operationsTrendOption(model)",
        "function operationsIssueOption(model)",
        "function operationsPlatformOption(model)",
        "function operationsDeliveryOption(model)",
        "function operationsResourceOption(model)",
        "function resourceSignalTone(status)",
        "function disposeOperationsChart(key)",
        "function operationsShortcutAttributes(card)",
        "class=\"operations-resource-entry\"",
        "class=\"operations-resource-card operations-resource-card-primary\"",
        "class=\"operations-resource-card operations-admin-chart operations-resource-card-secondary\"",
        "const home=data.operations_home || s.operations_home || legacyOperationsHome(s)",
        "function renderOperationsHome(home, options={})",
        "function operationsToneLabel(tone)",
        "function operationsMetricIcon(label)",
        "function operationsLegend(items=[], extraClass='')",
        "function operationsMetricCard(card)",
        "function renderOperationsResourceHealth(model)",
        "function renderOperationsAdminHealth(model)",
        "bindShortcutButtons(document.getElementById('dashboard'))",
        "class=\"operations-chart-legend",
        "class=\"operations-legend-item",
        "class=\"operations-metric-icon ",
        "const platformPalette=['#0f766e', '#2563eb', '#65a30d', '#d97706', '#7c3aed', '#0891b2', '#94a3b8'];",
        "data-shortcut-target=\"task_center_panel\"",
        "data-shortcut-grouped=\"1\"",
        "data-shortcut-tab=\"accounts\" data-menu-key=\"platform_accounts\"",
        "data-shortcut-tab=\"doctor\" data-menu-key=\"system_diagnostics\"",
        "资源由管理员维护",
        "任务",
        "运行",
        "报告",
        "邮件",
        "线索",
        "运行失败",
        "邮件失败",
        "待复核",
        "业务总量",
        "异常 / 待处理",
        "高风险",
        "监控走势",
        "问题分布",
        "平台分布",
        "交付 / 复核",
        "7天",
        "14天",
    ]:
        assert marker in page

    assert echarts_vendor.exists()
    assert echarts_vendor.stat().st_size > 500_000
    assert '<script src="/static/monitor/vendor/echarts.min.js"></script>' in page
    assert "/static/monitor/vendor/echarts.min.js" in page
    assert not re.search(
        r"<script[^>]+(?:https?:)?//[^>]*(?:echarts|cdn|unpkg|jsdelivr)",
        page,
        flags=re.IGNORECASE,
    )
    assert "https://cdn" not in page.lower()
    assert "unpkg" not in page.lower()
    assert "jsdelivr" not in page.lower()
    assert "echarts.init(" in page
    assert "echarts.init(" not in js

    for shortcut_label in [
        "<span>01</span><strong>新建任务</strong>",
        "<span>02</span><strong>任务中心</strong>",
        "<span>03</span><strong>运行记录</strong>",
        "<span>04</span><strong>资源处理</strong>",
        "<span>05</span><strong>系统诊断</strong>",
    ]:
        assert shortcut_label in dashboard_section

    assert "按任务、运行、报告和邮件交付汇总当前状态。" not in page
    assert "首页仅保留精简运营健康信号" not in page
    assert "operationsChipList(" not in page
    operations_block = page.split("const operationsOverviewState = {", 1)[1].split("function metricSkeletonGrid", 1)[0]
    assert "function operationsOverviewViewModel(home, trendPayload=null)" in page
    assert "function operationsIssueOption(model)" in page
    assert "function operationsPlatformOption(model)" in page
    assert "function operationsDeliveryOption(model)" in page
    for marker in [
        "operationsHealthSummary",
        "operationsIssueSeverityRank",
        "operationsPlatformFailureRows",
        "CR-106A data-aware signal refinement",
    ]:
        assert marker in operations_block
    assert "报告级邮件状态" in operations_block
    assert "report.email_status" in operations_block
    assert "email_delivery_logs" not in dashboard_section
    assert "email_delivery_logs" not in operations_block
    assert "operationsTrendLinePath" not in operations_block
    assert "operationsTrendAreaPath" not in operations_block
    assert 'class="operations-trend-svg"' not in operations_block
    assert "任务健康" not in operations_block
    assert "运行活动" not in operations_block
    assert "报告与复核" not in operations_block
    assert "邮件交付" not in operations_block
    assert "疑似负面线索" not in operations_block

    assert "operationsTrack(card.segments, card.tone)" in page
    assert "operationsChartUnavailable()" in page
    assert "buildOperationsFallbackTrend(home, operationsOverviewState.trendDays)" in page
    assert "api('/runs?'+runQs.toString())" in page
    assert "api('/reports?'+reportQs.toString())" in page
    assert "<button type=\"button\" class=\"operations-metric-card" in page
    assert 'data-chart-action="${esc(action)}"' in page
    assert "renderOperationsChartSurface('platforms', 'platforms', '平台分布')" in page
    assert "renderOperationsChartSurface('delivery', 'delivery', '交付 / 复核')" in page
    assert "renderOperationsChartSurface('issues', 'issues', '问题分布')" in page

    for selector in [
        ".operations-home",
        ".operations-home-meta",
        ".operations-context-bar",
        ".operations-metric-grid",
        ".operations-metric-card",
        ".operations-home-lower",
        ".operations-chart-card",
        ".operations-cockpit-trend",
        ".operations-trend-head",
        ".operations-trend-controls",
        ".operations-window-toggle",
        ".operations-chart-surface",
        ".operations-chart-fallback",
        ".operations-breakdown-grid",
        ".operations-breakdown-card",
        ".operations-chart-legend",
        ".operations-legend-item",
        ".operations-metric-icon",
        ".operations-resource-panel",
        ".operations-resource-entry",
        ".operations-resource-card",
        ".operations-resource-diagnostic-track",
        ".operations-resource-signals",
        ".operations-admin-health",
        ".operations-admin-chart",
        ".operations-quick-dock",
        "CR-098 final cascade: data-first Operations Home visual refit.",
        "CR-099 visual clarity: legend-first operations home alignment.",
        "CR-100 density compaction: content-sized operations home layout.",
        "CR-104 cockpit dashboard rebuild: single-screen chart-first Operations Home.",
        "CR-105 ECharts dashboard rebaseline: local chart modules.",
    ]:
        assert selector in css

    final_cr105_block = css.split("/* CR-105 ECharts dashboard rebaseline: local chart modules. */", 1)[1]
    for compact_dashboard_rule in [
        "body #dashboard.active .operations-home.is-user-overview .operations-home-lower {",
        "body #dashboard.active .operations-home.is-user-overview #operations_home_resource {\n  display: none;",
        "body #dashboard.active .operations-window-toggle {",
        "body #dashboard.active .operations-window-toggle button.is-active {",
        "body #dashboard.active .operations-chart-surface {",
        "body #dashboard.active .operations-cockpit-trend {",
        "body #dashboard.active .operations-breakdown-grid {\n  display: contents;",
        "body #dashboard.active .operations-resource-entry {",
        "body #dashboard.active .operations-resource-card {",
        "body #dashboard.active .operations-quick-dock {\n  display: none !important;",
        "body #dashboard.active > #operations_home_admin_health {\n  display: none !important;",
        "body #dashboard.active {\n    min-height: calc(100vh - 68px);\n    max-height: calc(100vh - 68px);",
        "body #dashboard.active > .operations-home {\n    height: calc(100vh - 144px);\n    max-height: calc(100vh - 144px);",
        'grid-template-areas:\n    "trend trend attention"\n    "platforms delivery resource";',
    ]:
        assert compact_dashboard_rule in final_cr105_block

    assert "if(status === 'empty') return 'is-neutral';" in page
    assert "if(status === 'empty') return 'neutral';" in page
    for old_chart_marker in [
        'class="operations-trend-svg"',
        "operationsTrendSeries(",
        "operationsTrendLinePath",
        "operationsTrendAreaPath",
        "operations-flow-line operations-trend-line",
        "operations-trend-point",
        "class=\"operations-platform-row",
        "class=\"operations-delivery-row",
        "class=\"operations-priority-node",
    ]:
        assert old_chart_marker not in operations_block

    assert "new Chart(" not in page
    assert "chart.js" not in page.lower()
    assert "email_delivery_logs" not in dashboard_section
    assert "job_snapshot_json" not in dashboard_section
    assert "crawl_runs.visibility" not in dashboard_section
    assert "function legacyOperationsHome(summary)" in page
    assert "资源由管理员维护" in page


def test_cr051_task_center_consolidates_report_grouping_without_separate_report_center():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    task_center = _monitor_section(page, "runs")

    assert '<button data-tab="runs" data-menu-key="run_center"' in page
    assert '<span>任务中心</span>' in page
    assert '<section id="reports"' not in page
    assert 'data-tab="reports"' not in page
    assert 'data-shortcut-tab="reports"' not in page
    assert "report_center" not in page

    assert 'id="task_group_view"' not in task_center
    assert 'id="run_records_view"' not in task_center
    assert 'id="task_center_panel"' in task_center
    assert 'id="task_group_toggle"' in task_center
    assert 'id="runs_table"' in task_center
    assert "按舆情任务分组" in task_center
    assert "按任务汇总运行、AI 评估、报告和交付状态" in task_center

    assert "function renderGroupedTaskRuns(runs)" in page
    assert "function groupRunsByTask(runs)" in page
    assert "function renderTaskRunGroup(group)" in page
    assert "function runTableHeaders(mode='flat')" in page
    assert "if(context.jobId) return `job:${context.jobId}`;" in page
    assert "const jobId=Number(run.job_id || summary.job_id || 0) || null;" in page
    assert "function runGroupMetricChips(group)" in page
    assert "function runGroupContextNote(group)" in page
    assert 'class="report-task-group-metrics" aria-label="运行分组汇总"' in page
    assert "旧运行缺少完整任务快照，可逐条进入详情。" in page
    assert "按舆情监控任务汇总运行记录。" not in page
    assert "function normalizeConsoleTab(tab)" in page
    assert "return tab==='reports' ? 'runs' : tab;" in page
    assert "function switchRunCenterView(viewId, options={})" in page
    assert "function loadTaskCenter()" in page
    assert "function refreshTaskCenter()" in page

    assert "'任务 ID','运行 ID','状态','任务名称 / 律所','平台','关键词摘要','开始时间'" in page
    assert "'运行 ID','状态','开始时间','耗时','采集数','新增数'" in page
    assert "任务 / 律所" in _monitor_section(page, "runs")


def test_task_center_single_grouping_switch_unifies_rows_and_actions():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    task_center = _monitor_section(page, "runs")
    run_detail_drawer = page[page.index('id="run_detail_drawer"') : page.index('id="report_preview_backdrop"')]

    assert 'id="task_group_view"' not in task_center
    assert 'id="run_records_view"' not in task_center
    assert 'id="task_center_panel"' in task_center
    assert 'id="task_group_toggle"' in task_center
    assert "setTaskCenterGrouped(this.checked)" in task_center
    assert "let taskCenterGrouped = true;" in page
    assert "taskCenterGrouped ? renderGroupedTaskRuns(runs) : renderFlatTaskRuns(runs)" in page
    assert "const headers=runTableHeaders('flat');" in page
    assert "return table(headers, runs.map(r=>runRow(r, headers)), {className:'run-record-table run-record-table-flat', minWidth:1720});" in page
    assert "const headers=runTableHeaders('grouped');" in page
    assert "table(headers, group.runs.map(r=>runRow(r, headers)), {className:'run-record-table run-record-table-grouped', minWidth:1320})" in page
    assert "return (headers||runTableHeaders()).map(header=>cells[header] ?? '');" in page
    assert "'任务 ID': r.job_id || '-'," in page
    assert "'关键词摘要': `<div class=\"truncate\"" in page
    assert "function tableColumnClass(header)" in page
    assert "if(['操作','详情'].includes(header)) classes.push('col-actions');" in page
    assert "if(header==='状态') classes.push('col-status');" in page
    assert "right: 0;" in css
    assert "th.col-actions" in css
    assert "td.col-status" in css
    assert "task_group_toggle" in task_center
    assert "刷新任务中心" not in task_center
    assert "onclick=\"loadRuns()\">刷新</button>" not in task_center
    assert ".content {\n  min-width: 0;\n  overflow: visible;\n}" in css
    assert ".content { min-width:0; overflow:visible; }" in page
    assert ".content { min-width:0; overflow:hidden; }" not in page
    assert "enhanceFilterSelects();" in page
    assert "enhanceFilterDateInputs();" in page
    assert 'id="run_status_filter"' in task_center
    assert 'id="run_platform_filter"' in task_center
    assert 'id="run_date_from" type="date"' in task_center
    assert 'id="run_date_to" type="date"' in task_center
    assert 'id="run_type_filter"' in task_center
    assert 'id="run_visibility_filter"' in task_center
    assert 'id="run_limit"' in task_center
    assert "filter-select-button" in css
    assert "filter-select-option" in css
    assert "filter-date-menu" in css
    assert "filter-date-day.is-selected" in css
    assert "const wrapper=activeFilterDateButton.closest('.filter-date-enhanced');" in page
    assert "if(wrapper && menu.parentElement !== wrapper)" in page
    assert "wrapper.appendChild(menu);" in page
    assert "const triggerWidth=Math.max(0, Math.round(activeFilterDateButton.getBoundingClientRect().width));" in page
    assert "menu.style.width=`${triggerWidth}px`;" in page
    assert "menu.style.setProperty('--filter-date-anchor-x', '50%');" in page
    assert "menu.style.position='absolute';" in page
    assert "menu.style.left='0';" in page
    assert "menu.style.top='calc(100% + 4px)';" in page
    assert "const visualViewportWidth=window.visualViewport?.width || 0;" not in page
    assert "const viewportWidth=visualViewportWidth || window.innerWidth || document.documentElement.clientWidth || 0;" not in page
    assert "const edgeMargin=6;" not in page
    assert "const attachedWidth=Math.max(0, viewportWidth - rect.left - edgeMargin);" not in page
    assert "const maxMenuWidth=Math.max(0, viewportWidth - edgeMargin * 2);" not in page
    assert "const triggerWidth=Math.max(0, Math.round(rect.width));" not in page
    assert "const minimumReadableWidth=Math.min(168, maxMenuWidth);" not in page
    assert "const menuWidth=Math.min(Math.max(triggerWidth, minimumReadableWidth), attachedWidth || maxMenuWidth, maxMenuWidth, 340);" not in page
    assert "const maxLeft=Math.max(edgeMargin, viewportWidth - menuWidth - edgeMargin);" not in page
    assert "let left=Math.max(edgeMargin, Math.min(maxLeft, rect.left));" not in page
    assert "const readableWidth=Math.max(236, Math.round(rect.width));" not in page
    assert "const menuWidth=Math.min(readableWidth, attachedWidth || maxMenuWidth, maxMenuWidth, 340);" not in page
    assert "const minAttachedWidth=Math.min(188, maxMenuWidth);" not in page
    assert "const attachedWidth=Math.min(menuWidth, Math.max(0, viewportWidth - rect.left - edgeMargin));" not in page
    assert "const canUseTriggerLeft=rect.left >= edgeMargin && rect.left + menuWidth <= viewportWidth - edgeMargin;" not in page
    assert "const preferredLeft=rect.right - menuWidth;" not in page
    assert "const anchorX=Math.max(18, Math.min(menuWidth - 18, rect.left + rect.width / 2 - left));" not in page
    assert "menu.style.setProperty('--filter-date-anchor-x', `${Math.round(anchorX)}px`);" not in page
    assert "menu.style.width=`${Math.round(menuWidth)}px`;" not in page
    assert "menu.style.left=`${Math.round(left)}px`;" not in page
    assert ".filter-date-menu::before" in css
    assert "grid-template-columns: repeat(7, minmax(0, 1fr));" in css
    assert ".filter-date-day {\n  width: 100%;" in css
    assert "  padding: 0;" in css
    assert "grid-template-columns: 24px 1fr 24px;" in css
    assert ".report-task-group-metrics" in css
    assert ".report-task-group-metric.is-danger" in css
    assert ".report-task-group-note" in css

    assert "function groupRunsByTask(runs)" in page
    assert "function runGroupingContext(run)" in page
    assert "const jobId=Number(run.job_id || summary.job_id || 0) || null;" in page
    assert "if(context.jobId) return `job:${context.jobId}`;" in page
    assert "历史运行：任务上下文有限" in page

    assert "viewRunLeads(${Number(r.id)})" not in page
    assert "loadRunLogs(${Number(r.id)})" not in page
    assert "stopRun(${Number(r.id)})" not in page
    assert "archiveRun(${Number(r.id)})" not in page
    assert "restoreRun(${Number(r.id)})" not in page
    assert '<button class="secondary" onclick="openRunDetail(${Number(r.id)})">详情</button>' in page

    assert "run_detail_actions" in run_detail_drawer
    assert "stopRunFromDetail" in page
    assert "archiveRunFromDetail" in page
    assert "restoreRunFromDetail" in page
    assert "previewReport(${Number(report.id)})" in page
    assert "jumpToReportAiEvaluations(${Number(report.id)})" in page
    assert "loadEmailDeliveryHistory(${Number(report.id)})" in page
    assert "resendReportEmail(${Number(report.id)})" in page
    assert "/reports/${Number(report.id)}/download?type=html" in page
    assert "/reports/${Number(report.id)}/download?type=excel" in page
    assert "/reports/${Number(report.id)}/download?type=markdown" in page


def test_phase_21l_task_center_visual_pass_preserves_current_structure():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    task_center = _monitor_section(page, "runs")
    header = page[page.index("<header>") : page.index("</header>", page.index("<header>"))]

    for marker in [
        '<button data-tab="runs" data-menu-key="run_center"',
        '<span>任务中心</span>',
        'data-page-entry="runs"',
        'id="task_center_panel"',
        'id="task_group_toggle"',
        "setTaskCenterGrouped(this.checked)",
        'id="run_task_filter"',
        'id="run_status_filter"',
        'id="run_platform_filter"',
        'id="run_date_from" type="date"',
        'id="run_date_to" type="date"',
        'id="run_type_filter"',
        'id="run_visibility_filter"',
        'id="run_limit"',
        "applyRunFilters()",
        "clearRunFilters()",
        'id="run_filter_summary"',
        'id="runs_table"',
        'id="run_pagination"',
        "renderGroupedTaskRuns(runs)",
        "renderFlatTaskRuns(runs)",
        "renderTaskRunGroup(group)",
        "runGroupMetricChips(group)",
        "runGroupMetaChips(group)",
        "runTableHeaders('grouped')",
        "runTableHeaders('flat')",
        "openRunDetail(${Number(r.id)})",
    ]:
        assert marker in task_center or marker in page

    for selector in [
        "body .task-center-panel {",
        "body #task_center_panel .account-ledger-head {",
        "body #task_center_panel .page-toolbar {",
        "body #running_summary .notice {",
        "body #run_filter_summary {",
        "body #runs_table {",
        "body .report-task-group {",
        "body .report-task-group-head:hover {",
        "body .report-task-group-head:focus-visible {",
        "body .report-task-group-metrics {",
        "body .report-task-group-metric.is-warn {",
        "body .report-task-group-metric.is-danger {",
        "body .report-task-group-meta span {",
        "body #runs_table .run-record-table {",
        "body #runs_table .run-record-table tbody tr:hover td {",
        "body .run-empty-state {",
        "body #run_pagination {",
    ]:
        assert selector in css

    assert 'id="global_refresh_button"' in header
    assert "refreshActiveSection(this)" in header
    for forbidden in [
        '<section id="reports"',
        'data-tab="reports"',
        'data-shortcut-tab="reports"',
        "report_center",
        'id="task_group_view"',
        'id="run_records_view"',
        ">刷新任务中心</button>",
        "onclick=\"loadRuns()\">刷新</button>",
        "viewRunLeads(${Number(r.id)})",
        "loadRunLogs(${Number(r.id)})",
        "stopRun(${Number(r.id)})",
        "archiveRun(${Number(r.id)})",
        "restoreRun(${Number(r.id)})",
        "<label>线索状态</label>",
        'id="report_risk"',
    ]:
        assert forbidden not in task_center

    assert '<button class="secondary" onclick="openRunDetail(${Number(r.id)})">详情</button>' in page
    assert "report-task-group-metric${tone}" in page
    assert "if(['操作','详情'].includes(header)) classes.push('col-actions');" in page
    assert "if(header==='状态') classes.push('col-status');" in page


def test_phase_21m_overlay_and_run_detail_freeze_gate_preserves_structure():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    frontend_source = page + "\n" + css
    run_detail_drawer = page[
        page.index('id="run_detail_drawer"') : page.index('id="report_preview_backdrop"')
    ]

    expected_tabs = [
        ('data-run-detail-tab="overview"', "概览"),
        ('data-run-detail-tab="logs"', "采集日志"),
        ('data-run-detail-tab="contents"', "采集内容"),
        ('data-run-detail-tab="ai"', "AI 评估"),
        ('data-run-detail-tab="reports"', "报告"),
        ('data-run-detail-tab="email"', "邮件交付"),
    ]
    previous_position = -1
    for tab_marker, label in expected_tabs:
        position = run_detail_drawer.index(tab_marker)
        assert position > previous_position
        assert f"{label}</button>" in run_detail_drawer[position : position + 180]
        previous_position = position

    for marker in [
        'id="run_detail_backdrop" class="drawer-backdrop" onclick="closeRunDetailDrawer()"',
        'id="run_detail_drawer" class="drawer run-detail-drawer"',
        "data-run-detail-panel",
        'id="run_detail_body"',
        'id="run_detail_actions"',
        "function switchRunDetailTab(tab)",
        "function openRunDetail(id)",
        "openDrawerChrome('run_detail_backdrop', 'run_detail_drawer')",
        "renderRunDetailBody(currentRunDetailState)",
        "run_detail_report_filter",
        "run-detail-ai-toolbar page-filter-region",
        'data-filter-region="run-detail-ai"',
    ]:
        assert marker in page

    for opener in [
        "openDrawerChrome('account_modal', 'account_dialog')",
        "openDrawerChrome('proxy_drawer_backdrop', 'proxy_drawer')",
        "openDrawerChrome('job_drawer_backdrop', 'job_drawer')",
        "openDrawerChrome('ai_rule_modal_backdrop', 'ai_rule_modal')",
        "openDrawerChrome('ai_connection_test_backdrop', 'ai_connection_test_modal', {lockBody:false})",
        "openDrawerChrome('ai_profile_drawer_backdrop', 'ai_profile_drawer')",
        "openDrawerChrome('mail_config_backdrop', 'mail_config_modal')",
        "openDrawerChrome('mail_test_backdrop', 'mail_test_modal')",
        "openDrawerChrome('email_template_drawer_backdrop', 'email_template_drawer')",
        "openDrawerChrome('run_log_backdrop', 'run_log_drawer')",
        "openDrawerChrome('run_detail_backdrop', 'run_detail_drawer')",
        "openDrawerChrome('report_preview_backdrop', 'report_preview_drawer')",
        "openDrawerChrome('email_delivery_history_backdrop', 'email_delivery_history_drawer')",
    ]:
        assert opener in page

    for marker in [
        "function normalizeDrawerScrollBodies(root=document)",
        "const footerSelector='.form-actions, .resource-modal-actions, .account-flow-actions, .ai-test-actions, .rule-modal-actions';",
        "const topChromeSelector='.run-detail-toolbar, .detail-tabs.run-detail-tabs, .report-leads-toolbar, .drawer-actions';",
        "body.className='drawer-scroll-body';",
        "action.classList.add('drawer-fixed-footer');",
        "drawer.classList.toggle('has-fixed-footer', !!drawer.querySelector(':scope > .drawer-fixed-footer'));",
        "drawer.classList.toggle('has-fixed-toolbar', !!drawer.querySelector(':scope > .drawer-fixed-toolbar'));",
        "body.setAttribute('data-scroll-owner', 'drawer-content');",
    ]:
        assert marker in page

    for selector in [
        "body .drawer > .drawer-fixed-toolbar {",
        "body .drawer > .drawer-fixed-toolbar.detail-tabs {",
        "body .drawer > .drawer-fixed-footer {",
        "body .drawer.has-fixed-toolbar > .drawer-scroll-body {",
        "body .drawer.has-fixed-footer > .drawer-scroll-body {",
        "body .run-detail-tabs button.active {",
        "body .run-detail-code {",
        "body .filter-select-option:focus-visible,",
        "body .filter-select-menu::-webkit-scrollbar-thumb,",
    ]:
        assert selector in css

    for marker in [
        "overflow: hidden;",
        "overflow-x: hidden;",
        "overflow-y: auto;",
        "scrollbar-gutter: stable;",
        "scrollbar-color: rgba(88, 106, 125, 0.42) transparent;",
        "--phase21-shadow-overlay-strong",
    ]:
        assert marker in frontend_source

    for drawer_id in [
        "job_drawer",
        "account_dialog",
        "proxy_drawer",
        "ai_connection_test_modal",
        "ai_profile_drawer",
        "ai_rule_modal",
        "mail_config_modal",
        "mail_test_modal",
        "email_template_drawer",
    ]:
        drawer_start = page.index(f'id="{drawer_id}"')
        next_drawer = page.find('class="drawer', drawer_start + 1)
        drawer_html = page[drawer_start : next_drawer if next_drawer != -1 else len(page)]
        if "drawer-fixed-footer" in drawer_html:
            assert 'class="drawer-scroll-body"' not in drawer_html[: drawer_html.index("drawer-fixed-footer")]

    for forbidden in [
        "report_leads_drawer",
        "document.getElementById('run_detail_drawer').classList.add('active')",
        "document.getElementById('job_drawer_backdrop').classList.add('active')",
        "document.getElementById('mail_config_modal').classList.add('active')",
        ".drawer::-webkit-scrollbar",
        "top: calc(var(--drawer-padding-y) + 80px);",
        'data-tab="reports"',
    ]:
        assert forbidden not in frontend_source

    assert "previewReport(${Number(report.id)})" in page
    assert "jumpToReportAiEvaluations(${Number(report.id)})" in page
    assert "loadEmailDeliveryHistory(${Number(report.id)})" in page
    assert "resendReportEmail(${Number(report.id)})" in page
    assert "copyCurrentRunLogs()" in page
    assert "downloadCurrentRunLogs()" in page


def test_phase_21n_system_diagnostics_visual_pass_preserves_actions_and_mounts():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    doctor_section = page[page.index('<section id="doctor">') : page.index('</section>', page.index('<section id="doctor">'))]

    for marker in [
        'data-page-entry="doctor"',
        'onclick="loadDoctor()">重新诊断</button>',
        'onclick="runSmokeCheck()">运行系统诊断</button>',
        'data-shortcut-tab="accounts" data-menu-key="platform_accounts">处理账号资源</button>',
        'id="doctor_summary"',
        'id="smoke_result"',
        'id="doctor_table"',
        'id="doctor_recommendations"',
        'id="readiness_summary"',
        'id="ops_checklist"',
        'id="readiness_actions"',
        'id="readiness_table"',
        'id="scheduler_status"',
        'id="platform_status_table"',
        'diagnostic-overview-layout',
        'diagnostic-primary-panel',
        'diagnostic-next-panel',
        'diagnostic-readiness-panel',
        'diagnostic-scheduler-panel',
        'diagnostic-platform-panel',
    ]:
        assert marker in doctor_section

    for marker in [
        "if(tab==='doctor') {\n        return Promise.all([\n          loadDoctor(),\n          loadReadiness(),\n          loadSchedulerStatus(),\n          loadPlatformStatus(),\n          loadSystemChecklist()\n        ]);\n      }",
        "addPermittedLoad(loads, '系统状态', 'system_diagnostics', loadSystemChecklist);",
        "addPermittedLoad(loads, '运行状态', 'system_diagnostics', loadReadiness);",
        "addPermittedLoad(loads, '调度状态', 'system_diagnostics', loadSchedulerStatus);",
        "addPermittedLoad(loads, '平台状态', 'system_diagnostics', loadPlatformStatus);",
        "api('/doctor')",
        "api('/readiness')",
        "api('/scheduler-status')",
        "api('/smoke',{method:'POST'})",
        "renderReadinessActions(data)",
        "platformStatusTable(data.platforms)",
    ]:
        assert marker in page

    for marker in [
        "基础配置",
        "诊断结果",
        "处理路径",
        "下一步处理",
        "运行就绪",
        "系统运行状态",
        "自动调度",
        "调度器状态",
        "平台资源",
        "平台状态",
        "验证范围：数据库、报告生成、附件和诊断汇总链路。",
    ]:
        assert marker in page

    for selector in [
        "body .diagnostic-overview-layout {",
        "body .diagnostic-section-title {",
        "body .diagnostic-summary-card {",
        "body .diagnostic-recommendation-stack {",
        "body .diagnostic-recommendation {",
        "body .diagnostic-scheduler-card {",
        "body .diagnostic-smoke-state {",
        "body #doctor .table-wrap {",
        ".diagnostic-overview-layout,",
    ]:
        assert selector in css

    for forbidden in [
        "href=\"/doctor\"",
        "data-tab=\"reports\"",
        "report_center",
        "simulate error state",
        "static data",
        "mock data",
    ]:
        assert forbidden not in doctor_section


def test_phase_21o_login_page_visual_pass_preserves_auth_flow():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    auth = page[
        page.index('<div id="auth_screen"') : page.index(
            '<svg aria-hidden="true"', page.index('<div id="auth_screen"')
        )
    ]

    for marker in [
        'id="auth_screen"',
        'class="auth-screen"',
        'class="auth-panel"',
        'role="dialog"',
        'aria-labelledby="auth_title"',
        'aria-describedby="auth_intro"',
        'class="auth-panel-head"',
        'class="auth-brand"',
        'class="auth-brand-mark"',
        'id="auth_title"',
        'id="auth_intro"',
        'id="login_form"',
        'for="login_email"',
        'id="login_email" type="email" autocomplete="username" required',
        'for="login_password"',
        'id="login_password" type="password" autocomplete="current-password" required',
        'id="login_error" class="auth-error" role="alert" aria-live="polite"',
        'id="login_submit" class="primary auth-submit" type="submit"',
        "<span>登录</span>",
    ]:
        assert marker in auth

    for forbidden in [
        "注册",
        "prototype",
        "mock",
        "debug",
        "simulate error state",
        "static data",
    ]:
        assert forbidden not in auth.lower()

    for marker in [
        "const errorEl=document.getElementById('login_error');",
        "if(errorEl) errorEl.textContent='';",
        "button.classList.add('is-loading');",
        "button.setAttribute('aria-busy','true');",
        'class="auth-button-spinner"',
        "if(errorEl) errorEl.textContent=await err(res);",
        "routeToOperationsHome();",
        "button.classList.remove('is-loading');",
        "button.removeAttribute('aria-busy');",
        "button.innerHTML=oldHtml;",
    ]:
        assert marker in page

    for selector in [
        "body .auth-screen",
        "body .auth-panel",
        "body .auth-panel-head",
        "body .auth-brand",
        "body .auth-brand-mark",
        "body .auth-kicker",
        "body .auth-field input:focus",
        "body .auth-error:not(:empty)",
        "body .auth-submit",
        "body .auth-submit.is-loading",
        "body .auth-button-spinner",
    ]:
        assert selector in css

    assert "box-shadow: var(--phase21-focus);" in css
    assert "animation: monitor-spin 0.8s linear infinite;" in css


def test_cr074_page_refresh_actions_are_deduplicated_icon_only():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    header = page[page.index("<header>") : page.index("</header>", page.index("<header>"))]

    assert 'id="global_refresh_button"' in header
    assert 'class="secondary btn-icon refresh-icon-button"' in header
    assert 'onclick="refreshActiveSection(this)"' in header
    assert 'aria-label="Refresh current page"' in header
    assert '<use href="#icon-refresh"></use>' in header

    for duplicate_button in [
        ">刷新当前页面</button>",
        ">刷新首页</button>",
        ">刷新账号</button>",
        ">刷新代理</button>",
        ">刷新 AI 接入</button>",
        ">刷新规则</button>",
        ">刷新配置</button>",
        ">刷新模板</button>",
        ">刷新策略</button>",
        ">刷新任务中心</button>",
        ">刷新历史</button>",
        ">刷新预览</button>",
        ">刷新日志</button>",
        ">刷新详情</button>",
        ">刷新调度时间</button>",
    ]:
        assert duplicate_button not in page

    for scoped_refresh in [
        'aria-label="Refresh schedule times"',
        'aria-label="Refresh delivery history"',
        'aria-label="Refresh preview"',
        'aria-label="Refresh logs"',
        'aria-label="Refresh run detail"',
    ]:
        assert scoped_refresh in page

    assert page.count('class="secondary btn-icon refresh-icon-button"') == 6
    assert "async function refreshActiveSection(button=null)" in page
    assert "await loadSectionData(tab);" in page
    assert "button || document.getElementById('global_refresh_button')" in page
    assert "if(!btn.classList.contains('refresh-icon-button'))" in page
    assert "const remaining=650 - (Date.now() - startedAt);" in page
    assert "async function refreshJobSchedule(button=null)" in page
    assert "async function refreshCurrentRunLogs(button=null)" in page
    assert "async function refreshSelectedEmailDeliveryHistory(button=null)" in page
    assert "async function refreshRunDetail(button=null)" in page
    assert ".refresh-icon-button {" in css
    assert ".refresh-icon-button.is-loading .refresh-icon" in css
    assert "animation: monitor-spin 0.8s linear infinite;" in css


def test_cr087_explanatory_helper_question_marks_are_removed_without_restoring_copy():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")

    for removed_helper_artifact in [
        "function helperTooltip(text, label='说明')",
        "function initializeHelperTooltips(root=document)",
        "function setHelperTooltipText(id, text)",
        "function closeHelperTooltips()",
        "helperTooltip(",
        "setHelperTooltipText(",
        "closeHelperTooltips();",
        "initializeHelperTooltips();",
        "helper.classList.toggle('is-open')",
        "if(helper && ['Enter',' '].includes(event.key))",
        'class="helper-tooltip"',
        "helper-tooltip",
        "data-tooltip",
        "label-with-help",
        "title-with-help",
        "status-inline-help",
        'id="ai_status" class="small" style="margin-top:8px"',
        "const status=document.getElementById('ai_status');",
        "if(status) status.textContent = '系统会把固定输出结构、评估规则和每条采集内容一起交给模型；AI 接入异常时，内容进入待人工复核。';",
        "if(!hasStructured && prompt && prompt !== defaultAIPrompt){",
        "if(status) status.textContent='已载入旧版完整 Prompt。建议按分段规则确认后重新保存，便于后续维护。';",
    ]:
        assert removed_helper_artifact not in page
        assert removed_helper_artifact not in css

    for removed_helper_copy in [
        "查看今日任务、运行、报告与邮件交付状态。",
        "新建和管理律所舆情监控任务。",
        "管理平台账号、登录状态和代理绑定。",
        "管理代理来源、并发上限和可用状态。",
        "管理模型连接资源和默认接入状态。",
        "管理舆情判断规则和默认初筛规则。",
        "按任务汇总运行、AI 评估、报告和交付状态。",
        "检查基础配置、运行状态和报告链路，便于管理员快速定位待处理项。",
        "当前首版以平台搜索词采集为主。平台搜索词会用于平台搜索；律所名称、别名和排除词用于评估、报告和采集后过滤。",
        "新建账号时选择平台；保存后平台不可变更。",
        "该登录方式只作用于当前账号，任务绑定此账号后会复用这里保存的登录态。",
        "保存账号会更新账号资料；登录是否可用以登录维护和账号检测结果为准。",
        "代理 URL 保存后仅显示掩码。",
        "密钥保存后不回显，连接测试会记录最近状态。",
        "查看规则状态、默认规则和最近测试结果。",
        "调整规则或样例内容后，点击“测试评估规则”查看模型返回的结构化结果。",
        "可修改规则或样例内容后点击“测试评估规则”。",
        "切换分组或运行记录，不影响详情查看。",
        "检测通过后自动刷新",
    ]:
        assert removed_helper_copy not in page

    assert '<section id="reports"' not in page
    assert 'id="run_detail_drawer" class="drawer run-detail-drawer"' in page
    assert ".drawer-scroll-body" in page
    assert 'id="global_refresh_button"' in page

    for preserved_operational_marker in [
        'id="social_account_platform"',
        'id="social_account_login_type"',
        'id="account_save_button"',
        'id="proxy_drawer"',
        'id="ai_profile_drawer"',
        'id="task_center_panel"',
        'data-run-detail-tab="overview"',
        'data-run-detail-tab="logs"',
        'data-run-detail-tab="contents"',
        'data-run-detail-tab="ai"',
        'data-run-detail-tab="reports"',
        'data-run-detail-tab="email"',
        "未保存密码",
        "点击“生成登录二维码”后在这里扫码。",
        "清空筛选",
        "保存任务",
        "保存代理",
        "保存 AI 接入",
        "保存配置",
        "保存模板",
    ]:
        assert preserved_operational_marker in page


def test_phase_13c_operations_home_responsive_role_views():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")

    dashboard_section = page.split('<section id="dashboard" class="active">', 1)[1].split('<section id="jobs">', 1)[0]
    doctor_section = page.split('<section id="doctor">', 1)[1].split('<div id="toast"', 1)[0]
    load_section = page.split("function loadSectionData(tab){", 1)[1].split("function canMenu", 1)[0]

    assert 'id="operations_home_admin_health" class="operations-admin-health admin-entry" data-menu-key="system_diagnostics"' in dashboard_section
    assert "renderOperationsAdminHealth(model)" in page
    assert "function renderOperationsAdminHealth(model)" in page
    assert 'data-shortcut-tab="doctor" data-menu-key="system_diagnostics">查看系统诊断' in page

    for diagnostic_mount in [
        'id="readiness_summary"',
        'id="ops_checklist"',
        'id="readiness_actions"',
        'id="readiness_table"',
        'id="scheduler_status"',
        'id="platform_status_table"',
    ]:
        assert diagnostic_mount not in dashboard_section
        assert diagnostic_mount in doctor_section

    assert "if(tab==='dashboard') {\n        return loadDashboard();\n      }" in load_section
    assert "if(tab==='doctor') {\n        return Promise.all([\n          loadDoctor(),\n          loadReadiness(),\n          loadSchedulerStatus(),\n          loadPlatformStatus(),\n          loadSystemChecklist()\n        ]);\n      }" in load_section
    assert "safeLoad('运行状态', loadReadiness)" not in page
    assert "addPermittedLoad(loads, '运行状态', 'system_diagnostics', loadReadiness)" in page

    for selector in [
        ".operations-admin-health",
        ".operations-resource-entry",
        ".operations-resource-card",
        ".operations-window-toggle",
        ".operations-breakdown-grid",
        ".operations-chart-card",
        ".operations-chart-surface",
    ]:
        assert selector in css

    tablet_block = css.split("@media (max-width: 1279px)", 1)[1].split("@media (max-width: 767px)", 1)[0]
    tablet_height_block = css.split("@media (min-width: 768px) and (max-width: 1279px)", 1)[1].split("@media (max-width: 767px)", 1)[0]
    mobile_block = css.split("@media (max-width: 767px)", 1)[1]
    final_cr105_block = css.split("/* CR-105 ECharts dashboard rebaseline: local chart modules. */", 1)[1]
    assert ".operations-home {\n    width: 100%;\n    overflow: hidden;\n  }" in tablet_block
    assert "body .operations-metric-grid,\n  .operations-metric-grid {\n    grid-template-columns: repeat(5, minmax(0, 1fr)) !important;" in tablet_block
    assert ".operations-resource-signals {\n    grid-template-columns: repeat(2, minmax(0, 1fr));\n  }" in tablet_block
    assert "body #dashboard.active {\n    min-height: calc(100vh - 68px);\n    max-height: calc(100vh - 68px);\n    overflow: hidden;" in tablet_height_block
    assert "body #dashboard.active > .operations-home {\n    height: calc(100vh - 140px);\n    max-height: calc(100vh - 140px);" in final_cr105_block
    assert "body #dashboard.active .operations-home-lower {\n    grid-template-columns: minmax(0, 1.28fr) minmax(180px, 0.62fr);" in final_cr105_block
    assert 'grid-template-areas:\n    "trend trend attention"\n    "platforms delivery resource";' in final_cr105_block
    assert ".account-menu-copy,\n  .account-menu-caret {\n    display: none;\n  }" in mobile_block
    assert "body header .header-actions {\n    display: contents !important;" in mobile_block
    top_status_block = mobile_block.split("body header #top_status {", 1)[1].split("}", 1)[0]
    assert "grid-area: status;" in top_status_block
    assert "body header .header-actions > .account-area {\n    min-width: 0;\n    width: auto;" in mobile_block
    assert ".operations-home-strip-head {\n    grid-template-columns: minmax(0, 1fr) auto;" in mobile_block
    assert ".operations-home-strip-head span:last-child {\n    display: none;" in mobile_block
    assert "body .operations-metric-grid,\n  .operations-metric-grid {\n    grid-template-columns: repeat(2, minmax(0, 1fr)) !important;" in mobile_block
    assert "body #dashboard.active .operations-home-lower {\n    grid-template-columns: minmax(0, 1fr);\n    grid-template-areas:\n      \"trend\"\n      \"attention\"\n      \"platforms\"\n      \"delivery\"\n      \"resource\";" in final_cr105_block
    assert "body #dashboard.active .operations-window-toggle {\n    margin-left: auto;" in final_cr105_block
    assert "body #dashboard.active .operations-chart-surface {\n    min-height: 168px;" in final_cr105_block
    assert "body #dashboard.active .operations-resource-signals {\n    grid-template-columns: repeat(2, minmax(0, 1fr));" in final_cr105_block
    assert "/* CR-105 ECharts dashboard rebaseline: final inline cascade guard. */" in page
    inline_cr105_block = page.split("/* CR-105 ECharts dashboard rebaseline: final inline cascade guard. */", 1)[1].split("</style>", 1)[0]
    assert "body #dashboard.active > .operations-home {\n        grid-template-columns:minmax(0,1fr);" in inline_cr105_block
    assert 'grid-template-areas:\n          "meta"\n          "metrics"\n          "lower";' in inline_cr105_block
    assert 'grid-template-areas:\n          "trend"\n          "attention"\n          "platforms"\n          "delivery"\n          "resource";' in inline_cr105_block
    assert "body #dashboard.active .operations-resource-card-primary .operations-chart-surface {\n      min-height:96px;" in inline_cr105_block
    assert "function resourceSignalTone(status)" in page

    assert "资源由管理员维护" in page
    assert "social_accounts_total" not in dashboard_section.split("function renderOperationsAdminHealth", 1)[-1]
    assert "new Chart(" not in page
    assert "chart.js" not in page.lower()
    assert "email_delivery_logs" not in dashboard_section
    assert "job_snapshot_json" not in page
    assert "crawl_runs.visibility" not in page


def test_phase_14_run_center_visibility_fields_migrate_and_backfill():
    init_db()
    jobs_snapshot = _snapshot_monitor_jobs()
    snapshots = {
        "reports": _snapshot_table("reports"),
        "crawl_runs": _snapshot_table("crawl_runs"),
    }
    _clear_monitor_jobs()
    try:
        with get_conn() as conn:
            columns = _table_columns(conn, "crawl_runs")
            assert {"visibility", "run_type", "archived_at", "archived_by"} <= columns
            index_names = {
                row["name"]
                for row in conn.execute("PRAGMA index_list(crawl_runs)").fetchall()
            }
            assert "idx_crawl_runs_visibility" in index_names
            assert "idx_crawl_runs_type_status" in index_names

        job = save_job(
            {
                "law_firm_name": "Phase14迁移律所",
                "aliases": [],
                "exclude_words": [],
                "keywords": ["Phase14迁移律所投诉"],
                "platforms": ["dy"],
                "recipients": [],
            }
        )
        run_id = create_run(job["id"], {"job_id": job["id"], "law_firm_name": job["law_firm_name"]})
        finish_run(run_id, "success", {"job_id": job["id"], "law_firm_name": job["law_firm_name"]})
        report = create_report(run_id, job, {"job_id": job["id"], "law_firm_name": job["law_firm_name"], "platforms": ["dy"]})

        run = get_run(run_id)
        assert run["visibility"] == "visible"
        assert run["run_type"] == "scheduled"
        assert run["archived_at"] is None
        assert run["archived_by"] is None
        assert any(item["id"] == run_id and item["visibility"] == "visible" for item in list_runs(0))
        assert get_report(report["id"])["run_id"] == run_id

        with get_conn() as conn:
            conn.execute("UPDATE crawl_runs SET visibility='', run_type='' WHERE id=?", (run_id,))
        init_db()

        migrated = get_run(run_id)
        assert migrated["visibility"] == "visible"
        assert migrated["run_type"] == "scheduled"
        with get_conn() as conn:
            observed_visibility = {
                row["visibility"]
                for row in conn.execute(
                    "SELECT DISTINCT visibility FROM crawl_runs WHERE visibility IS NOT NULL"
                ).fetchall()
            }
            observed_run_types = {
                row["run_type"]
                for row in conn.execute(
                    "SELECT DISTINCT run_type FROM crawl_runs WHERE run_type IS NOT NULL"
                ).fetchall()
            }
        assert observed_visibility <= {"visible", "archived"}
        assert observed_run_types <= {"scheduled", "manual", "test"}
    finally:
        _restore_table("reports", snapshots["reports"])
        _restore_table("crawl_runs", snapshots["crawl_runs"])
        _restore_monitor_jobs(jobs_snapshot)


def test_phase_15a_run_center_api_pagination_filters_archive_and_scope():
    from api import main as api_main

    init_db()
    snapshots = {
        "reports": _snapshot_table("reports"),
        "crawl_runs": _snapshot_table("crawl_runs"),
        "audit_logs": _snapshot_table("audit_logs"),
        "user_sessions": _snapshot_table("user_sessions"),
        "users": _snapshot_table("users"),
    }
    jobs_snapshot = _snapshot_monitor_jobs()
    try:
        with get_conn() as conn:
            for table in ["reports", "crawl_runs", "audit_logs", "user_sessions", "users"]:
                conn.execute(f"DELETE FROM {table}")
        _clear_monitor_jobs()

        admin = bootstrap_admin_from_env("phase15a-admin@example.com", "AdminPass123!", "Phase 15A Admin")
        user1 = save_user(
            {
                "email": "phase15a-user1@example.com",
                "display_name": "Phase 15A User One",
                "password": "UserPass123!",
                "role": "normal",
            },
            actor_id=int(admin["id"]),
        )
        user2 = save_user(
            {
                "email": "phase15a-user2@example.com",
                "display_name": "Phase 15A User Two",
                "password": "UserPass456!",
                "role": "normal",
            },
            actor_id=int(admin["id"]),
        )
        job1 = save_job(
            {
                "law_firm_name": "海安律所",
                "keywords": ["海安律所避雷"],
                "platforms": ["dy"],
                "recipients": [],
            },
            actor=user1,
        )
        job2 = save_job(
            {
                "law_firm_name": "恒泰律所",
                "keywords": ["恒泰律所投诉"],
                "platforms": ["ks"],
                "recipients": [],
            },
            actor=user2,
        )
        run_visible = create_run(job1["id"], {"job_id": job1["id"], "law_firm_name": job1["law_firm_name"], "platforms": ["dy"], "keywords": job1["keywords"]})
        finish_run(run_visible, "success", {"job_id": job1["id"], "law_firm_name": job1["law_firm_name"], "platforms": ["dy"], "keywords": job1["keywords"]})
        run_archived = create_run(job1["id"], {"job_id": job1["id"], "law_firm_name": job1["law_firm_name"], "platforms": ["ks"], "keywords": job1["keywords"]})
        finish_run(run_archived, "failed", {"job_id": job1["id"], "law_firm_name": job1["law_firm_name"], "platforms": ["ks"], "keywords": job1["keywords"]}, "登录态失效")
        run_other_user = create_run(job2["id"], {"job_id": job2["id"], "law_firm_name": job2["law_firm_name"], "platforms": ["ks"], "keywords": job2["keywords"]})
        finish_run(run_other_user, "success", {"job_id": job2["id"], "law_firm_name": job2["law_firm_name"], "platforms": ["ks"], "keywords": job2["keywords"]})
        report = create_report(run_visible, job1, {"job_id": job1["id"], "law_firm_name": job1["law_firm_name"], "platforms": ["dy"]})
        with get_conn() as conn:
            conn.execute("UPDATE crawl_runs SET visibility='archived', archived_at=?, archived_by=? WHERE id=?", (datetime.now(timezone.utc).isoformat(), admin["id"], run_archived))
            conn.execute("UPDATE crawl_runs SET run_type='manual' WHERE id=?", (run_visible,))
            conn.execute("UPDATE crawl_runs SET run_type='test' WHERE id=?", (run_other_user,))

        transport = httpx.ASGITransport(app=api_main.app)

        async def exercise() -> None:
            async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as normal_client:
                login = await normal_client.post(
                    "/api/auth/login",
                    json={"email": "phase15a-user1@example.com", "password": "UserPass123!"},
                )
                assert login.status_code == 200
                default_runs = await normal_client.get("/api/monitor/runs")
                assert default_runs.status_code == 200
                default_payload = default_runs.json()
                assert {item["id"] for item in default_payload["runs"]} == {run_visible}
                assert default_payload["pagination"]["total"] == 1
                assert default_payload["filters"]["visibility"] == "visible"
                assert default_payload["runs"][0]["run_type"] == "manual"
                assert "visibility" in default_payload["runs"][0]

                archived_forbidden = await normal_client.get("/api/monitor/runs", params={"visibility": "archived"})
                assert archived_forbidden.status_code == 403
                archive_forbidden = await normal_client.post(f"/api/monitor/runs/{run_visible}/archive")
                assert archive_forbidden.status_code == 403
                archived_logs = await normal_client.get(f"/api/monitor/runs/{run_archived}/logs")
                assert archived_logs.status_code == 404
                other_logs = await normal_client.get(f"/api/monitor/runs/{run_other_user}/logs")
                assert other_logs.status_code == 404

            async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as admin_client:
                login = await admin_client.post(
                    "/api/auth/login",
                    json={"email": "phase15a-admin@example.com", "password": "AdminPass123!"},
                )
                assert login.status_code == 200
                default_runs = await admin_client.get("/api/monitor/runs")
                assert {item["id"] for item in default_runs.json()["runs"]} == {run_other_user, run_visible}
                operational_runs = await admin_client.get("/api/monitor/runs", params={"run_type": "operational"})
                assert operational_runs.status_code == 200
                assert {item["id"] for item in operational_runs.json()["runs"]} == {run_visible}
                assert operational_runs.json()["filters"]["run_type"] == "operational"
                archived_runs = await admin_client.get("/api/monitor/runs", params={"visibility": "archived"})
                assert archived_runs.status_code == 200
                assert {item["id"] for item in archived_runs.json()["runs"]} == {run_archived}

                filtered = await admin_client.get(
                    "/api/monitor/runs",
                    params={
                        "visibility": "all",
                        "law_firm": "海安",
                        "status": "success",
                        "platform": "dy",
                        "run_type": "manual",
                        "page": 1,
                        "limit": 1,
                    },
                )
                filtered_payload = filtered.json()
                assert [item["id"] for item in filtered_payload["runs"]] == [run_visible]
                assert filtered_payload["pagination"] == {"page": 1, "per_page": 1, "total": 1, "total_pages": 1}
                assert filtered_payload["filters"]["visibility"] == "all"

                archived = await admin_client.post(f"/api/monitor/runs/{run_visible}/archive")
                assert archived.status_code == 200
                assert archived.json()["run"]["visibility"] == "archived"
                assert archived.json()["run"]["archived_at"]
                assert archived.json()["run"]["archived_by"] == admin["id"]
                default_after_archive = await admin_client.get("/api/monitor/runs")
                assert run_visible not in {item["id"] for item in default_after_archive.json()["runs"]}

                restored = await admin_client.post(f"/api/monitor/runs/{run_visible}/restore")
                assert restored.status_code == 200
                assert restored.json()["run"]["visibility"] == "visible"
                assert restored.json()["run"]["archived_at"] is None
                assert restored.json()["run"]["archived_by"] is None

                logs = await admin_client.get(f"/api/monitor/runs/{run_visible}/logs")
                assert logs.status_code == 200
                report_detail = await admin_client.get(f"/api/monitor/reports/{report['id']}")
                assert report_detail.status_code == 200
                assert report_detail.json()["report"]["run_id"] == run_visible

        asyncio.run(exercise())
    finally:
        _restore_table("reports", snapshots["reports"])
        _restore_table("crawl_runs", snapshots["crawl_runs"])
        _restore_table("audit_logs", snapshots["audit_logs"])
        _restore_table("user_sessions", snapshots["user_sessions"])
        _restore_table("users", snapshots["users"])
        _restore_monitor_jobs(jobs_snapshot)


def test_phase_15b_run_center_frontend_filters_pagination_archive_controls():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")

    assert 'id="run_task_filter"' in page
    assert 'id="run_status_filter"' in page
    assert 'id="run_platform_filter"' in page
    assert 'id="run_type_filter"' in page
    assert 'id="run_visibility_filter"' in page
    assert 'id="run_date_from"' in page
    assert 'id="run_date_to"' in page
    assert 'id="run_pagination"' in page
    assert 'id="run_filter_summary"' in page
    assert "function runQueryParams()" in page
    assert "qs.set('page', String(runListState.page || 1))" in page
    assert "qs.set('visibility', visibility)" in page
    assert "qs.set('task_id', query)" in page
    assert "qs.set('law_firm', query)" in page
    assert "qs.set('run_type', runType)" in page
    assert '<option value="operational" selected>运营记录</option>' in page
    assert "set('run_type_filter','operational')" in page
    assert "function renderRunPagination(pagination)" in page
    assert "function applyRunFilters()" in page
    assert "function clearRunFilters()" in page
    assert "function archiveRun(id)" in page
    assert "function restoreRun(id)" in page
    assert "confirm('确认归档这条运行记录？归档后默认列表将不再显示它。')" in page
    assert "confirm('确认恢复这条运行记录？恢复后它会回到默认可见列表。')" in page
    assert "applyRunCenterRoleMode()" in page
    assert "document.querySelectorAll('.run-admin-control')" in page
    assert "if(!admin)" in page and "set('run_visibility_filter','visible')" in page
    assert "runTypeBadge(r.run_type)" in page
    assert "runVisibilityBadge(r.visibility)" in page
    assert "测试/诊断" in page
    assert "默认可见" in page
    assert "已归档" in page
    assert "全部记录" in page
    assert "loadRunLogs" in page
    assert "loadRunLogs(${Number(r.id)})" not in page
    assert "viewRunLeads(${Number(r.id)})" not in page
    assert "function viewRunLeads(id)" not in page
    assert "new URLSearchParams({run_id:String(id), limit:'0', risk:val('lead_status_filter')})" not in page
    assert "function leadScopeForRun(id, leads)" not in page
    assert "copyCurrentRunLogs" in page
    assert "downloadCurrentRunLogs" in page
    assert "复制日志" in page
    assert "下载日志" in page
    assert "/runs/'+id+'/archive" in page
    assert "/runs/'+id+'/restore" in page
    assert ".run-center-meta" in css
    assert ".run-pagination" in css
    assert ".run-actions" in css
    task_center = _monitor_section(page, "runs")
    assert "email_delivery_logs" not in task_center
    assert "job_snapshot_json" not in page


def test_phase_19d_run_center_frontend_progress_polling_hooks():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")

    assert "let runPollInFlight = false;" in page
    assert "async function loadRuns(options={})" in page
    assert "const silent=!!options.silent;" in page
    assert "if(!silent){" in page
    assert "setLoadingMarkup('runs_table', tableSkeleton(5, 8));" in page
    assert "syncRunPolling(runs);" in page
    assert "return runs.some(run=>run.status==='running');" in page
    assert "function startRunPolling()" in page
    assert "loadRuns({silent:true})" in page
    assert "function stopRunPolling()" in page
    assert "function syncRunPolling(runs=[])" in page
    assert "function isRunsPageActive()" in page
    assert "if(!isRunsPageActive()){ stopRunPolling(); return; }" in page
    assert "if(!running){" in page
    assert "await Promise.all([loadJobs(), loadReports(), loadReadiness(), loadPlatformStatus(), loadDoctor()]);" not in page
    assert "await Promise.all([loadJobs(), loadRuns(), loadReadiness(), loadPlatformStatus(), loadDoctor()]);" in page
    assert "if(btn.dataset.tab!=='runs') stopRunPolling();" in page
    assert "if(target!=='run_records_view') stopRunPolling();" not in page
    assert "activeRunCenterView()==='run_records_view' ? loadRuns() : loadReports()" not in page
    run_poll_block = page[page.index("function startRunPolling()") : page.index("function stopRunPolling()")]
    assert "let rounds=40" not in run_poll_block
    assert "rounds++" not in run_poll_block
    assert "rounds--" not in run_poll_block

    assert "function runProgressDetails(run)" in page
    assert "if(run.status!=='running') return [];" in page
    assert "function compactRunProgressText(run, details=[])" in page
    assert "function compactRunStatusLabel(run)" in page
    assert "compactRunStatusLabel(run)" in page
    assert "run.display_status || formatRunStatus(status)" not in page
    assert "function runCollectionProgressText(progress)" in page
    assert "function runAiProgressText(progress)" in page
    assert "run.collection_progress || summary.collection_progress" in page
    assert "run.ai_progress || summary.ai_progress" in page
    assert "临时 ${esc(progressValue)}" in page
    assert "采集中：已观察" in page
    assert "采集中 ${Number(collection.raw_items_seen||0)}（临时）" in page
    assert "AI ${progress.final?'完成':'评估中'}" in page
    assert "AI ${Number(ai.evaluated_items||0)}/${Number(ai.total_candidates||0)}" in page
    assert "if(run.status==='success') return '';" in page
    assert "return run.status==='running' ? ((details||[])[0] || '') : '';" in page
    assert "报告生成中" in page
    assert "邮件发送中" in page
    assert "运行超时" in page
    assert "已取消" in page
    assert "已完成" in page
    assert "执行中断" in page
    assert "if(run.status==='success') parts.push('已完成')" not in page
    assert "if(run.status==='cancelled') parts.push('已取消')" not in page
    assert "if(run.status==='interrupted') parts.push('执行中断')" not in page
    assert "interrupted:'执行中断'" in page

    assert "viewRunLeads(${Number(r.id)})" not in page
    assert "loadRunLogs(${Number(r.id)})" not in page
    assert "stopRun(${Number(r.id)})" not in page
    assert "archiveRun(${Number(r.id)})" not in page
    assert "restoreRun(${Number(r.id)})" not in page
    assert "function stopRunFromDetail(id)" in page
    assert "function archiveRunFromDetail(id)" in page
    assert "function restoreRunFromDetail(id)" in page
    assert "const visibility=isAdminUser() ? (val('run_visibility_filter') || 'visible') : 'visible';" in page
    assert "applyRunCenterRoleMode()" in page

    assert ".run-status-stack" in css
    assert "width: fit-content;" in css
    assert ".run-status-badge" in css
    assert "display: inline-flex;" in css
    assert "#runs_table .run-record-table th.col-status" in css
    assert "#runs_table .run-record-table td.col-status" in css
    assert "max-width: 86px;" in css
    assert "run-status-badge ${cls}" in page
    assert "status run-status-badge" not in page
    assert ".run-progress-lines" in css
    assert ".run-progress-note" in css
    assert ".run-actions" in css
    assert "white-space: normal;" in css
    assert "flex-wrap: wrap;" in css


def test_phase_19d_run_display_status_labels_match_progress_ui_terms():
    init_db()
    jobs_snapshot = _snapshot_monitor_jobs()
    snapshots = {"crawl_runs": _snapshot_table("crawl_runs")}
    _clear_monitor_jobs()
    try:
        job = save_job(
            {
                "law_firm_name": "Phase19D状态律所",
                "keywords": ["Phase19D状态律所投诉"],
                "platforms": ["dy"],
                "enabled": True,
            }
        )
        expected = {
            "success": "已完成",
            "timeout": "运行超时",
            "cancelled": "已取消",
            "interrupted": "执行中断",
        }
        for status, label in expected.items():
            run_id = create_run(job["id"], {"job_id": job["id"], "platforms": ["dy"]})
            finish_run(run_id, status, {"job_id": job["id"], "platforms": ["dy"]})
            run = get_run(run_id)
            assert run and run["display_status"] == label
    finally:
        _restore_table("crawl_runs", snapshots["crawl_runs"])
        _restore_monitor_jobs(jobs_snapshot)


def test_phase_16_email_delivery_logs_schema_window_keys_and_compatibility():
    init_db()
    jobs_snapshot = _snapshot_monitor_jobs()
    snapshots = {
        "email_delivery_logs": _snapshot_table("email_delivery_logs"),
        "reports": _snapshot_table("reports"),
        "crawl_runs": _snapshot_table("crawl_runs"),
    }
    _clear_monitor_jobs()
    try:
        with get_conn() as conn:
            columns = _table_columns(conn, "email_delivery_logs")
            assert {
                "workspace_id",
                "job_id",
                "report_id",
                "send_window_key",
                "send_type",
                "sent_by",
                "sent_at",
                "status",
                "error_message",
                "recipients_json",
                "trigger_source",
                "effective_recipients_json",
                "effective_recipient_source",
                "email_template_id",
                "email_template_name",
                "email_template_source",
                "email_subject_template",
                "created_at",
            } <= columns
            index_names = {
                row["name"]
                for row in conn.execute("PRAGMA index_list(email_delivery_logs)").fetchall()
            }
            assert "idx_email_delivery_job_window" in index_names
            assert "idx_email_delivery_report" in index_names
            assert "idx_email_delivery_status" in index_names
            assert "idx_email_delivery_auto_window_unique" in index_names

        job = save_job(
            {
                "law_firm_name": "Phase16邮件律所",
                "keywords": ["Phase16邮件律所投诉"],
                "platforms": ["dy"],
                "recipients": ["ops@example.com"],
            }
        )
        run_id = create_run(job["id"], {"job_id": job["id"], "law_firm_name": job["law_firm_name"]})
        finish_run(run_id, "success", {"job_id": job["id"], "law_firm_name": job["law_firm_name"]})
        report = create_report(run_id, job, {"job_id": job["id"], "law_firm_name": job["law_firm_name"]})
        send_at = datetime(2026, 6, 16, 15, 45, tzinfo=timezone.utc)

        assert email_send_window_key(job["id"], "daily", send_at) == f"{job['id']}_2026-06-16"
        for frequency in ["6h", "12h", "cron"]:
            assert email_send_window_key(job["id"], frequency, send_at) == f"{job['id']}_2026-06-16_15"

        auto = record_email_delivery_log(
            {
                "workspace_id": job["workspace_id"],
                "job_id": job["id"],
                "report_id": report["id"],
                "send_window_key": email_send_window_key(job["id"], "daily", send_at),
                "send_type": "auto",
                "sent_at": send_at.isoformat(),
                "status": "sent",
                "error_message": "smtp_password=super-secret proxy_url=http://user:pass@example.test:8080",
                "recipients": ["ops@example.com"],
            }
        )
        assert auto["send_type"] == "auto"
        assert auto["status"] == "sent"
        assert auto["send_window_key"] == f"{job['id']}_2026-06-16"
        assert auto["recipients"] == ["ops@example.com"]
        assert auto["trigger_source"] == "scheduler_auto"
        assert auto["effective_recipients"] == []
        assert auto["effective_recipient_source"] == "limited_context"
        assert "super-secret" not in auto["error_message"]
        assert "pass@example" not in auto["error_message"]

        with pytest.raises(sqlite3.IntegrityError):
            record_email_delivery_log(
                {
                    "workspace_id": job["workspace_id"],
                    "job_id": job["id"],
                    "report_id": report["id"],
                    "send_window_key": auto["send_window_key"],
                    "send_type": "auto",
                    "status": "pending",
                    "recipients": ["ops@example.com"],
                }
            )

        failed_window = email_send_window_key(job["id"], "6h", datetime(2026, 6, 16, 16, 0, tzinfo=timezone.utc))
        failed_auto = record_email_delivery_log(
            {
                "workspace_id": job["workspace_id"],
                "job_id": job["id"],
                "report_id": report["id"],
                "send_window_key": failed_window,
                "send_type": "auto",
                "status": "failed",
                "error_message": "token=secret-token",
                "recipients_json": json.dumps(["ops@example.com"]),
            }
        )
        retried_auto = record_email_delivery_log(
            {
                "workspace_id": job["workspace_id"],
                "job_id": job["id"],
                "report_id": report["id"],
                "send_window_key": failed_window,
                "send_type": "auto",
                "status": "sending",
                "recipients": ["ops@example.com"],
            }
        )
        assert failed_auto["status"] == "failed"
        assert retried_auto["status"] == "sending"

        manual_one = record_email_delivery_log(
            {
                "workspace_id": job["workspace_id"],
                "job_id": job["id"],
                "report_id": report["id"],
                "send_window_key": auto["send_window_key"],
                "send_type": "manual_resend",
                "sent_by": 1,
                "status": "sent",
                "recipients": ["ops@example.com"],
            }
        )
        manual_two = record_email_delivery_log(
            {
                "workspace_id": job["workspace_id"],
                "job_id": job["id"],
                "report_id": report["id"],
                "send_window_key": auto["send_window_key"],
                "send_type": "manual_resend",
                "sent_by": 1,
                "status": "sent",
                "recipients": ["ops@example.com"],
            }
        )
        assert manual_one["send_type"] == "manual_resend"
        assert manual_two["send_type"] == "manual_resend"

        with get_conn() as conn:
            conn.execute(
                "UPDATE reports SET email_status='failed', email_error=? WHERE id=?",
                ("smtp_password=super-secret", report["id"]),
            )
            raw_log = conn.execute(
                "SELECT error_message, recipients_json FROM email_delivery_logs WHERE id=?",
                (auto["id"],),
            ).fetchone()
        assert "super-secret" not in raw_log["error_message"]
        assert "pass@example" not in raw_log["error_message"]
        assert "ops@example.com" in raw_log["recipients_json"]
        refreshed_report = get_report(report["id"])
        assert refreshed_report["email_status"] == "failed"
        assert "super-secret" not in refreshed_report["email_error"]
        assert any(item["id"] == report["id"] and item["email_status"] == "failed" for item in list_reports(0))
        logs = list_email_delivery_logs(report_id=report["id"], limit=10)
        assert {item["send_type"] for item in logs} == {"auto", "manual_resend"}
        assert all("super-secret" not in (item.get("error_message") or "") for item in logs)
    finally:
        _restore_table("email_delivery_logs", snapshots["email_delivery_logs"])
        _restore_table("reports", snapshots["reports"])
        _restore_table("crawl_runs", snapshots["crawl_runs"])
        _restore_monitor_jobs(jobs_snapshot)


def test_phase_17a_auto_delivery_is_idempotent_and_manual_resend_is_separate(monkeypatch):
    init_db()
    jobs_snapshot = _snapshot_monitor_jobs()
    snapshots = {
        "email_delivery_logs": _snapshot_table("email_delivery_logs"),
        "reports": _snapshot_table("reports"),
        "crawl_runs": _snapshot_table("crawl_runs"),
    }
    _clear_monitor_jobs()
    send_calls: list[int] = []
    try:
        with get_conn() as conn:
            for table in ["email_delivery_logs", "reports", "crawl_runs"]:
                conn.execute(f"DELETE FROM {table}")
        job = save_job(
            {
                "law_firm_name": "Phase17A交付律所",
                "keywords": ["Phase17A交付律所投诉"],
                "platforms": ["dy"],
                "recipients": ["ops@example.com"],
                "frequency": "daily",
                "email_time": "09:00",
            }
        )
        run_id = create_run(job["id"], {"job_id": job["id"], "law_firm_name": job["law_firm_name"]})
        finish_run(run_id, "success", {"job_id": job["id"], "law_firm_name": job["law_firm_name"]})
        report = create_report(run_id, job, {"job_id": job["id"], "law_firm_name": job["law_firm_name"]})
        send_at = datetime(2026, 6, 16, 9, 30, tzinfo=timezone.utc)

        def fake_send_report(job_arg, report_arg, allow_real_send=None):
            send_calls.append(int(report_arg["id"]))
            return True, None

        monkeypatch.setattr("api.monitoring.reporting.send_report", fake_send_report)

        ok_first, error_first, refreshed_first, log_first = send_report_with_delivery_log(
            job,
            report,
            send_type="auto",
            sent_at=send_at,
        )
        ok_second, error_second, refreshed_second, log_second = send_report_with_delivery_log(
            job,
            report,
            send_type="auto",
            sent_at=send_at,
        )
        manual_actor = {"id": 42, "workspace_id": job["workspace_id"], "role": "normal"}
        ok_manual, error_manual, refreshed_manual = resend_report_email(report["id"], actor=manual_actor)

        logs = list_email_delivery_logs(report_id=report["id"], limit=20)

        assert ok_first is True
        assert error_first in (None, "")
        assert refreshed_first["email_status"] == "sent"
        assert log_first and log_first["send_type"] == "auto"
        assert log_first["status"] == "sent"
        assert log_first["send_window_key"] == f"{job['id']}_2026-06-16"
        assert log_first["trigger_source"] == "scheduler_auto"
        assert log_first["recipients"] == ["ops@example.com"]
        assert log_first["effective_recipients"] == ["ops@example.com"]
        assert log_first["effective_recipient_source"] == "task_recipients"
        assert log_first["email_template_source"] in {"active_global_fallback", "default_renderer"}
        assert ok_second is False
        assert "已跳过重复发送" in (error_second or "")
        assert refreshed_second["email_status"] == "skipped"
        assert log_second is None
        assert ok_manual is True
        assert error_manual in (None, "")
        assert refreshed_manual["email_status"] == "sent"
        assert send_calls == [report["id"], report["id"]]
        assert sum(1 for item in logs if item["send_type"] == "auto" and item["status"] == "sent") == 1
        assert sum(1 for item in logs if item["send_type"] == "auto" and item["status"] == "skipped") == 1
        manual_logs = [item for item in logs if item["send_type"] == "manual_resend"]
        assert len(manual_logs) == 1
        assert manual_logs[0]["sent_by"] == manual_actor["id"]
        assert manual_logs[0]["status"] == "sent"
        assert manual_logs[0]["trigger_source"] == "manual_resend"
        assert manual_logs[0]["effective_recipients"] == ["ops@example.com"]
    finally:
        _restore_table("email_delivery_logs", snapshots["email_delivery_logs"])
        _restore_table("reports", snapshots["reports"])
        _restore_table("crawl_runs", snapshots["crawl_runs"])
        _restore_monitor_jobs(jobs_snapshot)


def test_phase_17_1_real_smtp_gate_skips_auto_delivery_with_default_recipients(monkeypatch):
    init_db()
    jobs_snapshot = _snapshot_monitor_jobs()
    snapshots = {
        "email_delivery_logs": _snapshot_table("email_delivery_logs"),
        "reports": _snapshot_table("reports"),
        "crawl_runs": _snapshot_table("crawl_runs"),
        "email_configs": _snapshot_singleton_table("email_configs"),
        "system_settings": _snapshot_table("system_settings"),
    }
    _clear_monitor_jobs()
    try:
        with get_conn() as conn:
            for table in ["email_delivery_logs", "reports", "crawl_runs"]:
                conn.execute(f"DELETE FROM {table}")
        save_runtime_settings({"real_email_delivery": False}, actor_id=1)
        monkeypatch.delenv("MONITOR_ALLOW_REAL_EMAIL_SEND", raising=False)
        save_email_config(
            {
                "smtp_host": "smtp.real-looking.example",
                "smtp_port": 465,
                "encryption": "ssl",
                "sender": "sender@example.com",
                "username": "sender@example.com",
                "password": "super-secret",
                "default_recipients": ["fallback@example.com"],
            }
        )
        job = save_job(
            {
                "law_firm_name": "Phase17.1安全律所",
                "keywords": ["Phase17.1安全律所投诉"],
                "platforms": ["dy"],
                "recipients": [],
                "frequency": "daily",
                "email_time": "09:00",
            }
        )
        run_id = create_run(job["id"], {"job_id": job["id"], "law_firm_name": job["law_firm_name"]})
        finish_run(run_id, "success", {"job_id": job["id"], "law_firm_name": job["law_firm_name"]})
        report = create_report(run_id, job, {"job_id": job["id"], "law_firm_name": job["law_firm_name"]})

        ok, error, refreshed, log = send_report_with_delivery_log(job, report, send_type="auto")

        assert ok is False
        assert error == REAL_EMAIL_BLOCKED_MESSAGE
        assert refreshed["email_status"] == "skipped"
        assert log and log["status"] == "skipped"
        assert log["recipients"] == []
        assert log["effective_recipients"] == ["fallback@example.com"]
        assert log["effective_recipient_source"] == "global_default_fallback"
        assert log["trigger_source"] == "scheduler_auto"
    finally:
        _restore_table("email_delivery_logs", snapshots["email_delivery_logs"])
        _restore_table("reports", snapshots["reports"])
        _restore_table("crawl_runs", snapshots["crawl_runs"])
        _restore_table("system_settings", snapshots["system_settings"])
        _restore_singleton_table("email_configs", snapshots["email_configs"])
        _restore_monitor_jobs(jobs_snapshot)


def test_phase_17_1_manual_resend_and_mail_test_are_blocked_without_opt_in(monkeypatch):
    init_db()
    jobs_snapshot = _snapshot_monitor_jobs()
    snapshots = {
        "email_delivery_logs": _snapshot_table("email_delivery_logs"),
        "reports": _snapshot_table("reports"),
        "crawl_runs": _snapshot_table("crawl_runs"),
        "email_configs": _snapshot_singleton_table("email_configs"),
        "system_settings": _snapshot_table("system_settings"),
    }
    _clear_monitor_jobs()
    try:
        with get_conn() as conn:
            for table in ["email_delivery_logs", "reports", "crawl_runs"]:
                conn.execute(f"DELETE FROM {table}")
        save_runtime_settings({"real_email_delivery": False}, actor_id=1)
        monkeypatch.delenv("MONITOR_ALLOW_REAL_EMAIL_SEND", raising=False)
        save_email_config(
            {
                "smtp_host": "smtp.real-looking.example",
                "smtp_port": 465,
                "encryption": "ssl",
                "sender": "sender@example.com",
                "username": "sender@example.com",
                "password": "super-secret",
                "default_recipients": ["ops@example.com"],
            }
        )
        job = save_job(
            {
                "law_firm_name": "Phase17.1重发律所",
                "keywords": ["Phase17.1重发律所投诉"],
                "platforms": ["dy"],
                "recipients": ["ops@example.com"],
                "frequency": "daily",
                "email_time": "09:00",
            }
        )
        run_id = create_run(job["id"], {"job_id": job["id"], "law_firm_name": job["law_firm_name"]})
        finish_run(run_id, "success", {"job_id": job["id"], "law_firm_name": job["law_firm_name"]})
        report = create_report(run_id, job, {"job_id": job["id"], "law_firm_name": job["law_firm_name"]})

        ok, error, refreshed = resend_report_email(report["id"], actor={"id": 7, "workspace_id": job["workspace_id"], "role": "administrator"})
        logs = list_email_delivery_logs(report_id=report["id"], limit=5)

        assert ok is False
        assert error == REAL_EMAIL_BLOCKED_MESSAGE
        assert refreshed["email_status"] == "skipped"
        assert logs and logs[0]["send_type"] == "manual_resend"
        assert logs[0]["status"] == "skipped"
        assert logs[0]["trigger_source"] == "manual_resend"
        with pytest.raises(ValueError, match="真实邮件发送未启用"):
            send_test_email({})
    finally:
        _restore_table("email_delivery_logs", snapshots["email_delivery_logs"])
        _restore_table("reports", snapshots["reports"])
        _restore_table("crawl_runs", snapshots["crawl_runs"])
        _restore_table("system_settings", snapshots["system_settings"])
        _restore_singleton_table("email_configs", snapshots["email_configs"])
        _restore_monitor_jobs(jobs_snapshot)


def test_phase_17_1_admin_real_email_toggle_reaches_mocked_smtp(monkeypatch, tmp_path):
    init_db()
    email_snapshot = _snapshot_singleton_table("email_configs")
    settings_snapshot = _snapshot_table("system_settings")
    sent_messages = []

    class FakeSMTP:
        def __init__(self, host, port, timeout=30):
            self.host = host
            self.port = port
            self.timeout = timeout

        def login(self, username, password):
            self.username = username
            self.password = password

        def send_message(self, msg):
            sent_messages.append(msg)

        def quit(self):
            return None

    try:
        monkeypatch.setattr(smtplib, "SMTP_SSL", FakeSMTP)
        save_runtime_settings({"real_email_delivery": True}, actor_id=1)
        save_email_config(
            {
                "smtp_host": "smtp.real-looking.example",
                "smtp_port": 465,
                "encryption": "ssl",
                "sender": "sender@example.com",
                "username": "sender@example.com",
                "password": "super-secret",
                "default_recipients": ["ops@example.com"],
            }
        )

        send_test_email({})
        html_path = tmp_path / "report.html"
        md_path = tmp_path / "report.md"
        xlsx_path = tmp_path / "report.xlsx"
        html_path.write_text("<article>显式验证报告</article>", encoding="utf-8")
        md_path.write_text("# 显式验证报告", encoding="utf-8")
        xlsx_path.write_bytes(b"placeholder")
        ok, error = send_report(
            {"law_firm_name": "Phase17.1显式发送律所", "recipients": ["report@example.com"], "platforms": ["dy"]},
            {
                "html_path": str(html_path),
                "markdown_path": str(md_path),
                "excel_path": str(xlsx_path),
                "summary": {"job_id": 1, "law_firm_name": "Phase17.1显式发送律所", "platforms": ["dy"]},
            },
        )

        assert ok is True
        assert error is None
        assert len(sent_messages) == 2
        assert sent_messages[0]["To"] == "ops@example.com"
        assert sent_messages[1]["To"] == "report@example.com"
    finally:
        _restore_table("system_settings", settings_snapshot)
        _restore_singleton_table("email_configs", email_snapshot)


def test_phase_17_1_smtp_refused_recipients_fail_delivery(monkeypatch, tmp_path):
    init_db()
    email_snapshot = _snapshot_singleton_table("email_configs")
    settings_snapshot = _snapshot_table("system_settings")

    class RefusingSMTP:
        def __init__(self, host, port, timeout=30):
            self.host = host
            self.port = port
            self.timeout = timeout

        def login(self, username, password):
            self.username = username
            self.password = password

        def send_message(self, msg):
            return {"ops@example.com": (550, b"mailbox unavailable")}

        def quit(self):
            return None

    try:
        monkeypatch.setattr(smtplib, "SMTP_SSL", RefusingSMTP)
        save_runtime_settings({"real_email_delivery": True}, actor_id=1)
        save_email_config(
            {
                "smtp_host": "smtp.real-looking.example",
                "smtp_port": 465,
                "encryption": "ssl",
                "sender": "sender@example.com",
                "username": "sender@example.com",
                "password": "super-secret",
                "default_recipients": ["ops@example.com"],
            }
        )
        html_path = tmp_path / "report.html"
        md_path = tmp_path / "report.md"
        xlsx_path = tmp_path / "report.xlsx"
        html_path.write_text("<article>拒收测试报告</article>", encoding="utf-8")
        md_path.write_text("# 拒收测试报告", encoding="utf-8")
        xlsx_path.write_bytes(b"placeholder")

        ok, error = send_report(
            {"law_firm_name": "Phase17.1拒收律所", "recipients": ["ops@example.com"], "platforms": ["dy"]},
            {
                "html_path": str(html_path),
                "markdown_path": str(md_path),
                "excel_path": str(xlsx_path),
                "summary": {"job_id": 1, "law_firm_name": "Phase17.1拒收律所", "platforms": ["dy"]},
            },
        )

        assert ok is False
        assert error and "SMTP 服务器拒收了 1 个收件人" in error
        assert "550" in error
        assert "ops@example.com" not in error
    finally:
        _restore_table("system_settings", settings_snapshot)
        _restore_singleton_table("email_configs", email_snapshot)


def test_cr043_admin_frontend_real_email_toggle_controls_mail_test(monkeypatch):
    from api import main as api_main

    init_db()
    snapshots = {
        "audit_logs": _snapshot_table("audit_logs"),
        "system_settings": _snapshot_table("system_settings"),
        "user_sessions": _snapshot_table("user_sessions"),
        "users": _snapshot_table("users"),
        "email_configs": _snapshot_singleton_table("email_configs"),
    }
    sent_messages = []

    class FakeSMTP:
        def __init__(self, host, port, timeout=30):
            self.host = host
            self.port = port
            self.timeout = timeout

        def login(self, username, password):
            self.username = username
            self.password = password

        def send_message(self, msg):
            sent_messages.append(msg)
            return {}

        def quit(self):
            return None

    try:
        with get_conn() as conn:
            for table in ["audit_logs", "system_settings", "user_sessions", "users"]:
                conn.execute(f"DELETE FROM {table}")
        monkeypatch.setattr(smtplib, "SMTP_SSL", FakeSMTP)
        admin = bootstrap_admin_from_env("cr043-admin@example.com", "AdminPass123!", "CR043 Admin")
        save_email_config(
            {
                "smtp_host": "smtp.real-looking.example",
                "smtp_port": 465,
                "encryption": "ssl",
                "sender": "sender@example.com",
                "username": "sender@example.com",
                "password": "super-secret",
                "default_recipients": ["ops@example.com"],
            }
        )
        transport = httpx.ASGITransport(app=api_main.app)

        async def exercise() -> None:
            async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as client:
                login = await client.post("/api/auth/login", json={"email": "cr043-admin@example.com", "password": "AdminPass123!"})
                assert login.status_code == 200

                closed_send = await client.post("/api/monitor/email-config/test", json={})
                assert closed_send.status_code == 400
                assert "真实邮件发送未启用" in closed_send.text
                assert sent_messages == []

                state = (await client.get("/api/monitor/email-validation-window")).json()["validation_window"]
                assert state["real_email_delivery"] is False
                assert state["real_email_admin_enabled"] is False
                assert state["recipient_summary"]["source"] == "global_default_fallback"
                assert state["recipient_summary"]["count"] == 1

                opened = await client.put("/api/monitor/runtime-settings", json={"real_email_delivery": True})
                assert opened.status_code == 200
                assert opened.json()["settings"]["real_email_delivery"]["value"] is True

                sent = await client.post("/api/monitor/email-config/test", json={})
                assert sent.status_code == 200
                assert sent.json()["validation_window"]["real_email_delivery"] is True
                assert len(sent_messages) == 1
                assert sent_messages[0]["To"] == "ops@example.com"

                second_send = await client.post("/api/monitor/email-config/test", json={})
                assert second_send.status_code == 200
                assert len(sent_messages) == 2

                closed = await client.put("/api/monitor/runtime-settings", json={"real_email_delivery": False})
                assert closed.status_code == 200
                blocked_again = await client.post("/api/monitor/email-config/test", json={})
                assert blocked_again.status_code == 400
                assert len(sent_messages) == 2

        asyncio.run(exercise())

        assert get_runtime_setting_value("real_email_delivery") is False
        with get_conn() as conn:
            actions = [row["action_type"] for row in conn.execute("SELECT action_type FROM audit_logs ORDER BY id").fetchall()]
            audit_payload = [dict(row) for row in conn.execute("SELECT * FROM audit_logs").fetchall()]
        assert "update_runtime_settings" in actions
        assert "test_email_config" in actions
        assert "super-secret" not in json.dumps(audit_payload, ensure_ascii=False)
    finally:
        _restore_table("users", snapshots["users"])
        _restore_table("user_sessions", snapshots["user_sessions"])
        _restore_table("system_settings", snapshots["system_settings"])
        _restore_table("audit_logs", snapshots["audit_logs"])
        _restore_singleton_table("email_configs", snapshots["email_configs"])


def test_cr044_mail_test_submits_all_default_recipients(monkeypatch):
    init_db()
    email_snapshot = _snapshot_singleton_table("email_configs")
    settings_snapshot = _snapshot_table("system_settings")
    sent_messages = []

    class FakeSMTP:
        def __init__(self, host, port, timeout=30):
            self.host = host
            self.port = port
            self.timeout = timeout

        def login(self, username, password):
            self.username = username
            self.password = password

        def send_message(self, msg):
            sent_messages.append(msg)
            return {}

        def quit(self):
            return None

    try:
        monkeypatch.setattr(smtplib, "SMTP_SSL", FakeSMTP)
        save_runtime_settings({"real_email_delivery": True}, actor_id=1)
        save_email_config(
            {
                "smtp_host": "smtp.real-looking.example",
                "smtp_port": 465,
                "encryption": "ssl",
                "sender": "sender@example.com",
                "username": "sender@example.com",
                "password": "super-secret",
                "default_recipients": ["ops@example.com", "owner@example.com"],
            }
        )

        result = send_test_email({})

        assert result["recipient_count"] == 2
        assert result["recipient_source"] == "global_default_recipients"
        assert len(sent_messages) == 1
        assert sent_messages[0]["To"] == "ops@example.com, owner@example.com"
    finally:
        _restore_table("system_settings", settings_snapshot)
        _restore_singleton_table("email_configs", email_snapshot)


def test_cr043_admin_real_email_toggle_does_not_require_scheduler_exclusion(monkeypatch):
    init_db()
    snapshot = _snapshot_table("system_settings")
    try:
        with get_conn() as conn:
            conn.execute("DELETE FROM system_settings")
        save_runtime_settings({"real_email_delivery": True, "scheduler_disabled": False}, actor_id=1)
        state = monitor_router._email_validation_window_view()

        assert state["real_email_delivery"] is True
        assert state["real_email_admin_enabled"] is True
        assert state["scheduler_excluded"] is False
    finally:
        _restore_table("system_settings", snapshot)


def test_cr043_real_email_runtime_setting_is_admin_editable(monkeypatch):
    from api import main as api_main

    init_db()
    snapshots = {
        "audit_logs": _snapshot_table("audit_logs"),
        "system_settings": _snapshot_table("system_settings"),
        "user_sessions": _snapshot_table("user_sessions"),
        "users": _snapshot_table("users"),
    }
    try:
        with get_conn() as conn:
            for table in ["audit_logs", "system_settings", "user_sessions", "users"]:
                conn.execute(f"DELETE FROM {table}")
        bootstrap_admin_from_env("cr043-gates-admin@example.com", "AdminPass123!", "CR043 Gates Admin")
        transport = httpx.ASGITransport(app=api_main.app)

        async def exercise() -> None:
            async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as client:
                login = await client.post("/api/auth/login", json={"email": "cr043-gates-admin@example.com", "password": "AdminPass123!"})
                assert login.status_code == 200

                opened = await client.put("/api/monitor/runtime-settings", json={"real_email_delivery": True})
                assert opened.status_code == 200
                assert opened.json()["settings"]["real_email_delivery"]["value"] is True
                assert opened.json()["settings"]["real_email_delivery"]["is_locked"] is False

                closed = await client.put("/api/monitor/runtime-settings", json={"real_email_delivery": False})
                assert closed.status_code == 200
                assert closed.json()["settings"]["real_email_delivery"]["value"] is False

        asyncio.run(exercise())
    finally:
        _restore_table("users", snapshots["users"])
        _restore_table("user_sessions", snapshots["user_sessions"])
        _restore_table("system_settings", snapshots["system_settings"])
        _restore_table("audit_logs", snapshots["audit_logs"])


def test_cr043_real_email_toggle_ignores_superseded_deployment_locks(monkeypatch):
    init_db()
    snapshot = _snapshot_table("system_settings")
    try:
        monkeypatch.setenv("MONITOR_ALLOW_REAL_EMAIL_SEND", "true")
        monkeypatch.setenv("MONITOR_ALLOW_FRONTEND_REAL_EMAIL_VALIDATION", "true")
        with get_conn() as conn:
            conn.execute("DELETE FROM system_settings")
            conn.execute(
                """
                INSERT INTO system_settings (
                    workspace_id, key, value_json, value_type, is_locked, source,
                    updated_by, updated_at
                ) VALUES (1, 'real_email_delivery', 'false', 'boolean', 1, 'environment', 1, ?)
                """,
                (datetime.now(timezone.utc).isoformat(),),
            )

        stale = list_runtime_settings()["real_email_delivery"]
        assert stale["value"] is False
        assert stale["is_locked"] is False
        assert stale["source"] == "database"
        assert stale["apply_scope"] == "immediate"
        assert stale["yaml_path"] == ""

        updated = save_runtime_settings({"real_email_delivery": True}, actor_id=1)

        assert updated["real_email_delivery"]["value"] is True
        assert updated["real_email_delivery"]["is_locked"] is False
        with get_conn() as conn:
            row = conn.execute(
                "SELECT value_json, is_locked, source FROM system_settings WHERE workspace_id=1 AND key='real_email_delivery'"
            ).fetchone()
        assert dict(row) == {"value_json": "true", "is_locked": 0, "source": "database"}
    finally:
        _restore_table("system_settings", snapshot)


def test_cr043_real_email_toggle_api_is_admin_only(monkeypatch):
    from api import main as api_main

    init_db()
    snapshots = {
        "audit_logs": _snapshot_table("audit_logs"),
        "system_settings": _snapshot_table("system_settings"),
        "user_sessions": _snapshot_table("user_sessions"),
        "users": _snapshot_table("users"),
    }
    try:
        with get_conn() as conn:
            for table in ["audit_logs", "system_settings", "user_sessions", "users"]:
                conn.execute(f"DELETE FROM {table}")
        admin = bootstrap_admin_from_env("cr043-admin-only@example.com", "AdminPass123!", "CR043 Admin")
        save_user(
            {
                "email": "cr043-normal@example.com",
                "display_name": "CR043 Normal",
                "password": "UserPass123!",
                "role": "normal",
            },
            actor_id=int(admin["id"]),
        )
        transport = httpx.ASGITransport(app=api_main.app)

        async def exercise() -> None:
            async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as client:
                login = await client.post("/api/auth/login", json={"email": "cr043-normal@example.com", "password": "UserPass123!"})
                assert login.status_code == 200
                assert (await client.get("/api/monitor/email-validation-window")).status_code == 403
                assert (await client.put("/api/monitor/runtime-settings", json={"real_email_delivery": True})).status_code == 403

        asyncio.run(exercise())
    finally:
        _restore_table("users", snapshots["users"])
        _restore_table("user_sessions", snapshots["user_sessions"])
        _restore_table("system_settings", snapshots["system_settings"])
        _restore_table("audit_logs", snapshots["audit_logs"])


def test_cr043_report_resend_api_follows_admin_real_email_toggle(monkeypatch, tmp_path):
    from api import main as api_main

    init_db()
    jobs_snapshot = _snapshot_monitor_jobs()
    snapshots = {
        "audit_logs": _snapshot_table("audit_logs"),
        "system_settings": _snapshot_table("system_settings"),
        "user_sessions": _snapshot_table("user_sessions"),
        "users": _snapshot_table("users"),
        "email_delivery_logs": _snapshot_table("email_delivery_logs"),
        "reports": _snapshot_table("reports"),
        "crawl_runs": _snapshot_table("crawl_runs"),
        "email_configs": _snapshot_singleton_table("email_configs"),
    }
    sent_messages = []

    class FakeSMTP:
        def __init__(self, host, port, timeout=30):
            self.host = host

        def login(self, username, password):
            self.username = username

        def send_message(self, msg):
            sent_messages.append(msg)
            return {}

        def quit(self):
            return None

    try:
        with get_conn() as conn:
            for table in ["audit_logs", "system_settings", "user_sessions", "users", "email_delivery_logs", "reports", "crawl_runs"]:
                conn.execute(f"DELETE FROM {table}")
        _clear_monitor_jobs()
        monkeypatch.setattr(smtplib, "SMTP_SSL", FakeSMTP)
        admin = bootstrap_admin_from_env("cr043-resend-admin@example.com", "AdminPass123!", "CR043 Admin")
        user = save_user(
            {
                "email": "cr043-resend-user@example.com",
                "display_name": "CR043 User",
                "password": "UserPass123!",
                "role": "normal",
            },
            actor_id=int(admin["id"]),
        )
        save_email_config(
            {
                "smtp_host": "smtp.real-looking.example",
                "smtp_port": 465,
                "encryption": "ssl",
                "sender": "sender@example.com",
                "username": "sender@example.com",
                "password": "super-secret",
                "default_recipients": ["fallback@example.com"],
            }
        )
        job = save_job(
            {
                "law_firm_name": "CR043重发律所",
                "keywords": ["CR043重发律所投诉"],
                "platforms": ["dy"],
                "recipients": ["report@example.com"],
                "frequency": "daily",
            },
            actor=user,
        )
        run_id = create_run(job["id"], {"job_id": job["id"], "law_firm_name": job["law_firm_name"]})
        finish_run(run_id, "success", {"job_id": job["id"], "law_firm_name": job["law_firm_name"]})
        report = create_report(run_id, job, {"job_id": job["id"], "law_firm_name": job["law_firm_name"]})
        transport = httpx.ASGITransport(app=api_main.app)

        async def exercise() -> None:
            async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as normal_client:
                login = await normal_client.post("/api/auth/login", json={"email": "cr043-resend-user@example.com", "password": "UserPass123!"})
                assert login.status_code == 200
                normal_resend = await normal_client.post(f"/api/monitor/reports/{report['id']}/resend-email")
                assert normal_resend.status_code == 200
                assert normal_resend.json()["ok"] is False
                assert sent_messages == []

            async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as admin_client:
                login = await admin_client.post("/api/auth/login", json={"email": "cr043-resend-admin@example.com", "password": "AdminPass123!"})
                assert login.status_code == 200
                blocked = await admin_client.post(f"/api/monitor/reports/{report['id']}/resend-email")
                assert blocked.status_code == 200
                assert blocked.json()["ok"] is False
                assert "真实邮件发送未启用" in blocked.json()["error"]
                assert sent_messages == []

                opened = await admin_client.put("/api/monitor/runtime-settings", json={"real_email_delivery": True})
                assert opened.status_code == 200
                resend = await admin_client.post(f"/api/monitor/reports/{report['id']}/resend-email")
                assert resend.status_code == 200
                payload = resend.json()
                assert payload["ok"] is True
                assert payload["validation_window"]["status"] == "enabled"
                assert payload["validation_window"]["real_email_delivery"] is True
                assert len(sent_messages) == 1

        asyncio.run(exercise())

        logs = list_email_delivery_logs(report_id=report["id"], limit=10)
        assert logs[0]["status"] == "sent"
        assert logs[0]["trigger_source"] == "manual_resend"
        assert logs[0]["effective_recipients"] == ["report@example.com"]
        assert any(item["status"] == "skipped" for item in logs)
    finally:
        _restore_singleton_table("email_configs", snapshots["email_configs"])
        _restore_table("crawl_runs", snapshots["crawl_runs"])
        _restore_table("reports", snapshots["reports"])
        _restore_table("email_delivery_logs", snapshots["email_delivery_logs"])
        _restore_table("users", snapshots["users"])
        _restore_table("user_sessions", snapshots["user_sessions"])
        _restore_table("system_settings", snapshots["system_settings"])
        _restore_table("audit_logs", snapshots["audit_logs"])
        _restore_monitor_jobs(jobs_snapshot)


def test_phase_17a_failed_auto_delivery_records_log_without_blocking_report(monkeypatch):
    init_db()
    jobs_snapshot = _snapshot_monitor_jobs()
    snapshots = {
        "email_delivery_logs": _snapshot_table("email_delivery_logs"),
        "reports": _snapshot_table("reports"),
        "crawl_runs": _snapshot_table("crawl_runs"),
        "raw_contents": _snapshot_table("raw_contents"),
        "raw_comments": _snapshot_table("raw_comments"),
        "ai_evaluations": _snapshot_table("ai_evaluations"),
    }
    _clear_monitor_jobs()
    content_id = "pytest_phase17a_auto_delivery_001"
    now_ts = int(datetime.now(timezone.utc).timestamp())
    try:
        with get_conn() as conn:
            for table in ["email_delivery_logs", "reports", "crawl_runs", "raw_contents", "raw_comments", "ai_evaluations"]:
                conn.execute(f"DELETE FROM {table}")
        job = save_job(
            {
                "law_firm_name": "Phase17A失败律所",
                "keywords": ["Phase17A失败律所投诉"],
                "platforms": ["dy"],
                "recipients": ["ops@example.com"],
                "frequency": "daily",
                "email_time": "09:00",
                "enable_comments": False,
            }
        )

        async def fake_run_platform(job_arg, run_id, platform, run_dir):
            return ingest_outputs(
                job_arg,
                run_id,
                platform,
                [
                    {
                        "aweme_id": content_id,
                        "title": "Phase17A失败律所投诉",
                        "desc": "收费争议需要人工复核",
                        "create_time": now_ts,
                        "share_url": "https://www.douyin.com/video/phase17a",
                    }
                ],
                [],
            )

        monkeypatch.setattr(runner_module, "run_platform", fake_run_platform)
        monkeypatch.setattr(
            "api.monitoring.reporting.send_report",
            lambda job_arg, report_arg: (False, "smtp_password=super-secret token=hidden"),
        )

        result = asyncio.run(run_monitor_job(job["id"]))
        report = get_report(int(result["report"]["id"]))
        logs = list_email_delivery_logs(report_id=report["id"], limit=10)

        assert result["status"] == "success"
        assert report and report["email_status"] == "failed"
        assert "super-secret" not in (report["email_error"] or "")
        assert result["summary"]["email_status"] == "failed"
        assert "email_delivery_log_id" in result["summary"]
        assert logs and logs[0]["send_type"] == "auto"
        assert logs[0]["status"] == "failed"
        assert logs[0]["send_window_key"].startswith(f"{job['id']}_")
        assert "super-secret" not in (logs[0].get("error_message") or "")
        assert "hidden" not in (logs[0].get("error_message") or "")
        assert logs[0]["recipients"] == ["ops@example.com"]
    finally:
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)
        _restore_monitor_jobs(jobs_snapshot)


def test_phase_17b_email_delivery_history_api_scope_and_safe_fields():
    from api import main as api_main

    init_db()
    jobs_snapshot = _snapshot_monitor_jobs()
    snapshots = {
        "email_delivery_logs": _snapshot_table("email_delivery_logs"),
        "reports": _snapshot_table("reports"),
        "crawl_runs": _snapshot_table("crawl_runs"),
        "users": _snapshot_table("users"),
        "user_sessions": _snapshot_table("user_sessions"),
        "audit_logs": _snapshot_table("audit_logs"),
    }
    _clear_monitor_jobs()
    try:
        with get_conn() as conn:
            for table in ["email_delivery_logs", "reports", "crawl_runs", "user_sessions", "audit_logs"]:
                conn.execute(f"DELETE FROM {table}")
        bootstrap_actor = bootstrap_admin_from_env("phase17b-bootstrap@example.com", "AdminPass123!", "Phase 17B Bootstrap")
        assert bootstrap_actor
        admin = save_user(
            {
                "email": "phase17b-admin@example.com",
                "display_name": "Phase 17B Admin",
                "password": "AdminPass123!",
                "role": "administrator",
            },
            actor_id=int(bootstrap_actor["id"]),
        )
        user1 = save_user(
            {
                "email": "phase17b-user1@example.com",
                "display_name": "Phase 17B User 1",
                "password": "UserPass123!",
                "role": "normal",
            },
            actor_id=int(admin["id"]),
        )
        user2 = save_user(
            {
                "email": "phase17b-user2@example.com",
                "display_name": "Phase 17B User 2",
                "password": "UserPass456!",
                "role": "normal",
            },
            actor_id=int(admin["id"]),
        )
        job1 = save_job(
            {
                "law_firm_name": "Phase17B用户一律所",
                "keywords": ["Phase17B用户一律所投诉"],
                "platforms": ["dy"],
                "recipients": ["user1@example.com"],
                "frequency": "daily",
            },
            actor=user1,
        )
        job2 = save_job(
            {
                "law_firm_name": "Phase17B用户二律所",
                "keywords": ["Phase17B用户二律所投诉"],
                "platforms": ["dy"],
                "recipients": ["user2@example.com"],
                "frequency": "daily",
            },
            actor=user2,
        )
        run1 = create_run(job1["id"], {"job_id": job1["id"], "law_firm_name": job1["law_firm_name"]})
        run2 = create_run(job2["id"], {"job_id": job2["id"], "law_firm_name": job2["law_firm_name"]})
        finish_run(run1, "success", {"job_id": job1["id"], "law_firm_name": job1["law_firm_name"]})
        finish_run(run2, "success", {"job_id": job2["id"], "law_firm_name": job2["law_firm_name"]})
        report1 = create_report(run1, job1, {"job_id": job1["id"], "law_firm_name": job1["law_firm_name"]})
        report2 = create_report(run2, job2, {"job_id": job2["id"], "law_firm_name": job2["law_firm_name"]})
        record_email_delivery_log(
            {
                "workspace_id": job1["workspace_id"],
                "job_id": job1["id"],
                "report_id": report1["id"],
                "send_window_key": f"{job1['id']}_2026-06-16",
                "send_type": "auto",
                "status": "failed",
                "sent_at": "2026-06-16T09:00:00+00:00",
                "error_message": "smtp_password=super-secret token=hidden",
                "recipients": ["user1@example.com"],
            }
        )
        record_email_delivery_log(
            {
                "workspace_id": job1["workspace_id"],
                "job_id": job1["id"],
                "report_id": report1["id"],
                "send_window_key": f"{job1['id']}_2026-06-16",
                "send_type": "manual_resend",
                "sent_by": user1["id"],
                "status": "sent",
                "sent_at": "2026-06-16T09:30:00+00:00",
                "recipients": ["user1@example.com"],
            }
        )
        record_email_delivery_log(
            {
                "workspace_id": job2["workspace_id"],
                "job_id": job2["id"],
                "report_id": report2["id"],
                "send_window_key": f"{job2['id']}_2026-06-16",
                "send_type": "auto",
                "status": "sent",
                "sent_at": "2026-06-16T09:00:00+00:00",
                "recipients": ["user2@example.com"],
            }
        )

        transport = httpx.ASGITransport(app=api_main.app)

        async def exercise() -> None:
            async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as user_client:
                login = await user_client.post(
                    "/api/auth/login",
                    json={"email": "phase17b-user1@example.com", "password": "UserPass123!"},
                )
                assert login.status_code == 200
                owned = await user_client.get(f"/api/monitor/reports/{report1['id']}/email-delivery-logs")
                assert owned.status_code == 200
                payload = owned.json()
                logs = payload["delivery_logs"]
                assert [item["send_type"] for item in logs] == ["manual_resend", "auto"]
                assert logs[0]["sent_by"] == user1["id"]
                assert logs[0]["recipients"] == ["user1@example.com"]
                assert "super-secret" not in str(payload)
                assert "hidden" not in str(payload)
                assert "smtp_password" not in str(payload)
                assert "recipients_json" not in str(payload)
                assert "workspace_id" not in str(payload)
                other = await user_client.get(f"/api/monitor/reports/{report2['id']}/email-delivery-logs")
                assert other.status_code == 404

            async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as admin_client:
                login = await admin_client.post(
                    "/api/auth/login",
                    json={"email": "phase17b-admin@example.com", "password": "AdminPass123!"},
                )
                assert login.status_code == 200
                admin_view = await admin_client.get(f"/api/monitor/reports/{report2['id']}/email-delivery-logs")
                assert admin_view.status_code == 200
                assert admin_view.json()["delivery_logs"][0]["recipients"] == ["user2@example.com"]

        asyncio.run(exercise())
    finally:
        _restore_table("email_delivery_logs", snapshots["email_delivery_logs"])
        _restore_table("reports", snapshots["reports"])
        _restore_table("crawl_runs", snapshots["crawl_runs"])
        _restore_table("users", snapshots["users"])
        _restore_table("user_sessions", snapshots["user_sessions"])
        _restore_table("audit_logs", snapshots["audit_logs"])
        _restore_monitor_jobs(jobs_snapshot)


def test_phase_18a_report_job_snapshots_persist_backfill_and_limited_context():
    init_db()
    jobs_snapshot = _snapshot_monitor_jobs()
    snapshots = {
        "reports": _snapshot_table("reports"),
        "crawl_runs": _snapshot_table("crawl_runs"),
    }
    _clear_monitor_jobs()
    try:
        with get_conn() as conn:
            for table in ["reports", "crawl_runs"]:
                conn.execute(f"DELETE FROM {table}")
        with get_conn() as conn:
            assert "job_snapshot_json" in _table_columns(conn, "reports")

        job = save_job(
            {
                "law_firm_name": "Phase18A快照律所",
                "keywords": ["Phase18A快照律所投诉", "Phase18A快照律所退费"],
                "platforms": ["dy", "xhs"],
                "recipients": ["ops@example.com"],
                "frequency": "daily",
                "email_time": "09:00",
            }
        )
        run_id = create_run(job["id"], {"job_id": job["id"], "law_firm_name": job["law_firm_name"]})
        finish_run(run_id, "success", {"job_id": job["id"], "law_firm_name": job["law_firm_name"]})
        report = create_report(
            run_id,
            job,
            {"job_id": job["id"], "law_firm_name": job["law_firm_name"], "platforms": job["platforms"]},
        )
        stored = get_report(report["id"])

        assert report["job_snapshot"]["job_id"] == job["id"]
        assert stored["job_snapshot"].items() >= {
            "job_id": job["id"],
            "law_firm_name": "Phase18A快照律所",
            "platforms": ["dy", "xhs"],
            "keywords": ["Phase18A快照律所投诉", "Phase18A快照律所退费"],
            "frequency": "daily",
            "deleted_at": None,
        }.items()
        assert stored["job_snapshot"]["email_template"]["source"] in {"active_global_fallback", "default_renderer"}
        assert stored["job_deleted"] is False
        assert stored["limited_context"] is False

        with get_conn() as conn:
            conn.execute("UPDATE reports SET job_snapshot_json=NULL WHERE id=?", (report["id"],))
        init_db()
        backfilled = get_report(report["id"])
        assert backfilled["job_snapshot"]["law_firm_name"] == "Phase18A快照律所"
        assert backfilled["job_snapshot"]["platforms"] == ["dy", "xhs"]
        assert backfilled["job_snapshot"]["keywords"] == ["Phase18A快照律所投诉", "Phase18A快照律所退费"]

        asyncio.run(monitor_router.remove_job(job["id"]))
        after_delete = get_report(report["id"])
        assert after_delete
        assert after_delete["job_id"] == job["id"]
        assert after_delete["job_deleted"] is True
        assert after_delete["display_law_firm_name"] == "Phase18A快照律所"
        assert after_delete["job_snapshot"]["deleted_at"]
        assert after_delete["limited_context"] is False

        now = datetime.now(timezone.utc).isoformat()
        with get_conn() as conn:
            cur = conn.execute(
                """
                INSERT INTO crawl_runs (workspace_id, job_id, status, started_at, finished_at, summary)
                VALUES (1, NULL, 'success', ?, ?, '{}')
                """,
                (now, now),
            )
            orphan_run_id = int(cur.lastrowid)
            cur = conn.execute(
                """
                INSERT INTO reports (
                    workspace_id, run_id, job_id, html_path, markdown_path, excel_path,
                    job_snapshot_json, summary, created_at
                ) VALUES (1, ?, NULL, ?, ?, ?, NULL, '{}', ?)
                """,
                (
                    orphan_run_id,
                    str(Path("monitor_data/reports/orphan.html")),
                    str(Path("monitor_data/reports/orphan.md")),
                    str(Path("monitor_data/reports/orphan.xlsx")),
                    now,
                ),
            )
            orphan_report_id = int(cur.lastrowid)
        orphan = get_report(orphan_report_id)
        assert orphan
        assert orphan["legacy_without_job_snapshot"] is True
        assert orphan["limited_context"] is True
        assert orphan["display_law_firm_name"] == "历史报告"

        missing_context_job = save_job(
            {
                "law_firm_name": "Phase18A缺失上下文律所",
                "keywords": ["Phase18A缺失上下文律所投诉"],
                "platforms": ["dy"],
                "recipients": [],
            }
        )
        missing_run_id = create_run(missing_context_job["id"], {})
        missing_report = create_report(missing_run_id, missing_context_job, {})
        with get_conn() as conn:
            conn.execute("UPDATE reports SET job_snapshot_json=NULL, summary='{}' WHERE id=?", (missing_report["id"],))
            conn.execute("DELETE FROM monitor_jobs WHERE id=?", (missing_context_job["id"],))
        missing = get_report(missing_report["id"])
        assert missing
        assert missing["job_deleted"] is True
        assert missing["legacy_without_job_snapshot"] is True
        assert missing["limited_context"] is True
    finally:
        _restore_table("reports", snapshots["reports"])
        _restore_table("crawl_runs", snapshots["crawl_runs"])
        _restore_monitor_jobs(jobs_snapshot)


def test_phase_18a_report_snapshot_does_not_bypass_owner_scope():
    init_db()
    jobs_snapshot = _snapshot_monitor_jobs()
    snapshots = {
        "reports": _snapshot_table("reports"),
        "crawl_runs": _snapshot_table("crawl_runs"),
        "user_sessions": _snapshot_table("user_sessions"),
        "audit_logs": _snapshot_table("audit_logs"),
        "users": _snapshot_table("users"),
    }
    _clear_monitor_jobs()
    try:
        with get_conn() as conn:
            for table in ["reports", "crawl_runs", "user_sessions", "audit_logs", "users"]:
                conn.execute(f"DELETE FROM {table}")
        admin = bootstrap_admin_from_env("phase18a-admin@example.com", "AdminPass123!", "Admin")
        user_one = save_user(
            {
                "email": "phase18a-user1@example.com",
                "display_name": "User One",
                "password": "UserPass123!",
                "role": "normal",
            },
            actor_id=int(admin["id"]),
        )
        user_two = save_user(
            {
                "email": "phase18a-user2@example.com",
                "display_name": "User Two",
                "password": "UserPass123!",
                "role": "normal",
            },
            actor_id=int(admin["id"]),
        )
        actor_one = {"id": user_one["id"], "role": "normal", "workspace_id": 1}
        actor_two = {"id": user_two["id"], "role": "normal", "workspace_id": 1}
        job = save_job(
            {
                "law_firm_name": "Phase18A权限律所",
                "keywords": ["Phase18A权限律所投诉"],
                "platforms": ["dy"],
                "recipients": [],
            },
            actor=actor_one,
        )
        run_id = create_run(job["id"], {"job_id": job["id"], "law_firm_name": job["law_firm_name"]})
        finish_run(run_id, "success", {"job_id": job["id"], "law_firm_name": job["law_firm_name"]})
        report = create_report(run_id, job, {"job_id": job["id"], "law_firm_name": job["law_firm_name"]})

        assert get_report(report["id"], actor=actor_one)
        assert get_report(report["id"], actor=actor_two) is None
        assert any(item["id"] == report["id"] for item in list_reports(0, actor=actor_one))
        assert all(item["id"] != report["id"] for item in list_reports(0, actor=actor_two))

        asyncio.run(monitor_router.remove_job(job["id"], user=actor_one))
        assert get_report(report["id"], actor=actor_one)
        assert get_report(report["id"], actor=actor_two) is None
        assert all(item["id"] != report["id"] for item in list_reports(0, actor=actor_two))
    finally:
        _restore_table("reports", snapshots["reports"])
        _restore_table("crawl_runs", snapshots["crawl_runs"])
        _restore_table("user_sessions", snapshots["user_sessions"])
        _restore_table("audit_logs", snapshots["audit_logs"])
        _restore_table("users", snapshots["users"])
        _restore_monitor_jobs(jobs_snapshot)


def test_phase_17b_report_center_delivery_history_frontend_hooks():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    frontend_source = page + "\n" + css

    reports_section = _task_group_view(page)
    delivery_drawer = page[page.index('id="email_delivery_history_drawer"') : page.index('id="email_template_drawer_backdrop"')]
    assert "邮件交付历史" in delivery_drawer
    assert "email_delivery_history" not in reports_section
    assert "email_delivery_history_scope" not in reports_section
    assert "email_delivery_history" in delivery_drawer
    assert "email_delivery_history_scope" in delivery_drawer
    assert "email_delivery_history_count" in delivery_drawer
    assert "data-report-delivery-panel" not in reports_section
    assert "data-report-delivery-panel" in delivery_drawer
    assert "loadEmailDeliveryHistory" in page
    assert "openEmailDeliveryHistoryDrawer" in page
    assert "closeEmailDeliveryHistoryDrawer" in page
    assert "refreshSelectedEmailDeliveryHistory" in page
    assert "renderEmailDeliveryHistory" in page
    assert "renderEmailDeliveryLog" in page
    assert "email-delivery-logs" in page
    assert "confirm('确认手动重发这份报告邮件？系统会单独记录本次重发历史。')" in page
    assert "await loadEmailDeliveryHistory(id, {silent:true})" in page
    assert "reportEmailStatusCell" not in page
    assert "emailDeliveryStatusLabel" in page
    assert "邮件已提交 SMTP，请人工确认收件箱或垃圾箱" in page
    assert "sent:'SMTP已接受'" in page
    assert "sent:'SMTP已接受，待收件确认'" in page
    assert "报告邮件已重新发送" not in page
    assert "sent:'发送成功'" not in page
    assert "recipients_json" not in reports_section
    assert "smtp_password" not in reports_section
    for selector in [
        ".email-delivery-history-drawer",
        ".email-delivery-history-content",
        ".email-status-button",
        ".email-delivery-latest",
        ".email-delivery-history-list",
        ".email-delivery-history-item",
    ]:
        assert selector in frontend_source
    assert "@media (max-width: 1279px)" in css
    assert "@media (max-width: 767px)" in css


def test_phase_17_1d_orphan_email_evidence_dry_run_helper_is_noop(tmp_path, capsys):
    init_db()
    jobs_snapshot = _snapshot_monitor_jobs()
    snapshots = {
        "email_delivery_logs": _snapshot_table("email_delivery_logs"),
        "reports": _snapshot_table("reports"),
        "crawl_runs": _snapshot_table("crawl_runs"),
    }
    _clear_monitor_jobs()
    artifact_root = tmp_path / "reports"
    artifact_root.mkdir()
    try:
        with get_conn() as conn:
            for table in ["email_delivery_logs", "reports", "crawl_runs"]:
                conn.execute(f"DELETE FROM {table}")
        job = save_job(
            {
                "law_firm_name": "Phase171D正常律所",
                "keywords": ["Phase171D正常律所投诉"],
                "platforms": ["dy"],
                "recipients": ["ops@example.com"],
                "frequency": "daily",
            }
        )
        run_id = create_run(job["id"], {"job_id": job["id"], "law_firm_name": job["law_firm_name"]})
        finish_run(run_id, "success", {"job_id": job["id"], "law_firm_name": job["law_firm_name"]})
        report = create_report(run_id, job, {"job_id": job["id"], "law_firm_name": job["law_firm_name"]})
        normal_log = record_email_delivery_log(
            {
                "workspace_id": job["workspace_id"],
                "job_id": job["id"],
                "report_id": report["id"],
                "send_window_key": f"{job['id']}_2026-06-18",
                "send_type": "auto",
                "status": "sent",
                "sent_at": "2026-06-18T01:00:00+00:00",
                "recipients": ["ops@example.com"],
            }
        )

        orphan_job_id = 9686
        orphan_run_id = 8380
        orphan_report_id = 3959
        for suffix in ("html", "md", "xlsx"):
            (artifact_root / f"job_{orphan_job_id}_run_{orphan_run_id}_20260616_152702.{suffix}").write_text(
                f"orphan {suffix}",
                encoding="utf-8",
            )
        with get_conn() as conn:
            conn.execute(
                """
                INSERT INTO email_delivery_logs (
                    workspace_id, job_id, report_id, send_window_key, send_type,
                    sent_by, sent_at, status, error_message, recipients_json,
                    trigger_source, effective_recipients_json, effective_recipient_source,
                    email_template_id, email_template_name, email_template_source,
                    email_subject_template, created_at
                ) VALUES (1, ?, ?, ?, 'auto', NULL, ?, 'sent', '', '["ops@example.com"]',
                    'scheduler_auto', '["ops@example.com"]', 'task_recipients',
                    NULL, '历史模板', 'task_bound', '日报 {law_firm_name}', ?)
                """,
                (
                    orphan_job_id,
                    orphan_report_id,
                    f"{orphan_job_id}_2026-06-16",
                    "2026-06-16T07:27:11+00:00",
                    "2026-06-16T07:27:11+00:00",
                ),
            )
            orphan_log_id = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
        before = {
            "email_delivery_logs": _snapshot_table("email_delivery_logs"),
            "reports": _snapshot_table("reports"),
            "crawl_runs": _snapshot_table("crawl_runs"),
            "artifacts": {path.name: path.read_text(encoding="utf-8") for path in artifact_root.iterdir()},
        }

        orphan_review = build_orphan_email_evidence_review(
            delivery_log_id=orphan_log_id,
            artifact_root=artifact_root,
        )
        normal_review = build_orphan_email_evidence_review(
            delivery_log_id=normal_log["id"],
            artifact_root=artifact_root,
        )
        assert orphan_review["mode"] == "dry_run"
        assert orphan_review["mutations_attempted"] == 0
        assert orphan_review["items"][0]["classification"] == "orphan_delivery_log"
        assert "detached_report_artifacts" in orphan_review["items"][0]["secondary_classifications"]
        assert orphan_review["items"][0]["exists"] == {"job": False, "report": False, "run": False}
        assert orphan_review["items"][0]["artifacts"]["existing_count"] == 3
        assert {
            "database_backup_required",
            "artifact_email_backup_required",
            "explicit_operator_approval_required",
            "rollback_plan_required",
        } <= set(orphan_review["items"][0]["required_before_any_mutation"])
        assert orphan_review["items"][0]["dry_run"]["proposed_effect"] == "review_only_no_changes"
        assert normal_review["items"][0]["classification"] == "normal"
        assert normal_review["items"][0]["exists"]["job"] is True
        assert normal_review["items"][0]["exists"]["report"] is True
        assert normal_review["items"][0]["exists"]["run"] is True

        exit_code = review_orphan_email_main(
            [
                "--delivery-log-id",
                str(orphan_log_id),
                "--artifact-root",
                str(artifact_root),
                "--json",
            ]
        )
        cli_payload = json.loads(capsys.readouterr().out)
        assert exit_code == 0
        assert cli_payload["items"][0]["classification"] == "orphan_delivery_log"
        assert cli_payload["mutations_attempted"] == 0

        after = {
            "email_delivery_logs": _snapshot_table("email_delivery_logs"),
            "reports": _snapshot_table("reports"),
            "crawl_runs": _snapshot_table("crawl_runs"),
            "artifacts": {path.name: path.read_text(encoding="utf-8") for path in artifact_root.iterdir()},
        }
        assert after == before
        assert list_email_delivery_logs(report_id=report["id"], limit=10)[0]["id"] == normal_log["id"]
        assert get_report(report["id"])["id"] == report["id"]
    finally:
        _restore_table("email_delivery_logs", snapshots["email_delivery_logs"])
        _restore_table("reports", snapshots["reports"])
        _restore_table("crawl_runs", snapshots["crawl_runs"])
        _restore_monitor_jobs(jobs_snapshot)


def test_cr048_report_center_lead_detail_requires_visible_scope_and_report_action():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    reports_section = _task_group_view(page)

    assert 'data-report-lead-panel' not in reports_section
    assert 'id="leads_scope"' not in reports_section
    assert 'id="leads_scope_count"' not in reports_section
    assert 'id="leads_table"' not in reports_section
    assert 'id="lead_status_filter"' not in reports_section
    assert '<label>线索状态</label>' not in reports_section
    assert 'id="report_leads_backdrop"' not in page
    assert 'id="report_leads_drawer"' not in page
    assert 'data-report-lead-panel' not in page
    assert 'id="leads_scope"' not in page
    assert 'id="leads_scope_count"' not in page
    assert 'id="leads_table"' not in page
    assert 'id="leads_scope_hint"' not in page
    assert 'id="lead_status_filter"' not in page
    assert 'function viewReportLeads(id)' not in page
    assert 'function viewRunLeads(id)' not in page
    assert 'function openReportLeadsDrawer(id, options={})' not in page
    assert 'function closeReportLeadsDrawer()' not in page
    assert "function reloadCurrentLeadDrawer()" not in page
    assert "run_detail_report_filter" in page
    assert "function setRunDetailAiFilters(filters={}, options={})" in page
    assert "function jumpToReportAiEvaluations(reportId)" in page
    assert "报告 #${esc(report.id)}" in page
    assert '<button class="secondary" onclick="viewReportLeads(${reportId})">查看线索</button>' not in page
    assert "closeReportActionMenu" not in page
    assert "jumpToReportAiEvaluations(${Number(report.id)})" in page
    assert 'renderLeads([], {' not in page
    assert "type:'none_selected'" not in page
    assert "当前筛选条件下的线索" not in page
    assert "document.getElementById('leads_scope').textContent = '正在加载当前报告线索...';" not in page
    assert ".report-leads-drawer" not in css
    assert ".report-leads-toolbar" not in css
    assert ".report-leads-content" not in css
    assert ".lead-detail-panel" not in css


def test_phase_20d_run_detail_frontend_hooks():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    runs_section = _monitor_section(page, "runs")
    reports_section = _task_group_view(page)
    run_detail_drawer = page[page.index('id="run_detail_drawer"') : page.index('id="report_preview_backdrop"')]

    assert 'data-run-detail-panel' not in runs_section
    assert 'data-run-detail-panel' not in reports_section
    assert 'id="run_detail_backdrop"' in page
    assert 'id="run_detail_drawer"' in page
    assert 'data-run-detail-panel' in run_detail_drawer
    assert 'id="run_detail_body"' in run_detail_drawer
    assert 'data-run-detail-tab="overview"' in run_detail_drawer
    assert 'data-run-detail-tab="logs"' in run_detail_drawer
    assert 'data-run-detail-tab="contents"' in run_detail_drawer
    assert 'data-run-detail-tab="ai"' in run_detail_drawer
    assert 'data-run-detail-tab="reports"' in run_detail_drawer
    assert 'data-run-detail-tab="email"' in run_detail_drawer
    assert 'onclick="openRunDetail(${Number(r.id)})">详情</button>' in page
    assert "function openRunDetail(id)" in page
    assert "function loadRunDetail(id)" in page
    assert "function renderRunDetailBody(detail)" in page
    assert "function renderRunAiEvaluations(detail)" in page
    assert "function syncRunDetailTabs()" in page
    assert "syncRunDetailTabs();" in page
    assert "function openRunEvaluationDetail(runId, evaluationId)" in page
    assert 'onclick="closeRunDetailDrawer()">返回列表</button>' in page
    assert 'onclick="renderRunDetailBody(currentRunDetailState)">返回列表</button>' not in page
    assert "/runs/${Number(id)}/detail" in page
    assert "/runs/${Number(runId)}/ai-evaluations/${Number(evaluationId)}" in page
    assert "日志和 AI 详情在同一面板内可达" in page
    assert "runDetailLogsText(logs)" in page
    assert "复制日志" in page
    assert "下载日志" in page
    assert 'id="run_detail_actions"' in run_detail_drawer
    assert "stopRunFromDetail" in page
    assert "archiveRunFromDetail" in page
    assert "restoreRunFromDetail" in page
    assert "run-detail-report-actions" in page
    assert "function setRunDetailAiFilters(filters={}, options={})" in page
    assert "function jumpToReportAiEvaluations(reportId)" in page
    assert "jumpToReportAiEvaluations(${Number(report.id)})" in page
    assert 'run-detail-ai-toolbar page-filter-region' in page
    assert 'data-filter-region="run-detail-ai"' in page
    assert "run_detail_report_filter" in page
    assert 'id="run_detail_report_filter" type="hidden"' in page
    assert "reports.length > 1 ? `" in page
    assert '<select id="run_detail_report_filter"' in page
    assert "reports.length > 1" in page
    assert "run-detail-ai-scope-note" in page
    assert "仅看报告 #${aiAppliedFilters.report_id}" in page
    assert "当前运行关联报告" in page
    assert "run_detail_ai_status_filter" in page
    assert "run_detail_ai_risk_filter" in page
    assert "run_detail_ai_platform_filter" in page
    assert "run_detail_ai_keyword_filter" in page
    assert "run_detail_ai_title_filter" in page
    assert "report_leads_drawer" not in page
    assert "report_leads_backdrop" not in page
    assert "closeReportLeadsDrawer" not in page
    assert "loadEmailDeliveryHistory(${Number(report.id)})" in page
    assert "resendReportEmail(${Number(report.id)})" in page
    assert "/reports/${Number(report.id)}/download?type=html" in page
    assert "返回 AI 评估列表" in page
    assert ".run-detail-drawer" in css
    assert ".run-detail-tabs" in css
    assert ".run-detail-grid" in css
    assert ".run-detail-code" in css
    assert ".run-detail-action-slot" in css
    assert ".run-detail-report-actions" in css
    assert ".run-detail-ai-toolbar" in css
    assert ".run-detail-ai-filters" in css
    assert ".run-detail-ai-filters .filter-select-button" in css
    assert ".run-detail-ai-filters .filter-select-enhanced" in css
    assert ".run-detail-ai-scope-note" in css
    assert ".run-detail-tabs {\n    display: grid;" in css
    assert "grid-template-columns: repeat(2, minmax(0, 1fr));" in css
    assert ".detail-tabs.run-detail-tabs {\n        display:grid;" in page


def test_phase_20d_run_detail_uses_final_summary_fallbacks():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")

    assert "summaryAi=summary.ai_evaluation || {}" in page
    assert "finalRawCount=Number(collection.final_raw_contents ?? summary.raw_contents ?? contents.length ?? 0)" in page
    assert "hasFinalCollection=!!collection.final || summary.raw_contents !== undefined || run.status !== 'running'" in page
    assert "aiTotal=Number(ai.total_candidates ?? summaryAi.total_candidates ?? (detail?.ai_pagination||{}).total ?? aiItems.length ?? 0)" in page
    assert "aiEvaluated=Number(ai.evaluated_items ?? summaryAi.evaluated ?? aiItems.filter(item=>item.evaluation_id).length ?? 0)" in page
    assert "aiManual=Number(ai.manual_review_count ?? summary.manual_review_count ?? summary.pending_review_count ?? aiItems.filter(item=>item.lead_status==='pending_review').length ?? 0)" in page
    assert "已采集 ${finalRawCount}" in page
    assert "临时 ${seenRawCount}" in page
    assert "aiTotal ? (aiFinal ? 'ok' : 'warn') : ''" in page


def test_phase_20e_report_leads_backlink_to_run_detail():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    reports_section = _task_group_view(page)
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")

    assert 'data-run-detail-panel' not in reports_section
    assert 'id="report_leads_drawer"' not in page
    assert "function leadLinks(item)" not in page
    assert "closeReportLeadsDrawer(); openRunDetail(${runId})" not in page
    assert "jumpToReportAiEvaluations(${Number(report.id)})" in page
    assert "report_id: reportId" in page
    assert "报告 #${filters.report_id}" in page
    assert "上下文有限</span>" in page
    assert '<label>线索状态</label>' not in reports_section
    assert 'id="leads_table"' not in reports_section
    assert ".report-leads-drawer" not in css
    assert "#report_leads_drawer.drawer.active" not in css
    assert "#report_leads_backdrop.active" not in css


def test_cr049_mail_and_delivery_history_action_hierarchy_frontend_hooks():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    email_section = page[page.index('<section id="email"') : page.index('<div id="mail_config_backdrop"')]
    reports_section = _task_group_view(page)

    assert "email-toolbar-switch" in email_section
    assert "真实邮件：已关闭" in email_section
    assert email_section.count("openMailConfigModal()") == 1
    assert email_section.count("openMailTestModal()") == 1
    assert "SMTP 与发送默认值" not in email_section
    assert "email_validation_status" not in email_section
    delivery_drawer = page[page.index('id="email_delivery_history_drawer"') : page.index('id="email_template_drawer_backdrop"')]
    assert 'data-report-delivery-panel' not in reports_section
    assert 'id="email_delivery_history"' not in reports_section
    assert 'data-report-delivery-panel' in delivery_drawer
    assert 'id="email_delivery_history"' in delivery_drawer
    assert 'id="email_delivery_history_scope"' in delivery_drawer
    assert 'id="email_delivery_history_count"' in delivery_drawer
    assert 'is-secondary-detail' not in reports_section
    assert '点击报告行的邮件状态或“更多 > 查看交付历史”会打开悬浮窗' not in reports_section
    assert "document.querySelector('[data-report-delivery-panel]')?.scrollIntoView" not in page
    assert "closeReportActionMenu" not in page
    assert "loadEmailDeliveryHistory(${Number(report.id)})" in page
    assert ".email-toolbar-switch" in css
    assert ".email-delivery-history-drawer" in css
    assert ".email-delivery-history-content" in css
    assert ".email-delivery-history-panel" not in css


def test_phase_17_1c_17_2bc_email_recipient_and_template_explanations():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")

    assert "HTML 模板必须包含 {report_html} 或 {report_body}" in page
    assert 'id="email_template_preset"' in page
    assert "标准日报" in page
    assert "紧凑摘要" in page
    assert "正式简报" in page
    assert "自定义 / 历史模板" in page
    assert 'id="email_template_guardrail"' in page
    assert "function emailTemplateHasReportBodyPlaceholder(template)" in page
    assert "function updateEmailTemplateGuardrail()" in page
    assert "function applyEmailTemplatePreset(preset)" in page
    assert "function emailTemplatePresetHtml(preset)" in page
    assert "templateDrawerActive" in page
    assert "set('email_template_preset', 'custom')" in page
    assert "set('email_template_preset', 'standard')" in page
    assert "resourceStat('正文保护'" in page
    assert "发送时模板：" in page
    assert "function emailTemplateSourceLabel(source)" in page
    assert "task_bound:'任务绑定模板'" in page
    assert "active_global_fallback:'发送时启用模板'" in page
    assert "default_renderer:'系统默认正文'" in page


def test_phase_18b_report_center_task_grouping_frontend_hooks():
    page = Path("api/monitor_web/index.html").read_text(encoding="utf-8")
    css = Path("api/webui/monitor/monitor.css").read_text(encoding="utf-8")
    frontend_source = page + "\n" + css

    reports_section = _task_group_view(page)
    assert "groupReportsByTask" not in page
    assert "renderReportTaskGroup" not in page
    assert "reportGroupSummaryText" not in page
    assert "formatReportFrequency" in page
    assert "function groupRunsByTask(runs)" in page
    assert "function renderTaskRunGroup(group)" in page
    assert 'return `<div class="task-run-groups">' in page
    assert "const jobId=Number(run.job_id || summary.job_id || 0) || null;" in page
    assert "if(context.jobId) return `job:${context.jobId}`;" in page
    assert "原任务已删除" in page
    assert "上下文有限" in page
    assert "历史运行：任务上下文有限" in page
    assert "平台：" in page
    assert "关键词：" in page
    assert "job_snapshot_json" not in reports_section
    assert "recipients_json" not in reports_section
    assert "smtp_password" not in reports_section
    assert "previewReport(${reportId})" not in page
    assert "previewReport(${Number(report.id)})" in page
    assert "toggleReportActionMenu" not in page
    assert "loadEmailDeliveryHistory(${Number(report.id)})" in page
    assert "/reports/${Number(report.id)}/download?type=html" in page
    assert "/leads?'+qs.toString()" not in page
    for selector in [
        ".report-task-groups",
        ".report-task-group",
        ".report-task-group-head",
        ".report-task-group-title-row",
        ".report-task-group-badges",
        ".report-task-group-count",
        ".report-task-group-meta",
        ".report-task-group.is-deleted",
        ".report-task-group.is-limited",
    ]:
        assert selector in frontend_source
    assert "@media (max-width: 1279px)" in css
    assert "@media (max-width: 767px)" in css


def test_cli_run_due_runs_only_due_enabled_jobs(monkeypatch):
    init_db()
    jobs_snapshot = _snapshot_monitor_jobs()
    run_calls: list[int] = []
    _clear_monitor_jobs()

    due_job = save_job(
        {
            "law_firm_name": "CLI到期测试律所",
            "aliases": [],
            "exclude_words": [],
            "keywords": ["CLI到期测试律所避雷"],
            "platforms": ["dy"],
            "recipients": [],
            "enable_comments": False,
            "time_window_type": "recent_1d",
            "frequency": "daily",
            "email_time": "08:00",
            "enabled": True,
        }
    )
    future_job = save_job(
        {
            "law_firm_name": "CLI未到期测试律所",
            "aliases": [],
            "exclude_words": [],
            "keywords": ["CLI未到期测试律所避雷"],
            "platforms": ["dy"],
            "recipients": [],
            "enable_comments": False,
            "time_window_type": "recent_1d",
            "frequency": "daily",
            "email_time": "23:00",
            "enabled": True,
        }
    )
    disabled_job = save_job(
        {
            "law_firm_name": "CLI暂停测试律所",
            "aliases": [],
            "exclude_words": [],
            "keywords": ["CLI暂停测试律所避雷"],
            "platforms": ["dy"],
            "recipients": [],
            "enable_comments": False,
            "time_window_type": "recent_1d",
            "frequency": "daily",
            "email_time": "08:00",
            "enabled": False,
        }
    )

    async def fake_run_job(job_id, source="manual"):
        run_calls.append(job_id)
        return {"run_id": 999, "status": "success", "summary": {}, "report": {}}

    try:
        monkeypatch.setattr(
            "api.monitoring.preflight.list_platform_status",
            lambda: [
                {"platform": "dy", "platform_label": "抖音", "profile_exists": True, "needs_login": False, "login_window_open": False},
                {"platform": "ks", "platform_label": "快手", "profile_exists": True, "needs_login": False, "login_window_open": False},
                {"platform": "xhs", "platform_label": "小红书", "profile_exists": True, "needs_login": False, "login_window_open": False},
            ],
        )
        monkeypatch.setattr(cli_module, "run_job", fake_run_job)
        result = asyncio.run(run_due_jobs(datetime(2026, 6, 12, 9, 0, 0)))
    finally:
        _restore_monitor_jobs(jobs_snapshot)

    assert result["ran"] == 1
    assert result["ok"] is True
    assert run_calls == [due_job["id"]]
    assert future_job["id"] not in run_calls
    assert disabled_job["id"] not in run_calls


def test_cli_run_due_skips_legacy_template_placeholder_jobs(monkeypatch):
    init_db()
    jobs_snapshot = _snapshot_monitor_jobs()
    run_calls: list[int] = []
    _clear_monitor_jobs()

    try:
        with get_conn() as conn:
            now = "2026-06-12T00:00:00+00:00"
            cur = conn.execute(
                """
                INSERT INTO monitor_jobs (
                    law_firm_name, aliases, exclude_words, enable_comments, time_window_type,
                    frequency, email_time, enabled, is_internal, created_at, updated_at
                ) VALUES (?, '[]', '[]', 0, 'recent_1d', 'daily', '08:00', 1, 0, ?, ?)
                """,
                ("请改成目标律所名称", now, now),
            )
            job_id = int(cur.lastrowid)
            conn.execute("INSERT INTO job_keywords (job_id, keyword) VALUES (?, ?)", (job_id, "目标律所避雷"))
            conn.execute("INSERT INTO job_platforms (job_id, platform) VALUES (?, ?)", (job_id, "dy"))

        async def fake_run_job(job_id, source="manual"):
            run_calls.append(job_id)
            return {"run_id": 999, "status": "success", "summary": {}, "report": {}}

        monkeypatch.setattr(cli_module, "run_job", fake_run_job)
        result = asyncio.run(run_due_jobs(datetime(2026, 6, 12, 9, 0, 0)))
    finally:
        _restore_monitor_jobs(jobs_snapshot)

    assert result["ran"] == 0
    assert result["skipped"] == 1
    assert result["ok"] is True
    assert run_calls == []
    assert "测试数据模板" in result["results"][0]["reason"]
    assert "平台搜索词" in result["results"][0]["reason"]


def test_cli_run_due_blocks_preflight_and_records_skipped_run(monkeypatch):
    init_db()
    jobs_snapshot = _snapshot_monitor_jobs()
    runs_snapshot = _snapshot_table("crawl_runs")
    run_calls: list[int] = []
    _clear_monitor_jobs()

    try:
        job = save_job(
            {
                "law_firm_name": "海安律所",
                "aliases": [],
                "exclude_words": [],
                "keywords": ["海安律所投诉"],
                "platforms": ["xhs"],
                "recipients": ["target@example.com"],
                "enable_comments": False,
                "time_window_type": "recent_1d",
                "frequency": "daily",
                "email_time": "08:00",
                "enabled": True,
            }
        )
        monkeypatch.setattr(
            "api.monitoring.preflight.list_platform_status",
            lambda: [
                {"platform": "dy", "platform_label": "抖音", "profile_exists": True, "needs_login": False, "login_window_open": False},
                {"platform": "ks", "platform_label": "快手", "profile_exists": True, "needs_login": False, "login_window_open": False},
                {"platform": "xhs", "platform_label": "小红书", "profile_exists": True, "needs_login": True, "login_window_open": False},
            ],
        )

        async def fake_run_job(job_id, source="manual"):
            run_calls.append(job_id)
            return {"run_id": 999, "status": "success", "summary": {}, "report": {}}

        monkeypatch.setattr(cli_module, "run_job", fake_run_job)
        result = asyncio.run(run_due_jobs(datetime(2026, 6, 12, 9, 0, 0)))
        skipped = get_run(int(result["results"][0]["run_id"]))
    finally:
        _restore_monitor_jobs(jobs_snapshot)
        _restore_table("crawl_runs", runs_snapshot)

    assert result["ran"] == 0
    assert result["skipped"] == 1
    assert result["ok"] is True
    assert run_calls == []
    assert result["results"][0]["status"] == "skipped"
    assert "重新登录" in result["results"][0]["reason"]
    assert skipped and skipped["status"] == "skipped"
    assert skipped["summary"]["source"] == "cli"
    assert skipped["summary"]["skip_type"] == "preflight_blocked"


def test_report_resend_email_updates_status(monkeypatch):
    init_db()
    job = save_job(
        {
            "law_firm_name": "重发测试律所",
            "aliases": [],
            "exclude_words": [],
            "keywords": ["重发测试律所避雷"],
            "platforms": ["dy"],
            "recipients": ["ops@example.com"],
            "enable_comments": False,
            "time_window_type": "recent_1d",
            "frequency": "daily",
            "email_time": "09:00",
            "enabled": True,
        }
    )
    run_id = create_run(job["id"])
    report = create_report(
        run_id,
        job,
        {"platforms": ["dy"], "failed_platforms": [], "new_contents": 0, "negative_count": 0, "high_count": 0},
    )

    def fake_send_report(job, report, allow_real_send=None):
        return False, "SMTP 配置未完成"

    try:
        monkeypatch.setattr("api.monitoring.reporting.send_report", fake_send_report)
        ok, error, refreshed = resend_report_email(report["id"])
        stored = get_report(report["id"])
    finally:
        _cleanup_test_records(job["id"], "")

    assert ok is False
    assert error == "SMTP 配置未完成"
    assert refreshed["email_status"] == "failed"
    assert stored and stored["email_status"] == "failed"
    assert stored["email_error"] == "SMTP 配置未完成"


def test_run_job_skips_when_cross_process_lock_exists():
    init_db()
    job = save_job(
        {
            "law_firm_name": "锁测试律所",
            "aliases": [],
            "exclude_words": [],
            "keywords": ["锁测试律所避雷"],
            "platforms": ["dy"],
            "recipients": [],
            "enable_comments": False,
            "time_window_type": "recent_1d",
            "frequency": "daily",
            "email_time": "09:00",
            "enabled": True,
        }
    )
    lock_path = runner_module.LOCKS_DIR / f"job_{job['id']}.lock"
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    lock_path.write_text("locked", encoding="utf-8")
    with get_conn() as conn:
        before = conn.execute("SELECT COUNT(*) AS n FROM crawl_runs WHERE job_id=?", (job["id"],)).fetchone()["n"]
    try:
        result = asyncio.run(run_monitor_job(job["id"]))
        with get_conn() as conn:
            after = conn.execute("SELECT COUNT(*) AS n FROM crawl_runs WHERE job_id=?", (job["id"],)).fetchone()["n"]
    finally:
        lock_path.unlink(missing_ok=True)
        _cleanup_test_records(job["id"], "")

    assert result["status"] == "already_running"
    assert result["run_id"] is None
    assert after == before


def test_run_history_keeps_job_snapshot_after_job_deleted():
    init_db()
    job = save_job(
        {
            "law_firm_name": "运行快照测试律所",
            "aliases": [],
            "exclude_words": [],
            "keywords": ["运行快照测试律所避雷"],
            "platforms": ["dy"],
            "recipients": ["ops@example.com"],
            "enable_comments": False,
            "time_window_type": "recent_1d",
            "frequency": "daily",
            "email_time": "09:00",
            "enabled": True,
        }
    )
    run_id = create_run(job["id"])
    finish_run(
        run_id,
        "success",
        {
            "job_id": job["id"],
            "law_firm_name": job["law_firm_name"],
            "keywords": job["keywords"],
            "platforms": job["platforms"],
        },
    )

    try:
        with get_conn() as conn:
            conn.execute("DELETE FROM monitor_jobs WHERE id=?", (job["id"],))
        run = get_run(run_id)
    finally:
        with get_conn() as conn:
            conn.execute("DELETE FROM reports WHERE run_id=?", (run_id,))
            conn.execute("DELETE FROM crawl_runs WHERE id=?", (run_id,))
        _cleanup_test_records(job["id"], "")

    assert run
    assert run["job_id"] == job["id"]
    assert run["law_firm_name"] == "运行快照测试律所"
    assert run["job_deleted"] is True


def test_report_history_keeps_law_firm_snapshot_after_job_deleted(monkeypatch):
    init_db()
    job = save_job(
        {
            "law_firm_name": "报告快照测试律所",
            "aliases": [],
            "exclude_words": [],
            "keywords": ["报告快照测试律所避雷"],
            "platforms": ["dy"],
            "recipients": [],
            "enable_comments": False,
            "time_window_type": "recent_1d",
            "frequency": "daily",
            "email_time": "09:00",
            "enabled": False,
        }
    )
    run_id = create_run(job["id"])
    report = create_report(
        run_id,
        job,
        {
            "job_id": job["id"],
            "law_firm_name": job["law_firm_name"],
            "platforms": job["platforms"],
            "keywords": job["keywords"],
            "failed_platforms": [],
        },
    )

    try:
        with get_conn() as conn:
            conn.execute("DELETE FROM monitor_jobs WHERE id=?", (job["id"],))
        stored = get_report(report["id"])
        listed = next(item for item in list_reports(200) if item["id"] == report["id"])
    finally:
        with get_conn() as conn:
            conn.execute("DELETE FROM reports WHERE id=?", (report["id"],))
            conn.execute("DELETE FROM crawl_runs WHERE id=?", (run_id,))
        _cleanup_test_records(job["id"], "")

    assert stored and stored["law_firm_name"] == "报告快照测试律所"
    assert listed["law_firm_name"] == "报告快照测试律所"
    assert stored["job_deleted"] is True


def test_list_runs_limit_zero_returns_all_recent_rows():
    init_db()
    job = save_job(
        {
            "law_firm_name": "全部运行测试律所",
            "aliases": [],
            "exclude_words": [],
            "keywords": ["全部运行测试律所避雷"],
            "platforms": ["dy"],
            "recipients": [],
            "enable_comments": False,
            "time_window_type": "recent_1d",
            "frequency": "daily",
            "email_time": "09:00",
            "enabled": True,
        }
    )
    run_ids = [create_run(job["id"]) for _ in range(2)]
    try:
        assert len([r for r in list_runs(1) if r["id"] in run_ids]) == 1
        assert len([r for r in list_runs(0) if r["id"] in run_ids]) == 2
    finally:
        _cleanup_test_records(job["id"], "")


def test_running_run_keeps_job_snapshot_before_finish():
    init_db()
    job = save_job(
        {
            "law_firm_name": "海安律所",
            "aliases": ["海安律师事务所"],
            "exclude_words": [],
            "keywords": ["海安律所避雷", "海安律所退费", "海安律所投诉"],
            "platforms": ["dy"],
            "recipients": [],
            "enable_comments": False,
            "time_window_type": "recent_1d",
            "frequency": "daily",
            "email_time": "09:00",
            "enabled": True,
        }
    )
    run_id = create_run(
        job["id"],
        {
            "job_id": job["id"],
            "law_firm_name": job["law_firm_name"],
            "platforms": job["platforms"],
            "keywords": job["keywords"],
        },
    )
    try:
        run = get_run(run_id)
    finally:
        _cleanup_test_records(job["id"], "")

    assert run
    assert run["status"] == "running"
    assert run["display_law_firm_name"] == "海安律所"
    assert run["summary"]["keywords"] == ["海安律所避雷", "海安律所退费", "海安律所投诉"]
    assert run["summary"]["duration_seconds"] >= 0


def test_list_reports_limit_zero_returns_all_recent_rows():
    init_db()
    job = save_job(
        {
            "law_firm_name": "海安律所",
            "aliases": [],
            "exclude_words": [],
            "keywords": ["海安律所避雷"],
            "platforms": ["dy"],
            "recipients": [],
            "enable_comments": False,
            "time_window_type": "recent_1d",
            "frequency": "daily",
            "email_time": "09:00",
            "enabled": True,
        }
    )
    run_ids = [create_run(job["id"]) for _ in range(2)]
    reports = []
    try:
        for run_id in run_ids:
            reports.append(
                create_report(
                    run_id,
                    job,
                    {
                        "job_id": job["id"],
                        "law_firm_name": job["law_firm_name"],
                        "platforms": ["dy"],
                        "failed_platforms": [],
                    },
                )
            )
        report_ids = {report["id"] for report in reports}
        assert len([r for r in list_reports(1) if r["id"] in report_ids]) == 1
        assert len([r for r in list_reports(0) if r["id"] in report_ids]) == 2
    finally:
        _cleanup_test_records(job["id"], "")


def test_list_leads_limit_zero_returns_all_recent_rows():
    init_db()
    job = save_job(
        {
            "law_firm_name": "海安律所",
            "aliases": [],
            "exclude_words": [],
            "keywords": ["海安律所避雷"],
            "platforms": ["dy"],
            "recipients": [],
            "enable_comments": False,
            "time_window_type": "recent_1d",
            "frequency": "daily",
            "email_time": "09:00",
            "enabled": True,
        }
    )
    run_id = create_run(job["id"])
    content_ids = ["pytest_all_leads_001", "pytest_all_leads_002"]
    now_ts = int(datetime.now(timezone.utc).timestamp())
    items = [
        {"aweme_id": content_id, "title": f"海安律所避雷 {index}", "create_time": now_ts}
        for index, content_id in enumerate(content_ids, start=1)
    ]
    try:
        ingest_outputs(job, run_id, "dy", items, [])
        stored_ids = {item["content_id"] for item in list_leads(0) if item["content_id"] in content_ids}
        assert len([item for item in list_leads(1) if item["content_id"] in content_ids]) == 1
        assert stored_ids == set(content_ids)
    finally:
        for content_id in content_ids:
            _cleanup_test_records(job["id"], content_id)


def test_leads_api_can_scope_items_to_selected_report():
    init_db()
    job = save_job(
        {
            "law_firm_name": "海安律所",
            "aliases": ["海安律师事务所"],
            "exclude_words": [],
            "keywords": ["海安律所避雷", "海安律所退费", "海安律所投诉"],
            "platforms": ["dy"],
            "recipients": [],
            "enable_comments": False,
            "time_window_type": "recent_1d",
            "frequency": "daily",
            "email_time": "09:00",
            "enabled": True,
        }
    )
    now_ts = int(datetime.now(timezone.utc).timestamp())
    first_id = "pytest_report_scope_001"
    second_id = "pytest_report_scope_002"
    try:
        run1 = create_run(job["id"])
        ingest_outputs(job, run1, "dy", [{"aweme_id": first_id, "title": "海安律所避雷", "create_time": now_ts}], [])
        report1 = create_report(run1, job, {"job_id": job["id"], "law_firm_name": job["law_firm_name"], "platforms": ["dy"], "failed_platforms": []})
        run2 = create_run(job["id"])
        ingest_outputs(job, run2, "dy", [{"aweme_id": second_id, "title": "海安律所退费", "create_time": now_ts}], [])
        create_report(run2, job, {"job_id": job["id"], "law_firm_name": job["law_firm_name"], "platforms": ["dy"], "failed_platforms": []})

        scoped = asyncio.run(monitor_router.leads(report_id=report1["id"], limit=0))["leads"]
    finally:
        _cleanup_test_records(job["id"], first_id)
        _cleanup_test_records(job["id"], second_id)

    assert [item["content_id"] for item in scoped] == [first_id]


def test_leads_api_can_scope_items_to_selected_run():
    init_db()
    job = save_job(
        {
            "law_firm_name": "海安律所",
            "aliases": [],
            "exclude_words": [],
            "keywords": ["海安律所避雷", "海安律所退费"],
            "platforms": ["dy"],
            "recipients": [],
            "enable_comments": False,
            "time_window_type": "recent_1d",
            "frequency": "daily",
            "email_time": "09:00",
            "enabled": True,
        }
    )
    now_ts = int(datetime.now(timezone.utc).timestamp())
    first_id = "pytest_run_scope_001"
    second_id = "pytest_run_scope_002"
    try:
        run1 = create_run(job["id"])
        ingest_outputs(job, run1, "dy", [{"aweme_id": first_id, "title": "海安律所避雷", "create_time": now_ts}], [])
        run2 = create_run(job["id"])
        ingest_outputs(job, run2, "dy", [{"aweme_id": second_id, "title": "海安律所退费", "create_time": now_ts}], [])

        scoped = asyncio.run(monitor_router.leads(run_id=run1, limit=0))["leads"]
    finally:
        _cleanup_test_records(job["id"], first_id)
        _cleanup_test_records(job["id"], second_id)

    assert [item["content_id"] for item in scoped] == [first_id]


def test_phase_19b_collection_progress_tolerates_partial_outputs_and_keeps_final_counts(tmp_path):
    init_db()
    job = save_job(
        {
            "law_firm_name": "进度测试律所",
            "aliases": [],
            "exclude_words": [],
            "keywords": ["进度测试律所避雷"],
            "platforms": ["dy"],
            "recipients": [],
            "enable_comments": False,
            "time_window_type": "recent_1d",
            "frequency": "daily",
            "email_time": "09:00",
            "enabled": True,
        }
    )
    run_id = create_run(job["id"], {"job_id": job["id"], "raw_contents": 0, "new_contents": 0}, timeout_seconds=120)
    json_dir = tmp_path / "douyin" / "json"
    jsonl_dir = tmp_path / "douyin" / "jsonl"
    json_dir.mkdir(parents=True)
    jsonl_dir.mkdir(parents=True)
    now_ts = int(datetime.now(timezone.utc).timestamp())
    (json_dir / "search_contents_valid.json").write_text(
        json.dumps(
            [
                {
                    "aweme_id": "pytest_phase19b_valid_001",
                    "title": "进度测试律所避雷",
                    "desc": "可正常入库",
                    "create_time": now_ts,
                }
            ],
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    (json_dir / "search_contents_empty.json").write_text("", encoding="utf-8")
    (json_dir / "search_contents_partial.json").write_text('{"aweme_id": "broken"', encoding="utf-8")
    (jsonl_dir / "search_contents_partial.jsonl").write_text(
        json.dumps(
            {
                "aweme_id": "pytest_phase19b_valid_002",
                "title": "进度测试律所避雷二",
                "desc": "JSONL 部分可读",
                "create_time": now_ts,
            },
            ensure_ascii=False,
        )
        + "\n{bad",
        encoding="utf-8",
    )
    try:
        provisional = runner_module._update_collection_progress(run_id, "dy", tmp_path, phase="collecting")
        assert provisional["collection_progress"]["provisional"] is True
        assert provisional["collection_progress"]["final"] is False
        assert provisional["collection_progress"]["raw_items_seen"] == 2
        assert provisional["collection_progress"]["malformed_files"] >= 2
        assert provisional["collection_progress"]["platforms"]["dy"]["empty_files"] >= 1
        assert provisional["raw_contents"] == 0
        assert provisional["new_contents"] == 0

        contents, comments = collect_platform_outputs(tmp_path, "dy")
        result = ingest_outputs(job, run_id, "dy", contents, comments)
        runner_module._finalize_collection_progress(run_id, "dy", result)
        summary = {"job_id": job["id"], **result}
        runner_module._sync_progress_from_stored_summary(run_id, summary, finalize=True)
        finish_run(run_id, "success", summary)
        run = get_run(run_id)
    finally:
        _cleanup_test_records(job["id"], "pytest_phase19b_valid_001")
        _cleanup_test_records(job["id"], "pytest_phase19b_valid_002")

    assert result["raw_contents"] == 2
    assert result["filtered_contents"] == 2
    assert result["new_contents"] == 2
    assert run
    assert run["summary"]["raw_contents"] == 2
    assert run["summary"]["new_contents"] == 2
    assert run["summary"]["collection_progress"]["final"] is True
    assert run["summary"]["collection_progress"]["final_raw_contents"] == 2
    assert run["summary"]["collection_progress"]["final_new_contents"] == 2
    assert run["collection_progress"]["platforms"]["dy"]["final"] is True


def test_phase_20b_ai_trace_persistence_redaction_truncation_and_retention():
    init_db()
    settings_snapshot = _snapshot_table("system_settings")
    traces_snapshot = _snapshot_table("ai_evaluation_traces")
    try:
        with get_conn() as conn:
            conn.execute("DELETE FROM system_settings")
            conn.execute("DELETE FROM ai_evaluation_traces")
        settings = list_runtime_settings()
        assert settings["ai_trace_retention_days"]["value"] == 30
        assert settings["ai_trace_retention_days"]["group"] == "Retention"
        updated = save_runtime_settings({"ai_trace_retention_days": 7}, actor_id=1)
        assert updated["ai_trace_retention_days"]["value"] == 7
        assert get_runtime_setting_value("ai_trace_retention_days") == 7

        trace = save_ai_evaluation_trace(
            {
                "workspace_id": 1,
                "run_id": 501,
                "raw_content_id": 601,
                "ai_evaluation_id": 701,
                "attempt_index": 1,
                "status": "ok",
                "provider": "openai",
                "model": "trace-model",
                "prompt_snapshot": "判断律所负面 " + ("P" * 20000),
                "input_payload": {
                    "law_firm_name": "Phase20B律所",
                    "profile_path": "C:\\secret\\profile",
                    "comments": [{"content": "C" * 800, "author_name": "用户"} for _ in range(25)],
                },
                "request_snapshot": {
                    "headers": {
                        "Authorization": "Bearer sk-secret-token",
                        "Cookie": "session=secret",
                    },
                    "json": {
                        "messages": [{"role": "user", "content": "hello"}],
                        "proxy_url": "http://user:pass@example.com:8080",
                    },
                },
                "response_snapshot": json.dumps({"raw": "R" * 30000, "api_key": "sk-response-secret"}),
                "parsed_result": {"is_related": True, "is_negative": True, "risk_level": "high"},
                "error_message": "authorization: Bearer sk-error-secret at C:\\server\\local\\path",
                "duration_ms": 123,
                "started_at": "2026-06-18T00:00:00Z",
                "finished_at": "2026-06-18T00:00:01Z",
            }
        )
        stored = get_ai_evaluation_trace(ai_evaluation_id=701)
        state = ai_evaluation_trace_state(ai_evaluation_id=701)
        missing_state = ai_evaluation_trace_state(ai_evaluation_id=999999)

        with get_conn() as conn:
            conn.execute(
                "UPDATE ai_evaluation_traces SET created_at=? WHERE id=?",
                ((datetime.now(timezone.utc) - timedelta(days=8)).isoformat(), trace["id"]),
            )
        cleanup = cleanup_ai_evaluation_traces()
        after_cleanup = get_ai_evaluation_trace(ai_evaluation_id=701)
    finally:
        _restore_table("ai_evaluation_traces", traces_snapshot)
        _restore_table("system_settings", settings_snapshot)

    combined = json.dumps(stored, ensure_ascii=False)
    assert stored["provider"] == "openai"
    assert stored["model"] == "trace-model"
    assert stored["duration_ms"] == 123
    assert stored["limited_context"] is False
    assert stored["parsed_result"]["_trace_meta"]["truncated"] is True
    assert "prompt_snapshot" in stored["parsed_result"]["_trace_meta"]["truncated_fields"]
    assert stored["input_payload"]["truncated"] is True
    assert len(stored["input_payload"]["comments"]) == 20
    assert "sk-secret" not in combined
    assert "session=secret" not in combined
    assert "user:pass" not in combined
    assert "C:\\secret" not in combined
    assert "C:\\server" not in combined
    assert state["status"] == "available"
    assert missing_state["limited_context"] is True
    assert "历史记录未保存完整入参/出参" in missing_state["message"]
    assert cleanup["deleted"] == 1
    assert cleanup["retention_days"] == 7
    assert after_cleanup is None


def test_phase_20b_evaluate_new_contents_persists_success_and_fallback_traces(monkeypatch):
    init_db()
    snapshots = {
        "reports": _snapshot_table("reports"),
        "crawl_runs": _snapshot_table("crawl_runs"),
        "raw_contents": _snapshot_table("raw_contents"),
        "raw_comments": _snapshot_table("raw_comments"),
        "ai_evaluations": _snapshot_table("ai_evaluations"),
        "ai_evaluation_traces": _snapshot_table("ai_evaluation_traces"),
    }
    jobs_snapshot = _snapshot_monitor_jobs()
    try:
        _clear_monitor_jobs()
        with get_conn() as conn:
            for table in ["reports", "crawl_runs", "raw_contents", "raw_comments", "ai_evaluations", "ai_evaluation_traces"]:
                conn.execute(f"DELETE FROM {table}")
        job = save_job(
            {
                "law_firm_name": "Phase20B追溯律所",
                "keywords": ["Phase20B追溯律所投诉"],
                "platforms": ["dy"],
                "recipients": [],
                "enable_comments": True,
                "time_window_type": "recent_1d",
                "enabled": True,
            }
        )
        run_id = create_run(job["id"], {"job_id": job["id"], "phase_20b_trace": True}, timeout_seconds=120)
        now_ts = int(datetime.now(timezone.utc).timestamp())
        ingested = ingest_outputs(
            job,
            run_id,
            "dy",
            [
                {
                    "aweme_id": "pytest_phase20b_trace_ok",
                    "title": "Phase20B追溯律所投诉",
                    "desc": "服务争议",
                    "create_time": now_ts,
                },
                {
                    "aweme_id": "pytest_phase20b_trace_fail",
                    "title": "Phase20B追溯律所待复核",
                    "desc": "需要人工判断",
                    "create_time": now_ts,
                },
            ],
            [
                {
                    "comment_id": "pytest_phase20b_comment_1",
                    "aweme_id": "pytest_phase20b_trace_ok",
                    "content": "Phase20B追溯律所沟通问题",
                    "create_time": now_ts,
                }
            ],
        )

        async def fake_evaluate(job_arg, content, comments):
            if content["content_id"] == "pytest_phase20b_trace_fail":
                raise RuntimeError("provider failed api_key=sk-secret Authorization: Bearer hidden")
            return {
                "status": "ok",
                "is_related": True,
                "is_negative": True,
                "risk_level": "high",
                "reason": "命中投诉",
                "evidence_quotes": [content["title"]],
                "recommended_action": "人工复核",
                "raw_response": '{"risk_level":"high","api_key":"sk-secret"}',
                "_ai_trace": {
                    "workspace_id": 1,
                    "run_id": run_id,
                    "raw_content_id": content["id"],
                    "attempt_index": 1,
                    "status": "ok",
                    "provider": "openai",
                    "model": "trace-model",
                    "prompt_snapshot": "prompt api_key=sk-secret",
                    "input_payload": {"title": content["title"], "comments": comments},
                    "request_snapshot": {"headers": {"Authorization": "Bearer sk-secret"}},
                    "response_snapshot": '{"risk_level":"high","api_key":"sk-secret"}',
                    "parsed_result": {"risk_level": "high"},
                    "duration_ms": 10,
                },
            }

        monkeypatch.setattr(runner_module, "evaluate_content", fake_evaluate)
        eval_summary = asyncio.run(evaluate_new_contents(job, run_id, ingested["content_db_ids"]))
        report = create_report(run_id, job, {"job_id": job["id"], "platforms": ["dy"], **ingested, **eval_summary})
        with get_conn() as conn:
            traces = [dict(row) for row in conn.execute("SELECT * FROM ai_evaluation_traces ORDER BY raw_content_id").fetchall()]
            evaluations = [dict(row) for row in conn.execute("SELECT * FROM ai_evaluations ORDER BY raw_content_id").fetchall()]
    finally:
        _restore_monitor_jobs(jobs_snapshot)
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)

    assert eval_summary["negative_count"] == 1
    assert eval_summary["high_count"] == 1
    assert eval_summary["pending_review_count"] == 1
    assert report["summary"]["pending_review_count"] == 1
    assert len(evaluations) == 2
    assert len(traces) == 2
    assert {row["status"] for row in traces} == {"ok", "pending_review"}
    combined = json.dumps(traces, ensure_ascii=False)
    assert "sk-secret" not in combined
    assert "Bearer hidden" not in combined
    assert "provider failed" in combined
    assert all(row["ai_evaluation_id"] for row in traces)


def test_phase_20b_trace_write_failure_does_not_block_evaluation_or_report(monkeypatch):
    init_db()
    snapshots = {
        "reports": _snapshot_table("reports"),
        "crawl_runs": _snapshot_table("crawl_runs"),
        "raw_contents": _snapshot_table("raw_contents"),
        "raw_comments": _snapshot_table("raw_comments"),
        "ai_evaluations": _snapshot_table("ai_evaluations"),
        "ai_evaluation_traces": _snapshot_table("ai_evaluation_traces"),
    }
    jobs_snapshot = _snapshot_monitor_jobs()
    try:
        _clear_monitor_jobs()
        with get_conn() as conn:
            for table in ["reports", "crawl_runs", "raw_contents", "raw_comments", "ai_evaluations", "ai_evaluation_traces"]:
                conn.execute(f"DELETE FROM {table}")
        job = save_job(
            {
                "law_firm_name": "Phase20B非阻塞律所",
                "keywords": ["Phase20B非阻塞律所投诉"],
                "platforms": ["dy"],
                "recipients": [],
                "enable_comments": False,
                "time_window_type": "recent_1d",
                "enabled": True,
            }
        )
        run_id = create_run(job["id"], {"job_id": job["id"]}, timeout_seconds=120)
        ingested = ingest_outputs(
            job,
            run_id,
            "dy",
            [{"aweme_id": "pytest_phase20b_trace_nonblocking", "title": "Phase20B非阻塞律所投诉", "create_time": int(datetime.now(timezone.utc).timestamp())}],
            [],
        )

        async def fake_evaluate(job_arg, content, comments):
            return {
                "status": "pending_review",
                "is_related": True,
                "is_negative": False,
                "risk_level": "low",
                "reason": "人工复核",
                "evidence_quotes": [content["title"]],
                "recommended_action": "人工复核",
                "raw_response": "",
                "_ai_trace": {
                    "workspace_id": 1,
                    "run_id": run_id,
                    "raw_content_id": content["id"],
                    "status": "pending_review",
                    "provider": "openai",
                    "model": "trace-model",
                },
            }

        def boom(trace):
            raise RuntimeError("trace write failed")

        monkeypatch.setattr(runner_module, "evaluate_content", fake_evaluate)
        monkeypatch.setattr(runner_module, "save_ai_evaluation_trace", boom)
        eval_summary = asyncio.run(evaluate_new_contents(job, run_id, ingested["content_db_ids"]))
        report = create_report(run_id, job, {"job_id": job["id"], "platforms": ["dy"], **ingested, **eval_summary})
        with get_conn() as conn:
            eval_count = conn.execute("SELECT COUNT(*) AS n FROM ai_evaluations WHERE run_id=?", (run_id,)).fetchone()["n"]
            trace_count = conn.execute("SELECT COUNT(*) AS n FROM ai_evaluation_traces WHERE run_id=?", (run_id,)).fetchone()["n"]
    finally:
        _restore_monitor_jobs(jobs_snapshot)
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)

    assert eval_summary["pending_review_count"] == 1
    assert eval_count == 1
    assert trace_count == 0
    assert Path(report["html_path"]).exists()


def test_phase_20c_run_detail_api_scope_filters_and_redacted_trace():
    from api import main as api_main

    init_db()
    snapshots = {
        "audit_logs": _snapshot_table("audit_logs"),
        "user_sessions": _snapshot_table("user_sessions"),
        "users": _snapshot_table("users"),
        "reports": _snapshot_table("reports"),
        "crawl_runs": _snapshot_table("crawl_runs"),
        "raw_contents": _snapshot_table("raw_contents"),
        "raw_comments": _snapshot_table("raw_comments"),
        "ai_evaluations": _snapshot_table("ai_evaluations"),
        "ai_evaluation_traces": _snapshot_table("ai_evaluation_traces"),
        "email_delivery_logs": _snapshot_table("email_delivery_logs"),
    }
    jobs_snapshot = _snapshot_monitor_jobs()
    try:
        _clear_monitor_jobs()
        with get_conn() as conn:
            for table in ["audit_logs", "user_sessions", "users", "reports", "crawl_runs", "raw_contents", "raw_comments", "ai_evaluations", "ai_evaluation_traces", "email_delivery_logs"]:
                conn.execute(f"DELETE FROM {table}")
        admin = bootstrap_admin_from_env("phase20c-admin@example.com", "AdminPass123!", "Phase20C Admin")
        user1 = save_user({"email": "phase20c-user1@example.com", "display_name": "User1", "password": "UserPass123!", "role": "normal"}, actor_id=int(admin["id"]))
        user2 = save_user({"email": "phase20c-user2@example.com", "display_name": "User2", "password": "UserPass123!", "role": "normal"}, actor_id=int(admin["id"]))
        job1 = save_job(
            {
                "law_firm_name": "Phase20C律所",
                "keywords": ["Phase20C律所投诉", "Phase20C律所口碑"],
                "platforms": ["dy", "xhs"],
                "recipients": ["ops@example.com"],
                "enabled": True,
            },
            actor=user1,
        )
        job2 = save_job(
            {
                "law_firm_name": "Other20C律所",
                "keywords": ["Other20C律所投诉"],
                "platforms": ["ks"],
                "recipients": ["ops@example.com"],
                "enabled": True,
            },
            actor=user2,
        )
        run1 = create_run(job1["id"], {"job_id": job1["id"], "law_firm_name": job1["law_firm_name"], "platforms": ["dy", "xhs"]})
        run2 = create_run(job2["id"], {"job_id": job2["id"], "law_firm_name": job2["law_firm_name"], "platforms": ["ks"]})
        now = datetime.now(timezone.utc).isoformat()
        now_ts = int(datetime.now(timezone.utc).timestamp())
        with get_conn() as conn:
            raw_ids: list[int] = []
            rows = [
                ("dy", "pytest_phase20c_high", "Phase20C律所投诉", "Phase20C律所多人投诉", "服务争议", user1["id"], run1, job1["id"]),
                ("xhs", "pytest_phase20c_pending", "Phase20C律所口碑", "Phase20C律所待复核", "待判断", user1["id"], run1, job1["id"]),
                ("dy", "pytest_phase20c_old", "Phase20C律所投诉", "Phase20C律所旧评估", "旧数据", user1["id"], run1, job1["id"]),
                ("ks", "pytest_phase20c_other", "Other20C律所投诉", "Other20C律所投诉", "其他用户数据", user2["id"], run2, job2["id"]),
            ]
            for platform_code, content_id, keyword, title_text, desc, owner_id, run_id_value, job_id_value in rows:
                conn.execute(
                    """
                    INSERT INTO raw_contents (
                        workspace_id, platform, content_id, job_id, run_id,
                        law_firm_name, source_keyword, title, description,
                        author_name, content_url, cover_url, publish_time,
                        comment_count, raw_json, first_seen_at, last_seen_at,
                        created_by, updated_by
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        1,
                        platform_code,
                        content_id,
                        job_id_value,
                        run_id_value,
                        "Phase20C律所" if run_id_value == run1 else "Other20C律所",
                        keyword,
                        title_text,
                        desc,
                        "用户",
                        "https://example.com/" + content_id,
                        "",
                        now_ts,
                        0,
                        "{}",
                        now,
                        now,
                        owner_id,
                        owner_id,
                    ),
                )
                raw_ids.append(int(conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]))
            eval_rows = [
                (raw_ids[0], run1, "ok", 1, 1, "high", "高风险投诉", '["多人投诉"]', "人工复核", user1["id"]),
                (raw_ids[1], run1, "pending_review", 1, 0, "low", "AI 超时，待复核", "[]", "人工复核", user1["id"]),
                (raw_ids[2], run1, "ok", 1, 0, "low", "旧评估无风险", "[]", "观察", user1["id"]),
                (raw_ids[3], run2, "ok", 1, 1, "high", "其他用户投诉", '["其他"]', "人工复核", user2["id"]),
            ]
            eval_ids: list[int] = []
            for raw_id, run_id_value, status_value, related, negative, risk_value, reason_value, quotes, action, owner_id in eval_rows:
                conn.execute(
                    """
                    INSERT INTO ai_evaluations (
                        workspace_id, raw_content_id, run_id, status, is_related,
                        is_negative, risk_level, reason, evidence_quotes,
                        recommended_action, raw_response, created_at, created_by,
                        updated_by
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (1, raw_id, run_id_value, status_value, related, negative, risk_value, reason_value, quotes, action, '{"api_key":"sk-raw-secret"}', now, owner_id, owner_id),
                )
                eval_ids.append(int(conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]))
        trace = save_ai_evaluation_trace(
            {
                "workspace_id": 1,
                "run_id": run1,
                "raw_content_id": raw_ids[0],
                "ai_evaluation_id": eval_ids[0],
                "status": "ok",
                "provider": "openai",
                "model": "phase20c-model",
                "prompt_snapshot": "prompt api_key=sk-secret C:\\server\\profile",
                "input_payload": {
                    "law_firm_name": "Phase20C律所",
                    "platform": "抖音",
                    "platform_code": "dy",
                    "source_keyword": "Phase20C律所投诉",
                    "title": "Phase20C律所多人投诉",
                    "description": "服务争议",
                    "comments": ["授权 Authorization: Bearer hidden"],
                    "profile_path": "C:\\server\\profile",
                },
                "request_snapshot": {"headers": {"Authorization": "Bearer sk-secret", "Cookie": "sid=secret"}, "smtp_password": "mail-secret"},
                "response_snapshot": '{"risk":"high","api_key":"sk-secret","cookie":"sid=secret"}',
                "parsed_result": {"status": "ok", "is_related": True, "is_negative": True, "risk_level": "high", "reason": "高风险投诉", "evidence_quotes": ["多人投诉"], "recommended_action": "人工复核"},
                "error_message": "proxy_url=http://user:pass@example.com:8080",
                "duration_ms": 321,
            }
        )
        report = create_report(run1, job1, {"job_id": job1["id"], "platforms": ["dy", "xhs"], "negative_count": 1, "high_count": 1, "pending_review_count": 1})
        other_report = create_report(run2, job2, {"job_id": job2["id"], "platforms": ["ks"], "negative_count": 1, "high_count": 1})
        record_email_delivery_log(
            {
                "workspace_id": 1,
                "job_id": job1["id"],
                "report_id": report["id"],
                "send_window_key": "phase20c_window",
                "send_type": "manual_resend",
                "status": "failed",
                "error_message": "smtp_password=mail-secret Authorization: Bearer hidden",
                "effective_recipients": ["ops@example.com"],
                "effective_recipient_source": "task_recipients",
            }
        )
        transport = httpx.ASGITransport(app=api_main.app)

        async def exercise() -> None:
            async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as user_client:
                login = await user_client.post("/api/auth/login", json={"email": "phase20c-user1@example.com", "password": "UserPass123!"})
                assert login.status_code == 200
                detail_response = await user_client.get(f"/api/monitor/runs/{run1}/detail", params={"ai_limit": 2, "risk": "high", "platform": "dy", "keyword": "投诉", "title": "多人"})
                assert detail_response.status_code == 200
                detail = detail_response.json()["detail"]
                assert detail["run"]["id"] == run1
                assert detail["ai_pagination"]["total"] == 1
                assert len(detail["ai_evaluations"]) == 1
                assert detail["ai_evaluations"][0]["evaluation_id"] == eval_ids[0]
                assert detail["reports"][0]["id"] == report["id"]
                assert detail["email_delivery_logs"][0]["status"] == "failed"
                report_detail_response = await user_client.get(
                    f"/api/monitor/runs/{run1}/detail",
                    params={
                        "report_id": report["id"],
                        "risk": "pending",
                        "platform": "xhs",
                        "keyword": "口碑",
                        "title": "待复核",
                    },
                )
                assert report_detail_response.status_code == 200
                report_detail = report_detail_response.json()["detail"]
                assert report_detail["ai_filters"]["report_id"] == report["id"]
                assert report_detail["ai_filters"]["risk"] == "pending"
                assert report_detail["ai_pagination"]["total"] == 1
                assert report_detail["ai_evaluations"][0]["evaluation_id"] == eval_ids[1]
                cross_report_filter = await user_client.get(f"/api/monitor/runs/{run1}/detail", params={"report_id": other_report["id"]})
                assert cross_report_filter.status_code == 404
                own_eval = await user_client.get(f"/api/monitor/runs/{run1}/ai-evaluations/{eval_ids[0]}")
                assert own_eval.status_code == 200
                own_payload = own_eval.json()["evaluation"]
                assert "debug" not in own_payload["trace"]
                assert own_payload["trace"]["business_input"]["title"] == "Phase20C律所多人投诉"
                assert own_payload["trace"]["structured_output"]["risk_level"] == "high"
                old_eval = await user_client.get(f"/api/monitor/runs/{run1}/ai-evaluations/{eval_ids[2]}")
                assert old_eval.status_code == 200
                assert old_eval.json()["evaluation"]["trace"]["limited_context"] is True
                other_detail = await user_client.get(f"/api/monitor/runs/{run2}/detail")
                assert other_detail.status_code == 404

            async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as user2_client:
                login = await user2_client.post("/api/auth/login", json={"email": "phase20c-user2@example.com", "password": "UserPass123!"})
                assert login.status_code == 200
                cross_eval = await user2_client.get(f"/api/monitor/runs/{run1}/ai-evaluations/{eval_ids[0]}")
                assert cross_eval.status_code == 404

            async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as admin_client:
                login = await admin_client.post("/api/auth/login", json={"email": "phase20c-admin@example.com", "password": "AdminPass123!"})
                assert login.status_code == 200
                detail_response = await admin_client.get(f"/api/monitor/runs/{run1}/detail", params={"ai_limit": 2, "ai_page": 1})
                assert detail_response.status_code == 200
                detail = detail_response.json()["detail"]
                assert detail["ai_pagination"]["total"] == 3
                assert len(detail["ai_evaluations"]) == 2
                second_page = await admin_client.get(f"/api/monitor/runs/{run1}/detail", params={"ai_limit": 2, "ai_page": 2})
                assert second_page.status_code == 200
                assert len(second_page.json()["detail"]["ai_evaluations"]) == 1
                admin_eval = await admin_client.get(f"/api/monitor/runs/{run1}/ai-evaluations/{eval_ids[0]}")
                assert admin_eval.status_code == 200
                admin_payload = admin_eval.json()["evaluation"]
                assert "debug" in admin_payload["trace"]
                assert admin_payload["trace"]["provider"] == "openai"
                combined = json.dumps({"detail": detail, "eval": admin_payload}, ensure_ascii=False)
                forbidden = [
                    "sk-secret",
                    "sk-raw-secret",
                    "Bearer hidden",
                    "sid=secret",
                    "mail-secret",
                    "user:pass",
                    "C:\\server",
                    "profile_path",
                    "smtp_password",
                    "Authorization: Bearer",
                ]
                for marker in forbidden:
                    assert marker not in combined
                assert "[REDACTED]" in combined or "[PATH_REDACTED]" in combined

        asyncio.run(exercise())
    finally:
        _restore_monitor_jobs(jobs_snapshot)
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)


def test_phase_19c_ai_progress_updates_and_final_counts_are_terminal_safe(monkeypatch):
    init_db()
    snapshots = {
        "reports": _snapshot_table("reports"),
        "crawl_runs": _snapshot_table("crawl_runs"),
        "raw_contents": _snapshot_table("raw_contents"),
        "raw_comments": _snapshot_table("raw_comments"),
        "ai_evaluations": _snapshot_table("ai_evaluations"),
    }
    jobs_snapshot = _snapshot_monitor_jobs()
    content_ids = [
        "pytest_phase_19c_high",
        "pytest_phase_19c_suspected",
        "pytest_phase_19c_fallback",
    ]
    seen_progress: list[dict] = []
    try:
        _clear_monitor_jobs()
        with get_conn() as conn:
            for table in ["reports", "crawl_runs", "raw_contents", "raw_comments", "ai_evaluations"]:
                conn.execute(f"DELETE FROM {table}")
        job = save_job(
            {
                "law_firm_name": "Phase19C进度律所",
                "keywords": ["Phase19C进度律所投诉"],
                "platforms": ["dy"],
                "recipients": [],
                "enable_comments": False,
                "time_window_type": "recent_1d",
                "enabled": True,
            }
        )
        run_id = create_run(job["id"], {"job_id": job["id"], "phase_7_1_lifecycle": True}, timeout_seconds=120)
        now_ts = int(datetime.now(timezone.utc).timestamp())
        ingested = ingest_outputs(
            job,
            run_id,
            "dy",
            [
                {
                    "aweme_id": content_ids[0],
                    "title": "Phase19C进度律所多人投诉",
                    "desc": "多人退费争议",
                    "create_time": now_ts,
                },
                {
                    "aweme_id": content_ids[1],
                    "title": "Phase19C进度律所收费争议",
                    "desc": "服务体验较差",
                    "create_time": now_ts,
                },
                {
                    "aweme_id": content_ids[2],
                    "title": "Phase19C进度律所待判断",
                    "desc": "需要人工复核",
                    "create_time": now_ts,
                },
            ],
            [],
        )

        async def fake_evaluate(job_arg, content, comments):
            run = get_run(run_id)
            seen_progress.append(run["summary"].get("ai_progress") or {})
            cid = content.get("content_id")
            if cid == content_ids[0]:
                return {
                    "status": "ok",
                    "is_related": True,
                    "is_negative": True,
                    "risk_level": "high",
                    "reason": "多人投诉",
                    "evidence_quotes": [content.get("title")],
                    "recommended_action": "人工复核",
                    "raw_response": "{}",
                }
            if cid == content_ids[1]:
                return {
                    "status": "ok",
                    "is_related": True,
                    "is_negative": True,
                    "risk_level": "medium",
                    "reason": "收费争议",
                    "evidence_quotes": [content.get("title")],
                    "recommended_action": "观察",
                    "raw_response": "{}",
                }
            raise RuntimeError("provider failed api_key=secret")

        monkeypatch.setattr(runner_module, "evaluate_content", fake_evaluate)

        eval_summary = asyncio.run(evaluate_new_contents(job, run_id, ingested["content_db_ids"]))
        run = get_run(run_id)
        final_progress = run["summary"]["ai_progress"]
        stale = runner_module._merge_ai_progress_summary(
            run_id,
            {
                "total_candidates": 3,
                "successful_evaluations": 0,
                "failed_fallback_evaluations": 0,
                "pending_review_items": 0,
                "manual_review_count": 0,
                "negative_count": 0,
                "high_count": 0,
                "unresolved_items": 3,
                "evaluated_items": 0,
                "final": False,
            },
            negative_count=0,
            high_count=0,
            pending_review_count=0,
        )
        finish_run(run_id, "success", {"job_id": job["id"], **ingested, **eval_summary})
        report = create_report(run_id, job, {"job_id": job["id"], "platforms": ["dy"], **ingested, **eval_summary})
        after_terminal = runner_module._merge_ai_progress_summary(
            run_id,
            {
                "total_candidates": 3,
                "successful_evaluations": 1,
                "failed_fallback_evaluations": 0,
                "pending_review_items": 0,
                "manual_review_count": 0,
                "negative_count": 1,
                "high_count": 0,
                "unresolved_items": 2,
                "evaluated_items": 1,
                "final": False,
            },
            negative_count=1,
            high_count=0,
            pending_review_count=0,
        )
        terminal_run = get_run(run_id)
        with get_conn() as conn:
            rows = conn.execute(
                "SELECT status, risk_level, reason FROM ai_evaluations WHERE run_id=? ORDER BY raw_content_id",
                (run_id,),
            ).fetchall()
    finally:
        _restore_monitor_jobs(jobs_snapshot)
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)

    assert seen_progress[0]["total_candidates"] == 3
    assert seen_progress[0]["evaluated_items"] == 0
    assert seen_progress[1]["evaluated_items"] == 1
    assert seen_progress[1]["negative_count"] == 1
    assert seen_progress[1]["high_count"] == 1
    assert seen_progress[2]["evaluated_items"] == 2
    assert seen_progress[2]["negative_count"] == 2
    assert eval_summary["negative_count"] == 2
    assert eval_summary["high_count"] == 1
    assert eval_summary["pending_review_count"] == 1
    assert final_progress["evaluated_items"] == 3
    assert final_progress["total_candidates"] == 3
    assert final_progress["negative_count"] == 2
    assert final_progress["high_count"] == 1
    assert final_progress["manual_review_count"] == 1
    assert final_progress["unresolved_items"] == 0
    assert final_progress["final"] is True
    assert run["summary"]["ai_progress_final"] is True
    assert "AI 评估已完成" in run["summary"]["progress_message"]
    assert stale["ai_progress"] == final_progress
    assert terminal_run["summary"]["ai_progress"] == final_progress
    assert after_terminal == {}
    assert report["summary"]["pending_review_count"] == 1
    assert Path(report["html_path"]).exists()
    assert [row["status"] for row in rows] == ["ok", "ok", "pending_review"]
    assert "secret" not in rows[2]["reason"]
    assert terminal_run["status"] == "success"


def test_delete_running_job_is_blocked_and_stop_job_marks_stale_run(monkeypatch):
    init_db()
    job = save_job(
        {
            "law_firm_name": "停止删除测试律所",
            "aliases": [],
            "exclude_words": [],
            "keywords": ["停止删除测试律所避雷"],
            "platforms": ["dy"],
            "recipients": [],
            "enable_comments": False,
            "time_window_type": "recent_1d",
            "frequency": "daily",
            "email_time": "09:00",
            "enabled": True,
        }
    )
    run_id = create_run(job["id"])
    try:
        monkeypatch.setattr(monitor_router, "running_job_ids", lambda: [])
        with pytest.raises(HTTPException) as exc:
            asyncio.run(monitor_router.remove_job(job["id"]))
        assert exc.value.status_code == 409

        result = asyncio.run(monitor_router.stop_job_now(job["id"]))
        run = get_run(run_id)
    finally:
        _cleanup_test_records(job["id"], "")

    assert result["status"] == "cancelled_stale_run"
    assert run and run["status"] == "cancelled"


def test_run_job_blocks_platform_when_login_window_is_open(monkeypatch):
    init_db()
    jobs_snapshot = _snapshot_monitor_jobs()
    job = save_job(
        {
            "law_firm_name": "海安律所",
            "aliases": [],
            "exclude_words": [],
            "keywords": ["海安律所避雷"],
            "platforms": ["dy"],
            "recipients": [],
            "enable_comments": False,
            "time_window_type": "recent_1d",
            "frequency": "daily",
            "email_time": "09:00",
            "enabled": True,
        }
    )

    def fail_if_called(*args, **kwargs):
        raise AssertionError("MediaCrawler subprocess should not start while login window is open")

    try:
        monkeypatch.setattr(
            runner_module,
            "list_platform_status",
            lambda: [
                {"platform": "dy", "platform_label": "抖音", "profile_exists": True, "needs_login": False, "login_window_open": True},
                {"platform": "ks", "platform_label": "快手", "profile_exists": True, "needs_login": False, "login_window_open": False},
                {"platform": "xhs", "platform_label": "小红书", "profile_exists": True, "needs_login": False, "login_window_open": False},
            ],
        )
        monkeypatch.setattr(runner_module.subprocess, "run", fail_if_called)
        result = asyncio.run(run_monitor_job(job["id"]))
    finally:
        _restore_monitor_jobs(jobs_snapshot)

    assert result["status"] == "partial_failed"
    assert result["summary"]["failed_platforms"] == ["dy"]
    assert "登录窗口未关闭" in result["summary"]["platform_results"]["dy"]["error"]


def test_run_platform_retries_transient_crawler_failure(tmp_path, monkeypatch):
    init_db()
    job = {
        "id": 9991,
        "law_firm_name": "重试测试律所",
        "keywords": ["重试测试律所避雷"],
        "enable_comments": False,
        "time_window_type": "recent_1d",
    }
    calls: list[Path] = []

    def fake_run_attempt(job_arg, platform_arg, out_dir, proxy_binding=None):
        calls.append(out_dir)
        if len(calls) == 1:
            (out_dir / "crawler.log").write_text("temporary network error", encoding="utf-8")
            raise RuntimeError(f"MediaCrawler exited with 1; see {out_dir / 'crawler.log'}")
        json_dir = out_dir / "douyin" / "json"
        json_dir.mkdir(parents=True)
        (json_dir / "search_contents_retry.json").write_text(
            json.dumps(
                [
                    {
                        "aweme_id": "pytest_retry_success_001",
                        "title": "重试测试律所避雷",
                        "desc": "第二次成功",
                        "create_time": int(datetime.now(timezone.utc).timestamp()),
                    }
                ],
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )

    monkeypatch.setenv("MONITOR_CRAWLER_MAX_RETRIES", "1")
    monkeypatch.setenv("MONITOR_CRAWLER_RETRY_DELAY_SECONDS", "0")
    monkeypatch.setattr(runner_module, "list_platform_status", lambda: [{"platform": "dy", "login_window_open": False}])
    monkeypatch.setattr(runner_module, "_run_crawler_attempt", fake_run_attempt)

    result = asyncio.run(runner_module.run_platform(job, 10001, "dy", tmp_path))

    _cleanup_test_records(job["id"], "pytest_retry_success_001")

    assert len(calls) == 2
    assert calls[0].name == "attempt_1"
    assert calls[1].name == "attempt_2"
    assert result["attempts"] == 2
    assert result["max_retries"] == 1
    assert result["new_contents"] == 1


def test_run_platform_attaches_bound_proxy_summary(tmp_path, monkeypatch):
    init_db()
    snapshots = {
        "proxy_profiles": _snapshot_table("proxy_profiles"),
        "social_accounts": _snapshot_table("social_accounts"),
    }
    job = {
        "id": 9993,
        "law_firm_name": "代理测试律所",
        "keywords": ["代理测试律所避雷"],
        "enable_comments": False,
        "time_window_type": "recent_1d",
    }
    seen: dict[str, Any] = {}

    def fake_run_attempt(job_arg, platform_arg, out_dir, proxy_binding=None):
        seen["proxy_binding"] = proxy_binding
        json_dir = out_dir / "douyin" / "json"
        json_dir.mkdir(parents=True)
        (json_dir / "search_contents_proxy.json").write_text(
            json.dumps(
                [
                    {
                        "aweme_id": "pytest_proxy_success_001",
                        "title": "代理测试律所避雷",
                        "desc": "代理绑定测试",
                        "create_time": int(datetime.now(timezone.utc).timestamp()),
                    }
                ],
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        return _phase_5_1d_result_for_plan(job_arg["_browser_environment_plan"])

    try:
        proxy = save_proxy_profile(
            {
                "name": "华东采集代理",
                "provider": "manual",
                "proxy_url": "http://user:pass@127.0.0.1:8081",
                "status": "active",
                "max_concurrency": 1,
            }
        )
        account = save_social_account(
            {
                "name": "抖音采集号",
                "platform": "dy",
                "login_type": "qrcode",
                "status": "active",
                "profile_path": str(tmp_path / "dy_account_profile"),
                "proxy_id": proxy["id"],
            }
        )
        Path(resolve_account_profile_path(account["profile_key"])).mkdir(parents=True, exist_ok=True)
        monkeypatch.setenv("MONITOR_CRAWLER_MAX_RETRIES", "0")
        monkeypatch.setenv("MONITOR_BROWSER_PROXY_PROBE_URL", "https://probe.invalid/region")
        monkeypatch.setattr(runner_module, "list_platform_status", lambda: [{"platform": "dy", "login_window_open": False}])
        monkeypatch.setattr(runner_module, "_run_crawler_attempt", fake_run_attempt)

        result = asyncio.run(runner_module.run_platform(job, 10003, "dy", tmp_path))
    finally:
        _cleanup_test_records(job["id"], "pytest_proxy_success_001")
        for table, snapshot in snapshots.items():
            _restore_table(table, snapshot)

    assert seen["proxy_binding"]["proxy_url"] == "http://user:pass@127.0.0.1:8081"
    assert seen["proxy_binding"]["profile_key"] == f"1/dy/acc_{account['id']}"
    assert seen["proxy_binding"]["profile_path"] == str(resolve_account_profile_path(seen["proxy_binding"]["profile_key"]))
    assert result["account"]["account_name"] == "抖音采集号"
    assert result["account"]["profile_key"] == f"1/dy/acc_{account['id']}"
    assert result["account"]["profile_configured"] is True
    assert "profile_path" not in result["account"]
    assert result["proxy"]["proxy_id"] == proxy["id"]
    assert "user:pass" not in result["proxy"]["proxy_url"]
    assert result["new_contents"] == 1


def test_run_platform_does_not_retry_login_required_error(tmp_path, monkeypatch):
    job = {
        "id": 9992,
        "law_firm_name": "登录失败测试律所",
        "keywords": ["登录失败测试律所避雷"],
        "enable_comments": False,
        "time_window_type": "recent_1d",
    }
    calls = 0

    def fake_run_attempt(job_arg, platform_arg, out_dir, proxy_binding=None):
        nonlocal calls
        calls += 1
        raise RuntimeError("MediaCrawler exited with 1；检测到登录态失效，请先重新登录该平台账号")

    monkeypatch.setenv("MONITOR_CRAWLER_MAX_RETRIES", "3")
    monkeypatch.setenv("MONITOR_CRAWLER_RETRY_DELAY_SECONDS", "0")
    monkeypatch.setattr(runner_module, "list_platform_status", lambda: [{"platform": "dy", "login_window_open": False}])
    monkeypatch.setattr(runner_module, "_run_crawler_attempt", fake_run_attempt)

    with pytest.raises(RuntimeError, match="failed after 1 attempt"):
        asyncio.run(runner_module.run_platform(job, 10002, "dy", tmp_path))

    assert calls == 1


def test_expired_cross_process_lock_is_replaced(tmp_path, monkeypatch):
    monkeypatch.setattr(runner_module, "LOCKS_DIR", tmp_path / "locks")
    monkeypatch.setattr(runner_module, "JOB_LOCK_TTL_SECONDS", 60)
    job_id = 98765
    lock_path = runner_module.LOCKS_DIR / f"job_{job_id}.lock"
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    old_created_at = (datetime.now(timezone.utc) - timedelta(minutes=5)).isoformat()
    lock_path.write_text(json.dumps({"job_id": job_id, "created_at": old_created_at}), encoding="utf-8")

    acquired = runner_module._acquire_job_lock(job_id)
    try:
        assert acquired == lock_path
        payload = json.loads(lock_path.read_text(encoding="utf-8"))
        assert payload["job_id"] == job_id
        assert payload["created_at"] != old_created_at
    finally:
        if acquired:
            runner_module._release_job_lock(acquired)


def test_readiness_requires_successful_douyin_report_for_mvp(monkeypatch):
    monkeypatch.setattr(
        readiness_module,
        "list_platform_status",
        lambda: [
            {"platform": "dy", "platform_label": "抖音", "profile_exists": True, "needs_login": False},
            {"platform": "ks", "platform_label": "快手", "profile_exists": True, "needs_login": False},
            {"platform": "xhs", "platform_label": "小红书", "profile_exists": True, "needs_login": False},
        ],
    )
    monkeypatch.setattr(
        readiness_module,
        "get_ai_config",
        lambda masked=True: {
            "provider": "openai",
            "base_url": "https://example.com",
            "api_key": "sk-********test",
            "model": "test-model",
            "last_test_status": "success",
            "last_test_at": "2026-06-11T00:00:00+00:00",
        },
    )
    monkeypatch.setattr(
        readiness_module,
        "get_email_config",
        lambda masked=True: {
            "smtp_host": "smtp.example.com",
            "sender": "sender@example.com",
            "default_recipients": ["target@example.com"],
            "last_test_status": "success",
            "last_test_at": "2026-06-11T00:00:00+00:00",
        },
    )

    partial_reports = [
        {
            "id": 1,
            "summary": {
                "platform_results": {
                    "dy": {"status": "success", "raw_contents": 2, "new_contents": 1},
                    "ks": {"status": "success", "raw_contents": 0, "new_contents": 0},
                }
            },
        },
        {"id": 2, "summary": {"selftest": True}},
    ]
    monkeypatch.setattr(readiness_module, "list_reports", lambda limit=200: partial_reports)
    partial = readiness_module.get_readiness_status()
    partial_real_check = next(check for check in partial["checks"] if check["key"] == "real_report")

    complete_reports = [
        {
            "id": 3,
            "summary": {
                "platform_results": {
                    "dy": {"status": "success", "raw_contents": 2, "new_contents": 1},
                    "ks": {"status": "success", "raw_contents": 3, "new_contents": 2},
                    "xhs": {"status": "success", "raw_contents": 1, "new_contents": 1},
                }
            },
        },
        {"id": 4, "summary": {"selftest": True}},
    ]
    monkeypatch.setattr(readiness_module, "list_reports", lambda limit=200: complete_reports)
    complete = readiness_module.get_readiness_status()
    complete_real_check = next(check for check in complete["checks"] if check["key"] == "real_report")

    assert partial_real_check["ok"] is True
    assert partial["real_platforms"] == ["dy"]
    assert partial["empty_real_platforms"] == []
    assert partial["missing_real_platforms"] == []
    assert "抖音采集闭环已完成" in partial_real_check["message"]
    assert complete_real_check["ok"] is True
    assert complete["missing_real_platforms"] == []
    assert complete["empty_real_platforms"] == []


def test_readiness_uses_all_reports_for_real_platform_audit(monkeypatch):
    seen: dict[str, int] = {}
    monkeypatch.setattr(
        readiness_module,
        "list_platform_status",
        lambda: [
            {"platform": "dy", "platform_label": "抖音", "profile_exists": True, "needs_login": False},
            {"platform": "ks", "platform_label": "快手", "profile_exists": True, "needs_login": False},
            {"platform": "xhs", "platform_label": "小红书", "profile_exists": True, "needs_login": False},
        ],
    )
    monkeypatch.setattr(
        readiness_module,
        "get_ai_config",
        lambda masked=True: {
            "base_url": "https://example.com",
            "api_key": "sk-********test",
            "model": "test-model",
            "last_test_status": "success",
        },
    )
    monkeypatch.setattr(
        readiness_module,
        "get_email_config",
        lambda masked=True: {
            "smtp_host": "smtp.example.com",
            "sender": "sender@example.com",
            "default_recipients": ["target@example.com"],
            "last_test_status": "success",
        },
    )

    def fake_list_reports(limit=100):
        seen["limit"] = limit
        return []

    monkeypatch.setattr(readiness_module, "list_reports", fake_list_reports)

    readiness_module.get_readiness_status()

    assert seen["limit"] == 0


async def _dedupe_and_report_check(monkeypatch):
    init_db()
    job = save_job(
        {
            "law_firm_name": "监控测试律所",
            "aliases": [],
            "exclude_words": [],
            "keywords": ["监控测试律所避雷"],
            "platforms": ["dy"],
            "recipients": ["test@example.com"],
            "enable_comments": False,
            "time_window_type": "recent_1d",
            "frequency": "daily",
            "email_time": "09:00",
            "enabled": True,
        }
    )
    now_ts = int(datetime.now(timezone.utc).timestamp())
    item = {
        "aweme_id": "pytest_monitor_dy_001",
        "title": "监控测试律所避雷",
        "desc": "收费争议",
        "aweme_url": "https://example.com/video",
        "cover_url": "https://example.com/cover.jpg",
        "create_time": now_ts,
    }

    async def pending_review(job, content, comments):
        return {
            "status": "pending_review",
            "is_related": True,
            "is_negative": False,
            "risk_level": "low",
            "reason": "AI 未完成判断，请人工复核",
            "evidence_quotes": [content.get("title") or ""],
            "recommended_action": "人工复核",
            "raw_response": "",
        }

    monkeypatch.setattr(runner_module, "evaluate_content", pending_review)
    run1 = create_run(job["id"])
    first = ingest_outputs(job, run1, "dy", [item], [])
    await evaluate_new_contents(job, run1, first["content_db_ids"])
    report = create_report(
        run1,
        job,
        {"platforms": ["dy"], "failed_platforms": [], "new_contents": first["new_contents"], "negative_count": 0, "high_count": 0},
    )
    html = Path(report["html_path"]).read_text(encoding="utf-8")
    markdown = Path(report["markdown_path"]).read_text(encoding="utf-8")

    run2 = create_run(job["id"])
    second = ingest_outputs(job, run2, "dy", [item], [])
    with get_conn() as conn:
        run2_rows = conn.execute("SELECT COUNT(*) AS n FROM raw_contents WHERE run_id=?", (run2,)).fetchone()["n"]

    _cleanup_test_records(job["id"], "pytest_monitor_dy_001")

    assert first["new_contents"] == 1
    assert second["new_contents"] == 0
    assert run2_rows == 0
    assert "待人工复核" in html
    assert "- 待人工复核：1" in markdown
    assert "https://example.com/cover.jpg" in html


async def _unrelated_negative_check(monkeypatch):
    init_db()
    job = save_job(
        {
            "law_firm_name": "相关性测试律所",
            "aliases": [],
            "exclude_words": [],
            "keywords": ["相关性测试律所避雷"],
            "platforms": ["dy"],
            "recipients": [],
            "enable_comments": False,
            "time_window_type": "recent_1d",
            "frequency": "daily",
            "email_time": "09:00",
            "enabled": True,
        }
    )
    now_ts = int(datetime.now(timezone.utc).timestamp())
    item = {
        "aweme_id": "pytest_unrelated_negative_001",
        "title": "其他机构避雷",
        "desc": "投诉内容很负面，但和目标律所无关",
        "create_time": now_ts,
    }

    async def fake_evaluate_content(job, content, comments):
        return {
            "status": "ok",
            "is_related": False,
            "is_negative": True,
            "risk_level": "high",
            "reason": "内容负面但不相关",
            "evidence_quotes": ["其他机构避雷"],
            "recommended_action": "忽略",
            "raw_response": "{}",
        }

    monkeypatch.setattr(runner_module, "evaluate_content", fake_evaluate_content)
    run_id = create_run(job["id"])
    ingested = ingest_outputs(job, run_id, "dy", [item], [])
    eval_summary = await evaluate_new_contents(job, run_id, ingested["content_db_ids"])
    report = create_report(run_id, job, {"platforms": ["dy"], "failed_platforms": [], **ingested, **eval_summary})
    html = Path(report["html_path"]).read_text(encoding="utf-8")

    _cleanup_test_records(job["id"], "pytest_unrelated_negative_001")

    assert eval_summary["negative_count"] == 0
    assert eval_summary["high_count"] == 0
    assert "本次未发现新增疑似负面线索" in html
    assert "其他机构避雷" not in html


async def _selftest_report_check():
    result = await create_sample_report()
    report = result["report"]
    summary = result["summary"]
    html_path = Path(report["html_path"])
    markdown_path = Path(report["markdown_path"])
    excel_path = Path(report["excel_path"])
    html = html_path.read_text(encoding="utf-8")
    markdown = markdown_path.read_text(encoding="utf-8")
    with get_conn() as conn:
        row = conn.execute("SELECT email_status, email_error FROM reports WHERE id=?", (report["id"],)).fetchone()
    _cleanup_test_records(result["job"]["id"], f"selftest_negative_{result['run_id']}")
    _cleanup_test_records(result["job"]["id"], f"selftest_excluded_{result['run_id']}")

    assert html_path.exists()
    assert markdown_path.exists()
    assert excel_path.exists()
    assert "海安律所" in html
    assert "待人工复核" in html
    assert "- 待人工复核：1" in markdown
    assert "AI 结果仅用于舆情线索筛查" in markdown
    assert summary["email_status"] == "skipped"
    assert row["email_status"] == "skipped"
    assert row["email_error"] == "本地自测不发送邮件"


def _cleanup_test_records(job_id: int, content_id: str) -> None:
    with get_conn() as conn:
        run_ids = [r["id"] for r in conn.execute("SELECT id FROM crawl_runs WHERE job_id=?", (job_id,)).fetchall()]
        raw_ids = [
            r["id"]
            for r in conn.execute(
                "SELECT id FROM raw_contents WHERE job_id=? AND content_id=?",
                (job_id, content_id),
            ).fetchall()
        ]
        if raw_ids:
            conn.execute("DELETE FROM ai_evaluations WHERE raw_content_id IN (%s)" % ",".join("?" for _ in raw_ids), raw_ids)
        conn.execute("DELETE FROM raw_comments WHERE content_id=?", (content_id,))
        conn.execute("DELETE FROM raw_contents WHERE job_id=? AND content_id=?", (job_id, content_id))
        if run_ids:
            conn.execute("DELETE FROM reports WHERE run_id IN (%s)" % ",".join("?" for _ in run_ids), run_ids)
            conn.execute("DELETE FROM crawl_runs WHERE id IN (%s)" % ",".join("?" for _ in run_ids), run_ids)
        conn.execute("DELETE FROM monitor_jobs WHERE id=?", (job_id,))


def _email_html_body(msg) -> str:
    for part in msg.walk():
        if part.get_content_type() == "text/html":
            return part.get_content()
    raise AssertionError("email html body not found")


def _snapshot_singleton_table(table: str) -> dict:
    with get_conn() as conn:
        return dict(conn.execute(f"SELECT * FROM {table} WHERE id=1").fetchone())


def _restore_singleton_table(table: str, snapshot: dict) -> None:
    columns = [key for key in snapshot.keys() if key != "id"]
    assignments = ", ".join(f"{key}=?" for key in columns)
    values = [snapshot[key] for key in columns] + [snapshot["id"]]
    with get_conn() as conn:
        conn.execute(f"UPDATE {table} SET {assignments} WHERE id=?", values)


def _snapshot_table(table: str) -> list[dict]:
    with get_conn() as conn:
        return [dict(row) for row in conn.execute(f"SELECT * FROM {table}").fetchall()]


def _table_columns(conn, table: str) -> set[str]:
    return {row["name"] for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}


def get_user_for_test_session(token: str) -> dict:
    from api.monitoring.database import get_user_for_session_token

    user = get_user_for_session_token(token)
    assert user
    return user


def _restore_table(table: str, snapshot: list[dict]) -> None:
    with get_conn() as conn:
        conn.execute("PRAGMA foreign_keys=OFF")
        try:
            if table == "monitor_jobs":
                _delete_rows_for_new_jobs(conn, {int(row["id"]) for row in snapshot if row.get("id") is not None})
            elif table == "crawl_runs":
                _delete_rows_for_new_runs(conn, {int(row["id"]) for row in snapshot if row.get("id") is not None})
            conn.execute(f"DELETE FROM {table}")
            if snapshot:
                columns = list(snapshot[0].keys())
                placeholders = ",".join("?" for _ in columns)
                conn.executemany(
                    f"INSERT INTO {table} ({','.join(columns)}) VALUES ({placeholders})",
                    [[row[col] for col in columns] for row in snapshot],
                )
            conn.commit()
        finally:
            conn.execute("PRAGMA foreign_keys=ON")


def _cmd_value(cmd: list[str], flag: str) -> str | None:
    if flag not in cmd:
        return None
    index = cmd.index(flag)
    return cmd[index + 1] if index + 1 < len(cmd) else None


def _login_test_account(platform: str, tmp_path: Path | None = None) -> dict[str, object]:
    label = {"dy": "抖音", "ks": "快手", "xhs": "小红书"}.get(platform, platform)
    profile_root = tmp_path or Path("monitor_data/test_profiles")
    profile_name = f"{platform}_login_profile_{uuid.uuid4().hex}"
    return save_social_account(
        {
            "name": f"海安律所{label}采集号",
            "platform": platform,
            "login_type": "qrcode",
            "status": "standby",
            "profile_path": str(profile_root / profile_name),
        }
    )


def _snapshot_monitor_jobs() -> dict[str, list[dict]]:
    tables = ["monitor_jobs", "job_keywords", "job_platforms", "job_recipients"]
    with get_conn() as conn:
        return {table: [dict(row) for row in conn.execute(f"SELECT * FROM {table}").fetchall()] for table in tables}


def _restore_monitor_jobs(snapshot: dict[str, list[dict]]) -> None:
    tables = ["job_recipients", "job_platforms", "job_keywords", "monitor_jobs"]
    snapshot_job_ids = {int(row["id"]) for row in snapshot.get("monitor_jobs", []) if row.get("id") is not None}
    with get_conn() as conn:
        conn.execute("PRAGMA foreign_keys=OFF")
        try:
            _delete_rows_for_new_jobs(conn, snapshot_job_ids)
            for table in tables:
                conn.execute(f"DELETE FROM {table}")
            for table in ["monitor_jobs", "job_keywords", "job_platforms", "job_recipients"]:
                rows = snapshot.get(table, [])
                if not rows:
                    continue
                columns = list(rows[0].keys())
                placeholders = ",".join("?" for _ in columns)
                conn.executemany(
                    f"INSERT INTO {table} ({','.join(columns)}) VALUES ({placeholders})",
                    [[row[col] for col in columns] for row in rows],
                )
            conn.commit()
        finally:
            conn.execute("PRAGMA foreign_keys=ON")


def _clear_monitor_jobs() -> None:
    with get_conn() as conn:
        conn.execute("PRAGMA foreign_keys=OFF")
        try:
            for table in ["job_recipients", "job_platforms", "job_keywords", "monitor_jobs"]:
                conn.execute(f"DELETE FROM {table}")
            conn.commit()
        finally:
            conn.execute("PRAGMA foreign_keys=ON")


def _delete_rows_for_new_jobs(conn, snapshot_job_ids: set[int]) -> None:
    current_job_ids = {
        int(row["id"])
        for row in conn.execute("SELECT id FROM monitor_jobs").fetchall()
        if row["id"] is not None
    }
    test_job_ids = current_job_ids - snapshot_job_ids
    if not test_job_ids:
        return
    placeholders = ",".join("?" for _ in test_job_ids)
    ids = sorted(test_job_ids)
    conn.execute(f"DELETE FROM reports WHERE job_id IN ({placeholders})", ids)
    conn.execute(f"DELETE FROM crawl_runs WHERE job_id IN ({placeholders})", ids)
    conn.execute(f"DELETE FROM job_recipients WHERE job_id IN ({placeholders})", ids)
    conn.execute(f"DELETE FROM job_platforms WHERE job_id IN ({placeholders})", ids)
    conn.execute(f"DELETE FROM job_keywords WHERE job_id IN ({placeholders})", ids)


def _delete_rows_for_new_runs(conn, snapshot_run_ids: set[int]) -> None:
    current_run_ids = {
        int(row["id"])
        for row in conn.execute("SELECT id FROM crawl_runs").fetchall()
        if row["id"] is not None
    }
    test_run_ids = current_run_ids - snapshot_run_ids
    if not test_run_ids:
        return
    placeholders = ",".join("?" for _ in test_run_ids)
    ids = sorted(test_run_ids)
    conn.execute(f"DELETE FROM reports WHERE run_id IN ({placeholders})", ids)


class _FakeClient:
    host = "127.0.0.1"


class _FakeRequest:
    headers = {"user-agent": "pytest"}
    client = _FakeClient()


class _FakeResponse:
    def __init__(self) -> None:
        self.cookies: dict[str, dict[str, object]] = {}
        self.deleted_cookies: set[str] = set()

    def set_cookie(self, key: str, value: str, **kwargs) -> None:
        self.cookies[key] = {"value": value, **kwargs}

    def delete_cookie(self, key: str, **kwargs) -> None:
        self.deleted_cookies.add(key)
